from flask import Flask, render_template, request, send_from_directory , url_for
from jinja2 import pass_eval_context
# from numpy import append
import shutil
import re
import PyPDF2
from werkzeug.utils import secure_filename
# from tkinter import *
from flask import flash
from openpyxl.worksheet.datavalidation import DataValidation 
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
#import pypiwin32
import itertools as it
import datetime
from datetime import datetime
# import tkinter as tk
# from tkinter.filedialog import askopenfilename
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl import *
import openpyxl
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import  Font
from openpyxl.styles import  Color
from openpyxl.styles import Alignment
import requests
import time

# from tkinter import filedialog

from openpyxl.chart import (
    LineChart,
    Series,
    Reference,
)

from openpyxl.cell import Cell
from openpyxl.descriptors import (
    String,
    Sequence,
    Integer,
)
from openpyxl.descriptors.serialisable import Serialisable
import sys, string, os
import os
from flask import Flask, render_template, request, send_from_directory
import datetime
from werkzeug.utils import secure_filename
import PyPDF2 as pf
# import tkinter
# import PIL
# from tkinter import *
# import tkinter as tk
# from tkinter import messagebox
import openpyxl
from openpyxl import *
from openpyxl.styles import Border, Side
from openpyxl import Workbook
from openpyxl.cell import cell
from openpyxl.descriptors import (
    String,
    Sequence,
    Integer,
)
from openpyxl.chart import LineChart, Reference
import string
# from PIL import ImageTk, Image
# from tkinter.filedialog import askopenfilename
from openpyxl.styles import Color, PatternFill, Font, borders
from openpyxl.worksheet.dimensions import ColumnDimension


# from openpyxl.styles import colors
# from PIL import ImageTk, Image
# import pandas as pd
from openpyxl.styles import Alignment, alignment

import os
import xml.etree.ElementTree as ET

import base64
# from tkinter import filedialog
import io

import webbrowser



today=datetime.datetime.now().date()
yesterday=today-datetime.timedelta(days=1)
print(yesterday.year)
print(yesterday.month)
print(yesterday.day)
curs="EUR"
curs2="USD"

listaluni=["1","2","3","4","5","6","7","8","9","10","11","12"]
listadenluni=['Ianuarie','Februarie','Martie','Aprilie','Mai','Iunie','Iulie','August','Septembrie','Octombrie','Noiembrie','Decembrie']
listadenluni2=['January','February','March','April','May','June','July','August','September','October','November','December']
listatrez=['Trezoreria operativă Agnita','Trezorerie operativa Alesd','Trezoreria operativă Avrig','Trezoreria operativă Babadag','Trezorerie operativa Babeni','Trezoreria operativă Baia','Trezoreria operativă Baia de Aramă','Trezorerie operativa Balcesti','Trezorerie operativa Beclean','Trezorerie operativa Beius','Trezoreria operativă Boldeşti - Scăeni','Trezorerie operativa Bozovici','Trezorerie operativa Budesti','Trezoreria operativă Buftea','Trezorerie operativa Buhusi','Trezoreria operativă Buşteni','Trezorerie operativa Buzias','Trezorerie operativa Campeni','Trezoreria operativă Cehu Silvaniei','Trezorerie operativa Chisinau Cris','Trezorerie operativa Codlea','Trezorerie operativa Comunala Podu Turcului','Trezorerie operativa Costesti','Trezorerie operativa Cugir','Trezorerie operativa Darabani','Trezorerie operativa Deta','Trezorerie operativa Eforie','Trezorerie operativa Faget','Trezorerie operativa Faurei','Trezoreria operativă Feteşti','Trezorerie operativa Gaiesti','Trezorerie operativa Gura Lotrului','Trezorerie operativa Harlau','Trezorerie operativa Harsova','Trezorerie operativa Hateg','Trezorerie operativa Horezu','Trezorerie operativa Huedin','Trezorerie operativa Ianca','Trezoreria operativă Ilfov','Trezorerie operativa Ineu','Trezorerie operativa Insuratei','Trezoreria operativă Jibou','Trezorerie operativa Jimbolia','Trezorerie operativa Lehliu Gara','Trezorerie operativa Lipova','Trezoreria operativă Luduş','Trezoreria operativă Măcin','Trezorerie operativa Marghita','Trezorerie operativa Mioveni','Trezoreria operativă Mizil','Trezorerie operativa Moinesti','Trezorerie operativa Moldova Noua','Trezorerie operativa Moreni','Trezorerie operativa Motru','Trezoreria operativă Municipiu Slatina','Trezorerie operativa Municipiul  Alba Iulia','Trezorerie operativa Municipiul Adjud','Trezorerie operativa Municipiul Aiud','Trezoreria operativă Municipiul Alexandria','Trezorerie operativa Municipiul Arad','Trezorerie operativa Municipiul Bacau','Trezoreria operativă Municipiul Baia Mare','Trezoreria operativă Municipiul Bârlad','Trezorerie operativa Municipiul Bistrita','Trezorerie operativa Municipiul Blaj','Trezorerie operativa Municipiul Botosani','Trezorerie operativa Municipiul Brad','Trezorerie operativa Municipiul Braila','Trezorerie operativa Municipiul Brasov','Trezorerie operativa Municipiul Bucuresti','Trezorerie operativa Municipiul Buzau','Trezorerie operativa Municipiul Calarasi','Trezoreria operativă Municipiul Câmpina','Trezorerie operativa Municipiul Campulung','Trezorerie operativa Municipiul Campulung Moldovenesc','Trezoreria operativă Municipiul Caracal','Trezorerie operativa Municipiul Caransebes','Trezorerie operativa Municipiul Carei','Trezorerie operativa Municipiul Cluj Napoca','Trezorerie operativa Municipiul Constanta','Trezorerie operativa Municipiul Curtea de Arges','Trezorerie operativa Municipiul Dej','Trezorerie operativa Municipiul Deva','Trezorerie operativa Municipiul Dorohoi','Trezorerie operativa Municipiul Dragasani','Trezoreria operativă Municipiul Drobeta-Turnu Severin','Trezorerie operativa Municipiul Fagaras','Trezorerie operativa Municipiul Falticeni','Trezorerie operativa Municipiul Focsani','Trezorerie operativa Municipiul Galati','Trezorerie operativa Municipiul Gherla','Trezorerie operativa Municipiul Gura Humorului','Trezorerie operativa Municipiul Hunedoara','Trezoreria operativă Municipiul Huşi','Trezorerie operativa Municipiul Iasi','Trezorerie operativa Municipiul Lugoj','Trezorerie operativa Municipiul Mangalia','Trezorerie operativa Municipiul Medgidia','Trezoreria operativă Municipiul Mediaş','Trezorerie operativa Municipiul Oltenita','Trezorerie operativa Municipiul Onesti','Trezorerie operativa Municipiul Oradea','Trezorerie operativa Municipiul Orastie','Trezorerie operativa Municipiul Pascani','Trezorerie operativa Municipiul Petrosani','Trezoreria operativă Municipiul Piatra Neamţ ','Trezorerie operativa Municipiul Pitesti','Trezorerie operativa Municipiul Ploiesti','Trezorerie operativa Municipiul Radauti','Trezorerie operativa Municipiul Ramnicu Sarat','Trezorerie operativa Municipiul Ramnicu Valcea','Trezoreria operativă Municipiul Reghin','Trezorerie operativa Municipiul Resita','Trezoreria operativă Municipiul Roman','Trezoreria operativă Municipiul Roşiori de Vede','Trezorerie operativa Municipiul Satu Mare','Trezoreria operativă Municipiul Sibiu','Trezoreria operativă Municipiul Sighetul Marmaţiei','Trezoreria operativă Municipiul Sighişoara','Trezorerie operativa Municipiul Siret','Trezoreria operativă Municipiul Slobozia','Trezorerie operativa Municipiul Suceava','Trezorerie operativa Municipiul Targoviste','Trezoreria operativă Municipiul Târgu Mureş','Trezoreria operativă Municipiul Târnăveni','Trezorerie operativa Municipiul Tecuci','Trezorerie operativa Municipiul Tg.Jiu','Trezorerie operativa Municipiul Timisoara','Trezoreria operativă Municipiul Tulcea','Trezorerie operativa Municipiul Turda','Trezoreria operativă Municipiul Turnu Măgurele','Trezoreria operativă Municipiul Urziceni','Trezoreria operativă Municipiul Vaslui','Trezoreria operativă Municipiul Zalău','Trezorerie operativa Nasaud','Trezoreria operativă Negreşti','Trezorerie operativa Negresti Oas','Trezorerie operativa Novaci','Trezoreria operativă Oraş Bicaz','Trezoreria operativă Oraş Târgu Neamţ','Trezoreria operativă Oraşul Balş','Trezoreria operativă Oraşul Corabia','Trezorerie operativa Oravita','Trezoreria operativă Orşova','Trezorerie operativa Otelul Rosu','Trezorerie operativa Panciu','Trezorerie operativa Patarlagele','Trezorerie operativa Pogoanele','Trezorerie operativa Pucioasa','Trezorerie operativa Raducaneni','Trezorerie operativa Rasnov','Trezorerie operativa Rovinari','Trezorerie operativa Rupea','Trezorerie operativa Sacele','Trezoreria operativă Sălişte','Trezorerie operativa Salonta','Trezorerie operativa Sangeorz Bai','Trezorerie operativa Sannicolau Mare','Trezorerie operativa Savarsin','Trezorerie operativa Saveni','Trezorerie operativa Sebes','Trezorerie operativa Sebis','Trezorerie operativa Sector 1','Trezorerie operativa Sector 2','Trezorerie operativa Sector 3','Trezorerie operativa Sector 4','Trezorerie operativa Sector 5','Trezorerie operativa Sector 6','Trezoreria operativă Şimleul Silvaniei','Trezoreria operativă Slănic','Trezoreria operativă Sovata','Trezoreria operativă Strehaia','Trezoreria operativă Sulina','Trezoreria operativă Târgu Lăpuş','Trezorerie operativa Tasnad','Trezorerie operativa Tg.Bujor','Trezorerie operativa Tg.Carbunesti','Trezorerie operativa Tg.Frumos','Trezorerie operativa Titu','Trezorerie operativa Topoloveni','Trezoreria operativă Vălenii de Munte','Trezoreria operativă Vânju Mare','Trezorerie operativa Vatra Dornei','Trezoreria operativă Videle','Trezoreria operativă Vişeul de Sus','Trezoreria operativă Zimnicea']
listaiban=['RO42TREZ57820A100101XTVA','RO69TREZ08120A100101XTVA','RO11TREZ58120A100101XTVA','RO61TREZ64520A100101XTVA','RO02TREZ68220A100101XTVA','RO83TREZ64620A100101XTVA','RO12TREZ46220A100101XTVA','RO33TREZ67920A100101XTVA','RO46TREZ10220A100101XTVA','RO03TREZ07820A100101XTVA','RO57TREZ53920A100101XTVA','RO42TREZ19020A100101XTVA','RO59TREZ20420A100101XTVA','RO05TREZ42220A100101XTVA','RO30TREZ06620A100101XTVA','RO31TREZ52920A100101XTVA','RO31TREZ62620A100101XTVA','RO68TREZ00620A100101XTVA','RO78TREZ56220A100101XTVA','RO32TREZ02220A100101XTVA','RO40TREZ13720A100101XTVA','RO43TREZ07120A100101XTVA','RO66TREZ05020A100101XTVA','RO37TREZ00920A100101XTVA','RO10TREZ11820A100101XTVA','RO84TREZ62420A100101XTVA','RO84TREZ23620A100101XTVA','RO75TREZ62820A100101XTVA','RO04TREZ15320A100101XTVA','RO24TREZ39220A100101XTVA','RO25TREZ27320A100101XTVA','RO64TREZ67620A100101XTVA','RO10TREZ40920A100101XTVA','RO31TREZ23820A100101XTVA','RO47TREZ37120A100101XTVA','RO20TREZ67420A100101XTVA','RO45TREZ22120A100101XTVA','RO26TREZ15420A100101XTVA','RO80TREZ42120A100101XTVA','RO76TREZ02420A100101XTVA','RO79TREZ15220A100101XTVA','RO25TREZ56420A100101XTVA','RO09TREZ62520A100101XTVA','RO37TREZ20320A100101XTVA','RO98TREZ02520A100101XTVA','RO20TREZ48020A100101XTVA','RO92TREZ64220A100101XTVA','RO91TREZ08220A100101XTVA','RO44TREZ04920A100101XTVA','RO18TREZ52420A100101XTVA','RO61TREZ06320A100101XTVA','RO29TREZ18520A100101XTVA','RO47TREZ27420A100101XTVA','RO97TREZ33820A100101XTVA','RO10TREZ50620A100101XTVA','RO77TREZ00220A100101XTVA','RO28TREZ69220A100101XTVA','RO24TREZ00420A100101XTVA','RO76TREZ60620A100101XTVA','RO10TREZ02120A100101XTVA','RO17TREZ06120A100101XTVA','RO22TREZ43620A100101XTVA','RO34TREZ65720A100101XTVA','RO24TREZ10120A100101XTVA','RO02TREZ00320A100101XTVA','RO63TREZ11620A100101XTVA','RO03TREZ36920A100101XTVA','RO57TREZ15120A100101XTVA','RO05TREZ13120A100101XTVA','RO10TREZ70020A100101XTVA','RO96TREZ16620A100101XTVA','RO90TREZ20120A100101XTVA','RO71TREZ52220A100101XTVA','RO97TREZ04720A100101XTVA','RO59TREZ59220A100101XTVA','RO32TREZ50720A100101XTVA','RO60TREZ18220A100101XTVA','RO39TREZ54720A100101XTVA','RO32TREZ21620A100101XTVA','RO71TREZ23120A100101XTVA','RO22TREZ04820A100101XTVA','RO54TREZ21720A100101XTVA','RO34TREZ36620A100101XTVA','RO85TREZ11720A100101XTVA','RO73TREZ67220A100101XTVA','RO87TREZ46120A100101XTVA','RO27TREZ13220A100101XTVA','RO81TREZ59320A100101XTVA','RO06TREZ69120A100101XTVA','RO72TREZ30620A100101XTVA','RO76TREZ21820A100101XTVA','RO06TREZ59420A100101XTVA','RO56TREZ36720A100101XTVA','RO56TREZ65820A100101XTVA','RO41TREZ40620A100101XTVA','RO62TREZ62320A100101XTVA','RO18TREZ23320A100101XTVA','RO93TREZ23220A100101XTVA','RO20TREZ57720A100101XTVA','RO15TREZ20220A100101XTVA','RO39TREZ06220A100101XTVA','RO56TREZ07620A100101XTVA','RO25TREZ37020A100101XTVA','RO63TREZ40720A100101XTVA','RO78TREZ36820A100101XTVA','RO68TREZ49120A100101XTVA','RO75TREZ04620A100101XTVA','RO49TREZ52120A100101XTVA','RO28TREZ59520A100101XTVA','RO21TREZ16720A100101XTVA','RO51TREZ67120A100101XTVA','RO51TREZ47720A100101XTVA','RO38TREZ18120A100101XTVA','RO90TREZ49220A100101XTVA','RO23TREZ60820A100101XTVA','RO17TREZ54620A100101XTVA','RO95TREZ57620A100101XTVA','RO44TREZ43720A100101XTVA','RO73TREZ47820A100101XTVA','RO50TREZ59620A100101XTVA','RO02TREZ39120A100101XTVA','RO37TREZ59120A100101XTVA','RO78TREZ27120A100101XTVA','RO29TREZ47620A100101XTVA','RO95TREZ47920A100101XTVA','RO94TREZ30720A100101XTVA','RO53TREZ33620A100101XTVA','RO18TREZ62120A100101XTVA','RO70TREZ64120A100101XTVA','RO98TREZ21920A100101XTVA','RO98TREZ60720A100101XTVA','RO46TREZ39320A100101XTVA','RO12TREZ65620A100101XTVA','RO56TREZ56120A100101XTVA','RO68TREZ10320A100101XTVA','RO78TREZ65920A100101XTVA','RO61TREZ54820A100101XTVA','RO22TREZ33920A100101XTVA','RO37TREZ49420A100101XTVA','RO15TREZ49320A100101XTVA','RO54TREZ50820A100101XTVA','RO76TREZ50920A100101XTVA','RO73TREZ18720A100101XTVA','RO34TREZ46320A100101XTVA','RO95TREZ18820A100101XTVA','RO72TREZ69420A100101XTVA','RO65TREZ16920A100101XTVA','RO87TREZ17020A100101XTVA','RO69TREZ27520A100101XTVA','RO32TREZ41020A100101XTVA','RO62TREZ13820A100101XTVA','RO44TREZ34020A100101XTVA','RO49TREZ13320A100101XTVA','RO18TREZ13620A100101XTVA','RO02TREZ58520A100101XTVA','RO16TREZ08320A100101XTVA','RO90TREZ10420A100101XTVA','RO53TREZ62720A100101XTVA','RO58TREZ03220A100101XTVA','RO32TREZ11920A100101XTVA','RO46TREZ00520A100101XTVA','RO67TREZ02820A100101XTVA','RO32TREZ70120A100101XTVA','RO54TREZ70220A100101XTVA','RO76TREZ70320A100101XTVA','RO98TREZ70420A100101XTVA','RO23TREZ70520A100101XTVA','RO45TREZ70620A100101XTVA','RO03TREZ56320A100101XTVA','RO62TREZ52620A100101XTVA','RO64TREZ48220A100101XTVA','RO56TREZ46420A100101XTVA','RO39TREZ64420A100101XTVA','RO13TREZ44020A100101XTVA','RO83TREZ54920A100101XTVA','RO19TREZ30820A100101XTVA','RO75TREZ33720A100101XTVA','RO85TREZ40820A100101XTVA','RO91TREZ27620A100101XTVA','RO88TREZ05120A100101XTVA','RO09TREZ52820A100101XTVA','RO78TREZ46520A100101XTVA','RO72TREZ59720A100101XTVA','RO45TREZ60920A100101XTVA','RO35TREZ44120A100101XTVA','RO67TREZ61020A100101XTVA']

listadend300eng=['LIC','Adjustments LIC','LIS','Intra-community services rendered','Adjustments of Intra-community services rendered','AIC','AIC (supplier registered for VAT purposes in the source state)','Adjustments AIC','AIS','Acquisitions of intra-Community services','Ajustments AIS','Supplies of goods and services (19%) ','Supplies of goods and services (9%) ','Supplies of goods and services (5%) ','Acquisition of goods and services subject to simplified measures (local reverse charge)','Acquisition of goods and services (19%) ','Acquisition of goods and services (9%) ','Acquisition of goods and services (5%) ','Supplies of goods and services subject to simplified measures (local reverse charge)','Supplies of goods and services exempted w/deduction right','Supplies of goods and services exempted w/o dduction right','Adjustments supplies of goods and services ','Intra-community supplies of services according to art 278 par(8) of the Fiscal code taxable in Romania','Adjustents of intra-community supplies of services according to art 278 par(8) of the Fiscal code taxable in Romania','Collected VAT','AIC','AIC (supplier registered for VAT purposes in the source state)','Adjustments AIC','AIS','Acquisitions of intra-Community services','Adjustment AIS','Acquisitions of goods and services (19%)','Acquisitions of goods and services (9%)','Acquisitions of goods and services (5%)','Acquisition of goods and services subject to simplified measures (local reverse charge)','Acquisition of goods and services (19%) ','Acquisition of goods and services (9%) ','Acquisition of goods and services (5%) ','Compensation in the flat fee for purchases of agricultural products and services from suppliers applying the special regime','Adjusments of compensation in the flat fee for purchases of agricultural products and services from suppliers applying the special regime','Acquisitions of goods and services VAT exempt or non-taxable, out of which:','Acquisitions of VAT exempt intra-community services (do not fill in fro the simplified method)','Deductible VAT','','Sub-total deductible VAT','VAT effectively refunded to foreign buyers, including the commission of the authorised bodies','Adjustments acquisitions','Adjustments according to pro rata / adjustments for capital goods','Total VAT deducted','Negative VAT amount during the reporting period','VAT payable during the reporting period','VAT payable from previous period not paid until the current submission deadline','VAT differences determined by the tax authorities by decision and unpaid until the submission of the current VAT return','Cumulated VAT payable','Negative VAT amount from previous periods ','Negative VAT differences determined by the tax authorities by decision until the submission of the current VAT return','Cumulated negative VAT amount','Balance of payable VAT at the end of the reporting period','Negative VAT to be reported for future periods','','A','A1','B','B1']
listadend300ro=['Livrari intracomunitare de bunuri, scutite conform art. 294 alin.(2)lit.a) si d) din Codul fiscal','Regularizari livrari intracomunitare scutite conform art. 294 alin.(2)lit.a) si d) din Codul fiscal','Livrari de bunuri sau prestari de servicii pentru care locul livrarii/ locul prestarii este in afara Romaniei (in UE sau in afara UE), precum si livrari intracomunitare de bunuri, scutite conformart. 294 alin.(2) lit.b) si c) din Codul fiscal, din care:','Prestari de servicii intracomunitare care nu beneficiaza de scutire in statul membru in care taxa este datorata','Regularizari privind prestarile de servicii intracomunitare care nu beneficiaza de scutire in statul membru in care taxa este datorata','Achizitii intracomunitare de bunuri pentru care cumparatorul este obligat la plata TVA (taxare inversa), din care:','Achizitii intracomunitare pentru care cumparatorul este obligat la plata TVA (taxare inversa), iar furnizorul este inregistrat in scopuri de TVA in statul membru din care a avut loc livrarea intracomunitara','Regularizari privind achizitiile intracomunitare de bunuri pentru care cumparatorul este obligat la plata TVA (taxare inversa)','Achizitii de bunuri, altele decat cele de la rd. 5 si 6 si achizitii de servicii pentru care beneficiarul din Romania este obligat la plata TVA (taxare inversa), din care:','Achizitii de servicii intracomunitare pentru care beneficiarul este obligat la plata TVA (taxare inversa)','Regularizari privind achizitii de servicii intracomunitare pentru care beneficiarul este obligat la plata TVA (taxare inversa)','Livrari de bunuri si prestari de servicii taxabile cu cota 19%','Livrari de bunuri si prestari de servicii taxabile cu cota 9%','Livrari de bunuri si prestari de servicii taxabile cu cota 5%','Achizitii de bunuri si servicii supuse masurilor de simplificare pentru care beneficiarul este obligat la plata TVA (taxare inversa), din care','Achizitii de bunuri si servicii, taxabile cu cota 19%','Achizitii de bunuri, taxabile cu cota 9%','Achizitii de bunuri, taxabile cu cota 5%','Livrari de bunuri si prestari de servicii supuse masurilor de simplificare (taxare inversa)','Livrari de bunuri si prestari de servicii scutite cu drept de deducere, altele decat cele de la rd. 1-3','Livrari de bunuri si prestari de servicii scutite fara drept de deducere','Regularizari taxa colectata','Prestari de servicii intracomunitare conform art.278 alin.(8)din Codul fiscal pentru care locul prestarii este în Romania','Regularizari privind prestari de servicii intracomunitare conform art.278 alin.(8) din Codul fiscal pentru care locul prestarii este in Romania','TOTAL TAXA COLECTATA (suma de la rd. 1 pana la rd. 18, cu exceptia celor de la rd. 3.1, 5.1 , 7.1, 12.1, 12.2 , 12.3)','Achizitii intracomunitare de bunuri pentru care cumparatorul este obligat la plata TVA (taxare inversa), din care:','Achizitii intracomunitare pentru care cumparatorul este obligat la plata TVA (taxare inversa), iar furnizorul este inregistrat in scopuri de TVA in statul membru din care a avut loc livrarea','Regularizari privind achizitiile intracomunitare de bunuri pentru care cumparatorul este obligat la plata TVA (taxare inversa)','Achizitii de bunuri, altele decat cele de la rd.20 şi 21 si achizitii de servicii pentru care beneficiarul din Romania este obligatla plata TVA (taxare inversa), din care:','Achizitii de servicii intracomunitare pentru care beneficiarul este obligat la plata TVA (taxare inversa)','Regularizari privind achizitii de servicii intracomunitare pentru care beneficiarul este obligat la plata TVA (taxare inversa)','Achizitii de bunuri si servicii taxabile cu cota de 19%, altele decat cele de la rd. 27','Achizitii de bunuri si servicii taxabile cu cota de 9%','Achizitii de bunuri si servicii taxabile cu cota de 5%','Achizitii de bunuri si servicii supuse masurilor de simplificare pentru care beneficiarul este obligat la plata TVA (taxare inversa), din care:','Achizitii de bunuri si servicii, taxabile cu cota 19%','Achizitii de bunuri, taxabile cu cota 9%','Achizitii de bunuri, taxabile cu cota 5%','Compensatia in cota forfetara pentru achizitii de produse si servicii agricole de la furnizori care aplica regimul special pentru agricultori','Regularizari privind compensatia in cota forfetara','Achizitii de bunuri si servicii scutite de taxa sau neimpozabile, din care:','Achizitii de servicii intracomunitare scutite de taxa','TOTAL TAXA DEDUCTIBILA (suma de la rd. 20 pana la rd. 29, cu exceptia celor de la rd. 20.1, 22.1, 27.1, 27.2, 27.3)','','SUB-TOTAL TAXA DEDUSA CONFORM ART. 297 SI ART. 298SAU ART. 300 SI ART. 298DIN CODUL FISCAL SI COMPENSATIE IN COTA FORFETARA','TVA efectiv restituita cumparatorilor straini, inclusiv comisionul unitatilor autorizate','Regularizari taxa dedusa','Ajustari conform pro-rata / ajustari de taxa','TOTAL TAXA DEDUSA (rd. 32 + rd. 33 + rd. 34 + rd. 35)','Suma negativa a TVA in perioada de raportare (rd. 36 - rd. 19)','Taxa de plata in perioada de raportare (rd. 19 - rd. 36)','Soldul TVA de plata din decontul perioadei fiscale precedente (rd. 45 din decontul perioadei fiscale precedente) neachitate pana la data depunerii decontului de TVA','Diferente de TVA de plata stabilite de organele de inspectie fiscala prin decizie comunicata si neachitate pana la data depunerii decontului de TVA','TVA de plata cumulat (rd. 38 + rd. 39 + rd. 40)','Soldul sumei negative a TVA reportate din perioada precedenta pentru care nu s-a solicitat rambursare (rd. 46 din decontul perioadei fiscale precedente)','Diferente negative de TVA stabilite de organele de inspectie fiscala prin decizie comunicata pana la data depunerii decontului de TVA','Suma negativa a TVA cumulate (rd. 37 + rd. 42 + rd. 43)','Sold TVA de plata la sfarsitul perioadei de raportare (rd. 41 - rd. 44)','Soldul sumei negative de TVA la sfarsitul perioadei de raportare (rd. 44 - rd. 41)','','Livrari de bunuri si prestari de servicii realizate a caror TVA aferenta a ramas neexigibila, existenta in sold la sfarsitul perioadei de raportare, ca urmare a aplicarii sistemului TVA la incasare, din care','Livrari de bunuri si prestari de servicii realizate in ultimele 6 luni/2 trimestre calendaristice','Achizitii de bunuri si servicii realizate pentru care nu s-a exercitat dreptul de deducere a TVA aferenta, existenta in sold la sfarsitul perioadei de raportare, ca urmare a aplicarii art. 297 alin (2) si (3) din Codul fiscal, din care:','Achizitii de bunuri si servicii realizate in ultimele 6 luni/2 trimestre calendaristice']
def get_fxrate(year):


    # if ccy != "RON":  #pentru valute diferite de ron
    r=requests.get('http://www.bnr.ro/files/xml/years/nbrfxrates'+str(year)+'.xml') #accesare xml bnr
    str_xml = r.text    #stocare text xml
    #     with open("D:/text.txt", "w", encoding="utf-8") as f:



    #         f.write(str_xml)
    #     print(str_xml)
    indexcurs=str_xml.rfind("EUR")
    index_maimare = str_xml.find(">",indexcurs)
    index_maimic=str_xml.find("<",index_maimare)
    valoarecurs=""
    # for chr in range(index_maimare+1,index_maimic):
    #     print(str_xml[chr]) 
    #     valoarecurs = valoarecurs + str_xml[chr]
    

    # while str_xml.find(str(year)+"-"+str(month)+"-"+str(day)) <= 0: #cautare cea mai apropiata data dispobibila
    #     date = datetime.datetime.strptime(str(day)+"/"+str(month)+"/"+str(year),'%d/%m/%Y') #transformam in date
    #     date = date - datetime.timedelta(days=1) #scadem o zi
    #     day = str(date.day) #stocam zi, luna, an
    #     month = str(date.month)
    #     year = str(date.year)
    #     # print(str(year) + "-" + str(month) + "-" + str(day))
    #     index_start = str_xml.find("Cube date="+'"'+str(year) + "-" + str(month) + "-" + str(day))
    #     print(index_start)      #cautarea cursului din aproape in aproape prin manipulare de sir de caractere
    #     index_end = str_xml.find("</Cube>",index_start)
    #     print(index_end)
    #     # for g in range(index_start,index_end):
    #     text=str_xml[index_start:index_end]
    #     print(text)
    #     index_curs = str_xml.find('EUR',index_start,index_end)
    #     print(index_curs)
    #     index_maimare = str_xml.find(">",index_curs,index_end)
    #     index_maimic = str_xml.find("<",index_maimare,index_end)
    #     print(index_maimic,index_maimare)
    #     substr=""
    #     print("sunt aici")
    valoarecurs=""
    for chr in range(index_maimare+1,index_maimic):
        print(str_xml[chr])
        valoarecurs = valoarecurs + str_xml[chr]
    else:
        substr = 1
    return valoarecurs

def make_archive(source, destination):
        base = os.path.basename(destination)
        name = base.split('.')[0]
        format = base.split('.')[1]
        archive_from = os.path.dirname(source)
        archive_to = os.path.basename(source.strip(os.sep))
        shutil.make_archive(name, format, archive_from, archive_to)
        shutil.move('%s.%s'%(name,format), destination)

thin = Side(border_style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_left = Border(left=thin, right=None, top=thin, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_right = Border(left=None, right=thin, top=thin, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_centered = Border(left=None, right=None, top=thin, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_upperleft = Border(left=thin, top=thin)

thin = Side(border_style='thin', color='000000')
border_lowerleft = Border(left=thin, right=None, top=None, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_upperright = Border(right=thin, top=thin)

thin = Side(border_style='thin', color='000000')
border_lowerright = Border(right=thin, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_left1 = Border(left=thin)

thin = Side(border_style='thin', color='000000')
border_right1 = Border(right=thin)

thin = Side(border_style='thin', color='000000')
border_top = Border(top=thin)

thin = Side(border_style='thin', color='000000')
border_bottom = Border(bottom=thin)

app=Flask(__name__, template_folder='template')
app.secret_key = "GT ROMANIA Delivery Center"

@app.route('/')
def my_form():
    return render_template('D3APPS dashboard.html')

global LL_g
@app.route('/D3APPS/PMG')
def my_form_D300():
	return render_template('D3APPS.html')

@app.route('/D3APPS/PMG', methods=['POST', 'GET'])
def D300xml():
	if request.method == 'POST':
		clientname=request.form.get('client')
		D300 = request.files["far"]
		val1 = request.form.get('D300')
		val2 = request.form.get('D390')
		val3 = request.form.get('D394')
		val4 = request.form.get('xyz')
		dropdown = request.form.get('trezorerie')
		dropdownlimba = request.form.get('limba')
		soldLunaTrecuta = request.form.get('largeAm')

	
		# #print(soldLunaTrecuta)
	if val1=="":
		# #print("Da")  # daca e bifat
		val1 = 1
	else:
		#print(val1)            
		val1 = 0
		# #print("Nu")

	if val2=="":  # daca e bifat
		val2 = 1
	else:
		# #print(val2)            
		val2 = 0

	if val3=="":  # daca e bifat
		val3 = 1
	else:
		#print(val3)            
		val3 = 0
		
	if str(dropdownlimba)=="Romana(RO)":
		option=1
	else:
		option=0	


	#print(val4)
	#print(val4,"--------------------------------------------")		
	cap_tabel_color_verde = PatternFill(start_color = '00B050', end_color ='00B050', fill_type = 'solid')
	cap_tabel_color_verde_deschis = PatternFill(start_color = '92D050', end_color ='92D050', fill_type = 'solid')
	cap_tabel_color_black = PatternFill(start_color = '000000', end_color ='000000', fill_type = 'solid')
	cap_tabel = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
	cap_tabelbold = Font(name='Calibri', size=10, bold=True)
	cap_tabeltitlu = Font(name='Calibri', size=15, bold=True,underline='single')
	scrisincredibildemare = Font(name='Calibri', size=30, bold=True)
	cap_tabel_color_GT_movdeschis = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')
	cap_tabel_color_GT_movinchis = PatternFill(start_color='CCC0DA', end_color='CCC0DA', fill_type='solid')

	temp = openpyxl.load_workbook(D300,data_only=True)
	ws = temp.active
	
	if(option==1):
		Sheet1=temp.create_sheet('Cover sheet')
		fonta = PatternFill(start_color = 'ffffff', end_color ='ffffff', fill_type = 'solid')
		fontg = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		font2 = Font(name = 'Georgia', size = 10, bold = True, color="000000")
		font1 = Font(name = 'Georgia', size = 10, color = "FFFFFF", bold = True,italic=True)
		font3 = Font(name = 'Georgia', size = 10, color = "000000",italic=True)
		culoare = PatternFill(start_color = '182A54', end_color ='182A54', fill_type = 'solid') 
		culoare2 = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		culoare3 = PatternFill(start_color = 'D9E1F2', end_color ='D9E1F2', fill_type = 'solid')
		culoare4 = PatternFill(start_color = 'E2EFDA', end_color ='E2EFDA', fill_type = 'solid')
		culoare5 = PatternFill(start_color = 'FFF2CC', end_color ='FFF2CC', fill_type = 'solid')
		culoare6 = PatternFill(start_color = '808080', end_color ='808080', fill_type = 'solid')
		font4 = Font(name = 'Georgia', size = 10, color = "000000",underline='single',bold=True)
		font5 = Font(name = 'Georgia', size = 10, color = "ffffff",underline='single',bold=True)
		border = Border(bottom=Side(style='dotted'))
		border2 = Border(top=Side(style='dotted'))
		border3 = Border(left=Side(style='dotted'))
		border4 = Border(right=Side(style='dotted'))
		border5 = Border(left=Side(style='dotted'),top=Side(style='dotted'))
		border6 = Border(left=Side(style='dotted'),bottom=Side(style='dotted'))
		border7 = Border(right=Side(style='dotted'),bottom=Side(style='dotted'))
		border8 = Border(right=Side(style='dotted'),top=Side(style='dotted'))
		# border9 = Border(right=Side(style='double'),bottom=Side(style='double'),top=Side(style='double'),left=Side(style='double'))


		Sheet1.cell(row=31, column=2).border=border2
		Sheet1.cell(row=31, column=3).border=border2
		Sheet1.cell(row=31, column=4).border=border2
		Sheet1.cell(row=31, column=5).border=border2
		Sheet1.cell(row=31, column=6).border=border2
		Sheet1.cell(row=31, column=7).border=border2
		Sheet1.cell(row=31, column=2).border=border5
		Sheet1.cell(row=31, column=7).border=border8


		Sheet1.cell(row=75, column=2).border=border
		Sheet1.cell(row=75, column=3).border=border
		Sheet1.cell(row=75, column=4).border=border
		Sheet1.cell(row=75, column=5).border=border
		Sheet1.cell(row=75, column=6).border=border
		Sheet1.cell(row=75, column=7).border=border
		Sheet1.cell(row=75, column=7).border=border7

		Sheet1.cell(row=32, column=2).border=border3
		Sheet1.cell(row=33, column=2).border=border3
		Sheet1.cell(row=34, column=2).border=border3
		Sheet1.cell(row=35, column=2).border=border3
		Sheet1.cell(row=36, column=2).border=border3
		Sheet1.cell(row=37, column=2).border=border3
		Sheet1.cell(row=38, column=2).border=border3
		Sheet1.cell(row=39, column=2).border=border3
		Sheet1.cell(row=40, column=2).border=border3
		Sheet1.cell(row=41, column=2).border=border3
		Sheet1.cell(row=42, column=2).border=border3
		Sheet1.cell(row=43, column=2).border=border3
		Sheet1.cell(row=44, column=2).border=border3
		Sheet1.cell(row=45, column=2).border=border3
		Sheet1.cell(row=46, column=2).border=border3
		Sheet1.cell(row=47, column=2).border=border3
		Sheet1.cell(row=48, column=2).border=border3
		Sheet1.cell(row=49, column=2).border=border3
		Sheet1.cell(row=50, column=2).border=border3
		Sheet1.cell(row=51, column=2).border=border3
		Sheet1.cell(row=52, column=2).border=border3
		Sheet1.cell(row=53, column=2).border=border3
		Sheet1.cell(row=54, column=2).border=border3
		Sheet1.cell(row=55, column=2).border=border3
		Sheet1.cell(row=75, column=2).border=border6
		Sheet1.cell(row=56, column=2).border=border3
		Sheet1.cell(row=57, column=2).border=border3
		Sheet1.cell(row=58, column=2).border=border3
		Sheet1.cell(row=59, column=2).border=border3
		Sheet1.cell(row=60, column=2).border=border3
		Sheet1.cell(row=61, column=2).border=border3
		Sheet1.cell(row=62, column=2).border=border3
		Sheet1.cell(row=63, column=2).border=border3
		Sheet1.cell(row=64, column=2).border=border3
		Sheet1.cell(row=65, column=2).border=border3
		Sheet1.cell(row=66, column=2).border=border3
		Sheet1.cell(row=67, column=2).border=border3
		Sheet1.cell(row=68, column=2).border=border3
		Sheet1.cell(row=69, column=2).border=border3
		Sheet1.cell(row=70, column=2).border=border3
		Sheet1.cell(row=71, column=2).border=border3
		Sheet1.cell(row=72, column=2).border=border3
		Sheet1.cell(row=73, column=2).border=border3
		Sheet1.cell(row=74, column=2).border=border3

		Sheet1.cell(row=32, column=7).border=border4
		Sheet1.cell(row=33, column=7).border=border4
		Sheet1.cell(row=34, column=7).border=border4
		Sheet1.cell(row=35, column=7).border=border4
		Sheet1.cell(row=36, column=7).border=border4
		Sheet1.cell(row=37, column=7).border=border4
		Sheet1.cell(row=38, column=7).border=border4
		Sheet1.cell(row=39, column=7).border=border4
		Sheet1.cell(row=40, column=7).border=border4
		Sheet1.cell(row=41, column=7).border=border4
		Sheet1.cell(row=42, column=7).border=border4
		Sheet1.cell(row=43, column=7).border=border4
		Sheet1.cell(row=44, column=7).border=border4
		Sheet1.cell(row=45, column=7).border=border4
		Sheet1.cell(row=46, column=7).border=border4
		Sheet1.cell(row=47, column=7).border=border4
		Sheet1.cell(row=48, column=7).border=border4
		Sheet1.cell(row=49, column=7).border=border4
		Sheet1.cell(row=50, column=7).border=border4
		Sheet1.cell(row=51, column=7).border=border4
		Sheet1.cell(row=52, column=7).border=border4
		Sheet1.cell(row=53, column=7).border=border4
		Sheet1.cell(row=54, column=7).border=border4
		Sheet1.cell(row=55, column=7).border=border4
		Sheet1.cell(row=56, column=7).border=border4
		Sheet1.cell(row=57, column=7).border=border4
		Sheet1.cell(row=58, column=7).border=border4
		Sheet1.cell(row=59, column=7).border=border4
		Sheet1.cell(row=60, column=7).border=border4
		Sheet1.cell(row=61, column=7).border=border4
		Sheet1.cell(row=62, column=7).border=border4
		Sheet1.cell(row=63, column=7).border=border4
		Sheet1.cell(row=64, column=7).border=border4
		Sheet1.cell(row=65, column=7).border=border4
		Sheet1.cell(row=66, column=7).border=border4
		Sheet1.cell(row=67, column=7).border=border4
		Sheet1.cell(row=68, column=7).border=border4
		Sheet1.cell(row=69, column=7).border=border4
		Sheet1.cell(row=70, column=7).border=border4
		Sheet1.cell(row=71, column=7).border=border4
		Sheet1.cell(row=72, column=7).border=border4
		Sheet1.cell(row=73, column=7).border=border4
		Sheet1.cell(row=74, column=7).border=border4
		info=temp['Other info']
		valluna=""
		vallunaurmatoare=""
		valIban=""
		okdecembrie=0
		for i in range(0,len(listaluni)):
			if(str(info.cell(row=3,column=3).value)=="12"):
				okdecembrie=1
				vallunaurmatoare=listadenluni[0]
				valluna=listadenluni[11]
			else:
				if(listaluni[i]==str(info.cell(row=3,column=3).value)):
					valluna=listadenluni[i]
					vallunaurmatoare=listadenluni[i+1]
		var=Sheet1.cell(row=12,column=4).value
		print(var)
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		for j in range(0,len(listatrez)):
			if(listatrez[j]==str(Sheet1.cell(row=10,column=4).value)):
				valIban=listaiban[j]
		Sheet1.cell(row=60, column=3).value='Perioada de plata: '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=61, column=3).value='="Suma de plata: " &D55&" RON "'
		Sheet1.cell(row=62, column=3).value="Moneda: RON"
		Sheet1.cell(row=63, column=3).value='Detalii plata: Decont TVA - '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		if(okdecembrie==1):
			Sheet1.cell(row=64, column=3).value='Data scadenta: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value+1)
		else:
			Sheet1.cell(row=64, column=3).value='Data scadenta: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=66, column=3).value='="Cod TVA: " & D8'
		Sheet1.cell(row=67, column=3).value='="Adresa: " &D7'
		Sheet1.cell(row=69, column=3).value="Beneficiar: BUGETUL DE STAT"
		Sheet1.cell(row=70, column=3).value='Cont IBAN: '+ str(valIban)
		Sheet1.cell(row=71, column=3).value="SWIFT / BIC: TREZROBU"
		Sheet1.cell(row=72, column=3).value="Deschis la:"+str(dropdown)
		Sheet1.cell(row=74, column=3).value="Nota: Orice taxe bancare legate de plata trebuie sa fie acoperite de catre platitor"
		Sheet1.cell(row=74, column=3).font=font2
		

		Sheet1.cell(row=14, column=3).value="Sumar"
		Sheet1.cell(row=14, column=3).font=font4
		Sheet1.cell(row=58, column=3).value="ORDIN DE PLATA"
		Sheet1.cell(row=58, column=3).font=font5

		for row in Sheet1['A1:N100']:
					for cell in row:
						cell.fill = fonta

		for row in Sheet1['N1:Z100']:
					for cell in row:
						cell.fill = fontg
		print(get_fxrate(today.year))

		Sheet1.cell(row = 6, column = 3).value = "Denumire"
		Sheet1.cell(row = 6, column = 4).value = "='Other info'!C4"
		Sheet1.cell(row = 7, column = 3).value = "Adresa"
		Sheet1.cell(row = 7, column = 4).value = "='Other info'!C6"
		Sheet1.cell(row = 8, column = 3).value = "CUI"
		Sheet1.cell(row = 8, column = 4).value = "='Other info'!C5"
		Sheet1['D8'].alignment = Alignment(wrapText=True, horizontal='left')
		Sheet1.cell(row = 9, column = 3).value = "Nr. Reg. Com."
		Sheet1.cell(row = 9, column = 4).value = "J08/1139/2017"
		Sheet1.cell(row = 10, column = 3).value = "Administratia de care apartine"
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		Sheet1.cell(row = 11, column = 3).value = "Frecventa depunere declaratie/plata"
		Sheet1.cell(row = 11, column = 4).value = "Monthly"
		Sheet1.cell(row = 12, column = 3).value = "Perioada de raportare"
		Sheet1.cell(row = 12, column = 4).value = "=date('Other info'!C2,'Other info'!C3,1)"
		Sheet1.cell(row = 12, column = 4).number_format = 'mmmm yyyy'

		Sheet1.cell(row = 43, column = 4).value="Yes"
		Sheet1.cell(row = 46, column = 4).value="N/a"
		Sheet1.cell(row = 47, column = 4).value="N/a"
		Sheet1.cell(row = 50, column = 4).value="N/a"
		Sheet1.cell(row = 51, column = 4).value="N/a"
		Sheet1.cell(row = 43, column = 4).font=font5
		Sheet1.cell(row = 46, column = 4).font=font5
		Sheet1.cell(row = 47, column = 4).font=font5
		Sheet1.cell(row = 50, column = 4).font=font5
		Sheet1.cell(row = 51, column = 4).font=font5


		Sheet1.cell(row = 6, column = 3).font=font1
		Sheet1.cell(row = 6, column = 4).font=font2
		Sheet1.cell(row = 7, column = 3).font=font1
		Sheet1.cell(row = 7, column = 4).font=font2
		Sheet1.cell(row = 8, column = 3).font=font1
		Sheet1.cell(row = 8, column = 4).font=font2
		Sheet1.cell(row = 9, column = 3).font=font1
		Sheet1.cell(row = 9, column = 4).font=font2
		Sheet1.cell(row = 10, column = 3).font=font1
		Sheet1.cell(row = 10, column = 4).font=font3
		Sheet1.cell(row = 11, column = 3).font=font1
		Sheet1.cell(row = 11, column = 4).font=font3
		Sheet1.cell(row = 12, column = 3).font=font1
		Sheet1.cell(row = 12, column = 4).font=font3

		Sheet1.cell(row = 6, column = 3).fill=culoare
		Sheet1.cell(row = 7, column = 3).fill=culoare
		Sheet1.cell(row = 8, column = 3).fill=culoare
		Sheet1.cell(row = 9, column = 3).fill=culoare
		Sheet1.cell(row = 10, column = 3).fill=culoare
		Sheet1.cell(row = 11, column = 3).fill=culoare
		Sheet1.cell(row = 12, column = 3).fill=culoare
		Sheet1.cell(row = 6, column = 4).fill=culoare2
		Sheet1.cell(row = 7, column = 4).fill=culoare2
		Sheet1.cell(row = 8, column = 4).fill=culoare2
		Sheet1.cell(row = 9, column = 4).fill=culoare2
		Sheet1.cell(row = 10, column = 4).fill=culoare2
		Sheet1.cell(row = 11, column = 4).fill=culoare2
		Sheet1.cell(row = 12, column = 4).fill=culoare2

		Sheet1.cell(row = 16, column = 3).fill=culoare
		Sheet1.cell(row=16, column=3).font=font1
		Sheet1.cell(row=16, column=3).value="  D300"
		Sheet1.cell(row=16, column=3).hyperlink="#'D300 draft figures'!A1"
		# Sheet1.cell(row=16, column=3).border=border9
		Sheet1.row_dimensions[18].height=8

		Sheet1.cell(row = 19, column = 3).fill=culoare
		Sheet1.cell(row=19, column=3).font=font1
		Sheet1.cell(row=19, column=3).value="  D390"
		Sheet1.cell(row=19, column=3).hyperlink="#'D390 workings'!A1"
		# Sheet1.cell(row=19, column=3).border=border9
		Sheet1.row_dimensions[21].height=8

		Sheet1.cell(row = 22, column = 3).fill=culoare
		Sheet1.cell(row=22, column=3).font=font1
		Sheet1.cell(row=22, column=3).value="  D394"
		Sheet1.cell(row=22, column=3).hyperlink="#'D394--->>>'!A1"
		# Sheet1.cell(row=22, column=3).border=border9
		Sheet1.row_dimensions[24].height=8

		Sheet1.cell(row = 25, column = 3).fill=culoare
		Sheet1.cell(row=25, column=3).font=font1
		Sheet1.cell(row=25, column=3).value="  Jurnal vanzari"
		Sheet1.cell(row=25, column=3).hyperlink="#'Sales'!A1"
		# Sheet1.cell(row=25, column=3).border=border9
		Sheet1.row_dimensions[27].height=8

		Sheet1.cell(row = 28, column = 3).fill=culoare
		Sheet1.cell(row=28, column=3).font=font1
		Sheet1.cell(row=28, column=3).value="  Jurnal cumparari"
		Sheet1.cell(row=28, column=3).hyperlink="#'Purchases'!A1"
		# Sheet1.cell(row=28, column=3).border=border9
		Sheet1.row_dimensions[30].height=8


		Sheet1.cell(row = 58, column = 3).fill=culoare6
		Sheet1.cell(row = 32, column = 3).fill=culoare3
		Sheet1.cell(row = 33, column = 3).fill=culoare3
		Sheet1.cell(row = 34, column = 3).fill=culoare3
		Sheet1.cell(row = 35, column = 3).fill=culoare3
		Sheet1.cell(row = 36, column = 3).fill=culoare3
		Sheet1.cell(row = 37, column = 3).fill=culoare3
		Sheet1.cell(row = 38, column = 3).fill=culoare3
		Sheet1.cell(row = 39, column = 3).fill=culoare3
		Sheet1.cell(row = 32, column = 4).fill=culoare3
		Sheet1.cell(row = 33, column = 4).fill=culoare3
		Sheet1.cell(row = 34, column = 4).fill=culoare3
		Sheet1.cell(row = 35, column = 4).fill=culoare3
		Sheet1.cell(row = 36, column = 4).fill=culoare3
		Sheet1.cell(row = 37, column = 4).fill=culoare3
		Sheet1.cell(row = 38, column = 4).fill=culoare3
		Sheet1.cell(row = 39, column = 4).fill=culoare3
		Sheet1.cell(row = 32, column = 6).fill=culoare3
		Sheet1.cell(row = 33, column = 6).fill=culoare3
		Sheet1.cell(row = 34, column = 6).fill=culoare3
		Sheet1.cell(row = 35, column = 6).fill=culoare3
		Sheet1.cell(row = 36, column = 6).fill=culoare3
		Sheet1.cell(row = 37, column = 6).fill=culoare3
		Sheet1.cell(row = 38, column = 6).fill=culoare3
		Sheet1.cell(row = 39, column = 6).fill=culoare3

		Sheet1.cell(row = 41, column = 6).fill=culoare4
		Sheet1.cell(row = 42, column = 6).fill=culoare4
		Sheet1.cell(row = 43, column = 6).fill=culoare4
		Sheet1.cell(row = 44, column = 6).fill=culoare4
		Sheet1.cell(row = 45, column = 6).fill=culoare4
		Sheet1.cell(row = 46, column = 6).fill=culoare4
		Sheet1.cell(row = 47, column = 6).fill=culoare4
		Sheet1.cell(row = 48, column = 6).fill=culoare4
		Sheet1.cell(row = 49, column = 6).fill=culoare4
		Sheet1.cell(row = 50, column = 6).fill=culoare4
		Sheet1.cell(row = 51, column = 6).fill=culoare4
		Sheet1.cell(row = 41, column = 3).fill=culoare4
		Sheet1.cell(row = 42, column = 3).fill=culoare4
		Sheet1.cell(row = 43, column = 3).fill=culoare4
		Sheet1.cell(row = 44, column = 3).fill=culoare4
		Sheet1.cell(row = 45, column = 3).fill=culoare4
		Sheet1.cell(row = 46, column = 3).fill=culoare4
		Sheet1.cell(row = 47, column = 3).fill=culoare4
		Sheet1.cell(row = 48, column = 3).fill=culoare4
		Sheet1.cell(row = 49, column = 3).fill=culoare4
		Sheet1.cell(row = 50, column = 3).fill=culoare4
		Sheet1.cell(row = 51, column = 3).fill=culoare4
		Sheet1.cell(row = 41, column = 4).fill=culoare4
		Sheet1.cell(row = 42, column = 4).fill=culoare4
		Sheet1.cell(row = 43, column = 4).fill=culoare6
		Sheet1.cell(row = 44, column = 4).fill=culoare4
		Sheet1.cell(row = 45, column = 4).fill=culoare4
		Sheet1.cell(row = 46, column = 4).fill=culoare6
		Sheet1.cell(row = 47, column = 4).fill=culoare6
		Sheet1.cell(row = 48, column = 4).fill=culoare4
		Sheet1.cell(row = 49, column = 4).fill=culoare4
		Sheet1.cell(row = 50, column = 4).fill=culoare6
		Sheet1.cell(row = 51, column = 4).fill=culoare6

		Sheet1.cell(row = 53, column = 4).fill=culoare5
		Sheet1.cell(row = 54, column = 4).fill=culoare5
		Sheet1.cell(row = 55, column = 4).fill=culoare5
		Sheet1.cell(row = 53, column = 3).fill=culoare5
		Sheet1.cell(row = 54, column = 3).fill=culoare5
		Sheet1.cell(row = 55, column = 3).fill=culoare5
		Sheet1.cell(row = 53, column = 6).fill=culoare5
		Sheet1.cell(row = 54, column = 6).fill=culoare5
		Sheet1.cell(row = 55, column = 6).fill=culoare5


		Sheet1.cell(row = 32, column = 3).value="Pozitia curenta din punct de vedere TVA"
		Sheet1.cell(row = 32, column = 3).font=font4
		Sheet1.cell(row = 34, column = 3).value="Input TVA perioada curenta"
		Sheet1.cell(row = 35, column = 3).value="Output TVA perioada curenta"
		Sheet1.cell(row = 36, column = 3).value="TVA de plata perioada curenta"
		Sheet1.cell(row = 37, column = 3).value="TVA de recuperat perioada curenta"
		Sheet1.cell(row = 38, column = 3).value="TVA in curs de decontare pentru achizitii"
		Sheet1.cell(row = 39, column = 3).value="TVA in curs de decontare pentru livrari"
		Sheet1.cell(row = 32, column = 4).value="RON"

		Sheet1.cell(row = 32, column = 6).value="Euro(@"+get_fxrate(2022)+")"
		Sheet1.cell(row = 34, column = 4).value="='D300 draft figures'!C56"
		Sheet1.cell(row = 34, column = 6).value="=IFERROR(D34/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=34, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=34, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row = 35, column = 4).value="='D300 draft figures'!C32"
		Sheet1.cell(row = 35, column = 6).value="=IFERROR(D35/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 36, column = 4).value="=IF('D300 draft figures'!C58<>0,'D300 draft figures'!C58,0)"
		Sheet1.cell(row = 36, column = 6).value="=IFERROR(D36/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=35, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=35, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=6).number_format = '#,##0_);(#,##0)'

		Sheet1.cell(row=37, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=37, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=6).number_format = '#,##0_);(#,##0)'				

		Sheet1.cell(row = 37, column = 4).value='''=IF('D300 draft figures'!C57<>0,'D300 draft figures'!C57,"nil")'''
		Sheet1.cell(row = 37, column = 6).value="=iferror(D37/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 38, column = 4).value="='D300 draft figures'!C70"
		Sheet1.cell(row = 38, column = 6).value="=iferror(D38/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 39, column = 4).value=0
		Sheet1.cell(row = 39, column = 6).value="=iferror(D39/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 55, column = 4).value='''=IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)=0),'Cover sheet'!D36,
IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)<>0),IF('Cover sheet'!D36>IFERROR(VALUE('Cover sheet'!D47),0),'Cover sheet'!D36-IFERROR(VALUE('Cover sheet'!D47),0),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="No"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D47),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="Yes"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D51),0),"N/A"))))'''
		Sheet1.cell(row = 55, column = 6).value="=iferror(D55/"+get_fxrate(2022)+",0)"

		Sheet1.cell(row = 41, column = 3).value="Pozitia reportata"
		Sheet1.cell(row = 41, column = 3).font=font4
		Sheet1.row_dimensions[42].height = 0.2
		Sheet1.cell(row = 43, column = 3).value="Solicitat la rambursare"
		Sheet1.cell(row = 44, column = 3).value
		Sheet1.cell(row = 45, column = 3).value="TVA de rambursat nesolicitat"
		Sheet1.cell(row = 46, column = 3).value="Perioada"
		Sheet1.cell(row = 47, column = 3).value="Suma"
		Sheet1.cell(row = 48, column = 3).value
		Sheet1.cell(row = 49, column = 3).value="TVA de rambursat solicitat si in curs de auditare"
		Sheet1.cell(row = 50, column = 3).value="Perioada"
		Sheet1.cell(row = 51, column = 3).value="Suma"


		Sheet1.cell(row = 53, column = 3).value="Pozitia balantei de TVA"
		Sheet1.cell(row = 53, column = 3).font=font4
		Sheet1.cell(row = 55, column = 3).value="Pozitia TVA in exercitiul curent"


		Sheet1['C16'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C19'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C22'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C25'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C28'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['D10'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D11'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D12'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['C58'].alignment = Alignment(wrapText=True, horizontal='center')

		Sheet1.column_dimensions['C'].width = 65
		Sheet1.column_dimensions['D'].width = 20
		Sheet1.column_dimensions['A'].width = 2
		Sheet1.column_dimensions['B'].width = 3
		Sheet1.column_dimensions['E'].width = 1
		Sheet1.column_dimensions['F'].width = 20
		Sheet1.column_dimensions['G'].width = 2

		# img= openpyxl.drawing.image.Image('test.png')
		# Sheet1.add_image(img,'C16')

		# img= openpyxl.drawing.image.Image('test2.png')
		# Sheet1.add_image(img,'C19')

		# img= openpyxl.drawing.image.Image('test10.png')
		# Sheet1.add_image(img,'C22')

		# img= openpyxl.drawing.image.Image('test6.png')
		# Sheet1.add_image(img,'C25')

		# img= openpyxl.drawing.image.Image('test7.png')
		# Sheet1.add_image(img,'C28')

		# img= openpyxl.drawing.image.Image('test6.png')
		# Sheet1.add_image(img,'D16')

		# img= openpyxl.drawing.image.Image('test7.png')
		# Sheet1.add_image(img,'D19')

		# img= openpyxl.drawing.image.Image('test8.png')
		# Sheet1.add_image(img,'D22')

		# img= openpyxl.drawing.image.Image('test9.png')
		# Sheet1.add_image(img,'D25')

		# img= openpyxl.drawing.image.Image('test5.png')
		# Sheet1.add_image(img,'D28')

		Sheet1.merge_cells(start_row=28, start_column=3, end_row=29, end_column=3)
		Sheet1.merge_cells(start_row=25, start_column=3, end_row=26, end_column=3)
		Sheet1.merge_cells(start_row=22, start_column=3, end_row=23, end_column=3)
		Sheet1.merge_cells(start_row=19, start_column=3, end_row=20, end_column=3)
		Sheet1.merge_cells(start_row=16, start_column=3, end_row=17, end_column=3)
		Sheet1.merge_cells(start_row=6, start_column=4, end_row=6, end_column=12)
		Sheet1.merge_cells(start_row=7, start_column=4, end_row=7, end_column=12)
		Sheet1.merge_cells(start_row=8, start_column=4, end_row=8, end_column=12)
		Sheet1.merge_cells(start_row=9, start_column=4, end_row=9, end_column=12)
		Sheet1.merge_cells(start_row=10, start_column=4, end_row=10, end_column=12)
		Sheet1.merge_cells(start_row=11, start_column=4, end_row=11, end_column=12)
		Sheet1.merge_cells(start_row=12, start_column=4, end_row=12, end_column=12)
		Sheet1.merge_cells(start_row=58, start_column=3, end_row=58, end_column=6)

	if(option==0):
		Sheet1=temp.create_sheet('Cover sheet')
		fonta = PatternFill(start_color = 'ffffff', end_color ='ffffff', fill_type = 'solid')
		fontg = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		font2 = Font(name = 'Georgia', size = 10, bold = True, color="000000")
		font1 = Font(name = 'Georgia', size = 10, color = "FFFFFF", bold = True,italic=True)
		font3 = Font(name = 'Georgia', size = 10, color = "000000",italic=True)
		culoare = PatternFill(start_color = '182A54', end_color ='182A54', fill_type = 'solid') 
		culoare2 = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		culoare3 = PatternFill(start_color = 'D9E1F2', end_color ='D9E1F2', fill_type = 'solid')
		culoare4 = PatternFill(start_color = 'E2EFDA', end_color ='E2EFDA', fill_type = 'solid')
		culoare5 = PatternFill(start_color = 'FFF2CC', end_color ='FFF2CC', fill_type = 'solid')
		culoare6 = PatternFill(start_color = '808080', end_color ='808080', fill_type = 'solid')
		font4 = Font(name = 'Georgia', size = 10, color = "000000",underline='single',bold=True)
		font5 = Font(name = 'Georgia', size = 10, color = "ffffff",underline='single',bold=True)
		border = Border(bottom=Side(style='dotted'))
		border2 = Border(top=Side(style='dotted'))
		border3 = Border(left=Side(style='dotted'))
		border4 = Border(right=Side(style='dotted'))
		border5 = Border(left=Side(style='dotted'),top=Side(style='dotted'))
		border6 = Border(left=Side(style='dotted'),bottom=Side(style='dotted'))
		border7 = Border(right=Side(style='dotted'),bottom=Side(style='dotted'))
		border8 = Border(right=Side(style='dotted'),top=Side(style='dotted'))
		# border9 = Border(right=Side(style='double'),bottom=Side(style='double'),top=Side(style='double'),left=Side(style='double'))

		Sheet1.cell(row=31, column=2).border=border2
		Sheet1.cell(row=31, column=3).border=border2
		Sheet1.cell(row=31, column=4).border=border2
		Sheet1.cell(row=31, column=5).border=border2
		Sheet1.cell(row=31, column=6).border=border2
		Sheet1.cell(row=31, column=7).border=border2
		Sheet1.cell(row=31, column=2).border=border5
		Sheet1.cell(row=31, column=7).border=border8


		Sheet1.cell(row=75, column=2).border=border
		Sheet1.cell(row=75, column=3).border=border
		Sheet1.cell(row=75, column=4).border=border
		Sheet1.cell(row=75, column=5).border=border
		Sheet1.cell(row=75, column=6).border=border
		Sheet1.cell(row=75, column=7).border=border
		Sheet1.cell(row=75, column=7).border=border7

		Sheet1.cell(row=32, column=2).border=border3
		Sheet1.cell(row=33, column=2).border=border3
		Sheet1.cell(row=34, column=2).border=border3
		Sheet1.cell(row=35, column=2).border=border3
		Sheet1.cell(row=36, column=2).border=border3
		Sheet1.cell(row=37, column=2).border=border3
		Sheet1.cell(row=38, column=2).border=border3
		Sheet1.cell(row=39, column=2).border=border3
		Sheet1.cell(row=40, column=2).border=border3
		Sheet1.cell(row=41, column=2).border=border3
		Sheet1.cell(row=42, column=2).border=border3
		Sheet1.cell(row=43, column=2).border=border3
		Sheet1.cell(row=44, column=2).border=border3
		Sheet1.cell(row=45, column=2).border=border3
		Sheet1.cell(row=46, column=2).border=border3
		Sheet1.cell(row=47, column=2).border=border3
		Sheet1.cell(row=48, column=2).border=border3
		Sheet1.cell(row=49, column=2).border=border3
		Sheet1.cell(row=50, column=2).border=border3
		Sheet1.cell(row=51, column=2).border=border3
		Sheet1.cell(row=52, column=2).border=border3
		Sheet1.cell(row=53, column=2).border=border3
		Sheet1.cell(row=54, column=2).border=border3
		Sheet1.cell(row=55, column=2).border=border3
		Sheet1.cell(row=75, column=2).border=border6
		Sheet1.cell(row=56, column=2).border=border3
		Sheet1.cell(row=57, column=2).border=border3
		Sheet1.cell(row=58, column=2).border=border3
		Sheet1.cell(row=59, column=2).border=border3
		Sheet1.cell(row=60, column=2).border=border3
		Sheet1.cell(row=61, column=2).border=border3
		Sheet1.cell(row=62, column=2).border=border3
		Sheet1.cell(row=63, column=2).border=border3
		Sheet1.cell(row=64, column=2).border=border3
		Sheet1.cell(row=65, column=2).border=border3
		Sheet1.cell(row=66, column=2).border=border3
		Sheet1.cell(row=67, column=2).border=border3
		Sheet1.cell(row=68, column=2).border=border3
		Sheet1.cell(row=69, column=2).border=border3
		Sheet1.cell(row=70, column=2).border=border3
		Sheet1.cell(row=71, column=2).border=border3
		Sheet1.cell(row=72, column=2).border=border3
		Sheet1.cell(row=73, column=2).border=border3
		Sheet1.cell(row=74, column=2).border=border3

		Sheet1.cell(row=32, column=7).border=border4
		Sheet1.cell(row=33, column=7).border=border4
		Sheet1.cell(row=34, column=7).border=border4
		Sheet1.cell(row=35, column=7).border=border4
		Sheet1.cell(row=36, column=7).border=border4
		Sheet1.cell(row=37, column=7).border=border4
		Sheet1.cell(row=38, column=7).border=border4
		Sheet1.cell(row=39, column=7).border=border4
		Sheet1.cell(row=40, column=7).border=border4
		Sheet1.cell(row=41, column=7).border=border4
		Sheet1.cell(row=42, column=7).border=border4
		Sheet1.cell(row=43, column=7).border=border4
		Sheet1.cell(row=44, column=7).border=border4
		Sheet1.cell(row=45, column=7).border=border4
		Sheet1.cell(row=46, column=7).border=border4
		Sheet1.cell(row=47, column=7).border=border4
		Sheet1.cell(row=48, column=7).border=border4
		Sheet1.cell(row=49, column=7).border=border4
		Sheet1.cell(row=50, column=7).border=border4
		Sheet1.cell(row=51, column=7).border=border4
		Sheet1.cell(row=52, column=7).border=border4
		Sheet1.cell(row=53, column=7).border=border4
		Sheet1.cell(row=54, column=7).border=border4
		Sheet1.cell(row=55, column=7).border=border4
		Sheet1.cell(row=56, column=7).border=border4
		Sheet1.cell(row=57, column=7).border=border4
		Sheet1.cell(row=58, column=7).border=border4
		Sheet1.cell(row=59, column=7).border=border4
		Sheet1.cell(row=60, column=7).border=border4
		Sheet1.cell(row=61, column=7).border=border4
		Sheet1.cell(row=62, column=7).border=border4
		Sheet1.cell(row=63, column=7).border=border4
		Sheet1.cell(row=64, column=7).border=border4
		Sheet1.cell(row=65, column=7).border=border4
		Sheet1.cell(row=66, column=7).border=border4
		Sheet1.cell(row=67, column=7).border=border4
		Sheet1.cell(row=68, column=7).border=border4
		Sheet1.cell(row=69, column=7).border=border4
		Sheet1.cell(row=70, column=7).border=border4
		Sheet1.cell(row=71, column=7).border=border4
		Sheet1.cell(row=72, column=7).border=border4
		Sheet1.cell(row=73, column=7).border=border4
		Sheet1.cell(row=74, column=7).border=border4
		

		Sheet1.cell(row=14, column=3).value="Summary"
		Sheet1.cell(row=14, column=3).font=font4
		Sheet1.cell(row=58, column=3).value="PAYMENT ORDER"
		Sheet1.cell(row=58, column=3).font=font5

		for row in Sheet1['A1:N100']:
					for cell in row:
						cell.fill = fonta

		for row in Sheet1['N1:Z100']:
					for cell in row:
						cell.fill = fontg


		Sheet1.cell(row = 6, column = 3).value = "Company"
		Sheet1.cell(row = 6, column = 4).value = "='Other info'!C4"
		Sheet1.cell(row = 7, column = 3).value = "Address"
		Sheet1.cell(row = 7, column = 4).value = "='Other info'!C6"
		Sheet1.cell(row = 8, column = 3).value = "VAT tax code"
		Sheet1.cell(row = 8, column = 4).value = "='Other info'!C5"
		Sheet1['D8'].alignment = Alignment(wrapText=True, horizontal='left')
		Sheet1.cell(row = 9, column = 3).value = "Registration no."
		Sheet1.cell(row = 9, column = 4).value = "J08/1139/2017"
		Sheet1.cell(row = 10, column = 3).value = "The administration it belongs to"
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		Sheet1.cell(row = 11, column = 3).value = "Frequency of declaration / payment"
		Sheet1.cell(row = 11, column = 4).value = "Monthly"
		Sheet1.cell(row = 12, column = 3).value = "Reporting period"
		Sheet1.cell(row = 12, column = 4).value = "=date('Other info'!C2,'Other info'!C3,1)"
		Sheet1.cell(row = 12, column = 4).number_format = 'mmmm yyyy'

		Sheet1.cell(row = 43, column = 4).value="Yes"
		Sheet1.cell(row = 46, column = 4).value="N/a"
		Sheet1.cell(row = 47, column = 4).value="N/a"
		Sheet1.cell(row = 50, column = 4).value="N/a"
		Sheet1.cell(row = 51, column = 4).value="N/a"
		Sheet1.cell(row = 43, column = 4).font=font5
		Sheet1.cell(row = 46, column = 4).font=font5
		Sheet1.cell(row = 47, column = 4).font=font5
		Sheet1.cell(row = 50, column = 4).font=font5
		Sheet1.cell(row = 51, column = 4).font=font5
		info=temp['Other info']
		valluna=""
		vallunaurmatoare=""
		valIban=""
		okdecembrie=0
		for i in range(0,len(listaluni)):
			if(str(info.cell(row=3,column=3).value)=="12"):
				okdecembrie=1
				vallunaurmatoare=listadenluni2[0]
				valluna=listadenluni2[11]
			else:
				if(listaluni[i]==str(info.cell(row=3,column=3).value)):
					valluna=listadenluni2[i]
					vallunaurmatoare=listadenluni2[i+1]
		var=Sheet1.cell(row=12,column=4).value
		print(var)
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		for j in range(0,len(listatrez)):
			if(listatrez[j]==str(Sheet1.cell(row=10,column=4).value)):
				valIban=listaiban[j]
		Sheet1.cell(row=60, column=3).value='Payment period: '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=61, column=3).value='="Suma de plata: " &D55&" RON "'
		Sheet1.cell(row=62, column=3).value="Currency: RON"
		Sheet1.cell(row=63, column=3).value='Payment details: VAT return - '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		if(okdecembrie==1):
			Sheet1.cell(row=64, column=3).value='Deadline: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value+1)
		else:
			Sheet1.cell(row=64, column=3).value='Deadline: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=66, column=3).value='="Payer TIN: " & D8'
		Sheet1.cell(row=67, column=3).value='="Payer address: " &D7'
		Sheet1.cell(row=69, column=3).value="Beneficiary: BUGETUL DE STAT"
		Sheet1.cell(row=70, column=3).value='IBAN: '+ str(valIban)
		Sheet1.cell(row=71, column=3).value="SWIFT / BIC: TREZROBU"
		Sheet1.cell(row=72, column=3).value="Bank / Treasury:"+str(dropdown)
		Sheet1.cell(row=74, column=3).value="Note: Any banking fees connected with the payment must be covered by the tax payer."
		Sheet1.cell(row=74, column=3).font=font2

		Sheet1.cell(row = 6, column = 3).font=font1
		Sheet1.cell(row = 6, column = 4).font=font2
		Sheet1.cell(row = 7, column = 3).font=font1
		Sheet1.cell(row = 7, column = 4).font=font2
		Sheet1.cell(row = 8, column = 3).font=font1
		Sheet1.cell(row = 8, column = 4).font=font2
		Sheet1.cell(row = 9, column = 3).font=font1
		Sheet1.cell(row = 9, column = 4).font=font2
		Sheet1.cell(row = 10, column = 3).font=font1
		Sheet1.cell(row = 10, column = 4).font=font3
		Sheet1.cell(row = 11, column = 3).font=font1
		Sheet1.cell(row = 11, column = 4).font=font3
		Sheet1.cell(row = 12, column = 3).font=font1
		Sheet1.cell(row = 12, column = 4).font=font3

		Sheet1.cell(row = 6, column = 3).fill=culoare
		Sheet1.cell(row = 7, column = 3).fill=culoare
		Sheet1.cell(row = 8, column = 3).fill=culoare
		Sheet1.cell(row = 9, column = 3).fill=culoare
		Sheet1.cell(row = 10, column = 3).fill=culoare
		Sheet1.cell(row = 11, column = 3).fill=culoare
		Sheet1.cell(row = 12, column = 3).fill=culoare
		Sheet1.cell(row = 6, column = 4).fill=culoare2
		Sheet1.cell(row = 7, column = 4).fill=culoare2
		Sheet1.cell(row = 8, column = 4).fill=culoare2
		Sheet1.cell(row = 9, column = 4).fill=culoare2
		Sheet1.cell(row = 10, column = 4).fill=culoare2
		Sheet1.cell(row = 11, column = 4).fill=culoare2
		Sheet1.cell(row = 12, column = 4).fill=culoare2

		Sheet1.cell(row = 58, column = 3).fill=culoare6
		Sheet1.cell(row = 32, column = 3).fill=culoare3
		Sheet1.cell(row = 33, column = 3).fill=culoare3
		Sheet1.cell(row = 34, column = 3).fill=culoare3
		Sheet1.cell(row = 35, column = 3).fill=culoare3
		Sheet1.cell(row = 36, column = 3).fill=culoare3
		Sheet1.cell(row = 37, column = 3).fill=culoare3
		Sheet1.cell(row = 38, column = 3).fill=culoare3
		Sheet1.cell(row = 39, column = 3).fill=culoare3
		Sheet1.cell(row = 32, column = 4).fill=culoare3
		Sheet1.cell(row = 33, column = 4).fill=culoare3
		Sheet1.cell(row = 34, column = 4).fill=culoare3
		Sheet1.cell(row = 35, column = 4).fill=culoare3
		Sheet1.cell(row = 36, column = 4).fill=culoare3
		Sheet1.cell(row = 37, column = 4).fill=culoare3
		Sheet1.cell(row = 38, column = 4).fill=culoare3
		Sheet1.cell(row = 39, column = 4).fill=culoare3
		Sheet1.cell(row = 32, column = 6).fill=culoare3
		Sheet1.cell(row = 33, column = 6).fill=culoare3
		Sheet1.cell(row = 34, column = 6).fill=culoare3
		Sheet1.cell(row = 35, column = 6).fill=culoare3
		Sheet1.cell(row = 36, column = 6).fill=culoare3
		Sheet1.cell(row = 37, column = 6).fill=culoare3
		Sheet1.cell(row = 38, column = 6).fill=culoare3
		Sheet1.cell(row = 39, column = 6).fill=culoare3

		Sheet1.cell(row = 41, column = 6).fill=culoare4
		Sheet1.cell(row = 42, column = 6).fill=culoare4
		Sheet1.cell(row = 43, column = 6).fill=culoare4
		Sheet1.cell(row = 44, column = 6).fill=culoare4
		Sheet1.cell(row = 45, column = 6).fill=culoare4
		Sheet1.cell(row = 46, column = 6).fill=culoare4
		Sheet1.cell(row = 47, column = 6).fill=culoare4
		Sheet1.cell(row = 48, column = 6).fill=culoare4
		Sheet1.cell(row = 49, column = 6).fill=culoare4
		Sheet1.cell(row = 50, column = 6).fill=culoare4
		Sheet1.cell(row = 51, column = 6).fill=culoare4
		Sheet1.cell(row = 41, column = 3).fill=culoare4
		Sheet1.cell(row = 42, column = 3).fill=culoare4
		Sheet1.cell(row = 43, column = 3).fill=culoare4
		Sheet1.cell(row = 44, column = 3).fill=culoare4
		Sheet1.cell(row = 45, column = 3).fill=culoare4
		Sheet1.cell(row = 46, column = 3).fill=culoare4
		Sheet1.cell(row = 47, column = 3).fill=culoare4
		Sheet1.cell(row = 48, column = 3).fill=culoare4
		Sheet1.cell(row = 49, column = 3).fill=culoare4
		Sheet1.cell(row = 50, column = 3).fill=culoare4
		Sheet1.cell(row = 51, column = 3).fill=culoare4
		Sheet1.cell(row = 41, column = 4).fill=culoare4
		Sheet1.cell(row = 42, column = 4).fill=culoare4
		Sheet1.cell(row = 43, column = 4).fill=culoare6
		Sheet1.cell(row = 44, column = 4).fill=culoare4
		Sheet1.cell(row = 45, column = 4).fill=culoare4
		Sheet1.cell(row = 46, column = 4).fill=culoare6
		Sheet1.cell(row = 47, column = 4).fill=culoare6
		Sheet1.cell(row = 48, column = 4).fill=culoare4
		Sheet1.cell(row = 49, column = 4).fill=culoare4
		Sheet1.cell(row = 50, column = 4).fill=culoare6
		Sheet1.cell(row = 51, column = 4).fill=culoare6

		Sheet1.cell(row = 53, column = 4).fill=culoare5
		Sheet1.cell(row = 54, column = 4).fill=culoare5
		Sheet1.cell(row = 55, column = 4).fill=culoare5
		Sheet1.cell(row = 53, column = 3).fill=culoare5
		Sheet1.cell(row = 54, column = 3).fill=culoare5
		Sheet1.cell(row = 55, column = 3).fill=culoare5
		Sheet1.cell(row = 53, column = 6).fill=culoare5
		Sheet1.cell(row = 54, column = 6).fill=culoare5
		Sheet1.cell(row = 55, column = 6).fill=culoare5


		Sheet1.cell(row = 32, column = 3).value="Current VAT position"
		Sheet1.cell(row = 32, column = 3).font=font4
		Sheet1.cell(row = 34, column = 3).value="Input VAT for the period"
		Sheet1.cell(row = 35, column = 3).value="Output VAT for the period"
		Sheet1.cell(row = 36, column = 3).value="VAT Payable for the period"
		Sheet1.cell(row = 37, column = 3).value="VAT Recoverable for the period"
		Sheet1.cell(row = 38, column = 3).value="VAT under settlement for purchases"
		Sheet1.cell(row = 39, column = 3).value="VAT under settlement for deliveries"
		Sheet1.cell(row = 32, column = 4).value="RON"

		Sheet1.cell(row = 32, column = 6).value="Euro(@"+get_fxrate(2022)+")"
		Sheet1.cell(row = 34, column = 4).value="='D300 draft figures'!C56"
		Sheet1.cell(row = 34, column = 6).value="=IFERROR(D34/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=34, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=34, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row = 35, column = 4).value="='D300 draft figures'!C32"
		Sheet1.cell(row = 35, column = 6).value="=IFERROR(D35/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 36, column = 4).value="=IF('D300 draft figures'!C58<>0,'D300 draft figures'!C58,0)"
		Sheet1.cell(row = 36, column = 6).value="=IFERROR(D36/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=35, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=35, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row = 37, column = 4).value='''=IF('D300 draft figures'!C57<>0,'D300 draft figures'!C57,"nil")'''
		Sheet1.cell(row = 37, column = 6).value="=iferror(D37/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 38, column = 4).value="='D300 draft figures'!C70"
		Sheet1.cell(row = 38, column = 6).value="=iferror(D38/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 39, column = 4).value=0
		Sheet1.cell(row = 39, column = 6).value="=iferror(D39/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 55, column = 4).value='''=IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)=0),'Cover sheet'!D36,
IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)<>0),IF('Cover sheet'!D36>IFERROR(VALUE('Cover sheet'!D47),0),'Cover sheet'!D36-IFERROR(VALUE('Cover sheet'!D47),0),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="No"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D47),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="Yes"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D51),0),"N/A"))))'''
		Sheet1.cell(row = 55, column = 6).value="=iferror(D55/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=37, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=37, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=6).number_format = '#,##0_);(#,##0)'


		Sheet1.cell(row = 41, column = 3).value="Carry Over position"
		Sheet1.cell(row = 41, column = 3).font=font4
		Sheet1.row_dimensions[42].height = 0.2
		Sheet1.cell(row = 43, column = 3).value="Requested for reimbursement"
		Sheet1.cell(row = 44, column = 3).value
		Sheet1.cell(row = 45, column = 3).value="VAT refundable not yet requested"
		Sheet1.cell(row = 46, column = 3).value="Period"
		Sheet1.cell(row = 47, column = 3).value="Amount"
		Sheet1.cell(row = 48, column = 3).value
		Sheet1.cell(row = 49, column = 3).value="VAT refundable requested and under audit"
		Sheet1.cell(row = 50, column = 3).value="Period"
		Sheet1.cell(row = 51, column = 3).value="Amount"


		Sheet1.cell(row = 53, column = 3).value="VAT balance position"
		Sheet1.cell(row = 53, column = 3).font=font4
		Sheet1.cell(row = 55, column = 3).value="VAT position in the current return"


		Sheet1['C16'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C19'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C22'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C25'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C28'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['D10'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D11'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D12'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['C58'].alignment = Alignment(wrapText=True, horizontal='center')
		Sheet1['D37'].alignment = Alignment(wrapText=True, horizontal='right')

		Sheet1.column_dimensions['C'].width = 65
		Sheet1.column_dimensions['D'].width = 20
		Sheet1.column_dimensions['A'].width = 2
		Sheet1.column_dimensions['B'].width = 3
		Sheet1.column_dimensions['E'].width = 1
		Sheet1.column_dimensions['F'].width = 20
		Sheet1.column_dimensions['G'].width = 2

		# img= openpyxl.drawing.image.Image('test.png')
		# Sheet1.add_image(img,'C16')

		# img= openpyxl.drawing.image.Image('test2.png')
		# Sheet1.add_image(img,'C19')

		# img= openpyxl.drawing.image.Image('test3.png')
		# Sheet1.add_image(img,'C22')

		# img= openpyxl.drawing.image.Image('test4.png')
		# Sheet1.add_image(img,'C25')

		# img= openpyxl.drawing.image.Image('test5.png')
		# Sheet1.add_image(img,'C28')

		# img= openpyxl.drawing.image.Image('test6.png')
		# Sheet1.add_image(img,'D16')

		# img= openpyxl.drawing.image.Image('test7.png')
		# Sheet1.add_image(img,'D19')

		# img= openpyxl.drawing.image.Image('test8.png')
		# Sheet1.add_image(img,'D22')

		# img= openpyxl.drawing.image.Image('test9.png')
		# Sheet1.add_image(img,'D25')
		Sheet1.cell(row = 16, column = 3).fill=culoare
		Sheet1.cell(row=16, column=3).font=font1
		Sheet1.cell(row=16, column=3).value="  D300"
		Sheet1.cell(row=16, column=3).hyperlink="#'D300 draft figures'!A1"
		# Sheet1.cell(row=16, column=3).border=border9
		Sheet1.row_dimensions[18].height=8

		Sheet1.cell(row = 19, column = 3).fill=culoare
		Sheet1.cell(row=19, column=3).font=font1
		Sheet1.cell(row=19, column=3).value="  D390"
		Sheet1.cell(row=19, column=3).hyperlink="#'D390 workings'!A1"
		# Sheet1.cell(row=19, column=3).border=border9
		Sheet1.row_dimensions[21].height=8

		Sheet1.cell(row = 22, column = 3).fill=culoare
		Sheet1.cell(row=22, column=3).font=font1
		Sheet1.cell(row=22, column=3).value="  D394"
		Sheet1.cell(row=22, column=3).hyperlink="#'D394--->>>'!A1"
		# Sheet1.cell(row=22, column=3).border=border9
		Sheet1.row_dimensions[24].height=8

		Sheet1.cell(row = 25, column = 3).fill=culoare
		Sheet1.cell(row=25, column=3).font=font1
		Sheet1.cell(row=25, column=3).value="  Sales Ledger"
		Sheet1.cell(row=25, column=3).hyperlink="#'Sales'!A1"
		# Sheet1.cell(row=25, column=3).border=border9
		Sheet1.row_dimensions[27].height=8

		Sheet1.cell(row = 28, column = 3).fill=culoare
		Sheet1.cell(row=28, column=3).font=font1
		Sheet1.cell(row=28, column=3).value="  Purchase Ledger"
		Sheet1.cell(row=28, column=3).hyperlink="#'Purchases'!A1"
		# Sheet1.cell(row=28, column=3).border=border9
		Sheet1.row_dimensions[30].height=8

		Sheet1.merge_cells(start_row=28, start_column=3, end_row=29, end_column=3)
		Sheet1.merge_cells(start_row=25, start_column=3, end_row=26, end_column=3)
		Sheet1.merge_cells(start_row=22, start_column=3, end_row=23, end_column=3)
		Sheet1.merge_cells(start_row=19, start_column=3, end_row=20, end_column=3)
		Sheet1.merge_cells(start_row=16, start_column=3, end_row=17, end_column=3)
		Sheet1.merge_cells(start_row=6, start_column=4, end_row=6, end_column=12)
		Sheet1.merge_cells(start_row=7, start_column=4, end_row=7, end_column=12)
		Sheet1.merge_cells(start_row=8, start_column=4, end_row=8, end_column=12)
		Sheet1.merge_cells(start_row=9, start_column=4, end_row=9, end_column=12)
		Sheet1.merge_cells(start_row=10, start_column=4, end_row=10, end_column=12)
		Sheet1.merge_cells(start_row=11, start_column=4, end_row=11, end_column=12)
		Sheet1.merge_cells(start_row=12, start_column=4, end_row=12, end_column=12)
		Sheet1.merge_cells(start_row=58, start_column=3, end_row=58, end_column=6)
	
	sales=temp['Sales']
	purchases=temp['Purchases']
	if(val1==1):
		sheetinutil1=temp.create_sheet('D300--->>>')
		sheetinutil1.sheet_view.showGridLines=False
		sheetinutil1.cell(row=2,column=1).value="Switch to next sheet for D300 Workings draft"
		sheetinutil1.cell(row=2,column=1).font=scrisincredibildemare
		amount=temp.create_sheet('D300 draft figures')
		amount.freeze_panes = 'A8'
		amount.auto_filter.ref = "A7:G71"
		amount.sheet_view.showGridLines = False
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Tax code":
					rand_tb = cell.row
					taxcodec = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in sales[taxcodec][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Tax code sales'")
			return render_template("index.html")
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "  Total doc.incl.VAT":
					rand_tb = cell.row
					tdocc = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in sales[tdocc][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Total sales'")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == " Doc. Date":
					rand_tb = cell.row
					tdat = cell.column
					lun = len(sales[cell.column])
		try:
			listdocdate = [b.value for b in sales[tdat][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Total sales'")
			return render_template("index.html")
		listacurentas=[]						
		for k in range(0,len(listdocdate)):
			# print(datadocument[k][3:4])
			# print(datadocument[k][3:5])
			if(str(listdocdate[k][4:5])=="0"):
				if(str(listdocdate[k][5:6])==str(info.cell(row=3,column=3).value)):
					listacurentas.append("Yes")
				else:
					listacurentas.append("No")

			else:
				if(str(listdocdate[k][4:6])==str(info.cell(row=3,column=3).value)):
					listacurentas.append("Yes")
				else:
					listacurentas.append("No")
		for kk in range(0,len(listacurentas)):
			sales.cell(row=2+kk,column=70).value=listacurentas[kk]

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == " Doc. Date":
					rand_tb = cell.row
					supplierCell = cell.column
					lun = len(purchases[cell.column])
		try:
			datadocument = [b.value for b in purchases[supplierCell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for ' Doc. Date' in Purchases sheet")
			return render_template("index.html")
		lunacurenta=[]
		for k in range(0,len(datadocument)):
			try:
				print(datadocument[k][3:4])
				print(datadocument[k][3:5])
				if(str(datadocument[k][3:4])=="0"):
					if(str(datadocument[k][4:5])==str(info.cell(row=3,column=3).value)):
						lunacurenta.append("Yes")
					else:
						lunacurenta.append("No")

				else:
					if(str(datadocument[k][3:5])==str(info.cell(row=3,column=3).value)):
						lunacurenta.append("Yes")
					else:
						lunacurenta.append("No")
			except:
				lunacurenta.append("Not applicable")
			# if(datadocument[k])
		for kk in range(0,len(lunacurenta)):
			purchases.cell(row=2+kk,column=70).value=lunacurenta[kk]			

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Tax code":
					rand_tb = cell.row
					taxcodea = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in purchases[taxcodea][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Tax code purchases'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "19%taxbase":
					rand_tb = cell.row
					tax19b = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in purchases[taxcodea][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Tax code purchases'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "19%taxamount":
					rand_tb = cell.row
					tax19vat = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[taxcodea][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Tax code purchases'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "9%taxbase":
					rand_tb = cell.row
					tax9b = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in purchases[taxcodea][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Tax code purchases'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "9%taxamount":
					rand_tb = cell.row
					tax9vat = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[taxcodea][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Tax code purchases'")
			return render_template("index.html")			

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "5%taxbase":
					rand_tb = cell.row
					tax5b = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in purchases[taxcodea][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Tax code purchases'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "5%taxamount":
					rand_tb = cell.row
					tax5vat = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[taxcodea][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Tax code purchases'")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Intra-commtaxbase":
					rand_tb = cell.row
					intracome1 = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[taxcodea][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Tax code purchases'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Intra-comtaxamount":
					rand_tb = cell.row
					intracomtaxe1 = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[taxcodea][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Tax code purchases'")
			return render_template("index.html")			

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TotaldocinclVAT":
					rand_tb = cell.row
					tdoca = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in purchases[tdoca][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Total purchases'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Nedeductibil":
					rand_tb = cell.row
					tdocneded = cell.column
					lun = len(sales[cell.column])
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Neexigibil BAZA 19%":
					rand_tb = cell.row
					tdocnexb = cell.column
					lun = len(sales[cell.column])
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Neexigibil TVA 19%":
					rand_tb = cell.row
					tdocnextva = cell.column
					lun = len(sales[cell.column])										
		try:
			listBazaL = [b.value for b in purchases[tdoca][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Total purchases'")
			return render_template("index.html")			
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Revtaxbase-art150":
					rand_tb = cell.row
					tdocsapte = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in purchases[tdoca][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Total purchases'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Revtaxamnt-art150":
					rand_tb = cell.row
					tdocvatsapte = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in purchases[tdoca][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Total purchases'")
			return render_template("index.html")
		if(option==1):	
			for po in range(0,len(listadend300ro)):
				amount.cell(row=po+8,column=8).value=listadend300ro[po]	
		else:
			for po in range(0,len(listadend300eng)):
				amount.cell(row=po+8,column=8).value=listadend300eng[po]

		amount.cell(row=6, column=2).value="1"
		amount.cell(row=6, column=3).value="2"
		amount.cell(row=7, column=1).value="Row"
		amount.cell(row=7, column=2).value="Taxable basis"
		amount.cell(row=7, column=3).value="VAT amount"
		amount.cell(row=7, column=4).value="Comments"
		amount.cell(row=7, column=5).value="Journal Source"
		amount.cell(row=7, column=6).value="Flag Suma Control"
		amount.cell(row=7, column=7).value="Suma Control"
		amount.cell(row=8, column=1).value="1"
		amount.cell(row=9, column=1).value="2"
		amount.cell(row=10, column=1).value="3"
		amount.cell(row=11, column=1).value="3.1"
		amount.cell(row=12, column=1).value="4"
		amount.cell(row=13, column=1).value="5"
		amount.cell(row=14, column=1).value="5.1"
		amount.cell(row=15, column=1).value="6"
		amount.cell(row=16, column=1).value="7"
		amount.cell(row=17, column=1).value="7.1"
		amount.cell(row=18, column=1).value="8"
		amount.cell(row=19, column=1).value="9"
		amount.cell(row=20, column=1).value="10"
		amount.cell(row=21, column=1).value="11"
		amount.cell(row=22, column=1).value="12"
		amount.cell(row=23, column=1).value="12.1"
		amount.cell(row=24, column=1).value="12.2"
		amount.cell(row=25, column=1).value="12.3"
		amount.cell(row=26, column=1).value="13"
		amount.cell(row=27, column=1).value="14"
		amount.cell(row=28, column=1).value="15"
		amount.cell(row=29, column=1).value="16"
		amount.cell(row=30, column=1).value="17"
		amount.cell(row=31, column=1).value="18"
		amount.cell(row=32, column=1).value="19"
		amount.cell(row=33, column=1).value="20"
		amount.cell(row=34, column=1).value="20.1"
		amount.cell(row=35, column=1).value="21"
		amount.cell(row=36, column=1).value="22"
		amount.cell(row=37, column=1).value="22.1"
		amount.cell(row=38, column=1).value="23"
		amount.cell(row=39, column=1).value="24"
		amount.cell(row=40, column=1).value="25"
		amount.cell(row=41, column=1).value="26"
		amount.cell(row=42, column=1).value="27"
		amount.cell(row=43, column=1).value="27.1"
		amount.cell(row=44, column=1).value="27.2"
		amount.cell(row=45, column=1).value="27.3"
		amount.cell(row=46, column=1).value="28"
		amount.cell(row=47, column=1).value="29"
		amount.cell(row=48, column=1).value="30"
		amount.cell(row=49, column=1).value="30.1"
		amount.cell(row=50, column=1).value="31"
		amount.cell(row=51, column=1).value="31.1"
		amount.cell(row=52, column=1).value="32"
		amount.cell(row=53, column=1).value="33"
		amount.cell(row=54, column=1).value="34"
		amount.cell(row=55, column=1).value="35"
		amount.cell(row=56, column=1).value="36"
		amount.cell(row=57, column=1).value="37"
		amount.cell(row=58, column=1).value="38"
		amount.cell(row=59, column=1).value="39"
		amount.cell(row=60, column=1).value="40"
		amount.cell(row=61, column=1).value="41"
		amount.cell(row=62, column=1).value="42"
		amount.cell(row=63, column=1).value="43"
		amount.cell(row=64, column=1).value="44"
		amount.cell(row=65, column=1).value="45"
		amount.cell(row=66, column=1).value="46"

		amount.cell(row=68, column=1).value="A"
		amount.cell(row=69, column=1).value="A1"
		amount.cell(row=70, column=1).value="B"
		amount.cell(row=71, column=1).value="B.1"


		amount.cell(row=8, column=2).value='=ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"Y1",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)+ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"Y3",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)'
		amount.cell(row=9, column=2).value='=ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"C2",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)'
		amount.cell(row=10, column=2).value='=ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"Y4",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)'		
		amount.cell(row=11, column=2).value='=ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"Y4",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)'
		amount.cell(row=12, column=2).value=0
		amount.cell(row=13, column=2).value='=(ROUND(SUMIFS(Purchases!'+str(intracome1)+":"+str(intracome1)+',Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"E1",Purchases!BR:BR,"Yes"),0))'
		amount.cell(row=14, column=2).value='=(ROUND(SUMIFS(Purchases!'+str(intracome1)+":"+str(intracome1)+',Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"E1",Purchases!BR:BR,"Yes"),0))'
		amount.cell(row=15, column=2).value='=(ROUND(SUMIFS(Purchases!'+str(tdoca)+":"+str(tdoca)+',Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"E1",Purchases!BR:BR,"No")/(119/100),0))'
		amount.cell(row=16, column=2).value='=(ROUND(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tdocsapte)+":"+str(tdocsapte)+')+SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"1M",Purchases!'+str(tax19b)+":"+str(tax19b)+'),0))'
		amount.cell(row=17, column=2).value='=(ROUND(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tdocsapte)+":"+str(tdocsapte)+'),0))-sumif(Purchases!'+str(taxcodea)+':'+str(taxcodea)+',"1M",Purchases!'+str(tdocsapte)+":"+str(tdocsapte)+')'	
		amount.cell(row=18, column=2).value='=(ROUND(SUMIFS(Purchases!'+str(tdoca)+":"+str(tdoca)+',Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"X1",Purchases!BR:BR,"No")/(119/100),0))'
		amount.cell(row=19, column=2).value='=round(ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"A1",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)/(119/100),0)'
		amount.cell(row=20, column=2).value=0
		amount.cell(row=21, column=2).value='=round(ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"5G",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)/(105/100),0)'
		# amount.cell(row=22, column=2).value='=ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"ZR",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)'
		amount.cell(row=23, column=2).value='=ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"ZR",Purchases!'+str(tdocc)+":"+str(tdocc)+'),0)'
		amount.cell(row=24, column=2).value=0
		amount.cell(row=25, column=2).value=0
		amount.cell(row=26, column=2).value='=ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"1V",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)'	
		amount.cell(row=27, column=2).value='=ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"A5",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)+ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"A4",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)+ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"Y8",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)'
		amount.cell(row=28, column=2).value=0
		amount.cell(row=30, column=2).value=0
		amount.cell(row=29, column=2).value='=ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"ZJ",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)'
		amount.cell(row=39, column=2).value='=round(SUM(Purchases!'+str(tax19b)+':'+str(tax19b)+')-SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax19b)+':'+str(tax19b)+')-SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"5Z",Purchases!'+str(tax19b)+":"+str(tax19b)+')-SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"1M",Purchases!'+str(tax19b)+":"+str(tax19b)+')+SUMIFS(Purchases!'+str(tax19b)+':'+str(tax19b)+',Purchases!'+str(taxcodea)+':'+str(taxcodea)+',"ZI",Purchases!BR:BR,"No"),0)'						
		amount.cell(row=40, column=2).value='=round((ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"V3",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)+ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"9S",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)+ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"W6",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0))/(109/100),0)'								
		amount.cell(row=41, column=2).value='=round((ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"W8",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)+ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"5S",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)+ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"6I",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0))/(105/100),0)'								

		
		amount.cell(row=22, column=2).value='=round(SUM(B23:B25),0)'


		amount.cell(row=31, column=2).value=0
		amount.cell(row=32, column=2).value='=B8+B10+B13+B16+B27+B15+B18+B31+B30+B29+B28+B26+B22+B21+B20+B19'
		amount.cell(row=33, column=2).value='=B13'
		amount.cell(row=34, column=2).value='=B14'
		amount.cell(row=35, column=2).value='=B15'
		amount.cell(row=36, column=2).value='=B16'
		amount.cell(row=37, column=2).value='=B17'
		amount.cell(row=38, column=2).value='=B18'

		
		amount.cell(row=42, column=2).value='=SUM(B43:B45)'
		amount.cell(row=43, column=2).value='=B23'
		amount.cell(row=44, column=2).value='=B24'
		amount.cell(row=45, column=2).value='=B25'
		amount.cell(row=46, column=2).value=0
		amount.cell(row=47, column=2).value=0
		amount.cell(row=48, column=2).value='=ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"V9",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)+ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"W0",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)+ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"AS",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)+ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"AS",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)'
		amount.cell(row=49, column=2).value=0
		amount.cell(row=50, column=2).value='=SUM(B33:B47)-B34-B37-SUM(B43:B45)'
		amount.cell(row=51, column=2).value='=round(SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax19b)+':'+str(tax19b)+')+SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax9b)+':'+str(tax9b)+')+SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax5b)+':'+str(tax5b)+')-SUMIFS(Purchases!'+str(tax19b)+':'+str(tax19b)+',Purchases!'+str(taxcodea)+':'+str(taxcodea)+',"ZI",Purchases!BR:BR,"No"),0)'
		amount.cell(row=52, column=2).value='=B50+B51'
		amount.cell(row=53, column=2).value=0
		amount.cell(row=54, column=2).value=0
		amount.cell(row=55, column=2).value=0
		amount.cell(row=56, column=2).value='=SUM(B52:B55)'
		amount.cell(row=57, column=2).value=0
		amount.cell(row=58, column=2).value=0
		amount.cell(row=59, column=2).value=0
		amount.cell(row=60, column=2).value=0
		amount.cell(row=61, column=2).value='=SUM(B58:B60)'
		amount.cell(row=62, column=2).value=0
		amount.cell(row=63, column=2).value=0
		amount.cell(row=64, column=2).value='=B57+B62+B63'
		amount.cell(row=65, column=2).value='=IF((B61-B664)<0,0,B61-B64)'
		amount.cell(row=66, column=2).value='=IF((B64-B61)<0,0,B64)'

		amount.cell(row=68, column=2).value=0
		amount.cell(row=69, column=2).value=0
		amount.cell(row=70, column=2).value='=ROUND(SUM(Purchases!'+str(tdocnexb)+":"+str(tdocnexb)+',),0)'
		amount.cell(row=71, column=2).value='=B70'
		
		#coloana TVA----------------------------------------------------

		for g in range(8, 13):
			amount.cell(row=g, column=3).value=0
		

		# for h in range(13, 19):
		amount.cell(row=13, column=3).value='=(ROUND(SUMIFS(Purchases!'+str(intracomtaxe1)+":"+str(intracomtaxe1)+',Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"E1",Purchases!BR:BR,"Yes"),0))'
		amount.cell(row=14, column=3).value='=(ROUND(SUMIFS(Purchases!'+str(intracomtaxe1)+":"+str(intracomtaxe1)+',Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"E1",Purchases!BR:BR,"Yes"),0))'
		amount.cell(row=15, column=3).value='=round(B15/100*19,0)'
		amount.cell(row=16, column=3).value='=(ROUND(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tdocvatsapte)+':'+str(tdocvatsapte)+')+SUMIF(Purchases!'+str(taxcodea)+':'+str(taxcodea)+',"1M",Purchases!'+str(tax19vat)+':'+str(tax19vat)+'),0))'
		amount.cell(row=17, column=3).value='=(ROUND(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tdocvatsapte)+":"+str(tdocvatsapte)+'),0))-sumif(Purchases!'+str(taxcodea)+':'+str(taxcodea)+',"1M",Purchases!'+str(tdocvatsapte)+":"+str(tdocvatsapte)+')'
		amount.cell(row=18, column=3).value='=round(B18/100*19,0)'
			
		# amount.cell(row=16,column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A16&"."&C$6,Purchases!$5:$5)-SUMIF(Purchases!$7:$7,$A18&"."&C$6,Purchases!$5:$5),0)'


		amount.cell(row=19, column=3).value='=round(B19/100*19,0)'
		amount.cell(row=20, column=3).value='=round(B20/100*9,0)'
		amount.cell(row=21, column=3).value='=round(B21/100*5,0)'
		amount.cell(row=22, column=3).value='=SUM(C23:C25)'

		amount.cell(row=23, column=3).value='=round(B23/100*19,0)'
		amount.cell(row=24, column=3).value='=round(B24/100*9,0)'
		amount.cell(row=25, column=3).value='=round(B25/100*5,0)'


		
		for k in range(26, 31):
			amount.cell(row=k, column=3).value=0
		amount.cell(row=31, column=3).value=0
		amount.cell(row=32, column=3).value='=C8+C10+C13+C16+C27+C15+C18+C31+C30+C29+C28+C26+C22+C21+C20+C19'
		amount.cell(row=33, column=3).value='=C13'
		amount.cell(row=34, column=3).value='=C14'
		amount.cell(row=35, column=3).value='=C15'
		amount.cell(row=36, column=3).value='=C16'
		amount.cell(row=37, column=3).value='=C17'
		amount.cell(row=38, column=3).value='=C18'


		amount.cell(row=39, column=3).value='=round(SUM(Purchases!'+str(tax19vat)+':'+str(tax19vat)+')-SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax19vat)+':'+str(tax19vat)+')-SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"5Z",Purchases!'+str(tax19vat)+":"+str(tax19vat)+')-SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"1M",Purchases!'+str(tax19vat)+":"+str(tax19vat)+')+SUMIFS(Purchases!'+str(tax19vat)+':'+str(tax19vat)+',Purchases!'+str(taxcodea)+':'+str(taxcodea)+',"ZI",Purchases!BR:BR,"No"),0)'
		amount.cell(row=40, column=3).value='=round(B40/100*9,0)'
		amount.cell(row=41, column=3).value='=round(B41/100*5,0)'			
		amount.cell(row=42, column=3).value='=round(SUM(C43:C45),0)'
		amount.cell(row=43, column=3).value='=C23'
		amount.cell(row=44, column=3).value='=C24'
		amount.cell(row=45, column=3).value='=C25'
		amount.cell(row=46, column=3).value=0
		amount.cell(row=47, column=3).value=0
		amount.cell(row=48, column=3).value=0
		amount.cell(row=49, column=3).value=0
		amount.cell(row=50, column=3).value='=SUM(C33:C47)-C34-C37-SUM(C43:C45)'
		amount.cell(row=51, column=3).value='=Round(SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax19vat)+':'+str(tax19vat)+')+SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax9vat)+':'+str(tax9vat)+')+SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax5vat)+':'+str(tax5vat)+')-SUMIFS(Purchases!'+str(tax19vat)+':'+str(tax19vat)+',Purchases!'+str(taxcodea)+':'+str(taxcodea)+',"ZI",Purchases!BR:BR,"No"),0)'
		amount.cell(row=52, column=3).value='=C50+C51-SUM(Purchases!'+str(tdocneded)+':'+str(tdocneded)+')'
		amount.cell(row=53, column=3).value=0
		amount.cell(row=54, column=3).value=0
		amount.cell(row=55, column=3).value=0
		amount.cell(row=56, column=3).value='=SUM(C52:C55)'
		amount.cell(row=57, column=3).value='=IF((C56-C32)<0,0,C56-C32)'
		amount.cell(row=58, column=3).value='=IF((C32-C56)<0,0,C32-C56)'
		amount.cell(row=59, column=3).value=0
		amount.cell(row=60, column=3).value=0
		amount.cell(row=61, column=3).value='=SUM(C58:C60)'	
		if soldLunaTrecuta == None or soldLunaTrecuta == "" or soldLunaTrecuta == " ":
			amount.cell(row=62, column=3).value=0
		else:
			amount.cell(row=62, column=3).value=int(soldLunaTrecuta) 
		#print(soldLunaTrecuta, "sold luna trecuta")
		amount.cell(row=63, column=3).value=0
		amount.cell(row=64, column=3).value='=C57+C62+C63'
		amount.cell(row=65, column=3).value='=IF((C61-C64)<0,0,C61-C64)'
		amount.cell(row=66, column=3).value='=IF((C64-C61)<0,0,C64-C61)'
		amount.cell(row=68, column=3).value=0
		amount.cell(row=69, column=3).value=0

		amount.cell(row=70, column=3).value='=ROUND(SUM(Purchases!'+str(tdocnextva)+":"+str(tdocnextva)+',),0)'
		amount.cell(row=71, column=3).value='=C70'

		amount.cell(row=73, column=1).value='Informații privind valoarea totală, fără TVA, a operațiunilor prevăzute la art. 2781 alin. (1) lit. b) din Codul fiscal, respectiv a vânzărilor intracomunitare de bunuri la distanță și a prestărilor de servicii de telecomunicaţii, de radiodifuziune şi televiziune, precum și servicii furnizate pe cale electronică, către persoane neimpozabile din alte state membre UE'
		amount.cell(row=73, column=2).value='Total an precedent'
		amount.cell(row=73, column=3).value='An curent (inclusiv perioada de raportare)'

		amount.cell(row=74, column=2).value=0
		amount.cell(row=74, column=3).value=0


		amount.cell(row=22, column=5).value='Total'

		for m in it.chain(range(8, 13), range(19, 22), range(26, 32)):
			amount.cell(row=m, column=5).value='SALES'

		for n in it.chain(range(13, 19), range(23, 26), range(39, 42), range(46, 50), range(53, 56)):
			amount.cell(row=n, column=5).value='Purchases'
		
		for o in range(32, 39):
			amount.cell(row=o, column=5).value='Total'

		for p in range(42, 46):
			amount.cell(row=p, column=5).value='Total'
		
		amount.cell(row=50, column=5).value='Total'
		amount.cell(row=51, column=5).value='Purchases'
		amount.cell(row=52, column=5).value='Total'
		amount.cell(row=66, column=5).value='Purchases'

		for q in range(68, 71):
			amount.cell(row=q, column=5).value='Total'

		for r in range(56, 67):
			amount.cell(row=r, column=5).value='Total'

		amount.cell(row=71, column=5).value='Purchases'

		for s in it.chain(range(8, 13), range(26, 29), range(48, 50)):
			amount.cell(row=s, column=6).value='no VAT'

		for t in it.chain(range(13, 26), range(29, 46)):
			amount.cell(row=t, column=6).value='Add all'

		for u in it.chain(range(46, 48), range(52, 54), range(55, 67)):
			amount.cell(row=u, column=6).value='No basis'
		
		amount.cell(row=50, column=6).value='Add all'
		amount.cell(row=51, column=6).value='Se pune 0'
		amount.cell(row=54, column=6).value='Add all'

		for a1 in range(8, 13):
			amount.cell(row=a1, column=7).value='=B{0}'.format(a1)

		for b1 in range(13, 26):
			amount.cell(row=b1, column=7).value='=SUM(B{0}:C{0})'.format(b1)

		for c1 in range(26, 29):
			amount.cell(row=c1, column=7).value='=B{0}'.format(c1)
		
		for d1 in range(29, 46):
			amount.cell(row=d1, column=7).value='=SUM(B{0}:C{0})'.format(d1)

		amount.cell(row=46, column=7).value='=C46'
		amount.cell(row=47, column=7).value='=C47'
		amount.cell(row=48, column=7).value='=B48'
		amount.cell(row=49, column=7).value='=B49'
		amount.cell(row=50, column=7).value='=SUM(B50:C50)'
		amount.cell(row=51, column=7).value='=C51'
		amount.cell(row=52, column=7).value='=C52'
		amount.cell(row=53, column=7).value='=SUM(B53:C53)'
		amount.cell(row=54, column=7).value='=SUM(B54:C54)'
		amount.cell(row=2,column=1).value="D300 draft figures "
		amount.cell(row=2,column=1).font=cap_tabeltitlu
		amount.row_dimensions[5].hidden = True
		for e1 in range(55, 67):
			amount.cell(row=e1, column=7).value='=C{0}'.format(e1)
		
		for row in amount['B8:B74']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'
		
		for row in amount['C8:C74']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'

		#------foratare D300

		for row in amount['A7:C7']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.alignment=Alignment(horizontal='center',vertical='center')
		for row in amount['D7:G7']:
			for cell in row:
				cell.fill=cap_tabel_color_black
		
		for row in amount['A7:G7']:
			for cell in row:
				cell.font=cap_tabel

		listanoua=['A','B','C','D','E','F','G']

		for column in listanoua:
			for i in listanoua:
				if (column==i):
					amount.column_dimensions[column].width = 17
		
		amount.column_dimensions['D'].hidden = True
		amount.column_dimensions['E'].hidden = True
		amount.column_dimensions['F'].hidden = True
		amount.column_dimensions['G'].hidden = True
		info=temp['Other info']
	
		listaMapare=["L", "T", "S", "A"]
		Poz="10"
		Fix01="01"
		Fix0000="0000"
		dictMapare={'L':'301', 'T':'302', 'S':'303', 'A':'304'}

		elementMapare=""
		if info.cell(row=53, column=3).value=="L":
			elementMapare="301"
		else:
			if info.cell(row=53, column=3).value == "T":
				elementMapare="302"
			else:
				if info.cell(row=53, column=3).value == "S":
					elementMapare="303"
				else:
					if info.cell(row=53, column=3).value == "A":
						elementMapare="304"
			
		# #print(elementMapare, 'MAPARE')

		LL=""
		AA=""

		if len(str(info.cell(row=3, column=3).value))==2:
			LL=str(info.cell(row=3, column=3).value)
			LL_g=LL
		else:
			LL="0"+str(info.cell(row=3, column=3).value)

		AA=str(info.cell(row=2, column=3).value)[2:]
		strYear=str(info.cell(row=2, column=3).value)[2:]
		intYear=int(strYear)+1
		year=str(intYear)
		
		# #print(year, 'an')

		LLAA=LL+AA
		LL2=""
		AA2=""

		if int(info.cell(row=3, column=3).value)==12:
			LL2="1"
		if len(str(info.cell(row=3, column=3).value))==1:
			if(int(info.cell(row=3,column=3).value)==9):
				LL2=str(int(info.cell(row=3,column=3).value)+1)
			else:
				LL2="0"+str(int(info.cell(row=3, column=3).value)+1)
		# #print(LL2, 'LL2')
		
		if int(info.cell(row=3, column=3).value)==12:
			AA2=year 
		else:
			AA2=strYear

		ZZLLAA="25"+str(LL2)+str(AA2)
		# #print(ZZLLAA, 'ZZLLAA')

		poz1=""
		LTSA=""
		fix1=""
		LLAA1=""
		ZZLLAA1=""
		fix01=""
		control=""

		poz1 = sum(int(digit) for digit in str(Poz))
		LTSA = sum(int(digit) for digit in str(elementMapare))
		fix1=sum(int(digit) for digit in str(Fix01))
		LLAA1=sum(int(digit) for digit in str(LLAA))
		ZZLLAA1=sum(int(digit) for digit in str(ZZLLAA))
		fix01=sum(int(digit) for digit in str(Fix0000))

		control=poz1+LTSA+fix1+LLAA1+ZZLLAA1+fix01
		

		nrEvidenta=Poz+elementMapare+Fix01+LLAA+ZZLLAA+Fix0000+str(control)
		#print(nrEvidenta, 'nrEvidenta')
			
		
		info.cell(row=18, column=3).value=nrEvidenta
		info.cell(row=52, column=3).value="=SUM('D300 draft figures'!G8:G66)"
		try:
			info.cell(row=50, column=3).value=lenbfi
			info.cell(row=51, column=3).value=sumabfi
			info.cell(row=52, column=3).value=sumatbfi
		except:
			pass
		try:
			info.cell(row=53, column=3).value=sumabif
			info.cell(row=54, column=3).value=sumatbif
			info.cell(row=55, column=3).value=lentbif
		except:
			pass
		try: 
			# info.cell(row=56, column=3).value=lenbfr
			info.cell(row=57, column=3).value=sumabfr
			info.cell(row=58, column=3).value=sumatbfr
		except:
			pass


		if(val2==1):
			sheetinutil2=temp.create_sheet('D390--->>>')
			sheetinutil2.sheet_view.showGridLines=False
			sheetinutil2.cell(row=2,column=1).value="Switch to next sheet for D390 Workings draft"
			sheetinutil2.cell(row=2,column=1).font=scrisincredibildemare		

			workings=temp.create_sheet('D390 workings')
			workings.cell(row=1,column=1).value="D390 workings"
			workings.cell(row=1,column=1).font=cap_tabelbold
			workings.freeze_panes = 'A4'
			workings.auto_filter.ref = "A3:I10000"
			workings.sheet_view.showGridLines = False
			workings.column_dimensions['I'].hidden = True
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "Tax code":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				taxcodes = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Tax code sales'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "  Total doc.incl.VAT":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(sales[cell.column])
			try:
				totals = [b.value for b in sales[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total sales'")
				return render_template("index.html")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "Business PartnerName":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(sales[cell.column])
			try:
				denumires = [b.value for b in sales[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total sales'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "VAT Registration No.":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(sales[cell.column])
			try:
				vats = [b.value for b in sales[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total sales'")
				return render_template("index.html")			

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Tax code":
						rand_tb = cell.row
						taxcodea = cell.column
						lun = len(purchases[cell.column])
			try:
				taxcodeach = [b.value for b in purchases[taxcodea][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Tax code purchases'")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Business PartnerName":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(purchases[cell.column])
			try:
				denumirea = [b.value for b in purchases[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total sales'")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "VAT Registration No.":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(purchases[cell.column])
			try:
				vata = [b.value for b in purchases[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total sales'")
				return render_template("index.html")				
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "TotaldocinclVAT":
						rand_tb = cell.row
						tdoca = cell.column
						lun = len(purchases[cell.column])
			try:
				totala = [b.value for b in purchases[tdoca][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total purchases'")
				return render_template("index.html")

			workings.cell(row=3, column=1).value='TIP'
			workings.cell(row=3, column=2).value='ŢARA'
			workings.cell(row=3, column=3).value='COD OPERATOR INTRACOMUNITAR'
			workings.cell(row=3, column=4).value='DENUMIRE'
			workings.cell(row=3, column=5).value='BAZA IMPOZABILĂ'
			workings.cell(row=3, column=6).value='CIF'
			workings.cell(row=3, column=7).value='Country Code'
			workings.cell(row=3, column=8).value='BAZA IMPOZABILĂ'
			workings.cell(row=3, column=9).value='Cheie extragere - filtreaza 1'

			# bazaA_Furnizor=get_column_letter(7)
			# bazaA_index=get_column_letter(bazaA)
			# bazaA_literaA=get_column_letter(bazaA)
			# #print(bazaA_Furnizor, "litera pentru furnizor/ client")
			# #print(bazaA_index, "ASTA E NUMARUL LUI A")
			# #print(bazaA_literaA, "LITERA PENTRU TIP")
			print(taxcodeach)
			a=3
			for x in range(0, len(taxcodeach)):
				if(lunacurenta[x]=="Yes"):
					if str(taxcodeach[x])=="C3" or str(taxcodeach[x])=="E1" :
						a=a+1
						workings.cell(row=a, column=1).value="A"
						workings.cell(row=a, column=4).value=denumirea[x]
						workings.cell(row=a, column=6).value=vata[x]
						workings.cell(row=a, column=3).value=vata[x][2:]
						workings.cell(row=a, column=7).value=vata[x][0:2]
						# workings.cell(row=a, column=8).value=listaBazaA[x]
						# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
						workings.cell(row=a, column=8).value=totala[x]
						workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
						workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)


			for x in range(0, len(taxcodeach)):
				if(lunacurenta[x]=="Yes"):				
					if str(taxcodeach[x])=="X1":
						a=a+1
						workings.cell(row=a, column=1).value="S"
						workings.cell(row=a, column=4).value=denumirea[x]
						workings.cell(row=a, column=6).value=vata[x]
						workings.cell(row=a, column=3).value=vata[x][2:]
						workings.cell(row=a, column=7).value=vata[x][0:2]
						# workings.cell(row=a, column=8).value=listaBazaA[x]
						# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
						workings.cell(row=a, column=8).value=totala[x]
						workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
						workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)

			for x in range(0, len(taxcodes)):
				if(listacurentas[x]=="Yes"):				
					if str(taxcodes[x])=="Y3" or str(taxcodes[x])=="Y1":
						a=a+1
						workings.cell(row=a, column=1).value="L"
						workings.cell(row=a, column=4).value=denumires[x]
						workings.cell(row=a, column=6).value=vats[x]
						workings.cell(row=a, column=3).value=vats[x][2:]
						workings.cell(row=a, column=7).value=vats[x][0:2]
						# workings.cell(row=a, column=8).value=listaBazaA[x]
						# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
						workings.cell(row=a, column=8).value=totals[x]
						workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
						workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)
			for x in range(0, len(taxcodes)):
				if(listacurentas[x]=="Yes"):				
					if str(taxcodes[x])=="Y4":
						a=a+1
						workings.cell(row=a, column=1).value="P"
						workings.cell(row=a, column=4).value=denumires[x]
						workings.cell(row=a, column=6).value=vats[x]
						workings.cell(row=a, column=3).value=vats[x][2:]
						workings.cell(row=a, column=7).value=vats[x][0:2]
						# workings.cell(row=a, column=8).value=listaBazaA[x]
						# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
						workings.cell(row=a, column=8).value=totals[x]
						workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
						workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)
			
			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "TIP":
						rand_tb = cell.row
						tip = cell.column
						lun = len(workings[cell.column])
			try:
				listaTip = [b.value for b in workings[tip][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TIP' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "CIF":
						rand_tb = cell.row
						cod = cell.column
						lun = len(workings[cell.column])
			try:
				codPartener = [b.value for b in workings[cod][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'CIF' in Workings sheet")
				return render_template("index.html")


			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "Country Code":
						rand_tb = cell.row
						country = cell.column
						lun = len(workings[cell.column])
			try:
				countryCode = [b.value for b in workings[country][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Country Code' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "DENUMIRE":
						rand_tb = cell.row
						numep = cell.column
						lun = len(workings[cell.column])
			try:
				partnerName = [b.value for b in workings[numep][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'DENUMIRE' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "BAZA IMPOZABILĂ":
						rand_tb = cell.row
						suma = cell.column
						lun = len(workings[cell.column])
			try:
				sumaTot = [b.value for b in workings[suma][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'BAZA IMPOZABILĂ' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "Cheie extragere - filtreaza 1":
						rand_tb = cell.row
						cheie_sort = cell.column
						lun = len(workings[cell.column])
			try:
				cheie = [b.value for b in workings[cheie_sort][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Cheie extragere - filtreaza 1' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "COD OPERATOR INTRACOMUNITAR":
						rand_tb = cell.row
						coi = cell.column
						lun = len(workings[cell.column])
			try:
				listaCOI = [b.value for b in workings[coi][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'BAZA IMPOZABILĂ' in Workings sheet")
				return render_template("index.html")


			nomenclatorTari={'AT':'Austria', 'BE':'Belgia', 'BG':'Bulgaria','CY':'Cipru','DK':'Danemarca','EE':'Estonia', 'FI':'Finlanda','FR':'Franta', 'DE':'Germania','HR':'Croatia',
							'GR':'Grecia','IE':'Irlanda','IT':'Italia','LV':'Letonia','LT':'Lituania','LU':'Luxemburg','MT':'Malta','XI':'Irlanda de Nord - Regatul Unit','NL':'Olanda',
							'PL':'Polonia','PT':'Portugalia','CZ':'Republica Ceha','RO':'Romania','SK':'Slovacia','SI':'Slovavia','ES':'Spania','SE':'Suedia','HU':'Ungaria'}
			b=3
			for i in countryCode:
				if i in nomenclatorTari:
					b=b+1
					workings.cell(row=b, column=2).value=nomenclatorTari[i]

			for row in workings['A3:I3']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')
			for row in workings['E4:E10000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'
			
			for row in workings['H4:H10000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'

			workings.column_dimensions['B'].width=20
			workings.column_dimensions['D'].width=35				

			forxml = temp.create_sheet('D390 for XML')
			forxml.cell(row=1,column=1).value="D390 for XML"
			forxml.cell(row=1,column=1).font=cap_tabelbold
			forxml.freeze_panes = 'A4'
			forxml.auto_filter.ref = "A3:F10000"
			forxml.sheet_view.showGridLines = False

			forxml.cell(row=3, column=1).value="III.B"
			forxml.cell(row=3, column=2).value="TIP"
			forxml.cell(row=3, column=3).value="ŢARA"
			forxml.cell(row=3, column=4).value="COD OPERATOR INTRACOMUNITAR"
			forxml.cell(row=3, column=5).value="Denumire"
			forxml.cell(row=3, column=6).value="BAZA IMPOZABILĂ"

			codeAndType=[]
			codeAndName=[]
			typeAndName=[]
			# typeCodeName=[]

			# for k in range(0,len(listaTip)):
			# 	codeAndType.append(str(listaTip[k])+" "+str(codPartener[k]))
			# 	codeAndName.append(str(listaTip[k])+" "+str(partnerName[k]))

			# #print(codeAndType,'codeandtyp')
			# codeAndTypeUnique=list(set(codeAndType))
			# codeAndNameUnique=list(set(codeAndName))

			for i in range(0, len(listaTip)):
				typeAndName.append(str(listaTip[i])+";;;"+str(partnerName[i])+";;;"+str(countryCode[i])+";;;"+str(listaCOI[i]))
			#print(typeAndName, 'TYPEAndNAME')

			typeAndNameUni=list(set(typeAndName))

			typeAndNameUni=list(set(typeAndName))

			for i in it.chain(range(0, len(typeAndNameUni))):
				x=typeAndNameUni[i].split(";;;")
				forxml.cell(row=4+i, column=2).value=str(x[0])
				forxml.cell(row=4+i, column=3).value=str(x[2])
				forxml.cell(row=4+i, column=4).value=str(x[3])
				forxml.cell(row=4+i, column=5).value=str(x[1])
				forxml.cell(row=4+i, column=6).value="=SUMIFS('D390 workings'!H:H,'D390 workings'!A:A,B{0},'D390 workings'!C:C,D{0},'D390 workings'!G:G,C{0})".format(4+i)


			for row in forxml['A3:F3']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in forxml['F4:F10000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'

			forxml.column_dimensions['D'].width=27
			forxml.column_dimensions['E'].width=35
			forxml.column_dimensions['F'].width=15


	#---------------------------NR DE EVIDENTA
		if(val3==1):
			sheetinutil3=temp.create_sheet('D394--->>>')
			sheetinutil3.sheet_view.showGridLines=False
			sheetinutil3.cell(row=2,column=1).value="Switch to next sheet for D394 Workings draft"
			sheetinutil3.cell(row=2,column=1).font=scrisincredibildemare		
			nomenclatorTari={'AT':'Austrie', 'BE':'Belgia', 'BG':'Bulgaria','CY':'Cipru','DK':'Danemarca','EE':'Estonia', 'FI':'Finlanda','FR':'Franta', 'DE':'Germania','HR':'Croatia',
							'GR':'Grecia','IE':'Irlanda','IT':'Italia','LV':'Letonia','LT':'Lituania','LU':'Luxemburg','MT':'Malta','XI':'Irlanda de Nord - Regatul Unit','NL':'Olanda',
							'PL':'Polonia','PT':'Portugalia','CZ':'Republica Ceha','RO':'Romania','SK':'Slovacia','SI':'Slovavia','ES':'Spania','SE':'Suedia','HU':'Ungaria'}


			salesExcel=temp.create_sheet("Mapping tranzactii")
			salesExcel.sheet_view.showGridLines = False
			salesExcel.cell(row=2,column=1).value="Mapping tranzactii"
			salesExcel.cell(row=2,column=1).font=cap_tabeltitlu	
			salesExcel.freeze_panes = 'A10'
			
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value=="Declarat anterior":
						rand_tb = cell.row
						declarateanteriorp = cell.column
						lun = len(purchases[cell.column])
			try:
				listadeclantp = [b.value for b in purchases[declarateanteriorp][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Declarat anterior' in Purchases sheet")
				return render_template("index.html")		
			# except:
			# 	listadeclant=[]
			listadeclantp_1=[]
			# #print(listadeclantp,"---------")
			for c in range(0, len(listadeclantp)):
				if listadeclantp[c] == None:
					listadeclantp_1.append("No")
				else:
					listadeclantp_1.append(listadeclantp[c])
			# #print("-----",listadeclantp_1,"------")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value=="Declarat anterior":
						rand_tb = cell.row
						declarateanterior = cell.column
						lun = len(sales[cell.column])
			try:
				listadeclant = [b.value for b in sales[declarateanterior][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Declarat anterior' in Sales sheet")
				return render_template("index.html")
			# except:
			# 	listadeclant=[]
			listadeclant_1=[]
			# #print(len(listadeclant))
			for c in range(0, len(listadeclant)):
				if listadeclant[c] == None:
					listadeclant_1.append("No")
				else:
					listadeclant_1.append(listadeclant[c])
			for row in sales.iter_rows():
				for cell in row:
					if cell.value=="Tax code":
						rand_tb = cell.row
						declarateanteriorp = cell.column
						lun = len(sales[cell.column])
			try:
				ltaxcode = [b.value for b in sales[declarateanteriorp][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Declarat anterior' in Purchases sheet")
				return render_template("index.html")			

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "Business PartnerName":
						rand_tb = cell.row
						clientCell = cell.column
						lun = len(sales[cell.column])
			try:
				listaClient = [b.value for b in sales[clientCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Denumirea clientului Client name' in Sales sheet")
				return render_template("index.html")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "VAT Registration No.":
						rand_tb = cell.row
						coloanaClientID = cell.column
						lun = len(sales[cell.column])
			try:
				listaCUISales = [b.value for b in sales[coloanaClientID][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Codul de inregistrare in scopuri de TVA al clientului Client VAT ID' in Sales sheet")
				return render_template("index.html")

			listaCUISales1=[]
			# listadeclant_1=[]
			for val in listaCUISales:
				if val != None:
					# listadeclant_1.append("")
					listaCUISales1.append(val)
				else:
					listaCUISales1.append("US111")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "Document No.":
						rand_tb = cell.row
						docNumber = cell.column
						lun = len(sales[cell.column])
			try:
				docNoSales = [b.value for b in sales[docNumber][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Nr/ document Document no' in Sales sheet")
				return render_template("index.html")




			for row in sales.iter_rows():
				for cell in row:
					if cell.value == " 19% tax base amount":
						rand_tb = cell.row
						taxBaseL19 = cell.column
						lun = len(sales[cell.column])
			try:
				taxBaseL19 = [b.value for b in sales[taxBaseL19][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)- Livrari locale-Taxable base (RON)-Local supplies (19%)' in Sales sheet")
				return render_template("index.html")
			#print(taxBaseL19)

			taxBaseL19_1=[]
			for c in range(0, len(taxBaseL19)):
				if taxBaseL19[c] == None:
					taxBaseL19_1.append(0)
				else:
					taxBaseL19_1.append(taxBaseL19[c])

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "      19% tax amount":
						rand_tb = cell.row
						vatBaseL19 = cell.column
						lun = len(sales[cell.column])
			try:
				vatL19 = [b.value for b in sales[vatBaseL19][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-Livrari locale-VAT (RON)-Local supplies (19%)' in Sales sheet")
				return render_template("index.html")

			vatL19_1=[]
			for c in range(0, len(vatL19)):
				if vatL19[c] == None:
					vatL19_1.append(0)
				else:
					vatL19_1.append(vatL19[c])

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "  9% tax base amount":
						rand_tb = cell.row
						taxBaseL9 = cell.column
						lun = len(sales[cell.column])
			try:
				taxBaseL9 = [b.value for b in sales[taxBaseL9][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)-Livrari locale-Taxable base (RON)-Local supplies (9%)' in Sales sheet")
				return render_template("index.html")


			taxBaseL9_1=[]
			for c in range(0, len(taxBaseL9)):
				if taxBaseL9[c] == None:
					taxBaseL9_1.append(0)
				else:
					taxBaseL9_1.append(taxBaseL9[c])


			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "       9% tax amount":
						rand_tb = cell.row
						vatBaseL9 = cell.column
						lun = len(sales[cell.column])
			try:
				vatL9 = [b.value for b in sales[vatBaseL9][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-Livrari locale-VAT (RON)-Local supplies (9%)' in Sales sheet")
				return render_template("index.html")

			vatL9_1=[]
			for c in range(0, len(vatL9)):
				if vatL9[c] == None:
					vatL9_1.append(0)
				else:
					vatL9_1.append(vatL9[c])

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "         5% tax base":
						rand_tb = cell.row
						taxBaseL5 = cell.column
						lun = len(sales[cell.column])
			try:
				taxBaseL5 = [b.value for b in sales[taxBaseL5][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (EUR/USD)-Livrari locale-Taxable base (EUR/USD)-Local supplies (5%) ' in Sales sheet")
				return render_template("index.html")


			taxBaseL5_1=[]
			for c in range(0, len(taxBaseL5)):
				if taxBaseL5[c] == None:
					taxBaseL5_1.append(0)
				else:
					taxBaseL5_1.append(taxBaseL5[c])


			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "       5% tax amount":
						rand_tb = cell.row
						vatBaseL5 = cell.column
						lun = len(sales[cell.column])
			try:
				vatL5 = [b.value for b in sales[vatBaseL5][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-Livrari locale-VAT (RON)-Local supplies (5%)' in Sales sheet")
				return render_template("index.html")

			vatL5_1=[]
			for c in range(0, len(vatL5)):
				if vatL5[c] == None:
					vatL5_1.append(0)
				else:
					vatL5_1.append(vatL5[c])
			# #print(vatL5_1)

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "   Inversal tax base":
						rand_tb = cell.row
						taxBV = cell.column
						lun = len(sales[cell.column])
			try:
				taxBaseV = [b.value for b in sales[taxBV][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Bază de impozitare (RON)-Livrari locale taxare inversa-Taxable base (RON)-Local supplies reverse charge' in Sales sheet")
				return render_template("index.html")



			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "  Total doc.incl.VAT":
						rand_tb = cell.row
						totdoc = cell.column
						lun = len(sales[cell.column])
			totdocuments = [b.value for b in sales[totdoc][rand_tb:lun+1]]
			taxBaseV_1=[]
			for c in range(0, len(taxBaseV)):
				if taxBaseV[c] == None:
					taxBaseV_1.append(0)
				else:
					taxBaseV_1.append(taxBaseV[c])
			# #print(taxBaseV_1)

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "Outside RO deductabl":
						rand_tb = cell.row
						taxBi = cell.column
						lun = len(sales[cell.column])
			try:
				taxBaseIntracom = [b.value for b in sales[taxBi][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services' in Sales sheet")
				return render_template("index.html")

			taxBaseIntracom_1=[]
			for c in range(0, len(taxBaseIntracom)):
				if taxBaseIntracom[c] == None:
					taxBaseIntracom_1.append(0)
				else:
					taxBaseIntracom_1.append(taxBaseIntracom[c])

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "Outside RO nonDeduct":
						rand_tb = cell.row
						taxBiSc = cell.column
						lun = len(sales[cell.column])
			try:
				taxBaseIntracomScutit = [b.value for b in sales[taxBiSc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services' in Sales sheet")
				return render_template("index.html")


			taxBaseIntracomScutit_1=[]
			for c in range(0, len(taxBaseIntracomScutit)):
				if taxBaseIntracomScutit[c] == None:
					taxBaseIntracomScutit_1.append(0)
				else:
					taxBaseIntracomScutit_1.append(taxBaseIntracomScutit[c])
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "   Exempt nonDeduct.":
						rand_tb = cell.row
						taxBiSc = cell.column
						lun = len(sales[cell.column])
			try:
				taxBaseIntracomScutit = [b.value for b in sales[taxBiSc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services' in Sales sheet")
				return render_template("index.html")

			serieCuiSales=[]
			codTaraCuiSales=[]
			for i in listaCUISales1:
				if(str(i)[:1].isalpha()):
					r = re.compile("([a-zA-Z]+)([0-9]+)")
					m = r.match(str(i))
					try:
						serieCuiSales.append(m.group(2))
						codTaraCuiSales.append(m.group(1))

					except:
						codTaraCuiSales.append(None)
						serieCuiSales.append(str(i))	

				else:
					codTaraCuiSales.append(None)
					serieCuiSales.append(str(i))
				# tara,oras=i.split(',',1)
				# serieCuiSales.append(oras)
				# codTaraCuiSales.append(tara)
			# #print(codTaraCuiSales)
			# TIP Furnizor!!!!!
			# #print(len(codTaraCuiSales))
			print(serieCuiSales,codTaraCuiSales)
			coteTVAsales=[]
			for i in range(0, len(docNoSales)):
				if (ltaxcode[i]=="A1"):
					coteTVAsales.append(19)
				else:
					if (ltaxcode[i]=="A8"):
						coteTVAsales.append(9)
					else:
						coteTVAsales.append(0)

			codTranzactieSales=[]
			for i in range(0, len(codTaraCuiSales)):
				if str(serieCuiSales[i])[1:2].isalpha():
					codTranzactieSales.append(2)
				else:
					if codTaraCuiSales[i] == "RO":
						# #print("RO")
						codTranzactieSales.append(1)
					else:	
						if codTaraCuiSales[i] in nomenclatorTari:
							# #print("UE")
							codTranzactieSales.append(3)
						else:
							# #print("nonUE")
							codTranzactieSales.append(4)
			#Cote TVAA




			#TIP TRANZACTIE
			storno=[]
			tipTranzSale = []
			# #print(docNoSales)
			# #print(len(docNoSales),len(codTranzactieSales))
			# print(len(docNoSales),len(codTranzactieSales))
			# print(codTranzactieSales)
			for i in range(0, len(docNoSales)):
				print(ltaxcode[i])
				# print(docNoSales[i],print(codTranzactieSales[i]))
				if(listadeclant_1[i]=="Yes"):
					tipTranzSale.append("Declarat anterior")
				else:
					if int(codTranzactieSales[i]) == 1:
						# #print(docNoSales[i]," ",listaCUISales1[i], "", taxBaseL19_1[i], " ", taxBaseL9_1[i], " ", taxBaseL5_1[i])
						# if (int(taxBaseL19_1[i])>0 and int(vatL19_1[i])>0) or (int(taxBaseL9_1[i])>0 and int(vatL9_1[i])> 0) or (int(taxBaseL5_1[i])>0 and int(vatL5_1[i])>0):
						if (ltaxcode[i]=='A1'):
							print("Yes")
							tipTranzSale.append('L')
							storno.append("")
						else:
							# None
							if (ltaxcode[i]=='A8'):
								print("Yes")
								tipTranzSale.append('L')
							else:
								if ltaxcode[i]=='A5' or ltaxcode[i]=='A2' or ltaxcode[i]=='A4':
									print("Yes")
									tipTranzSale.append("V")
									storno.append("")
					else:
						if int(codTranzactieSales[i]) == 2:
							if (ltaxcode[i]=='A1'):
								tipTranzSale.append('L')
								storno.append("")
							else:
								tipTranzSale.append("Not applicable for D394")
						else:
							if int(codTranzactieSales[i]) == 3:
								if (ltaxcode[i]=='E1' or ltaxcode[i]=='X1' or ltaxcode[i]=='Y3'):
									tipTranzSale.append('Not applicable for D394')
									storno.append("")
								else:
									if (ltaxcode[i]=='A1' or ltaxcode[i]=='A8'):
										tipTranzSale.append('L')
										storno.append("")
									else:
										tipTranzSale.append('Not applicable for D394')
							else:
								if int(codTranzactieSales[i]) == 4:
									if (ltaxcode[i]=='A1' or ltaxcode[i]=='A8'):
										tipTranzSale.append('L')
										storno.append("")
									else:
										if (ltaxcode[i]=='E1' or ltaxcode[i]=='X1' or ltaxcode[i]=='Y3'):
											tipTranzSale.append('Not applicable for d394')
										else:
											tipTranzSale.append('L')
											storno.append("")


			# #print(docNoSales)
			#Scriere in excel

			salesExcel.cell(row=9, column=1).value = "Cod tara"
			salesExcel.cell(row=9, column=2).value = "Serie cui"
			salesExcel.cell(row=9, column=3).value = "Numar document"
			salesExcel.cell(row=9, column=4).value = "CUI"
			salesExcel.cell(row=9, column=5).value = "Clasa tranzactie"
			salesExcel.cell(row=9, column=6).value = "Tip tranzactie"
			salesExcel.cell(row=9, column=7).value = "Cota TVA"
			salesExcel.cell(row=9, column=8).value = "Total document"
			salesExcel.cell(row=9, column=9).value = "Tip jurnal"
			salesExcel.cell(row=9, column=10).value = "Nume partener"
			salesExcel.cell(row=9, column=11).value = "Check"
			salesExcel.cell(row=9, column=12).value = "Cod si denumire NC produs(TIP V)"

			# dv = DataValidation(
				# type='list', formula1='"Yes,No"', allow_blank=True)


			listahelp=["1002--Secara","1003--Orz","1005--Porumb","1201--Boabe de soia"," 1205--Seminte de rapita sau de rapita salbatica","120600--Seminte de floarea soarelui","121291--Sfecla de zahar","1001-Grau si meslin","1004--Ovaz","10086000--Triticale","22-deseuri feroase si neferoase","23-masa lemnoasa","32-terenuri","33-constructii","34-alte bunuri","35-servicii","24-certificate de emisii de gaze cu efect de sera","25-energie electrica","26-certificate verzi","27-constructii/terenuri","28-aur de investitii","29-telefoane mobile","30-microprocesoare","31-console de jocuri tablete PC si laptopuri"]
			sheethelp=temp.create_sheet("Validation")
			sheethelp.sheet_state = 'hidden'

			# dv = DataValidation(
			# 	type="list", formula1="", allow_blank=True)
			# salesExcel.add_data_validation(dv)

			# dv.add(salesExcel["L2"])



			print(len(tipTranzSale),len(codTranzactieSales))
			for i in range(0, len(codTaraCuiSales)):
				print(serieCuiSales[i],docNoSales[i],listaCUISales1[i],codTranzactieSales[i],tipTranzSale[i])
				salesExcel.cell(row=10 + i, column=1).value = codTaraCuiSales[i]
				salesExcel.cell(row=10 + i, column=2).value = serieCuiSales[i]
				salesExcel.cell(row=10 + i, column=3).value = docNoSales[i]
				salesExcel.cell(row=10 + i, column=4).value = listaCUISales1[i]
				salesExcel.cell(row=10 + i, column=5).value = codTranzactieSales[i]
				# if(listadeclant_1[i]!=""):
				salesExcel.cell(row=10 + i, column=6).value = tipTranzSale[i]
				# else:
					# salesExcel.cell(row=10 + i, column=6).value = listadeclant_1[i]

				if(tipTranzSale[i]=='V'):
					salesExcel.cell(row=10+i,column=12).value="V"
				else:
					salesExcel.cell(row=10+i,column=12).value="N/A, valid only for V trans."

				salesExcel.cell(row=10 + i, column=8).value = totdocuments[i]
				salesExcel.cell(row=10 + i, column=9).value = "Jurnal vanzari"
				salesExcel.cell(row=10 + i, column=10).value = listaClient[i]

			for i in range(0, len(coteTVAsales)):
				salesExcel.cell(row=10 + i, column=7).value = coteTVAsales[i]
				salesExcel.cell(row=10+i,column=18).value="=B{0}&E{0}&F{0}&G{0}".format(i+10)

			#FORMATARE------------------------------------------------------------------
			red_color = 'ffc7ce'
			green_color='99ff99'
			red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
			green_fill = styles.PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
			row=salesExcel.max_row	
			salesExcel.conditional_formatting.add('K10:K'+str(row-1), formatting.rule.CellIsRule(operator='notEqual', formula=['"OK"'], fill=red_fill))
			for row in salesExcel['A9:L9']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in salesExcel['A9:L9']:
				for cell in row:
					cell.font = cap_tabel

			# for row in salesExcel['A9:K9']:
			# 	for cell in row:
			# 		cell.border = border_thin

			
			salesExcel.freeze_panes = 'A10'

			salesExcel.column_dimensions['B'].width = 20
			salesExcel.column_dimensions['C'].width = 20
			salesExcel.column_dimensions['D'].width = 20
			salesExcel.column_dimensions['E'].width = 16
			salesExcel.column_dimensions['F'].width = 16
			salesExcel.column_dimensions['F'].width = 20		
			salesExcel.column_dimensions['H'].width = 14
			salesExcel.column_dimensions['J'].width = 35
			salesExcel.column_dimensions['L'].width = 35		
			purchases = temp['Purchases']

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Business PartnerName":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				supplierName = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Furnizor Supplier' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == " Doc. Date":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				datadocument = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for ' Doc. Date' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Aplica TVA la incasare VAT cash-in system":
						rand_tb = cell.row
						vatCashinSys = cell.column
						lun = len(purchases[cell.column])
			try:
				vatApplies = [b.value for b in purchases[vatCashinSys][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Aplica TVA la incasare VAT cash-in system' in Purchases sheet")
				return render_template("index.html")

			vatApplies_1=[]


			for val in vatApplies:
				if val != None:
					vatApplies_1.append(val)


			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "VAT Registration No.":
						rand_tb = cell.row
						suppID = cell.column
						lun = len(purchases[cell.column])
			try:
				suppIDPurch = [b.value for b in purchases[suppID][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Codul de înregistrare în scopuri de TVA VAT number' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Tax code":
						rand_tb = cell.row
						suppID = cell.column
						lun = len(purchases[cell.column])
			try:
				taxcodep = [b.value for b in purchases[suppID][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Codul de înregistrare în scopuri de TVA VAT number' in Purchases sheet")
				return render_template("index.html")		
			

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Nr. document Document no":
						rand_tb = cell.row
						docNumberPurch = cell.column
						lun = len(purchases[cell.column])
			try:
				docNoPurch = [b.value for b in purchases[docNumberPurch][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Nr. document Document no' in Purchases sheet")
				return render_template("index.html")
			
			docNoPurch1 = []
			for val in docNoPurch:

				docNoPurch1.append(val)
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "19%taxbase":
						rand_tb = cell.row
						taxBaseAch19cell = cell.column
						lun = len(purchases[cell.column])
			try:
				taxBaseAch19 = [b.value for b in purchases[taxBaseAch19cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON) -achizitii locale-Taxable base (RON) - local acquisition (19%)' in Purchases sheet")
				return render_template("index.html")

			taxBaseAch19_1=[]
			for i in range(0, len(taxBaseAch19)):
				if taxBaseAch19[i] == None:
					taxBaseAch19_1.append(0)
				else:
					taxBaseAch19_1.append(taxBaseAch19[i])

			# #print(taxBaseAch19,taxBaseAch19_1)
			# #print(taxBaseAch19_1)
			# #print(taxBaseAch19)
			# for item in taxBaseAch19:
			#    #print(type(item))
			# #print(type(taxBaseAch19))

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "19%taxamount":
						rand_tb = cell.row
						vatAch19cell = cell.column
						lun = len(purchases[cell.column])
			try:
				vatAch19 = [b.value for b in purchases[vatAch19cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-achizitii locale-VAT (RON)-local acquisition (19%)' in Purchases sheet")
				return render_template("index.html")

			vatAch19_1=[]
			for i in range(0, len(vatAch19)):
				if vatAch19[i] == None or vatAch19[i]=="":
					vatAch19_1.append(0)
				else:
					vatAch19_1.append(vatAch19[i])
			# #print(vatAch19_1)

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "9%taxbase":
						rand_tb = cell.row
						taxBaseAch9cell = cell.column
						lun = len(purchases[cell.column])
			try:
				taxBaseAch9 = [b.value for b in purchases[taxBaseAch9cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)-achizitii locale- Taxable base (RON)-local acquisition (9%)' in Purchases sheet")
				return render_template("index.html")

			taxBaseAch9_1=[]
			for i in range(0, len(taxBaseAch9)):
				if taxBaseAch9[i] == None:
					taxBaseAch9_1.append(0)
				else:
					taxBaseAch9_1.append(taxBaseAch9[i])
			# #print(taxBaseAch9_1)

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "9%taxamount":
						rand_tb = cell.row
						vatAch9cell = cell.column
						lun = len(purchases[cell.column])
			try:
				vatAch9 = [b.value for b in purchases[vatAch9cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-achizitii locale-VAT (RON)-local acquisition (9%)' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "TotaldocinclVAT":
						rand_tb = cell.row
						totdocp = cell.column
						lun = len(purchases[cell.column])
			try:
				totdocumentp = [b.value for b in purchases[totdocp][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total document (inclusiv TVA)-RON' in Purchases sheet")
				return render_template("index.html")

			vatAch9_1=[]
			for i in range(0, len(vatAch9)):
				if vatAch9[i] == None:
					vatAch9_1.append(0)
				else:
					vatAch9_1.append(vatAch9[i])
			# #print(vatAch9_1)

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "5%taxbase":
						rand_tb = cell.row
						taxBaseAch5cell = cell.column
						lun = len(purchases[cell.column])
			try:
				taxBaseAch5 = [b.value for b in purchases[taxBaseAch5cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)-achizitii locale Taxable base (RON) local acquisition (5%)' in Purchases sheet")
				return render_template("index.html")

			taxBaseAch5_1=[]
			for i in range(0, len(taxBaseAch5)):
				if taxBaseAch5[i] == None:
					taxBaseAch5_1.append(0)
				else:
					taxBaseAch5_1.append(taxBaseAch5[i])
			# #print(taxBaseAch5_1)

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "5%taxamount":
						rand_tb = cell.row
						vatAch5cell = cell.column
						lun = len(purchases[cell.column])
			try:
				vatAch5 = [b.value for b in purchases[vatAch5cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-achizitii locale-VAT (RON)-local acquisition (5%)' in Purchases sheet")
				return render_template("index.html")

			vatAch5_1=[]
			for i in range(0, len(vatAch5)):
				if vatAch5[i] == None:
					vatAch5_1.append(0)
				else:
					vatAch5_1.append(vatAch5[i])


			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Exemptint+impbase":
						rand_tb = cell.row
						vatExemptLocAcq = cell.column
						lun = len(purchases[cell.column])
			try:
				vatExempt = [b.value for b in purchases[vatExemptLocAcq][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
				return render_template("index.html")

			vatExempt_1=[]
			for i in range(0, len(vatExempt)):
				if vatExempt[i] == None:
					vatExempt_1.append(0)
				else:
					vatExempt_1.append(vatExempt[i])

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Intra-commtaxbase":
						rand_tb = cell.row
						nonChartb = cell.column
						lun = len(purchases[cell.column])
			try:
				nonCharTaxBase = [b.value for b in purchases[nonChartb][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
				return render_template("index.html")

			nonCharTaxBase_1=[]
			for i in range(0, len(nonCharTaxBase)):
				if nonCharTaxBase[i] == None:
					nonCharTaxBase_1.append(0)
				else:
					nonCharTaxBase_1.append(nonCharTaxBase[i])
			# #print(nonCharTaxBase_1)

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Intra-comtaxamount":
						rand_tb = cell.row
						nonChartVATtb = cell.column
						lun = len(purchases[cell.column])
			try:
				nonChartVATBase = [b.value for b in purchases[nonChartVATtb][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA nedeductibila/ Non-deductible VAT (RON)' in Purchases sheet")
				return render_template("index.html")

			nonChartVATBase_1=[]
			for i in range(0, len(nonChartVATBase)):
				if nonChartVATBase[i] == None:
					nonChartVATBase_1.append(0)
				else:
					nonChartVATBase_1.append(nonChartVATBase[i])
			# #print(nonChartVATBase_1)

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Revtaxbase-art150":
						rand_tb = cell.row
						revTaxBaseAch19cell = cell.column
						lun = len(purchases[cell.column])
			try:
				revTaxBaseAch19 = [b.value for b in purchases[revTaxBaseAch19cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge (19%)' in Purchases sheet")
				return render_template("index.html")

			revTaxBaseAch19_1=[]
			for i in range(0, len(revTaxBaseAch19)):
				if revTaxBaseAch19[i] == None:
					revTaxBaseAch19_1.append(0)
				else:
					revTaxBaseAch19_1.append(revTaxBaseAch19[i])
			# #print(revTaxBaseAch19_1)

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Revtaxamnt-art150":
						rand_tb = cell.row
						revVatAch19cell = cell.column
						lun = len(purchases[cell.column])
			try:
				revVatAch19 = [b.value for b in purchases[revVatAch19cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-Achizitii locale taxare inversa-VAT (RON)-Local acquisition reverse charge (19%)' in Purchases sheet")
				return render_template("index.html")

			revVatAch19_1=[]
			for i in range(0, len(revVatAch19)):
				if revVatAch19[i] == None:
					revVatAch19_1.append(0)
				else:
					revVatAch19_1.append(revVatAch19[i])
			# #print(revVatAch19_1)

			# for row in purchases.iter_rows():
			# 	for cell in row:
			# 		if cell.value == "Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge (9%)":
			# 			rand_tb = cell.row
			# 			revTaxBaseAch9cell = cell.column
			# 			lun = len(purchases[cell.column])
			# try:
			# 	revTaxBaseAch9 = [b.value for b in purchases[revTaxBaseAch9cell][rand_tb:lun]]
			# except:
			# 	flash("Please insert the correct header for 'Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge (9%)' in Purchases sheet")
			# 	return render_template("index.html")

			# revTaxBaseAch9_1=[]
			# for i in range(0, len(revTaxBaseAch9)):
			# 	if revTaxBaseAch9[i] == None:
			# 		revTaxBaseAch9_1.append(0)
			# 	else:
			# 		revTaxBaseAch9_1.append(revTaxBaseAch9[i])
			# # #print(revTaxBaseAch9_1)

			# for row in purchases.iter_rows():
			# 	for cell in row:
			# 		if cell.value == "TVA (RON)-Achizitii locale taxare inversa-VAT (RON)-Local acquisition reverse charge (9%)":
			# 			rand_tb = cell.row
			# 			revVatAch9cell = cell.column
			# 			lun = len(purchases[cell.column])
			# try:
			# 	revVatAch9 = [b.value for b in purchases[revVatAch9cell][rand_tb:lun]]
			# except:
			# 	flash("Please insert the correct header for 'TVA (RON)-Achizitii locale taxare inversa-VAT (RON)-Local acquisition reverse charge (9%)' in Purchases sheet")
			# 	return render_template("index.html")

			# revVatAch9_1=[]
			# for i in range(0, len(revVatAch9)):
			# 	if revVatAch9[i] == None:
			# 		revVatAch9_1.append(0)
			# 	else:
			# 		revVatAch9_1.append(revVatAch9[i])
			# # #print(revVatAch9_1)

			# for row in purchases.iter_rows():
			# 	for cell in row:
			# 		if cell.value == "Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge(5%)":
			# 			rand_tb = cell.row
			# 			revTaxBaseAch5cell = cell.column
			# 			lun = len(purchases[cell.column])
			# try:
			# 	revTaxBaseAch5 = [b.value for b in purchases[revTaxBaseAch5cell][rand_tb:lun]]
			# except:
			# 	flash("Please insert the correct header for 'Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge(5%)' in Purchases sheet")
			# 	return render_template("index.html")

			# revTaxBaseAch5_1=[]
			# for i in range(0, len(revTaxBaseAch5)):
			# 	if revTaxBaseAch5[i] == None:
			# 		revTaxBaseAch5_1.append(0)
			# 	else:
			# 		revTaxBaseAch5_1.append(revTaxBaseAch5[i])
			# # #print(revTaxBaseAch5_1)

			# for row in purchases.iter_rows():
			# 	for cell in row:
			# 		if cell.value == "TVA (RON)-Achizitii locale taxare inversa VAT (RON)-Local acquisition reverse charge (5%)":
			# 			rand_tb = cell.row
			# 			revVatAch5cell = cell.column
			# 			lun = len(purchases[cell.column])
			# try:
			# 	revVatAch5 = [b.value for b in purchases[revVatAch5cell][rand_tb:lun]]
			# except:
			# 	flash("Please insert the correct header for 'TVA (RON)-Achizitii locale taxare inversa VAT (RON)-Local acquisition reverse charge (5%)' in Purchases sheet")
			# 	return render_template("index.html")

			# revVatAch5_1=[]
			# for i in range(0, len(revVatAch5)):
			# 	if revVatAch5[i] == None:
			# 		revVatAch5_1.append(0)
			# 	else:
			# 		revVatAch5_1.append(revVatAch5[i])
			# # #print(revVatAch5_1)

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Intra-communexempt":
						rand_tb = cell.row
						nonUEGoodscell = cell.column
						lun = len(purchases[cell.column])
			try:
				taxNonUEgoods = [b.value for b in purchases[nonUEGoodscell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Bază de impozitare (RON)-bunuri nonUE-Taxable base (RON)-nonUE goods' in Purchases sheet")
				return render_template("index.html")
			taxNonUEgoods_1=[]
			for i in range(0, len(taxNonUEgoods)):
				if taxNonUEgoods[i] == None:
					taxNonUEgoods_1.append(0)
				else:
					taxNonUEgoods_1.append(taxNonUEgoods[i])

			# #print(datadocument)
			# for row in purchases.iter_rows():
			# 	for cell in row:
			# 		if cell.value == "TVA (RON)-bunuri nonUE-VAT (RON)-nonUE goods":
			# 			rand_tb = cell.row
			# 			vatNonUEGoodscell = cell.column
			# 			lun = len(purchases[cell.column])
			# try:
			# 	vatNonUEGoods = [b.value for b in purchases[vatNonUEGoodscell][rand_tb:lun]]
			# except:
			# 	flash("Please insert the correct header for 'TVA (RON)-bunuri nonUE-VAT (RON)-nonUE goods' in Purchases sheet")
			# 	return render_template("index.html")

			# vatNonUEGoods_1=[]
			# for i in range(0, len(vatNonUEGoods)):
			# 	if vatNonUEGoods[i] == None:
			# 		vatNonUEGoods_1.append(0)
			# 	else:
			# 		vatNonUEGoods_1.append(vatNonUEGoods[i])


			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Intra-communexempt":
						rand_tb = cell.row
						nonUEServcell = cell.column
						lun = len(purchases[cell.column])
			try:
				taxNonUEservices = [b.value for b in purchases[nonUEServcell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Bază de impozitare (RON)-servicii nonUE-Taxable base (RON)-nonUE services' in Purchases sheet")
				return render_template("index.html")

			# taxNonUEservices_1=[]
			taxNonUEservices_1=[]
			for i in range(0, len(taxNonUEservices)):
				if taxNonUEservices[i] == None:
					taxNonUEservices_1.append(0)
				else:
					taxNonUEservices_1.append(taxNonUEservices[i])


			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Intra-commnon-tax":
						rand_tb = cell.row
						vatNonUEservcell = cell.column
						lun = len(purchases[cell.column])
			try:
				vatNonUEservices = [b.value for b in purchases[vatNonUEservcell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-servicii nonUE-VAT (RON)-nonUE services' in Purchases sheet")
				return render_template("index.html")

			vatNonUEservices_1=[]
			for i in range(0, len(vatNonUEservices)):
				if vatNonUEservices[i] == None:
					vatNonUEservices_1.append(0)
				else:
					vatNonUEservices_1.append(vatNonUEservices[i])

			serieCuiPurch = []
			codTaraCuiPurch = []
			# #print(suppIDPurch)
			for i in suppIDPurch:
				# #print(i)
				if(str(i)[:1].isalpha()):	
					r = re.compile("([a-zA-Z]+)([0-9]+)")
					m = r.match(str(i))
					try:
						serieCuiPurch.append(m.group(2))
					except:
						serieCuiPurch.append(" ")
					try:
						codTaraCuiPurch.append(m.group(1))
					except:
						codTaraCuiPurch.append(" ")
				else:
					codTaraCuiPurch.append(None)
					serieCuiPurch.append(i)
			# #print(codTaraCuiPurch,serieCuiPurch)
			tipTranzactiePurchases=[]
			#print("Aici vat -------",vatAch19_1,"-----Achizitii")
			#Tip furnizor
			for i in range(0, len(codTaraCuiPurch)):
				if codTaraCuiPurch[i] == "RO":
					# #print("RO")
					tipTranzactiePurchases.append(1)
				else:
					if serieCuiPurch[i] in suppIDPurch and int(nonCharTaxBase_1[i])>0:
						tipTranzactiePurchases.append(2)
					else:
						if codTaraCuiPurch[i] in nomenclatorTari:
							# #print("UE")
							tipTranzactiePurchases.append(3)
						else:
							# #print("nonUE")
							tipTranzactiePurchases.append(4)

			#Cote TVA
			coteTVApurchases=[]
			for i in range(0, len(docNoPurch1)):
				if (taxcodep[i]=="ZI" or taxcodep[i]=="C3" or taxcodep[i]=="ZD" or taxcodep[i]=="E1" or taxcodep[i]=='V1' or taxcodep[i]=="5H" or taxcodep[i]=="X1" or taxcodep[i]=="1L" or taxcodep[i]=="1M" or taxcodep[i]=="3S" or taxcodep[i]=="5B" or taxcodep[i]=="5H"):
					coteTVApurchases.append('19')
				else:
					if (taxcodep[i]=="I9" or taxcodep[i]=="W8" or taxcodep[i]=="W6" or taxcodep[i]=="J2" or taxcodep[i]=="V3" or taxcodep[i]=="9S"):
						coteTVApurchases.append('9')
					else:
						if (taxcodep[i]=="I7" or taxcodep[i]=="5D" or taxcodep[i]=="W8" or taxcodep[i]=="6I" or taxcodep[i]=="5S"):
							coteTVApurchases.append('5')
						else:
							coteTVApurchases.append('0')

			#mapare tip tranzactie
			# #print(len(docNoPurch1),len(tipTranzactiePurchases),len(vatApplies))
			# #print(vatApplies)
			tipTranzPurch=[]

			# for i in range(0,len(suppIDPurch)):
				# #print(suppIDPurch[i],tipTranzactiePurchases[i])
			# #print(len(docNoPurch),len(tipTranzactiePurchases))
			#print(len(docNoPurch1),len(listadeclantp_1),"--------------len de lista")
			for i in range(0, len(docNoPurch1)):
				if(listadeclantp_1[i]=="Yes"):
					tipTranzPurch.append("Declarat anterior")
				else:
					# #print(docNoPurch1[i])
					if int(tipTranzactiePurchases[i]) == 1:
						if (taxcodep[i]=="V1" or taxcodep[i]=="W8" or taxcodep[i]=="V3" or taxcodep[i]=="5H" or taxcodep[i]=="5B"):
							tipTranzPurch.append('A')
						else:
							if taxcodep[i]=="ZI" or taxcodep[i]=="5D" or taxcodep[i]=="ZD" or taxcodep[i]=="I7" or taxcodep[i]=="W8" or taxcodep[i]=="I9" or taxcodep[i]=="W6" or taxcodep[i]=="6I":
								tipTranzPurch.append("AI")
								#print(docNoPurch1[i],";;;;;es 3")
							else:
								if (taxcodep[i]=="ZR"):
									tipTranzPurch.append("C")
									#print("Yes 5")
								else:
									if (taxcodep[i]=="AS"):
										tipTranzPurch.append("AS")
										#print("Yes 7")
									else:
										tipTranzPurch.append("Not applicable for D394")
					else:
						if int(tipTranzactiePurchases[i]) == 2:
							if taxcodep[i]=="7N" or taxcodep[i]=="8N" or taxcodep[i]=="A3" :
								tipTranzPurch.append("N")
								#print(docNoPurch1[i],";;;;es 9")
						else:
							if int(tipTranzactiePurchases[i]) == 3:
								if (taxcodep[i]=="V1" or taxcodep[i]=="W8" or taxcodep[i]=="V3"):
											tipTranzPurch.append('A')
											#print(docNoPurch1[i],";;;;;es 12")
								else:
										#print(docNoPurch1[i],";;;;;es 13")
									if (taxcodep[i]=="ZR"):
										tipTranzPurch.append("C")
										#print("Yes 14")
									else:
										tipTranzPurch.append("Not applicable for D394")
							else:
									if int(tipTranzactiePurchases[i]) == 4:
										if (taxcodep[i]=="V1" or taxcodep[i]=="W8" or taxcodep[i]=="V3"):
											tipTranzPurch.append('A')
											#print(docNoPurch1[i],";;;;;es 12")
										else:
											if (taxcodep[i]=="ZR"):
												tipTranzPurch.append("C")
											else:
												tipTranzPurch.append("Not applicable for D394")
										#print("Yes 16")
				# #print(docNoPurch1[i],tipTranzPurch[i],docNoPurch[i+1])
			ma=salesExcel.max_row+1
			for i in range(0, len(codTaraCuiPurch)):
				salesExcel.cell(row=ma + i, column=1).value = codTaraCuiPurch[i]

			for i in range(0, len(serieCuiPurch)):
				salesExcel.cell(row=ma + i, column=2).value = serieCuiPurch[i]

			for i in range(0, len(docNoPurch1)):
				salesExcel.cell(row=ma+ i, column=3).value = docNoPurch1[i]

			for i in range(0, len(suppIDPurch)):
				salesExcel.cell(row=ma + i, column=4).value = suppIDPurch[i]

			for i in range(0, len(tipTranzactiePurchases)):
				salesExcel.cell(row=ma+ i, column=5).value = tipTranzactiePurchases[i]

			for i in range(0, len(tipTranzPurch)):
				# if(listadeclantp_1!=""):
				salesExcel.cell(row=ma+ i, column=6).value = tipTranzPurch[i]
				# else:
					# salesExcel.cell(row=ma+ i, column=6).value = "Declarate anterior"
				if(tipTranzPurch[i]=="V"):
					salesExcel.cell(row=ma+i,column=12).value="Add type of tranzactie"
				else:
					salesExcel.cell(row=ma+i,column=12).value="N/A"

			for i in range(0, len(coteTVApurchases)):
				#print(coteTVApurchases[i])
				salesExcel.cell(row=ma+ i, column=7).value = coteTVApurchases[i]
				salesExcel.cell(row=ma+ i, column=8).value = totdocumentp[i]
				salesExcel.cell(row=ma+ i, column=9).value = "Jurnal cumparari"
				salesExcel.cell(row=ma+ i, column=10).value = supplierName[i]

			codTaraCUItotal=codTaraCuiPurch+codTaraCuiSales
			for i in range(0, len(codTaraCUItotal)):
				salesExcel.cell(row=10 + i, column=11).value = '=IFERROR(IF(VLOOKUP(B{0}&E{0}&F{0}&G{0},Tranzactii!K:K,1,0)=B{0}&E{0}&F{0}&G{0},"OK","Mapped missing in Tranzactii sheet"),"Mapped missing inTranzactiisheet")'.format(10+i)
			salesExcel.auto_filter.ref = "A9:L9"
			for row in salesExcel['H10:F1000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'		
			tranzactii=temp.create_sheet("Tranzactii")
			tranzactii.freeze_panes = 'A6'
			tranzactii.sheet_view.showGridLines = False

													
			setSalesCUI=set(listaCUISales1)
			idSalesUnique=list(setSalesCUI)

			setPurchCUI=set(suppIDPurch)
			idPurchUnique=list(setPurchCUI)

			listaCUIUnique=idSalesUnique+idPurchUnique

			setlistaClient=set(listaClient)
			listaClientUnique=list(setlistaClient)

			setSupplierName=set(supplierName)
			supplierNameUnique=list(setSupplierName)

			# for k in range(0,len(setlistaClient)):
			#    count
			#    for j in range(0,len(listaClient)):


			# print(len(supplierName),len(tipTranzPurch))
			listanouaappendpurch=[]
			# for i in range(0,len(supplierName)):
			for p in range(0,len(serieCuiPurch)):
				print(serieCuiPurch[p],tipTranzPurch[p],coteTVApurchases[p],tipTranzactiePurchases[p])
			# print()
			print(len(serieCuiPurch),len(tipTranzPurch),len(coteTVApurchases),len(tipTranzactiePurchases))
			for k in range(0,len(serieCuiPurch)):
				try:
					print(serieCuiPurch[k],tipTranzPurch[k])
				except:
					print(serieCuiPurch[k])
				listanouaappendpurch.append(str(serieCuiPurch[k])+";"+str(tipTranzPurch[k])+";"+str(coteTVApurchases[k])+";"+str(tipTranzactiePurchases[k])+";"+str(listadeclantp_1[k]))

			listanouasetpurch=list(set(listanouaappendpurch))


			listanouaappendsales=[]

			for k in range(0,len(serieCuiSales)):
				listanouaappendsales.append(str(serieCuiSales[k])+";"+str(tipTranzSale[k])+";"+str(coteTVAsales[k])+";"+str(codTranzactieSales[k])+";"+listadeclant_1[k])

			listanouasetsales=list(set(listanouaappendsales))

			countsales=[]
			for p in range(0,len(listanouasetsales)):
				count=0
				for k in range(0,len(listanouaappendsales)):
					if(listanouaappendsales[k]==listanouasetsales[p]):
						count=count+1
				countsales.append(count)
			countpurch=[]
			for p in range(0,len(listanouasetpurch)):
				count=0
				for k in range(0,len(listanouaappendpurch)):
					if(listanouaappendpurch[k]==listanouasetpurch[p]):
						count=count+1
				countpurch.append(count)


			#print(listanouasetsales)

			tranzactii.cell(row=5,column=1).value="Cui partener"
			tranzactii.cell(row=5,column=2).value="Nume partener"
			tranzactii.cell(row=5,column=3).value="Tip partener"
			tranzactii.cell(row=5,column=4).value="Tip tranzactie"
			tranzactii.cell(row=5,column=5).value="Cota TVA"
			tranzactii.cell(row=5,column=6).value="Baza TVA"
			tranzactii.cell(row=5,column=7).value="TVA"
			tranzactii.cell(row=5,column=8).value="Nr Facturi"
			tranzactii.cell(row=5,column=9).value="Neexigibile - nu se vor raporta"
			tranzactii.cell(row=5,column=10).value="Cod si denumire NC produs(TIP V)"
			tranzactii.cell(row=2,column=1).value="Tranzactii"
			tranzactii.cell(row=2,column=1).font=cap_tabeltitlu
			counts=0
			for i in range(0, len(listanouasetsales)):
				x=listanouasetsales[i].split(";")
				#print(x[3],x[4])
				if(int(x[3])<3 or int(x[2])>0):
					counts=counts+1
					y=tranzactii.max_row
					tranzactii.cell(row=y+1,column=1).value=x[0]
					tranzactii.cell(row=y+1,column=2).value="=VLOOKUP(A"+str(y+1)+",'Mapping tranzactii'!B:J,9,0)"
					tranzactii.cell(row=y+1,column=3).value=x[3]
					tranzactii.cell(row=y+1,column=4).value=x[1]
					tranzactii.cell(row=y+1,column=5).value=x[2]
					tranzactii.cell(row=y+1,column=9).value=x[4]
					# tranzactii.cell(row=y+1,column=10).value="=xlookup(K"+str(y+1)+",'Mapping tranzactii'!R:R,'Mapping tranzactii'!L:L)"
			countp=0
			for i in range(0, len(listanouasetpurch)):
				x=listanouasetpurch[i].split(";")
				if(int(x[3])<3 or int(x[2])>0):
					countp=countp+1
					y=tranzactii.max_row
					tranzactii.cell(row=y+1,column=1).value=x[0]
					tranzactii.cell(row=y+1,column=2).value="=VLOOKUP(A"+str(y+1)+",'Mapping tranzactii'!B:J,9,0)"   
					try:
						tranzactii.cell(row=y+1,column=3).value=x[3]
					except:
						tranzactii.cell(row=y+1,column=3).value=""
					tranzactii.cell(row=y+1,column=4).value=x[1]
					tranzactii.cell(row=y+1,column=5).value=x[2]
					tranzactii.cell(row=y+1,column=9).value=x[4]
					# tranzactii.cell(row=y+1,column=10).value="=xlookup(K"+str(y+1)+",'Mapping tranzactii'!R:R,'Mapping tranzactii'!L:L)"

			countmare=countp+counts
			for i in range(0, countmare):
				tranzactii.cell(row=i+6,column=6).value="=SUMIFS('Mapping tranzactii'!H:H,'Mapping tranzactii'!B:B,A{0},'Mapping tranzactii'!E:E,C{0},'Mapping tranzactii'!F:F,D{0},'Mapping tranzactii'!G:G,E{0})/((100+E{0})/100)".format(6+i)
				tranzactii.cell(row=i+6,column=7).value="=F{0}/100*E{0}".format(6+i)
				tranzactii.cell(row=i+6,column=8).value="=COUNTIFS('Mapping tranzactii'!B:B,A{0},'Mapping tranzactii'!E:E,C{0},'Mapping tranzactii'!F:F,D{0},'Mapping tranzactii'!G:G,E{0})".format(6+i)
				tranzactii.cell(row=i+6,column=11).value="=A{0}&C{0}&D{0}&E{0}".format(6+i)


			#---------FORMAT-----------------
			for row in tranzactii['A5:J5']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in tranzactii['A5:J5']:
				for cell in row:
					cell.font = cap_tabel
		
			tranzactii.column_dimensions['K'].hidden = True
			tranzactii.column_dimensions['A'].width = 20
			tranzactii.column_dimensions['I'].width = 27		
			tranzactii.column_dimensions['B'].width = 35
			tranzactii.column_dimensions['C'].width = 13
			tranzactii.column_dimensions['D'].width = 13
			tranzactii.column_dimensions['H'].width = 13
			tranzactii.auto_filter.ref = "A5:H5"
			for row in tranzactii['F6:F'+str(tranzactii.max_row)]:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'
			for row in tranzactii['G6:G'+str(tranzactii.max_row)]:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'

			saf=temp.create_sheet("Facturi storno si anulate")
			saf.sheet_view.showGridLines = False
			saf.freeze_panes = 'A4'		
			saf.cell(row=1,column=1).value="Facturi storno/anulate"
			saf.cell(row=1,column=1).font=cap_tabelbold
			saf.cell(row=3,column=1).value="Tip"
			saf.cell(row=3,column=2).value="Serie"
			saf.cell(row=3,column=3).value="Numar"
			for row in saf['A3:C3']:
				for cell in row:
					cell.font = cap_tabel
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			ind=saf.max_row
			for k in range(0,len(storno)):
				if(storno[k]=="Yes"):
					saf.cell(row=ind+1,column=1).value="Stornata"
					saf.cell(row=ind+1,column=2).value=""
					saf.cell(row=ind+1,column=3).value=docNoSales[k]
					ind=ind+1
			xx=saf.max_row		
			saf.cell(row=xx+1,column=1).value="Anulata"
			saf.cell(row=xx+1,column=2).value="Please input the cancelled invoice number"	
			bonuri=temp.create_sheet("Bonuri fiscale")
			bonuri.cell(row=4,column=1).value="Luna"
			bonuri.cell(row=4,column=2).value="Nr. bon fiscal"
			bonuri.cell(row=4,column=3).value="Baza 5%"
			bonuri.cell(row=4,column=4).value="TVA 5%"
			bonuri.cell(row=4,column=5).value="Baza 9%"
			bonuri.cell(row=4,column=6).value="TVA 9%"
			bonuri.cell(row=4,column=7).value="Baza 19%"
			bonuri.cell(row=4,column=8).value="TVA 19%"
			bonuri.cell(row=4,column=9).value="Baza 20%"
			bonuri.cell(row=4,column=10).value="TVA 20%"

			for row in bonuri['A4:K4']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in bonuri['A4:K4']:
				for cell in row:
					cell.font = cap_tabel
			bonuri.sheet_view.showGridLines = False

			facturi=temp.create_sheet("Sectiunea 2.1&2.2")
			facturi.sheet_view.showGridLines = False
			facturi.column_dimensions['A'].width = 12
			facturi.column_dimensions['B'].width = 18
			facturi.column_dimensions['C'].width = 15

			facturi.cell(row=1,column=1).value="Serie Emise"
			facturi.cell(row=1,column=2).value="Inceput Emise"
			facturi.cell(row=1,column=3).value="Final Emise"
			facturi.cell(row=1,column=4).value="Tip Emise"		

			docNoSales2=[]
			seriefacturi=[]
			for i in range(0,len(docNoSales)):
				# docNoSales[i].replaceAll("[^a-zA-Z0-9]", ")
				if(int(codTranzactieSales[i])<3):
					# try:
					numere=re.sub("[^0-9]", "",str(docNoSales[i]))
					# except:
						# print(docNoSales[i])
					result = ''.join([i for i in str(docNoSales[i]) if not i.isdigit()])
					docNoSales2.append(numere)
					seriefacturi.append(result)
			#print(seriefacturi)
			# print(docNoSales2)
			initial=0
			final=0
			docNoSales2.sort()
			docNo=[]
			for k in range(0,len(docNoSales2)):
				docNo.append(str(docNoSales2[k]))
			docNo.sort()
			listaunica=list(set(docNoSales2))
			listaunica.sort()
	# print(listaunica)

			for i in range(0,len(listaunica)):
				listafacturi=[]
				print(listaunica[i])
				for j in range(0,len(docNoSales2)):
					if(listaunica[i]==docNoSales2[j]):
						try:
							listafacturi.append(int(docNoSales2[j]))
						except:
							print("nu este factura de vanzare")
			listafacturi=list(set(listafacturi))
			print(listafacturi)
			listafacturi.sort()
			start=[]
			start.append(listaunica[0])
			stop=[]
			try:
				if(int(listaunica[1])-int(listaunica[0])>1):
					stop.append(listaunica[0])
				for k in range(1,len(listaunica)):

					if(int(listaunica[k])-int(listaunica[k-1])==1):
						print("ok")
					else:
						stop.append(listaunica[k-1])
						start.append(listaunica[k])
			except:
				stop.append(listaunica[0])
			if(len(stop)==len(start)):
				print("ok")
			else:
				stop.append(listaunica[len(listaunica)-1])
			print(start,stop)

			# #print(docNoSales)
			for k in range(0,len(start)):
				facturi.cell(row=2+k,column=2).value=start[k]
				facturi.cell(row=2+k,column=3).value=stop[k]
				facturi.cell(row=2+k,column=4).value=2

			# for p in range(0,len(docNoSales2)-1):
			# 	#print(docNo[p])
			# 	if(p==0):
			# 		initial=initial+1
			# 		# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 		facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 		if(int(docNo[p])-int(docNo[p+1])< -1):
			# 			final=final+1
			# 			facturi.cell(row=1+final,column=3).value=docNo[p]
			# 	else:
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])==1 and int(docNo[p])-int(docNo[p+1])==-1):
			# 				print("bailando")
						
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])==1 and int(docNo[p][3:])-int(docNo[p+1][3:])==-1):
			# 					print("bailando")
			# 					None
			# 			except:
			# 				print(None)
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])>1 and int(docNo[p])-int(docNo[p+1])==-1):
			# 				initial=initial+1
			# 				# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 				facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])>1 and int(docNo[p][3:])-int(docNo[p+1][3:])==-1):
			# 					initial=initial+1
			# 					# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 					facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 			except:
			# 				print(docNo[p])
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])==1 and int(docNo[p])-int(docNo[p+1])<-1):
			# 				final=final+1
			# 				facturi.cell(row=1+final,column=3).value=docNo[p]
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])==1 and int(docNo[p][3:])-int(docNo[p+1][3:])<-1):
			# 					final=final+1
			# 					facturi.cell(row=1+final,column=3).value=docNo[p]
			# 			except:
			# 				print("none")
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])>1 and int(docNo[p])-int(docNo[p+1])<-1):
			# 				initial=initial+1
			# 				# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 				facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 				final=final+1
			# 				facturi.cell(row=1+final,column=3).value=docNo[p]
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])>1 and int(docNo[p][3:])-int(docNo[p+1][3:])<-1):
			# 					initial=initial+1
			# 					# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 					facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 					final=final+1
			# 					facturi.cell(row=1+final,column=3).value=docNo[p]
			# 			except:
			# 				print("none")

			x=facturi.max_row
			facturi.auto_filter.ref = "A1:C1"
			# if(int(docNoSales2[len(docNoSales2)-1])-int(docNoSales2[len(docNoSales2)-2])>1):
			# 	facturi.cell(row=x+1,column=1).value=seriefacturi[0]
			# 	facturi.cell(row=x+1,column=2).value=docNoSales2[len(docNoSales2)-1]
			# 	facturi.cell(row=x+1,column=3).value=docNoSales2[len(docNoSales2)-1]
			# else:
			# 	facturi.cell(row=x+1,column=1).value=seriefacturi[0]
			# 	facturi.cell(row=x, column=3).value = docNoSales2[len(docNoSales2) - 1]


			yy=facturi.max_row+2
			facturi.cell(row=yy,column=1).value="Serie Alocate"
			facturi.cell(row=yy,column=2).value="Inceput Alocate"
			facturi.cell(row=yy,column=3).value="Final Alocate"
			facturi.cell(row=yy,column=4).value="Tip Alocate"		
			for kk in range(1,5):
				facturi.cell(row=yy,column=kk).font=cap_tabel
				facturi.cell(row=yy,column=kk).fill=cap_tabel_color_black	
			for pp in range(2,yy):
				facturi.cell(row=yy+pp-1,column=2).value=facturi.cell(row=pp,column=2).value
				facturi.cell(row=yy+pp-1,column=3).value=facturi.cell(row=pp,column=3).value
				facturi.cell(row=yy+pp-1,column=4).value=1						
			a23=temp.create_sheet("Sectiunea 2.3,2.4")
			dv = DataValidation(
				type='list', formula1='"Yes,No"', allow_blank=True,showDropDown=False)		
			dv.add(a23["A24"])			
			a23.sheet_view.showGridLines = False
			a23.column_dimensions['A'].width=18
			a23.column_dimensions['B'].width=13		
			a23.cell(row=1,column=1).value="Sectiunea 2.3"
			a23.cell(row=3,column=1).value="Denumire beneficiar"
			a23.cell(row=3,column=2).value="CUI beneficiar"
			a23.cell(row=3,column=1).fill=cap_tabel_color_black
			a23.cell(row=3,column=1).font=cap_tabel		
			a23.cell(row=36,column=1).border=border_bottom
			a23.cell(row=36,column=2).border=border_lowerleft
			a23.cell(row=35,column=2).border=border_right
			a23.cell(row=34,column=2).border=border_right
			a23.cell(row=33,column=2).border=border_right
			a23.cell(row=32,column=2).border=border_right						
			a23.cell(row=3,column=2).fill=cap_tabel_color_black
			a23.cell(row=3,column=2).font=cap_tabel


			a23.cell(row=6,column=1).value="Seria"
			a23.cell(row=6,column=2).value="De la"
			a23.cell(row=6,column=3).value="La"



			for row in a23['A6:C6']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in a23['A14:C14']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			a23.cell(row=9,column=1).value="Sectiunea 2.4"
			a23.cell(row=9,column=1).font=cap_tabelbold
			a23.cell(row=1,column=1).font=cap_tabelbold		
			a23.cell(row=28,column=1).font=cap_tabelbold		
			a23.cell(row=4,column=1).border=border_lowerright
			a23.cell(row=4,column=2).border=border_lowerright		
			a23.cell(row=12,column=1).border=border_lowerright
			a23.cell(row=12,column=2).border=border_lowerright		
			a23.cell(row=11,column=1).fill=cap_tabel_color_black
			a23.cell(row=11,column=1).font=cap_tabel		
			a23.cell(row=7,column=1).border=border_lowerright
			a23.cell(row=7,column=2).border=border_lowerright
			a23.cell(row=7,column=3).border=border_lowerright
			a23.cell(row=15,column=1).border=border_lowerright
			a23.cell(row=15,column=2).border=border_lowerright
			a23.cell(row=15,column=3).border=border_lowerright				
			a23.cell(row=11,column=2).fill=cap_tabel_color_black
			a23.cell(row=11,column=2).font=cap_tabel
			a23.cell(row=11,column=1).value="Denumire tert"

			a23.cell(row=11,column=2).value="CUI tert"

			a23.cell(row=14,column=1).value="Seria"
			a23.cell(row=14,column=2).value="De la"
			a23.cell(row=14,column=3).value="La"
			for row in a23['A31:B31']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
			try:
				if int(val4)== 1:
					a3=temp.create_sheet("Sectiunea 3")
					a3.cell(row=1,column=1).value="Sectiunea 3"
					a3.cell(row=1,column=1).font=cap_tabeltitlu
					a3.cell(row=3,column=1).value="In cazul in care soldul sumei negative inregistrate in decontul de TVA aferent perioadei de raportare este solicitat la rambursare , se vor selecta datele cu privire la natura operatiunilor din care provine acesta"
					a3.cell(row=7,column=1).value="Achizitii de bunuri si servicii legate direct de bunurile imobile din urmatoarele categorii"
					a3.cell(row=8,column=1).value="a) parcuri eoliene"
					a3.cell(row=9,column=1).value="b) constructii rezidentiale"



					a3.cell(row=10,column=1).value="c) cladiri de birouri"
					a3.cell(row=11,column=1).value="d) constructii industriale"
					a3.cell(row=12,column=1).value="e) altele"
					a3.cell(row=14,column=1).value="Achizitii de bunuri, cu exceptia celor legate direct de bunuri imobile:"
					a3.cell(row=15,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=16,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=17,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=18,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=19,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=21,column=1).value="Achizitii de servicii, cu exceptia celor legate direct de bunurile imobile:"
					a3.cell(row=22,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=23,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=24,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=25,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=26,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=28,column=1).value="Importuri de bunuri"
					a3.cell(row=30,column=1).value="Achizitii imobilizari necorporale"
					a3.cell(row=32,column=1).value="Livrari de bunuri imobile"
					a3.cell(row=34,column=1).value="Livrari de bunuri, cu exceptia bunurilor imobile:"
					a3.cell(row=35,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=36,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=37,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=38,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=39,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=41,column=1).value="Livrari de bunuri scutite de TVA"
					a3.cell(row=43,column=1).value="Livrari de bunuri/prestari de servicii pt care se aplica taxarea inversa"
					a3.cell(row=45,column=1).value="Prestari de servicii:"
					a3.cell(row=46,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=47,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=48,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=49,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=50,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=52,column=1).value="Prestari de servicii scutite de TVA"
					a3.cell(row=54,column=1).value="Livrari intracomunitare de bunuri"
					a3.cell(row=56,column=1).value="Prestari intracomunitare de servicii"
					a3.cell(row=58,column=1).value="Exporturi de bunuri"
					a3.cell(row=60,column=1).value="Livrari imobilizari necorporale"
					a3.cell(row=62,column=1).value="Persoana impozabila nu a efectuat livrari de bunuri/prestari de servicii in perioada de raportare"																					
			except:
				pass


																																																																						
			a5=temp.create_sheet("Sectiunea 5")
			a5.sheet_view.showGridLines = False
			a5.cell(row=2,column=1).value="Sectiunea 5"
			a5.cell(row=2,column=1).font=cap_tabeltitlu
			a5.cell(row=4,column=1).font=cap_tabelbold
			a5.cell(row=6,column=1).font=cap_tabelbold
			a5.column_dimensions['B'].width=13
			a5.merge_cells('A6:J7')
			a5.merge_cells('A18:J19')		
			
			for pop in range(1,11):
				a5.cell(row=5,column=pop).border=border_bottom
				a5.cell(row=7,column=pop).border=border_bottom
				a5.cell(row=17,column=pop).border=border_bottom
				a5.cell(row=19,column=pop).border=border_bottom
			a5.cell(row=6,column=10).border=border_upperright
			a5.cell(row=7,column=10).border=border_lowerright
			a5.cell(row=18,column=10).border=border_upperright
			a5.cell(row=19,column=10).border=border_lowerright									
			a5['A6'].alignment=Alignment(wrap_text=True)
			a5['A18'].alignment=Alignment(wrap_text=True)
			a5['A30'].alignment=Alignment(wrap_text=True)						
			a5.cell(row=4,column=1).value="Sectiune 5.2"
			a5.cell(row=16,column=1).font=cap_tabelbold		
			a5.cell(row=16,column=1).value="Sectiune 5.3"
			a5.cell(row=28,column=1).font=cap_tabelbold		
			a5.cell(row=10,column=1).value="cota 24%"
			a5.cell(row=11,column=1).value="cota 20%"
			a5.cell(row=12,column=1).value="cota 19%"
			a5.cell(row=13,column=1).value="cota 9%"
			a5.cell(row=14,column=1).value="cota 5%"
			a5.cell(row=10,column=2).value=0
			a5.cell(row=11,column=2).value=0
			a5.cell(row=12,column=2).value=0
			a5.cell(row=13,column=2).value=0
			a5.cell(row=14,column=2).value=0

			a5.cell(row=9,column=2).value="Valoare TVA"
			a5.cell(row=9,column=1).value="Cota"		

			a5.cell(row=6,column=1).value="5.2 TVA deductibila aferenta facturilor achitate in perioada de raportare indiferent de data in care acestea au fost primite de la persoane impozabile care aplica sistemul normal de TVA, defalcata pe fiecare cota de TVA"
			a5.cell(row=18,column=1).font=cap_tabelbold		

			a5.cell(row=22,column=1).value="cota 24%"
			a5.cell(row=23,column=1).value="cota 20%"
			a5.cell(row=24,column=1).value="cota 19%"
			a5.cell(row=25,column=1).value="cota 9%"
			a5.cell(row=26,column=1).value="cota 5%"

			a5.cell(row=22,column=2).value=0
			a5.cell(row=23,column=2).value=0
			a5.cell(row=24,column=2).value=0
			a5.cell(row=25,column=2).value=0
			a5.cell(row=26,column=2).value=0

			a5.cell(row=21,column=1).value="Cota"		
			a5.cell(row=21,column=2).value="Valoare TVA"
			for row in a5['A9:B9']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in a5['A21:B21']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
								
			a5.cell(row=18,column=1).value="5.3 TVA deductibila aferenta facturilor achitate in perioada de raportare indiferent de data in care acestea au fost primite de la persoane impozabile care aplica sistemul de TVA la incasare, defalcata pe fiecare cota de TVA"

			for kk in range(22,27):
				a5.cell(row=kk,column=1).border=border_lowerright
				a5.cell(row=kk,column=2).border=border_lowerright						
			for kk in range(10,15):
				a5.cell(row=kk,column=1).border=border_lowerright
				a5.cell(row=kk,column=2).border=border_lowerright		
			a6=temp.create_sheet("Sectiunea 6")
			a6.sheet_view.showGridLines = False
			a6.cell(row=2,column=1).value="Sectiunea 6"
			a6.cell(row=2,column=1).font=cap_tabeltitlu
			a6.cell(row=4,column=1).font=cap_tabelbold
			a6.cell(row=6,column=1).font=cap_tabelbold
			a6.cell(row=15,column=1).font=cap_tabelbold
			a6.cell(row=13,column=1).font=cap_tabelbold
			a6.column_dimensions['A'].width=17			
			a6.column_dimensions['B'].width=28
			a6.column_dimensions['C'].width=17
																
			a6.cell(row=4,column=1).value="Sectiunea 6.1"		
			a6.cell(row=6,column=1).value="6.1 Persoanele impozabile care aplica regimul special pt agentiile de turism, vor completa:"
			a6.merge_cells('A6:F7')
			for l in range(1,7):
				a6.cell(row=5,column=l).border=border_bottom
				a6.cell(row=7,column=l).border=border_bottom
			a6.cell(row=6,column=6).border=border_upperright
			a6.cell(row=7,column=6).border=border_lowerright						

			a6.merge_cells('A15:G16')

			for l in range(1,8):
				a6.cell(row=14,column=l).border=border_bottom
				a6.cell(row=16,column=l).border=border_bottom
			a6.cell(row=15,column=7).border=border_upperright
			a6.cell(row=16,column=7).border=border_lowerright




			a6.cell(row=9,column=1).value="Incasarile agentiei"
			a6.cell(row=9,column=2).value="Costurile agentiei de turism"
			a6.cell(row=9,column=3).value="Marja de profit"
			a6.cell(row=9,column=4).value="TVA"

			a6.cell(row=13,column=1).value="Sectiunea 6.2"

			a6.cell(row=15,column=1).value="6.2 Persoanele impozabile care aplica regimul special pt bunurile second-hand, opere de arta, obiecte de colectie si antichitati , vor completa:"

			for row in a6['A18:D18']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in a6['A9:D9']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')										


			a6.cell(row=18,column=1).value="Pret de vanzare"
			a6.cell(row=18,column=2).value="Pret de cumparare"
			a6.cell(row=18,column=3).value="Marja de profit"	
			a6.cell(row=18,column=4).value="TVA"
			for oo in range(1,5):
				a6.cell(row=10,column=oo).border=border_lowerright
				a6.cell(row=19,column=oo).border=border_lowerright		

			a6['A6'].alignment=Alignment(wrap_text=True)

			a6['A15'].alignment=Alignment(wrap_text=True)
			a7=temp.create_sheet(" Sectiunea 7 ")

			a7.cell(row=2,column=1).value="Sectiunea 7"
			a7.sheet_view.showGridLines = False
			a7.cell(row=2,column=1).font=cap_tabeltitlu
			a7.cell(row=5,column=1).value="7. In situatia in care ati desfasurat, in perioada de raportare, activitati dintre cele inscrise in lista veti selecta activitatea corespunzatoare si veti inscrie valoarea livrarilor/prestarilor, precum si TVA aferenta"
			a7['A5'].alignment=Alignment(wrap_text=True)
			a7.merge_cells('A5:I6')
			a7.cell(row=5,column=1).font=cap_tabelbold
			for ii in range(1,10):
				a7.cell(row=4,column=ii).border=border_bottom
				a7.cell(row=6,column=ii).border=border_bottom
			a7.cell(row=5,column=9).border=border_upperright
			a7.cell(row=6,column=9).border=border_lowerright		
			a7.cell(row=5,column=8).border=border_right
			a7.cell(row=6,column=8).border=border_right

			for jj in range(1,4):
				a7.cell(row=9,column=jj).border=border_bottom
				a7.cell(row=9,column=jj).border=border_right
			
			for pp in range(12,17):
				a7.cell(row=pp,column=2).border=border_lowerright
				a7.cell(row=pp,column=1).border=border_lowerright
			a7.cell(row=8,column=1).value="Activitate"
			a7.cell(row=8,column=2).value="Tip operatiune"
			a7.cell(row=8,column=3).value="Valoarea livrarilor/prestarilor"

			a7.cell(row=11,column=1).value="Cota"
			a7.cell(row=12,column=1).value="cota 24%"
			a7.cell(row=13,column=1).value="cota 20%"
			a7.cell(row=14,column=1).value="cota 19%"
			a7.cell(row=15,column=1).value="cota 9%"
			a7.cell(row=16,column=1).value="cota 5%"
			for row in a7['A8:C8']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font = cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')										
			for row in a7['A11:B11']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font = cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			a7.cell(row=11,column=2).value="Valoare TVA"

			for row in facturi['A1:D1']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in facturi['A1:D1']:
				for cell in row:
					cell.font = cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			facturi.column_dimensions['C'].width = 14
			facturi.column_dimensions['D'].width = 14
			a23.column_dimensions['A'].width = 18
			a23.column_dimensions['B'].width = 13
			a23.column_dimensions['C'].width = 10
			a6.column_dimensions['A'].width = 17
			a6.column_dimensions['B'].width = 25
			a6.column_dimensions['C'].width = 15				
			sumaryG=temp.create_sheet("Sectiunea G. Manual input")
			sumaryG.sheet_view.showGridLines = False
			sumaryG.cell(row=2,column=1).value="Sectiunea G"
			sumaryG.cell(row=2,column=1).font=cap_tabeltitlu		
			sumaryG.column_dimensions['A'].width =45
			sumaryG.cell(row=5,column=1).value="Total Nr. Bonuri Fiscale"
			sumaryG.cell(row=5,column=1).font=cap_tabelbold
			sumaryG.cell(row=6,column=1).font=cap_tabelbold
			sumaryG.cell(row=7,column=1).font=cap_tabelbold						
			sumaryG.cell(row=6,column=1).value="Total incasari in perioada de raportare prin intermediul AMEF ( aparate de marcatelectronice fiscale ) inclusiv incasarile prin intermediul bonurilor fiscale care indeplinesc conditiile unei facturi simplificate indiferent daca au/nu au inscris codul de inregistrare in scopuri de TVA al beneficiarului (i1)"
			sumaryG['A6'].alignment=Alignment(wrap_text=True)
			sumaryG['A7'].alignment=Alignment(wrap_text=True)				
			sumaryG.cell(row=7,column=1).value="Total incasari in perioada de raportare efectuate din activitati exceptate de la obligatia utilizarii AMEF***) (i2) conform prevederilor legale in vigoare )"
			sumaryG.merge_cells('A9:H11')
			sumaryG.cell(row=9,column=1).font=cap_tabelbold
			for ii in range(1,9):
				sumaryG.cell(row=8,column=ii).border=border_bottom
				sumaryG.cell(row=11,column=ii).border=border_bottom
			sumaryG.cell(row=9,column=8).border=border_upperright
			sumaryG.cell(row=10,column=8).border=border_right		
			sumaryG.cell(row=11,column=8).border=border_lowerright		
			for ii in range(1,9):	
				sumaryG.cell(row=19,column=ii).border=border_bottom
				sumaryG.cell(row=20,column=ii).border=border_bottom			
			sumaryG.cell(row=5,column=2).border=border_right
			sumaryG.cell(row=6,column=2).border=border_right
			sumaryG.cell(row=7,column=2).border=border_right
			sumaryG.cell(row=5,column=1).border=border_right
			sumaryG.cell(row=6,column=1).border=border_right
			sumaryG.cell(row=7,column=1).border=border_right
			sumaryG.cell(row=4,column=1).border=border_bottom
			sumaryG.cell(row=5,column=1).border=border_bottom
			sumaryG.cell(row=6,column=1).border=border_bottom
			sumaryG.cell(row=7,column=1).border=border_bottom
			sumaryG.cell(row=5,column=2).value=0
			sumaryG.cell(row=6,column=2).value=0
			sumaryG.cell(row=7,column=2).value=0
			sumaryG.cell(row=13,column=1).value="Cota"		
			sumaryG.cell(row=13,column=2).value="Total baza impozabila"
			sumaryG.cell(row=13,column=3).value="TVA"
			sumaryG.cell(row=23,column=1).value="Cota"		
			sumaryG.cell(row=23,column=2).value="Total baza impozabila"
			sumaryG.cell(row=23,column=3).value="TVA"		
			for ii in range(1,4):
				sumaryG.cell(row=17,column=ii).border=border_lowerright
				sumaryG.cell(row=16,column=ii).border=border_lowerright
				sumaryG.cell(row=15,column=ii).border=border_lowerright
				sumaryG.cell(row=14,column=ii).border=border_lowerright
				sumaryG.cell(row=24,column=ii).border=border_lowerright
				sumaryG.cell(row=25,column=ii).border=border_lowerright
				sumaryG.cell(row=26,column=ii).border=border_lowerright
				sumaryG.cell(row=27,column=ii).border=border_lowerright																					
							
			for row in sumaryG['A13:C13']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryG['A23:C23']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')										
			sumaryG.cell(row=9,column=1).value="Incasari in perioada de raportare prin intermediul AMEF ( aparate de marcat electronice fiscale ) inclusiv incasarile prin intermediul bonurilor fiscale care indeplinesc conditiile unei facturi simplificate indiferent daca au/nu au inscris codul de inregistrare in scopuri de TVA al beneficiarului (i1)"
			sumaryG.cell(row=20,column=1).value="Incasari in perioada de raportare efectuate din activitati exceptate de la obligatia utilizarii AMEF***) (i2) conform prevederilor legale in vigoare )"
			sumaryG['A9'].alignment=Alignment(wrap_text=True)		
			sumaryG.cell(row=14,column=1).value="Cota 20%"
			sumaryG.cell(row=15,column=1).value="Cota 19%"
			sumaryG.cell(row=16,column=1).value="Cota 9%"
			sumaryG.cell(row=17,column=1).value="Cota 5%"

			sumaryG.cell(row=24,column=1).value="Cota 20%"
			sumaryG.cell(row=25,column=1).value="Cota 19%"
			sumaryG.cell(row=26,column=1).value="Cota 9%"
			sumaryG.cell(row=27,column=1).value="Cota 5%"

			sumaryG.cell(row=14,column=2).value=0
			sumaryG.cell(row=15,column=2).value=0
			sumaryG.cell(row=16,column=2).value=0
			sumaryG.cell(row=17,column=2).value=0


			sumaryG.cell(row=14,column=3).value=0
			sumaryG.cell(row=15,column=3).value=0
			sumaryG.cell(row=16,column=3).value=0
			sumaryG.cell(row=17,column=3).value=0


			sumaryG.cell(row=24,column=2).value=0
			sumaryG.cell(row=25,column=2).value=0
			sumaryG.cell(row=26,column=2).value=0
			sumaryG.cell(row=27,column=2).value=0

			sumaryG.cell(row=24,column=3).value=0
			sumaryG.cell(row=25,column=3).value=0
			sumaryG.cell(row=26,column=3).value=0
			sumaryG.cell(row=27,column=3).value=0


			sumaryI=temp.create_sheet("Sectiunea I 1. Manual input")

			for row in sumaryI['A7:C7']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A17:C17']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A27:C27']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A37:C37']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A47:C47']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')																
			sumaryI.cell(row=2,column=1).value="Sectiunea I"
			sumaryI.cell(row=2,column=1).font=cap_tabeltitlu
			sumaryI.sheet_view.showGridLines = False
			sumaryI.cell(row=7,column=1).value="Cota"
			sumaryI.cell(row=7,column=2).value="Baza impozabila"
			sumaryI.cell(row=7,column=3).value="TVA"

			sumaryI.cell(row=7,column=1).value="Cota"
			sumaryI.cell(row=7,column=2).value="Baza impozabila"
			sumaryI.cell(row=7,column=3).value="TVA"
			sumaryI.cell(row=17,column=1).value="Cota"
			sumaryI.cell(row=17,column=2).value="Baza impozabila"
			sumaryI.cell(row=17,column=3).value="TVA"
			sumaryI.cell(row=27,column=1).value="Cota"
			sumaryI.cell(row=27,column=2).value="Baza impozabila"
			sumaryI.cell(row=27,column=3).value="TVA"
			sumaryI.cell(row=37,column=1).value="Cota"
			sumaryI.cell(row=37,column=2).value="Baza impozabila"
			sumaryI.cell(row=37,column=3).value="TVA"								
			sumaryI.cell(row=47,column=1).value="Cota"
			sumaryI.cell(row=47,column=2).value="Baza impozabila"
			sumaryI.cell(row=47,column=3).value="TVA"
			sumaryI.cell(row=4,column=1).value="1.1 Livrari de bunuri/prestari de servicii pentru care s-au emis facturi simplificate care au inscris codul de inregistrare in scopuri de TVA al beneficiarului %"
			sumaryI.cell(row=4,column=1).font=cap_tabelbold
			sumaryI.cell(row=14,column=1).font=cap_tabelbold
			sumaryI.cell(row=24,column=1).font=cap_tabelbold
			sumaryI.cell(row=34,column=1).font=cap_tabelbold
			sumaryI.cell(row=44,column=1).font=cap_tabelbold								
			sumaryI.cell(row=14,column=1).value="1.2 Livrari de bunuri/prestari de servicii pentru care s-au emis facturi simplificate fara a avea inscris codul de inregistrare in scopuri de TVA al beneficiarului %"
			sumaryI['A4'].alignment=Alignment(wrap_text=True)
			sumaryI['A14'].alignment=Alignment(wrap_text=True)
			sumaryI['A24'].alignment=Alignment(wrap_text=True)
			sumaryI['A34'].alignment=Alignment(wrap_text=True)
			sumaryI['A44'].alignment=Alignment(wrap_text=True)								
			sumaryI.cell(row=24,column=1).value="1.3 Achizitii de bunuri si servicii pentru care s-au primit facturi simplificate de la persoane impozabile care aplica sistemul normal de TVA si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

			sumaryI.cell(row=34,column=1).value="1.4 Achizitii de bunuri si servicii pentru care s-au primit facturi simplificate de la persoane impozabile care aplica sistemul de TVA la incasare si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

			sumaryI.cell(row=44,column=1).value="1.5 Achizitii de bunuri si servicii pentru care s-au primit bonuri fiscale care indeplinesc conditiile unei facturi simplificate si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

			for ip in range(1,11):
				sumaryI.cell(row=3,column=ip).border=border_bottom
				sumaryI.cell(row=6,column=ip).border=border_top
				sumaryI.cell(row=13,column=ip).border=border_bottom
				sumaryI.cell(row=16,column=ip).border=border_top
				sumaryI.cell(row=23,column=ip).border=border_bottom
				sumaryI.cell(row=26,column=ip).border=border_top
				sumaryI.cell(row=33,column=ip).border=border_bottom
				sumaryI.cell(row=36,column=ip).border=border_top
				sumaryI.cell(row=43,column=ip).border=border_bottom
				sumaryI.cell(row=46,column=ip).border=border_top
			sumaryI.cell(row=4,column=10).border=border_upperright
			sumaryI.cell(row=5,column=10).border=border_lowerright
			sumaryI.cell(row=14,column=10).border=border_upperright
			sumaryI.cell(row=15,column=10).border=border_lowerright
			sumaryI.cell(row=24,column=10).border=border_upperright
			sumaryI.cell(row=25,column=10).border=border_lowerright
			sumaryI.cell(row=34,column=10).border=border_upperright
			sumaryI.cell(row=35,column=10).border=border_lowerright
			sumaryI.cell(row=44,column=10).border=border_upperright
			sumaryI.cell(row=45,column=10).border=border_lowerright						

			sumaryI.cell(row=4,column=10).border=border_upperright
			sumaryI.cell(row=5,column=10).border=border_lowerright
			sumaryI.cell(row=14,column=10).border=border_upperright
			sumaryI.cell(row=15,column=10).border=border_lowerright
			sumaryI.cell(row=24,column=10).border=border_upperright
			sumaryI.cell(row=25,column=10).border=border_lowerright
			sumaryI.cell(row=34,column=10).border=border_upperright
			sumaryI.cell(row=35,column=10).border=border_lowerright
			sumaryI.cell(row=44,column=10).border=border_upperright
			sumaryI.cell(row=45,column=10).border=border_lowerright										
			sumaryI.cell(row=4,column=10).border=border_right
			sumaryI.cell(row=5,column=10).border=border_right
			sumaryI.cell(row=14,column=10).border=border_right
			sumaryI.cell(row=15,column=10).border=border_right
			sumaryI.cell(row=24,column=10).border=border_right
			sumaryI.cell(row=25,column=10).border=border_right
			sumaryI.cell(row=34,column=10).border=border_right
			sumaryI.cell(row=35,column=10).border=border_right
			sumaryI.cell(row=44,column=10).border=border_right
			sumaryI.cell(row=45,column=10).border=border_right																																																	
			for io in range(8,13):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(18,23):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(28,33):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(38,43):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(48,53):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright																		
			sumaryI.merge_cells('A4:J5')
			sumaryI.merge_cells('A14:J15')
			sumaryI.merge_cells('A24:J25')
			sumaryI.merge_cells('A34:J35')
			sumaryI.merge_cells('A44:J45')
									
			sumaryI.cell(row=8,column=1).value="Cota 24%"
			sumaryI.cell(row=9,column=1).value="Cota 20%"
			sumaryI.cell(row=10,column=1).value="Cota 19%"
			sumaryI.cell(row=11,column=1).value="Cota 9%"
			sumaryI.cell(row=12,column=1).value="Cota 5%"

			sumaryI.cell(row=8,column=2).value=0
			sumaryI.cell(row=9,column=2).value=0
			sumaryI.cell(row=10,column=2).value=0
			sumaryI.cell(row=11,column=2).value=0
			sumaryI.cell(row=12,column=2).value=0

			sumaryI.cell(row=18,column=2).value=0
			sumaryI.cell(row=19,column=2).value=0
			sumaryI.cell(row=20,column=2).value=0
			sumaryI.cell(row=21,column=2).value=0
			sumaryI.cell(row=22,column=2).value=0

			sumaryI.cell(row=28,column=2).value=0
			sumaryI.cell(row=29,column=2).value=0
			sumaryI.cell(row=30,column=2).value=0
			sumaryI.cell(row=31,column=2).value=0
			sumaryI.cell(row=32,column=2).value=0

			sumaryI.cell(row=38,column=2).value=0
			sumaryI.cell(row=39,column=2).value=0
			sumaryI.cell(row=40,column=2).value=0
			sumaryI.cell(row=41,column=2).value=0
			sumaryI.cell(row=42,column=2).value=0

			sumaryI.cell(row=48,column=2).value=0
			sumaryI.cell(row=49,column=2).value=0
			sumaryI.cell(row=50,column=2).value=0
			sumaryI.cell(row=51,column=2).value=0
			sumaryI.cell(row=52,column=2).value=0

			sumaryI.cell(row=8,column=3).value=0
			sumaryI.cell(row=9,column=3).value=0
			sumaryI.cell(row=10,column=3).value=0
			sumaryI.cell(row=11,column=3).value=0
			sumaryI.cell(row=12,column=3).value=0

			sumaryI.cell(row=18,column=3).value=0
			sumaryI.cell(row=19,column=3).value=0
			sumaryI.cell(row=20,column=3).value=0
			sumaryI.cell(row=21,column=3).value=0
			sumaryI.cell(row=22,column=3).value=0

			sumaryI.cell(row=28,column=3).value=0
			sumaryI.cell(row=29,column=3).value=0
			sumaryI.cell(row=30,column=3).value=0
			sumaryI.cell(row=31,column=3).value=0
			sumaryI.cell(row=32,column=3).value=0

			sumaryI.cell(row=38,column=3).value=0
			sumaryI.cell(row=39,column=3).value=0
			sumaryI.cell(row=40,column=3).value=0
			sumaryI.cell(row=41,column=3).value=0
			sumaryI.cell(row=42,column=3).value=0

			sumaryI.cell(row=48,column=3).value=0
			sumaryI.cell(row=49,column=3).value=0
			sumaryI.cell(row=50,column=3).value=0
			sumaryI.cell(row=51,column=3).value=0
			sumaryI.cell(row=52,column=3).value=0		

			sumaryI.cell(row=8,column=1).value="Cota 24%"
			sumaryI.cell(row=9,column=1).value="Cota 20%"
			sumaryI.cell(row=10,column=1).value="Cota 19%"
			sumaryI.cell(row=11,column=1).value="Cota 9%"
			sumaryI.cell(row=12,column=1).value="Cota 5%"

			sumaryI.cell(row=18,column=1).value="Cota 24%"
			sumaryI.cell(row=19,column=1).value="Cota 20%"
			sumaryI.cell(row=20,column=1).value="Cota 19%"
			sumaryI.cell(row=21,column=1).value="Cota 9%"
			sumaryI.cell(row=22,column=1).value="Cota 5%"

			sumaryI.cell(row=28,column=1).value="Cota 24%"
			sumaryI.cell(row=29,column=1).value="Cota 20%"
			sumaryI.cell(row=30,column=1).value="Cota 19%"
			sumaryI.cell(row=31,column=1).value="Cota 9%"
			sumaryI.cell(row=32,column=1).value="Cota 5%"

			sumaryI.cell(row=38,column=1).value="Cota 24%"
			sumaryI.cell(row=39,column=1).value="Cota 20%"
			sumaryI.cell(row=40,column=1).value="Cota 19%"
			sumaryI.cell(row=41,column=1).value="Cota 9%"
			sumaryI.cell(row=42,column=1).value="Cota 5%"

			sumaryI.cell(row=48,column=1).value="Cota 24%"
			sumaryI.cell(row=49,column=1).value="Cota 20%"
			sumaryI.cell(row=50,column=1).value="Cota 19%"
			sumaryI.cell(row=51,column=1).value="Cota 9%"
			sumaryI.cell(row=52,column=1).value="Cota 5%"

			sumaryI.column_dimensions['B'].width = 15	
			sumaryG.column_dimensions['B'].width = 20	
			a7.column_dimensions['B'].width = 12						
			a7.column_dimensions['C'].width = 22
			a6.column_dimensions['B'].width = 22						
			a6.column_dimensions['C'].width = 12


			# for row in sumary['A5:D34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['F5:I34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['K5:N34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['P5:S34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['U5:X34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# listanoua=['A','B','C','D','F','G','H','I','K','L','M','N','P','Q','R','S','U','V','W','W','X']
			# for column in listanoua:
			# 	for i in listanoua:
			# 		if (column==i):
	# 			sumary.column_dimensions[column].width = 15

			
			# for i in range(0 ,len(tip)):
		# folderpath="D:/D300 to XML/docs"
		folderpath="/home/mirus_app/storage_spreadsheet"
		# folderpath="C:/Users/Cristian.Iordache/Documents/D300 to XML Final CI/D300 to XML 2/storage"
		file_pathFS = os.path.join(folderpath, "One VAT app spreadsheets " +str(clientname)+".xlsx")
		temp.save(file_pathFS)
		# return send_from_directory("D:/D300 to XML/docs","One VAT app spreadsheets.xlsx",as_attachment=True)
		return send_from_directory("/home/mirus_app/storage_spreadsheet","One VAT app spreadsheets " +str(clientname)+".xlsx",as_attachment=True)
		return render_template('D3APPS2')@app.route('/D3APPS2')
def my_form2():
    return render_template('D3APPS second step.html')


#====================================================================STRAUMANN=============================================================================================
@app.route('/')
def my_form_straumann():
    return render_template('D3APPS dashboard.html')

global LL_g
@app.route('/D3APPS/STRAUMANN')
def my_form_D300_straumann():
	return render_template('D3APPS.html')

@app.route('/D3APPS/STRAUMANN', methods=['POST', 'GET'])
def D300xml_straumann():
	if request.method == 'POST':
		clientname=request.form.get('client')
		D300 = request.files["far"]
		val1 = request.form.get('D300')
		val2 = request.form.get('D390')
		val3 = request.form.get('D394')
		val4 = request.form.get('xyz')
		dropdown = request.form.get('trezorerie')
		dropdownlimba = request.form.get('limba')
		soldLunaTrecuta = request.form.get('largeAm')

	
		# #print(soldLunaTrecuta)
	if val1=="":
		# #print("Da")  # daca e bifat
		val1 = 1
	else:
		#print(val1)            
		val1 = 0
		# #print("Nu")

	if val2=="":  # daca e bifat
		val2 = 1
	else:
		# #print(val2)            
		val2 = 0

	if val3=="":  # daca e bifat
		val3 = 1
	else:
		#print(val3)            
		val3 = 0
		
	if str(dropdownlimba)=="Romana(RO)":
		option=1
	else:
		option=0	


	#print(val4)
	#print(val4,"--------------------------------------------")		
	cap_tabel_color_verde = PatternFill(start_color = '00B050', end_color ='00B050', fill_type = 'solid')
	cap_tabel_color_verde_deschis = PatternFill(start_color = '92D050', end_color ='92D050', fill_type = 'solid')
	cap_tabel_color_black = PatternFill(start_color = '000000', end_color ='000000', fill_type = 'solid')
	cap_tabel = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
	cap_tabelbold = Font(name='Calibri', size=10, bold=True)
	cap_tabeltitlu = Font(name='Calibri', size=15, bold=True,underline='single')
	scrisincredibildemare = Font(name='Calibri', size=30, bold=True)
	cap_tabel_color_GT_movdeschis = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')
	cap_tabel_color_GT_movinchis = PatternFill(start_color='CCC0DA', end_color='CCC0DA', fill_type='solid')

	temp = openpyxl.load_workbook(D300,data_only=True)
	ws = temp.active
	
	if(option==1):
		Sheet1=temp.create_sheet('Cover sheet')
		fonta = PatternFill(start_color = 'ffffff', end_color ='ffffff', fill_type = 'solid')
		fontg = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		font2 = Font(name = 'Georgia', size = 10, bold = True, color="000000")
		font1 = Font(name = 'Georgia', size = 10, color = "FFFFFF", bold = True,italic=True)
		font3 = Font(name = 'Georgia', size = 10, color = "000000",italic=True)
		culoare = PatternFill(start_color = '182A54', end_color ='182A54', fill_type = 'solid') 
		culoare2 = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		culoare3 = PatternFill(start_color = 'D9E1F2', end_color ='D9E1F2', fill_type = 'solid')
		culoare4 = PatternFill(start_color = 'E2EFDA', end_color ='E2EFDA', fill_type = 'solid')
		culoare5 = PatternFill(start_color = 'FFF2CC', end_color ='FFF2CC', fill_type = 'solid')
		culoare6 = PatternFill(start_color = '808080', end_color ='808080', fill_type = 'solid')
		font4 = Font(name = 'Georgia', size = 10, color = "000000",underline='single',bold=True)
		font5 = Font(name = 'Georgia', size = 10, color = "ffffff",underline='single',bold=True)
		border = Border(bottom=Side(style='dotted'))
		border2 = Border(top=Side(style='dotted'))
		border3 = Border(left=Side(style='dotted'))
		border4 = Border(right=Side(style='dotted'))
		border5 = Border(left=Side(style='dotted'),top=Side(style='dotted'))
		border6 = Border(left=Side(style='dotted'),bottom=Side(style='dotted'))
		border7 = Border(right=Side(style='dotted'),bottom=Side(style='dotted'))
		border8 = Border(right=Side(style='dotted'),top=Side(style='dotted'))
		# border9 = Border(right=Side(style='double'),bottom=Side(style='double'),top=Side(style='double'),left=Side(style='double'))


		Sheet1.cell(row=31, column=2).border=border2
		Sheet1.cell(row=31, column=3).border=border2
		Sheet1.cell(row=31, column=4).border=border2
		Sheet1.cell(row=31, column=5).border=border2
		Sheet1.cell(row=31, column=6).border=border2
		Sheet1.cell(row=31, column=7).border=border2
		Sheet1.cell(row=31, column=2).border=border5
		Sheet1.cell(row=31, column=7).border=border8


		Sheet1.cell(row=75, column=2).border=border
		Sheet1.cell(row=75, column=3).border=border
		Sheet1.cell(row=75, column=4).border=border
		Sheet1.cell(row=75, column=5).border=border
		Sheet1.cell(row=75, column=6).border=border
		Sheet1.cell(row=75, column=7).border=border
		Sheet1.cell(row=75, column=7).border=border7

		Sheet1.cell(row=32, column=2).border=border3
		Sheet1.cell(row=33, column=2).border=border3
		Sheet1.cell(row=34, column=2).border=border3
		Sheet1.cell(row=35, column=2).border=border3
		Sheet1.cell(row=36, column=2).border=border3
		Sheet1.cell(row=37, column=2).border=border3
		Sheet1.cell(row=38, column=2).border=border3
		Sheet1.cell(row=39, column=2).border=border3
		Sheet1.cell(row=40, column=2).border=border3
		Sheet1.cell(row=41, column=2).border=border3
		Sheet1.cell(row=42, column=2).border=border3
		Sheet1.cell(row=43, column=2).border=border3
		Sheet1.cell(row=44, column=2).border=border3
		Sheet1.cell(row=45, column=2).border=border3
		Sheet1.cell(row=46, column=2).border=border3
		Sheet1.cell(row=47, column=2).border=border3
		Sheet1.cell(row=48, column=2).border=border3
		Sheet1.cell(row=49, column=2).border=border3
		Sheet1.cell(row=50, column=2).border=border3
		Sheet1.cell(row=51, column=2).border=border3
		Sheet1.cell(row=52, column=2).border=border3
		Sheet1.cell(row=53, column=2).border=border3
		Sheet1.cell(row=54, column=2).border=border3
		Sheet1.cell(row=55, column=2).border=border3
		Sheet1.cell(row=75, column=2).border=border6
		Sheet1.cell(row=56, column=2).border=border3
		Sheet1.cell(row=57, column=2).border=border3
		Sheet1.cell(row=58, column=2).border=border3
		Sheet1.cell(row=59, column=2).border=border3
		Sheet1.cell(row=60, column=2).border=border3
		Sheet1.cell(row=61, column=2).border=border3
		Sheet1.cell(row=62, column=2).border=border3
		Sheet1.cell(row=63, column=2).border=border3
		Sheet1.cell(row=64, column=2).border=border3
		Sheet1.cell(row=65, column=2).border=border3
		Sheet1.cell(row=66, column=2).border=border3
		Sheet1.cell(row=67, column=2).border=border3
		Sheet1.cell(row=68, column=2).border=border3
		Sheet1.cell(row=69, column=2).border=border3
		Sheet1.cell(row=70, column=2).border=border3
		Sheet1.cell(row=71, column=2).border=border3
		Sheet1.cell(row=72, column=2).border=border3
		Sheet1.cell(row=73, column=2).border=border3
		Sheet1.cell(row=74, column=2).border=border3

		Sheet1.cell(row=32, column=7).border=border4
		Sheet1.cell(row=33, column=7).border=border4
		Sheet1.cell(row=34, column=7).border=border4
		Sheet1.cell(row=35, column=7).border=border4
		Sheet1.cell(row=36, column=7).border=border4
		Sheet1.cell(row=37, column=7).border=border4
		Sheet1.cell(row=38, column=7).border=border4
		Sheet1.cell(row=39, column=7).border=border4
		Sheet1.cell(row=40, column=7).border=border4
		Sheet1.cell(row=41, column=7).border=border4
		Sheet1.cell(row=42, column=7).border=border4
		Sheet1.cell(row=43, column=7).border=border4
		Sheet1.cell(row=44, column=7).border=border4
		Sheet1.cell(row=45, column=7).border=border4
		Sheet1.cell(row=46, column=7).border=border4
		Sheet1.cell(row=47, column=7).border=border4
		Sheet1.cell(row=48, column=7).border=border4
		Sheet1.cell(row=49, column=7).border=border4
		Sheet1.cell(row=50, column=7).border=border4
		Sheet1.cell(row=51, column=7).border=border4
		Sheet1.cell(row=52, column=7).border=border4
		Sheet1.cell(row=53, column=7).border=border4
		Sheet1.cell(row=54, column=7).border=border4
		Sheet1.cell(row=55, column=7).border=border4
		Sheet1.cell(row=56, column=7).border=border4
		Sheet1.cell(row=57, column=7).border=border4
		Sheet1.cell(row=58, column=7).border=border4
		Sheet1.cell(row=59, column=7).border=border4
		Sheet1.cell(row=60, column=7).border=border4
		Sheet1.cell(row=61, column=7).border=border4
		Sheet1.cell(row=62, column=7).border=border4
		Sheet1.cell(row=63, column=7).border=border4
		Sheet1.cell(row=64, column=7).border=border4
		Sheet1.cell(row=65, column=7).border=border4
		Sheet1.cell(row=66, column=7).border=border4
		Sheet1.cell(row=67, column=7).border=border4
		Sheet1.cell(row=68, column=7).border=border4
		Sheet1.cell(row=69, column=7).border=border4
		Sheet1.cell(row=70, column=7).border=border4
		Sheet1.cell(row=71, column=7).border=border4
		Sheet1.cell(row=72, column=7).border=border4
		Sheet1.cell(row=73, column=7).border=border4
		Sheet1.cell(row=74, column=7).border=border4
		info=temp['Other info']
		valluna=""
		vallunaurmatoare=""
		valIban=""
		okdecembrie=0
		for i in range(0,len(listaluni)):
			if(str(info.cell(row=3,column=3).value)=="12"):
				okdecembrie=1
				vallunaurmatoare=listadenluni[0]
				valluna=listadenluni[11]
			else:
				if(listaluni[i]==str(info.cell(row=3,column=3).value)):
					valluna=listadenluni[i]
					vallunaurmatoare=listadenluni[i+1]
		var=Sheet1.cell(row=12,column=4).value
		print(var)
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		for j in range(0,len(listatrez)):
			if(listatrez[j]==str(Sheet1.cell(row=10,column=4).value)):
				valIban=listaiban[j]
		Sheet1.cell(row=60, column=3).value='Perioada de plata: '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=61, column=3).value='="Suma de plata: " &D55&" RON "'
		Sheet1.cell(row=62, column=3).value="Moneda: RON"
		Sheet1.cell(row=63, column=3).value='Detalii plata: Decont TVA - '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		if(okdecembrie==1):
			Sheet1.cell(row=64, column=3).value='Data scadenta: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value+1)
		else:
			Sheet1.cell(row=64, column=3).value='Data scadenta: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=66, column=3).value='="Cod TVA: " & D8'
		Sheet1.cell(row=67, column=3).value='="Adresa: " &D7'
		Sheet1.cell(row=69, column=3).value="Beneficiar: BUGETUL DE STAT"
		Sheet1.cell(row=70, column=3).value='Cont IBAN: '+ str(valIban)
		Sheet1.cell(row=71, column=3).value="SWIFT / BIC: TREZROBU"
		Sheet1.cell(row=72, column=3).value="Deschis la:"+str(dropdown)
		Sheet1.cell(row=74, column=3).value="Nota: Orice taxe bancare legate de plata trebuie sa fie acoperite de catre platitor"
		Sheet1.cell(row=74, column=3).font=font2
		

		Sheet1.cell(row=14, column=3).value="Sumar"
		Sheet1.cell(row=14, column=3).font=font4
		Sheet1.cell(row=58, column=3).value="ORDIN DE PLATA"
		Sheet1.cell(row=58, column=3).font=font5

		for row in Sheet1['A1:N100']:
					for cell in row:
						cell.fill = fonta

		for row in Sheet1['N1:Z100']:
					for cell in row:
						cell.fill = fontg
		print(get_fxrate(today.year))

		Sheet1.cell(row = 6, column = 3).value = "Denumire"
		Sheet1.cell(row = 6, column = 4).value = "='Other info'!C4"
		Sheet1.cell(row = 7, column = 3).value = "Adresa"
		Sheet1.cell(row = 7, column = 4).value = "='Other info'!C6"
		Sheet1.cell(row = 8, column = 3).value = "CUI"
		Sheet1.cell(row = 8, column = 4).value = "='Other info'!C5"
		Sheet1['D8'].alignment = Alignment(wrapText=True, horizontal='left')
		Sheet1.cell(row = 9, column = 3).value = "Nr. Reg. Com."
		Sheet1.cell(row = 9, column = 4).value = "J08/1139/2017"
		Sheet1.cell(row = 10, column = 3).value = "Administratia de care apartine"
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		Sheet1.cell(row = 11, column = 3).value = "Frecventa depunere declaratie/plata"
		Sheet1.cell(row = 11, column = 4).value = "Monthly"
		Sheet1.cell(row = 12, column = 3).value = "Perioada de raportare"
		Sheet1.cell(row = 12, column = 4).value = "=date('Other info'!C2,'Other info'!C3,1)"
		Sheet1.cell(row = 12, column = 4).number_format = 'mmmm yyyy'

		Sheet1.cell(row = 43, column = 4).value="Yes"
		Sheet1.cell(row = 46, column = 4).value="N/a"
		Sheet1.cell(row = 47, column = 4).value="N/a"
		Sheet1.cell(row = 50, column = 4).value="N/a"
		Sheet1.cell(row = 51, column = 4).value="N/a"
		Sheet1.cell(row = 43, column = 4).font=font5
		Sheet1.cell(row = 46, column = 4).font=font5
		Sheet1.cell(row = 47, column = 4).font=font5
		Sheet1.cell(row = 50, column = 4).font=font5
		Sheet1.cell(row = 51, column = 4).font=font5


		Sheet1.cell(row = 6, column = 3).font=font1
		Sheet1.cell(row = 6, column = 4).font=font2
		Sheet1.cell(row = 7, column = 3).font=font1
		Sheet1.cell(row = 7, column = 4).font=font2
		Sheet1.cell(row = 8, column = 3).font=font1
		Sheet1.cell(row = 8, column = 4).font=font2
		Sheet1.cell(row = 9, column = 3).font=font1
		Sheet1.cell(row = 9, column = 4).font=font2
		Sheet1.cell(row = 10, column = 3).font=font1
		Sheet1.cell(row = 10, column = 4).font=font3
		Sheet1.cell(row = 11, column = 3).font=font1
		Sheet1.cell(row = 11, column = 4).font=font3
		Sheet1.cell(row = 12, column = 3).font=font1
		Sheet1.cell(row = 12, column = 4).font=font3

		Sheet1.cell(row = 6, column = 3).fill=culoare
		Sheet1.cell(row = 7, column = 3).fill=culoare
		Sheet1.cell(row = 8, column = 3).fill=culoare
		Sheet1.cell(row = 9, column = 3).fill=culoare
		Sheet1.cell(row = 10, column = 3).fill=culoare
		Sheet1.cell(row = 11, column = 3).fill=culoare
		Sheet1.cell(row = 12, column = 3).fill=culoare
		Sheet1.cell(row = 6, column = 4).fill=culoare2
		Sheet1.cell(row = 7, column = 4).fill=culoare2
		Sheet1.cell(row = 8, column = 4).fill=culoare2
		Sheet1.cell(row = 9, column = 4).fill=culoare2
		Sheet1.cell(row = 10, column = 4).fill=culoare2
		Sheet1.cell(row = 11, column = 4).fill=culoare2
		Sheet1.cell(row = 12, column = 4).fill=culoare2

		Sheet1.cell(row = 16, column = 3).fill=culoare
		Sheet1.cell(row=16, column=3).font=font1
		Sheet1.cell(row=16, column=3).value="  D300"
		Sheet1.cell(row=16, column=3).hyperlink="#'D300 draft figures'!A1"
		# Sheet1.cell(row=16, column=3).border=border9
		Sheet1.row_dimensions[18].height=8

		Sheet1.cell(row = 19, column = 3).fill=culoare
		Sheet1.cell(row=19, column=3).font=font1
		Sheet1.cell(row=19, column=3).value="  D390"
		Sheet1.cell(row=19, column=3).hyperlink="#'D390 workings'!A1"
		# Sheet1.cell(row=19, column=3).border=border9
		Sheet1.row_dimensions[21].height=8

		Sheet1.cell(row = 22, column = 3).fill=culoare
		Sheet1.cell(row=22, column=3).font=font1
		Sheet1.cell(row=22, column=3).value="  D394"
		Sheet1.cell(row=22, column=3).hyperlink="#'D394--->>>'!A1"
		# Sheet1.cell(row=22, column=3).border=border9
		Sheet1.row_dimensions[24].height=8

		Sheet1.cell(row = 25, column = 3).fill=culoare
		Sheet1.cell(row=25, column=3).font=font1
		Sheet1.cell(row=25, column=3).value="  Jurnal vanzari"
		Sheet1.cell(row=25, column=3).hyperlink="#'Sales'!A1"
		# Sheet1.cell(row=25, column=3).border=border9
		Sheet1.row_dimensions[27].height=8

		Sheet1.cell(row = 28, column = 3).fill=culoare
		Sheet1.cell(row=28, column=3).font=font1
		Sheet1.cell(row=28, column=3).value="  Jurnal cumparari"
		Sheet1.cell(row=28, column=3).hyperlink="#'Purchases'!A1"
		# Sheet1.cell(row=28, column=3).border=border9
		Sheet1.row_dimensions[30].height=8


		Sheet1.cell(row = 58, column = 3).fill=culoare6
		Sheet1.cell(row = 32, column = 3).fill=culoare3
		Sheet1.cell(row = 33, column = 3).fill=culoare3
		Sheet1.cell(row = 34, column = 3).fill=culoare3
		Sheet1.cell(row = 35, column = 3).fill=culoare3
		Sheet1.cell(row = 36, column = 3).fill=culoare3
		Sheet1.cell(row = 37, column = 3).fill=culoare3
		Sheet1.cell(row = 38, column = 3).fill=culoare3
		Sheet1.cell(row = 39, column = 3).fill=culoare3
		Sheet1.cell(row = 32, column = 4).fill=culoare3
		Sheet1.cell(row = 33, column = 4).fill=culoare3
		Sheet1.cell(row = 34, column = 4).fill=culoare3
		Sheet1.cell(row = 35, column = 4).fill=culoare3
		Sheet1.cell(row = 36, column = 4).fill=culoare3
		Sheet1.cell(row = 37, column = 4).fill=culoare3
		Sheet1.cell(row = 38, column = 4).fill=culoare3
		Sheet1.cell(row = 39, column = 4).fill=culoare3
		Sheet1.cell(row = 32, column = 6).fill=culoare3
		Sheet1.cell(row = 33, column = 6).fill=culoare3
		Sheet1.cell(row = 34, column = 6).fill=culoare3
		Sheet1.cell(row = 35, column = 6).fill=culoare3
		Sheet1.cell(row = 36, column = 6).fill=culoare3
		Sheet1.cell(row = 37, column = 6).fill=culoare3
		Sheet1.cell(row = 38, column = 6).fill=culoare3
		Sheet1.cell(row = 39, column = 6).fill=culoare3

		Sheet1.cell(row = 41, column = 6).fill=culoare4
		Sheet1.cell(row = 42, column = 6).fill=culoare4
		Sheet1.cell(row = 43, column = 6).fill=culoare4
		Sheet1.cell(row = 44, column = 6).fill=culoare4
		Sheet1.cell(row = 45, column = 6).fill=culoare4
		Sheet1.cell(row = 46, column = 6).fill=culoare4
		Sheet1.cell(row = 47, column = 6).fill=culoare4
		Sheet1.cell(row = 48, column = 6).fill=culoare4
		Sheet1.cell(row = 49, column = 6).fill=culoare4
		Sheet1.cell(row = 50, column = 6).fill=culoare4
		Sheet1.cell(row = 51, column = 6).fill=culoare4
		Sheet1.cell(row = 41, column = 3).fill=culoare4
		Sheet1.cell(row = 42, column = 3).fill=culoare4
		Sheet1.cell(row = 43, column = 3).fill=culoare4
		Sheet1.cell(row = 44, column = 3).fill=culoare4
		Sheet1.cell(row = 45, column = 3).fill=culoare4
		Sheet1.cell(row = 46, column = 3).fill=culoare4
		Sheet1.cell(row = 47, column = 3).fill=culoare4
		Sheet1.cell(row = 48, column = 3).fill=culoare4
		Sheet1.cell(row = 49, column = 3).fill=culoare4
		Sheet1.cell(row = 50, column = 3).fill=culoare4
		Sheet1.cell(row = 51, column = 3).fill=culoare4
		Sheet1.cell(row = 41, column = 4).fill=culoare4
		Sheet1.cell(row = 42, column = 4).fill=culoare4
		Sheet1.cell(row = 43, column = 4).fill=culoare6
		Sheet1.cell(row = 44, column = 4).fill=culoare4
		Sheet1.cell(row = 45, column = 4).fill=culoare4
		Sheet1.cell(row = 46, column = 4).fill=culoare6
		Sheet1.cell(row = 47, column = 4).fill=culoare6
		Sheet1.cell(row = 48, column = 4).fill=culoare4
		Sheet1.cell(row = 49, column = 4).fill=culoare4
		Sheet1.cell(row = 50, column = 4).fill=culoare6
		Sheet1.cell(row = 51, column = 4).fill=culoare6

		Sheet1.cell(row = 53, column = 4).fill=culoare5
		Sheet1.cell(row = 54, column = 4).fill=culoare5
		Sheet1.cell(row = 55, column = 4).fill=culoare5
		Sheet1.cell(row = 53, column = 3).fill=culoare5
		Sheet1.cell(row = 54, column = 3).fill=culoare5
		Sheet1.cell(row = 55, column = 3).fill=culoare5
		Sheet1.cell(row = 53, column = 6).fill=culoare5
		Sheet1.cell(row = 54, column = 6).fill=culoare5
		Sheet1.cell(row = 55, column = 6).fill=culoare5


		Sheet1.cell(row = 32, column = 3).value="Pozitia curenta din punct de vedere TVA"
		Sheet1.cell(row = 32, column = 3).font=font4
		Sheet1.cell(row = 34, column = 3).value="Input TVA perioada curenta"
		Sheet1.cell(row = 35, column = 3).value="Output TVA perioada curenta"
		Sheet1.cell(row = 36, column = 3).value="TVA de plata perioada curenta"
		Sheet1.cell(row = 37, column = 3).value="TVA de recuperat perioada curenta"
		Sheet1.cell(row = 38, column = 3).value="TVA in curs de decontare pentru achizitii"
		Sheet1.cell(row = 39, column = 3).value="TVA in curs de decontare pentru livrari"
		Sheet1.cell(row = 32, column = 4).value="RON"

		Sheet1.cell(row = 32, column = 6).value="Euro(@"+get_fxrate(2022)+")"
		Sheet1.cell(row = 34, column = 4).value="='D300 draft figures'!C56"
		Sheet1.cell(row = 34, column = 6).value="=IFERROR(D34/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=34, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=34, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row = 35, column = 4).value="='D300 draft figures'!C32"
		Sheet1.cell(row = 35, column = 6).value="=IFERROR(D35/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 36, column = 4).value="=IF('D300 draft figures'!C58<>0,'D300 draft figures'!C58,0)"
		Sheet1.cell(row = 36, column = 6).value="=IFERROR(D36/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=35, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=35, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=6).number_format = '#,##0_);(#,##0)'

		Sheet1.cell(row=37, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=37, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=6).number_format = '#,##0_);(#,##0)'				

		Sheet1.cell(row = 37, column = 4).value='''=IF('D300 draft figures'!C57<>0,'D300 draft figures'!C57,"nil")'''
		Sheet1.cell(row = 37, column = 6).value="=iferror(D37/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 38, column = 4).value="='D300 draft figures'!C70"
		Sheet1.cell(row = 38, column = 6).value="=iferror(D38/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 39, column = 4).value=0
		Sheet1.cell(row = 39, column = 6).value="=iferror(D39/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 55, column = 4).value='''=IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)=0),'Cover sheet'!D36,
IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)<>0),IF('Cover sheet'!D36>IFERROR(VALUE('Cover sheet'!D47),0),'Cover sheet'!D36-IFERROR(VALUE('Cover sheet'!D47),0),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="No"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D47),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="Yes"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D51),0),"N/A"))))'''
		Sheet1.cell(row = 55, column = 6).value="=iferror(D55/"+get_fxrate(2022)+",0)"

		Sheet1.cell(row = 41, column = 3).value="Pozitia reportata"
		Sheet1.cell(row = 41, column = 3).font=font4
		Sheet1.row_dimensions[42].height = 0.2
		Sheet1.cell(row = 43, column = 3).value="Solicitat la rambursare"
		Sheet1.cell(row = 44, column = 3).value
		Sheet1.cell(row = 45, column = 3).value="TVA de rambursat nesolicitat"
		Sheet1.cell(row = 46, column = 3).value="Perioada"
		Sheet1.cell(row = 47, column = 3).value="Suma"
		Sheet1.cell(row = 48, column = 3).value
		Sheet1.cell(row = 49, column = 3).value="TVA de rambursat solicitat si in curs de auditare"
		Sheet1.cell(row = 50, column = 3).value="Perioada"
		Sheet1.cell(row = 51, column = 3).value="Suma"


		Sheet1.cell(row = 53, column = 3).value="Pozitia balantei de TVA"
		Sheet1.cell(row = 53, column = 3).font=font4
		Sheet1.cell(row = 55, column = 3).value="Pozitia TVA in exercitiul curent"


		Sheet1['C16'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C19'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C22'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C25'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C28'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['D10'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D11'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D12'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['C58'].alignment = Alignment(wrapText=True, horizontal='center')

		Sheet1.column_dimensions['C'].width = 65
		Sheet1.column_dimensions['D'].width = 20
		Sheet1.column_dimensions['A'].width = 2
		Sheet1.column_dimensions['B'].width = 3
		Sheet1.column_dimensions['E'].width = 1
		Sheet1.column_dimensions['F'].width = 20
		Sheet1.column_dimensions['G'].width = 2

		# img= openpyxl.drawing.image.Image('test.png')
		# Sheet1.add_image(img,'C16')

		# img= openpyxl.drawing.image.Image('test2.png')
		# Sheet1.add_image(img,'C19')

		# img= openpyxl.drawing.image.Image('test10.png')
		# Sheet1.add_image(img,'C22')

		# img= openpyxl.drawing.image.Image('test6.png')
		# Sheet1.add_image(img,'C25')

		# img= openpyxl.drawing.image.Image('test7.png')
		# Sheet1.add_image(img,'C28')

		# img= openpyxl.drawing.image.Image('test6.png')
		# Sheet1.add_image(img,'D16')

		# img= openpyxl.drawing.image.Image('test7.png')
		# Sheet1.add_image(img,'D19')

		# img= openpyxl.drawing.image.Image('test8.png')
		# Sheet1.add_image(img,'D22')

		# img= openpyxl.drawing.image.Image('test9.png')
		# Sheet1.add_image(img,'D25')

		# img= openpyxl.drawing.image.Image('test5.png')
		# Sheet1.add_image(img,'D28')

		Sheet1.merge_cells(start_row=28, start_column=3, end_row=29, end_column=3)
		Sheet1.merge_cells(start_row=25, start_column=3, end_row=26, end_column=3)
		Sheet1.merge_cells(start_row=22, start_column=3, end_row=23, end_column=3)
		Sheet1.merge_cells(start_row=19, start_column=3, end_row=20, end_column=3)
		Sheet1.merge_cells(start_row=16, start_column=3, end_row=17, end_column=3)
		Sheet1.merge_cells(start_row=6, start_column=4, end_row=6, end_column=12)
		Sheet1.merge_cells(start_row=7, start_column=4, end_row=7, end_column=12)
		Sheet1.merge_cells(start_row=8, start_column=4, end_row=8, end_column=12)
		Sheet1.merge_cells(start_row=9, start_column=4, end_row=9, end_column=12)
		Sheet1.merge_cells(start_row=10, start_column=4, end_row=10, end_column=12)
		Sheet1.merge_cells(start_row=11, start_column=4, end_row=11, end_column=12)
		Sheet1.merge_cells(start_row=12, start_column=4, end_row=12, end_column=12)
		Sheet1.merge_cells(start_row=58, start_column=3, end_row=58, end_column=6)

	if(option==0):
		Sheet1=temp.create_sheet('Cover sheet')
		fonta = PatternFill(start_color = 'ffffff', end_color ='ffffff', fill_type = 'solid')
		fontg = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		font2 = Font(name = 'Georgia', size = 10, bold = True, color="000000")
		font1 = Font(name = 'Georgia', size = 10, color = "FFFFFF", bold = True,italic=True)
		font3 = Font(name = 'Georgia', size = 10, color = "000000",italic=True)
		culoare = PatternFill(start_color = '182A54', end_color ='182A54', fill_type = 'solid') 
		culoare2 = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		culoare3 = PatternFill(start_color = 'D9E1F2', end_color ='D9E1F2', fill_type = 'solid')
		culoare4 = PatternFill(start_color = 'E2EFDA', end_color ='E2EFDA', fill_type = 'solid')
		culoare5 = PatternFill(start_color = 'FFF2CC', end_color ='FFF2CC', fill_type = 'solid')
		culoare6 = PatternFill(start_color = '808080', end_color ='808080', fill_type = 'solid')
		font4 = Font(name = 'Georgia', size = 10, color = "000000",underline='single',bold=True)
		font5 = Font(name = 'Georgia', size = 10, color = "ffffff",underline='single',bold=True)
		border = Border(bottom=Side(style='dotted'))
		border2 = Border(top=Side(style='dotted'))
		border3 = Border(left=Side(style='dotted'))
		border4 = Border(right=Side(style='dotted'))
		border5 = Border(left=Side(style='dotted'),top=Side(style='dotted'))
		border6 = Border(left=Side(style='dotted'),bottom=Side(style='dotted'))
		border7 = Border(right=Side(style='dotted'),bottom=Side(style='dotted'))
		border8 = Border(right=Side(style='dotted'),top=Side(style='dotted'))
		# border9 = Border(right=Side(style='double'),bottom=Side(style='double'),top=Side(style='double'),left=Side(style='double'))

		Sheet1.cell(row=31, column=2).border=border2
		Sheet1.cell(row=31, column=3).border=border2
		Sheet1.cell(row=31, column=4).border=border2
		Sheet1.cell(row=31, column=5).border=border2
		Sheet1.cell(row=31, column=6).border=border2
		Sheet1.cell(row=31, column=7).border=border2
		Sheet1.cell(row=31, column=2).border=border5
		Sheet1.cell(row=31, column=7).border=border8


		Sheet1.cell(row=75, column=2).border=border
		Sheet1.cell(row=75, column=3).border=border
		Sheet1.cell(row=75, column=4).border=border
		Sheet1.cell(row=75, column=5).border=border
		Sheet1.cell(row=75, column=6).border=border
		Sheet1.cell(row=75, column=7).border=border
		Sheet1.cell(row=75, column=7).border=border7

		Sheet1.cell(row=32, column=2).border=border3
		Sheet1.cell(row=33, column=2).border=border3
		Sheet1.cell(row=34, column=2).border=border3
		Sheet1.cell(row=35, column=2).border=border3
		Sheet1.cell(row=36, column=2).border=border3
		Sheet1.cell(row=37, column=2).border=border3
		Sheet1.cell(row=38, column=2).border=border3
		Sheet1.cell(row=39, column=2).border=border3
		Sheet1.cell(row=40, column=2).border=border3
		Sheet1.cell(row=41, column=2).border=border3
		Sheet1.cell(row=42, column=2).border=border3
		Sheet1.cell(row=43, column=2).border=border3
		Sheet1.cell(row=44, column=2).border=border3
		Sheet1.cell(row=45, column=2).border=border3
		Sheet1.cell(row=46, column=2).border=border3
		Sheet1.cell(row=47, column=2).border=border3
		Sheet1.cell(row=48, column=2).border=border3
		Sheet1.cell(row=49, column=2).border=border3
		Sheet1.cell(row=50, column=2).border=border3
		Sheet1.cell(row=51, column=2).border=border3
		Sheet1.cell(row=52, column=2).border=border3
		Sheet1.cell(row=53, column=2).border=border3
		Sheet1.cell(row=54, column=2).border=border3
		Sheet1.cell(row=55, column=2).border=border3
		Sheet1.cell(row=75, column=2).border=border6
		Sheet1.cell(row=56, column=2).border=border3
		Sheet1.cell(row=57, column=2).border=border3
		Sheet1.cell(row=58, column=2).border=border3
		Sheet1.cell(row=59, column=2).border=border3
		Sheet1.cell(row=60, column=2).border=border3
		Sheet1.cell(row=61, column=2).border=border3
		Sheet1.cell(row=62, column=2).border=border3
		Sheet1.cell(row=63, column=2).border=border3
		Sheet1.cell(row=64, column=2).border=border3
		Sheet1.cell(row=65, column=2).border=border3
		Sheet1.cell(row=66, column=2).border=border3
		Sheet1.cell(row=67, column=2).border=border3
		Sheet1.cell(row=68, column=2).border=border3
		Sheet1.cell(row=69, column=2).border=border3
		Sheet1.cell(row=70, column=2).border=border3
		Sheet1.cell(row=71, column=2).border=border3
		Sheet1.cell(row=72, column=2).border=border3
		Sheet1.cell(row=73, column=2).border=border3
		Sheet1.cell(row=74, column=2).border=border3

		Sheet1.cell(row=32, column=7).border=border4
		Sheet1.cell(row=33, column=7).border=border4
		Sheet1.cell(row=34, column=7).border=border4
		Sheet1.cell(row=35, column=7).border=border4
		Sheet1.cell(row=36, column=7).border=border4
		Sheet1.cell(row=37, column=7).border=border4
		Sheet1.cell(row=38, column=7).border=border4
		Sheet1.cell(row=39, column=7).border=border4
		Sheet1.cell(row=40, column=7).border=border4
		Sheet1.cell(row=41, column=7).border=border4
		Sheet1.cell(row=42, column=7).border=border4
		Sheet1.cell(row=43, column=7).border=border4
		Sheet1.cell(row=44, column=7).border=border4
		Sheet1.cell(row=45, column=7).border=border4
		Sheet1.cell(row=46, column=7).border=border4
		Sheet1.cell(row=47, column=7).border=border4
		Sheet1.cell(row=48, column=7).border=border4
		Sheet1.cell(row=49, column=7).border=border4
		Sheet1.cell(row=50, column=7).border=border4
		Sheet1.cell(row=51, column=7).border=border4
		Sheet1.cell(row=52, column=7).border=border4
		Sheet1.cell(row=53, column=7).border=border4
		Sheet1.cell(row=54, column=7).border=border4
		Sheet1.cell(row=55, column=7).border=border4
		Sheet1.cell(row=56, column=7).border=border4
		Sheet1.cell(row=57, column=7).border=border4
		Sheet1.cell(row=58, column=7).border=border4
		Sheet1.cell(row=59, column=7).border=border4
		Sheet1.cell(row=60, column=7).border=border4
		Sheet1.cell(row=61, column=7).border=border4
		Sheet1.cell(row=62, column=7).border=border4
		Sheet1.cell(row=63, column=7).border=border4
		Sheet1.cell(row=64, column=7).border=border4
		Sheet1.cell(row=65, column=7).border=border4
		Sheet1.cell(row=66, column=7).border=border4
		Sheet1.cell(row=67, column=7).border=border4
		Sheet1.cell(row=68, column=7).border=border4
		Sheet1.cell(row=69, column=7).border=border4
		Sheet1.cell(row=70, column=7).border=border4
		Sheet1.cell(row=71, column=7).border=border4
		Sheet1.cell(row=72, column=7).border=border4
		Sheet1.cell(row=73, column=7).border=border4
		Sheet1.cell(row=74, column=7).border=border4
		

		Sheet1.cell(row=14, column=3).value="Summary"
		Sheet1.cell(row=14, column=3).font=font4
		Sheet1.cell(row=58, column=3).value="PAYMENT ORDER"
		Sheet1.cell(row=58, column=3).font=font5

		for row in Sheet1['A1:N100']:
					for cell in row:
						cell.fill = fonta

		for row in Sheet1['N1:Z100']:
					for cell in row:
						cell.fill = fontg


		Sheet1.cell(row = 6, column = 3).value = "Company"
		Sheet1.cell(row = 6, column = 4).value = "='Other info'!C4"
		Sheet1.cell(row = 7, column = 3).value = "Address"
		Sheet1.cell(row = 7, column = 4).value = "='Other info'!C6"
		Sheet1.cell(row = 8, column = 3).value = "VAT tax code"
		Sheet1.cell(row = 8, column = 4).value = "='Other info'!C5"
		Sheet1['D8'].alignment = Alignment(wrapText=True, horizontal='left')
		Sheet1.cell(row = 9, column = 3).value = "Registration no."
		Sheet1.cell(row = 9, column = 4).value = "J08/1139/2017"
		Sheet1.cell(row = 10, column = 3).value = "The administration it belongs to"
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		Sheet1.cell(row = 11, column = 3).value = "Frequency of declaration / payment"
		Sheet1.cell(row = 11, column = 4).value = "Monthly"
		Sheet1.cell(row = 12, column = 3).value = "Reporting period"
		Sheet1.cell(row = 12, column = 4).value = "=date('Other info'!C2,'Other info'!C3,1)"
		Sheet1.cell(row = 12, column = 4).number_format = 'mmmm yyyy'

		Sheet1.cell(row = 43, column = 4).value="Yes"
		Sheet1.cell(row = 46, column = 4).value="N/a"
		Sheet1.cell(row = 47, column = 4).value="N/a"
		Sheet1.cell(row = 50, column = 4).value="N/a"
		Sheet1.cell(row = 51, column = 4).value="N/a"
		Sheet1.cell(row = 43, column = 4).font=font5
		Sheet1.cell(row = 46, column = 4).font=font5
		Sheet1.cell(row = 47, column = 4).font=font5
		Sheet1.cell(row = 50, column = 4).font=font5
		Sheet1.cell(row = 51, column = 4).font=font5
		info=temp['Other info']
		valluna=""
		vallunaurmatoare=""
		valIban=""
		okdecembrie=0
		for i in range(0,len(listaluni)):
			if(str(info.cell(row=3,column=3).value)=="12"):
				okdecembrie=1
				vallunaurmatoare=listadenluni2[0]
				valluna=listadenluni2[11]
			else:
				if(listaluni[i]==str(info.cell(row=3,column=3).value)):
					valluna=listadenluni2[i]
					vallunaurmatoare=listadenluni2[i+1]
		var=Sheet1.cell(row=12,column=4).value
		print(var)
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		for j in range(0,len(listatrez)):
			if(listatrez[j]==str(Sheet1.cell(row=10,column=4).value)):
				valIban=listaiban[j]
		Sheet1.cell(row=60, column=3).value='Payment period: '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=61, column=3).value='="Suma de plata: " &D55&" RON "'
		Sheet1.cell(row=62, column=3).value="Currency: RON"
		Sheet1.cell(row=63, column=3).value='Payment details: VAT return - '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		if(okdecembrie==1):
			Sheet1.cell(row=64, column=3).value='Deadline: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value+1)
		else:
			Sheet1.cell(row=64, column=3).value='Deadline: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=66, column=3).value='="Payer TIN: " & D8'
		Sheet1.cell(row=67, column=3).value='="Payer address: " &D7'
		Sheet1.cell(row=69, column=3).value="Beneficiary: BUGETUL DE STAT"
		Sheet1.cell(row=70, column=3).value='IBAN: '+ str(valIban)
		Sheet1.cell(row=71, column=3).value="SWIFT / BIC: TREZROBU"
		Sheet1.cell(row=72, column=3).value="Bank / Treasury:"+str(dropdown)
		Sheet1.cell(row=74, column=3).value="Note: Any banking fees connected with the payment must be covered by the tax payer."
		Sheet1.cell(row=74, column=3).font=font2

		Sheet1.cell(row = 6, column = 3).font=font1
		Sheet1.cell(row = 6, column = 4).font=font2
		Sheet1.cell(row = 7, column = 3).font=font1
		Sheet1.cell(row = 7, column = 4).font=font2
		Sheet1.cell(row = 8, column = 3).font=font1
		Sheet1.cell(row = 8, column = 4).font=font2
		Sheet1.cell(row = 9, column = 3).font=font1
		Sheet1.cell(row = 9, column = 4).font=font2
		Sheet1.cell(row = 10, column = 3).font=font1
		Sheet1.cell(row = 10, column = 4).font=font3
		Sheet1.cell(row = 11, column = 3).font=font1
		Sheet1.cell(row = 11, column = 4).font=font3
		Sheet1.cell(row = 12, column = 3).font=font1
		Sheet1.cell(row = 12, column = 4).font=font3

		Sheet1.cell(row = 6, column = 3).fill=culoare
		Sheet1.cell(row = 7, column = 3).fill=culoare
		Sheet1.cell(row = 8, column = 3).fill=culoare
		Sheet1.cell(row = 9, column = 3).fill=culoare
		Sheet1.cell(row = 10, column = 3).fill=culoare
		Sheet1.cell(row = 11, column = 3).fill=culoare
		Sheet1.cell(row = 12, column = 3).fill=culoare
		Sheet1.cell(row = 6, column = 4).fill=culoare2
		Sheet1.cell(row = 7, column = 4).fill=culoare2
		Sheet1.cell(row = 8, column = 4).fill=culoare2
		Sheet1.cell(row = 9, column = 4).fill=culoare2
		Sheet1.cell(row = 10, column = 4).fill=culoare2
		Sheet1.cell(row = 11, column = 4).fill=culoare2
		Sheet1.cell(row = 12, column = 4).fill=culoare2

		Sheet1.cell(row = 58, column = 3).fill=culoare6
		Sheet1.cell(row = 32, column = 3).fill=culoare3
		Sheet1.cell(row = 33, column = 3).fill=culoare3
		Sheet1.cell(row = 34, column = 3).fill=culoare3
		Sheet1.cell(row = 35, column = 3).fill=culoare3
		Sheet1.cell(row = 36, column = 3).fill=culoare3
		Sheet1.cell(row = 37, column = 3).fill=culoare3
		Sheet1.cell(row = 38, column = 3).fill=culoare3
		Sheet1.cell(row = 39, column = 3).fill=culoare3
		Sheet1.cell(row = 32, column = 4).fill=culoare3
		Sheet1.cell(row = 33, column = 4).fill=culoare3
		Sheet1.cell(row = 34, column = 4).fill=culoare3
		Sheet1.cell(row = 35, column = 4).fill=culoare3
		Sheet1.cell(row = 36, column = 4).fill=culoare3
		Sheet1.cell(row = 37, column = 4).fill=culoare3
		Sheet1.cell(row = 38, column = 4).fill=culoare3
		Sheet1.cell(row = 39, column = 4).fill=culoare3
		Sheet1.cell(row = 32, column = 6).fill=culoare3
		Sheet1.cell(row = 33, column = 6).fill=culoare3
		Sheet1.cell(row = 34, column = 6).fill=culoare3
		Sheet1.cell(row = 35, column = 6).fill=culoare3
		Sheet1.cell(row = 36, column = 6).fill=culoare3
		Sheet1.cell(row = 37, column = 6).fill=culoare3
		Sheet1.cell(row = 38, column = 6).fill=culoare3
		Sheet1.cell(row = 39, column = 6).fill=culoare3

		Sheet1.cell(row = 41, column = 6).fill=culoare4
		Sheet1.cell(row = 42, column = 6).fill=culoare4
		Sheet1.cell(row = 43, column = 6).fill=culoare4
		Sheet1.cell(row = 44, column = 6).fill=culoare4
		Sheet1.cell(row = 45, column = 6).fill=culoare4
		Sheet1.cell(row = 46, column = 6).fill=culoare4
		Sheet1.cell(row = 47, column = 6).fill=culoare4
		Sheet1.cell(row = 48, column = 6).fill=culoare4
		Sheet1.cell(row = 49, column = 6).fill=culoare4
		Sheet1.cell(row = 50, column = 6).fill=culoare4
		Sheet1.cell(row = 51, column = 6).fill=culoare4
		Sheet1.cell(row = 41, column = 3).fill=culoare4
		Sheet1.cell(row = 42, column = 3).fill=culoare4
		Sheet1.cell(row = 43, column = 3).fill=culoare4
		Sheet1.cell(row = 44, column = 3).fill=culoare4
		Sheet1.cell(row = 45, column = 3).fill=culoare4
		Sheet1.cell(row = 46, column = 3).fill=culoare4
		Sheet1.cell(row = 47, column = 3).fill=culoare4
		Sheet1.cell(row = 48, column = 3).fill=culoare4
		Sheet1.cell(row = 49, column = 3).fill=culoare4
		Sheet1.cell(row = 50, column = 3).fill=culoare4
		Sheet1.cell(row = 51, column = 3).fill=culoare4
		Sheet1.cell(row = 41, column = 4).fill=culoare4
		Sheet1.cell(row = 42, column = 4).fill=culoare4
		Sheet1.cell(row = 43, column = 4).fill=culoare6
		Sheet1.cell(row = 44, column = 4).fill=culoare4
		Sheet1.cell(row = 45, column = 4).fill=culoare4
		Sheet1.cell(row = 46, column = 4).fill=culoare6
		Sheet1.cell(row = 47, column = 4).fill=culoare6
		Sheet1.cell(row = 48, column = 4).fill=culoare4
		Sheet1.cell(row = 49, column = 4).fill=culoare4
		Sheet1.cell(row = 50, column = 4).fill=culoare6
		Sheet1.cell(row = 51, column = 4).fill=culoare6

		Sheet1.cell(row = 53, column = 4).fill=culoare5
		Sheet1.cell(row = 54, column = 4).fill=culoare5
		Sheet1.cell(row = 55, column = 4).fill=culoare5
		Sheet1.cell(row = 53, column = 3).fill=culoare5
		Sheet1.cell(row = 54, column = 3).fill=culoare5
		Sheet1.cell(row = 55, column = 3).fill=culoare5
		Sheet1.cell(row = 53, column = 6).fill=culoare5
		Sheet1.cell(row = 54, column = 6).fill=culoare5
		Sheet1.cell(row = 55, column = 6).fill=culoare5


		Sheet1.cell(row = 32, column = 3).value="Current VAT position"
		Sheet1.cell(row = 32, column = 3).font=font4
		Sheet1.cell(row = 34, column = 3).value="Input VAT for the period"
		Sheet1.cell(row = 35, column = 3).value="Output VAT for the period"
		Sheet1.cell(row = 36, column = 3).value="VAT Payable for the period"
		Sheet1.cell(row = 37, column = 3).value="VAT Recoverable for the period"
		Sheet1.cell(row = 38, column = 3).value="VAT under settlement for purchases"
		Sheet1.cell(row = 39, column = 3).value="VAT under settlement for deliveries"
		Sheet1.cell(row = 32, column = 4).value="RON"

		Sheet1.cell(row = 32, column = 6).value="Euro(@"+get_fxrate(2022)+")"
		Sheet1.cell(row = 34, column = 4).value="='D300 draft figures'!C56"
		Sheet1.cell(row = 34, column = 6).value="=IFERROR(D34/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=34, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=34, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row = 35, column = 4).value="='D300 draft figures'!C32"
		Sheet1.cell(row = 35, column = 6).value="=IFERROR(D35/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 36, column = 4).value="=IF('D300 draft figures'!C58<>0,'D300 draft figures'!C58,0)"
		Sheet1.cell(row = 36, column = 6).value="=IFERROR(D36/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=35, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=35, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row = 37, column = 4).value='''=IF('D300 draft figures'!C57<>0,'D300 draft figures'!C57,"nil")'''
		Sheet1.cell(row = 37, column = 6).value="=iferror(D37/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 38, column = 4).value="='D300 draft figures'!C70"
		Sheet1.cell(row = 38, column = 6).value="=iferror(D38/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 39, column = 4).value=0
		Sheet1.cell(row = 39, column = 6).value="=iferror(D39/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 55, column = 4).value='''=IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)=0),'Cover sheet'!D36,
IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)<>0),IF('Cover sheet'!D36>IFERROR(VALUE('Cover sheet'!D47),0),'Cover sheet'!D36-IFERROR(VALUE('Cover sheet'!D47),0),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="No"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D47),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="Yes"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D51),0),"N/A"))))'''
		Sheet1.cell(row = 55, column = 6).value="=iferror(D55/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=37, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=37, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=6).number_format = '#,##0_);(#,##0)'


		Sheet1.cell(row = 41, column = 3).value="Carry Over position"
		Sheet1.cell(row = 41, column = 3).font=font4
		Sheet1.row_dimensions[42].height = 0.2
		Sheet1.cell(row = 43, column = 3).value="Requested for reimbursement"
		Sheet1.cell(row = 44, column = 3).value
		Sheet1.cell(row = 45, column = 3).value="VAT refundable not yet requested"
		Sheet1.cell(row = 46, column = 3).value="Period"
		Sheet1.cell(row = 47, column = 3).value="Amount"
		Sheet1.cell(row = 48, column = 3).value
		Sheet1.cell(row = 49, column = 3).value="VAT refundable requested and under audit"
		Sheet1.cell(row = 50, column = 3).value="Period"
		Sheet1.cell(row = 51, column = 3).value="Amount"


		Sheet1.cell(row = 53, column = 3).value="VAT balance position"
		Sheet1.cell(row = 53, column = 3).font=font4
		Sheet1.cell(row = 55, column = 3).value="VAT position in the current return"


		Sheet1['C16'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C19'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C22'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C25'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C28'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['D10'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D11'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D12'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['C58'].alignment = Alignment(wrapText=True, horizontal='center')
		Sheet1['D37'].alignment = Alignment(wrapText=True, horizontal='right')

		Sheet1.column_dimensions['C'].width = 65
		Sheet1.column_dimensions['D'].width = 20
		Sheet1.column_dimensions['A'].width = 2
		Sheet1.column_dimensions['B'].width = 3
		Sheet1.column_dimensions['E'].width = 1
		Sheet1.column_dimensions['F'].width = 20
		Sheet1.column_dimensions['G'].width = 2

		# img= openpyxl.drawing.image.Image('test.png')
		# Sheet1.add_image(img,'C16')

		# img= openpyxl.drawing.image.Image('test2.png')
		# Sheet1.add_image(img,'C19')

		# img= openpyxl.drawing.image.Image('test3.png')
		# Sheet1.add_image(img,'C22')

		# img= openpyxl.drawing.image.Image('test4.png')
		# Sheet1.add_image(img,'C25')

		# img= openpyxl.drawing.image.Image('test5.png')
		# Sheet1.add_image(img,'C28')

		# img= openpyxl.drawing.image.Image('test6.png')
		# Sheet1.add_image(img,'D16')

		# img= openpyxl.drawing.image.Image('test7.png')
		# Sheet1.add_image(img,'D19')

		# img= openpyxl.drawing.image.Image('test8.png')
		# Sheet1.add_image(img,'D22')

		# img= openpyxl.drawing.image.Image('test9.png')
		# Sheet1.add_image(img,'D25')
		Sheet1.cell(row = 16, column = 3).fill=culoare
		Sheet1.cell(row=16, column=3).font=font1
		Sheet1.cell(row=16, column=3).value="  D300"
		Sheet1.cell(row=16, column=3).hyperlink="#'D300 draft figures'!A1"
		# Sheet1.cell(row=16, column=3).border=border9
		Sheet1.row_dimensions[18].height=8

		Sheet1.cell(row = 19, column = 3).fill=culoare
		Sheet1.cell(row=19, column=3).font=font1
		Sheet1.cell(row=19, column=3).value="  D390"
		Sheet1.cell(row=19, column=3).hyperlink="#'D390 workings'!A1"
		# Sheet1.cell(row=19, column=3).border=border9
		Sheet1.row_dimensions[21].height=8

		Sheet1.cell(row = 22, column = 3).fill=culoare
		Sheet1.cell(row=22, column=3).font=font1
		Sheet1.cell(row=22, column=3).value="  D394"
		Sheet1.cell(row=22, column=3).hyperlink="#'D394--->>>'!A1"
		# Sheet1.cell(row=22, column=3).border=border9
		Sheet1.row_dimensions[24].height=8

		Sheet1.cell(row = 25, column = 3).fill=culoare
		Sheet1.cell(row=25, column=3).font=font1
		Sheet1.cell(row=25, column=3).value="  Sales Ledger"
		Sheet1.cell(row=25, column=3).hyperlink="#'Sales'!A1"
		# Sheet1.cell(row=25, column=3).border=border9
		Sheet1.row_dimensions[27].height=8

		Sheet1.cell(row = 28, column = 3).fill=culoare
		Sheet1.cell(row=28, column=3).font=font1
		Sheet1.cell(row=28, column=3).value="  Purchase Ledger"
		Sheet1.cell(row=28, column=3).hyperlink="#'Purchases'!A1"
		# Sheet1.cell(row=28, column=3).border=border9
		Sheet1.row_dimensions[30].height=8

		Sheet1.merge_cells(start_row=28, start_column=3, end_row=29, end_column=3)
		Sheet1.merge_cells(start_row=25, start_column=3, end_row=26, end_column=3)
		Sheet1.merge_cells(start_row=22, start_column=3, end_row=23, end_column=3)
		Sheet1.merge_cells(start_row=19, start_column=3, end_row=20, end_column=3)
		Sheet1.merge_cells(start_row=16, start_column=3, end_row=17, end_column=3)
		Sheet1.merge_cells(start_row=6, start_column=4, end_row=6, end_column=12)
		Sheet1.merge_cells(start_row=7, start_column=4, end_row=7, end_column=12)
		Sheet1.merge_cells(start_row=8, start_column=4, end_row=8, end_column=12)
		Sheet1.merge_cells(start_row=9, start_column=4, end_row=9, end_column=12)
		Sheet1.merge_cells(start_row=10, start_column=4, end_row=10, end_column=12)
		Sheet1.merge_cells(start_row=11, start_column=4, end_row=11, end_column=12)
		Sheet1.merge_cells(start_row=12, start_column=4, end_row=12, end_column=12)
		Sheet1.merge_cells(start_row=58, start_column=3, end_row=58, end_column=6)
	
	sales=temp['Sales']
	purchases=temp['Purchases']
	if(val1==1):
		listadescrieri=['Intra-community supplies of goods, exempted according to art. 294 par.(2)let.a) and d) of the Fiscal code','Adjustments of VAT exempt intra-community supplies according to art. 294 par. (2) let. a) and d) of the Fiscal code','Supplies of goods or services for which the place of supply is outside Romania (in EU or outside EU), as well as intra-community supplies of goods, exempted according to art. 294 par.(2) let.b) and c) of the Fiscal code, out of which:','Intra-community supplies of services which are not VAT exempt in the Member State where the tax is due','Adjustments of intra-community supplies of services which are not VAT exempt in the Member State where the tax is due','Intra-community acquisitions of goods for which the buyer is liable to pay VAT (reverse charge), out of which:','Intra-community acquisitions for which the buyer is liable to pay VAT (reverse charge), and the supplier is registered for VAT purposes in the Member State where the intra-community supply took place','Adjustments of intra-community acquisitions of goods for which the buyer is liable to pay VAT (reverse charge)','Acquisitions of goods, other than those in rows 5 and 6, and acquisitions of services for which the beneficiary in Romania is liable to pay VAT (reverse charge), out of which:','Acquisitions of intra-community services for which the beneficiary is liable to pay VAT (reverse charge)','Adjustments of intra-community acquisitions of services for which the beneficiary is liable to pay VAT (reverse charge)','Supplies of goods and services, taxable with 19% VAT rate','Supplies of goods and services, taxable with 9% VAT rate','Supplies of goods taxable with 5% VAT rate','Acquisitions of goods and services subject to simplification measures for which the beneficiary is liable to pay VAT (reverse charge)','Acquisitions of goods and services taxable with 20% VAT rate','Acquisitions of goods taxable with 9% VAT rate','Acquisitions of goods taxable with 5% VAT rate','Supplies of goods and services subject to simplification measures (reverse charge)','Supplies of goods and services VAT exempt with deduction right, other than those in rows 1 - 3','Supplies of goods and services exempt without deduction right','Adjustments of output tax','Intra-community supplies of services according to art 278 par(8) of the Fiscal code taxable in Romania','Adjustents of intra-community supplies of services according to art 278 par(8) of the Fiscal code taxable in Romania','Total output tax (sum for rows 1 to 16, except 3.1, 5.1, 7.1, 12.1, 12.2, 12.3)','Intra-community acquisitions of goods for which the buyer is liable to pay VAT (reverse charge)(row 18=row 5), out of which:','Intra-community acquisitions for which the buyer is liable to pay VAT (reverse charge), whereas the supplier is registered for VAT purposes in the Member State where the supply took place (row 18.1=row 5.1)','Adjustments of intra-community acquisitions of goods for which the buyer is liable to pay VAT (reverse charge) (row 19=row 6)','Acquisitions of goods, other than those in rows 18 and 19, and acquisitions of services for which the beneficiary in Romania is liable to pay VAT (reverse charge)(row 20=row 7), out of which:','Intra-community acquisitions of services for which the beneficiary is liable to pay VAT (reverse charge)(row 20.1=row 7.1)','Adjustments of intra-community acquisitions of services for which the beneficiary in Romania is liable to pay VAT (reverse charge)(row 21= row 8)','Acquisitions of goods and services, taxable with 19% VAT rate, other than those in row 25','Acquisitions of goods and services, taxable with 9% VAT rate','Acquisitions of goods taxable with 5% VAT rate','Acquisitions of goods and services subject to simplification measures for which the beneficiary is liable to pay VAT (reverse charge) ,out of which ','Acquisitions of goods and services taxable with 20% VAT rate ','Acquisitions of goods taxable with 9% VAT rate ','Acquisitions of goods taxable with 5% VAT rate','Compensation in the flat fee for purchases of agricultural products and services from suppliers applying the special regime','Adjusments of compensation in the flat fee for purchases of agricultural products and services from suppliers applying the special regime','Acquisitions of goods and services VAT exempt or non-taxable, out of which:','Acquisitions of VAT exempt intra-community services (do not fill in fro the simplified method)','TOTAL DEDUCTIBLE VAT (sum for rows 18 to 25, except 18.1, 20.1, 25.1, 25.2, 25.3):','','SUB-TOTAL DEDUCTED VAT ACCORDING TO ART. 297 AND ART. 298 OR ART. 300 AND ART. 298 OF THE FISCAL CODE','VAT effectively refunded to foreign buyers, including the commission of the authorised bodies','Adjustments of input VAT','Adjustments according to pro rata / adjustments for capital goods','TOTAL DEDUCTED VAT (row 28+row 29+row 30+row 31)','Refundable VAT amount in the reporting period (row 32-row 17)','Payable VAT in the reporting period (row 17-row 32)','Balance of payable VAT from the previous VAT return (row 41 in the previous VAT return) unpaid until the date of the submission of the VAT return','Payable VAT differences assessed by the fiscal inspection bodies through a communicated decision and unpaid until the date of submission of the VAT return ','Cumulated VAT payable (row 34+row 35+row 36)','Balance of the VAT refundable amount reported from the previous period for which the refund was not claimed (row 46 in the previous VAT return)','Differences of VAT refundable amounts assessed by the fiscal inspection bodies decision through a communicated decision until the date of submission of the VAT return','Cumulated refundable VAT amount (row 33+row 38+row 39)','Balance of payable VAT at the end of the reporting period (row 37-row 40)','Balance of refundable VAT at the end of the reporting period (row 40-row 37)','','Supplies of goods and services performed for which the related VAT remained non-chargeable, existing in the balance at the end of the reporting period, as a consequence of the application of the VAT cash accouting system','Supplies of goods and services performed in the last 6 months/2 calendar quarters','Acquisitions of goods and services performed for which has not been applied the deduction right, existing in the balance at the end of the reporting period, as a consequence of the application of art. 145 para (1¹) and (1²) of the Fiscal Code','Acquisitions of goods and services performed in the last 6 months/ 2 calendar quarters ']
		sheetinutil1=temp.create_sheet('D300--->>>')
		sheetinutil1.sheet_view.showGridLines=False
		sheetinutil1.cell(row=2,column=1).value="Switch to next sheet for D300 Workings draft"
		sheetinutil1.cell(row=2,column=1).font=scrisincredibildemare
		amount=temp.create_sheet('D300 draft figures')
		amount.freeze_panes = 'A8'
		amount.auto_filter.ref = "A7:G71"
		amount.sheet_view.showGridLines = False
		taxcode=[]
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "                 S01":
					rand_tb = cell.row
					A12col = cell.column
					lun = len(sales[cell.column])
		try:
			bazaA12 = [b.value for b in sales[A12col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                 S01'")
			return render_template("index.html")
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "                 S02":
					rand_tb = cell.row
					A12tva= cell.column
					lun = len(sales[cell.column])
		try:
			tvaA12 = [b.value for b in sales[A12tva][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                 S02'")
			return render_template("index.html")
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "                 S03":
					rand_tb = cell.row
					A8col = cell.column
					lun = len(sales[cell.column])
		try:
			bazaA8 = [b.value for b in sales[A8col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                 S03'")
			return render_template("index.html")
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "                 S04":
					rand_tb = cell.row
					A8tva = cell.column
					lun = len(sales[cell.column])
		try:
			tvaA8 = [b.value for b in sales[A8tva][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                 S04'")
			return render_template("index.html")
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "                 S25":
					rand_tb = cell.row
					Y8col = cell.column
					lun = len(sales[cell.column])
		try:
			valY8 = [b.value for b in sales[Y8col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                 S25'")
			return render_template("index.html")			

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "                 S21":
					rand_tb = cell.row
					E2col = cell.column
					lun = len(sales[cell.column])
		try:
			valE2 = [b.value for b in sales[E2col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                 S25'")
			return render_template("index.html")
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "                 S22":
					rand_tb = cell.row
					E2colvat = cell.column
					lun = len(sales[cell.column])
		try:
			valE2 = [b.value for b in sales[E2colvat][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                 S25'")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "  Total doc.incl.VAT":
					rand_tb = cell.row
					tdocc = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in sales[tdocc][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '  Total doc.incl.VAT'")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Posting Date":
					rand_tb = cell.row
					tdat = cell.column
					lun = len(sales[cell.column])
		try:
			listdocdate = [b.value for b in sales[tdat][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Posting Date'")
			return render_template("index.html")
		listacurentas=[]						
		for k in range(0,len(listdocdate)):
			# print(datadocument[k][3:4])
			# print(datadocument[k][3:5])
			if(str(listdocdate[k])[3:4]=="0"):
				if(str(listdocdate[k])[4:5]==str(info.cell(row=3,column=3).value)):
					listacurentas.append("Yes")
				else:
					listacurentas.append("No")

			else:
				if(str(listdocdate[k])[3:5]==str(info.cell(row=3,column=3).value)):
					listacurentas.append("Yes")
				else:
					listacurentas.append("No")
		for kk in range(0,len(listacurentas)):
			sales.cell(row=18+kk,column=70).value=listacurentas[kk]
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                   1":
					rand_tb = cell.row
					bazaV1col = cell.column
					lun = len(purchases[cell.column])
		try:
			bazaV1 = [b.value for b in purchases[bazaV1col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                   1' in Purchases sheet")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                   2":
					rand_tb = cell.row
					tvaV1col = cell.column
					lun = len(purchases[cell.column])
		try:
			tvaV1 = [b.value for b in purchases[tvaV1col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                    2' in Purchases sheet")
			return render_template("index.html")			

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                   3":
					rand_tb = cell.row
					bazaW7col = cell.column
					lun = len(purchases[cell.column])
		try:
			bazaW7 = [b.value for b in purchases[bazaW7col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                   3' in Purchases sheet")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                   4":
					rand_tb = cell.row
					tvaW7col = cell.column
					lun = len(purchases[cell.column])
		try:
			tvaW7 = [b.value for b in purchases[tvaW7col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                   4' in Purchases sheet")
			return render_template("index.html")		

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                   5":
					rand_tb = cell.row
					bazaW8col = cell.column
					lun = len(purchases[cell.column])
		try:
			bazaW8 = [b.value for b in purchases[bazaW8col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                   5' in Purchases sheet")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                   6":
					rand_tb = cell.row
					tvaw8col = cell.column
					lun = len(purchases[cell.column])
		try:
			tvaW8 = [b.value for b in purchases[tvaw8col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                   6' in Purchases sheet")
			return render_template("index.html")	

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                   7":
					rand_tb = cell.row
					baza1Lcol = cell.column
					lun = len(purchases[cell.column])
		try:
			baza1L = [b.value for b in purchases[baza1Lcol][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                   7' in Purchases sheet")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                   8":
					rand_tb = cell.row
					tva1Lcol = cell.column
					lun = len(purchases[cell.column])
		try:
			tva1L = [b.value for b in purchases[tva1Lcol][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                   8' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  10":
					rand_tb = cell.row
					baza1Jcol= cell.column
					lun = len(purchases[cell.column])
		try:
			baza1J = [b.value for b in purchases[baza1Jcol][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  10' in Purchases sheet")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  11":
					rand_tb = cell.row
					tva1Jcol = cell.column
					lun = len(purchases[cell.column])
		try:
			tva1J = [b.value for b in purchases[tva1Jcol][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  11' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  22":
					rand_tb = cell.row
					W0col = cell.column
					lun = len(purchases[cell.column])
		try:
			bazaW0 = [b.value for b in purchases[W0col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  22' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  23":
					rand_tb = cell.row
					V9col = cell.column
					lun = len(purchases[cell.column])
		try:
			bazaV9 = [b.value for b in purchases[V9col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  23' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  24":
					rand_tb = cell.row
					E2col = cell.column
					lun = len(purchases[cell.column])
		try:
			bazaE2 = [b.value for b in purchases[E2col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  24' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  25":
					rand_tb = cell.row
					E2tvacol = cell.column
					lun = len(purchases[cell.column])
		try:
			tvaE2 = [b.value for b in purchases[E2tvacol][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  25' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  26":
					rand_tb = cell.row
					I0col = cell.column
					lun = len(purchases[cell.column])
		try:
			baza1O = [b.value for b in purchases[I0col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  26' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  27":
					rand_tb = cell.row
					tva10col = cell.column
					lun = len(purchases[cell.column])
		try:
			tva1O = [b.value for b in purchases[tva10col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  27' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  30":
					rand_tb = cell.row
					M1col = cell.column
					lun = len(purchases[cell.column])
		try:
			baza1M = [b.value for b in purchases[M1col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  30' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  31":
					rand_tb = cell.row
					tva1Mcol = cell.column
					lun = len(purchases[cell.column])
		try:
			tva1M = [b.value for b in purchases[tva1Mcol][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  31' in Purchases sheet")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  34":
					rand_tb = cell.row
					tvaN1col = cell.column
					lun = len(purchases[cell.column])
		try:
			bazaN1 = [b.value for b in purchases[tvaN1col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  31' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "                  35":
					rand_tb = cell.row
					tvanedcol = cell.column
					lun = len(purchases[cell.column])
		try:
			neded = [b.value for b in purchases[tvaN1col][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '                  31' in Purchases sheet")
			return render_template("index.html")


		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == " Doc..Date  ":
					rand_tb = cell.row
					supplierCell = cell.column
					lun = len(purchases[cell.column])
		try:
			datadocument = [b.value for b in purchases[supplierCell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for ' Doc..Date  ' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "VAT Registration No.":
					rand_tb = cell.row
					supplierCell = cell.column
					lun = len(purchases[cell.column])
		try:
			cuipurch = [b.value for b in purchases[supplierCell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for ' Doc..Date  ' in Purchases sheet")
			return render_template("index.html")
		nomenclatorTari={'AT':'Austria', 'BE':'Belgia', 'BG':'Bulgaria','CY':'Cipru','DK':'Danemarca','EE':'Estonia', 'FI':'Finlanda','FR':'Franta', 'DE':'Germania','HR':'Croatia',
				'GR':'Grecia','IE':'Irlanda','IT':'Italia','LV':'Letonia','LT':'Lituania','LU':'Luxemburg','MT':'Malta','XI':'Irlanda de Nord - Regatul Unit','NL':'Olanda',
				'PL':'Polonia','PT':'Portugalia','CZ':'Republica Ceha','RO':'Romania','SK':'Slovacia','SI':'Slovavia','ES':'Spania','SE':'Suedia','HU':'Ungaria'}
		
		codintracomunitar=[]
		for k in range(0,len(cuipurch)):
			if(cuipurch[k]!=None):
				if(str(cuipurch[k])[:2] in nomenclatorTari and str(cuipurch[k])[:2]!="RO"):
					codintracomunitar.append("UE")
				else:
					codintracomunitar.append("Not applicable")
			else:
				codintracomunitar.append("Not applicable")

		lunacurenta=[]
		for k in range(0,len(datadocument)):
			try:
				print(datadocument[k][3:4])
				print(datadocument[k][3:5])
				if(str(datadocument[k][4:5])=="0"):
					if(str(datadocument[k][5:6])==str(info.cell(row=3,column=3).value)):
						lunacurenta.append("Yes")
					else:
						lunacurenta.append("No")

				else:
					if(str(datadocument[k][4:6])==str(info.cell(row=3,column=3).value)):
						lunacurenta.append("Yes")
					else:
						lunacurenta.append("No")
			except:
				lunacurenta.append("Not applicable")
			# if(datadocument[k])
		for kk in range(0,len(lunacurenta)):
			purchases.cell(row=rand_tb+1+kk,column=70).value=lunacurenta[kk]
			purchases.cell(row=rand_tb+1+kk,column=71).value=codintracomunitar[kk]								
		purchases.cell(row=rand_tb,column=70).value="Perioada curenta"
		purchases.cell(row=rand_tb,column=71).value="Intracomunitar"		



		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "  Total doc.incl.VAT":
					rand_tb = cell.row
					tdoca = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in purchases[tdoca][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '  Total doc.incl.VAT'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Nedeductibil":
					rand_tb = cell.row
					tdocneded = cell.column
					lun = len(sales[cell.column])
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Neexigibil BAZA 19%":
					rand_tb = cell.row
					tdocnexb = cell.column
					lun = len(sales[cell.column])
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Neexigibil TVA 19%":
					rand_tb = cell.row
					tdocnextva = cell.column
					lun = len(sales[cell.column])		

		amount.cell(row=6, column=2).value="1"
		amount.cell(row=6, column=3).value="2"
		amount.cell(row=7, column=1).value="Row"
		amount.cell(row=7, column=2).value="Taxable basis"
		amount.cell(row=7, column=3).value="VAT amount"
		amount.cell(row=7, column=4).value="Comments"
		amount.cell(row=7, column=5).value="Journal Source"
		amount.cell(row=7, column=6).value="Flag Suma Control"
		amount.cell(row=7, column=7).value="Suma Control"
		amount.cell(row=8, column=1).value="1"
		amount.cell(row=9, column=1).value="2"
		amount.cell(row=10, column=1).value="3"
		amount.cell(row=11, column=1).value="3.1"
		amount.cell(row=12, column=1).value="4"
		amount.cell(row=13, column=1).value="5"
		amount.cell(row=14, column=1).value="5.1"
		amount.cell(row=15, column=1).value="6"
		amount.cell(row=16, column=1).value="7"
		amount.cell(row=17, column=1).value="7.1"
		amount.cell(row=18, column=1).value="8"
		amount.cell(row=19, column=1).value="9"
		amount.cell(row=20, column=1).value="10"
		amount.cell(row=21, column=1).value="11"
		amount.cell(row=22, column=1).value="12"
		amount.cell(row=23, column=1).value="12.1"
		amount.cell(row=24, column=1).value="12.2"
		amount.cell(row=25, column=1).value="12.3"
		amount.cell(row=26, column=1).value="13"
		amount.cell(row=27, column=1).value="14"
		amount.cell(row=28, column=1).value="15"
		amount.cell(row=29, column=1).value="16"
		amount.cell(row=30, column=1).value="17"
		amount.cell(row=31, column=1).value="18"
		amount.cell(row=32, column=1).value="19"
		amount.cell(row=33, column=1).value="20"
		amount.cell(row=34, column=1).value="20.1"
		amount.cell(row=35, column=1).value="21"
		amount.cell(row=36, column=1).value="22"
		amount.cell(row=37, column=1).value="22.1"
		amount.cell(row=38, column=1).value="23"
		amount.cell(row=39, column=1).value="24"
		amount.cell(row=40, column=1).value="25"
		amount.cell(row=41, column=1).value="26"
		amount.cell(row=42, column=1).value="27"
		amount.cell(row=43, column=1).value="27.1"
		amount.cell(row=44, column=1).value="27.2"
		amount.cell(row=45, column=1).value="27.3"
		amount.cell(row=46, column=1).value="28"
		amount.cell(row=47, column=1).value="29"
		amount.cell(row=48, column=1).value="30"
		amount.cell(row=49, column=1).value="30.1"
		amount.cell(row=50, column=1).value="31"
		amount.cell(row=51, column=1).value="31.1"
		amount.cell(row=52, column=1).value="32"
		amount.cell(row=53, column=1).value="33"
		amount.cell(row=54, column=1).value="34"
		amount.cell(row=55, column=1).value="35"
		amount.cell(row=56, column=1).value="36"
		amount.cell(row=57, column=1).value="37"
		amount.cell(row=58, column=1).value="38"
		amount.cell(row=59, column=1).value="39"
		amount.cell(row=60, column=1).value="40"
		amount.cell(row=61, column=1).value="41"
		amount.cell(row=62, column=1).value="42"
		amount.cell(row=63, column=1).value="43"
		amount.cell(row=64, column=1).value="44"
		amount.cell(row=65, column=1).value="45"
		amount.cell(row=66, column=1).value="46"

		amount.cell(row=68, column=1).value="A"
		amount.cell(row=69, column=1).value="A1"
		amount.cell(row=70, column=1).value="B"
		amount.cell(row=71, column=1).value="B.1"
		for w in range(0,len(listadescrieri)):
			amount.cell(row=8+w,column=8).value=listadescrieri[w]

		amount.cell(row=8, column=2).value='0'
		amount.cell(row=9, column=2).value='0'
		amount.cell(row=10, column=2).value='0'		
		amount.cell(row=11, column=2).value='0'
		amount.cell(row=12, column=2).value='0'
		amount.cell(row=13, column=2).value='=ROUND(SUMIFS(Purchases!'+str(W0col)+":"+str(W0col)+',Purchases!$BR:$BR,"Yes",Purchases!$BS:$BS,"UE")+SUMIFS(Purchases!'+str(E2col)+":"+str(E2col)+',Purchases!$BR:$BR,"Yes",Purchases!$BS:$BS,"UE"),0)'
		amount.cell(row=14, column=2).value='=ROUND(SUMIFS(Purchases!'+str(W0col)+":"+str(W0col)+',Purchases!$BR:$BR,"Yes",Purchases!$BS:$BS,"UE")+SUMIFS(Purchases!'+str(E2col)+":"+str(E2col)+',Purchases!$BR:$BR,"Yes",Purchases!$BS:$BS,"UE"),0)'
		amount.cell(row=15, column=2).value='0'
		amount.cell(row=16, column=2).value='=round(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(I0col)+':'+str(I0col)+')+SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(M1col)+':'+str(M1col)+'),0)'
		amount.cell(row=17, column=2).value='=round(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(I0col)+':'+str(I0col)+'),0)'
		amount.cell(row=18, column=2).value='=round(SUMIF(Purchases!BR:BR,"No",Purchases!'+str(I0col)+':'+str(I0col)+'),0)'		
		amount.cell(row=19,column=2).value='=round(sum(Sales!'+str(A12col)+":"+str(A12col)+"),0)"
		amount.cell(row=20,column=2).value='=round(sum(Sales!'+str(A8col)+":"+str(A8col)+"),0)"
				
		amount.cell(row=21, column=2).value=0
		amount.cell(row=22, column=2).value=0
		amount.cell(row=23, column=2).value=0
		amount.cell(row=24, column=2).value=0
		amount.cell(row=25, column=2).value=0
		amount.cell(row=26, column=2).value=0
		amount.cell(row=27, column=2).value=0
		amount.cell(row=28, column=2).value=0
		amount.cell(row=29, column=2).value=0
		amount.cell(row=30, column=2).value=0
		amount.cell(row=31, column=2).value=0												

		# # amount.cell(row=21, column=2).value='=round(ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"5G",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)/(105/100),0)'
		# # amount.cell(row=22, column=2).value='=ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"ZR",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)'
		# # amount.cell(row=23, column=2).value='=ROUND(SUMIF(Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"ZR",Purchases!'+str(tdoca)+":"+str(tdoca)+'),0)'
		# # amount.cell(row=24, column=2).value=0
		# # amount.cell(row=25, column=2).value=0
		# # amount.cell(row=26, column=2).value='=ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"1V",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)'	
		# # amount.cell(row=27, column=2).value='=ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"A5",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)+ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"A4",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)+ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"Y8",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)'
		# amount.cell(row=28, column=2).value=0
		# amount.cell(row=30, column=2).value=0
		# amount.cell(row=29, column=2).value='=ROUND(SUMIF(Sales!'+str(taxcodec)+":"+str(taxcodec)+',"ZJ",Sales!'+str(tdocc)+":"+str(tdocc)+'),0)'
		amount.cell(row=39, column=2).value='=round(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(bazaV1col)+':'+str(bazaV1col)+'),0)'						
		amount.cell(row=40, column=2).value='=round(SUM(Purchases!'+str(bazaW7col)+":"+str(bazaW7col)+"),0)"								
		amount.cell(row=41, column=2).value='=round(SUM(Purchases!'+str(bazaW8col)+":"+str(bazaW8col)+"),0)"
										

		
		# amount.cell(row=22, column=2).value='=round(SUM(B23:B25),0)'


		# amount.cell(row=31, column=2).value=0
		amount.cell(row=32, column=2).value='=B8+B10+B13+B16+B27+B15+B18+B31+B30+B29+B28+B26+B22+B21+B20+B19'
		amount.cell(row=33, column=2).value='=B13'
		amount.cell(row=34, column=2).value='=B14'
		amount.cell(row=35, column=2).value='=B15'
		amount.cell(row=36, column=2).value='=B16'
		amount.cell(row=37, column=2).value='=B17'
		amount.cell(row=38, column=2).value='=B18'

		
		amount.cell(row=42, column=2).value='=SUM(B43:B45)'
		amount.cell(row=43, column=2).value='=B23'
		amount.cell(row=44, column=2).value='=B24'
		amount.cell(row=45, column=2).value='=B25'
		amount.cell(row=46, column=2).value=0
		amount.cell(row=47, column=2).value=0
		amount.cell(row=48, column=2).value='=round(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(bazaV1col)+':'+str(bazaV1col)+')+SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(bazaW7col)+':'+str(bazaW7col)+')+SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(bazaW8col)+':'+str(bazaW8col)+'),0)'
		amount.cell(row=49, column=2).value=0
		amount.cell(row=50, column=2).value='=SUM(B33:B47)-B34-B37-SUM(B43:B45)'
		amount.cell(row=51, column=2).value='0'
		amount.cell(row=52, column=2).value='=B50+B51'
		amount.cell(row=53, column=2).value=0
		amount.cell(row=54, column=2).value='=round(SUMIFS(Purchases!'+str(bazaV1col)+':'+str(bazaV1col)+',Purchases!BQ:BQ,"No",Purchases!BR:BR,"Not applicable")+SUMIFS(Purchases!'+str(bazaW7col)+':'+str(bazaW7col)+',Purchases!BQ:BQ,"No",Purchases!BR:BR,"Not applicable"),0)'
		amount.cell(row=55, column=2).value=0
		amount.cell(row=56, column=2).value='=SUM(B52:B55)'
		amount.cell(row=57, column=2).value=0
		amount.cell(row=58, column=2).value=0
		amount.cell(row=59, column=2).value=0
		amount.cell(row=60, column=2).value=0
		amount.cell(row=61, column=2).value='=SUM(B58:B60)'
		amount.cell(row=62, column=2).value=0
		amount.cell(row=63, column=2).value=0
		amount.cell(row=64, column=2).value='=B57+B62+B63'
		amount.cell(row=65, column=2).value='=IF((B61-B664)<0,0,B61-B64)'
		amount.cell(row=66, column=2).value='=IF((B64-B61)<0,0,B64)'

		amount.cell(row=68, column=2).value=0
		amount.cell(row=69, column=2).value=0
		amount.cell(row=70, column=2).value='0'
		amount.cell(row=71, column=2).value='=B70'
		
		# #coloana TVA----------------------------------------------------

		for g in range(8, 13):
			amount.cell(row=g, column=3).value=0
		

		# for h in range(13, 19):
		amount.cell(row=13, column=3).value='=round(B13/100*19,0)'
		amount.cell(row=14, column=3).value='=round(B14/100*19,0)'
		# amount.cell(row=13, column=3).value='=(ROUND(SUMIFS(Purchases!'+str(intracomtaxe1)+":"+str(intracomtaxe1)+',Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"E1",Purchases!BR:BR,"Yes"),0))'
		# amount.cell(row=14, column=3).value='=(ROUND(SUMIFS(Purchases!'+str(intracomtaxe1)+":"+str(intracomtaxe1)+',Purchases!'+str(taxcodea)+":"+str(taxcodea)+',"E1",Purchases!BR:BR,"Yes"),0))'
		amount.cell(row=15, column=3).value='=round(B15/100*19,0)'
		amount.cell(row=16, column=3).value='=round(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tva10col)+':'+str(tva10col)+')+SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tva1Mcol)+':'+str(tva1Mcol)+'),0)'
		amount.cell(row=17, column=3).value='=round(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tva10col)+':'+str(tva10col)+'),0)'
		amount.cell(row=18, column=3).value='=round(SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tva10col)+':'+str(tva10col)+'),0)'		
		amount.cell(row=19,column=3).value='=round(sum(Sales!'+str(A12tva)+":"+str(A12tva)+"),0)"
		amount.cell(row=20,column=3).value='=round(sum(Sales!'+str(A8tva)+":"+str(A8tva)+"),0)"
			
		# # amount.cell(row=16,column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A16&"."&C$6,Purchases!$5:$5)-SUMIF(Purchases!$7:$7,$A18&"."&C$6,Purchases!$5:$5),0)'


		# amount.cell(row=19, column=3).value='=round(B19/100*19,0)'
		# amount.cell(row=20, column=3).value='=round(B20/100*9,0)'
		amount.cell(row=21, column=3).value='=round(B21/100*5,0)'
		amount.cell(row=22, column=3).value='0'
		amount.cell(row=23, column=3).value='0'
		amount.cell(row=24, column=3).value='0'
		amount.cell(row=25, column=3).value='0'
		amount.cell(row=26, column=3).value='0'
		amount.cell(row=27, column=3).value='0'
		amount.cell(row=28, column=3).value='0'
		amount.cell(row=29, column=3).value='0'
		amount.cell(row=30, column=3).value='0'
		amount.cell(row=31, column=3).value='0'																		

		# amount.cell(row=23, column=3).value='=round(B23/100*19,0)'
		# amount.cell(row=24, column=3).value='=round(B24/100*9,0)'
		# amount.cell(row=25, column=3).value='=round(B25/100*5,0)'


		
		# for k in range(26, 31):
		# 	amount.cell(row=k, column=3).value=0
		# amount.cell(row=31, column=3).value=0
		amount.cell(row=32, column=3).value='=C8+C10+C13+C16+C27+C15+C18+C31+C30+C29+C28+C26+C22+C21+C20+C19'
		amount.cell(row=33, column=3).value='=C13'
		amount.cell(row=34, column=3).value='=C14'
		amount.cell(row=35, column=3).value='=C15'
		amount.cell(row=36, column=3).value='=C16'
		amount.cell(row=37, column=3).value='=C17'
		amount.cell(row=38, column=3).value='=C18'
		amount.cell(row=39, column=3).value='=round(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tvaV1col)+':'+str(tvaV1col)+'),0)'						
		amount.cell(row=40, column=3).value='=round(SUM(Purchases!'+str(tvaW7col)+":"+str(tvaW7col)+"),0)"								
		amount.cell(row=41, column=3).value='=round(SUM(Purchases!'+str(tvaw8col)+":"+str(tvaw8col)+"),0)"

		# amount.cell(row=39, column=3).value='=SUM(Purchases!'+str(tax19vat)+':'+str(tax19vat)+')-SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax19vat)+':'+str(tax19vat)+')+SUMIFS(Purchases!'+str(tax19vat)+':'+str(tax19vat)+',Purchases!'+str(taxcodea)+':'+str(taxcodea)+',"ZI",Purchases!BR:BR,"No")'
		# amount.cell(row=40, column=3).value='=round(B40/100*9,0)'
		# amount.cell(row=41, column=3).value='=round(B41/100*5,0)'			
		amount.cell(row=42, column=3).value='=round(SUM(C43:C45),0)'
		amount.cell(row=43, column=3).value='=C23'
		amount.cell(row=44, column=3).value='=C24'
		amount.cell(row=45, column=3).value='=C25'
		amount.cell(row=46, column=3).value=0
		amount.cell(row=47, column=3).value=0
		amount.cell(row=48, column=3).value=0
		amount.cell(row=49, column=3).value=0
		amount.cell(row=50, column=3).value='=SUM(C33:C47)-C34-C37-SUM(C43:C45)'
		amount.cell(row=51, column=3).value='0'
		amount.cell(row=52, column=3).value='=C51+C50-sum(Purchases!'+str(tvanedcol)+":"+str(tvanedcol)+')'
		amount.cell(row=53, column=3).value=0
		amount.cell(row=54, column=3).value='=round(SUMIFS(Purchases!'+str(tvaV1col)+':'+str(tvaV1col)+',Purchases!BQ:BQ,"No",Purchases!BR:BR,"Not applicable")+SUMIFS(Purchases!'+str(tvaW7col)+':'+str(tvaW7col)+',Purchases!BQ:BQ,"No",Purchases!BR:BR,"Not applicable"),0)'
		amount.cell(row=55, column=3).value=0
		amount.cell(row=56, column=3).value='=SUM(C52:C55)'
		amount.cell(row=57, column=3).value='=IF((C56-C32)<0,0,C56-C32)'
		amount.cell(row=58, column=3).value='=IF((C32-C56)<0,0,C32-C56)'
		amount.cell(row=59, column=3).value=0
		amount.cell(row=60, column=3).value=0
		amount.cell(row=61, column=3).value='=SUM(C58:C60)'	
		if soldLunaTrecuta == None or soldLunaTrecuta == "" or soldLunaTrecuta == " ":
			amount.cell(row=62, column=3).value=0
		else:
			amount.cell(row=62, column=3).value=int(soldLunaTrecuta) 
		#print(soldLunaTrecuta, "sold luna trecuta")
		amount.cell(row=63, column=3).value=0
		amount.cell(row=64, column=3).value='=C57+C62+C63'
		amount.cell(row=65, column=3).value='=IF((C61-C64)<0,0,C61-C64)'
		amount.cell(row=66, column=3).value='=IF((C64-C61)<0,0,C64-C61)'
		amount.cell(row=68, column=3).value=0
		amount.cell(row=69, column=3).value=0

		amount.cell(row=70, column=3).value='0'
		amount.cell(row=71, column=3).value='=C70'

		amount.cell(row=73, column=1).value='Informații privind valoarea totală, fără TVA, a operațiunilor prevăzute la art. 2781 alin. (1) lit. b) din Codul fiscal, respectiv a vânzărilor intracomunitare de bunuri la distanță și a prestărilor de servicii de telecomunicaţii, de radiodifuziune şi televiziune, precum și servicii furnizate pe cale electronică, către persoane neimpozabile din alte state membre UE'
		amount.cell(row=73, column=2).value='Total an precedent'
		amount.cell(row=73, column=3).value='An curent (inclusiv perioada de raportare)'

		amount.cell(row=74, column=2).value=0
		amount.cell(row=74, column=3).value=0


		amount.cell(row=22, column=5).value='Total'

		for m in it.chain(range(8, 13), range(19, 22), range(26, 32)):
			amount.cell(row=m, column=5).value='SALES'

		for n in it.chain(range(13, 19), range(23, 26), range(39, 42), range(46, 50), range(53, 56)):
			amount.cell(row=n, column=5).value='Purchases'
		
		for o in range(32, 39):
			amount.cell(row=o, column=5).value='Total'

		for p in range(42, 46):
			amount.cell(row=p, column=5).value='Total'
		
		amount.cell(row=50, column=5).value='Total'
		amount.cell(row=51, column=5).value='Purchases'
		amount.cell(row=52, column=5).value='Total'
		amount.cell(row=66, column=5).value='Purchases'

		for q in range(68, 71):
			amount.cell(row=q, column=5).value='Total'

		for r in range(56, 67):
			amount.cell(row=r, column=5).value='Total'

		amount.cell(row=71, column=5).value='Purchases'

		for s in it.chain(range(8, 13), range(26, 29), range(48, 50)):
			amount.cell(row=s, column=6).value='no VAT'

		for t in it.chain(range(13, 26), range(29, 46)):
			amount.cell(row=t, column=6).value='Add all'

		for u in it.chain(range(46, 48), range(52, 54), range(55, 67)):
			amount.cell(row=u, column=6).value='No basis'
		
		amount.cell(row=50, column=6).value='Add all'
		amount.cell(row=51, column=6).value='Se pune 0'
		amount.cell(row=54, column=6).value='Add all'

		for a1 in range(8, 13):
			amount.cell(row=a1, column=7).value='=B{0}'.format(a1)

		for b1 in range(13, 26):
			amount.cell(row=b1, column=7).value='=SUM(B{0}:C{0})'.format(b1)

		for c1 in range(26, 29):
			amount.cell(row=c1, column=7).value='=B{0}'.format(c1)
		
		for d1 in range(29, 46):
			amount.cell(row=d1, column=7).value='=SUM(B{0}:C{0})'.format(d1)

		amount.cell(row=46, column=7).value='=C46'
		amount.cell(row=47, column=7).value='=C47'
		amount.cell(row=48, column=7).value='=B48'
		amount.cell(row=49, column=7).value='=B49'
		amount.cell(row=50, column=7).value='=SUM(B50:C50)'
		amount.cell(row=51, column=7).value='=C51'
		amount.cell(row=52, column=7).value='=C52'
		amount.cell(row=53, column=7).value='=SUM(B53:C53)'
		amount.cell(row=54, column=7).value='=SUM(B54:C54)'
		amount.cell(row=2,column=1).value="D300 draft figures "
		amount.cell(row=2,column=1).font=cap_tabeltitlu
		amount.row_dimensions[5].hidden = True
		for e1 in range(55, 67):
			amount.cell(row=e1, column=7).value='=C{0}'.format(e1)
		
		for row in amount['B8:B74']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'
		
		for row in amount['C8:C74']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'

		#------foratare D300

		for row in amount['A7:C7']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.alignment=Alignment(horizontal='center',vertical='center')
		for row in amount['D7:G7']:
			for cell in row:
				cell.fill=cap_tabel_color_black
		
		for row in amount['A7:G7']:
			for cell in row:
				cell.font=cap_tabel

		listanoua=['A','B','C','D','E','F','G']

		for column in listanoua:
			for i in listanoua:
				if (column==i):
					amount.column_dimensions[column].width = 17
		
		amount.column_dimensions['D'].hidden = True
		amount.column_dimensions['E'].hidden = True
		amount.column_dimensions['F'].hidden = True
		amount.column_dimensions['G'].hidden = True
		info=temp['Other info']
	
		listaMapare=["L", "T", "S", "A"]
		Poz="10"
		Fix01="01"
		Fix0000="0000"
		dictMapare={'L':'301', 'T':'302', 'S':'303', 'A':'304'}

		elementMapare=""
		if info.cell(row=53, column=3).value=="L":
			elementMapare="301"
		else:
			if info.cell(row=53, column=3).value == "T":
				elementMapare="302"
			else:
				if info.cell(row=53, column=3).value == "S":
					elementMapare="303"
				else:
					if info.cell(row=53, column=3).value == "A":
						elementMapare="304"
			
		# #print(elementMapare, 'MAPARE')

		LL=""
		AA=""

		if len(str(info.cell(row=3, column=3).value))==2:
			LL=str(info.cell(row=3, column=3).value)
			LL_g=LL
		else:
			LL="0"+str(info.cell(row=3, column=3).value)

		AA=str(info.cell(row=2, column=3).value)[2:]
		strYear=str(info.cell(row=2, column=3).value)[2:]
		intYear=int(strYear)+1
		year=str(intYear)
		
		# #print(year, 'an')

		LLAA=LL+AA
		LL2=""
		AA2=""

		if int(info.cell(row=3, column=3).value)==12:
			LL2="1"
		if len(str(info.cell(row=3, column=3).value))==1:
			if(int(info.cell(row=3,column=3).value)==9):
				LL2=str(int(info.cell(row=3,column=3).value)+1)
			else:
				LL2="0"+str(int(info.cell(row=3, column=3).value)+1)
			# #print(LL2, 'LL2')
		
		if int(info.cell(row=3, column=3).value)==12:
			AA2=year 
		else:
			AA2=strYear

		ZZLLAA="25"+str(LL2)+str(AA2)
		# #print(ZZLLAA, 'ZZLLAA')

		poz1=""
		LTSA=""
		fix1=""
		LLAA1=""
		ZZLLAA1=""
		fix01=""
		control=""

		poz1 = sum(int(digit) for digit in str(Poz))
		LTSA = sum(int(digit) for digit in str(elementMapare))
		fix1=sum(int(digit) for digit in str(Fix01))
		LLAA1=sum(int(digit) for digit in str(LLAA))
		ZZLLAA1=sum(int(digit) for digit in str(ZZLLAA))
		fix01=sum(int(digit) for digit in str(Fix0000))

		control=poz1+LTSA+fix1+LLAA1+ZZLLAA1+fix01
		

		nrEvidenta=Poz+elementMapare+Fix01+LLAA+ZZLLAA+Fix0000+str(control)
		#print(nrEvidenta, 'nrEvidenta')
			
		
		info.cell(row=18, column=3).value=nrEvidenta
		info.cell(row=52, column=3).value="=SUM('D300 draft figures'!G8:G66)"
		try:
			info.cell(row=50, column=3).value=lenbfi
			info.cell(row=51, column=3).value=sumabfi
			info.cell(row=52, column=3).value=sumatbfi
		except:
			pass
		try:
			info.cell(row=53, column=3).value=sumabif
			info.cell(row=54, column=3).value=sumatbif
			info.cell(row=55, column=3).value=lentbif
		except:
			pass
		try: 
			# info.cell(row=56, column=3).value=lenbfr
			info.cell(row=57, column=3).value=sumabfr
			info.cell(row=58, column=3).value=sumatbfr
		except:
			pass


		if(val2==1):
			sheetinutil2=temp.create_sheet('D390--->>>')
			sheetinutil2.sheet_view.showGridLines=False
			sheetinutil2.cell(row=2,column=1).value="Switch to next sheet for D390 Workings draft"
			sheetinutil2.cell(row=2,column=1).font=scrisincredibildemare		

			workings=temp.create_sheet('D390 workings')
			workings.cell(row=1,column=1).value="D390 workings"
			workings.cell(row=1,column=1).font=cap_tabelbold
			workings.freeze_panes = 'A4'
			workings.auto_filter.ref = "A3:I10000"
			workings.sheet_view.showGridLines = False
			workings.column_dimensions['I'].hidden = True
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "                 S01":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				bazaA12 = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                 S01'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "                 S02":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				tvaA12 = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '    S02'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "                 S03":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				bazaA8 = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                 S03'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "                 S04":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				tvaA8 = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                 S04'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "                 S25":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				valY8 = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                 S25'")
				return render_template("index.html")			
										
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "  Total doc.incl.VAT":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(sales[cell.column])
			try:
				listBazaS = [b.value for b in sales[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '  Total doc.incl.VAT'")
				return render_template("index.html")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "Posting Date":
						rand_tb = cell.row
						tdat = cell.column
						lun = len(sales[cell.column])
			try:
				listdocdate = [b.value for b in sales[tdat][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Posting Date'")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   1":
						rand_tb = cell.row
						V1cell = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaV1 = [b.value for b in purchases[V1cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   1' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   2":
						rand_tb = cell.row
						V1tva = cell.column
						lun = len(purchases[cell.column])
			try:
				tvaV1 = [b.value for b in purchases[V1tva][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                    2' in Purchases sheet")
				return render_template("index.html")			

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   3":
						rand_tb = cell.row
						W7col = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaW7 = [b.value for b in purchases[W7col][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   3' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   4":
						rand_tb = cell.row
						W7tva = cell.column
						lun = len(purchases[cell.column])
			try:
				tvaW7 = [b.value for b in purchases[W7tva][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   4' in Purchases sheet")
				return render_template("index.html")		

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   5":
						rand_tb = cell.row
						W8col = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaW8 = [b.value for b in purchases[W8col][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   5' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   6":
						rand_tb = cell.row
						W8tva = cell.column
						lun = len(purchases[cell.column])
			try:
				tvaW8 = [b.value for b in purchases[W8tva][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   6' in Purchases sheet")
				return render_template("index.html")	

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   7":
						rand_tb = cell.row
						col1L = cell.column
						lun = len(purchases[cell.column])
			try:
				baza1L = [b.value for b in purchases[col1L][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   7' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   8":
						rand_tb = cell.row
						col1Ltva = cell.column
						lun = len(purchases[cell.column])
			try:
				tva1L = [b.value for b in purchases[col1Ltva][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   8' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  10":
						rand_tb = cell.row
						col1J = cell.column
						lun = len(purchases[cell.column])
			try:
				baza1J = [b.value for b in purchases[col1J][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  10' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  11":
						rand_tb = cell.row
						col1JTva = cell.column
						lun = len(purchases[cell.column])
			try:
				tva1J = [b.value for b in purchases[col1JTva][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  11' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  22":
						rand_tb = cell.row
						W0col = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaW0 = [b.value for b in purchases[W0col][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  22' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  23":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaV9 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  23' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  24":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaE2 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  24' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  25":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tvaE2 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  25' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  26":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				baza1O = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  26' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  27":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tva1O = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  27' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  30":
						rand_tb = cell.row
						M1col = cell.column
						lun = len(purchases[cell.column])
			try:
				baza1M = [b.value for b in purchases[M1col][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  30' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  31":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tva1M = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  31' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == " Doc..Date  ":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				datadocument = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for ' Doc..Date  ' in Purchases sheet")
				return render_template("index.html")

			




			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "  Total doc.incl.VAT":
						rand_tb = cell.row
						tdoca = cell.column
						lun = len(sales[cell.column])
			try:
				listBazaL = [b.value for b in purchases[tdoca][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '  Total doc.incl.VAT'")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Nedeductibil":
						rand_tb = cell.row
						tdocneded = cell.column
						lun = len(sales[cell.column])
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Neexigibil BAZA 19%":
						rand_tb = cell.row
						tdocnexb = cell.column
						lun = len(sales[cell.column])
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Neexigibil TVA 19%":
						rand_tb = cell.row
						tdocnextva = cell.column
						lun = len(sales[cell.column])
			workings.cell(row=3, column=1).value='TIP'
			workings.cell(row=3, column=2).value='ŢARA'
			workings.cell(row=3, column=3).value='COD OPERATOR INTRACOMUNITAR'
			workings.cell(row=3, column=4).value='DENUMIRE'
			workings.cell(row=3, column=5).value='BAZA IMPOZABILĂ'
			workings.cell(row=3, column=6).value='CIF'
			workings.cell(row=3, column=7).value='Country Code'
			workings.cell(row=3, column=8).value='BAZA IMPOZABILĂ'
			workings.cell(row=3, column=9).value='Cheie extragere - filtreaza 1'

			# bazaA_Furnizor=get_column_letter(7)
			# bazaA_index=get_column_letter(bazaA)
			# bazaA_literaA=get_column_letter(bazaA)
			# #print(bazaA_Furnizor, "litera pentru furnizor/ client")
			# #print(bazaA_index, "ASTA E NUMARUL LUI A")
			# #print(bazaA_literaA, "LITERA PENTRU TIP")
			taxcodeach=[]
			for i in range(0,len(listBazaL)):
				if(int(bazaV1[i])>0 or int(bazaV1[i]<0)):
					taxcodeach.append("V1")
				else:
					if(int(bazaW7[i])>0 or int(bazaW7[i]<0)):
						taxcodeach.append("W7")
					else:
						if(int(bazaW8[i])>0 or int(bazaW8[i]<0)):
							taxcodeach.append("W8")
						else:
							if(int(baza1L[i])>0 or int(baza1L[i]<0)):
								taxcodeach.append("1L")
							else:
								if(int(baza1J[i])>0 or int(baza1J[i]<0)):
									taxcodeach.append("1J")
								else:
									if(int(bazaW0[i])>0 or int(bazaW0[i]<0)):
										taxcodeach.append("W0")
									else:
										if(int(bazaV9[i])>0 or int(bazaV9[i]<0)):
											taxcodeach.append("V9")
										else:
											if(int(bazaE2[i])>0 or int(bazaE2[i]<0)):
												taxcodeach.append("E2")
											else:
												if(int(baza1O[i])>0 or int(baza1O[i]<0)):
													taxcodeach.append("1O")
												else:
													if(int(baza1M[i])>0 or int(baza1M[i]<0)):
														taxcodeach.append("1M")
													else:
														if(int(bazaN1[i])>0 or int(bazaN1[i]<0)):
															taxcodeach.append("N1")
			taxcodes=[]
			for j in range(0,len(listBazaS)):
				if(int(bazaA12[j])>0 or int(bazaA12[i])<0):
					taxcodes.append("A1")
				else:
					if(int(bazaA8[j])>0 or int(bazaA8[i])<0):
						taxcodes.append("A8")
					else:
						if(int(valY8[j])>0 or int(valY8[i])<0):
							taxcodes.append("A8")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Business PartnerName":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				denumirea = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Business PartnerName' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "VAT Registration No.":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				vata = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'VAT Registration No.' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "  Total doc.incl.VAT":
						rand_tb = cell.row
						totaldoc = cell.column
						lun = len(purchases[cell.column])
			try:
				totala = [b.value for b in purchases[totaldoc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '  Total doc.incl.VAT' in Purchases sheet")
				return render_template("index.html")				

			print(len(taxcodeach))
			for k in range(0,len(taxcodeach)):
				purchases.cell(row=19+k,column=46).value=taxcodeach[k]
			a=3
			for x in range(0, len(taxcodeach)):
				print(taxcodeach[x])
				if(lunacurenta[x]=="Yes" and codintracomunitar[x]=="UE"):
					
					if str(taxcodeach[x])=="W0" or str(taxcodeach[x])=="E2" :
						a=a+1
						workings.cell(row=a, column=1).value="A"
						workings.cell(row=a, column=4).value=denumirea[x]
						workings.cell(row=a, column=6).value=vata[x]
						workings.cell(row=a, column=3).value=vata[x][2:]
						workings.cell(row=a, column=7).value=vata[x][0:2]
						# workings.cell(row=a, column=8).value=listaBazaA[x]
						# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
						
						if(taxcodeach[x]=="W0"):
							workings.cell(row=a, column=8).value=bazaW0[x]
						else:
							workings.cell(row=a,column=8).value=bazaE2[x]
						workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
						workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)


			for x in range(0, len(taxcodeach)):
				if lunacurenta[x]=="Yes":				
					if str(taxcodeach[x])=="1O":
						a=a+1
						workings.cell(row=a, column=1).value="S"
						workings.cell(row=a, column=4).value=denumirea[x]
						workings.cell(row=a, column=6).value=vata[x]
						workings.cell(row=a, column=3).value=vata[x][2:]
						workings.cell(row=a, column=7).value=vata[x][0:2]
						# workings.cell(row=a, column=8).value=listaBazaA[x]
						# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
						workings.cell(row=a, column=8).value=baza1O[x]
						workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
						workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)

			for x in range(0, len(taxcodes)):
				if(listacurentas[x]=="Yes"):				
					if str(taxcodes[x])=="Y1":
						a=a+1
						workings.cell(row=a, column=1).value="L"
						workings.cell(row=a, column=4).value=denumires[x]
						workings.cell(row=a, column=6).value=vats[x]
						workings.cell(row=a, column=3).value=vats[x][2:]
						workings.cell(row=a, column=7).value=vats[x][0:2]
						# workings.cell(row=a, column=8).value=listaBazaA[x]
						# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
						workings.cell(row=a, column=8).value=totals[x]
						workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
						workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)
			for x in range(0, len(taxcodes)):
				if(listacurentas[x]=="Yes"):				
					if str(taxcodes[x])=="Y4":
						a=a+1
						workings.cell(row=a, column=1).value="P"
						workings.cell(row=a, column=4).value=denumires[x]
						workings.cell(row=a, column=6).value=vats[x]
						workings.cell(row=a, column=3).value=vats[x][2:]
						workings.cell(row=a, column=7).value=vats[x][0:2]
						# workings.cell(row=a, column=8).value=listaBazaA[x]
						# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
						workings.cell(row=a, column=8).value=totals[x]
						workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
						workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)
			
			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "TIP":
						rand_tb = cell.row
						tip = cell.column
						lun = len(workings[cell.column])
			try:
				listaTip = [b.value for b in workings[tip][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TIP' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "CIF":
						rand_tb = cell.row
						cod = cell.column
						lun = len(workings[cell.column])
			try:
				codPartener = [b.value for b in workings[cod][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'CIF' in Workings sheet")
				return render_template("index.html")


			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "Country Code":
						rand_tb = cell.row
						country = cell.column
						lun = len(workings[cell.column])
			try:
				countryCode = [b.value for b in workings[country][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Country Code' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "DENUMIRE":
						rand_tb = cell.row
						numep = cell.column
						lun = len(workings[cell.column])
			try:
				partnerName = [b.value for b in workings[numep][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'DENUMIRE' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "BAZA IMPOZABILĂ":
						rand_tb = cell.row
						suma = cell.column
						lun = len(workings[cell.column])
			try:
				sumaTot = [b.value for b in workings[suma][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'BAZA IMPOZABILĂ' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "Cheie extragere - filtreaza 1":
						rand_tb = cell.row
						cheie_sort = cell.column
						lun = len(workings[cell.column])
			try:
				cheie = [b.value for b in workings[cheie_sort][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Cheie extragere - filtreaza 1' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "COD OPERATOR INTRACOMUNITAR":
						rand_tb = cell.row
						coi = cell.column
						lun = len(workings[cell.column])
			try:
				listaCOI = [b.value for b in workings[coi][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'BAZA IMPOZABILĂ' in Workings sheet")
				return render_template("index.html")


			nomenclatorTari={'AT':'Austria', 'BE':'Belgia', 'BG':'Bulgaria','CY':'Cipru','DK':'Danemarca','EE':'Estonia', 'FI':'Finlanda','FR':'Franta', 'DE':'Germania','HR':'Croatia',
							'GR':'Grecia','IE':'Irlanda','IT':'Italia','LV':'Letonia','LT':'Lituania','LU':'Luxemburg','MT':'Malta','XI':'Irlanda de Nord - Regatul Unit','NL':'Olanda',
							'PL':'Polonia','PT':'Portugalia','CZ':'Republica Ceha','RO':'Romania','SK':'Slovacia','SI':'Slovavia','ES':'Spania','SE':'Suedia','HU':'Ungaria'}
			b=3
			for i in countryCode:
				if i in nomenclatorTari:
					b=b+1
					workings.cell(row=b, column=2).value=nomenclatorTari[i]

			for row in workings['A3:I3']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')
			for row in workings['E4:E10000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'
			
			for row in workings['H4:H10000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'

			workings.column_dimensions['B'].width=20
			workings.column_dimensions['D'].width=35				

			forxml = temp.create_sheet('D390 for XML')
			forxml.cell(row=1,column=1).value="D390 for XML"
			forxml.cell(row=1,column=1).font=cap_tabelbold
			forxml.freeze_panes = 'A4'
			forxml.auto_filter.ref = "A3:F10000"
			forxml.sheet_view.showGridLines = False

			forxml.cell(row=3, column=1).value="III.B"
			forxml.cell(row=3, column=2).value="TIP"
			forxml.cell(row=3, column=3).value="ŢARA"
			forxml.cell(row=3, column=4).value="COD OPERATOR INTRACOMUNITAR"
			forxml.cell(row=3, column=5).value="Denumire"
			forxml.cell(row=3, column=6).value="BAZA IMPOZABILĂ"

			codeAndType=[]
			codeAndName=[]
			typeAndName=[]
			# typeCodeName=[]

			# for k in range(0,len(listaTip)):
			# 	codeAndType.append(str(listaTip[k])+" "+str(codPartener[k]))
			# 	codeAndName.append(str(listaTip[k])+" "+str(partnerName[k]))

			# #print(codeAndType,'codeandtyp')
			# codeAndTypeUnique=list(set(codeAndType))
			# codeAndNameUnique=list(set(codeAndName))

			for i in range(0, len(listaTip)):
				typeAndName.append(str(listaTip[i])+";;;"+str(partnerName[i])+";;;"+str(countryCode[i])+";;;"+str(listaCOI[i]))
			#print(typeAndName, 'TYPEAndNAME')

			typeAndNameUni=list(set(typeAndName))

			typeAndNameUni=list(set(typeAndName))

			for i in it.chain(range(0, len(typeAndNameUni))):
				x=typeAndNameUni[i].split(";;;")
				forxml.cell(row=4+i, column=2).value=str(x[0])
				forxml.cell(row=4+i, column=3).value=str(x[2])
				forxml.cell(row=4+i, column=4).value=str(x[3])
				forxml.cell(row=4+i, column=5).value=str(x[1])
				forxml.cell(row=4+i, column=6).value="=SUMIFS('D390 workings'!H:H,'D390 workings'!A:A,B{0},'D390 workings'!C:C,D{0},'D390 workings'!G:G,C{0})".format(4+i)


			for row in forxml['A3:F3']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in forxml['F4:F10000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'

			forxml.column_dimensions['D'].width=27
			forxml.column_dimensions['E'].width=35
			forxml.column_dimensions['F'].width=15


	#---------------------------NR DE EVIDENTA
		if(val3==1):
			sheetinutil3=temp.create_sheet('D394--->>>')
			sheetinutil3.sheet_view.showGridLines=False
			sheetinutil3.cell(row=2,column=1).value="Switch to next sheet for D394 Workings draft"
			sheetinutil3.cell(row=2,column=1).font=scrisincredibildemare		
			nomenclatorTari={'AT':'Austrie', 'BE':'Belgia', 'BG':'Bulgaria','CY':'Cipru','DK':'Danemarca','EE':'Estonia', 'FI':'Finlanda','FR':'Franta', 'DE':'Germania','HR':'Croatia',
							'GR':'Grecia','IE':'Irlanda','IT':'Italia','LV':'Letonia','LT':'Lituania','LU':'Luxemburg','MT':'Malta','XI':'Irlanda de Nord - Regatul Unit','NL':'Olanda',
							'PL':'Polonia','PT':'Portugalia','CZ':'Republica Ceha','RO':'Romania','SK':'Slovacia','SI':'Slovavia','ES':'Spania','SE':'Suedia','HU':'Ungaria'}


			salesExcel=temp.create_sheet("Mapping tranzactii")
			salesExcel.sheet_view.showGridLines = False
			salesExcel.cell(row=2,column=1).value="Mapping tranzactii"
			salesExcel.cell(row=2,column=1).font=cap_tabeltitlu	
			salesExcel.freeze_panes = 'A10'
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "                 S01":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				bazaA12 = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                 S01'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "                 S02":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				tvaA12 = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                 S02'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "                 S03":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				bazaA8 = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                 S03'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "                 S04":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				tvaA8 = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                 S04'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "                 S25":
						rand_tb = cell.row
						taxcodec = cell.column
						lun = len(sales[cell.column])
			try:
				valY8 = [b.value for b in sales[taxcodec][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                 S25'")
				return render_template("index.html")			
										
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "  Total doc.incl.VAT":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(sales[cell.column])
			try:
				listBazaS = [b.value for b in sales[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '  Total docinclVAT'")
				return render_template("index.html")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "Posting Date":
						rand_tb = cell.row
						tdat = cell.column
						lun = len(sales[cell.column])
			try:
				listdocdate = [b.value for b in sales[tdat][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Posting Date'")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   1":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaV1 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   1' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   2":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tvaV1 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                    2' in Purchases sheet")
				return render_template("index.html")			

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   3":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaW7 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   3' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   4":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tvaW7 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   4' in Purchases sheet")
				return render_template("index.html")		

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   5":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaW8 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   5' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   6":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tvaW8 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   6' in Purchases sheet")
				return render_template("index.html")	

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   7":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				baza1L = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   7' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                   8":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tva1L = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                   8' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  10":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				baza1J = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  10' in Purchases sheet")
				return render_template("index.html")
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  16":
						rand_tb = cell.row
						baza1I = cell.column
						lun = len(purchases[cell.column])
			try:
				baza1I = [b.value for b in purchases[baza1I][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  10' in Purchases sheet")
				return render_template("index.html")				
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  11":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tva1J = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  11' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  22":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaW0 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  22' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  23":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaV9 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  23' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  24":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				bazaE2 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  24' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  25":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tvaE2 = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  25' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  26":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				baza1O = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  26' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  27":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tva1O = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  27' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  30":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				baza1M = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  30' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "                  31":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				tva1M = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '                  31' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == " Doc..Date  ":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				datadocument = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for ' Doc..Date  ' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Document No.    ":
						rand_tb = cell.row
						docnoc = cell.column
						lun = len(purchases[cell.column])
			try:
				docNoPurch = [b.value for b in purchases[docnoc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Document No.    ' in Purchases sheet")
				return render_template("index.html")			




			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "  Total doc.incl.VAT":
						rand_tb = cell.row
						tdoca = cell.column
						lun = len(sales[cell.column])
			try:
				listBazaL = [b.value for b in purchases[tdoca][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '  Total doc.incl.VAT'")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value=="Declarat anterior":
						rand_tb = cell.row
						declarateanteriorp = cell.column
						lun = len(purchases[cell.column])
			try:
				listadeclantp = [b.value for b in purchases[declarateanteriorp][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Declarat anterior' in Purchases sheet")
				return render_template("index.html")		
			# except:
			# 	listadeclant=[]
			listadeclantp_1=[]
			# #print(listadeclantp,"---------")
			for c in range(0, len(listadeclantp)):
				if listadeclantp[c] == None:
					listadeclantp_1.append("No")
				else:
					listadeclantp_1.append(listadeclantp[c])
			# #print("-----",listadeclantp_1,"------")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value=="Declarat anterior":
						rand_tb = cell.row
						declarateanterior = cell.column
						lun = len(sales[cell.column])
			try:
				listadeclant = [b.value for b in sales[declarateanterior][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Declarat anterior' in Sales sheet")
				return render_template("index.html")
			# except:
			# 	listadeclant=[]
			listadeclant_1=[]
			# #print(len(listadeclant))
			for c in range(0, len(listadeclant)):
				if listadeclant[c] == None:
					listadeclant_1.append("No")
				else:
					listadeclant_1.append(listadeclant[c])			

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "Business PartnerName":
						rand_tb = cell.row
						clientCell = cell.column
						lun = len(sales[cell.column])
			try:
				listaClient = [b.value for b in sales[clientCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Denumirea clientului Client name' in Sales sheet")
				return render_template("index.html")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "VAT Registration No.":
						rand_tb = cell.row
						coloanaClientID = cell.column
						lun = len(sales[cell.column])
			try:
				listaCUISales = [b.value for b in sales[coloanaClientID][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Tax Number 1    ' in Sales sheet")
				return render_template("index.html")

			listaCUISales1=[]
			# listadeclant_1=[]
			for val in listaCUISales:
				if val != None:
					# listadeclant_1.append("")
					listaCUISales1.append(val)
				else:
					listaCUISales1.append("US111")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "Document No.    ":
						rand_tb = cell.row
						docNumber = cell.column
						lun = len(sales[cell.column])
			try:
				docNoSales = [b.value for b in sales[docNumber][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'DocumentNo' in Sales sheet")
				return render_template("index.html")





			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "  Total doc.incl.VAT":
						rand_tb = cell.row
						totdoc = cell.column
						lun = len(sales[cell.column])
			totdocuments = [b.value for b in sales[totdoc][rand_tb:lun+1]]


			serieCuiSales=[]
			codTaraCuiSales=[]
			# print(listaCUISales1)
			for i in range(0,len(listaCUISales1)):
				if(str(listaCUISales1[i])[:1].isalpha()):
					r = re.compile("([a-zA-Z]+)([0-9]+)")
					m = r.match(str(listaCUISales1[i]))
					try:
						serieCuiSales.append(m.group(2))
						codTaraCuiSales.append(m.group(1))

					except:
						codTaraCuiSales.append(None)
						serieCuiSales.append(listaCUISales1[i])	

				else:
					codTaraCuiSales.append(None)
					serieCuiSales.append(listaCUISales1[i])
				# tara,oras=i.split(',',1)
				# serieCuiSales.append(oras)
				# codTaraCuiSales.append(tara)
			# #print(codTaraCuiSales)
			# TIP Furnizor!!!!!
			# #print(len(codTaraCuiSales))
			# print(serieCuiSales,codTaraCuiSales)
			coteTVAsales=[]
			taxcodeach=[]
			for i in range(0,len(listBazaL)):
				if(int(bazaV1[i])>0 or int(bazaV1[i]<0)):
					taxcodeach.append("V1")
				else:
					if(int(bazaW7[i])>0 or int(bazaW7[i]<0)):
						taxcodeach.append("W7")
					else:
						if(int(bazaW8[i])>0 or int(bazaW8[i]<0)):
							taxcodeach.append("W8")
						else:
							if(int(baza1L[i])>0 or int(baza1L[i]<0)):
								taxcodeach.append("1L")
							else:
								if(int(baza1J[i])>0 or int(baza1J[i]<0)):
									taxcodeach.append("1J")
								else:
									if(int(bazaW0[i])>0 or int(bazaW0[i]<0)):
										taxcodeach.append("W0")
									else:
										if(int(bazaV9[i])>0 or int(bazaV9[i]<0)):
											taxcodeach.append("V9")
										else:
											if(int(bazaE2[i])>0 or int(bazaE2[i]<0)):
												taxcodeach.append("E2")
											else:
												if(int(baza1O[i])>0 or int(baza1O[i]<0)):
													taxcodeach.append("1O")
												else:
													if(int(baza1M[i])>0 or int(baza1M[i]<0)):
														taxcodeach.append("1M")
													else:
														if(int(bazaN1[i])>0 or int(bazaN1[i]<0)):
															taxcodeach.append("N1")
														else:
															if(int(baza1I[i])>0 or int(baza1I[i]<0)):
																taxcodeach.append("1I")
															else:
																print(docNoPurch[i])
														
			taxcodes=[]
			# print(len(bazaA12),len(docNoSales))
			for j in range(0,len(docNoSales)):
				if(int(bazaA12[j])>0 or int(bazaA12[i])<0):
					taxcodes.append("A1")
				else:
					if(int(bazaA8[j])>0 or int(bazaA8[i])<0):
						taxcodes.append("A8")
					else:
						if(int(valY8[j])>0 or int(valY8[i])<0):
							taxcodes.append("Y8")
						else:
							if(int(valE2[j])>0):
								taxcodes.append("E1")
							else:
								taxcodes.append("E2")
			for i in range(0, len(docNoSales)):
				if (taxcodes[i]=="A1"):
					coteTVAsales.append(19)
				else:
					if (taxcodes[i]=="A8"):
						coteTVAsales.append(9)
					else:
						coteTVAsales.append(0)

			codTranzactieSales=[]
			for i in range(0, len(codTaraCuiSales)):
				# print(str(serieCuiSales[i])[1:2])
				if str(serieCuiSales[i])[1:2].isalpha():
					codTranzactieSales.append(1)
				else:
					if codTaraCuiSales[i] == "RO":
						# #print("RO")
						codTranzactieSales.append(1)
					else:	
						if codTaraCuiSales[i] in nomenclatorTari:
							# #print("UE")
							codTranzactieSales.append(3)
						else:
							# #print("nonUE")
							codTranzactieSales.append(4)
			#Cote TVAA




			#TIP TRANZACTIE
			storno=[]
			tipTranzSale = []
			# #print(docNoSales)
			# #print(len(docNoSales),len(codTranzactieSales))
			# print(len(docNoSales),len(codTranzactieSales))
			# print(codTranzactieSales)
			for i in range(0, len(docNoSales)):
				# print(taxcodes[i])
				# print(docNoSales[i],print(codTranzactieSales[i]))
				if(listadeclant_1[i]=="Yes"):
					tipTranzSale.append("Declarat anterior")
				else:
					if int(codTranzactieSales[i]) == 1:
						# #print(docNoSales[i]," ",listaCUISales1[i], "", taxBaseL19_1[i], " ", taxBaseL9_1[i], " ", taxBaseL5_1[i])
						# if (int(taxBaseL19_1[i])>0 and int(vatL19_1[i])>0) or (int(taxBaseL9_1[i])>0 and int(vatL9_1[i])> 0) or (int(taxBaseL5_1[i])>0 and int(vatL5_1[i])>0):
						if (taxcodes[i]=='A1'):
							# print("Yes")
							tipTranzSale.append('L')
							storno.append("")
						else:
							# None
							if (taxcodes[i]=='A8'):
								# print("Yes")
								tipTranzSale.append('L')
							else:
								if taxcodes[i]=='A5' or taxcodes[i]=='A2' or taxcodes[i]=='A4':
									# print("Yes")
									tipTranzSale.append("V")
									storno.append("")
								else:
									tipTranzSale.append("Not applicable for D394")
					else:
						if int(codTranzactieSales[i]) == 2:
							if (taxcodes[i]=='A1'):
								tipTranzSale.append('L')
								storno.append("")
							else:
								tipTranzSale.append("Not applicable for D394")
						else:
							if int(codTranzactieSales[i]) == 3:
								if (taxcodes[i]=='E1' or taxcodes[i]=='X1' or taxcodes[i]=='Y3'):
									tipTranzSale.append('Not applicable for D394')
									storno.append("")
								else:
									if (taxcodes[i]=='A1' or taxcodes[i]=='A8'):
										tipTranzSale.append('L')
										storno.append("")
									else:
										tipTranzSale.append('Not applicable for D394')
							else:
								if int(codTranzactieSales[i]) == 4:
									if (taxcodes[i]=='A1' or taxcodes[i]=='A8'):
										tipTranzSale.append('L')
										storno.append("")
									else:
										if (taxcodes[i]=='E1' or taxcodes[i]=='X1' or taxcodes[i]=='Y3'):
											tipTranzSale.append('Not applicable for d394')
										else:
											tipTranzSale.append('L')
											storno.append("")


			# #print(docNoSales)
			#Scriere in excel

			salesExcel.cell(row=9, column=1).value = "Cod tara"
			salesExcel.cell(row=9, column=2).value = "Serie cui"
			salesExcel.cell(row=9, column=3).value = "Numar document"
			salesExcel.cell(row=9, column=4).value = "CUI"
			salesExcel.cell(row=9, column=5).value = "Clasa tranzactie"
			salesExcel.cell(row=9, column=6).value = "Tip tranzactie"
			salesExcel.cell(row=9, column=7).value = "Cota TVA"
			salesExcel.cell(row=9, column=8).value = "Total document"
			salesExcel.cell(row=9, column=9).value = "Tip jurnal"
			salesExcel.cell(row=9, column=10).value = "Nume partener"
			salesExcel.cell(row=9, column=11).value = "Check"
			salesExcel.cell(row=9, column=12).value = "Cod si denumire NC produs(TIP V)"

			# dv = DataValidation(
				# type='list', formula1='"Yes,No"', allow_blank=True)


			listahelp=["1002--Secara","1003--Orz","1005--Porumb","1201--Boabe de soia"," 1205--Seminte de rapita sau de rapita salbatica","120600--Seminte de floarea soarelui","121291--Sfecla de zahar","1001-Grau si meslin","1004--Ovaz","10086000--Triticale","22-deseuri feroase si neferoase","23-masa lemnoasa","32-terenuri","33-constructii","34-alte bunuri","35-servicii","24-certificate de emisii de gaze cu efect de sera","25-energie electrica","26-certificate verzi","27-constructii/terenuri","28-aur de investitii","29-telefoane mobile","30-microprocesoare","31-console de jocuri tablete PC si laptopuri"]
			sheethelp=temp.create_sheet("Validation")
			sheethelp.sheet_state = 'hidden'

			# dv = DataValidation(
			# 	type="list", formula1="", allow_blank=True)
			# salesExcel.add_data_validation(dv)

			# dv.add(salesExcel["L2"])



			# print(len(tipTranzSale),len(codTranzactieSales))
			for i in range(0, len(codTaraCuiSales)):
				# print(serieCuiSales[i],docNoSales[i],listaCUISales1[i],codTranzactieSales[i],tipTranzSale[i])
				salesExcel.cell(row=10 + i, column=1).value = codTaraCuiSales[i]
				salesExcel.cell(row=10 + i, column=2).value = serieCuiSales[i]
				salesExcel.cell(row=10 + i, column=3).value = docNoSales[i]
				salesExcel.cell(row=10 + i, column=4).value = listaCUISales1[i]
				salesExcel.cell(row=10 + i, column=5).value = codTranzactieSales[i]
				# if(listadeclant_1[i]!=""):
				salesExcel.cell(row=10 + i, column=6).value = tipTranzSale[i]
				# else:
					# salesExcel.cell(row=10 + i, column=6).value = listadeclant_1[i]

				if(tipTranzSale[i]=='V'):
					salesExcel.cell(row=10+i,column=12).value="V"
				else:
					salesExcel.cell(row=10+i,column=12).value="N/A, valid only for V trans."

				salesExcel.cell(row=10 + i, column=8).value = totdocuments[i]
				salesExcel.cell(row=10 + i, column=9).value = "Jurnal vanzari"
				salesExcel.cell(row=10 + i, column=10).value = listaClient[i]

			for i in range(0, len(coteTVAsales)):
				salesExcel.cell(row=10 + i, column=7).value = coteTVAsales[i]
				salesExcel.cell(row=10+i,column=18).value="=B{0}&E{0}&F{0}&G{0}".format(i+10)

			#FORMATARE------------------------------------------------------------------
			red_color = 'ffc7ce'
			green_color='99ff99'
			red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
			green_fill = styles.PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
			row=salesExcel.max_row	
			salesExcel.conditional_formatting.add('K10:K'+str(row-1), formatting.rule.CellIsRule(operator='notEqual', formula=['"OK"'], fill=red_fill))
			for row in salesExcel['A9:L9']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in salesExcel['A9:L9']:
				for cell in row:
					cell.font = cap_tabel

			# for row in salesExcel['A9:K9']:
			# 	for cell in row:
			# 		cell.border = border_thin

			
			salesExcel.freeze_panes = 'A10'

			salesExcel.column_dimensions['B'].width = 20
			salesExcel.column_dimensions['C'].width = 20
			salesExcel.column_dimensions['D'].width = 20
			salesExcel.column_dimensions['E'].width = 16
			salesExcel.column_dimensions['F'].width = 16
			salesExcel.column_dimensions['F'].width = 20		
			salesExcel.column_dimensions['H'].width = 14
			salesExcel.column_dimensions['J'].width = 35
			salesExcel.column_dimensions['L'].width = 35		
			purchases = temp['Purchases']

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "Business PartnerName":
						rand_tb = cell.row
						codPartener = cell.column
						lun = len(purchases[cell.column])
			try:
				supplierName = [b.value for b in purchases[codPartener][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Furnizor Supplier' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "  Total doc.incl.VAT":
						rand_tb = cell.row
						totaldocp = cell.column
						lun = len(purchases[cell.column])
			try:
				totdocumentp = [b.value for b in purchases[totaldocp][rand_tb:lun]]
			except:
				flash("Please insert the correct header for '  Total doc.incl.VAT' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "VAT Registration No.":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				suppIDPurch = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'VAT Registration No.' in Purchases sheet")
				return render_template("index.html")
				


			serieCuiPurch = []
			codTaraCuiPurch = []
			# #print(suppIDPurch)
			for i in suppIDPurch:
				if(i==None):
					serieCuiPurch.append(None)
					codTaraCuiPurch.append(None)
				else:
				# #print(i)
					if(str(i)[:1].isalpha()):	
						r = re.compile("([a-zA-Z]+)([0-9]+)")
						m = r.match(str(i))
						try:
							serieCuiPurch.append(m.group(2))
						except:
							serieCuiPurch.append(" ")
						try:
							codTaraCuiPurch.append(m.group(1))
						except:
							codTaraCuiPurch.append(" ")
					else:
						codTaraCuiPurch.append(None)
						serieCuiPurch.append(i)
			# #print(codTaraCuiPurch,serieCuiPurch)
			tipTranzactiePurchases=[]
			#print("Aici vat -------",vatAch19_1,"-----Achizitii")
			#Tip furnizor
			for i in range(0, len(codTaraCuiPurch)):
				if(codTaraCuiPurch[i]==None):
					tipTranzactiePurchases.append(4)
				else:
					if codTaraCuiPurch[i] == "RO":
						# #print("RO")
						tipTranzactiePurchases.append(1)
					else:
						if serieCuiPurch[i] in suppIDPurch and int(nonCharTaxBase_1[i])>0:
							tipTranzactiePurchases.append(2)
						else:
							if codTaraCuiPurch[i] in nomenclatorTari:
								# #print("UE")
								tipTranzactiePurchases.append(3)
							else:
								# #print("nonUE")
								tipTranzactiePurchases.append(4)

			#Cote TVA
			coteTVApurchases=[]
			print(len(taxcodeach),len(docNoPurch))
			for i in range(0, len(docNoPurch)):
				if (taxcodeach[i]=="ZI" or taxcodeach[i]=="C3" or taxcodeach[i]=="ZD" or taxcodeach[i]=="E1" or taxcodeach[i]=='V1' or taxcodeach[i]=="5H" or taxcodeach[i]=="X1" or taxcodeach[i]=="1L" or taxcodeach[i]=="1M" or taxcodeach[i]=="3S" or taxcodeach[i]=="5B" or taxcodeach[i]=="5H" or taxcodeach[i]=="1I"):
					coteTVApurchases.append('19')
				else:
					if (taxcodeach[i]=="I9" or taxcodeach[i]=="W8" or taxcodeach[i]=="W6" or taxcodeach[i]=="J2" or taxcodeach[i]=="V3" or taxcodeach[i]=="9S"):
						coteTVApurchases.append('9')
					else:
						if (taxcodeach[i]=="I7" or taxcodeach[i]=="5D" or taxcodeach[i]=="W8" or taxcodeach[i]=="6I" or taxcodeach[i]=="5S"):
							coteTVApurchases.append('5')
						else:
							coteTVApurchases.append('0')

			#mapare tip tranzactie
			# #print(len(docNoPurch1),len(tipTranzactiePurchases),len(vatApplies))
			# #print(vatApplies)
			tipTranzPurch=[]

			# for i in range(0,len(suppIDPurch)):
				# #print(suppIDPurch[i],tipTranzactiePurchases[i])
			print(len(docNoPurch),len(tipTranzactiePurchases))
			#print(len(docNoPurch1),len(listadeclantp_1),"--------------len de lista")
			for i in range(0, len(docNoPurch)):
				if(listadeclantp_1[i]=="Yes"):
					tipTranzPurch.append("Declarat anterior")
				else:
					# #print(docNoPurch1[i])
					if int(tipTranzactiePurchases[i]) == 1:
						if (taxcodeach[i]=="V1" or taxcodeach[i]=="W8" or taxcodeach[i]=="V3" or taxcodeach[i]=="5H" or taxcodeach[i]=="5B"):
							tipTranzPurch.append('A')
						else:
							if taxcodeach[i]=="ZI" or taxcodeach[i]=="5D" or taxcodeach[i]=="1J" or taxcodeach[i]=="ZD" or taxcodeach[i]=="I7" or taxcodeach[i]=="W8" or taxcodeach[i]=="I9" or taxcodeach[i]=="1I" or taxcodeach[i]=="W6" or taxcodeach[i]=="6I":
								tipTranzPurch.append("AI")
								#print(docNoPurch1[i],";;;;;es 3")
							else:
								if (taxcodeach[i]=="1M"):
									tipTranzPurch.append("C")
									#print("Yes 5")
								else:
									if (taxcodeach[i]=="AS"):
										tipTranzPurch.append("AS")
										#print("Yes 7")
									else:
										tipTranzPurch.append("Not applicable for D394")
					else:
						if int(tipTranzactiePurchases[i]) == 2:
							if taxcodeach[i]=="7N" or taxcodeach[i]=="8N" or taxcodeach[i]=="A3" :
								tipTranzPurch.append("N")
								#print(docNoPurch1[i],";;;;es 9")
						else:
							if int(tipTranzactiePurchases[i]) == 3:
								if (taxcodeach[i]=="V1" or taxcodeach[i]=="W8" or taxcodeach[i]=="V3"):
											tipTranzPurch.append('A')
											#print(docNoPurch1[i],";;;;;es 12")
								else:
										#print(docNoPurch1[i],";;;;;es 13")
									if (taxcodeach[i]=="1M"):
										tipTranzPurch.append("C")
										#print("Yes 14")
									else:
										tipTranzPurch.append("Not applicable for D394")
							else:
									if int(tipTranzactiePurchases[i]) == 4:
										if (taxcodeach[i]=="V1" or taxcodeach[i]=="W8" or taxcodeach[i]=="V3"):
											tipTranzPurch.append('A')
											#print(docNoPurch1[i],";;;;;es 12")
										else:
											if (taxcodeach[i]=="1M"):
												tipTranzPurch.append("C")
											else:
												tipTranzPurch.append("Not applicable for D394")
										#print("Yes 16")
				# #print(docNoPurch1[i],tipTranzPurch[i],docNoPurch[i+1])
			ma=salesExcel.max_row+1
			for i in range(0, len(codTaraCuiPurch)):
				salesExcel.cell(row=ma + i, column=1).value = codTaraCuiPurch[i]

			for i in range(0, len(serieCuiPurch)):
				salesExcel.cell(row=ma + i, column=2).value = serieCuiPurch[i]

			for i in range(0, len(docNoPurch)):
				salesExcel.cell(row=ma+ i, column=3).value = docNoPurch[i]

			for i in range(0, len(suppIDPurch)):
				salesExcel.cell(row=ma + i, column=4).value = suppIDPurch[i]

			for i in range(0, len(tipTranzactiePurchases)):
				salesExcel.cell(row=ma+ i, column=5).value = tipTranzactiePurchases[i]

			for i in range(0, len(tipTranzPurch)):
				# if(listadeclantp_1!=""):
				salesExcel.cell(row=ma+ i, column=6).value = tipTranzPurch[i]
				# else:
					# salesExcel.cell(row=ma+ i, column=6).value = "Declarate anterior"
				if(tipTranzPurch[i]=="V"):
					salesExcel.cell(row=ma+i,column=12).value="Add type of tranzactie"
				else:
					salesExcel.cell(row=ma+i,column=12).value="N/A"

			for i in range(0, len(coteTVApurchases)):
				#print(coteTVApurchases[i])
				salesExcel.cell(row=ma+ i, column=7).value = coteTVApurchases[i]
				salesExcel.cell(row=ma+ i, column=8).value = totdocumentp[i]
				salesExcel.cell(row=ma+ i, column=9).value = "Jurnal cumparari"
				salesExcel.cell(row=ma+ i, column=10).value = denumirea[i]

			codTaraCUItotal=codTaraCuiPurch+codTaraCuiSales
			for i in range(0, len(codTaraCUItotal)):
				salesExcel.cell(row=10 + i, column=11).value = '=IFERROR(IF(VLOOKUP(B{0}&E{0}&F{0}&G{0},Tranzactii!K:K,1,0)=B{0}&E{0}&F{0}&G{0},"OK","Mapped missing in Tranzactii sheet"),"Mapped missing inTranzactiisheet")'.format(10+i)
			salesExcel.auto_filter.ref = "A9:L9"
			for row in salesExcel['H10:F1000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'		
			tranzactii=temp.create_sheet("Tranzactii")
			tranzactii.freeze_panes = 'A6'
			tranzactii.sheet_view.showGridLines = False

													
			setSalesCUI=set(listaCUISales1)
			idSalesUnique=list(setSalesCUI)

			setPurchCUI=set(suppIDPurch)
			idPurchUnique=list(setPurchCUI)

			listaCUIUnique=idSalesUnique+idPurchUnique

			setlistaClient=set(listaClient)
			listaClientUnique=list(setlistaClient)

			setSupplierName=set(supplierName)
			supplierNameUnique=list(setSupplierName)

			# for k in range(0,len(setlistaClient)):
			#    count
			#    for j in range(0,len(listaClient)):


			# print(len(supplierName),len(tipTranzPurch))
			listanouaappendpurch=[]
			# for i in range(0,len(supplierName)):
			# for p in range(0,len(serieCuiPurch)):
			# 	# print(serieCuiPurch[p],tipTranzPurch[p],coteTVApurchases[p],tipTranzactiePurchases[p])
			# print()
			# print(len(serieCuiPurch),len(tipTranzPurch),len(coteTVApurchases),len(tipTranzactiePurchases))
			for k in range(0,len(serieCuiPurch)):
				try:
					print(serieCuiPurch[k],tipTranzPurch[k])
				except:
					print(serieCuiPurch[k])
				listanouaappendpurch.append(str(serieCuiPurch[k])+";"+str(tipTranzPurch[k])+";"+str(coteTVApurchases[k])+";"+str(tipTranzactiePurchases[k])+";"+str(listadeclantp_1[k]))

			listanouasetpurch=list(set(listanouaappendpurch))


			listanouaappendsales=[]

			for k in range(0,len(serieCuiSales)):
				listanouaappendsales.append(str(serieCuiSales[k])+";"+str(tipTranzSale[k])+";"+str(coteTVAsales[k])+";"+str(codTranzactieSales[k])+";"+listadeclant_1[k])

			listanouasetsales=list(set(listanouaappendsales))

			countsales=[]
			for p in range(0,len(listanouasetsales)):
				count=0
				for k in range(0,len(listanouaappendsales)):
					if(listanouaappendsales[k]==listanouasetsales[p]):
						count=count+1
				countsales.append(count)
			countpurch=[]
			for p in range(0,len(listanouasetpurch)):
				count=0
				for k in range(0,len(listanouaappendpurch)):
					if(listanouaappendpurch[k]==listanouasetpurch[p]):
						count=count+1
				countpurch.append(count)


			#print(listanouasetsales)

			tranzactii.cell(row=5,column=1).value="Cui partener"
			tranzactii.cell(row=5,column=2).value="Nume partener"
			tranzactii.cell(row=5,column=3).value="Tip partener"
			tranzactii.cell(row=5,column=4).value="Tip tranzactie"
			tranzactii.cell(row=5,column=5).value="Cota TVA"
			tranzactii.cell(row=5,column=6).value="Baza TVA"
			tranzactii.cell(row=5,column=7).value="TVA"
			tranzactii.cell(row=5,column=8).value="Nr Facturi"
			tranzactii.cell(row=5,column=9).value="Neexigibile - nu se vor raporta"
			tranzactii.cell(row=5,column=10).value="Cod si denumire NC produs(TIP V)"
			tranzactii.cell(row=2,column=1).value="Tranzactii"
			tranzactii.cell(row=2,column=1).font=cap_tabeltitlu
			counts=0
			for i in range(0, len(listanouasetsales)):
				x=listanouasetsales[i].split(";")
				#print(x[3],x[4])
				if(int(x[3])<3 or int(x[2])>0):
					counts=counts+1
					y=tranzactii.max_row
					tranzactii.cell(row=y+1,column=1).value=x[0]
					tranzactii.cell(row=y+1,column=2).value="=VLOOKUP(A"+str(y+1)+",'Mapping tranzactii'!B:J,9,0)"
					tranzactii.cell(row=y+1,column=3).value=x[3]
					tranzactii.cell(row=y+1,column=4).value=x[1]
					tranzactii.cell(row=y+1,column=5).value=x[2]
					tranzactii.cell(row=y+1,column=9).value=x[4]
					# tranzactii.cell(row=y+1,column=10).value="=xlookup(K"+str(y+1)+",'Mapping tranzactii'!R:R,'Mapping tranzactii'!L:L)"
			countp=0
			for i in range(0, len(listanouasetpurch)):
				x=listanouasetpurch[i].split(";")
				if(int(x[3])<3 or int(x[2])>0):
					countp=countp+1
					y=tranzactii.max_row
					tranzactii.cell(row=y+1,column=1).value=x[0]
					tranzactii.cell(row=y+1,column=2).value="=VLOOKUP(A"+str(y+1)+",'Mapping tranzactii'!B:J,9,0)"   
					try:
						tranzactii.cell(row=y+1,column=3).value=x[3]
					except:
						tranzactii.cell(row=y+1,column=3).value=""
					tranzactii.cell(row=y+1,column=4).value=x[1]
					tranzactii.cell(row=y+1,column=5).value=x[2]
					tranzactii.cell(row=y+1,column=9).value=x[4]
					# tranzactii.cell(row=y+1,column=10).value="=xlookup(K"+str(y+1)+",'Mapping tranzactii'!R:R,'Mapping tranzactii'!L:L)"

			countmare=countp+counts
			for i in range(0, countmare):
				tranzactii.cell(row=i+6,column=6).value="=SUMIFS('Mapping tranzactii'!H:H,'Mapping tranzactii'!B:B,A{0},'Mapping tranzactii'!E:E,C{0},'Mapping tranzactii'!F:F,D{0},'Mapping tranzactii'!G:G,E{0})/((100+E{0})/100)".format(6+i)
				tranzactii.cell(row=i+6,column=7).value="=F{0}/100*E{0}".format(6+i)
				tranzactii.cell(row=i+6,column=8).value="=COUNTIFS('Mapping tranzactii'!B:B,A{0},'Mapping tranzactii'!E:E,C{0},'Mapping tranzactii'!F:F,D{0},'Mapping tranzactii'!G:G,E{0})".format(6+i)
				tranzactii.cell(row=i+6,column=11).value="=A{0}&C{0}&D{0}&E{0}".format(6+i)


			#---------FORMAT-----------------
			for row in tranzactii['A5:J5']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in tranzactii['A5:J5']:
				for cell in row:
					cell.font = cap_tabel
		
			tranzactii.column_dimensions['K'].hidden = True
			tranzactii.column_dimensions['A'].width = 20
			tranzactii.column_dimensions['I'].width = 27		
			tranzactii.column_dimensions['B'].width = 35
			tranzactii.column_dimensions['C'].width = 13
			tranzactii.column_dimensions['D'].width = 13
			tranzactii.column_dimensions['H'].width = 13
			tranzactii.auto_filter.ref = "A5:H5"
			for row in tranzactii['F6:F'+str(tranzactii.max_row)]:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'
			for row in tranzactii['G6:G'+str(tranzactii.max_row)]:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'

			saf=temp.create_sheet("Facturi storno si anulate")
			saf.sheet_view.showGridLines = False
			saf.freeze_panes = 'A4'		
			saf.cell(row=1,column=1).value="Facturi storno/anulate"
			saf.cell(row=1,column=1).font=cap_tabelbold
			saf.cell(row=3,column=1).value="Tip"
			saf.cell(row=3,column=2).value="Serie"
			saf.cell(row=3,column=3).value="Numar"
			for row in saf['A3:C3']:
				for cell in row:
					cell.font = cap_tabel
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			ind=saf.max_row
			for k in range(0,len(storno)):
				if(storno[k]=="Yes"):
					saf.cell(row=ind+1,column=1).value="Stornata"
					saf.cell(row=ind+1,column=2).value=""
					saf.cell(row=ind+1,column=3).value=docNoSales[k]
					ind=ind+1
			xx=saf.max_row		
			saf.cell(row=xx+1,column=1).value="Anulata"
			saf.cell(row=xx+1,column=2).value="Please input the cancelled invoice number"	
			bonuri=temp.create_sheet("Bonuri fiscale")
			bonuri.cell(row=4,column=1).value="Luna"
			bonuri.cell(row=4,column=2).value="Nr. bon fiscal"
			bonuri.cell(row=4,column=3).value="Baza 5%"
			bonuri.cell(row=4,column=4).value="TVA 5%"
			bonuri.cell(row=4,column=5).value="Baza 9%"
			bonuri.cell(row=4,column=6).value="TVA 9%"
			bonuri.cell(row=4,column=7).value="Baza 19%"
			bonuri.cell(row=4,column=8).value="TVA 19%"
			bonuri.cell(row=4,column=9).value="Baza 20%"
			bonuri.cell(row=4,column=10).value="TVA 20%"

			for row in bonuri['A4:K4']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in bonuri['A4:K4']:
				for cell in row:
					cell.font = cap_tabel
			bonuri.sheet_view.showGridLines = False

			facturi=temp.create_sheet("Sectiunea 2.1&2.2")
			facturi.sheet_view.showGridLines = False
			facturi.column_dimensions['A'].width = 12
			facturi.column_dimensions['B'].width = 18
			facturi.column_dimensions['C'].width = 15

			facturi.cell(row=1,column=1).value="Serie Emise"
			facturi.cell(row=1,column=2).value="Inceput Emise"
			facturi.cell(row=1,column=3).value="Final Emise"
			facturi.cell(row=1,column=4).value="Tip Emise"		

			docNoSales2=[]
			seriefacturi=[]
			print(codTranzactieSales)
			# print(docNoSales)
			for i in range(0,len(docNoSales)):
				# print(docNoSales[i])

				# docNoSales[i].replaceAll("[^a-zA-Z0-9]", ")
				if(int(codTranzactieSales[i])<3):
					# try:
					numere=re.sub("[^0-9]", "",str(docNoSales[i]))
					# except:
						# print(docNoSales[i])
					result = ''.join([i for i in str(docNoSales[i]) if not i.isdigit()])
					docNoSales2.append(numere)
					seriefacturi.append(result)
			# print(seriefacturi)
			# print(docNoSales2)
			initial=0
			final=0
			docNoSales2.sort()
			docNo=[]
			for k in range(0,len(docNoSales2)):
				docNo.append(str(docNoSales2[k]))
			docNo.sort()
			listaunica=list(set(docNoSales2))
			listaunica.sort()
			print(listaunica)

			for i in range(0,len(listaunica)):
				listafacturi=[]
				print(listaunica[i])
				for j in range(0,len(docNoSales)):
					if(listaunica[i]==docNoSales[j]):

						listafacturi.append(int(docNoSales[j]))
			listafacturi=list(set(listafacturi))
			print(listafacturi)
			listafacturi.sort()
			start=[]
			start.append(listaunica[0])
			stop=[]
			try:
				if(int(listaunica[1])-int(listaunica[0])>1):
					stop.append(listaunica[0])
				for k in range(1,len(listaunica)):

					if(int(listaunica[k])-int(listaunica[k-1])==1):
						print("ok")
					else:
						stop.append(listaunica[k-1])
						start.append(listaunica[k])
			except:
				stop.append(listaunica[0])
			if(len(stop)==len(start)):
				print("ok")
			else:
				stop.append(listaunica[len(listaunica)-1])
			print(start,stop)

			# #print(docNoSales)
			for k in range(0,len(start)):
				facturi.cell(row=2+k,column=2).value=start[k]
				facturi.cell(row=2+k,column=3).value=stop[k]
				facturi.cell(row=2+k,column=4).value=2

			# for p in range(0,len(docNoSales2)-1):
			# 	#print(docNo[p])
			# 	if(p==0):
			# 		initial=initial+1
			# 		# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 		facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 		if(int(docNo[p])-int(docNo[p+1])< -1):
			# 			final=final+1
			# 			facturi.cell(row=1+final,column=3).value=docNo[p]
			# 	else:
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])==1 and int(docNo[p])-int(docNo[p+1])==-1):
			# 				print("bailando")
						
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])==1 and int(docNo[p][3:])-int(docNo[p+1][3:])==-1):
			# 					print("bailando")
			# 					None
			# 			except:
			# 				print(None)
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])>1 and int(docNo[p])-int(docNo[p+1])==-1):
			# 				initial=initial+1
			# 				# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 				facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])>1 and int(docNo[p][3:])-int(docNo[p+1][3:])==-1):
			# 					initial=initial+1
			# 					# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 					facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 			except:
			# 				print(docNo[p])
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])==1 and int(docNo[p])-int(docNo[p+1])<-1):
			# 				final=final+1
			# 				facturi.cell(row=1+final,column=3).value=docNo[p]
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])==1 and int(docNo[p][3:])-int(docNo[p+1][3:])<-1):
			# 					final=final+1
			# 					facturi.cell(row=1+final,column=3).value=docNo[p]
			# 			except:
			# 				print("none")
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])>1 and int(docNo[p])-int(docNo[p+1])<-1):
			# 				initial=initial+1
			# 				# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 				facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 				final=final+1
			# 				facturi.cell(row=1+final,column=3).value=docNo[p]
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])>1 and int(docNo[p][3:])-int(docNo[p+1][3:])<-1):
			# 					initial=initial+1
			# 					# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 					facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 					final=final+1
			# 					facturi.cell(row=1+final,column=3).value=docNo[p]
			# 			except:
			# 				print("none")

			x=facturi.max_row
			facturi.auto_filter.ref = "A1:C1"
			# if(int(docNoSales2[len(docNoSales2)-1])-int(docNoSales2[len(docNoSales2)-2])>1):
			# 	facturi.cell(row=x+1,column=1).value=seriefacturi[0]
			# 	facturi.cell(row=x+1,column=2).value=docNoSales2[len(docNoSales2)-1]
			# 	facturi.cell(row=x+1,column=3).value=docNoSales2[len(docNoSales2)-1]
			# else:
			# 	facturi.cell(row=x+1,column=1).value=seriefacturi[0]
			# 	facturi.cell(row=x, column=3).value = docNoSales2[len(docNoSales2) - 1]


			yy=facturi.max_row+2
			facturi.cell(row=yy,column=1).value="Serie Alocate"
			facturi.cell(row=yy,column=2).value="Inceput Alocate"
			facturi.cell(row=yy,column=3).value="Final Alocate"
			facturi.cell(row=yy,column=4).value="Tip Alocate"		
			for kk in range(1,5):
				facturi.cell(row=yy,column=kk).font=cap_tabel
				facturi.cell(row=yy,column=kk).fill=cap_tabel_color_black	
			for pp in range(2,yy):
				facturi.cell(row=yy+pp-1,column=2).value=facturi.cell(row=pp,column=2).value
				facturi.cell(row=yy+pp-1,column=3).value=facturi.cell(row=pp,column=3).value
				facturi.cell(row=yy+pp-1,column=4).value=1						
			a23=temp.create_sheet("Sectiunea 2.3,2.4")
			dv = DataValidation(
				type='list', formula1='"Yes,No"', allow_blank=True,showDropDown=False)		
			dv.add(a23["A24"])			
			a23.sheet_view.showGridLines = False
			a23.column_dimensions['A'].width=18
			a23.column_dimensions['B'].width=13		
			a23.cell(row=1,column=1).value="Sectiunea 2.3"
			a23.cell(row=3,column=1).value="Denumire beneficiar"
			a23.cell(row=3,column=2).value="CUI beneficiar"
			a23.cell(row=3,column=1).fill=cap_tabel_color_black
			a23.cell(row=3,column=1).font=cap_tabel		
			a23.cell(row=36,column=1).border=border_bottom
			a23.cell(row=36,column=2).border=border_lowerleft
			a23.cell(row=35,column=2).border=border_right
			a23.cell(row=34,column=2).border=border_right
			a23.cell(row=33,column=2).border=border_right
			a23.cell(row=32,column=2).border=border_right						
			a23.cell(row=3,column=2).fill=cap_tabel_color_black
			a23.cell(row=3,column=2).font=cap_tabel


			a23.cell(row=6,column=1).value="Seria"
			a23.cell(row=6,column=2).value="De la"
			a23.cell(row=6,column=3).value="La"



			for row in a23['A6:C6']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in a23['A14:C14']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			a23.cell(row=9,column=1).value="Sectiunea 2.4"
			a23.cell(row=9,column=1).font=cap_tabelbold
			a23.cell(row=1,column=1).font=cap_tabelbold		
			a23.cell(row=28,column=1).font=cap_tabelbold		
			a23.cell(row=4,column=1).border=border_lowerright
			a23.cell(row=4,column=2).border=border_lowerright		
			a23.cell(row=12,column=1).border=border_lowerright
			a23.cell(row=12,column=2).border=border_lowerright		
			a23.cell(row=11,column=1).fill=cap_tabel_color_black
			a23.cell(row=11,column=1).font=cap_tabel		
			a23.cell(row=7,column=1).border=border_lowerright
			a23.cell(row=7,column=2).border=border_lowerright
			a23.cell(row=7,column=3).border=border_lowerright
			a23.cell(row=15,column=1).border=border_lowerright
			a23.cell(row=15,column=2).border=border_lowerright
			a23.cell(row=15,column=3).border=border_lowerright				
			a23.cell(row=11,column=2).fill=cap_tabel_color_black
			a23.cell(row=11,column=2).font=cap_tabel
			a23.cell(row=11,column=1).value="Denumire tert"

			a23.cell(row=11,column=2).value="CUI tert"

			a23.cell(row=14,column=1).value="Seria"
			a23.cell(row=14,column=2).value="De la"
			a23.cell(row=14,column=3).value="La"
			for row in a23['A31:B31']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
			try:
				if int(val4)== 1:
					a3=temp.create_sheet("Sectiunea 3")
					a3.cell(row=1,column=1).value="Sectiunea 3"
					a3.cell(row=1,column=1).font=cap_tabeltitlu
					a3.cell(row=3,column=1).value="In cazul in care soldul sumei negative inregistrate in decontul de TVA aferent perioadei de raportare este solicitat la rambursare , se vor selecta datele cu privire la natura operatiunilor din care provine acesta"
					a3.cell(row=7,column=1).value="Achizitii de bunuri si servicii legate direct de bunurile imobile din urmatoarele categorii"
					a3.cell(row=8,column=1).value="a) parcuri eoliene"
					a3.cell(row=9,column=1).value="b) constructii rezidentiale"



					a3.cell(row=10,column=1).value="c) cladiri de birouri"
					a3.cell(row=11,column=1).value="d) constructii industriale"
					a3.cell(row=12,column=1).value="e) altele"
					a3.cell(row=14,column=1).value="Achizitii de bunuri, cu exceptia celor legate direct de bunuri imobile:"
					a3.cell(row=15,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=16,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=17,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=18,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=19,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=21,column=1).value="Achizitii de servicii, cu exceptia celor legate direct de bunurile imobile:"
					a3.cell(row=22,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=23,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=24,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=25,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=26,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=28,column=1).value="Importuri de bunuri"
					a3.cell(row=30,column=1).value="Achizitii imobilizari necorporale"
					a3.cell(row=32,column=1).value="Livrari de bunuri imobile"
					a3.cell(row=34,column=1).value="Livrari de bunuri, cu exceptia bunurilor imobile:"
					a3.cell(row=35,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=36,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=37,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=38,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=39,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=41,column=1).value="Livrari de bunuri scutite de TVA"
					a3.cell(row=43,column=1).value="Livrari de bunuri/prestari de servicii pt care se aplica taxarea inversa"
					a3.cell(row=45,column=1).value="Prestari de servicii:"
					a3.cell(row=46,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=47,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=48,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=49,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=50,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=52,column=1).value="Prestari de servicii scutite de TVA"
					a3.cell(row=54,column=1).value="Livrari intracomunitare de bunuri"
					a3.cell(row=56,column=1).value="Prestari intracomunitare de servicii"
					a3.cell(row=58,column=1).value="Exporturi de bunuri"
					a3.cell(row=60,column=1).value="Livrari imobilizari necorporale"
					a3.cell(row=62,column=1).value="Persoana impozabila nu a efectuat livrari de bunuri/prestari de servicii in perioada de raportare"																					
			except:
				pass


																																																																						
			a5=temp.create_sheet("Sectiunea 5")
			a5.sheet_view.showGridLines = False
			a5.cell(row=2,column=1).value="Sectiunea 5"
			a5.cell(row=2,column=1).font=cap_tabeltitlu
			a5.cell(row=4,column=1).font=cap_tabelbold
			a5.cell(row=6,column=1).font=cap_tabelbold
			a5.column_dimensions['B'].width=13
			a5.merge_cells('A6:J7')
			a5.merge_cells('A18:J19')		
			
			for pop in range(1,11):
				a5.cell(row=5,column=pop).border=border_bottom
				a5.cell(row=7,column=pop).border=border_bottom
				a5.cell(row=17,column=pop).border=border_bottom
				a5.cell(row=19,column=pop).border=border_bottom
			a5.cell(row=6,column=10).border=border_upperright
			a5.cell(row=7,column=10).border=border_lowerright
			a5.cell(row=18,column=10).border=border_upperright
			a5.cell(row=19,column=10).border=border_lowerright									
			a5['A6'].alignment=Alignment(wrap_text=True)
			a5['A18'].alignment=Alignment(wrap_text=True)
			a5['A30'].alignment=Alignment(wrap_text=True)						
			a5.cell(row=4,column=1).value="Sectiune 5.2"
			a5.cell(row=16,column=1).font=cap_tabelbold		
			a5.cell(row=16,column=1).value="Sectiune 5.3"
			a5.cell(row=28,column=1).font=cap_tabelbold		
			a5.cell(row=10,column=1).value="cota 24%"
			a5.cell(row=11,column=1).value="cota 20%"
			a5.cell(row=12,column=1).value="cota 19%"
			a5.cell(row=13,column=1).value="cota 9%"
			a5.cell(row=14,column=1).value="cota 5%"
			a5.cell(row=10,column=2).value=0
			a5.cell(row=11,column=2).value=0
			a5.cell(row=12,column=2).value=0
			a5.cell(row=13,column=2).value=0
			a5.cell(row=14,column=2).value=0

			a5.cell(row=9,column=2).value="Valoare TVA"
			a5.cell(row=9,column=1).value="Cota"		

			a5.cell(row=6,column=1).value="5.2 TVA deductibila aferenta facturilor achitate in perioada de raportare indiferent de data in care acestea au fost primite de la persoane impozabile care aplica sistemul normal de TVA, defalcata pe fiecare cota de TVA"
			a5.cell(row=18,column=1).font=cap_tabelbold		

			a5.cell(row=22,column=1).value="cota 24%"
			a5.cell(row=23,column=1).value="cota 20%"
			a5.cell(row=24,column=1).value="cota 19%"
			a5.cell(row=25,column=1).value="cota 9%"
			a5.cell(row=26,column=1).value="cota 5%"

			a5.cell(row=22,column=2).value=0
			a5.cell(row=23,column=2).value=0
			a5.cell(row=24,column=2).value=0
			a5.cell(row=25,column=2).value=0
			a5.cell(row=26,column=2).value=0

			a5.cell(row=21,column=1).value="Cota"		
			a5.cell(row=21,column=2).value="Valoare TVA"
			for row in a5['A9:B9']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in a5['A21:B21']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
								
			a5.cell(row=18,column=1).value="5.3 TVA deductibila aferenta facturilor achitate in perioada de raportare indiferent de data in care acestea au fost primite de la persoane impozabile care aplica sistemul de TVA la incasare, defalcata pe fiecare cota de TVA"

			for kk in range(22,27):
				a5.cell(row=kk,column=1).border=border_lowerright
				a5.cell(row=kk,column=2).border=border_lowerright						
			for kk in range(10,15):
				a5.cell(row=kk,column=1).border=border_lowerright
				a5.cell(row=kk,column=2).border=border_lowerright		
			a6=temp.create_sheet("Sectiunea 6")
			a6.sheet_view.showGridLines = False
			a6.cell(row=2,column=1).value="Sectiunea 6"
			a6.cell(row=2,column=1).font=cap_tabeltitlu
			a6.cell(row=4,column=1).font=cap_tabelbold
			a6.cell(row=6,column=1).font=cap_tabelbold
			a6.cell(row=15,column=1).font=cap_tabelbold
			a6.cell(row=13,column=1).font=cap_tabelbold
			a6.column_dimensions['A'].width=17			
			a6.column_dimensions['B'].width=28
			a6.column_dimensions['C'].width=17
																
			a6.cell(row=4,column=1).value="Sectiunea 6.1"		
			a6.cell(row=6,column=1).value="6.1 Persoanele impozabile care aplica regimul special pt agentiile de turism, vor completa:"
			a6.merge_cells('A6:F7')
			for l in range(1,7):
				a6.cell(row=5,column=l).border=border_bottom
				a6.cell(row=7,column=l).border=border_bottom
			a6.cell(row=6,column=6).border=border_upperright
			a6.cell(row=7,column=6).border=border_lowerright						

			a6.merge_cells('A15:G16')

			for l in range(1,8):
				a6.cell(row=14,column=l).border=border_bottom
				a6.cell(row=16,column=l).border=border_bottom
			a6.cell(row=15,column=7).border=border_upperright
			a6.cell(row=16,column=7).border=border_lowerright




			a6.cell(row=9,column=1).value="Incasarile agentiei"
			a6.cell(row=9,column=2).value="Costurile agentiei de turism"
			a6.cell(row=9,column=3).value="Marja de profit"
			a6.cell(row=9,column=4).value="TVA"

			a6.cell(row=13,column=1).value="Sectiunea 6.2"

			a6.cell(row=15,column=1).value="6.2 Persoanele impozabile care aplica regimul special pt bunurile second-hand, opere de arta, obiecte de colectie si antichitati , vor completa:"

			for row in a6['A18:D18']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in a6['A9:D9']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')										


			a6.cell(row=18,column=1).value="Pret de vanzare"
			a6.cell(row=18,column=2).value="Pret de cumparare"
			a6.cell(row=18,column=3).value="Marja de profit"	
			a6.cell(row=18,column=4).value="TVA"
			for oo in range(1,5):
				a6.cell(row=10,column=oo).border=border_lowerright
				a6.cell(row=19,column=oo).border=border_lowerright		

			a6['A6'].alignment=Alignment(wrap_text=True)

			a6['A15'].alignment=Alignment(wrap_text=True)
			a7=temp.create_sheet(" Sectiunea 7 ")

			a7.cell(row=2,column=1).value="Sectiunea 7"
			a7.sheet_view.showGridLines = False
			a7.cell(row=2,column=1).font=cap_tabeltitlu
			a7.cell(row=5,column=1).value="7. In situatia in care ati desfasurat, in perioada de raportare, activitati dintre cele inscrise in lista veti selecta activitatea corespunzatoare si veti inscrie valoarea livrarilor/prestarilor, precum si TVA aferenta"
			a7['A5'].alignment=Alignment(wrap_text=True)
			a7.merge_cells('A5:I6')
			a7.cell(row=5,column=1).font=cap_tabelbold
			for ii in range(1,10):
				a7.cell(row=4,column=ii).border=border_bottom
				a7.cell(row=6,column=ii).border=border_bottom
			a7.cell(row=5,column=9).border=border_upperright
			a7.cell(row=6,column=9).border=border_lowerright		
			a7.cell(row=5,column=8).border=border_right
			a7.cell(row=6,column=8).border=border_right

			for jj in range(1,4):
				a7.cell(row=9,column=jj).border=border_bottom
				a7.cell(row=9,column=jj).border=border_right
			
			for pp in range(12,17):
				a7.cell(row=pp,column=2).border=border_lowerright
				a7.cell(row=pp,column=1).border=border_lowerright
			a7.cell(row=8,column=1).value="Activitate"
			a7.cell(row=8,column=2).value="Tip operatiune"
			a7.cell(row=8,column=3).value="Valoarea livrarilor/prestarilor"

			a7.cell(row=11,column=1).value="Cota"
			a7.cell(row=12,column=1).value="cota 24%"
			a7.cell(row=13,column=1).value="cota 20%"
			a7.cell(row=14,column=1).value="cota 19%"
			a7.cell(row=15,column=1).value="cota 9%"
			a7.cell(row=16,column=1).value="cota 5%"
			for row in a7['A8:C8']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font = cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')										
			for row in a7['A11:B11']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font = cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			a7.cell(row=11,column=2).value="Valoare TVA"

			for row in facturi['A1:D1']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in facturi['A1:D1']:
				for cell in row:
					cell.font = cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			facturi.column_dimensions['C'].width = 14
			facturi.column_dimensions['D'].width = 14
			a23.column_dimensions['A'].width = 18
			a23.column_dimensions['B'].width = 13
			a23.column_dimensions['C'].width = 10
			a6.column_dimensions['A'].width = 17
			a6.column_dimensions['B'].width = 25
			a6.column_dimensions['C'].width = 15				
			sumaryG=temp.create_sheet("Sectiunea G. Manual input")
			sumaryG.sheet_view.showGridLines = False
			sumaryG.cell(row=2,column=1).value="Sectiunea G"
			sumaryG.cell(row=2,column=1).font=cap_tabeltitlu		
			sumaryG.column_dimensions['A'].width =45
			sumaryG.cell(row=5,column=1).value="Total Nr. Bonuri Fiscale"
			sumaryG.cell(row=5,column=1).font=cap_tabelbold
			sumaryG.cell(row=6,column=1).font=cap_tabelbold
			sumaryG.cell(row=7,column=1).font=cap_tabelbold						
			sumaryG.cell(row=6,column=1).value="Total incasari in perioada de raportare prin intermediul AMEF ( aparate de marcatelectronice fiscale ) inclusiv incasarile prin intermediul bonurilor fiscale care indeplinesc conditiile unei facturi simplificate indiferent daca au/nu au inscris codul de inregistrare in scopuri de TVA al beneficiarului (i1)"
			sumaryG['A6'].alignment=Alignment(wrap_text=True)
			sumaryG['A7'].alignment=Alignment(wrap_text=True)				
			sumaryG.cell(row=7,column=1).value="Total incasari in perioada de raportare efectuate din activitati exceptate de la obligatia utilizarii AMEF***) (i2) conform prevederilor legale in vigoare )"
			sumaryG.merge_cells('A9:H11')
			sumaryG.cell(row=9,column=1).font=cap_tabelbold
			for ii in range(1,9):
				sumaryG.cell(row=8,column=ii).border=border_bottom
				sumaryG.cell(row=11,column=ii).border=border_bottom
			sumaryG.cell(row=9,column=8).border=border_upperright
			sumaryG.cell(row=10,column=8).border=border_right		
			sumaryG.cell(row=11,column=8).border=border_lowerright		
			for ii in range(1,9):	
				sumaryG.cell(row=19,column=ii).border=border_bottom
				sumaryG.cell(row=20,column=ii).border=border_bottom			
			sumaryG.cell(row=5,column=2).border=border_right
			sumaryG.cell(row=6,column=2).border=border_right
			sumaryG.cell(row=7,column=2).border=border_right
			sumaryG.cell(row=5,column=1).border=border_right
			sumaryG.cell(row=6,column=1).border=border_right
			sumaryG.cell(row=7,column=1).border=border_right
			sumaryG.cell(row=4,column=1).border=border_bottom
			sumaryG.cell(row=5,column=1).border=border_bottom
			sumaryG.cell(row=6,column=1).border=border_bottom
			sumaryG.cell(row=7,column=1).border=border_bottom
			sumaryG.cell(row=5,column=2).value=0
			sumaryG.cell(row=6,column=2).value=0
			sumaryG.cell(row=7,column=2).value=0
			sumaryG.cell(row=13,column=1).value="Cota"		
			sumaryG.cell(row=13,column=2).value="Total baza impozabila"
			sumaryG.cell(row=13,column=3).value="TVA"
			sumaryG.cell(row=23,column=1).value="Cota"		
			sumaryG.cell(row=23,column=2).value="Total baza impozabila"
			sumaryG.cell(row=23,column=3).value="TVA"		
			for ii in range(1,4):
				sumaryG.cell(row=17,column=ii).border=border_lowerright
				sumaryG.cell(row=16,column=ii).border=border_lowerright
				sumaryG.cell(row=15,column=ii).border=border_lowerright
				sumaryG.cell(row=14,column=ii).border=border_lowerright
				sumaryG.cell(row=24,column=ii).border=border_lowerright
				sumaryG.cell(row=25,column=ii).border=border_lowerright
				sumaryG.cell(row=26,column=ii).border=border_lowerright
				sumaryG.cell(row=27,column=ii).border=border_lowerright																					
							
			for row in sumaryG['A13:C13']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryG['A23:C23']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')										
			sumaryG.cell(row=9,column=1).value="Incasari in perioada de raportare prin intermediul AMEF ( aparate de marcat electronice fiscale ) inclusiv incasarile prin intermediul bonurilor fiscale care indeplinesc conditiile unei facturi simplificate indiferent daca au/nu au inscris codul de inregistrare in scopuri de TVA al beneficiarului (i1)"
			sumaryG.cell(row=20,column=1).value="Incasari in perioada de raportare efectuate din activitati exceptate de la obligatia utilizarii AMEF***) (i2) conform prevederilor legale in vigoare )"
			sumaryG['A9'].alignment=Alignment(wrap_text=True)		
			sumaryG.cell(row=14,column=1).value="Cota 20%"
			sumaryG.cell(row=15,column=1).value="Cota 19%"
			sumaryG.cell(row=16,column=1).value="Cota 9%"
			sumaryG.cell(row=17,column=1).value="Cota 5%"

			sumaryG.cell(row=24,column=1).value="Cota 20%"
			sumaryG.cell(row=25,column=1).value="Cota 19%"
			sumaryG.cell(row=26,column=1).value="Cota 9%"
			sumaryG.cell(row=27,column=1).value="Cota 5%"

			sumaryG.cell(row=14,column=2).value=0
			sumaryG.cell(row=15,column=2).value=0
			sumaryG.cell(row=16,column=2).value=0
			sumaryG.cell(row=17,column=2).value=0


			sumaryG.cell(row=14,column=3).value=0
			sumaryG.cell(row=15,column=3).value=0
			sumaryG.cell(row=16,column=3).value=0
			sumaryG.cell(row=17,column=3).value=0


			sumaryG.cell(row=24,column=2).value=0
			sumaryG.cell(row=25,column=2).value=0
			sumaryG.cell(row=26,column=2).value=0
			sumaryG.cell(row=27,column=2).value=0

			sumaryG.cell(row=24,column=3).value=0
			sumaryG.cell(row=25,column=3).value=0
			sumaryG.cell(row=26,column=3).value=0
			sumaryG.cell(row=27,column=3).value=0


			sumaryI=temp.create_sheet("Sectiunea I 1. Manual input")

			for row in sumaryI['A7:C7']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A17:C17']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A27:C27']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A37:C37']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A47:C47']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')																
			sumaryI.cell(row=2,column=1).value="Sectiunea I"
			sumaryI.cell(row=2,column=1).font=cap_tabeltitlu
			sumaryI.sheet_view.showGridLines = False
			sumaryI.cell(row=7,column=1).value="Cota"
			sumaryI.cell(row=7,column=2).value="Baza impozabila"
			sumaryI.cell(row=7,column=3).value="TVA"

			sumaryI.cell(row=7,column=1).value="Cota"
			sumaryI.cell(row=7,column=2).value="Baza impozabila"
			sumaryI.cell(row=7,column=3).value="TVA"
			sumaryI.cell(row=17,column=1).value="Cota"
			sumaryI.cell(row=17,column=2).value="Baza impozabila"
			sumaryI.cell(row=17,column=3).value="TVA"
			sumaryI.cell(row=27,column=1).value="Cota"
			sumaryI.cell(row=27,column=2).value="Baza impozabila"
			sumaryI.cell(row=27,column=3).value="TVA"
			sumaryI.cell(row=37,column=1).value="Cota"
			sumaryI.cell(row=37,column=2).value="Baza impozabila"
			sumaryI.cell(row=37,column=3).value="TVA"								
			sumaryI.cell(row=47,column=1).value="Cota"
			sumaryI.cell(row=47,column=2).value="Baza impozabila"
			sumaryI.cell(row=47,column=3).value="TVA"
			sumaryI.cell(row=4,column=1).value="1.1 Livrari de bunuri/prestari de servicii pentru care s-au emis facturi simplificate care au inscris codul de inregistrare in scopuri de TVA al beneficiarului %"
			sumaryI.cell(row=4,column=1).font=cap_tabelbold
			sumaryI.cell(row=14,column=1).font=cap_tabelbold
			sumaryI.cell(row=24,column=1).font=cap_tabelbold
			sumaryI.cell(row=34,column=1).font=cap_tabelbold
			sumaryI.cell(row=44,column=1).font=cap_tabelbold								
			sumaryI.cell(row=14,column=1).value="1.2 Livrari de bunuri/prestari de servicii pentru care s-au emis facturi simplificate fara a avea inscris codul de inregistrare in scopuri de TVA al beneficiarului %"
			sumaryI['A4'].alignment=Alignment(wrap_text=True)
			sumaryI['A14'].alignment=Alignment(wrap_text=True)
			sumaryI['A24'].alignment=Alignment(wrap_text=True)
			sumaryI['A34'].alignment=Alignment(wrap_text=True)
			sumaryI['A44'].alignment=Alignment(wrap_text=True)								
			sumaryI.cell(row=24,column=1).value="1.3 Achizitii de bunuri si servicii pentru care s-au primit facturi simplificate de la persoane impozabile care aplica sistemul normal de TVA si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

			sumaryI.cell(row=34,column=1).value="1.4 Achizitii de bunuri si servicii pentru care s-au primit facturi simplificate de la persoane impozabile care aplica sistemul de TVA la incasare si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

			sumaryI.cell(row=44,column=1).value="1.5 Achizitii de bunuri si servicii pentru care s-au primit bonuri fiscale care indeplinesc conditiile unei facturi simplificate si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

			for ip in range(1,11):
				sumaryI.cell(row=3,column=ip).border=border_bottom
				sumaryI.cell(row=6,column=ip).border=border_top
				sumaryI.cell(row=13,column=ip).border=border_bottom
				sumaryI.cell(row=16,column=ip).border=border_top
				sumaryI.cell(row=23,column=ip).border=border_bottom
				sumaryI.cell(row=26,column=ip).border=border_top
				sumaryI.cell(row=33,column=ip).border=border_bottom
				sumaryI.cell(row=36,column=ip).border=border_top
				sumaryI.cell(row=43,column=ip).border=border_bottom
				sumaryI.cell(row=46,column=ip).border=border_top
			sumaryI.cell(row=4,column=10).border=border_upperright
			sumaryI.cell(row=5,column=10).border=border_lowerright
			sumaryI.cell(row=14,column=10).border=border_upperright
			sumaryI.cell(row=15,column=10).border=border_lowerright
			sumaryI.cell(row=24,column=10).border=border_upperright
			sumaryI.cell(row=25,column=10).border=border_lowerright
			sumaryI.cell(row=34,column=10).border=border_upperright
			sumaryI.cell(row=35,column=10).border=border_lowerright
			sumaryI.cell(row=44,column=10).border=border_upperright
			sumaryI.cell(row=45,column=10).border=border_lowerright						

			sumaryI.cell(row=4,column=10).border=border_upperright
			sumaryI.cell(row=5,column=10).border=border_lowerright
			sumaryI.cell(row=14,column=10).border=border_upperright
			sumaryI.cell(row=15,column=10).border=border_lowerright
			sumaryI.cell(row=24,column=10).border=border_upperright
			sumaryI.cell(row=25,column=10).border=border_lowerright
			sumaryI.cell(row=34,column=10).border=border_upperright
			sumaryI.cell(row=35,column=10).border=border_lowerright
			sumaryI.cell(row=44,column=10).border=border_upperright
			sumaryI.cell(row=45,column=10).border=border_lowerright										
			sumaryI.cell(row=4,column=10).border=border_right
			sumaryI.cell(row=5,column=10).border=border_right
			sumaryI.cell(row=14,column=10).border=border_right
			sumaryI.cell(row=15,column=10).border=border_right
			sumaryI.cell(row=24,column=10).border=border_right
			sumaryI.cell(row=25,column=10).border=border_right
			sumaryI.cell(row=34,column=10).border=border_right
			sumaryI.cell(row=35,column=10).border=border_right
			sumaryI.cell(row=44,column=10).border=border_right
			sumaryI.cell(row=45,column=10).border=border_right																																																	
			for io in range(8,13):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(18,23):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(28,33):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(38,43):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(48,53):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright																		
			sumaryI.merge_cells('A4:J5')
			sumaryI.merge_cells('A14:J15')
			sumaryI.merge_cells('A24:J25')
			sumaryI.merge_cells('A34:J35')
			sumaryI.merge_cells('A44:J45')
									
			sumaryI.cell(row=8,column=1).value="Cota 24%"
			sumaryI.cell(row=9,column=1).value="Cota 20%"
			sumaryI.cell(row=10,column=1).value="Cota 19%"
			sumaryI.cell(row=11,column=1).value="Cota 9%"
			sumaryI.cell(row=12,column=1).value="Cota 5%"

			sumaryI.cell(row=8,column=2).value=0
			sumaryI.cell(row=9,column=2).value=0
			sumaryI.cell(row=10,column=2).value=0
			sumaryI.cell(row=11,column=2).value=0
			sumaryI.cell(row=12,column=2).value=0

			sumaryI.cell(row=18,column=2).value=0
			sumaryI.cell(row=19,column=2).value=0
			sumaryI.cell(row=20,column=2).value=0
			sumaryI.cell(row=21,column=2).value=0
			sumaryI.cell(row=22,column=2).value=0

			sumaryI.cell(row=28,column=2).value=0
			sumaryI.cell(row=29,column=2).value=0
			sumaryI.cell(row=30,column=2).value=0
			sumaryI.cell(row=31,column=2).value=0
			sumaryI.cell(row=32,column=2).value=0

			sumaryI.cell(row=38,column=2).value=0
			sumaryI.cell(row=39,column=2).value=0
			sumaryI.cell(row=40,column=2).value=0
			sumaryI.cell(row=41,column=2).value=0
			sumaryI.cell(row=42,column=2).value=0

			sumaryI.cell(row=48,column=2).value=0
			sumaryI.cell(row=49,column=2).value=0
			sumaryI.cell(row=50,column=2).value=0
			sumaryI.cell(row=51,column=2).value=0
			sumaryI.cell(row=52,column=2).value=0

			sumaryI.cell(row=8,column=3).value=0
			sumaryI.cell(row=9,column=3).value=0
			sumaryI.cell(row=10,column=3).value=0
			sumaryI.cell(row=11,column=3).value=0
			sumaryI.cell(row=12,column=3).value=0

			sumaryI.cell(row=18,column=3).value=0
			sumaryI.cell(row=19,column=3).value=0
			sumaryI.cell(row=20,column=3).value=0
			sumaryI.cell(row=21,column=3).value=0
			sumaryI.cell(row=22,column=3).value=0

			sumaryI.cell(row=28,column=3).value=0
			sumaryI.cell(row=29,column=3).value=0
			sumaryI.cell(row=30,column=3).value=0
			sumaryI.cell(row=31,column=3).value=0
			sumaryI.cell(row=32,column=3).value=0

			sumaryI.cell(row=38,column=3).value=0
			sumaryI.cell(row=39,column=3).value=0
			sumaryI.cell(row=40,column=3).value=0
			sumaryI.cell(row=41,column=3).value=0
			sumaryI.cell(row=42,column=3).value=0

			sumaryI.cell(row=48,column=3).value=0
			sumaryI.cell(row=49,column=3).value=0
			sumaryI.cell(row=50,column=3).value=0
			sumaryI.cell(row=51,column=3).value=0
			sumaryI.cell(row=52,column=3).value=0		

			sumaryI.cell(row=8,column=1).value="Cota 24%"
			sumaryI.cell(row=9,column=1).value="Cota 20%"
			sumaryI.cell(row=10,column=1).value="Cota 19%"
			sumaryI.cell(row=11,column=1).value="Cota 9%"
			sumaryI.cell(row=12,column=1).value="Cota 5%"

			sumaryI.cell(row=18,column=1).value="Cota 24%"
			sumaryI.cell(row=19,column=1).value="Cota 20%"
			sumaryI.cell(row=20,column=1).value="Cota 19%"
			sumaryI.cell(row=21,column=1).value="Cota 9%"
			sumaryI.cell(row=22,column=1).value="Cota 5%"

			sumaryI.cell(row=28,column=1).value="Cota 24%"
			sumaryI.cell(row=29,column=1).value="Cota 20%"
			sumaryI.cell(row=30,column=1).value="Cota 19%"
			sumaryI.cell(row=31,column=1).value="Cota 9%"
			sumaryI.cell(row=32,column=1).value="Cota 5%"

			sumaryI.cell(row=38,column=1).value="Cota 24%"
			sumaryI.cell(row=39,column=1).value="Cota 20%"
			sumaryI.cell(row=40,column=1).value="Cota 19%"
			sumaryI.cell(row=41,column=1).value="Cota 9%"
			sumaryI.cell(row=42,column=1).value="Cota 5%"

			sumaryI.cell(row=48,column=1).value="Cota 24%"
			sumaryI.cell(row=49,column=1).value="Cota 20%"
			sumaryI.cell(row=50,column=1).value="Cota 19%"
			sumaryI.cell(row=51,column=1).value="Cota 9%"
			sumaryI.cell(row=52,column=1).value="Cota 5%"

			sumaryI.column_dimensions['B'].width = 15	
			sumaryG.column_dimensions['B'].width = 20	
			a7.column_dimensions['B'].width = 12						
			a7.column_dimensions['C'].width = 22
			a6.column_dimensions['B'].width = 22						
			a6.column_dimensions['C'].width = 12


			# for row in sumary['A5:D34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['F5:I34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['K5:N34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['P5:S34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['U5:X34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# listanoua=['A','B','C','D','F','G','H','I','K','L','M','N','P','Q','R','S','U','V','W','W','X']
			# for column in listanoua:
			# 	for i in listanoua:
			# 		if (column==i):
	# 			sumary.column_dimensions[column].width = 15

			
			# for i in range(0 ,len(tip)):
		# folderpath="D:/D300 to XML/docs"
		folderpath="/home/mirus_app/storage_spreadsheet"
		# folderpath="C:/Users/Cristian.Iordache/Documents/D300 to XML Final CI/D300 to XML 2/storage"
		file_pathFS = os.path.join(folderpath, "One VAT app spreadsheets " +str(clientname)+".xlsx")
		temp.save(file_pathFS)
		# return send_from_directory("D:/D300 to XML/docs","One VAT app spreadsheets.xlsx",as_attachment=True)
		return send_from_directory("/home/mirus_app/storage_spreadsheet","One VAT app spreadsheets " +str(clientname)+".xlsx",as_attachment=True)
		return render_template('D3APPS second step.html')




#====================================================================THALES=============================================================================================
@app.route('/')
def my_form_thales():
    return render_template('D3APPS dashboard.html')

global LL_g
@app.route('/D3APPS/CIEL-THALES')
def my_form_D300_thales():
	return render_template('D3APPS.html')

@app.route('/D3APPS/CIEL-THALES', methods=['POST', 'GET'])
def D300xml_thales():
	if request.method == 'POST':
		clientname=request.form.get('client')
		D300 = request.files["far"]
		val1 = request.form.get('D300')
		val2 = request.form.get('D390')
		val3 = request.form.get('D394')
		val4 = request.form.get('xyz')
		dropdown = request.form.get('trezorerie')
		dropdownlimba = request.form.get('limba')
		soldLunaTrecuta = request.form.get('largeAm')

	
		# #print(soldLunaTrecuta)
	if val1=="":
		# #print("Da")  # daca e bifat
		val1 = 1
	else:
		#print(val1)            
		val1 = 0
		# #print("Nu")

	if val2=="":  # daca e bifat
		val2 = 1
	else:
		# #print(val2)            
		val2 = 0

	if val3=="":  # daca e bifat
		val3 = 1
	else:
		#print(val3)            
		val3 = 0
		
	if str(dropdownlimba)=="Romana(RO)":
		option=1
	else:
		option=0	


	#print(val4)
	#print(val4,"--------------------------------------------")		
	cap_tabel_color_verde = PatternFill(start_color = '00B050', end_color ='00B050', fill_type = 'solid')
	cap_tabel_color_verde_deschis = PatternFill(start_color = '92D050', end_color ='92D050', fill_type = 'solid')
	cap_tabel_color_black = PatternFill(start_color = '000000', end_color ='000000', fill_type = 'solid')
	cap_tabel = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
	cap_tabelbold = Font(name='Calibri', size=10, bold=True)
	cap_tabeltitlu = Font(name='Calibri', size=15, bold=True,underline='single')
	scrisincredibildemare = Font(name='Calibri', size=30, bold=True)
	cap_tabel_color_GT_movdeschis = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')
	cap_tabel_color_GT_movinchis = PatternFill(start_color='CCC0DA', end_color='CCC0DA', fill_type='solid')

	temp = openpyxl.load_workbook(D300,data_only=True)
	ws = temp.active
	
	if(option==1):
		Sheet1=temp.create_sheet('Cover sheet')
		fonta = PatternFill(start_color = 'ffffff', end_color ='ffffff', fill_type = 'solid')
		fontg = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		font2 = Font(name = 'Georgia', size = 10, bold = True, color="000000")
		font1 = Font(name = 'Georgia', size = 10, color = "FFFFFF", bold = True,italic=True)
		font3 = Font(name = 'Georgia', size = 10, color = "000000",italic=True)
		culoare = PatternFill(start_color = '182A54', end_color ='182A54', fill_type = 'solid') 
		culoare2 = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		culoare3 = PatternFill(start_color = 'D9E1F2', end_color ='D9E1F2', fill_type = 'solid')
		culoare4 = PatternFill(start_color = 'E2EFDA', end_color ='E2EFDA', fill_type = 'solid')
		culoare5 = PatternFill(start_color = 'FFF2CC', end_color ='FFF2CC', fill_type = 'solid')
		culoare6 = PatternFill(start_color = '808080', end_color ='808080', fill_type = 'solid')
		font4 = Font(name = 'Georgia', size = 10, color = "000000",underline='single',bold=True)
		font5 = Font(name = 'Georgia', size = 10, color = "ffffff",underline='single',bold=True)
		border = Border(bottom=Side(style='dotted'))
		border2 = Border(top=Side(style='dotted'))
		border3 = Border(left=Side(style='dotted'))
		border4 = Border(right=Side(style='dotted'))
		border5 = Border(left=Side(style='dotted'),top=Side(style='dotted'))
		border6 = Border(left=Side(style='dotted'),bottom=Side(style='dotted'))
		border7 = Border(right=Side(style='dotted'),bottom=Side(style='dotted'))
		border8 = Border(right=Side(style='dotted'),top=Side(style='dotted'))
		# border9 = Border(right=Side(style='double'),bottom=Side(style='double'),top=Side(style='double'),left=Side(style='double'))


		Sheet1.cell(row=31, column=2).border=border2
		Sheet1.cell(row=31, column=3).border=border2
		Sheet1.cell(row=31, column=4).border=border2
		Sheet1.cell(row=31, column=5).border=border2
		Sheet1.cell(row=31, column=6).border=border2
		Sheet1.cell(row=31, column=7).border=border2
		Sheet1.cell(row=31, column=2).border=border5
		Sheet1.cell(row=31, column=7).border=border8


		Sheet1.cell(row=75, column=2).border=border
		Sheet1.cell(row=75, column=3).border=border
		Sheet1.cell(row=75, column=4).border=border
		Sheet1.cell(row=75, column=5).border=border
		Sheet1.cell(row=75, column=6).border=border
		Sheet1.cell(row=75, column=7).border=border
		Sheet1.cell(row=75, column=7).border=border7

		Sheet1.cell(row=32, column=2).border=border3
		Sheet1.cell(row=33, column=2).border=border3
		Sheet1.cell(row=34, column=2).border=border3
		Sheet1.cell(row=35, column=2).border=border3
		Sheet1.cell(row=36, column=2).border=border3
		Sheet1.cell(row=37, column=2).border=border3
		Sheet1.cell(row=38, column=2).border=border3
		Sheet1.cell(row=39, column=2).border=border3
		Sheet1.cell(row=40, column=2).border=border3
		Sheet1.cell(row=41, column=2).border=border3
		Sheet1.cell(row=42, column=2).border=border3
		Sheet1.cell(row=43, column=2).border=border3
		Sheet1.cell(row=44, column=2).border=border3
		Sheet1.cell(row=45, column=2).border=border3
		Sheet1.cell(row=46, column=2).border=border3
		Sheet1.cell(row=47, column=2).border=border3
		Sheet1.cell(row=48, column=2).border=border3
		Sheet1.cell(row=49, column=2).border=border3
		Sheet1.cell(row=50, column=2).border=border3
		Sheet1.cell(row=51, column=2).border=border3
		Sheet1.cell(row=52, column=2).border=border3
		Sheet1.cell(row=53, column=2).border=border3
		Sheet1.cell(row=54, column=2).border=border3
		Sheet1.cell(row=55, column=2).border=border3
		Sheet1.cell(row=75, column=2).border=border6
		Sheet1.cell(row=56, column=2).border=border3
		Sheet1.cell(row=57, column=2).border=border3
		Sheet1.cell(row=58, column=2).border=border3
		Sheet1.cell(row=59, column=2).border=border3
		Sheet1.cell(row=60, column=2).border=border3
		Sheet1.cell(row=61, column=2).border=border3
		Sheet1.cell(row=62, column=2).border=border3
		Sheet1.cell(row=63, column=2).border=border3
		Sheet1.cell(row=64, column=2).border=border3
		Sheet1.cell(row=65, column=2).border=border3
		Sheet1.cell(row=66, column=2).border=border3
		Sheet1.cell(row=67, column=2).border=border3
		Sheet1.cell(row=68, column=2).border=border3
		Sheet1.cell(row=69, column=2).border=border3
		Sheet1.cell(row=70, column=2).border=border3
		Sheet1.cell(row=71, column=2).border=border3
		Sheet1.cell(row=72, column=2).border=border3
		Sheet1.cell(row=73, column=2).border=border3
		Sheet1.cell(row=74, column=2).border=border3

		Sheet1.cell(row=32, column=7).border=border4
		Sheet1.cell(row=33, column=7).border=border4
		Sheet1.cell(row=34, column=7).border=border4
		Sheet1.cell(row=35, column=7).border=border4
		Sheet1.cell(row=36, column=7).border=border4
		Sheet1.cell(row=37, column=7).border=border4
		Sheet1.cell(row=38, column=7).border=border4
		Sheet1.cell(row=39, column=7).border=border4
		Sheet1.cell(row=40, column=7).border=border4
		Sheet1.cell(row=41, column=7).border=border4
		Sheet1.cell(row=42, column=7).border=border4
		Sheet1.cell(row=43, column=7).border=border4
		Sheet1.cell(row=44, column=7).border=border4
		Sheet1.cell(row=45, column=7).border=border4
		Sheet1.cell(row=46, column=7).border=border4
		Sheet1.cell(row=47, column=7).border=border4
		Sheet1.cell(row=48, column=7).border=border4
		Sheet1.cell(row=49, column=7).border=border4
		Sheet1.cell(row=50, column=7).border=border4
		Sheet1.cell(row=51, column=7).border=border4
		Sheet1.cell(row=52, column=7).border=border4
		Sheet1.cell(row=53, column=7).border=border4
		Sheet1.cell(row=54, column=7).border=border4
		Sheet1.cell(row=55, column=7).border=border4
		Sheet1.cell(row=56, column=7).border=border4
		Sheet1.cell(row=57, column=7).border=border4
		Sheet1.cell(row=58, column=7).border=border4
		Sheet1.cell(row=59, column=7).border=border4
		Sheet1.cell(row=60, column=7).border=border4
		Sheet1.cell(row=61, column=7).border=border4
		Sheet1.cell(row=62, column=7).border=border4
		Sheet1.cell(row=63, column=7).border=border4
		Sheet1.cell(row=64, column=7).border=border4
		Sheet1.cell(row=65, column=7).border=border4
		Sheet1.cell(row=66, column=7).border=border4
		Sheet1.cell(row=67, column=7).border=border4
		Sheet1.cell(row=68, column=7).border=border4
		Sheet1.cell(row=69, column=7).border=border4
		Sheet1.cell(row=70, column=7).border=border4
		Sheet1.cell(row=71, column=7).border=border4
		Sheet1.cell(row=72, column=7).border=border4
		Sheet1.cell(row=73, column=7).border=border4
		Sheet1.cell(row=74, column=7).border=border4
		info=temp['Other info']
		valluna=""
		vallunaurmatoare=""
		valIban=""
		okdecembrie=0
		for i in range(0,len(listaluni)):
			if(str(info.cell(row=3,column=3).value)=="12"):
				okdecembrie=1
				vallunaurmatoare=listadenluni[0]
				valluna=listadenluni[11]
			else:
				if(listaluni[i]==str(info.cell(row=3,column=3).value)):
					valluna=listadenluni[i]
					vallunaurmatoare=listadenluni[i+1]
		var=Sheet1.cell(row=12,column=4).value
		# print(var)
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		for j in range(0,len(listatrez)):
			if(listatrez[j]==str(Sheet1.cell(row=10,column=4).value)):
				valIban=listaiban[j]
		Sheet1.cell(row=60, column=3).value='Perioada de plata: '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=61, column=3).value='="Suma de plata: " &D55&" RON "'
		Sheet1.cell(row=62, column=3).value="Moneda: RON"
		Sheet1.cell(row=63, column=3).value='Detalii plata: Decont TVA - '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		if(okdecembrie==1):
			Sheet1.cell(row=64, column=3).value='Data scadenta: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value+1)
		else:
			Sheet1.cell(row=64, column=3).value='Data scadenta: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=66, column=3).value='="Cod TVA: " & D8'
		Sheet1.cell(row=67, column=3).value='="Adresa: " &D7'
		Sheet1.cell(row=69, column=3).value="Beneficiar: BUGETUL DE STAT"
		Sheet1.cell(row=70, column=3).value='Cont IBAN: '+ str(valIban)
		Sheet1.cell(row=71, column=3).value="SWIFT / BIC: TREZROBU"
		Sheet1.cell(row=72, column=3).value="Deschis la:"+str(dropdown)
		Sheet1.cell(row=74, column=3).value="Nota: Orice taxe bancare legate de plata trebuie sa fie acoperite de catre platitor"
		Sheet1.cell(row=74, column=3).font=font2
		

		Sheet1.cell(row=14, column=3).value="Sumar"
		Sheet1.cell(row=14, column=3).font=font4
		Sheet1.cell(row=58, column=3).value="ORDIN DE PLATA"
		Sheet1.cell(row=58, column=3).font=font5

		for row in Sheet1['A1:N100']:
					for cell in row:
						cell.fill = fonta

		for row in Sheet1['N1:Z100']:
					for cell in row:
						cell.fill = fontg
		# print(get_fxrate(today.year))

		Sheet1.cell(row = 6, column = 3).value = "Denumire"
		Sheet1.cell(row = 6, column = 4).value = "='Other info'!C4"
		Sheet1.cell(row = 7, column = 3).value = "Adresa"
		Sheet1.cell(row = 7, column = 4).value = "='Other info'!C6"
		Sheet1.cell(row = 8, column = 3).value = "CUI"
		Sheet1.cell(row = 8, column = 4).value = "='Other info'!C5"
		Sheet1['D8'].alignment = Alignment(wrapText=True, horizontal='left')
		Sheet1.cell(row = 9, column = 3).value = "Nr. Reg. Com."
		Sheet1.cell(row = 9, column = 4).value = "J08/1139/2017"
		Sheet1.cell(row = 10, column = 3).value = "Administratia de care apartine"
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		Sheet1.cell(row = 11, column = 3).value = "Frecventa depunere declaratie/plata"
		Sheet1.cell(row = 11, column = 4).value = "Monthly"
		Sheet1.cell(row = 12, column = 3).value = "Perioada de raportare"
		Sheet1.cell(row = 12, column = 4).value = "=date('Other info'!C2,'Other info'!C3,1)"
		Sheet1.cell(row = 12, column = 4).number_format = 'mmmm yyyy'

		Sheet1.cell(row = 43, column = 4).value="Yes"
		Sheet1.cell(row = 46, column = 4).value="N/a"
		Sheet1.cell(row = 47, column = 4).value="N/a"
		Sheet1.cell(row = 50, column = 4).value="N/a"
		Sheet1.cell(row = 51, column = 4).value="N/a"
		Sheet1.cell(row = 43, column = 4).font=font5
		Sheet1.cell(row = 46, column = 4).font=font5
		Sheet1.cell(row = 47, column = 4).font=font5
		Sheet1.cell(row = 50, column = 4).font=font5
		Sheet1.cell(row = 51, column = 4).font=font5


		Sheet1.cell(row = 6, column = 3).font=font1
		Sheet1.cell(row = 6, column = 4).font=font2
		Sheet1.cell(row = 7, column = 3).font=font1
		Sheet1.cell(row = 7, column = 4).font=font2
		Sheet1.cell(row = 8, column = 3).font=font1
		Sheet1.cell(row = 8, column = 4).font=font2
		Sheet1.cell(row = 9, column = 3).font=font1
		Sheet1.cell(row = 9, column = 4).font=font2
		Sheet1.cell(row = 10, column = 3).font=font1
		Sheet1.cell(row = 10, column = 4).font=font3
		Sheet1.cell(row = 11, column = 3).font=font1
		Sheet1.cell(row = 11, column = 4).font=font3
		Sheet1.cell(row = 12, column = 3).font=font1
		Sheet1.cell(row = 12, column = 4).font=font3

		Sheet1.cell(row = 6, column = 3).fill=culoare
		Sheet1.cell(row = 7, column = 3).fill=culoare
		Sheet1.cell(row = 8, column = 3).fill=culoare
		Sheet1.cell(row = 9, column = 3).fill=culoare
		Sheet1.cell(row = 10, column = 3).fill=culoare
		Sheet1.cell(row = 11, column = 3).fill=culoare
		Sheet1.cell(row = 12, column = 3).fill=culoare
		Sheet1.cell(row = 6, column = 4).fill=culoare2
		Sheet1.cell(row = 7, column = 4).fill=culoare2
		Sheet1.cell(row = 8, column = 4).fill=culoare2
		Sheet1.cell(row = 9, column = 4).fill=culoare2
		Sheet1.cell(row = 10, column = 4).fill=culoare2
		Sheet1.cell(row = 11, column = 4).fill=culoare2
		Sheet1.cell(row = 12, column = 4).fill=culoare2

		Sheet1.cell(row = 16, column = 3).fill=culoare
		Sheet1.cell(row=16, column=3).font=font1
		Sheet1.cell(row=16, column=3).value="  D300"
		Sheet1.cell(row=16, column=3).hyperlink="#'D300 draft figures'!A1"
		# Sheet1.cell(row=16, column=3).border=border9
		Sheet1.row_dimensions[18].height=8

		Sheet1.cell(row = 19, column = 3).fill=culoare
		Sheet1.cell(row=19, column=3).font=font1
		Sheet1.cell(row=19, column=3).value="  D390"
		Sheet1.cell(row=19, column=3).hyperlink="#'D390 workings'!A1"
		# Sheet1.cell(row=19, column=3).border=border9
		Sheet1.row_dimensions[21].height=8

		Sheet1.cell(row = 22, column = 3).fill=culoare
		Sheet1.cell(row=22, column=3).font=font1
		Sheet1.cell(row=22, column=3).value="  D394"
		Sheet1.cell(row=22, column=3).hyperlink="#'D394--->>>'!A1"
		# Sheet1.cell(row=22, column=3).border=border9
		Sheet1.row_dimensions[24].height=8

		Sheet1.cell(row = 25, column = 3).fill=culoare
		Sheet1.cell(row=25, column=3).font=font1
		Sheet1.cell(row=25, column=3).value="  Jurnal vanzari"
		Sheet1.cell(row=25, column=3).hyperlink="#'Sales'!A1"
		# Sheet1.cell(row=25, column=3).border=border9
		Sheet1.row_dimensions[27].height=8

		Sheet1.cell(row = 28, column = 3).fill=culoare
		Sheet1.cell(row=28, column=3).font=font1
		Sheet1.cell(row=28, column=3).value="  Jurnal cumparari"
		Sheet1.cell(row=28, column=3).hyperlink="#'Purchases'!A1"
		# Sheet1.cell(row=28, column=3).border=border9
		Sheet1.row_dimensions[30].height=8


		Sheet1.cell(row = 58, column = 3).fill=culoare6
		Sheet1.cell(row = 32, column = 3).fill=culoare3
		Sheet1.cell(row = 33, column = 3).fill=culoare3
		Sheet1.cell(row = 34, column = 3).fill=culoare3
		Sheet1.cell(row = 35, column = 3).fill=culoare3
		Sheet1.cell(row = 36, column = 3).fill=culoare3
		Sheet1.cell(row = 37, column = 3).fill=culoare3
		Sheet1.cell(row = 38, column = 3).fill=culoare3
		Sheet1.cell(row = 39, column = 3).fill=culoare3
		Sheet1.cell(row = 32, column = 4).fill=culoare3
		Sheet1.cell(row = 33, column = 4).fill=culoare3
		Sheet1.cell(row = 34, column = 4).fill=culoare3
		Sheet1.cell(row = 35, column = 4).fill=culoare3
		Sheet1.cell(row = 36, column = 4).fill=culoare3
		Sheet1.cell(row = 37, column = 4).fill=culoare3
		Sheet1.cell(row = 38, column = 4).fill=culoare3
		Sheet1.cell(row = 39, column = 4).fill=culoare3
		Sheet1.cell(row = 32, column = 6).fill=culoare3
		Sheet1.cell(row = 33, column = 6).fill=culoare3
		Sheet1.cell(row = 34, column = 6).fill=culoare3
		Sheet1.cell(row = 35, column = 6).fill=culoare3
		Sheet1.cell(row = 36, column = 6).fill=culoare3
		Sheet1.cell(row = 37, column = 6).fill=culoare3
		Sheet1.cell(row = 38, column = 6).fill=culoare3
		Sheet1.cell(row = 39, column = 6).fill=culoare3

		Sheet1.cell(row = 41, column = 6).fill=culoare4
		Sheet1.cell(row = 42, column = 6).fill=culoare4
		Sheet1.cell(row = 43, column = 6).fill=culoare4
		Sheet1.cell(row = 44, column = 6).fill=culoare4
		Sheet1.cell(row = 45, column = 6).fill=culoare4
		Sheet1.cell(row = 46, column = 6).fill=culoare4
		Sheet1.cell(row = 47, column = 6).fill=culoare4
		Sheet1.cell(row = 48, column = 6).fill=culoare4
		Sheet1.cell(row = 49, column = 6).fill=culoare4
		Sheet1.cell(row = 50, column = 6).fill=culoare4
		Sheet1.cell(row = 51, column = 6).fill=culoare4
		Sheet1.cell(row = 41, column = 3).fill=culoare4
		Sheet1.cell(row = 42, column = 3).fill=culoare4
		Sheet1.cell(row = 43, column = 3).fill=culoare4
		Sheet1.cell(row = 44, column = 3).fill=culoare4
		Sheet1.cell(row = 45, column = 3).fill=culoare4
		Sheet1.cell(row = 46, column = 3).fill=culoare4
		Sheet1.cell(row = 47, column = 3).fill=culoare4
		Sheet1.cell(row = 48, column = 3).fill=culoare4
		Sheet1.cell(row = 49, column = 3).fill=culoare4
		Sheet1.cell(row = 50, column = 3).fill=culoare4
		Sheet1.cell(row = 51, column = 3).fill=culoare4
		Sheet1.cell(row = 41, column = 4).fill=culoare4
		Sheet1.cell(row = 42, column = 4).fill=culoare4
		Sheet1.cell(row = 43, column = 4).fill=culoare6
		Sheet1.cell(row = 44, column = 4).fill=culoare4
		Sheet1.cell(row = 45, column = 4).fill=culoare4
		Sheet1.cell(row = 46, column = 4).fill=culoare6
		Sheet1.cell(row = 47, column = 4).fill=culoare6
		Sheet1.cell(row = 48, column = 4).fill=culoare4
		Sheet1.cell(row = 49, column = 4).fill=culoare4
		Sheet1.cell(row = 50, column = 4).fill=culoare6
		Sheet1.cell(row = 51, column = 4).fill=culoare6

		Sheet1.cell(row = 53, column = 4).fill=culoare5
		Sheet1.cell(row = 54, column = 4).fill=culoare5
		Sheet1.cell(row = 55, column = 4).fill=culoare5
		Sheet1.cell(row = 53, column = 3).fill=culoare5
		Sheet1.cell(row = 54, column = 3).fill=culoare5
		Sheet1.cell(row = 55, column = 3).fill=culoare5
		Sheet1.cell(row = 53, column = 6).fill=culoare5
		Sheet1.cell(row = 54, column = 6).fill=culoare5
		Sheet1.cell(row = 55, column = 6).fill=culoare5


		Sheet1.cell(row = 32, column = 3).value="Pozitia curenta din punct de vedere TVA"
		Sheet1.cell(row = 32, column = 3).font=font4
		Sheet1.cell(row = 34, column = 3).value="Input TVA perioada curenta"
		Sheet1.cell(row = 35, column = 3).value="Output TVA perioada curenta"
		Sheet1.cell(row = 36, column = 3).value="TVA de plata perioada curenta"
		Sheet1.cell(row = 37, column = 3).value="TVA de recuperat perioada curenta"
		Sheet1.cell(row = 38, column = 3).value="TVA in curs de decontare pentru achizitii"
		Sheet1.cell(row = 39, column = 3).value="TVA in curs de decontare pentru livrari"
		Sheet1.cell(row = 32, column = 4).value="RON"

		Sheet1.cell(row = 32, column = 6).value="Euro(@"+get_fxrate(2022)+")"
		Sheet1.cell(row = 34, column = 4).value="='D300 draft figures'!C56"
		Sheet1.cell(row = 34, column = 6).value="=IFERROR(D34/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=34, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=34, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row = 35, column = 4).value="='D300 draft figures'!C32"
		Sheet1.cell(row = 35, column = 6).value="=IFERROR(D35/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 36, column = 4).value="=IF('D300 draft figures'!C58<>0,'D300 draft figures'!C58,0)"
		Sheet1.cell(row = 36, column = 6).value="=IFERROR(D36/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=35, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=35, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=6).number_format = '#,##0_);(#,##0)'

		Sheet1.cell(row=37, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=37, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=6).number_format = '#,##0_);(#,##0)'				

		Sheet1.cell(row = 37, column = 4).value='''=IF('D300 draft figures'!C57<>0,'D300 draft figures'!C57,"nil")'''
		Sheet1.cell(row = 37, column = 6).value="=iferror(D37/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 38, column = 4).value="='D300 draft figures'!C70"
		Sheet1.cell(row = 38, column = 6).value="=iferror(D38/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 39, column = 4).value=0
		Sheet1.cell(row = 39, column = 6).value="=iferror(D39/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 55, column = 4).value='''=IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)=0),'Cover sheet'!D36,
IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)<>0),IF('Cover sheet'!D36>IFERROR(VALUE('Cover sheet'!D47),0),'Cover sheet'!D36-IFERROR(VALUE('Cover sheet'!D47),0),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="No"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D47),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="Yes"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D51),0),"N/A"))))'''
		Sheet1.cell(row = 55, column = 6).value="=iferror(D55/"+get_fxrate(2022)+",0)"

		Sheet1.cell(row = 41, column = 3).value="Pozitia reportata"
		Sheet1.cell(row = 41, column = 3).font=font4
		Sheet1.row_dimensions[42].height = 0.2
		Sheet1.cell(row = 43, column = 3).value="Solicitat la rambursare"
		Sheet1.cell(row = 44, column = 3).value
		Sheet1.cell(row = 45, column = 3).value="TVA de rambursat nesolicitat"
		Sheet1.cell(row = 46, column = 3).value="Perioada"
		Sheet1.cell(row = 47, column = 3).value="Suma"
		Sheet1.cell(row = 48, column = 3).value
		Sheet1.cell(row = 49, column = 3).value="TVA de rambursat solicitat si in curs de auditare"
		Sheet1.cell(row = 50, column = 3).value="Perioada"
		Sheet1.cell(row = 51, column = 3).value="Suma"


		Sheet1.cell(row = 53, column = 3).value="Pozitia balantei de TVA"
		Sheet1.cell(row = 53, column = 3).font=font4
		Sheet1.cell(row = 55, column = 3).value="Pozitia TVA in exercitiul curent"


		Sheet1['C16'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C19'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C22'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C25'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C28'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['D10'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D11'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D12'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['C58'].alignment = Alignment(wrapText=True, horizontal='center')

		Sheet1.column_dimensions['C'].width = 65
		Sheet1.column_dimensions['D'].width = 20
		Sheet1.column_dimensions['A'].width = 2
		Sheet1.column_dimensions['B'].width = 3
		Sheet1.column_dimensions['E'].width = 1
		Sheet1.column_dimensions['F'].width = 20
		Sheet1.column_dimensions['G'].width = 2

		# img= openpyxl.drawing.image.Image('test.png')
		# Sheet1.add_image(img,'C16')

		# img= openpyxl.drawing.image.Image('test2.png')
		# Sheet1.add_image(img,'C19')

		# img= openpyxl.drawing.image.Image('test10.png')
		# Sheet1.add_image(img,'C22')

		# img= openpyxl.drawing.image.Image('test6.png')
		# Sheet1.add_image(img,'C25')

		# img= openpyxl.drawing.image.Image('test7.png')
		# Sheet1.add_image(img,'C28')

		# img= openpyxl.drawing.image.Image('test6.png')
		# Sheet1.add_image(img,'D16')

		# img= openpyxl.drawing.image.Image('test7.png')
		# Sheet1.add_image(img,'D19')

		# img= openpyxl.drawing.image.Image('test8.png')
		# Sheet1.add_image(img,'D22')

		# img= openpyxl.drawing.image.Image('test9.png')
		# Sheet1.add_image(img,'D25')

		# img= openpyxl.drawing.image.Image('test5.png')
		# Sheet1.add_image(img,'D28')

		Sheet1.merge_cells(start_row=28, start_column=3, end_row=29, end_column=3)
		Sheet1.merge_cells(start_row=25, start_column=3, end_row=26, end_column=3)
		Sheet1.merge_cells(start_row=22, start_column=3, end_row=23, end_column=3)
		Sheet1.merge_cells(start_row=19, start_column=3, end_row=20, end_column=3)
		Sheet1.merge_cells(start_row=16, start_column=3, end_row=17, end_column=3)
		Sheet1.merge_cells(start_row=6, start_column=4, end_row=6, end_column=12)
		Sheet1.merge_cells(start_row=7, start_column=4, end_row=7, end_column=12)
		Sheet1.merge_cells(start_row=8, start_column=4, end_row=8, end_column=12)
		Sheet1.merge_cells(start_row=9, start_column=4, end_row=9, end_column=12)
		Sheet1.merge_cells(start_row=10, start_column=4, end_row=10, end_column=12)
		Sheet1.merge_cells(start_row=11, start_column=4, end_row=11, end_column=12)
		Sheet1.merge_cells(start_row=12, start_column=4, end_row=12, end_column=12)
		Sheet1.merge_cells(start_row=58, start_column=3, end_row=58, end_column=6)

	if(option==0):
		Sheet1=temp.create_sheet('Cover sheet')
		fonta = PatternFill(start_color = 'ffffff', end_color ='ffffff', fill_type = 'solid')
		fontg = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		font2 = Font(name = 'Georgia', size = 10, bold = True, color="000000")
		font1 = Font(name = 'Georgia', size = 10, color = "FFFFFF", bold = True,italic=True)
		font3 = Font(name = 'Georgia', size = 10, color = "000000",italic=True)
		culoare = PatternFill(start_color = '182A54', end_color ='182A54', fill_type = 'solid') 
		culoare2 = PatternFill(start_color = 'EDEDED', end_color ='EDEDED', fill_type = 'solid')
		culoare3 = PatternFill(start_color = 'D9E1F2', end_color ='D9E1F2', fill_type = 'solid')
		culoare4 = PatternFill(start_color = 'E2EFDA', end_color ='E2EFDA', fill_type = 'solid')
		culoare5 = PatternFill(start_color = 'FFF2CC', end_color ='FFF2CC', fill_type = 'solid')
		culoare6 = PatternFill(start_color = '808080', end_color ='808080', fill_type = 'solid')
		font4 = Font(name = 'Georgia', size = 10, color = "000000",underline='single',bold=True)
		font5 = Font(name = 'Georgia', size = 10, color = "ffffff",underline='single',bold=True)
		border = Border(bottom=Side(style='dotted'))
		border2 = Border(top=Side(style='dotted'))
		border3 = Border(left=Side(style='dotted'))
		border4 = Border(right=Side(style='dotted'))
		border5 = Border(left=Side(style='dotted'),top=Side(style='dotted'))
		border6 = Border(left=Side(style='dotted'),bottom=Side(style='dotted'))
		border7 = Border(right=Side(style='dotted'),bottom=Side(style='dotted'))
		border8 = Border(right=Side(style='dotted'),top=Side(style='dotted'))
		# border9 = Border(right=Side(style='double'),bottom=Side(style='double'),top=Side(style='double'),left=Side(style='double'))

		Sheet1.cell(row=31, column=2).border=border2
		Sheet1.cell(row=31, column=3).border=border2
		Sheet1.cell(row=31, column=4).border=border2
		Sheet1.cell(row=31, column=5).border=border2
		Sheet1.cell(row=31, column=6).border=border2
		Sheet1.cell(row=31, column=7).border=border2
		Sheet1.cell(row=31, column=2).border=border5
		Sheet1.cell(row=31, column=7).border=border8


		Sheet1.cell(row=75, column=2).border=border
		Sheet1.cell(row=75, column=3).border=border
		Sheet1.cell(row=75, column=4).border=border
		Sheet1.cell(row=75, column=5).border=border
		Sheet1.cell(row=75, column=6).border=border
		Sheet1.cell(row=75, column=7).border=border
		Sheet1.cell(row=75, column=7).border=border7

		Sheet1.cell(row=32, column=2).border=border3
		Sheet1.cell(row=33, column=2).border=border3
		Sheet1.cell(row=34, column=2).border=border3
		Sheet1.cell(row=35, column=2).border=border3
		Sheet1.cell(row=36, column=2).border=border3
		Sheet1.cell(row=37, column=2).border=border3
		Sheet1.cell(row=38, column=2).border=border3
		Sheet1.cell(row=39, column=2).border=border3
		Sheet1.cell(row=40, column=2).border=border3
		Sheet1.cell(row=41, column=2).border=border3
		Sheet1.cell(row=42, column=2).border=border3
		Sheet1.cell(row=43, column=2).border=border3
		Sheet1.cell(row=44, column=2).border=border3
		Sheet1.cell(row=45, column=2).border=border3
		Sheet1.cell(row=46, column=2).border=border3
		Sheet1.cell(row=47, column=2).border=border3
		Sheet1.cell(row=48, column=2).border=border3
		Sheet1.cell(row=49, column=2).border=border3
		Sheet1.cell(row=50, column=2).border=border3
		Sheet1.cell(row=51, column=2).border=border3
		Sheet1.cell(row=52, column=2).border=border3
		Sheet1.cell(row=53, column=2).border=border3
		Sheet1.cell(row=54, column=2).border=border3
		Sheet1.cell(row=55, column=2).border=border3
		Sheet1.cell(row=75, column=2).border=border6
		Sheet1.cell(row=56, column=2).border=border3
		Sheet1.cell(row=57, column=2).border=border3
		Sheet1.cell(row=58, column=2).border=border3
		Sheet1.cell(row=59, column=2).border=border3
		Sheet1.cell(row=60, column=2).border=border3
		Sheet1.cell(row=61, column=2).border=border3
		Sheet1.cell(row=62, column=2).border=border3
		Sheet1.cell(row=63, column=2).border=border3
		Sheet1.cell(row=64, column=2).border=border3
		Sheet1.cell(row=65, column=2).border=border3
		Sheet1.cell(row=66, column=2).border=border3
		Sheet1.cell(row=67, column=2).border=border3
		Sheet1.cell(row=68, column=2).border=border3
		Sheet1.cell(row=69, column=2).border=border3
		Sheet1.cell(row=70, column=2).border=border3
		Sheet1.cell(row=71, column=2).border=border3
		Sheet1.cell(row=72, column=2).border=border3
		Sheet1.cell(row=73, column=2).border=border3
		Sheet1.cell(row=74, column=2).border=border3

		Sheet1.cell(row=32, column=7).border=border4
		Sheet1.cell(row=33, column=7).border=border4
		Sheet1.cell(row=34, column=7).border=border4
		Sheet1.cell(row=35, column=7).border=border4
		Sheet1.cell(row=36, column=7).border=border4
		Sheet1.cell(row=37, column=7).border=border4
		Sheet1.cell(row=38, column=7).border=border4
		Sheet1.cell(row=39, column=7).border=border4
		Sheet1.cell(row=40, column=7).border=border4
		Sheet1.cell(row=41, column=7).border=border4
		Sheet1.cell(row=42, column=7).border=border4
		Sheet1.cell(row=43, column=7).border=border4
		Sheet1.cell(row=44, column=7).border=border4
		Sheet1.cell(row=45, column=7).border=border4
		Sheet1.cell(row=46, column=7).border=border4
		Sheet1.cell(row=47, column=7).border=border4
		Sheet1.cell(row=48, column=7).border=border4
		Sheet1.cell(row=49, column=7).border=border4
		Sheet1.cell(row=50, column=7).border=border4
		Sheet1.cell(row=51, column=7).border=border4
		Sheet1.cell(row=52, column=7).border=border4
		Sheet1.cell(row=53, column=7).border=border4
		Sheet1.cell(row=54, column=7).border=border4
		Sheet1.cell(row=55, column=7).border=border4
		Sheet1.cell(row=56, column=7).border=border4
		Sheet1.cell(row=57, column=7).border=border4
		Sheet1.cell(row=58, column=7).border=border4
		Sheet1.cell(row=59, column=7).border=border4
		Sheet1.cell(row=60, column=7).border=border4
		Sheet1.cell(row=61, column=7).border=border4
		Sheet1.cell(row=62, column=7).border=border4
		Sheet1.cell(row=63, column=7).border=border4
		Sheet1.cell(row=64, column=7).border=border4
		Sheet1.cell(row=65, column=7).border=border4
		Sheet1.cell(row=66, column=7).border=border4
		Sheet1.cell(row=67, column=7).border=border4
		Sheet1.cell(row=68, column=7).border=border4
		Sheet1.cell(row=69, column=7).border=border4
		Sheet1.cell(row=70, column=7).border=border4
		Sheet1.cell(row=71, column=7).border=border4
		Sheet1.cell(row=72, column=7).border=border4
		Sheet1.cell(row=73, column=7).border=border4
		Sheet1.cell(row=74, column=7).border=border4
		

		Sheet1.cell(row=14, column=3).value="Summary"
		Sheet1.cell(row=14, column=3).font=font4
		Sheet1.cell(row=58, column=3).value="PAYMENT ORDER"
		Sheet1.cell(row=58, column=3).font=font5

		for row in Sheet1['A1:N100']:
					for cell in row:
						cell.fill = fonta

		for row in Sheet1['N1:Z100']:
					for cell in row:
						cell.fill = fontg


		Sheet1.cell(row = 6, column = 3).value = "Company"
		Sheet1.cell(row = 6, column = 4).value = "='Other info'!C4"
		Sheet1.cell(row = 7, column = 3).value = "Address"
		Sheet1.cell(row = 7, column = 4).value = "='Other info'!C6"
		Sheet1.cell(row = 8, column = 3).value = "VAT tax code"
		Sheet1.cell(row = 8, column = 4).value = "='Other info'!C5"
		Sheet1['D8'].alignment = Alignment(wrapText=True, horizontal='left')
		Sheet1.cell(row = 9, column = 3).value = "Registration no."
		Sheet1.cell(row = 9, column = 4).value = "J08/1139/2017"
		Sheet1.cell(row = 10, column = 3).value = "The administration it belongs to"
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		Sheet1.cell(row = 11, column = 3).value = "Frequency of declaration / payment"
		Sheet1.cell(row = 11, column = 4).value = "Monthly"
		Sheet1.cell(row = 12, column = 3).value = "Reporting period"
		Sheet1.cell(row = 12, column = 4).value = "=date('Other info'!C2,'Other info'!C3,1)"
		Sheet1.cell(row = 12, column = 4).number_format = 'mmmm yyyy'

		Sheet1.cell(row = 43, column = 4).value="Yes"
		Sheet1.cell(row = 46, column = 4).value="N/a"
		Sheet1.cell(row = 47, column = 4).value="N/a"
		Sheet1.cell(row = 50, column = 4).value="N/a"
		Sheet1.cell(row = 51, column = 4).value="N/a"
		Sheet1.cell(row = 43, column = 4).font=font5
		Sheet1.cell(row = 46, column = 4).font=font5
		Sheet1.cell(row = 47, column = 4).font=font5
		Sheet1.cell(row = 50, column = 4).font=font5
		Sheet1.cell(row = 51, column = 4).font=font5
		info=temp['Other info']
		valluna=""
		vallunaurmatoare=""
		valIban=""
		okdecembrie=0
		for i in range(0,len(listaluni)):
			if(str(info.cell(row=3,column=3).value)=="12"):
				okdecembrie=1
				vallunaurmatoare=listadenluni2[0]
				valluna=listadenluni2[11]
			else:
				if(listaluni[i]==str(info.cell(row=3,column=3).value)):
					valluna=listadenluni2[i]
					vallunaurmatoare=listadenluni2[i+1]
		var=Sheet1.cell(row=12,column=4).value
		# print(var)
		Sheet1.cell(row = 10, column = 4).value = str(dropdown)
		for j in range(0,len(listatrez)):
			if(listatrez[j]==str(Sheet1.cell(row=10,column=4).value)):
				valIban=listaiban[j]
		Sheet1.cell(row=60, column=3).value='Payment period: '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=61, column=3).value='="Suma de plata: " &D55&" RON "'
		Sheet1.cell(row=62, column=3).value="Currency: RON"
		Sheet1.cell(row=63, column=3).value='Payment details: VAT return - '+ str(valluna)+' '+ str(info.cell(row=2,column=3).value)
		if(okdecembrie==1):
			Sheet1.cell(row=64, column=3).value='Deadline: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value+1)
		else:
			Sheet1.cell(row=64, column=3).value='Deadline: 25-'+ str(vallunaurmatoare)+' '+ str(info.cell(row=2,column=3).value)
		Sheet1.cell(row=66, column=3).value='="Payer TIN: " & D8'
		Sheet1.cell(row=67, column=3).value='="Payer address: " &D7'
		Sheet1.cell(row=69, column=3).value="Beneficiary: BUGETUL DE STAT"
		Sheet1.cell(row=70, column=3).value='IBAN: '+ str(valIban)
		Sheet1.cell(row=71, column=3).value="SWIFT / BIC: TREZROBU"
		Sheet1.cell(row=72, column=3).value="Bank / Treasury:"+str(dropdown)
		Sheet1.cell(row=74, column=3).value="Note: Any banking fees connected with the payment must be covered by the tax payer."
		Sheet1.cell(row=74, column=3).font=font2

		Sheet1.cell(row = 6, column = 3).font=font1
		Sheet1.cell(row = 6, column = 4).font=font2
		Sheet1.cell(row = 7, column = 3).font=font1
		Sheet1.cell(row = 7, column = 4).font=font2
		Sheet1.cell(row = 8, column = 3).font=font1
		Sheet1.cell(row = 8, column = 4).font=font2
		Sheet1.cell(row = 9, column = 3).font=font1
		Sheet1.cell(row = 9, column = 4).font=font2
		Sheet1.cell(row = 10, column = 3).font=font1
		Sheet1.cell(row = 10, column = 4).font=font3
		Sheet1.cell(row = 11, column = 3).font=font1
		Sheet1.cell(row = 11, column = 4).font=font3
		Sheet1.cell(row = 12, column = 3).font=font1
		Sheet1.cell(row = 12, column = 4).font=font3

		Sheet1.cell(row = 6, column = 3).fill=culoare
		Sheet1.cell(row = 7, column = 3).fill=culoare
		Sheet1.cell(row = 8, column = 3).fill=culoare
		Sheet1.cell(row = 9, column = 3).fill=culoare
		Sheet1.cell(row = 10, column = 3).fill=culoare
		Sheet1.cell(row = 11, column = 3).fill=culoare
		Sheet1.cell(row = 12, column = 3).fill=culoare
		Sheet1.cell(row = 6, column = 4).fill=culoare2
		Sheet1.cell(row = 7, column = 4).fill=culoare2
		Sheet1.cell(row = 8, column = 4).fill=culoare2
		Sheet1.cell(row = 9, column = 4).fill=culoare2
		Sheet1.cell(row = 10, column = 4).fill=culoare2
		Sheet1.cell(row = 11, column = 4).fill=culoare2
		Sheet1.cell(row = 12, column = 4).fill=culoare2

		Sheet1.cell(row = 58, column = 3).fill=culoare6
		Sheet1.cell(row = 32, column = 3).fill=culoare3
		Sheet1.cell(row = 33, column = 3).fill=culoare3
		Sheet1.cell(row = 34, column = 3).fill=culoare3
		Sheet1.cell(row = 35, column = 3).fill=culoare3
		Sheet1.cell(row = 36, column = 3).fill=culoare3
		Sheet1.cell(row = 37, column = 3).fill=culoare3
		Sheet1.cell(row = 38, column = 3).fill=culoare3
		Sheet1.cell(row = 39, column = 3).fill=culoare3
		Sheet1.cell(row = 32, column = 4).fill=culoare3
		Sheet1.cell(row = 33, column = 4).fill=culoare3
		Sheet1.cell(row = 34, column = 4).fill=culoare3
		Sheet1.cell(row = 35, column = 4).fill=culoare3
		Sheet1.cell(row = 36, column = 4).fill=culoare3
		Sheet1.cell(row = 37, column = 4).fill=culoare3
		Sheet1.cell(row = 38, column = 4).fill=culoare3
		Sheet1.cell(row = 39, column = 4).fill=culoare3
		Sheet1.cell(row = 32, column = 6).fill=culoare3
		Sheet1.cell(row = 33, column = 6).fill=culoare3
		Sheet1.cell(row = 34, column = 6).fill=culoare3
		Sheet1.cell(row = 35, column = 6).fill=culoare3
		Sheet1.cell(row = 36, column = 6).fill=culoare3
		Sheet1.cell(row = 37, column = 6).fill=culoare3
		Sheet1.cell(row = 38, column = 6).fill=culoare3
		Sheet1.cell(row = 39, column = 6).fill=culoare3

		Sheet1.cell(row = 41, column = 6).fill=culoare4
		Sheet1.cell(row = 42, column = 6).fill=culoare4
		Sheet1.cell(row = 43, column = 6).fill=culoare4
		Sheet1.cell(row = 44, column = 6).fill=culoare4
		Sheet1.cell(row = 45, column = 6).fill=culoare4
		Sheet1.cell(row = 46, column = 6).fill=culoare4
		Sheet1.cell(row = 47, column = 6).fill=culoare4
		Sheet1.cell(row = 48, column = 6).fill=culoare4
		Sheet1.cell(row = 49, column = 6).fill=culoare4
		Sheet1.cell(row = 50, column = 6).fill=culoare4
		Sheet1.cell(row = 51, column = 6).fill=culoare4
		Sheet1.cell(row = 41, column = 3).fill=culoare4
		Sheet1.cell(row = 42, column = 3).fill=culoare4
		Sheet1.cell(row = 43, column = 3).fill=culoare4
		Sheet1.cell(row = 44, column = 3).fill=culoare4
		Sheet1.cell(row = 45, column = 3).fill=culoare4
		Sheet1.cell(row = 46, column = 3).fill=culoare4
		Sheet1.cell(row = 47, column = 3).fill=culoare4
		Sheet1.cell(row = 48, column = 3).fill=culoare4
		Sheet1.cell(row = 49, column = 3).fill=culoare4
		Sheet1.cell(row = 50, column = 3).fill=culoare4
		Sheet1.cell(row = 51, column = 3).fill=culoare4
		Sheet1.cell(row = 41, column = 4).fill=culoare4
		Sheet1.cell(row = 42, column = 4).fill=culoare4
		Sheet1.cell(row = 43, column = 4).fill=culoare6
		Sheet1.cell(row = 44, column = 4).fill=culoare4
		Sheet1.cell(row = 45, column = 4).fill=culoare4
		Sheet1.cell(row = 46, column = 4).fill=culoare6
		Sheet1.cell(row = 47, column = 4).fill=culoare6
		Sheet1.cell(row = 48, column = 4).fill=culoare4
		Sheet1.cell(row = 49, column = 4).fill=culoare4
		Sheet1.cell(row = 50, column = 4).fill=culoare6
		Sheet1.cell(row = 51, column = 4).fill=culoare6

		Sheet1.cell(row = 53, column = 4).fill=culoare5
		Sheet1.cell(row = 54, column = 4).fill=culoare5
		Sheet1.cell(row = 55, column = 4).fill=culoare5
		Sheet1.cell(row = 53, column = 3).fill=culoare5
		Sheet1.cell(row = 54, column = 3).fill=culoare5
		Sheet1.cell(row = 55, column = 3).fill=culoare5
		Sheet1.cell(row = 53, column = 6).fill=culoare5
		Sheet1.cell(row = 54, column = 6).fill=culoare5
		Sheet1.cell(row = 55, column = 6).fill=culoare5


		Sheet1.cell(row = 32, column = 3).value="Current VAT position"
		Sheet1.cell(row = 32, column = 3).font=font4
		Sheet1.cell(row = 34, column = 3).value="Input VAT for the period"
		Sheet1.cell(row = 35, column = 3).value="Output VAT for the period"
		Sheet1.cell(row = 36, column = 3).value="VAT Payable for the period"
		Sheet1.cell(row = 37, column = 3).value="VAT Recoverable for the period"
		Sheet1.cell(row = 38, column = 3).value="VAT under settlement for purchases"
		Sheet1.cell(row = 39, column = 3).value="VAT under settlement for deliveries"
		Sheet1.cell(row = 32, column = 4).value="RON"

		Sheet1.cell(row = 32, column = 6).value="Euro(@"+get_fxrate(2022)+")"
		Sheet1.cell(row = 34, column = 4).value="='D300 draft figures'!C56"
		Sheet1.cell(row = 34, column = 6).value="=IFERROR(D34/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=34, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=34, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row = 35, column = 4).value="='D300 draft figures'!C32"
		Sheet1.cell(row = 35, column = 6).value="=IFERROR(D35/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 36, column = 4).value="=IF('D300 draft figures'!C58<>0,'D300 draft figures'!C58,0)"
		Sheet1.cell(row = 36, column = 6).value="=IFERROR(D36/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=35, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=35, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=36, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row = 37, column = 4).value='''=IF('D300 draft figures'!C57<>0,'D300 draft figures'!C57,"nil")'''
		Sheet1.cell(row = 37, column = 6).value="=iferror(D37/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 38, column = 4).value="='D300 draft figures'!C70"
		Sheet1.cell(row = 38, column = 6).value="=iferror(D38/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 39, column = 4).value=0
		Sheet1.cell(row = 39, column = 6).value="=iferror(D39/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row = 55, column = 4).value='''=IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)=0),'Cover sheet'!D36,
IF(AND('Cover sheet'!D36<>"nil",IFERROR(VALUE('Cover sheet'!D47),0)<>0),IF('Cover sheet'!D36>IFERROR(VALUE('Cover sheet'!D47),0),'Cover sheet'!D36-IFERROR(VALUE('Cover sheet'!D47),0),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="No"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D47),0),
IF(AND('Cover sheet'!D47<>"nil",'Cover sheet'!D43="Yes"),'Cover sheet'!D47+IFERROR(VALUE('Cover sheet'!D51),0),"N/A"))))'''
		Sheet1.cell(row = 55, column = 6).value="=iferror(D55/"+get_fxrate(2022)+",0)"
		Sheet1.cell(row=37, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=37, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=38, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=39, column=6).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=4).number_format = '#,##0_);(#,##0)'
		Sheet1.cell(row=55, column=6).number_format = '#,##0_);(#,##0)'


		Sheet1.cell(row = 41, column = 3).value="Carry Over position"
		Sheet1.cell(row = 41, column = 3).font=font4
		Sheet1.row_dimensions[42].height = 0.2
		Sheet1.cell(row = 43, column = 3).value="Requested for reimbursement"
		Sheet1.cell(row = 44, column = 3).value
		Sheet1.cell(row = 45, column = 3).value="VAT refundable not yet requested"
		Sheet1.cell(row = 46, column = 3).value="Period"
		Sheet1.cell(row = 47, column = 3).value="Amount"
		Sheet1.cell(row = 48, column = 3).value
		Sheet1.cell(row = 49, column = 3).value="VAT refundable requested and under audit"
		Sheet1.cell(row = 50, column = 3).value="Period"
		Sheet1.cell(row = 51, column = 3).value="Amount"


		Sheet1.cell(row = 53, column = 3).value="VAT balance position"
		Sheet1.cell(row = 53, column = 3).font=font4
		Sheet1.cell(row = 55, column = 3).value="VAT position in the current return"


		Sheet1['C16'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C19'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C22'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C25'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['C28'].alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
		Sheet1['D10'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D11'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['D12'].alignment = Alignment(wrapText=True, horizontal='right')
		Sheet1['C58'].alignment = Alignment(wrapText=True, horizontal='center')
		Sheet1['D37'].alignment = Alignment(wrapText=True, horizontal='right')

		Sheet1.column_dimensions['C'].width = 65
		Sheet1.column_dimensions['D'].width = 20
		Sheet1.column_dimensions['A'].width = 2
		Sheet1.column_dimensions['B'].width = 3
		Sheet1.column_dimensions['E'].width = 1
		Sheet1.column_dimensions['F'].width = 20
		Sheet1.column_dimensions['G'].width = 2

		# img= openpyxl.drawing.image.Image('test.png')
		# Sheet1.add_image(img,'C16')

		# img= openpyxl.drawing.image.Image('test2.png')
		# Sheet1.add_image(img,'C19')

		# img= openpyxl.drawing.image.Image('test3.png')
		# Sheet1.add_image(img,'C22')

		# img= openpyxl.drawing.image.Image('test4.png')
		# Sheet1.add_image(img,'C25')

		# img= openpyxl.drawing.image.Image('test5.png')
		# Sheet1.add_image(img,'C28')

		# img= openpyxl.drawing.image.Image('test6.png')
		# Sheet1.add_image(img,'D16')

		# img= openpyxl.drawing.image.Image('test7.png')
		# Sheet1.add_image(img,'D19')

		# img= openpyxl.drawing.image.Image('test8.png')
		# Sheet1.add_image(img,'D22')

		# img= openpyxl.drawing.image.Image('test9.png')
		# Sheet1.add_image(img,'D25')
		Sheet1.cell(row = 16, column = 3).fill=culoare
		Sheet1.cell(row=16, column=3).font=font1
		Sheet1.cell(row=16, column=3).value="  D300"
		Sheet1.cell(row=16, column=3).hyperlink="#'D300 draft figures'!A1"
		# Sheet1.cell(row=16, column=3).border=border9
		Sheet1.row_dimensions[18].height=8

		Sheet1.cell(row = 19, column = 3).fill=culoare
		Sheet1.cell(row=19, column=3).font=font1
		Sheet1.cell(row=19, column=3).value="  D390"
		Sheet1.cell(row=19, column=3).hyperlink="#'D390 workings'!A1"
		# Sheet1.cell(row=19, column=3).border=border9
		Sheet1.row_dimensions[21].height=8

		Sheet1.cell(row = 22, column = 3).fill=culoare
		Sheet1.cell(row=22, column=3).font=font1
		Sheet1.cell(row=22, column=3).value="  D394"
		Sheet1.cell(row=22, column=3).hyperlink="#'D394--->>>'!A1"
		# Sheet1.cell(row=22, column=3).border=border9
		Sheet1.row_dimensions[24].height=8

		Sheet1.cell(row = 25, column = 3).fill=culoare
		Sheet1.cell(row=25, column=3).font=font1
		Sheet1.cell(row=25, column=3).value="  Sales Ledger"
		Sheet1.cell(row=25, column=3).hyperlink="#'Sales'!A1"
		# Sheet1.cell(row=25, column=3).border=border9
		Sheet1.row_dimensions[27].height=8

		Sheet1.cell(row = 28, column = 3).fill=culoare
		Sheet1.cell(row=28, column=3).font=font1
		Sheet1.cell(row=28, column=3).value="  Purchase Ledger"
		Sheet1.cell(row=28, column=3).hyperlink="#'Purchases'!A1"
		# Sheet1.cell(row=28, column=3).border=border9
		Sheet1.row_dimensions[30].height=8

		Sheet1.merge_cells(start_row=28, start_column=3, end_row=29, end_column=3)
		Sheet1.merge_cells(start_row=25, start_column=3, end_row=26, end_column=3)
		Sheet1.merge_cells(start_row=22, start_column=3, end_row=23, end_column=3)
		Sheet1.merge_cells(start_row=19, start_column=3, end_row=20, end_column=3)
		Sheet1.merge_cells(start_row=16, start_column=3, end_row=17, end_column=3)
		Sheet1.merge_cells(start_row=6, start_column=4, end_row=6, end_column=12)
		Sheet1.merge_cells(start_row=7, start_column=4, end_row=7, end_column=12)
		Sheet1.merge_cells(start_row=8, start_column=4, end_row=8, end_column=12)
		Sheet1.merge_cells(start_row=9, start_column=4, end_row=9, end_column=12)
		Sheet1.merge_cells(start_row=10, start_column=4, end_row=10, end_column=12)
		Sheet1.merge_cells(start_row=11, start_column=4, end_row=11, end_column=12)
		Sheet1.merge_cells(start_row=12, start_column=4, end_row=12, end_column=12)
		Sheet1.merge_cells(start_row=58, start_column=3, end_row=58, end_column=6)
	
	sales=temp['Sales']
	purchases=temp['Purchases']
	if(val1==1):
		sheetinutil1=temp.create_sheet('D300--->>>')
		sheetinutil1.sheet_view.showGridLines=False
		sheetinutil1.cell(row=2,column=1).value="Switch to next sheet for D300 Workings draft"
		sheetinutil1.cell(row=2,column=1).font=scrisincredibildemare
		amount=temp.create_sheet('D300 draft figures')
		amount.freeze_panes = 'A8'
		amount.auto_filter.ref = "A7:G71"
		amount.sheet_view.showGridLines = False
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "5":
					rand_tb = cell.row
					tdocc = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in sales[tdocc][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Total sales'")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "2":
					rand_tb = cell.row
					tdat = cell.column
					lun = len(sales[cell.column])
		try:
			listdocdate = [b.value for b in sales[tdat][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Total sales'")
			return render_template("index.html")
		listacurentas=[]						
		for k in range(0,len(listdocdate)):
			# print(datadocument[k][3:4])
			# print(datadocument[k][3:5])
			if(str(listdocdate[k])[5:6]=="0"):
				if(str(listdocdate[k])[6:7]==str(info.cell(row=3,column=3).value)):
					listacurentas.append("Yes")
				else:
					listacurentas.append("No")

			else:
				if(str(listdocdate[k])[5:7]==info.cell(row=3,column=3).value):
					listacurentas.append("Yes")
				else:
					listacurentas.append("No")
		for kk in range(0,len(listacurentas)):
			sales.cell(row=4+kk,column=70).value=listacurentas[kk]

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "2":
					rand_tb = cell.row
					supplierCell = cell.column
					lun = len(purchases[cell.column])
		try:
			datadocument = [b.value for b in purchases[supplierCell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for ' Doc. Date' in Purchases sheet")
			return render_template("index.html")
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "7.3":
					rand_tb = cell.row
					sal19c = cell.column
					lun = len(purchases[cell.column])
		try:
			salescol = [b.value for b in purchases[supplierCell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for ' Doc. Date' in Purchases sheet")
			return render_template("index.html")
		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "7.4":
					rand_tb = cell.row
					sal19tva = cell.column
					lun = len(purchases[cell.column])
		try:
			salescol = [b.value for b in purchases[supplierCell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for ' Doc. Date' in Purchases sheet")
			return render_template("index.html")						
		lunacurenta=[]
		for k in range(0,len(datadocument)):
			print(str(datadocument[k])[5:7])
			try:
				# print(datadocument[k][3:4])
				# print(datadocument[k][3:5])
				if(str(datadocument[k])[5:6]=="0"):
					if(str(datadocument[k])[6:7]==str(info.cell(row=3,column=3).value)):
						lunacurenta.append("Yes")
					else:
						lunacurenta.append("No")

				else:
					if(str(datadocument[k])[5:7]==info.cell(row=3,column=3).value):
						lunacurenta.append("Yes")
					else:
						lunacurenta.append("No")
			except:
				lunacurenta.append("Not applicable")
			# if(datadocument[k])
		for kk in range(0,len(lunacurenta)):
			purchases.cell(row=4+kk,column=70).value=lunacurenta[kk]			

		print(purchases.cell(row=3,column=10).value)
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "7.3":
					rand_tb = cell.row
					tax19b = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in purchases[tax19b][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '7.3'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "7.4":
					rand_tb = cell.row
					tax19vat = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[tax19vat][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '7.4'")
			return render_template("index.html")
		
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "12":
					rand_tb = cell.row
					tax5b = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in purchases[tax5b][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '12'")
			return render_template("index.html")
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "13":
					rand_tb = cell.row
					tax5vatb = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[tax5vatb][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '13'")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "13.3":
					rand_tb = cell.row
					tax19nexvat = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[tax19nexvat][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '13.3'")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "13.4":
					rand_tb = cell.row
					tax19nexvatb = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[tax19nexvatb][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '13.4'")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "22":
					rand_tb = cell.row
					taxareinvbun = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[taxareinvbun][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '22'")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "23":
					rand_tb = cell.row
					tvainvbun = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[tvainvbun][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '23'")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "24":
					rand_tb = cell.row
					taxareinvserv = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[taxareinvserv][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '24'")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "25":
					rand_tb = cell.row
					tvainvserv = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[tvainvserv][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '25'")
			return render_template("index.html")


		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "27":
					rand_tb = cell.row
					scutite = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[scutite][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '27'")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "28":
					rand_tb = cell.row
					neimp = cell.column
					lun = len(sales[cell.column])

		try:
			listBazaL = [b.value for b in purchases[neimp][rand_tb:lun]]
		except:
			flash("Please insert the correct header for '28'")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "34":
					rand_tb = cell.row
					tdocneded = cell.column
					lun = len(sales[cell.column])
		amount.cell(row=6, column=2).value="1"
		amount.cell(row=6, column=3).value="2"
		amount.cell(row=7, column=1).value="Row"
		amount.cell(row=7, column=2).value="Taxable basis"
		amount.cell(row=7, column=3).value="VAT amount"
		amount.cell(row=7, column=4).value="Comments"
		amount.cell(row=7, column=5).value="Journal Source"
		amount.cell(row=7, column=6).value="Flag Suma Control"
		amount.cell(row=7, column=7).value="Suma Control"
		amount.cell(row=8, column=1).value="1"
		amount.cell(row=9, column=1).value="2"
		amount.cell(row=10, column=1).value="3"
		amount.cell(row=11, column=1).value="3.1"
		amount.cell(row=12, column=1).value="4"
		amount.cell(row=13, column=1).value="5"
		amount.cell(row=14, column=1).value="5.1"
		amount.cell(row=15, column=1).value="6"
		amount.cell(row=16, column=1).value="7"
		amount.cell(row=17, column=1).value="7.1"
		amount.cell(row=18, column=1).value="8"
		amount.cell(row=19, column=1).value="9"
		amount.cell(row=20, column=1).value="10"
		amount.cell(row=21, column=1).value="11"
		amount.cell(row=22, column=1).value="12"
		amount.cell(row=23, column=1).value="12.1"
		amount.cell(row=24, column=1).value="12.2"
		amount.cell(row=25, column=1).value="12.3"
		amount.cell(row=26, column=1).value="13"
		amount.cell(row=27, column=1).value="14"
		amount.cell(row=28, column=1).value="15"
		amount.cell(row=29, column=1).value="16"
		amount.cell(row=30, column=1).value="17"
		amount.cell(row=31, column=1).value="18"
		amount.cell(row=32, column=1).value="19"
		amount.cell(row=33, column=1).value="20"
		amount.cell(row=34, column=1).value="20.1"
		amount.cell(row=35, column=1).value="21"
		amount.cell(row=36, column=1).value="22"
		amount.cell(row=37, column=1).value="22.1"
		amount.cell(row=38, column=1).value="23"
		amount.cell(row=39, column=1).value="24"
		amount.cell(row=40, column=1).value="25"
		amount.cell(row=41, column=1).value="26"
		amount.cell(row=42, column=1).value="27"
		amount.cell(row=43, column=1).value="27.1"
		amount.cell(row=44, column=1).value="27.2"
		amount.cell(row=45, column=1).value="27.3"
		amount.cell(row=46, column=1).value="28"
		amount.cell(row=47, column=1).value="29"
		amount.cell(row=48, column=1).value="30"
		amount.cell(row=49, column=1).value="30.1"
		amount.cell(row=50, column=1).value="31"
		amount.cell(row=51, column=1).value="31.1"
		amount.cell(row=52, column=1).value="32"
		amount.cell(row=53, column=1).value="33"
		amount.cell(row=54, column=1).value="34"
		amount.cell(row=55, column=1).value="35"
		amount.cell(row=56, column=1).value="36"
		amount.cell(row=57, column=1).value="37"
		amount.cell(row=58, column=1).value="38"
		amount.cell(row=59, column=1).value="39"
		amount.cell(row=60, column=1).value="40"
		amount.cell(row=61, column=1).value="41"
		amount.cell(row=62, column=1).value="42"
		amount.cell(row=63, column=1).value="43"
		amount.cell(row=64, column=1).value="44"
		amount.cell(row=65, column=1).value="45"
		amount.cell(row=66, column=1).value="46"

		amount.cell(row=68, column=1).value="A"
		amount.cell(row=69, column=1).value="A1"
		amount.cell(row=70, column=1).value="B"
		amount.cell(row=71, column=1).value="B.1"


		amount.cell(row=8, column=2).value=0
		amount.cell(row=9, column=2).value=0
		amount.cell(row=10, column=2).value=0		
		amount.cell(row=11, column=2).value=0
		amount.cell(row=12, column=2).value=0
		amount.cell(row=13, column=2).value='=round(sum(Purchases!'+str(taxareinvbun)+":"+str(taxareinvbun)+'),0)'
		amount.cell(row=14, column=2).value='=round(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(taxareinvbun)+":"+str(taxareinvbun)+'),0)'
		amount.cell(row=15, column=2).value=0
		amount.cell(row=16, column=2).value='=round(sum(Purchases!'+str(taxareinvserv)+":"+str(taxareinvserv)+'),0)'
		amount.cell(row=17, column=2).value='=round(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(taxareinvserv)+":"+str(taxareinvserv)+'),0)'	
		amount.cell(row=18, column=2).value=0
		amount.cell(row=19, column=2).value='=round(sum(Sales!'+str(sal19c)+":"+str(sal19c)+'),0)'
		amount.cell(row=20, column=2).value=0
		amount.cell(row=21, column=2).value=0
		amount.cell(row=22, column=2).value=0
		amount.cell(row=23, column=2).value=0
		amount.cell(row=24, column=2).value=0
		amount.cell(row=25, column=2).value=0
		amount.cell(row=26, column=2).value=0	
		amount.cell(row=27, column=2).value=0
		amount.cell(row=28, column=2).value=0
		amount.cell(row=30, column=2).value=0
		amount.cell(row=29, column=2).value=0
		amount.cell(row=39, column=2).value='=round(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tax19b)+':'+str(tax19b)+'),0)'						
		amount.cell(row=40, column=2).value=0								
		amount.cell(row=41, column=2).value='=round(sum(Purchases!'+str(tax5b)+":"+str(tax5b)+'),0)'								

		
		# amount.cell(row=22, column=2).value='=round(SUM(B23:B25),0)'


		amount.cell(row=31, column=2).value=0
		amount.cell(row=32, column=2).value='=B8+B10+B13+B16+B27+B15+B18+B31+B30+B29+B28+B26+B22+B21+B20+B19'
		amount.cell(row=33, column=2).value='=B13'
		amount.cell(row=34, column=2).value='=B14'
		amount.cell(row=35, column=2).value='=B15'
		amount.cell(row=36, column=2).value='=B16'
		amount.cell(row=37, column=2).value='=B17'
		amount.cell(row=38, column=2).value='=B18'

		
		amount.cell(row=42, column=2).value='=SUM(B43:B45)'
		amount.cell(row=43, column=2).value='=B23'
		amount.cell(row=44, column=2).value='=B24'
		amount.cell(row=45, column=2).value='=B25'
		amount.cell(row=46, column=2).value=0
		amount.cell(row=47, column=2).value=0
		amount.cell(row=48, column=2).value='=round(SUM(Purchases!'+str(scutite)+":"+str(scutite)+')+SUM(Purchases!'+str(neimp)+":"+str(neimp)+'),0)'
		amount.cell(row=49, column=2).value=0
		amount.cell(row=50, column=2).value='=SUM(B33:B47)-B34-B37-SUM(B43:B45)'
		amount.cell(row=51, column=2).value='=round(sumif(Purchases!BR:BR,"No",Purchases!'+str(tax19nexvat)+':'+str(tax19nexvat)+'),0)'
		amount.cell(row=52, column=2).value='=B50-B51'
		amount.cell(row=53, column=2).value=0
		amount.cell(row=54, column=2).value='=round(SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax19b)+':'+str(tax19b)+'),0)'
		amount.cell(row=55, column=2).value=0
		amount.cell(row=56, column=2).value='=SUM(B52:B55)'
		amount.cell(row=57, column=2).value=0
		amount.cell(row=58, column=2).value=0
		amount.cell(row=59, column=2).value=0
		amount.cell(row=60, column=2).value=0
		amount.cell(row=61, column=2).value='=SUM(B58:B60)'
		amount.cell(row=62, column=2).value=0
		amount.cell(row=63, column=2).value=0
		amount.cell(row=64, column=2).value='=B57+B62+B63'
		amount.cell(row=65, column=2).value='=IF((B61-B664)<0,0,B61-B64)'
		amount.cell(row=66, column=2).value='=IF((B64-B61)<0,0,B64)'

		amount.cell(row=68, column=2).value=0
		amount.cell(row=69, column=2).value=0
		amount.cell(row=70, column=2).value='=ROUND(SUM(Purchases!'+str(tax19nexvat)+":"+str(tax19nexvat)+',),0)'
		amount.cell(row=71, column=2).value='=B70'
		
		# #coloana TVA----------------------------------------------------

		for g in range(8, 13):
			amount.cell(row=g, column=3).value=0
		

		# # for h in range(13, 19):
		amount.cell(row=13, column=3).value='=round(sum(Purchases!'+str(tvainvbun)+":"+str(tvainvbun)+'),0)'
		amount.cell(row=14, column=3).value='=ROUND(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tvainvbun)+":"+str(tvainvbun)+'),0)'
		amount.cell(row=15, column=3).value=0
		amount.cell(row=16, column=3).value='=round(sum(Purchases!'+str(tvainvserv)+":"+str(tvainvserv)+'),0)'
		amount.cell(row=17, column=3).value='=ROUND(SUMIF(Purchases!BR:BR,"Yes",Purchases!'+str(tvainvserv)+":"+str(tvainvserv)+'),0)'
		amount.cell(row=18, column=3).value=0
			
		# # amount.cell(row=16,column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A16&"."&C$6,Purchases!$5:$5)-SUMIF(Purchases!$7:$7,$A18&"."&C$6,Purchases!$5:$5),0)'


		amount.cell(row=19, column=3).value='=round(SUM(Sales!'+str(sal19tva)+':'+str(sal19tva)+'),0)'
		amount.cell(row=20, column=3).value=0
		amount.cell(row=21, column=3).value=0
		amount.cell(row=22, column=3).value=0

		amount.cell(row=23, column=3).value=0
		amount.cell(row=24, column=3).value=0
		amount.cell(row=25, column=3).value=0


		
		for k in range(26, 31):
			amount.cell(row=k, column=3).value=0
		amount.cell(row=31, column=3).value=0
		amount.cell(row=32, column=3).value='=C8+C10+C13+C16+C27+C15+C18+C31+C30+C29+C28+C26+C22+C21+C20+C19'
		amount.cell(row=33, column=3).value='=C13'
		amount.cell(row=34, column=3).value='=C14'
		amount.cell(row=35, column=3).value='=C15'
		amount.cell(row=36, column=3).value='=C16'
		amount.cell(row=37, column=3).value='=C17'
		amount.cell(row=38, column=3).value='=C18'


		amount.cell(row=39, column=3).value='=round(sumif(Purchases!BR:BR,"Yes",Purchases!'+str(tax19vat)+':'+str(tax19vat)+'),0)'
		amount.cell(row=40, column=3).value=0
		amount.cell(row=41, column=3).value='=round(sum(Purchases!'+str(tax5vatb)+':'+str(tax5vatb)+'),0)'		
		amount.cell(row=42, column=3).value='=round(SUM(C43:C45),0)'
		amount.cell(row=43, column=3).value='=C23'
		amount.cell(row=44, column=3).value='=C24'
		amount.cell(row=45, column=3).value='=C25'
		amount.cell(row=46, column=3).value=0
		amount.cell(row=47, column=3).value=0
		amount.cell(row=48, column=3).value=0
		amount.cell(row=49, column=3).value=0
		amount.cell(row=50, column=3).value='=SUM(C33:C47)-C34-C37-SUM(C43:C45)'
		amount.cell(row=51, column=3).value='=SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax19nexvatb)+':'+str(tax19nexvatb)+')'
		amount.cell(row=52, column=3).value='=C51+C50-SUM(Purchases!'+str(tdocneded)+':'+str(tdocneded)+')'
		amount.cell(row=53, column=3).value=0
		amount.cell(row=54, column=3).value='=SUMIF(Purchases!BR:BR,"No",Purchases!'+str(tax19vat)+':'+str(tax19vat)+')'
		amount.cell(row=55, column=3).value=0
		amount.cell(row=56, column=3).value='=SUM(C52:C55)'
		amount.cell(row=57, column=3).value='=IF((C56-C32)<0,0,C56-C32)'
		amount.cell(row=58, column=3).value='=IF((C32-C56)<0,0,C32-C56)'
		amount.cell(row=59, column=3).value=0
		amount.cell(row=60, column=3).value=0
		amount.cell(row=61, column=3).value='=SUM(C58:C60)'	
		if soldLunaTrecuta == None or soldLunaTrecuta == "" or soldLunaTrecuta == " ":
			amount.cell(row=62, column=3).value=0
		else:
			amount.cell(row=62, column=3).value=int(soldLunaTrecuta) 
		#print(soldLunaTrecuta, "sold luna trecuta")
		amount.cell(row=63, column=3).value=0
		amount.cell(row=64, column=3).value='=C57+C62+C63'
		amount.cell(row=65, column=3).value='=IF((C61-C64)<0,0,C61-C64)'
		amount.cell(row=66, column=3).value='=IF((C64-C61)<0,0,C64-C61)'
		amount.cell(row=68, column=3).value=0
		amount.cell(row=69, column=3).value=0

		amount.cell(row=70, column=3).value='=ROUND(SUM(Purchases!'+str(tax19nexvatb)+":"+str(tax19nexvatb)+',),0)'
		amount.cell(row=71, column=3).value='=C70'

		# amount.cell(row=73, column=1).value='Informații privind valoarea totală, fără TVA, a operațiunilor prevăzute la art. 2781 alin. (1) lit. b) din Codul fiscal, respectiv a vânzărilor intracomunitare de bunuri la distanță și a prestărilor de servicii de telecomunicaţii, de radiodifuziune şi televiziune, precum și servicii furnizate pe cale electronică, către persoane neimpozabile din alte state membre UE'
		# amount.cell(row=73, column=2).value='Total an precedent'
		# amount.cell(row=73, column=3).value='An curent (inclusiv perioada de raportare)'

		# amount.cell(row=74, column=2).value=0
		# amount.cell(row=74, column=3).value=0


		amount.cell(row=22, column=5).value='Total'

		for m in it.chain(range(8, 13), range(19, 22), range(26, 32)):
			amount.cell(row=m, column=5).value='SALES'

		for n in it.chain(range(13, 19), range(23, 26), range(39, 42), range(46, 50), range(53, 56)):
			amount.cell(row=n, column=5).value='Purchases'
		
		for o in range(32, 39):
			amount.cell(row=o, column=5).value='Total'

		for p in range(42, 46):
			amount.cell(row=p, column=5).value='Total'
		
		amount.cell(row=50, column=5).value='Total'
		amount.cell(row=51, column=5).value='Purchases'
		amount.cell(row=52, column=5).value='Total'
		amount.cell(row=66, column=5).value='Purchases'

		for q in range(68, 71):
			amount.cell(row=q, column=5).value='Total'

		for r in range(56, 67):
			amount.cell(row=r, column=5).value='Total'

		amount.cell(row=71, column=5).value='Purchases'

		for s in it.chain(range(8, 13), range(26, 29), range(48, 50)):
			amount.cell(row=s, column=6).value='no VAT'

		for t in it.chain(range(13, 26), range(29, 46)):
			amount.cell(row=t, column=6).value='Add all'

		for u in it.chain(range(46, 48), range(52, 54), range(55, 67)):
			amount.cell(row=u, column=6).value='No basis'
		
		amount.cell(row=50, column=6).value='Add all'
		amount.cell(row=51, column=6).value='Se pune 0'
		amount.cell(row=54, column=6).value='Add all'

		for a1 in range(8, 13):
			amount.cell(row=a1, column=7).value='=B{0}'.format(a1)

		for b1 in range(13, 26):
			amount.cell(row=b1, column=7).value='=SUM(B{0}:C{0})'.format(b1)

		for c1 in range(26, 29):
			amount.cell(row=c1, column=7).value='=B{0}'.format(c1)
		
		for d1 in range(29, 46):
			amount.cell(row=d1, column=7).value='=SUM(B{0}:C{0})'.format(d1)

		amount.cell(row=46, column=7).value='=C46'
		amount.cell(row=47, column=7).value='=C47'
		amount.cell(row=48, column=7).value='=B48'
		amount.cell(row=49, column=7).value='=B49'
		amount.cell(row=50, column=7).value='=SUM(B50:C50)'
		amount.cell(row=51, column=7).value='=C51'
		amount.cell(row=52, column=7).value='=C52'
		amount.cell(row=53, column=7).value='=SUM(B53:C53)'
		amount.cell(row=54, column=7).value='=SUM(B54:C54)'
		amount.cell(row=2,column=1).value="D300 draft figures "
		amount.cell(row=2,column=1).font=cap_tabeltitlu
		amount.row_dimensions[5].hidden = True
		for e1 in range(55, 67):
			amount.cell(row=e1, column=7).value='=C{0}'.format(e1)
		
		for row in amount['B8:B74']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'
		
		for row in amount['C8:C74']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'

		#------foratare D300

		for row in amount['A7:C7']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.alignment=Alignment(horizontal='center',vertical='center')
		for row in amount['D7:G7']:
			for cell in row:
				cell.fill=cap_tabel_color_black
		
		for row in amount['A7:G7']:
			for cell in row:
				cell.font=cap_tabel

		listanoua=['A','B','C','D','E','F','G']

		for column in listanoua:
			for i in listanoua:
				if (column==i):
					amount.column_dimensions[column].width = 17
		
		amount.column_dimensions['D'].hidden = True
		amount.column_dimensions['E'].hidden = True
		amount.column_dimensions['F'].hidden = True
		amount.column_dimensions['G'].hidden = True
		info=temp['Other info']
	
		listaMapare=["L", "T", "S", "A"]
		Poz="10"
		Fix01="01"
		Fix0000="0000"
		dictMapare={'L':'301', 'T':'302', 'S':'303', 'A':'304'}

		elementMapare=""
		if info.cell(row=53, column=3).value=="L":
			elementMapare="301"
		else:
			if info.cell(row=53, column=3).value == "T":
				elementMapare="302"
			else:
				if info.cell(row=53, column=3).value == "S":
					elementMapare="303"
				else:
					if info.cell(row=53, column=3).value == "A":
						elementMapare="304"
			
		# #print(elementMapare, 'MAPARE')

		LL=""
		AA=""

		if len(str(info.cell(row=3, column=3).value))==2:
			LL=str(info.cell(row=3, column=3).value)
			LL_g=LL
		else:
			LL="0"+str(info.cell(row=3, column=3).value)

		AA=str(info.cell(row=2, column=3).value)[2:]
		strYear=str(info.cell(row=2, column=3).value)[2:]
		intYear=int(strYear)+1
		year=str(intYear)
		
		# #print(year, 'an')

		LLAA=LL+AA
		LL2=""
		AA2=""

		if int(info.cell(row=3, column=3).value)==12:
			LL2="1"
		if len(str(info.cell(row=3, column=3).value))==1:
			if(int(info.cell(row=3,column=3).value)==9):
				LL2=str(int(info.cell(row=3,column=3).value)+1)
			else:
				LL2="0"+str(int(info.cell(row=3, column=3).value)+1)
		# #print(LL2, 'LL2')
		
		if int(info.cell(row=3, column=3).value)==12:
			AA2=year 
		else:
			AA2=strYear

		ZZLLAA="25"+str(LL2)+str(AA2)
		# #print(ZZLLAA, 'ZZLLAA')

		poz1=""
		LTSA=""
		fix1=""
		LLAA1=""
		ZZLLAA1=""
		fix01=""
		control=""

		poz1 = sum(int(digit) for digit in str(Poz))
		LTSA = sum(int(digit) for digit in str(elementMapare))
		fix1=sum(int(digit) for digit in str(Fix01))
		LLAA1=sum(int(digit) for digit in str(LLAA))
		ZZLLAA1=sum(int(digit) for digit in str(ZZLLAA))
		fix01=sum(int(digit) for digit in str(Fix0000))

		control=poz1+LTSA+fix1+LLAA1+ZZLLAA1+fix01
		

		nrEvidenta=Poz+elementMapare+Fix01+LLAA+ZZLLAA+Fix0000+str(control)
		#print(nrEvidenta, 'nrEvidenta')
			
		
		info.cell(row=18, column=3).value=nrEvidenta
		info.cell(row=52, column=3).value="=SUM('D300 draft figures'!G8:G66)"
		try:
			info.cell(row=50, column=3).value=lenbfi
			info.cell(row=51, column=3).value=sumabfi
			info.cell(row=52, column=3).value=sumatbfi
		except:
			pass
		try:
			info.cell(row=53, column=3).value=sumabif
			info.cell(row=54, column=3).value=sumatbif
			info.cell(row=55, column=3).value=lentbif
		except:
			pass
		try: 
			# info.cell(row=56, column=3).value=lenbfr
			info.cell(row=57, column=3).value=sumabfr
			info.cell(row=58, column=3).value=sumatbfr
		except:
			pass


		if(val2==1):
			sheetinutil2=temp.create_sheet('D390--->>>')
			sheetinutil2.sheet_view.showGridLines=False
			sheetinutil2.cell(row=2,column=1).value="Switch to next sheet for D390 Workings draft"
			sheetinutil2.cell(row=2,column=1).font=scrisincredibildemare		

			workings=temp.create_sheet('D390 workings')
			workings.cell(row=1,column=1).value="D390 workings"
			workings.cell(row=1,column=1).font=cap_tabelbold
			workings.freeze_panes = 'A4'
			workings.auto_filter.ref = "A3:I10000"
			workings.sheet_view.showGridLines = False
			workings.column_dimensions['I'].hidden = True
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "5":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(sales[cell.column])
			try:
				totals = [b.value for b in sales[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total sales'")
				return render_template("index.html")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "3":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(sales[cell.column])
			try:
				denumires = [b.value for b in sales[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total sales'")
				return render_template("index.html")
			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "4":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(sales[cell.column])
			try:
				vats = [b.value for b in sales[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total sales'")
				return render_template("index.html")			

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "3":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(purchases[cell.column])
			try:
				denumirea = [b.value for b in purchases[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total sales'")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "4":
						rand_tb = cell.row
						tdocc = cell.column
						lun = len(purchases[cell.column])
			try:
				vata = [b.value for b in purchases[tdocc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total sales'")
				return render_template("index.html")				
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "5":
						rand_tb = cell.row
						tdoca = cell.column
						lun = len(purchases[cell.column])
			try:
				totala = [b.value for b in purchases[tdoca][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total purchases'")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "22":
						rand_tb = cell.row
						aiccol = cell.column
						lun = len(purchases[cell.column])
			try:
				aic = [b.value for b in purchases[aiccol][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total purchases'")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "24":
						rand_tb = cell.row
						aiscol = cell.column
						lun = len(purchases[cell.column])
			try:
				ais = [b.value for b in purchases[aiscol][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total purchases'")
				return render_template("index.html")

			workings.cell(row=3, column=1).value='TIP'
			workings.cell(row=3, column=2).value='ŢARA'
			workings.cell(row=3, column=3).value='COD OPERATOR INTRACOMUNITAR'
			workings.cell(row=3, column=4).value='DENUMIRE'
			workings.cell(row=3, column=5).value='BAZA IMPOZABILĂ'
			workings.cell(row=3, column=6).value='CIF'
			workings.cell(row=3, column=7).value='Country Code'
			workings.cell(row=3, column=8).value='BAZA IMPOZABILĂ'
			workings.cell(row=3, column=9).value='Cheie extragere - filtreaza 1'

			# bazaA_Furnizor=get_column_letter(7)
			# bazaA_index=get_column_letter(bazaA)
			# bazaA_literaA=get_column_letter(bazaA)
			# #print(bazaA_Furnizor, "litera pentru furnizor/ client")
			# #print(bazaA_index, "ASTA E NUMARUL LUI A")
			# #print(bazaA_literaA, "LITERA PENTRU TIP")
			# print(taxcodeach)
			a=3
			for jj in range(0, len(denumirea)):
				if(lunacurenta[jj]=="Yes"):
					if aic[jj]!=None :
						# print(aic[x])
						a=a+1
						workings.cell(row=a, column=1).value="A"
						workings.cell(row=a, column=4).value=denumirea[jj]
						workings.cell(row=a, column=6).value=vata[jj]
						workings.cell(row=a, column=3).value=vata[jj][2:]
						workings.cell(row=a, column=7).value=vata[jj][0:2]
						# workings.cell(row=a, column=8).value=listaBazaA[x]
						# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
						workings.cell(row=a, column=8).value=aic[jj]
						workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
						workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)


			for x in range(0, len(denumirea)):
				if(lunacurenta[x]=="Yes"):				
					if ais[x]!=None:
						a=a+1
						workings.cell(row=a, column=1).value="S"
						workings.cell(row=a, column=4).value=denumirea[x]
						workings.cell(row=a, column=6).value=vata[x]
						workings.cell(row=a, column=3).value=vata[x][2:]
						workings.cell(row=a, column=7).value=vata[x][0:2]
						# workings.cell(row=a, column=8).value=listaBazaA[x]
						# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
						workings.cell(row=a, column=8).value=totala[x]
						workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
						workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)

			# for x in range(0, len(taxcodes)):
			# 	if(listacurentas[x]=="Yes"):				
			# 		if str(taxcodes[x])=="Y3" or str(taxcodes[x])=="Y1":
			# 			a=a+1
			# 			workings.cell(row=a, column=1).value="L"
			# 			workings.cell(row=a, column=4).value=denumires[x]
			# 			workings.cell(row=a, column=6).value=vats[x]
			# 			workings.cell(row=a, column=3).value=vats[x][2:]
			# 			workings.cell(row=a, column=7).value=vats[x][0:2]
			# 			# workings.cell(row=a, column=8).value=listaBazaA[x]
			# 			# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
			# 			workings.cell(row=a, column=8).value=totals[x]
			# 			workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
			# 			workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)
			# for x in range(0, len(taxcodes)):
			# 	if(listacurentas[x]=="Yes"):				
			# 		if str(taxcodes[x])=="Y4":
			# 			a=a+1
			# 			workings.cell(row=a, column=1).value="P"
			# 			workings.cell(row=a, column=4).value=denumires[x]
			# 			workings.cell(row=a, column=6).value=vats[x]
			# 			workings.cell(row=a, column=3).value=vats[x][2:]
			# 			workings.cell(row=a, column=7).value=vats[x][0:2]
			# 			# workings.cell(row=a, column=8).value=listaBazaA[x]
			# 			# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
			# 			workings.cell(row=a, column=8).value=totals[x]
			# 			workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
			# 			workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)
			
			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "TIP":
						rand_tb = cell.row
						tip = cell.column
						lun = len(workings[cell.column])
			try:
				listaTip = [b.value for b in workings[tip][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TIP' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "CIF":
						rand_tb = cell.row
						cod = cell.column
						lun = len(workings[cell.column])
			try:
				codPartener = [b.value for b in workings[cod][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'CIF' in Workings sheet")
				return render_template("index.html")


			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "Country Code":
						rand_tb = cell.row
						country = cell.column
						lun = len(workings[cell.column])
			try:
				countryCode = [b.value for b in workings[country][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Country Code' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "DENUMIRE":
						rand_tb = cell.row
						numep = cell.column
						lun = len(workings[cell.column])
			try:
				partnerName = [b.value for b in workings[numep][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'DENUMIRE' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "BAZA IMPOZABILĂ":
						rand_tb = cell.row
						suma = cell.column
						lun = len(workings[cell.column])
			try:
				sumaTot = [b.value for b in workings[suma][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'BAZA IMPOZABILĂ' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "Cheie extragere - filtreaza 1":
						rand_tb = cell.row
						cheie_sort = cell.column
						lun = len(workings[cell.column])
			try:
				cheie = [b.value for b in workings[cheie_sort][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Cheie extragere - filtreaza 1' in Workings sheet")
				return render_template("index.html")

			for row in workings.iter_rows():
				for cell in row:
					if cell.value == "COD OPERATOR INTRACOMUNITAR":
						rand_tb = cell.row
						coi = cell.column
						lun = len(workings[cell.column])
			try:
				listaCOI = [b.value for b in workings[coi][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'BAZA IMPOZABILĂ' in Workings sheet")
				return render_template("index.html")


			nomenclatorTari={'AT':'Austria', 'BE':'Belgia', 'BG':'Bulgaria','CY':'Cipru','DK':'Danemarca','EE':'Estonia', 'FI':'Finlanda','FR':'Franta', 'DE':'Germania','HR':'Croatia',
							'GR':'Grecia','IE':'Irlanda','IT':'Italia','LV':'Letonia','LT':'Lituania','LU':'Luxemburg','MT':'Malta','XI':'Irlanda de Nord - Regatul Unit','NL':'Olanda',
							'PL':'Polonia','PT':'Portugalia','CZ':'Republica Ceha','RO':'Romania','SK':'Slovacia','SI':'Slovavia','ES':'Spania','SE':'Suedia','HU':'Ungaria'}
			b=3
			for i in countryCode:
				if i in nomenclatorTari:
					b=b+1
					workings.cell(row=b, column=2).value=nomenclatorTari[i]

			for row in workings['A3:I3']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')
			for row in workings['E4:E10000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'
			
			for row in workings['H4:H10000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'

			workings.column_dimensions['B'].width=20
			workings.column_dimensions['D'].width=35				

			forxml = temp.create_sheet('D390 for XML')
			forxml.cell(row=1,column=1).value="D390 for XML"
			forxml.cell(row=1,column=1).font=cap_tabelbold
			forxml.freeze_panes = 'A4'
			forxml.auto_filter.ref = "A3:F10000"
			forxml.sheet_view.showGridLines = False

			forxml.cell(row=3, column=1).value="III.B"
			forxml.cell(row=3, column=2).value="TIP"
			forxml.cell(row=3, column=3).value="ŢARA"
			forxml.cell(row=3, column=4).value="COD OPERATOR INTRACOMUNITAR"
			forxml.cell(row=3, column=5).value="Denumire"
			forxml.cell(row=3, column=6).value="BAZA IMPOZABILĂ"

			codeAndType=[]
			codeAndName=[]
			typeAndName=[]
			# typeCodeName=[]

			# for k in range(0,len(listaTip)):
			# 	codeAndType.append(str(listaTip[k])+" "+str(codPartener[k]))
			# 	codeAndName.append(str(listaTip[k])+" "+str(partnerName[k]))

			# #print(codeAndType,'codeandtyp')
			# codeAndTypeUnique=list(set(codeAndType))
			# codeAndNameUnique=list(set(codeAndName))

			for i in range(0, len(listaTip)):
				typeAndName.append(str(listaTip[i])+";;;"+str(partnerName[i])+";;;"+str(countryCode[i])+";;;"+str(listaCOI[i]))
			#print(typeAndName, 'TYPEAndNAME')

			typeAndNameUni=list(set(typeAndName))

			typeAndNameUni=list(set(typeAndName))

			for i in it.chain(range(0, len(typeAndNameUni))):
				x=typeAndNameUni[i].split(";;;")
				forxml.cell(row=4+i, column=2).value=str(x[0])
				forxml.cell(row=4+i, column=3).value=str(x[2])
				forxml.cell(row=4+i, column=4).value=str(x[3])
				forxml.cell(row=4+i, column=5).value=str(x[1])
				forxml.cell(row=4+i, column=6).value="=SUMIFS('D390 workings'!H:H,'D390 workings'!A:A,B{0},'D390 workings'!C:C,D{0},'D390 workings'!G:G,C{0})".format(4+i)


			for row in forxml['A3:F3']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in forxml['F4:F10000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'

			forxml.column_dimensions['D'].width=27
			forxml.column_dimensions['E'].width=35
			forxml.column_dimensions['F'].width=15


	#---------------------------NR DE EVIDENTA
		if(val3==1):
			sheetinutil3=temp.create_sheet('D394--->>>')
			sheetinutil3.sheet_view.showGridLines=False
			sheetinutil3.cell(row=2,column=1).value="Switch to next sheet for D394 Workings draft"
			sheetinutil3.cell(row=2,column=1).font=scrisincredibildemare		
			nomenclatorTari={'AT':'Austrie', 'BE':'Belgia', 'BG':'Bulgaria','CY':'Cipru','DK':'Danemarca','EE':'Estonia', 'FI':'Finlanda','FR':'Franta', 'DE':'Germania','HR':'Croatia',
							'GR':'Grecia','IE':'Irlanda','IT':'Italia','LV':'Letonia','LT':'Lituania','LU':'Luxemburg','MT':'Malta','XI':'Irlanda de Nord - Regatul Unit','NL':'Olanda',
							'PL':'Polonia','PT':'Portugalia','CZ':'Republica Ceha','RO':'Romania','SK':'Slovacia','SI':'Slovavia','ES':'Spania','SE':'Suedia','HU':'Ungaria'}


			salesExcel=temp.create_sheet("Mapping tranzactii")
			salesExcel.sheet_view.showGridLines = False
			salesExcel.cell(row=2,column=1).value="Mapping tranzactii"
			salesExcel.cell(row=2,column=1).font=cap_tabeltitlu	
			salesExcel.freeze_panes = 'A10'
			
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value=="Declarat anterior":
						rand_tb = cell.row
						declarateanteriorp = cell.column
						lun = len(purchases[cell.column])
			try:
				listadeclantp = [b.value for b in purchases[declarateanteriorp][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Declarat anterior' in Purchases sheet")
				return render_template("index.html")		
			# except:
			# 	listadeclant=[]
			listadeclantp_1=[]
			# #print(listadeclantp,"---------")
			for c in range(0, len(listadeclantp)):
				if listadeclantp[c] == None:
					listadeclantp_1.append("No")
				else:
					listadeclantp_1.append(listadeclantp[c])
			# #print("-----",listadeclantp_1,"------")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value=="Declarat anterior":
						rand_tb = cell.row
						declarateanterior = cell.column
						lun = len(sales[cell.column])
			try:
				listadeclant = [b.value for b in sales[declarateanterior][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Declarat anterior' in Sales sheet")
				return render_template("index.html")
			# except:
			# 	listadeclant=[]
			listadeclant_1=[]
			# #print(len(listadeclant))
			for c in range(0, len(listadeclant)):
				if listadeclant[c] == None:
					listadeclant_1.append("No")
				else:
					listadeclant_1.append(listadeclant[c])

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "3":
						rand_tb = cell.row
						clientCell = cell.column
						lun = len(sales[cell.column])
			try:
				listaClient = [b.value for b in sales[clientCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Denumirea clientului Client name' in Sales sheet")
				return render_template("index.html")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "4":
						rand_tb = cell.row
						coloanaClientID = cell.column
						lun = len(sales[cell.column])
			try:
				listaCUISales = [b.value for b in sales[coloanaClientID][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Codul de inregistrare in scopuri de TVA al clientului Client VAT ID' in Sales sheet")
				return render_template("index.html")

			listaCUISales1=[]
			# listadeclant_1=[]
			for val in listaCUISales:
				if val != None:
					# listadeclant_1.append("")
					listaCUISales1.append(val)
				else:
					listaCUISales1.append("US111")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "1":
						rand_tb = cell.row
						docNumber = cell.column
						lun = len(sales[cell.column])
			try:
				docNoSales = [b.value for b in sales[docNumber][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Nr/ document Document no' in Sales sheet")
				return render_template("index.html")




			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "7.3":
						rand_tb = cell.row
						taxBaseL19 = cell.column
						lun = len(sales[cell.column])
			try:
				taxBaseL19 = [b.value for b in sales[taxBaseL19][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)- Livrari locale-Taxable base (RON)-Local supplies (19%)' in Sales sheet")
				return render_template("index.html")
			#print(taxBaseL19)

			taxBaseL19_1=[]
			for c in range(0, len(taxBaseL19)):
				if taxBaseL19[c] == None:
					taxBaseL19_1.append(0)
				else:
					taxBaseL19_1.append(taxBaseL19[c])

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "7.4":
						rand_tb = cell.row
						vatBaseL19 = cell.column
						lun = len(sales[cell.column])
			try:
				vatL19 = [b.value for b in sales[vatBaseL19][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-Livrari locale-VAT (RON)-Local supplies (19%)' in Sales sheet")
				return render_template("index.html")

			vatL19_1=[]
			for c in range(0, len(vatL19)):
				if vatL19[c] == None:
					vatL19_1.append(0)
				else:
					vatL19_1.append(vatL19[c])



			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "22":
						rand_tb = cell.row
						taxBV = cell.column
						lun = len(sales[cell.column])
			try:
				taxBaseInv = [b.value for b in sales[taxBV][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Bază de impozitare (RON)-Livrari locale taxare inversa-Taxable base (RON)-Local supplies reverse charge' in Sales sheet")
				return render_template("index.html")

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "23":
						rand_tb = cell.row
						taxBV = cell.column
						lun = len(sales[cell.column])
			try:
				taxAmountInv = [b.value for b in sales[taxBV][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Bază de impozitare (RON)-Livrari locale taxare inversa-Taxable base (RON)-Local supplies reverse charge' in Sales sheet")
				return render_template("index.html")


			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "5":
						rand_tb = cell.row
						totdoc = cell.column
						lun = len(sales[cell.column])
			totdocuments = [b.value for b in sales[totdoc][rand_tb:lun+1]]
			taxBaseV_1=[]
			# #print(taxBaseV_1)

			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "24":
						rand_tb = cell.row
						taxBi = cell.column
						lun = len(sales[cell.column])
			try:
				taxBaseserv = [b.value for b in sales[taxBi][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services' in Sales sheet")
				return render_template("index.html")


			for row in sales.iter_rows():
				for cell in row:
					if cell.value == "25":
						rand_tb = cell.row
						taxBiSc = cell.column
						lun = len(sales[cell.column])
			try:
				taxamountserv = [b.value for b in sales[taxBiSc][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services' in Sales sheet")
				return render_template("index.html")


			serieCuiSales=[]
			codTaraCuiSales=[]
			for i in listaCUISales1:
				if(str(i)[:1].isalpha()):
					r = re.compile("([a-zA-Z]+)([0-9]+)")
					m = r.match(str(i))
					try:
						serieCuiSales.append(m.group(2))
						codTaraCuiSales.append(m.group(1))

					except:
						codTaraCuiSales.append(None)
						serieCuiSales.append(str(i))	

				else:
					codTaraCuiSales.append(None)
					serieCuiSales.append(str(i))
				# tara,oras=i.split(',',1)
				# serieCuiSales.append(oras)
				# codTaraCuiSales.append(tara)
			# #print(codTaraCuiSales)
			# TIP Furnizor!!!!!
			# #print(len(codTaraCuiSales))
			print(serieCuiSales,codTaraCuiSales)
			coteTVAsales=[]
			for i in range(0, len(docNoSales)):
				if taxBaseL19[i]!=None:
					coteTVAsales.append(19)
				else:
					if taxBaseInv[i]!=None or taxBaseserv[i]!=None:
						coteTVAsales.append(19)
					else:
						coteTVAsales.append(0)

			codTranzactieSales=[]
			for i in range(0, len(codTaraCuiSales)):
				if str(serieCuiSales[i])[1:2].isalpha():
					codTranzactieSales.append(2)
				else:
					if codTaraCuiSales[i] == "RO":
						# #print("RO")
						codTranzactieSales.append(1)
					else:	
						if codTaraCuiSales[i] in nomenclatorTari:
							# #print("UE")
							codTranzactieSales.append(3)
						else:
							# #print("nonUE")
							codTranzactieSales.append(4)
			#Cote TVAA




			#TIP TRANZACTIE
			storno=[]
			tipTranzSale = []
			# #print(docNoSales)
			# #print(len(docNoSales),len(codTranzactieSales))
			# print(len(docNoSales),len(codTranzactieSales))
			# print(codTranzactieSales)
			for i in range(0, len(docNoSales)):
				# print(docNoSales[i],print(codTranzactieSales[i]))
				if(listadeclant_1[i]=="Yes"):
					tipTranzSale.append("Declarat anterior")
				else:
					if int(codTranzactieSales[i]) == 1:
						# #print(docNoSales[i]," ",listaCUISales1[i], "", taxBaseL19_1[i], " ", taxBaseL9_1[i], " ", taxBaseL5_1[i])
						# if (int(taxBaseL19_1[i])>0 and int(vatL19_1[i])>0) or (int(taxBaseL9_1[i])>0 and int(vatL9_1[i])> 0) or (int(taxBaseL5_1[i])>0 and int(vatL5_1[i])>0):
						if (taxBaseL19[i]!=None):
							print("Yes")
							tipTranzSale.append('L')
							storno.append("")
						else:
							# None
							if (taxBaseInv[i]!=None or taxBaseserv[i]!=None):
								tipTranzSale.append('Not applicable for D394')
					else:
						if int(codTranzactieSales[i]) == 2:
							if (taxBaseL19[i]!=None):
								tipTranzSale.append('L')
								storno.append("")
							else:
								tipTranzSale.append("Not applicable for D394")
						else:
							if int(codTranzactieSales[i]) == 3:
								if (taxBaseL19[i]!=None):
									print("Yes")
									tipTranzSale.append('L')
									storno.append("")
								else:
									# None
									if (taxBaseInv[i]!=None or taxBaseserv[i]!=None):
										tipTranzSale.append('Not applicable for D394')
								
							else:
								if int(codTranzactieSales[i]) == 4:
									if (taxBaseL19[i]!=None):
										tipTranzSale.append('L')
										storno.append("")
									else:
										tipTranzSale.append("Not applicable for D394")

			# #print(docNoSales)
			#Scriere in excel

			salesExcel.cell(row=9, column=1).value = "Cod tara"
			salesExcel.cell(row=9, column=2).value = "Serie cui"
			salesExcel.cell(row=9, column=3).value = "Numar document"
			salesExcel.cell(row=9, column=4).value = "CUI"
			salesExcel.cell(row=9, column=5).value = "Clasa tranzactie"
			salesExcel.cell(row=9, column=6).value = "Tip tranzactie"
			salesExcel.cell(row=9, column=7).value = "Cota TVA"
			salesExcel.cell(row=9, column=8).value = "Total document"
			salesExcel.cell(row=9, column=9).value = "Tip jurnal"
			salesExcel.cell(row=9, column=10).value = "Nume partener"
			salesExcel.cell(row=9, column=11).value = "Check"
			salesExcel.cell(row=9, column=12).value = "Cod si denumire NC produs(TIP V)"

			# dv = DataValidation(
				# type='list', formula1='"Yes,No"', allow_blank=True)


			listahelp=["1002--Secara","1003--Orz","1005--Porumb","1201--Boabe de soia"," 1205--Seminte de rapita sau de rapita salbatica","120600--Seminte de floarea soarelui","121291--Sfecla de zahar","1001-Grau si meslin","1004--Ovaz","10086000--Triticale","22-deseuri feroase si neferoase","23-masa lemnoasa","32-terenuri","33-constructii","34-alte bunuri","35-servicii","24-certificate de emisii de gaze cu efect de sera","25-energie electrica","26-certificate verzi","27-constructii/terenuri","28-aur de investitii","29-telefoane mobile","30-microprocesoare","31-console de jocuri tablete PC si laptopuri"]
			sheethelp=temp.create_sheet("Validation")
			sheethelp.sheet_state = 'hidden'

			# dv = DataValidation(
			# 	type="list", formula1="", allow_blank=True)
			# salesExcel.add_data_validation(dv)

			# dv.add(salesExcel["L2"])



			print(len(tipTranzSale),len(codTranzactieSales))
			for i in range(0, len(codTaraCuiSales)):
				print(serieCuiSales[i],docNoSales[i],listaCUISales1[i],codTranzactieSales[i],tipTranzSale[i])
				salesExcel.cell(row=10 + i, column=1).value = codTaraCuiSales[i]
				salesExcel.cell(row=10 + i, column=2).value = serieCuiSales[i]
				salesExcel.cell(row=10 + i, column=3).value = docNoSales[i]
				salesExcel.cell(row=10 + i, column=4).value = listaCUISales1[i]
				salesExcel.cell(row=10 + i, column=5).value = codTranzactieSales[i]
				# if(listadeclant_1[i]!=""):
				salesExcel.cell(row=10 + i, column=6).value = tipTranzSale[i]
				# else:
					# salesExcel.cell(row=10 + i, column=6).value = listadeclant_1[i]

				if(tipTranzSale[i]=='V'):
					salesExcel.cell(row=10+i,column=12).value="V"
				else:
					salesExcel.cell(row=10+i,column=12).value="N/A, valid only for V trans."

				salesExcel.cell(row=10 + i, column=8).value = totdocuments[i]
				salesExcel.cell(row=10 + i, column=9).value = "Jurnal vanzari"
				salesExcel.cell(row=10 + i, column=10).value = listaClient[i]

			for i in range(0, len(coteTVAsales)):
				salesExcel.cell(row=10 + i, column=7).value = coteTVAsales[i]
				salesExcel.cell(row=10+i,column=18).value="=B{0}&E{0}&F{0}&G{0}".format(i+10)

			#FORMATARE------------------------------------------------------------------
			red_color = 'ffc7ce'
			green_color='99ff99'
			red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
			green_fill = styles.PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
			row=salesExcel.max_row	
			salesExcel.conditional_formatting.add('K10:K'+str(row-1), formatting.rule.CellIsRule(operator='notEqual', formula=['"OK"'], fill=red_fill))
			for row in salesExcel['A9:L9']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in salesExcel['A9:L9']:
				for cell in row:
					cell.font = cap_tabel

			# for row in salesExcel['A9:K9']:
			# 	for cell in row:
			# 		cell.border = border_thin

			
			salesExcel.freeze_panes = 'A10'

			salesExcel.column_dimensions['B'].width = 20
			salesExcel.column_dimensions['C'].width = 20
			salesExcel.column_dimensions['D'].width = 20
			salesExcel.column_dimensions['E'].width = 16
			salesExcel.column_dimensions['F'].width = 16
			salesExcel.column_dimensions['F'].width = 20		
			salesExcel.column_dimensions['H'].width = 14
			salesExcel.column_dimensions['J'].width = 35
			salesExcel.column_dimensions['L'].width = 35		
			purchases = temp['Purchases']

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "3":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				supplierName = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Furnizor Supplier' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "2":
						rand_tb = cell.row
						supplierCell = cell.column
						lun = len(purchases[cell.column])
			try:
				datadocument = [b.value for b in purchases[supplierCell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for ' Doc. Date' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "6":
						rand_tb = cell.row
						vatCashinSys = cell.column
						lun = len(purchases[cell.column])
			try:
				vatApplies = [b.value for b in purchases[vatCashinSys][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Aplica TVA la incasare VAT cash-in system' in Purchases sheet")
				return render_template("index.html")


			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "4":
						rand_tb = cell.row
						suppID = cell.column
						lun = len(purchases[cell.column])
			try:
				suppIDPurch = [b.value for b in purchases[suppID][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Codul de înregistrare în scopuri de TVA VAT number' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "1":
						rand_tb = cell.row
						docNumberPurch = cell.column
						lun = len(purchases[cell.column])
			try:
				docNoPurch = [b.value for b in purchases[docNumberPurch][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Nr. document Document no' in Purchases sheet")
				return render_template("index.html")
			
			docNoPurch1 = []
			for val in docNoPurch:

				docNoPurch1.append(val)
			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "7.3":
						rand_tb = cell.row
						taxBaseAch19cell = cell.column
						lun = len(purchases[cell.column])
			try:
				taxBaseAch19 = [b.value for b in purchases[taxBaseAch19cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON) -achizitii locale-Taxable base (RON) - local acquisition (19%)' in Purchases sheet")
				return render_template("index.html")

			# #print(taxBaseAch19,taxBaseAch19_1)
			# #print(taxBaseAch19_1)
			# #print(taxBaseAch19)
			# for item in taxBaseAch19:
			#    #print(type(item))
			# #print(type(taxBaseAch19))

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "7.4":
						rand_tb = cell.row
						vatAch19cell = cell.column
						lun = len(purchases[cell.column])
			try:
				vatAch19 = [b.value for b in purchases[vatAch19cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-achizitii locale-VAT (RON)-local acquisition (19%)' in Purchases sheet")
				return render_template("index.html")


			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "13.3":
						rand_tb = cell.row
						taxBaseAch19cell = cell.column
						lun = len(purchases[cell.column])
			try:
				taxBaseneexAch19 = [b.value for b in purchases[taxBaseAch19cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON) -achizitii locale-Taxable base (RON) - local acquisition (19%)' in Purchases sheet")
				return render_template("index.html")


			# #print(taxBaseAch19,taxBaseAch19_1)
			# #print(taxBaseAch19_1)
			# #print(taxBaseAch19)
			# for item in taxBaseAch19:
			#    #print(type(item))
			# #print(type(taxBaseAch19))

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "13.4":
						rand_tb = cell.row
						vatAch19cell = cell.column
						lun = len(purchases[cell.column])
			try:
				vatnexAch19 = [b.value for b in purchases[vatAch19cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-achizitii locale-VAT (RON)-local acquisition (19%)' in Purchases sheet")
				return render_template("index.html")



			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "5":
						rand_tb = cell.row
						totdocp = cell.column
						lun = len(purchases[cell.column])
			try:
				totdocumentp = [b.value for b in purchases[totdocp][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Total document (inclusiv TVA)-RON' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "12":
						rand_tb = cell.row
						taxBaseAch5cell = cell.column
						lun = len(purchases[cell.column])
			try:
				taxBaseAch5 = [b.value for b in purchases[taxBaseAch5cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Baza de impozitare (RON)-achizitii locale Taxable base (RON) local acquisition (5%)' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "13":
						rand_tb = cell.row
						vatAch5cell = cell.column
						lun = len(purchases[cell.column])
			try:
				vatAch5 = [b.value for b in purchases[vatAch5cell][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'TVA (RON)-achizitii locale-VAT (RON)-local acquisition (5%)' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "22":
						rand_tb = cell.row
						vatExemptLocAcq = cell.column
						lun = len(purchases[cell.column])
			try:
				taxinversbunuri = [b.value for b in purchases[vatExemptLocAcq][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "23":
						rand_tb = cell.row
						vatExemptLocAcq = cell.column
						lun = len(purchases[cell.column])
			try:
				taxvatinversbunuri = [b.value for b in purchases[vatExemptLocAcq][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
				return render_template("index.html")


			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "24":
						rand_tb = cell.row
						vatExemptLocAcq = cell.column
						lun = len(purchases[cell.column])
			try:
				taxinversserv = [b.value for b in purchases[vatExemptLocAcq][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "25":
						rand_tb = cell.row
						vatExemptLocAcq = cell.column
						lun = len(purchases[cell.column])
			try:
				taxvatinversserv = [b.value for b in purchases[vatExemptLocAcq][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "27":
						rand_tb = cell.row
						vatExemptLocAcq = cell.column
						lun = len(purchases[cell.column])
			try:
				scutite = [b.value for b in purchases[vatExemptLocAcq][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "28":
						rand_tb = cell.row
						vatExemptLocAcq = cell.column
						lun = len(purchases[cell.column])
			try:
				neimpoz = [b.value for b in purchases[vatExemptLocAcq][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
				return render_template("index.html")

			for row in purchases.iter_rows():
				for cell in row:
					if cell.value == "34":
						rand_tb = cell.row
						vatExemptLocAcq = cell.column
						lun = len(purchases[cell.column])
			try:
				tvaneded = [b.value for b in purchases[vatExemptLocAcq][rand_tb:lun]]
			except:
				flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
				return render_template("index.html")


			# #print(revVatAch19_1)

			# for row in purchases.iter_rows():
			# 	for cell in row:
			# 		if cell.value == "Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge (9%)":
			# 			rand_tb = cell.row
			# 			revTaxBaseAch9cell = cell.column
			# 			lun = len(purchases[cell.column])
			# try:
			# 	revTaxBaseAch9 = [b.value for b in purchases[revTaxBaseAch9cell][rand_tb:lun]]
			# except:
			# 	flash("Please insert the correct header for 'Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge (9%)' in Purchases sheet")
			# 	return render_template("index.html")

			# revTaxBaseAch9_1=[]
			# for i in range(0, len(revTaxBaseAch9)):
			# 	if revTaxBaseAch9[i] == None:
			# 		revTaxBaseAch9_1.append(0)
			# 	else:
			# 		revTaxBaseAch9_1.append(revTaxBaseAch9[i])
			# # #print(revTaxBaseAch9_1)

			# for row in purchases.iter_rows():
			# 	for cell in row:
			# 		if cell.value == "TVA (RON)-Achizitii locale taxare inversa-VAT (RON)-Local acquisition reverse charge (9%)":
			# 			rand_tb = cell.row
			# 			revVatAch9cell = cell.column
			# 			lun = len(purchases[cell.column])
			# try:
			# 	revVatAch9 = [b.value for b in purchases[revVatAch9cell][rand_tb:lun]]
			# except:
			# 	flash("Please insert the correct header for 'TVA (RON)-Achizitii locale taxare inversa-VAT (RON)-Local acquisition reverse charge (9%)' in Purchases sheet")
			# 	return render_template("index.html")

			# revVatAch9_1=[]
			# for i in range(0, len(revVatAch9)):
			# 	if revVatAch9[i] == None:
			# 		revVatAch9_1.append(0)
			# 	else:
			# 		revVatAch9_1.append(revVatAch9[i])
			# # #print(revVatAch9_1)

			# for row in purchases.iter_rows():
			# 	for cell in row:
			# 		if cell.value == "Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge(5%)":
			# 			rand_tb = cell.row
			# 			revTaxBaseAch5cell = cell.column
			# 			lun = len(purchases[cell.column])
			# try:
			# 	revTaxBaseAch5 = [b.value for b in purchases[revTaxBaseAch5cell][rand_tb:lun]]
			# except:
			# 	flash("Please insert the correct header for 'Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge(5%)' in Purchases sheet")
			# 	return render_template("index.html")

			# revTaxBaseAch5_1=[]
			# for i in range(0, len(revTaxBaseAch5)):
			# 	if revTaxBaseAch5[i] == None:
			# 		revTaxBaseAch5_1.append(0)
			# 	else:
			# 		revTaxBaseAch5_1.append(revTaxBaseAch5[i])
			# # #print(revTaxBaseAch5_1)

			# for row in purchases.iter_rows():
			# 	for cell in row:
			# 		if cell.value == "TVA (RON)-Achizitii locale taxare inversa VAT (RON)-Local acquisition reverse charge (5%)":
			# 			rand_tb = cell.row
			# 			revVatAch5cell = cell.column
			# 			lun = len(purchases[cell.column])
			# try:
			# 	revVatAch5 = [b.value for b in purchases[revVatAch5cell][rand_tb:lun]]
			# except:
			# 	flash("Please insert the correct header for 'TVA (RON)-Achizitii locale taxare inversa VAT (RON)-Local acquisition reverse charge (5%)' in Purchases sheet")
			# 	return render_template("index.html")

			# revVatAch5_1=[]
			# for i in range(0, len(revVatAch5)):
			# 	if revVatAch5[i] == None:
			# 		revVatAch5_1.append(0)
			# 	else:
			# 		revVatAch5_1.append(revVatAch5[i])
			# # #print(revVatAch5_1)


			# #print(datadocument)
			# for row in purchases.iter_rows():
			# 	for cell in row:
			# 		if cell.value == "TVA (RON)-bunuri nonUE-VAT (RON)-nonUE goods":
			# 			rand_tb = cell.row
			# 			vatNonUEGoodscell = cell.column
			# 			lun = len(purchases[cell.column])
			# try:
			# 	vatNonUEGoods = [b.value for b in purchases[vatNonUEGoodscell][rand_tb:lun]]
			# except:
			# 	flash("Please insert the correct header for 'TVA (RON)-bunuri nonUE-VAT (RON)-nonUE goods' in Purchases sheet")
			# 	return render_template("index.html")

			# vatNonUEGoods_1=[]
			# for i in range(0, len(vatNonUEGoods)):
			# 	if vatNonUEGoods[i] == None:
			# 		vatNonUEGoods_1.append(0)
			# 	else:
			# 		vatNonUEGoods_1.append(vatNonUEGoods[i])

			serieCuiPurch = []
			codTaraCuiPurch = []
			# #print(suppIDPurch)
			for i in suppIDPurch:
				# #print(i)
				if(str(i)[:1].isalpha()):	
					r = re.compile("([a-zA-Z]+)([0-9]+)")
					m = r.match(str(i))
					try:
						serieCuiPurch.append(m.group(2))
					except:
						serieCuiPurch.append(" ")
					try:
						codTaraCuiPurch.append(m.group(1))
					except:
						codTaraCuiPurch.append(" ")
				else:
					codTaraCuiPurch.append(None)
					serieCuiPurch.append(i)
			# #print(codTaraCuiPurch,serieCuiPurch)
			tipTranzactiePurchases=[]
			#print("Aici vat -------",vatAch19_1,"-----Achizitii")
			#Tip furnizor
			for i in range(0, len(codTaraCuiPurch)):
				if codTaraCuiPurch[i] == "RO":
					# #print("RO")
					tipTranzactiePurchases.append(1)
				else:
					if codTaraCuiPurch[i]==None:
						tipTranzactiePurchases.append(2)
					else:
						if codTaraCuiPurch[i] in nomenclatorTari:
							# #print("UE")
							tipTranzactiePurchases.append(3)
						else:
							# #print("nonUE")
							tipTranzactiePurchases.append(4)

			#Cote TVA
			coteTVApurchases=[]
			for i in range(0, len(docNoPurch1)):
				if (taxBaseAch19[i]!= None  or taxBaseneexAch19[i]!= None or taxinversbunuri[i]!= None or taxinversserv[i]!=None):
					coteTVApurchases.append('19')
				else:
					if (taxBaseAch5[i]!=None):
						coteTVApurchases.append('5')
					else:
						coteTVApurchases.append('0')

			#mapare tip tranzactie
			# #print(len(docNoPurch1),len(tipTranzactiePurchases),len(vatApplies))
			# #print(vatApplies)
			tipTranzPurch=[]

			# for i in range(0,len(suppIDPurch)):
				# #print(suppIDPurch[i],tipTranzactiePurchases[i])
			# #print(len(docNoPurch),len(tipTranzactiePurchases))
			# print(len(docNoPurch1),len(listadeclantp_1),"--------------len de lista")
			for i in range(0, len(docNoPurch1)):
				if(listadeclantp_1[i]=="Yes"):
					tipTranzPurch.append("Declarat anterior")
				else:
					# #print(docNoPurch1[i])
					if int(tipTranzactiePurchases[i]) == 1:
						if(vatApplies[i]!=None):
							if (taxBaseAch19[i]!=None or taxBaseneexAch19[i]!=None):
								tipTranzPurch.append('AI')
						else:
							if (taxBaseAch19[i]!=None or taxBaseneexAch19[i]!=None):
								tipTranzPurch.append('A')
							else:
								tipTranzPurch.append('Not applicable for D394')
					else:
						if int(tipTranzactiePurchases[i]) == 2:
							if neimpoz[i]!= None or scutite[i]!= None:
								tipTranzPurch.append("N")
							else:
								if(taxBaseAch19[i]!=None):
									tipTranzPurch.append("A")

								#print(docNoPurch1[i],";;;;es 9")
						else:
							if int(tipTranzactiePurchases[i]) == 3:
								if (taxBaseAch19[i]!=None or taxBaseneexAch19[i]!=None):
									tipTranzPurch.append('A')
								else:
									tipTranzPurch.append("Not applicable for D394")
											#print(docNoPurch1[i],";;;;;es 12")
							
							else:
									if int(tipTranzactiePurchases[i]) == 4:
										if (taxBaseAch19[i]!=None or taxBaseneexAch19[i]!=None):
											tipTranzPurch.append('A')
											#print(docNoPurch1[i],";;;;;es 12")
										else:
											tipTranzPurch.append("Not applicable for D394")
										#print("Yes 16")
				# #print(docNoPurch1[i],tipTranzPurch[i],docNoPurch[i+1])
			ma=salesExcel.max_row+1
			for i in range(0, len(codTaraCuiPurch)):
				salesExcel.cell(row=ma + i, column=1).value = codTaraCuiPurch[i]

			for i in range(0, len(serieCuiPurch)):
				salesExcel.cell(row=ma + i, column=2).value = serieCuiPurch[i]

			for i in range(0, len(docNoPurch1)):
				salesExcel.cell(row=ma+ i, column=3).value = docNoPurch1[i]

			for i in range(0, len(suppIDPurch)):
				salesExcel.cell(row=ma + i, column=4).value = suppIDPurch[i]

			for i in range(0, len(tipTranzactiePurchases)):
				salesExcel.cell(row=ma+ i, column=5).value = tipTranzactiePurchases[i]

			for i in range(0, len(tipTranzPurch)):
				# if(listadeclantp_1!=""):
				salesExcel.cell(row=ma+ i, column=6).value = tipTranzPurch[i]
				# else:
					# salesExcel.cell(row=ma+ i, column=6).value = "Declarate anterior"
				if(tipTranzPurch[i]=="V"):
					salesExcel.cell(row=ma+i,column=12).value="Add type of tranzactie"
				else:
					salesExcel.cell(row=ma+i,column=12).value="N/A"

			for i in range(0, len(coteTVApurchases)):
				#print(coteTVApurchases[i])
				salesExcel.cell(row=ma+ i, column=7).value = coteTVApurchases[i]
				salesExcel.cell(row=ma+ i, column=8).value = totdocumentp[i]
				salesExcel.cell(row=ma+ i, column=9).value = "Jurnal cumparari"
				salesExcel.cell(row=ma+ i, column=10).value = supplierName[i]

			codTaraCUItotal=codTaraCuiPurch+codTaraCuiSales
			for i in range(0, len(codTaraCUItotal)):
				salesExcel.cell(row=10 + i, column=11).value = '=IFERROR(IF(VLOOKUP(B{0}&E{0}&F{0}&G{0},Tranzactii!K:K,1,0)=B{0}&E{0}&F{0}&G{0},"OK","Mapped missing in Tranzactii sheet"),"Mapped missing inTranzactiisheet")'.format(10+i)
			salesExcel.auto_filter.ref = "A9:L9"
			for row in salesExcel['H10:F1000']:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'		
			tranzactii=temp.create_sheet("Tranzactii")
			tranzactii.freeze_panes = 'A6'
			tranzactii.sheet_view.showGridLines = False

													
			setSalesCUI=set(listaCUISales1)
			idSalesUnique=list(setSalesCUI)

			setPurchCUI=set(suppIDPurch)
			idPurchUnique=list(setPurchCUI)

			listaCUIUnique=idSalesUnique+idPurchUnique

			setlistaClient=set(listaClient)
			listaClientUnique=list(setlistaClient)

			setSupplierName=set(supplierName)
			supplierNameUnique=list(setSupplierName)

			# for k in range(0,len(setlistaClient)):
			#    count
			#    for j in range(0,len(listaClient)):


			print(len(supplierName),len(tipTranzPurch))
			listanouaappendpurch=[]
			# for i in range(0,len(supplierName)):
			for p in range(0,len(serieCuiPurch)):
				print(serieCuiPurch[p],tipTranzPurch[p],coteTVApurchases[p],tipTranzactiePurchases[p])
			# print()
			print(len(serieCuiPurch),len(tipTranzPurch),len(coteTVApurchases),len(tipTranzactiePurchases))
			for k in range(0,len(serieCuiPurch)):
				try:
					print(serieCuiPurch[k],tipTranzPurch[k])
				except:
					print(serieCuiPurch[k])
				listanouaappendpurch.append(str(serieCuiPurch[k])+";"+str(tipTranzPurch[k])+";"+str(coteTVApurchases[k])+";"+str(tipTranzactiePurchases[k])+";"+str(listadeclantp_1[k]))

			listanouasetpurch=list(set(listanouaappendpurch))


			listanouaappendsales=[]

			for k in range(0,len(serieCuiSales)):
				listanouaappendsales.append(str(serieCuiSales[k])+";"+str(tipTranzSale[k])+";"+str(coteTVAsales[k])+";"+str(codTranzactieSales[k])+";"+listadeclant_1[k])

			listanouasetsales=list(set(listanouaappendsales))

			countsales=[]
			for p in range(0,len(listanouasetsales)):
				count=0
				for k in range(0,len(listanouaappendsales)):
					if(listanouaappendsales[k]==listanouasetsales[p]):
						count=count+1
				countsales.append(count)
			countpurch=[]
			for p in range(0,len(listanouasetpurch)):
				count=0
				for k in range(0,len(listanouaappendpurch)):
					if(listanouaappendpurch[k]==listanouasetpurch[p]):
						count=count+1
				countpurch.append(count)


			#print(listanouasetsales)

			tranzactii.cell(row=5,column=1).value="Cui partener"
			tranzactii.cell(row=5,column=2).value="Nume partener"
			tranzactii.cell(row=5,column=3).value="Tip partener"
			tranzactii.cell(row=5,column=4).value="Tip tranzactie"
			tranzactii.cell(row=5,column=5).value="Cota TVA"
			tranzactii.cell(row=5,column=6).value="Baza TVA"
			tranzactii.cell(row=5,column=7).value="TVA"
			tranzactii.cell(row=5,column=8).value="Nr Facturi"
			tranzactii.cell(row=5,column=9).value="Neexigibile - nu se vor raporta"
			tranzactii.cell(row=5,column=10).value="Cod si denumire NC produs(TIP V)"
			tranzactii.cell(row=2,column=1).value="Tranzactii"
			tranzactii.cell(row=2,column=1).font=cap_tabeltitlu
			counts=0
			for i in range(0, len(listanouasetsales)):
				x=listanouasetsales[i].split(";")
				#print(x[3],x[4])
				if(int(x[3])<3 or int(x[2])>0):
					counts=counts+1
					y=tranzactii.max_row
					tranzactii.cell(row=y+1,column=1).value=x[0]
					tranzactii.cell(row=y+1,column=2).value="=VLOOKUP(A"+str(y+1)+",'Mapping tranzactii'!B:J,9,0)"
					tranzactii.cell(row=y+1,column=3).value=x[3]
					tranzactii.cell(row=y+1,column=4).value=x[1]
					tranzactii.cell(row=y+1,column=5).value=x[2]
					tranzactii.cell(row=y+1,column=9).value=x[4]
					# tranzactii.cell(row=y+1,column=10).value="=xlookup(K"+str(y+1)+",'Mapping tranzactii'!R:R,'Mapping tranzactii'!L:L)"
			countp=0
			for i in range(0, len(listanouasetpurch)):
				x=listanouasetpurch[i].split(";")
				if(int(x[3])<3 or int(x[2])>0):
					countp=countp+1
					y=tranzactii.max_row
					tranzactii.cell(row=y+1,column=1).value=x[0]
					tranzactii.cell(row=y+1,column=2).value="=VLOOKUP(A"+str(y+1)+",'Mapping tranzactii'!B:J,9,0)"   
					try:
						tranzactii.cell(row=y+1,column=3).value=x[3]
					except:
						tranzactii.cell(row=y+1,column=3).value=""
					tranzactii.cell(row=y+1,column=4).value=x[1]
					tranzactii.cell(row=y+1,column=5).value=x[2]
					tranzactii.cell(row=y+1,column=9).value=x[4]
					# tranzactii.cell(row=y+1,column=10).value="=xlookup(K"+str(y+1)+",'Mapping tranzactii'!R:R,'Mapping tranzactii'!L:L)"

			countmare=countp+counts
			for i in range(0, countmare):
				tranzactii.cell(row=i+6,column=6).value="=SUMIFS('Mapping tranzactii'!H:H,'Mapping tranzactii'!B:B,A{0},'Mapping tranzactii'!E:E,C{0},'Mapping tranzactii'!F:F,D{0},'Mapping tranzactii'!G:G,E{0})/((100+E{0})/100)".format(6+i)
				tranzactii.cell(row=i+6,column=7).value="=F{0}/100*E{0}".format(6+i)
				tranzactii.cell(row=i+6,column=8).value="=COUNTIFS('Mapping tranzactii'!B:B,A{0},'Mapping tranzactii'!E:E,C{0},'Mapping tranzactii'!F:F,D{0},'Mapping tranzactii'!G:G,E{0})".format(6+i)
				tranzactii.cell(row=i+6,column=11).value="=A{0}&C{0}&D{0}&E{0}".format(6+i)


			#---------FORMAT-----------------
			for row in tranzactii['A5:J5']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in tranzactii['A5:J5']:
				for cell in row:
					cell.font = cap_tabel
		
			tranzactii.column_dimensions['K'].hidden = True
			tranzactii.column_dimensions['A'].width = 20
			tranzactii.column_dimensions['I'].width = 27		
			tranzactii.column_dimensions['B'].width = 35
			tranzactii.column_dimensions['C'].width = 13
			tranzactii.column_dimensions['D'].width = 13
			tranzactii.column_dimensions['H'].width = 13
			tranzactii.auto_filter.ref = "A5:H5"
			for row in tranzactii['F6:F'+str(tranzactii.max_row)]:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'
			for row in tranzactii['G6:G'+str(tranzactii.max_row)]:
				for cell in row:
					cell.number_format='#,##0_);(#,##0)'

			saf=temp.create_sheet("Facturi storno si anulate")
			saf.sheet_view.showGridLines = False
			saf.freeze_panes = 'A4'		
			saf.cell(row=1,column=1).value="Facturi storno/anulate"
			saf.cell(row=1,column=1).font=cap_tabelbold
			saf.cell(row=3,column=1).value="Tip"
			saf.cell(row=3,column=2).value="Serie"
			saf.cell(row=3,column=3).value="Numar"
			for row in saf['A3:C3']:
				for cell in row:
					cell.font = cap_tabel
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			ind=saf.max_row
			for k in range(0,len(storno)):
				if(storno[k]=="Yes"):
					saf.cell(row=ind+1,column=1).value="Stornata"
					saf.cell(row=ind+1,column=2).value=""
					saf.cell(row=ind+1,column=3).value=docNoSales[k]
					ind=ind+1
			xx=saf.max_row		
			saf.cell(row=xx+1,column=1).value="Anulata"
			saf.cell(row=xx+1,column=2).value="Please input the cancelled invoice number"	
			bonuri=temp.create_sheet("Bonuri fiscale")
			bonuri.cell(row=4,column=1).value="Luna"
			bonuri.cell(row=4,column=2).value="Nr. bon fiscal"
			bonuri.cell(row=4,column=3).value="Baza 5%"
			bonuri.cell(row=4,column=4).value="TVA 5%"
			bonuri.cell(row=4,column=5).value="Baza 9%"
			bonuri.cell(row=4,column=6).value="TVA 9%"
			bonuri.cell(row=4,column=7).value="Baza 19%"
			bonuri.cell(row=4,column=8).value="TVA 19%"
			bonuri.cell(row=4,column=9).value="Baza 20%"
			bonuri.cell(row=4,column=10).value="TVA 20%"

			for row in bonuri['A4:K4']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in bonuri['A4:K4']:
				for cell in row:
					cell.font = cap_tabel
			bonuri.sheet_view.showGridLines = False

			facturi=temp.create_sheet("Sectiunea 2.1&2.2")
			facturi.sheet_view.showGridLines = False
			facturi.column_dimensions['A'].width = 12
			facturi.column_dimensions['B'].width = 18
			facturi.column_dimensions['C'].width = 15

			facturi.cell(row=1,column=1).value="Serie Emise"
			facturi.cell(row=1,column=2).value="Inceput Emise"
			facturi.cell(row=1,column=3).value="Final Emise"
			facturi.cell(row=1,column=4).value="Tip Emise"		

			docNoSales2=[]
			seriefacturi=[]
			for i in range(0,len(docNoSales)):
				# docNoSales[i].replaceAll("[^a-zA-Z0-9]", ")
				if(int(codTranzactieSales[i])<3):
					# try:
					numere=re.sub("[^0-9]", "",str(docNoSales[i]))
					# except:
						# print(docNoSales[i])
					result = ''.join([i for i in str(docNoSales[i]) if not i.isdigit()])
					docNoSales2.append(numere)
					seriefacturi.append(result)
			#print(seriefacturi)
			# print(docNoSales2)
			initial=0
			final=0
			docNoSales2.sort()
			docNo=[]
			for k in range(0,len(docNoSales2)):
				docNo.append(str(docNoSales2[k]))
			docNo.sort()
			listaunica=list(set(docNoSales2))
			listaunica.sort()
	# print(listaunica)

			for i in range(0,len(listaunica)):
				listafacturi=[]
				print(listaunica[i])
				for j in range(0,len(docNoSales2)):
					if(listaunica[i]==docNoSales2[j]):

						listafacturi.append(int(docNoSales2[j]))
			listafacturi=list(set(listafacturi))
			print(listafacturi)
			listafacturi.sort()
			start=[]
			start.append(listaunica[0])
			stop=[]
			try:
				if(int(listaunica[1])-int(listaunica[0])>1):
					stop.append(listaunica[0])
				for k in range(1,len(listaunica)):

					if(int(listaunica[k])-int(listaunica[k-1])==1):
						print("ok")
					else:
						stop.append(listaunica[k-1])
						start.append(listaunica[k])
			except:
				stop.append(listaunica[0])
			if(len(stop)==len(start)):
				print("ok")
			else:
				stop.append(listaunica[len(listaunica)-1])
			print(start,stop)

			# #print(docNoSales)
			for k in range(0,len(start)):
				facturi.cell(row=2+k,column=2).value=start[k]
				facturi.cell(row=2+k,column=3).value=stop[k]
				facturi.cell(row=2+k,column=4).value=2

			# for p in range(0,len(docNoSales2)-1):
			# 	#print(docNo[p])
			# 	if(p==0):
			# 		initial=initial+1
			# 		# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 		facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 		if(int(docNo[p])-int(docNo[p+1])< -1):
			# 			final=final+1
			# 			facturi.cell(row=1+final,column=3).value=docNo[p]
			# 	else:
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])==1 and int(docNo[p])-int(docNo[p+1])==-1):
			# 				print("bailando")
						
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])==1 and int(docNo[p][3:])-int(docNo[p+1][3:])==-1):
			# 					print("bailando")
			# 					None
			# 			except:
			# 				print(None)
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])>1 and int(docNo[p])-int(docNo[p+1])==-1):
			# 				initial=initial+1
			# 				# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 				facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])>1 and int(docNo[p][3:])-int(docNo[p+1][3:])==-1):
			# 					initial=initial+1
			# 					# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 					facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 			except:
			# 				print(docNo[p])
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])==1 and int(docNo[p])-int(docNo[p+1])<-1):
			# 				final=final+1
			# 				facturi.cell(row=1+final,column=3).value=docNo[p]
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])==1 and int(docNo[p][3:])-int(docNo[p+1][3:])<-1):
			# 					final=final+1
			# 					facturi.cell(row=1+final,column=3).value=docNo[p]
			# 			except:
			# 				print("none")
			# 		try:
			# 			if(int(docNo[p])-int(docNo[p-1])>1 and int(docNo[p])-int(docNo[p+1])<-1):
			# 				initial=initial+1
			# 				# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 				facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 				final=final+1
			# 				facturi.cell(row=1+final,column=3).value=docNo[p]
			# 		except:
			# 			try:
			# 				if(int(docNo[p][3:])-int(docNo[p-1][3:])>1 and int(docNo[p][3:])-int(docNo[p+1][3:])<-1):
			# 					initial=initial+1
			# 					# facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
			# 					facturi.cell(row=1+initial,column=2).value=docNo[p]
			# 					final=final+1
			# 					facturi.cell(row=1+final,column=3).value=docNo[p]
			# 			except:
			# 				print("none")

			x=facturi.max_row
			facturi.auto_filter.ref = "A1:C1"
			# if(int(docNoSales2[len(docNoSales2)-1])-int(docNoSales2[len(docNoSales2)-2])>1):
			# 	facturi.cell(row=x+1,column=1).value=seriefacturi[0]
			# 	facturi.cell(row=x+1,column=2).value=docNoSales2[len(docNoSales2)-1]
			# 	facturi.cell(row=x+1,column=3).value=docNoSales2[len(docNoSales2)-1]
			# else:
			# 	facturi.cell(row=x+1,column=1).value=seriefacturi[0]
			# 	facturi.cell(row=x, column=3).value = docNoSales2[len(docNoSales2) - 1]


			yy=facturi.max_row+2
			facturi.cell(row=yy,column=1).value="Serie Alocate"
			facturi.cell(row=yy,column=2).value="Inceput Alocate"
			facturi.cell(row=yy,column=3).value="Final Alocate"
			facturi.cell(row=yy,column=4).value="Tip Alocate"		
			for kk in range(1,5):
				facturi.cell(row=yy,column=kk).font=cap_tabel
				facturi.cell(row=yy,column=kk).fill=cap_tabel_color_black	
			for pp in range(2,yy):
				facturi.cell(row=yy+pp-1,column=2).value=facturi.cell(row=pp,column=2).value
				facturi.cell(row=yy+pp-1,column=3).value=facturi.cell(row=pp,column=3).value
				facturi.cell(row=yy+pp-1,column=4).value=1						
			a23=temp.create_sheet("Sectiunea 2.3,2.4")
			dv = DataValidation(
				type='list', formula1='"Yes,No"', allow_blank=True,showDropDown=False)		
			dv.add(a23["A24"])			
			a23.sheet_view.showGridLines = False
			a23.column_dimensions['A'].width=18
			a23.column_dimensions['B'].width=13		
			a23.cell(row=1,column=1).value="Sectiunea 2.3"
			a23.cell(row=3,column=1).value="Denumire beneficiar"
			a23.cell(row=3,column=2).value="CUI beneficiar"
			a23.cell(row=3,column=1).fill=cap_tabel_color_black
			a23.cell(row=3,column=1).font=cap_tabel		
			a23.cell(row=36,column=1).border=border_bottom
			a23.cell(row=36,column=2).border=border_lowerleft
			a23.cell(row=35,column=2).border=border_right
			a23.cell(row=34,column=2).border=border_right
			a23.cell(row=33,column=2).border=border_right
			a23.cell(row=32,column=2).border=border_right						
			a23.cell(row=3,column=2).fill=cap_tabel_color_black
			a23.cell(row=3,column=2).font=cap_tabel


			a23.cell(row=6,column=1).value="Seria"
			a23.cell(row=6,column=2).value="De la"
			a23.cell(row=6,column=3).value="La"



			for row in a23['A6:C6']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in a23['A14:C14']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			a23.cell(row=9,column=1).value="Sectiunea 2.4"
			a23.cell(row=9,column=1).font=cap_tabelbold
			a23.cell(row=1,column=1).font=cap_tabelbold		
			a23.cell(row=28,column=1).font=cap_tabelbold		
			a23.cell(row=4,column=1).border=border_lowerright
			a23.cell(row=4,column=2).border=border_lowerright		
			a23.cell(row=12,column=1).border=border_lowerright
			a23.cell(row=12,column=2).border=border_lowerright		
			a23.cell(row=11,column=1).fill=cap_tabel_color_black
			a23.cell(row=11,column=1).font=cap_tabel		
			a23.cell(row=7,column=1).border=border_lowerright
			a23.cell(row=7,column=2).border=border_lowerright
			a23.cell(row=7,column=3).border=border_lowerright
			a23.cell(row=15,column=1).border=border_lowerright
			a23.cell(row=15,column=2).border=border_lowerright
			a23.cell(row=15,column=3).border=border_lowerright				
			a23.cell(row=11,column=2).fill=cap_tabel_color_black
			a23.cell(row=11,column=2).font=cap_tabel
			a23.cell(row=11,column=1).value="Denumire tert"

			a23.cell(row=11,column=2).value="CUI tert"

			a23.cell(row=14,column=1).value="Seria"
			a23.cell(row=14,column=2).value="De la"
			a23.cell(row=14,column=3).value="La"
			for row in a23['A31:B31']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
			try:
				if int(val4)== 1:
					a3=temp.create_sheet("Sectiunea 3")
					a3.cell(row=1,column=1).value="Sectiunea 3"
					a3.cell(row=1,column=1).font=cap_tabeltitlu
					a3.cell(row=3,column=1).value="In cazul in care soldul sumei negative inregistrate in decontul de TVA aferent perioadei de raportare este solicitat la rambursare , se vor selecta datele cu privire la natura operatiunilor din care provine acesta"
					a3.cell(row=7,column=1).value="Achizitii de bunuri si servicii legate direct de bunurile imobile din urmatoarele categorii"
					a3.cell(row=8,column=1).value="a) parcuri eoliene"
					a3.cell(row=9,column=1).value="b) constructii rezidentiale"



					a3.cell(row=10,column=1).value="c) cladiri de birouri"
					a3.cell(row=11,column=1).value="d) constructii industriale"
					a3.cell(row=12,column=1).value="e) altele"
					a3.cell(row=14,column=1).value="Achizitii de bunuri, cu exceptia celor legate direct de bunuri imobile:"
					a3.cell(row=15,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=16,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=17,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=18,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=19,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=21,column=1).value="Achizitii de servicii, cu exceptia celor legate direct de bunurile imobile:"
					a3.cell(row=22,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=23,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=24,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=25,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=26,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=28,column=1).value="Importuri de bunuri"
					a3.cell(row=30,column=1).value="Achizitii imobilizari necorporale"
					a3.cell(row=32,column=1).value="Livrari de bunuri imobile"
					a3.cell(row=34,column=1).value="Livrari de bunuri, cu exceptia bunurilor imobile:"
					a3.cell(row=35,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=36,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=37,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=38,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=39,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=41,column=1).value="Livrari de bunuri scutite de TVA"
					a3.cell(row=43,column=1).value="Livrari de bunuri/prestari de servicii pt care se aplica taxarea inversa"
					a3.cell(row=45,column=1).value="Prestari de servicii:"
					a3.cell(row=46,column=1).value="a) cu cota de TVA de 24%"
					a3.cell(row=47,column=1).value="b) cu cota standard de TVA de 20%"
					a3.cell(row=48,column=1).value="c) cu cota de TVA de 19%"
					a3.cell(row=49,column=1).value="d) cu cota de TVA de 9%"
					a3.cell(row=50,column=1).value="e) cu cota de TVA de 5%"
					a3.cell(row=52,column=1).value="Prestari de servicii scutite de TVA"
					a3.cell(row=54,column=1).value="Livrari intracomunitare de bunuri"
					a3.cell(row=56,column=1).value="Prestari intracomunitare de servicii"
					a3.cell(row=58,column=1).value="Exporturi de bunuri"
					a3.cell(row=60,column=1).value="Livrari imobilizari necorporale"
					a3.cell(row=62,column=1).value="Persoana impozabila nu a efectuat livrari de bunuri/prestari de servicii in perioada de raportare"																					
			except:
				pass


																																																																						
			a5=temp.create_sheet("Sectiunea 5")
			a5.sheet_view.showGridLines = False
			a5.cell(row=2,column=1).value="Sectiunea 5"
			a5.cell(row=2,column=1).font=cap_tabeltitlu
			a5.cell(row=4,column=1).font=cap_tabelbold
			a5.cell(row=6,column=1).font=cap_tabelbold
			a5.column_dimensions['B'].width=13
			a5.merge_cells('A6:J7')
			a5.merge_cells('A18:J19')		
			
			for pop in range(1,11):
				a5.cell(row=5,column=pop).border=border_bottom
				a5.cell(row=7,column=pop).border=border_bottom
				a5.cell(row=17,column=pop).border=border_bottom
				a5.cell(row=19,column=pop).border=border_bottom
			a5.cell(row=6,column=10).border=border_upperright
			a5.cell(row=7,column=10).border=border_lowerright
			a5.cell(row=18,column=10).border=border_upperright
			a5.cell(row=19,column=10).border=border_lowerright									
			a5['A6'].alignment=Alignment(wrap_text=True)
			a5['A18'].alignment=Alignment(wrap_text=True)
			a5['A30'].alignment=Alignment(wrap_text=True)						
			a5.cell(row=4,column=1).value="Sectiune 5.2"
			a5.cell(row=16,column=1).font=cap_tabelbold		
			a5.cell(row=16,column=1).value="Sectiune 5.3"
			a5.cell(row=28,column=1).font=cap_tabelbold		
			a5.cell(row=10,column=1).value="cota 24%"
			a5.cell(row=11,column=1).value="cota 20%"
			a5.cell(row=12,column=1).value="cota 19%"
			a5.cell(row=13,column=1).value="cota 9%"
			a5.cell(row=14,column=1).value="cota 5%"
			a5.cell(row=10,column=2).value=0
			a5.cell(row=11,column=2).value=0
			a5.cell(row=12,column=2).value=0
			a5.cell(row=13,column=2).value=0
			a5.cell(row=14,column=2).value=0

			a5.cell(row=9,column=2).value="Valoare TVA"
			a5.cell(row=9,column=1).value="Cota"		

			a5.cell(row=6,column=1).value="5.2 TVA deductibila aferenta facturilor achitate in perioada de raportare indiferent de data in care acestea au fost primite de la persoane impozabile care aplica sistemul normal de TVA, defalcata pe fiecare cota de TVA"
			a5.cell(row=18,column=1).font=cap_tabelbold		

			a5.cell(row=22,column=1).value="cota 24%"
			a5.cell(row=23,column=1).value="cota 20%"
			a5.cell(row=24,column=1).value="cota 19%"
			a5.cell(row=25,column=1).value="cota 9%"
			a5.cell(row=26,column=1).value="cota 5%"

			a5.cell(row=22,column=2).value=0
			a5.cell(row=23,column=2).value=0
			a5.cell(row=24,column=2).value=0
			a5.cell(row=25,column=2).value=0
			a5.cell(row=26,column=2).value=0

			a5.cell(row=21,column=1).value="Cota"		
			a5.cell(row=21,column=2).value="Valoare TVA"
			for row in a5['A9:B9']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in a5['A21:B21']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
								
			a5.cell(row=18,column=1).value="5.3 TVA deductibila aferenta facturilor achitate in perioada de raportare indiferent de data in care acestea au fost primite de la persoane impozabile care aplica sistemul de TVA la incasare, defalcata pe fiecare cota de TVA"

			for kk in range(22,27):
				a5.cell(row=kk,column=1).border=border_lowerright
				a5.cell(row=kk,column=2).border=border_lowerright						
			for kk in range(10,15):
				a5.cell(row=kk,column=1).border=border_lowerright
				a5.cell(row=kk,column=2).border=border_lowerright		
			a6=temp.create_sheet("Sectiunea 6")
			a6.sheet_view.showGridLines = False
			a6.cell(row=2,column=1).value="Sectiunea 6"
			a6.cell(row=2,column=1).font=cap_tabeltitlu
			a6.cell(row=4,column=1).font=cap_tabelbold
			a6.cell(row=6,column=1).font=cap_tabelbold
			a6.cell(row=15,column=1).font=cap_tabelbold
			a6.cell(row=13,column=1).font=cap_tabelbold
			a6.column_dimensions['A'].width=17			
			a6.column_dimensions['B'].width=28
			a6.column_dimensions['C'].width=17
																
			a6.cell(row=4,column=1).value="Sectiunea 6.1"		
			a6.cell(row=6,column=1).value="6.1 Persoanele impozabile care aplica regimul special pt agentiile de turism, vor completa:"
			a6.merge_cells('A6:F7')
			for l in range(1,7):
				a6.cell(row=5,column=l).border=border_bottom
				a6.cell(row=7,column=l).border=border_bottom
			a6.cell(row=6,column=6).border=border_upperright
			a6.cell(row=7,column=6).border=border_lowerright						

			a6.merge_cells('A15:G16')

			for l in range(1,8):
				a6.cell(row=14,column=l).border=border_bottom
				a6.cell(row=16,column=l).border=border_bottom
			a6.cell(row=15,column=7).border=border_upperright
			a6.cell(row=16,column=7).border=border_lowerright




			a6.cell(row=9,column=1).value="Incasarile agentiei"
			a6.cell(row=9,column=2).value="Costurile agentiei de turism"
			a6.cell(row=9,column=3).value="Marja de profit"
			a6.cell(row=9,column=4).value="TVA"

			a6.cell(row=13,column=1).value="Sectiunea 6.2"

			a6.cell(row=15,column=1).value="6.2 Persoanele impozabile care aplica regimul special pt bunurile second-hand, opere de arta, obiecte de colectie si antichitati , vor completa:"

			for row in a6['A18:D18']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in a6['A9:D9']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')										


			a6.cell(row=18,column=1).value="Pret de vanzare"
			a6.cell(row=18,column=2).value="Pret de cumparare"
			a6.cell(row=18,column=3).value="Marja de profit"	
			a6.cell(row=18,column=4).value="TVA"
			for oo in range(1,5):
				a6.cell(row=10,column=oo).border=border_lowerright
				a6.cell(row=19,column=oo).border=border_lowerright		

			a6['A6'].alignment=Alignment(wrap_text=True)

			a6['A15'].alignment=Alignment(wrap_text=True)
			a7=temp.create_sheet(" Sectiunea 7 ")

			a7.cell(row=2,column=1).value="Sectiunea 7"
			a7.sheet_view.showGridLines = False
			a7.cell(row=2,column=1).font=cap_tabeltitlu
			a7.cell(row=5,column=1).value="7. In situatia in care ati desfasurat, in perioada de raportare, activitati dintre cele inscrise in lista veti selecta activitatea corespunzatoare si veti inscrie valoarea livrarilor/prestarilor, precum si TVA aferenta"
			a7['A5'].alignment=Alignment(wrap_text=True)
			a7.merge_cells('A5:I6')
			a7.cell(row=5,column=1).font=cap_tabelbold
			for ii in range(1,10):
				a7.cell(row=4,column=ii).border=border_bottom
				a7.cell(row=6,column=ii).border=border_bottom
			a7.cell(row=5,column=9).border=border_upperright
			a7.cell(row=6,column=9).border=border_lowerright		
			a7.cell(row=5,column=8).border=border_right
			a7.cell(row=6,column=8).border=border_right

			for jj in range(1,4):
				a7.cell(row=9,column=jj).border=border_bottom
				a7.cell(row=9,column=jj).border=border_right
			
			for pp in range(12,17):
				a7.cell(row=pp,column=2).border=border_lowerright
				a7.cell(row=pp,column=1).border=border_lowerright
			a7.cell(row=8,column=1).value="Activitate"
			a7.cell(row=8,column=2).value="Tip operatiune"
			a7.cell(row=8,column=3).value="Valoarea livrarilor/prestarilor"

			a7.cell(row=11,column=1).value="Cota"
			a7.cell(row=12,column=1).value="cota 24%"
			a7.cell(row=13,column=1).value="cota 20%"
			a7.cell(row=14,column=1).value="cota 19%"
			a7.cell(row=15,column=1).value="cota 9%"
			a7.cell(row=16,column=1).value="cota 5%"
			for row in a7['A8:C8']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font = cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')										
			for row in a7['A11:B11']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.font = cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			a7.cell(row=11,column=2).value="Valoare TVA"

			for row in facturi['A1:D1']:
				for cell in row:
					cell.fill = cap_tabel_color_black
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			for row in facturi['A1:D1']:
				for cell in row:
					cell.font = cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				

			facturi.column_dimensions['C'].width = 14
			facturi.column_dimensions['D'].width = 14
			a23.column_dimensions['A'].width = 18
			a23.column_dimensions['B'].width = 13
			a23.column_dimensions['C'].width = 10
			a6.column_dimensions['A'].width = 17
			a6.column_dimensions['B'].width = 25
			a6.column_dimensions['C'].width = 15				
			sumaryG=temp.create_sheet("Sectiunea G. Manual input")
			sumaryG.sheet_view.showGridLines = False
			sumaryG.cell(row=2,column=1).value="Sectiunea G"
			sumaryG.cell(row=2,column=1).font=cap_tabeltitlu		
			sumaryG.column_dimensions['A'].width =45
			sumaryG.cell(row=5,column=1).value="Total Nr. Bonuri Fiscale"
			sumaryG.cell(row=5,column=1).font=cap_tabelbold
			sumaryG.cell(row=6,column=1).font=cap_tabelbold
			sumaryG.cell(row=7,column=1).font=cap_tabelbold						
			sumaryG.cell(row=6,column=1).value="Total incasari in perioada de raportare prin intermediul AMEF ( aparate de marcatelectronice fiscale ) inclusiv incasarile prin intermediul bonurilor fiscale care indeplinesc conditiile unei facturi simplificate indiferent daca au/nu au inscris codul de inregistrare in scopuri de TVA al beneficiarului (i1)"
			sumaryG['A6'].alignment=Alignment(wrap_text=True)
			sumaryG['A7'].alignment=Alignment(wrap_text=True)				
			sumaryG.cell(row=7,column=1).value="Total incasari in perioada de raportare efectuate din activitati exceptate de la obligatia utilizarii AMEF***) (i2) conform prevederilor legale in vigoare )"
			sumaryG.merge_cells('A9:H11')
			sumaryG.cell(row=9,column=1).font=cap_tabelbold
			for ii in range(1,9):
				sumaryG.cell(row=8,column=ii).border=border_bottom
				sumaryG.cell(row=11,column=ii).border=border_bottom
			sumaryG.cell(row=9,column=8).border=border_upperright
			sumaryG.cell(row=10,column=8).border=border_right		
			sumaryG.cell(row=11,column=8).border=border_lowerright		
			for ii in range(1,9):	
				sumaryG.cell(row=19,column=ii).border=border_bottom
				sumaryG.cell(row=20,column=ii).border=border_bottom			
			sumaryG.cell(row=5,column=2).border=border_right
			sumaryG.cell(row=6,column=2).border=border_right
			sumaryG.cell(row=7,column=2).border=border_right
			sumaryG.cell(row=5,column=1).border=border_right
			sumaryG.cell(row=6,column=1).border=border_right
			sumaryG.cell(row=7,column=1).border=border_right
			sumaryG.cell(row=4,column=1).border=border_bottom
			sumaryG.cell(row=5,column=1).border=border_bottom
			sumaryG.cell(row=6,column=1).border=border_bottom
			sumaryG.cell(row=7,column=1).border=border_bottom
			sumaryG.cell(row=5,column=2).value=0
			sumaryG.cell(row=6,column=2).value=0
			sumaryG.cell(row=7,column=2).value=0
			sumaryG.cell(row=13,column=1).value="Cota"		
			sumaryG.cell(row=13,column=2).value="Total baza impozabila"
			sumaryG.cell(row=13,column=3).value="TVA"
			sumaryG.cell(row=23,column=1).value="Cota"		
			sumaryG.cell(row=23,column=2).value="Total baza impozabila"
			sumaryG.cell(row=23,column=3).value="TVA"		
			for ii in range(1,4):
				sumaryG.cell(row=17,column=ii).border=border_lowerright
				sumaryG.cell(row=16,column=ii).border=border_lowerright
				sumaryG.cell(row=15,column=ii).border=border_lowerright
				sumaryG.cell(row=14,column=ii).border=border_lowerright
				sumaryG.cell(row=24,column=ii).border=border_lowerright
				sumaryG.cell(row=25,column=ii).border=border_lowerright
				sumaryG.cell(row=26,column=ii).border=border_lowerright
				sumaryG.cell(row=27,column=ii).border=border_lowerright																					
							
			for row in sumaryG['A13:C13']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryG['A23:C23']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')										
			sumaryG.cell(row=9,column=1).value="Incasari in perioada de raportare prin intermediul AMEF ( aparate de marcat electronice fiscale ) inclusiv incasarile prin intermediul bonurilor fiscale care indeplinesc conditiile unei facturi simplificate indiferent daca au/nu au inscris codul de inregistrare in scopuri de TVA al beneficiarului (i1)"
			sumaryG.cell(row=20,column=1).value="Incasari in perioada de raportare efectuate din activitati exceptate de la obligatia utilizarii AMEF***) (i2) conform prevederilor legale in vigoare )"
			sumaryG['A9'].alignment=Alignment(wrap_text=True)		
			sumaryG.cell(row=14,column=1).value="Cota 20%"
			sumaryG.cell(row=15,column=1).value="Cota 19%"
			sumaryG.cell(row=16,column=1).value="Cota 9%"
			sumaryG.cell(row=17,column=1).value="Cota 5%"

			sumaryG.cell(row=24,column=1).value="Cota 20%"
			sumaryG.cell(row=25,column=1).value="Cota 19%"
			sumaryG.cell(row=26,column=1).value="Cota 9%"
			sumaryG.cell(row=27,column=1).value="Cota 5%"

			sumaryG.cell(row=14,column=2).value=0
			sumaryG.cell(row=15,column=2).value=0
			sumaryG.cell(row=16,column=2).value=0
			sumaryG.cell(row=17,column=2).value=0


			sumaryG.cell(row=14,column=3).value=0
			sumaryG.cell(row=15,column=3).value=0
			sumaryG.cell(row=16,column=3).value=0
			sumaryG.cell(row=17,column=3).value=0


			sumaryG.cell(row=24,column=2).value=0
			sumaryG.cell(row=25,column=2).value=0
			sumaryG.cell(row=26,column=2).value=0
			sumaryG.cell(row=27,column=2).value=0

			sumaryG.cell(row=24,column=3).value=0
			sumaryG.cell(row=25,column=3).value=0
			sumaryG.cell(row=26,column=3).value=0
			sumaryG.cell(row=27,column=3).value=0


			sumaryI=temp.create_sheet("Sectiunea I 1. Manual input")

			for row in sumaryI['A7:C7']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A17:C17']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A27:C27']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A37:C37']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')				
			for row in sumaryI['A47:C47']:
				for cell in row:
					cell.fill=cap_tabel_color_black
					cell.font=cap_tabel
					cell.alignment=Alignment(horizontal='center',vertical='center')																
			sumaryI.cell(row=2,column=1).value="Sectiunea I"
			sumaryI.cell(row=2,column=1).font=cap_tabeltitlu
			sumaryI.sheet_view.showGridLines = False
			sumaryI.cell(row=7,column=1).value="Cota"
			sumaryI.cell(row=7,column=2).value="Baza impozabila"
			sumaryI.cell(row=7,column=3).value="TVA"

			sumaryI.cell(row=7,column=1).value="Cota"
			sumaryI.cell(row=7,column=2).value="Baza impozabila"
			sumaryI.cell(row=7,column=3).value="TVA"
			sumaryI.cell(row=17,column=1).value="Cota"
			sumaryI.cell(row=17,column=2).value="Baza impozabila"
			sumaryI.cell(row=17,column=3).value="TVA"
			sumaryI.cell(row=27,column=1).value="Cota"
			sumaryI.cell(row=27,column=2).value="Baza impozabila"
			sumaryI.cell(row=27,column=3).value="TVA"
			sumaryI.cell(row=37,column=1).value="Cota"
			sumaryI.cell(row=37,column=2).value="Baza impozabila"
			sumaryI.cell(row=37,column=3).value="TVA"								
			sumaryI.cell(row=47,column=1).value="Cota"
			sumaryI.cell(row=47,column=2).value="Baza impozabila"
			sumaryI.cell(row=47,column=3).value="TVA"
			sumaryI.cell(row=4,column=1).value="1.1 Livrari de bunuri/prestari de servicii pentru care s-au emis facturi simplificate care au inscris codul de inregistrare in scopuri de TVA al beneficiarului %"
			sumaryI.cell(row=4,column=1).font=cap_tabelbold
			sumaryI.cell(row=14,column=1).font=cap_tabelbold
			sumaryI.cell(row=24,column=1).font=cap_tabelbold
			sumaryI.cell(row=34,column=1).font=cap_tabelbold
			sumaryI.cell(row=44,column=1).font=cap_tabelbold								
			sumaryI.cell(row=14,column=1).value="1.2 Livrari de bunuri/prestari de servicii pentru care s-au emis facturi simplificate fara a avea inscris codul de inregistrare in scopuri de TVA al beneficiarului %"
			sumaryI['A4'].alignment=Alignment(wrap_text=True)
			sumaryI['A14'].alignment=Alignment(wrap_text=True)
			sumaryI['A24'].alignment=Alignment(wrap_text=True)
			sumaryI['A34'].alignment=Alignment(wrap_text=True)
			sumaryI['A44'].alignment=Alignment(wrap_text=True)								
			sumaryI.cell(row=24,column=1).value="1.3 Achizitii de bunuri si servicii pentru care s-au primit facturi simplificate de la persoane impozabile care aplica sistemul normal de TVA si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

			sumaryI.cell(row=34,column=1).value="1.4 Achizitii de bunuri si servicii pentru care s-au primit facturi simplificate de la persoane impozabile care aplica sistemul de TVA la incasare si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

			sumaryI.cell(row=44,column=1).value="1.5 Achizitii de bunuri si servicii pentru care s-au primit bonuri fiscale care indeplinesc conditiile unei facturi simplificate si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

			for ip in range(1,11):
				sumaryI.cell(row=3,column=ip).border=border_bottom
				sumaryI.cell(row=6,column=ip).border=border_top
				sumaryI.cell(row=13,column=ip).border=border_bottom
				sumaryI.cell(row=16,column=ip).border=border_top
				sumaryI.cell(row=23,column=ip).border=border_bottom
				sumaryI.cell(row=26,column=ip).border=border_top
				sumaryI.cell(row=33,column=ip).border=border_bottom
				sumaryI.cell(row=36,column=ip).border=border_top
				sumaryI.cell(row=43,column=ip).border=border_bottom
				sumaryI.cell(row=46,column=ip).border=border_top
			sumaryI.cell(row=4,column=10).border=border_upperright
			sumaryI.cell(row=5,column=10).border=border_lowerright
			sumaryI.cell(row=14,column=10).border=border_upperright
			sumaryI.cell(row=15,column=10).border=border_lowerright
			sumaryI.cell(row=24,column=10).border=border_upperright
			sumaryI.cell(row=25,column=10).border=border_lowerright
			sumaryI.cell(row=34,column=10).border=border_upperright
			sumaryI.cell(row=35,column=10).border=border_lowerright
			sumaryI.cell(row=44,column=10).border=border_upperright
			sumaryI.cell(row=45,column=10).border=border_lowerright						

			sumaryI.cell(row=4,column=10).border=border_upperright
			sumaryI.cell(row=5,column=10).border=border_lowerright
			sumaryI.cell(row=14,column=10).border=border_upperright
			sumaryI.cell(row=15,column=10).border=border_lowerright
			sumaryI.cell(row=24,column=10).border=border_upperright
			sumaryI.cell(row=25,column=10).border=border_lowerright
			sumaryI.cell(row=34,column=10).border=border_upperright
			sumaryI.cell(row=35,column=10).border=border_lowerright
			sumaryI.cell(row=44,column=10).border=border_upperright
			sumaryI.cell(row=45,column=10).border=border_lowerright										
			sumaryI.cell(row=4,column=10).border=border_right
			sumaryI.cell(row=5,column=10).border=border_right
			sumaryI.cell(row=14,column=10).border=border_right
			sumaryI.cell(row=15,column=10).border=border_right
			sumaryI.cell(row=24,column=10).border=border_right
			sumaryI.cell(row=25,column=10).border=border_right
			sumaryI.cell(row=34,column=10).border=border_right
			sumaryI.cell(row=35,column=10).border=border_right
			sumaryI.cell(row=44,column=10).border=border_right
			sumaryI.cell(row=45,column=10).border=border_right																																																	
			for io in range(8,13):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(18,23):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(28,33):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(38,43):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright			
			for io in range(48,53):
				sumaryI.cell(row=io,column=1).border=border_lowerright
				sumaryI.cell(row=io,column=2).border=border_lowerright
				sumaryI.cell(row=io,column=3).border=border_lowerright																		
			sumaryI.merge_cells('A4:J5')
			sumaryI.merge_cells('A14:J15')
			sumaryI.merge_cells('A24:J25')
			sumaryI.merge_cells('A34:J35')
			sumaryI.merge_cells('A44:J45')
									
			sumaryI.cell(row=8,column=1).value="Cota 24%"
			sumaryI.cell(row=9,column=1).value="Cota 20%"
			sumaryI.cell(row=10,column=1).value="Cota 19%"
			sumaryI.cell(row=11,column=1).value="Cota 9%"
			sumaryI.cell(row=12,column=1).value="Cota 5%"

			sumaryI.cell(row=8,column=2).value=0
			sumaryI.cell(row=9,column=2).value=0
			sumaryI.cell(row=10,column=2).value=0
			sumaryI.cell(row=11,column=2).value=0
			sumaryI.cell(row=12,column=2).value=0

			sumaryI.cell(row=18,column=2).value=0
			sumaryI.cell(row=19,column=2).value=0
			sumaryI.cell(row=20,column=2).value=0
			sumaryI.cell(row=21,column=2).value=0
			sumaryI.cell(row=22,column=2).value=0

			sumaryI.cell(row=28,column=2).value=0
			sumaryI.cell(row=29,column=2).value=0
			sumaryI.cell(row=30,column=2).value=0
			sumaryI.cell(row=31,column=2).value=0
			sumaryI.cell(row=32,column=2).value=0

			sumaryI.cell(row=38,column=2).value=0
			sumaryI.cell(row=39,column=2).value=0
			sumaryI.cell(row=40,column=2).value=0
			sumaryI.cell(row=41,column=2).value=0
			sumaryI.cell(row=42,column=2).value=0

			sumaryI.cell(row=48,column=2).value=0
			sumaryI.cell(row=49,column=2).value=0
			sumaryI.cell(row=50,column=2).value=0
			sumaryI.cell(row=51,column=2).value=0
			sumaryI.cell(row=52,column=2).value=0

			sumaryI.cell(row=8,column=3).value=0
			sumaryI.cell(row=9,column=3).value=0
			sumaryI.cell(row=10,column=3).value=0
			sumaryI.cell(row=11,column=3).value=0
			sumaryI.cell(row=12,column=3).value=0

			sumaryI.cell(row=18,column=3).value=0
			sumaryI.cell(row=19,column=3).value=0
			sumaryI.cell(row=20,column=3).value=0
			sumaryI.cell(row=21,column=3).value=0
			sumaryI.cell(row=22,column=3).value=0

			sumaryI.cell(row=28,column=3).value=0
			sumaryI.cell(row=29,column=3).value=0
			sumaryI.cell(row=30,column=3).value=0
			sumaryI.cell(row=31,column=3).value=0
			sumaryI.cell(row=32,column=3).value=0

			sumaryI.cell(row=38,column=3).value=0
			sumaryI.cell(row=39,column=3).value=0
			sumaryI.cell(row=40,column=3).value=0
			sumaryI.cell(row=41,column=3).value=0
			sumaryI.cell(row=42,column=3).value=0

			sumaryI.cell(row=48,column=3).value=0
			sumaryI.cell(row=49,column=3).value=0
			sumaryI.cell(row=50,column=3).value=0
			sumaryI.cell(row=51,column=3).value=0
			sumaryI.cell(row=52,column=3).value=0		

			sumaryI.cell(row=8,column=1).value="Cota 24%"
			sumaryI.cell(row=9,column=1).value="Cota 20%"
			sumaryI.cell(row=10,column=1).value="Cota 19%"
			sumaryI.cell(row=11,column=1).value="Cota 9%"
			sumaryI.cell(row=12,column=1).value="Cota 5%"

			sumaryI.cell(row=18,column=1).value="Cota 24%"
			sumaryI.cell(row=19,column=1).value="Cota 20%"
			sumaryI.cell(row=20,column=1).value="Cota 19%"
			sumaryI.cell(row=21,column=1).value="Cota 9%"
			sumaryI.cell(row=22,column=1).value="Cota 5%"

			sumaryI.cell(row=28,column=1).value="Cota 24%"
			sumaryI.cell(row=29,column=1).value="Cota 20%"
			sumaryI.cell(row=30,column=1).value="Cota 19%"
			sumaryI.cell(row=31,column=1).value="Cota 9%"
			sumaryI.cell(row=32,column=1).value="Cota 5%"

			sumaryI.cell(row=38,column=1).value="Cota 24%"
			sumaryI.cell(row=39,column=1).value="Cota 20%"
			sumaryI.cell(row=40,column=1).value="Cota 19%"
			sumaryI.cell(row=41,column=1).value="Cota 9%"
			sumaryI.cell(row=42,column=1).value="Cota 5%"

			sumaryI.cell(row=48,column=1).value="Cota 24%"
			sumaryI.cell(row=49,column=1).value="Cota 20%"
			sumaryI.cell(row=50,column=1).value="Cota 19%"
			sumaryI.cell(row=51,column=1).value="Cota 9%"
			sumaryI.cell(row=52,column=1).value="Cota 5%"

			sumaryI.column_dimensions['B'].width = 15	
			sumaryG.column_dimensions['B'].width = 20	
			a7.column_dimensions['B'].width = 12						
			a7.column_dimensions['C'].width = 22
			a6.column_dimensions['B'].width = 22						
			a6.column_dimensions['C'].width = 12


			# for row in sumary['A5:D34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['F5:I34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['K5:N34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['P5:S34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# for row in sumary['U5:X34']:
			# 	for cell in row:
			# 		cell.border = border_thin

			# listanoua=['A','B','C','D','F','G','H','I','K','L','M','N','P','Q','R','S','U','V','W','W','X']
			# for column in listanoua:
			# 	for i in listanoua:
			# 		if (column==i):
	# 			sumary.column_dimensions[column].width = 15

			
			# for i in range(0 ,len(tip)):
		# folderpath="D:/D300 to XML/docs"
		folderpath="/home/mirus_app/storage_spreadsheet"
		# folderpath="C:/Users/Cristian.Iordache/Documents/D300 to XML Final CI/D300 to XML 2/storage"
		file_pathFS = os.path.join(folderpath, "One VAT app spreadsheets " +str(clientname)+".xlsx")
		temp.save(file_pathFS)
		# return send_from_directory("D:/D300 to XML/docs","One VAT app spreadsheets.xlsx",as_attachment=True)
		return send_from_directory("/home/mirus_app/storage_spreadsheet","One VAT app spreadsheets " +str(clientname)+".xlsx",as_attachment=True)
		return render_template('D3APPS2')
@app.route('/D3APPS2')
def my_form2():
    return render_template('D3APPS second step.html')

@app.route('/D3APPS2', methods=['POST', 'GET'])
def D300_thales():
	if request.method == 'POST':
		D300_2= request.files["d300file2"]
	temp = openpyxl.load_workbook(D300_2,data_only=True)
	try:
		amount=temp['D300 draft figures']
		info=temp['Other info']
		an=info.cell(row=20,column=3).value
		#print(an)
		luna=info.cell(row=3,column=3).value
		# temei=info.cell(row=8,column=2).value
		cif=info.cell(row=27,column=3).value
		den=info.cell(row=26,column=3).value

		# #print(an, 'an')

		# cif=info.cell(row=11,column=2).value
		# adresa="strada: "str(info.cell(row=12,column=2).value)+str(info.cell(row=13,column=2).value)+ ",localitate: "+str(info.cell(row=14,column=2).value)+" "+str(info.cell(row=16,column=2).value)+",judet: "+str(info.cell(row=15,column=2).value)+",cod postal: "+str(info.cell(row=21,column=2).value)
		strada=info.cell(row=28,column=3).value
		numar=info.cell(row=29,column=3).value
		localitate=info.cell(row=30,column=3).value
		judet=info.cell(row=31,column=3).value

		sector=info.cell(row=32,column=3).value
		if(sector==None):
			sector=""
		bloc=info.cell(row=33,column=3).value
		if(bloc==None):
			bloc=""
		scara=info.cell(row=34,column=3).value
		if(scara==None):
			scara=""
		etaj=info.cell(row=35,column=3).value
		if(etaj==None):
			etaj=""
		apt=info.cell(row=36,column=3).value
		if(apt==None):
			apt=""
		strada=strada
		# codpostal=info.cell(row=37,column=3).value
		telefon=info.cell(row=29,column=3).value
		email=info.cell(row=31,column=3).value
		banca=info.cell(row=32,column=3).value
		contban=info.cell(row=33,column=3).value
		Caen=info.cell(row=34,column=3).value
		prorata=info.cell(row=35,column=3).value
		cereale=info.cell(row=36,column=3).value
		telmob=info.cell(row=37,column=3).value
		disp=info.cell(row=38,column=3).value
		nr_evid=int(info.cell(row=18,column=3).value)
		cons=info.cell(row=39,column=3).value
		ramburs=info.cell(row=40,column=3).value
		if(ramburs=="N"):
			ramburs2=0
		else:
			ramburs2=1
		# total_plata=R1_1+R2_1+R3_1+R3_1_1+R4_1+R5_1+R5_1_1+R6_1+R7_1+R7_1_1+R8_1+R9_1+R10_1+R11_1+R12_1+R12_1_1+R12_2_1+R12_3_1+R13_1+R14_1+R15_1+R16_1+R64_1+R65_1+R17_1+R18_1+R18_1_1+R19_1+R20_1+R20_1_1+R21_1+R22_1+R23_1+R24_1+R25_1+R25_1_1+R25_2_1+R25_3_1+R43_1+R44_1+R26_1+R26_1_1+R27_1+R28_1+R29_1+R30_1+R31_1+R32_1+R33_1+R34_1+R35_1+R36_1+R37_1+R38_1+R39_1+R40_1+R41_1+R42_1+R1_2+R2_2+R3_2+R3_1_2+R4_2+R5_2+R5_1_2+R6_2+R7_2+R7_1_2+R8_2+R9_2+R10_2+R11_2+R12_2+R12_1_2+R12_2_2+R12_3_2+R13_2+R14_2+R15_2+R16_2+R64_2+R65_2+R17_2+R18_2+R18_1_2+R19_2+R20_2+R20_1_2+R21_2+R22_2+R23_2+R24_2+R25_2+R25_1_2+R25_2_2+R25_3_2+R43_2+R44_2+R26_2+R26_1_2+R27_2+R28_2+R29_2+R30_2+R31_2+R32_2+R33_2+R34_2+R35_2+R36_2+R37_2+R38_2+R39_2+R40_2+R41_2+R42_2+valoare_a+valoare_a1+valoare_b+valoare_b1+tva_a+tva_a1+tva_b+tva_b1

		ramburs=info.cell(row=40,column=3).value
		nrfact=info.cell(row=41,column=3).value
		baza=info.cell(row=42,column=3).value
		tva=info.cell(row=43,column=3).value
		factprimite=info.cell(row=44,column=3).value
		bazaprimite=info.cell(row=45,column=3).value
		tvaprimite=info.cell(row=46,column=3).value
		valoare_a=amount.cell(row=68,column=2).value
		valoare_a1=amount.cell(row=69,column=2).value
		tva_a=amount.cell(row=68,column=3).value
		tva_a1=amount.cell(row=69,column=3).value
		valoare_b=amount.cell(row=70,column=2).value
		valoare_b1=amount.cell(row=71,column=2).value
		tva_b=amount.cell(row=70,column=3).value
		tva_b1=amount.cell(row=71,column=3).value
		nrfactemise=info.cell(row=47,column=3).value
		total_baza=info.cell(row=48,column=3).value
		total_tva=info.cell(row=49,column=3).value
		total_precedent=info.cell(row=50,column=3).value
		total_curent=info.cell(row=51,column=3).value
		tip=info.cell(row=53,column=3).value
		pren=info.cell(row=54,column=3).value
		nume=info.cell(row=55,column=3).value
		funct=info.cell(row=56,column=3).value
		total_precedent=amount.cell(row=74,column=2).value
		total_curent=amount.cell(row=74,column=3).value
		totalp=info.cell(row=52,column=3).value

		for row in amount.iter_rows():
			for cell in row:
				if cell.value == "Taxable basis":
					rand_tb = cell.row
					suma1 = cell.column
					lun = len(amount[cell.column])
		coloana = [b.value for b in amount[suma1][rand_tb:lun]]


		R1_1=amount.cell(row=8,column=2).value
		R2_1=amount.cell(row=9,column=2).value
		R3_1=amount.cell(row=10,column=2).value
		R3_1_1=amount.cell(row=11,column=2).value
		R4_1=amount.cell(row=12,column=2).value
		R5_1=amount.cell(row=13,column=2).value
		R5_1_1=amount.cell(row=14,column=2).value
		R6_1=amount.cell(row=15,column=2).value
		R7_1=amount.cell(row=16,column=2).value
		R7_1_1=amount.cell(row=17,column=2).value
		R8_1=amount.cell(row=18,column=2).value
		R9_1=amount.cell(row=19,column=2).value
		R10_1=amount.cell(row=20,column=2).value
		R11_1=amount.cell(row=21,column=2).value
		R12_1=amount.cell(row=22,column=2).value
		R12_1_1=amount.cell(row=23,column=2).value
		R12_2_1=amount.cell(row=24,column=2).value
		R12_3_1=amount.cell(row=25,column=2).value
		R13_1=amount.cell(row=26,column=2).value
		R14_1=amount.cell(row=27,column=2).value
		R15_1=amount.cell(row=28,column=2).value
		R16_1=amount.cell(row=29,column=2).value
		R64_1=amount.cell(row=30,column=2).value
		R65_1=amount.cell(row=31,column=2).value
		R17_1=amount.cell(row=32,column=2).value
		R18_1=amount.cell(row=33,column=2).value
		R18_1_1=amount.cell(row=34,column=2).value
		R19_1=amount.cell(row=35,column=2).value
		R20_1=amount.cell(row=36,column=2).value
		R20_1_1=amount.cell(row=37,column=2).value
		R21_1=amount.cell(row=38,column=2).value
		R22_1=amount.cell(row=39,column=2).value
		R23_1=amount.cell(row=40,column=2).value
		R24_1=amount.cell(row=41,column=2).value
		R25_1=amount.cell(row=42,column=2).value
		R25_1_1=amount.cell(row=43,column=2).value
		R25_2_1=amount.cell(row=44,column=2).value
		R25_3_1=amount.cell(row=45,column=2).value
		R43_1=amount.cell(row=46,column=2).value
		R44_1=amount.cell(row=47,column=2).value
		R26_1=amount.cell(row=48,column=2).value
		R26_1_1=amount.cell(row=49,column=2).value
		R27_1=amount.cell(row=50,column=2).value
		R28_1=amount.cell(row=52,column=2).value
		R29_1=amount.cell(row=53,column=2).value
		R30_1=amount.cell(row=54,column=2).value
		R31_1=amount.cell(row=55,column=2).value
		R32_1=amount.cell(row=56,column=2).value
		R33_1=amount.cell(row=57,column=2).value
		R34_1=amount.cell(row=58,column=2).value
		R35_1=amount.cell(row=59,column=2).value
		R36_1=amount.cell(row=60,column=2).value
		R37_1=amount.cell(row=61,column=2).value
		R38_1=amount.cell(row=62,column=2).value
		R39_1=amount.cell(row=63,column=2).value
		R40_1=amount.cell(row=64,column=2).value
		R41_1=amount.cell(row=65,column=2).value
		R42_1=amount.cell(row=66,column=2).value
		R1_2=amount.cell(row=8,column=3).value
		R2_2=amount.cell(row=9,column=3).value
		R3_2=amount.cell(row=10,column=3).value
		R3_1_2=amount.cell(row=11,column=3).value
		R4_2=amount.cell(row=12,column=3).value
		R5_2=amount.cell(row=13,column=3).value
		R5_1_2=amount.cell(row=14,column=3).value
		R6_2=amount.cell(row=15,column=3).value
		R7_2=amount.cell(row=16,column=3).value
		R7_1_2=amount.cell(row=17,column=3).value
		R8_2=amount.cell(row=18,column=3).value
		R9_2=amount.cell(row=19,column=3).value
		R10_2=amount.cell(row=20,column=3).value
		R11_2=amount.cell(row=21,column=3).value
		R12_2=amount.cell(row=22,column=3).value


		R12_1_2=amount.cell(row=23,column=3).value
		R12_2_2=amount.cell(row=24,column=3).value
		R12_3_2=amount.cell(row=25,column=3).value
		R13_2=amount.cell(row=26,column=3).value
		R14_2=amount.cell(row=27,column=3).value
		R15_2=amount.cell(row=28,column=3).value
		R16_2=amount.cell(row=29,column=3).value
		R64_2=amount.cell(row=30,column=3).value
		R65_2=amount.cell(row=31,column=3).value
		R17_2=amount.cell(row=32,column=3).value
		R18_2=amount.cell(row=33,column=3).value
		R18_1_2=amount.cell(row=34,column=3).value
		R19_2=amount.cell(row=35,column=3).value
		R20_2=amount.cell(row=36,column=3).value
		R20_1_2=amount.cell(row=37,column=3).value
		R21_2=amount.cell(row=38,column=3).value
		R22_2=amount.cell(row=39,column=3).value
		R23_2=amount.cell(row=40,column=3).value
		R24_2=amount.cell(row=41,column=3).value
		R25_2=amount.cell(row=42,column=3).value
		R25_1_2=amount.cell(row=43,column=3).value
		R25_2_2=amount.cell(row=44,column=3).value
		R25_3_2=amount.cell(row=45,column=3).value


		R43_2=amount.cell(row=46,column=3).value
		R44_2=amount.cell(row=47,column=3).value
		R26_2=amount.cell(row=48,column=3).value
		R26_1_2=amount.cell(row=49,column=3).value
		R27_2=amount.cell(row=50,column=3).value


		R28_2=amount.cell(row=52,column=3).value
		R29_2=amount.cell(row=53,column=3).value
		R30_2=amount.cell(row=54,column=3).value
		R31_2=amount.cell(row=55,column=3).value
		R32_2=amount.cell(row=56,column=3).value


		R33_2=amount.cell(row=57,column=3).value
		R34_2=amount.cell(row=58,column=3).value
		R35_2=amount.cell(row=59,column=3).value
		R36_2=amount.cell(row=60,column=3).value
		R37_2=amount.cell(row=61,column=3).value
		R38_2=amount.cell(row=62,column=3).value
		R39_2=amount.cell(row=63,column=3).value
		R40_2=amount.cell(row=64,column=3).value
		R41_2=amount.cell(row=65,column=3).value
		R42_2=amount.cell(row=66,column=3).value
		


		# for i in range(0 ,len(tip)):
		# folderpath="D:/apps/TEST D3APPS/Test 21.03.2022/output"
		folderpath="/home/mirus_app/storage"
		# folderpath="C:/Users/Cristian.Iordache/Documents/D300 to XML Final CI/D300 to XML 2/storage"
		# temp.save(folderpath+".xlsx")
		text='<?xml version="1.0"?> <declaratie300  luna="'+str(luna)+'" an="'+str(an)+'" depusReprezentant="'+str(ramburs2)+'" bifa_interne="0" temei="0" prenume_declar="'+str(pren)+'" nume_declar="'+str(nume)+'" functie_declar="'+str(funct)+'" cui="'+str(cif)+'" den="'+str(den)+'" adresa="'+str(strada)+'" telefon="'+str(telefon)+'" mail="'+str(email)+'" banca="'+str(banca)+'" cont="'+str(contban)+'" caen="'+str(Caen)+'" tip_decont="'+str(tip)+'" pro_rata="'+str(prorata)+'" bifa_cereale="'+str(cereale)+'" bifa_mob="'+str(telmob)+'" bifa_disp="'+str(disp)+'" bifa_cons="'+str(cons)+'" solicit_ramb="'+str(ramburs)+'" nr_evid="'+str(nr_evid)+'" totalPlata_A="'+str(totalp)+'" R1_1="'+str(R1_1)+'" R2_1="'+str(R2_1)+'" R3_1="'+str(R3_1)+'" R3_1_1="'+str(R3_1_1)+'" R4_1="'+str(R4_1)+'" R5_1="'+str(R5_1)+'" R5_2="'+str(R5_2)+'" R5_1_1="'+str(R5_1_1)+'" R5_1_2="'+str(R5_1_2)+'" R6_1="0" R6_2="'+str(R6_2)+'" R7_1="'+str(R7_1)+'" R7_2="'+str(R7_2)+'" R7_1_1="'+str(R7_1_1)+'" R7_1_2="'+str(R7_1_2)+'" R8_1="'+str(R8_1)+'" R8_2="'+str(R8_2)+'" R9_1="'+str(R9_1)+'" R9_2="'+str(R9_2)+'" R10_1="'+str(R10_1)+'" R10_2="'+str(R10_2)+'" R11_1="'+str(R11_1)+'" R11_2="'+str(R11_2)+'" R12_1="'+str(R12_1)+'" R12_2="'+str(R12_2)+'" R12_1_1="'+str(R12_1_1)+'" R12_1_2="'+str(R12_1_2)+'" R12_2_1="'+str(R12_2_1)+'" R12_2_2="'+str(R12_2_2)+'" R12_3_1="'+str(R12_3_1)+'" R12_3_2="'+str(R12_3_2)+'" R13_1="'+str(R13_1)+'" R14_1="'+str(R14_1)+'" R15_1="'+str(R15_1)+'" R16_1="'+str(R16_1)+'" R16_2="'+str(R16_2)+'" R64_1="'+str(R64_1)+'" R64_2="'+str(R64_2)+'" R65_1="'+str(R65_1)+'" R65_2="'+str(R65_2)+'" R17_1="'+str(R17_1)+'" R17_2="'+str(R17_2)+'" R18_1="'+str(R18_1)+'" R18_2="'+str(R18_2)+'" R18_1_1="'+str(R18_1_1)+'" R18_1_2="'+str(R18_1_2)+'" R19_1="'+str(R19_1)+'" R19_2="'+str(R19_2)+'" R20_1="'+str(R20_1)+'" R20_2="'+str(R20_2)+'" R20_1_1="'+str(R20_1_1)+'" R20_1_2="'+str(R20_1_2)+'" R21_1="'+str(R21_1)+'" R21_2="'+str(R21_2)+'" R22_1="'+str(R22_1)+'" R22_2="'+str(R22_2)+'" R23_1="'+str(R23_1)+'" R23_2="'+str(R23_2)+'" R24_1="'+str(R24_1)+'" R24_2="'+str(R24_2)+'" R25_1="'+str(R25_1)+'" R25_2="'+str(R25_2)+'" R25_1_1="'+str(R25_1_1)+'" R25_1_2="'+str(R25_1_2)+'" R25_2_1="'+str(R25_2_1)+'" R25_2_2="'+str(R25_2_2)+'" R25_3_1="'+str(R25_3_1)+'" R25_3_2="'+str(R25_3_2)+'" R43_2="'+str(R43_2)+'" R44_2="'+str(R44_2)+'" R26_1="'+str(R26_1)+'" R26_1_1="'+str(R26_1_1)+'" R27_1="'+str(R27_1)+'" R27_2="'+str(R27_2)+'" R28_2="'+str(R28_2)+'" R29_2="'+str(R29_2)+'" R30_1="'+str(R30_1)+'" R30_2="'+str(R30_2)+'" R31_2="'+str(R31_2)+'" R32_2="'+str(R32_2)+'" R33_2="'+str(R33_2)+'" R34_2="'+str(R34_2)+'" R35_2="'+str(R35_2)+'" R36_2="'+str(R36_2)+'" R37_2="'+str(R37_2)+'" R38_2="'+str(R38_2)+'" R39_2="'+str(R39_2)+'" R40_2="'+str(R40_2)+'" R41_2="'+str(R41_2)+'" R42_2="'+str(R42_2)+'" nr_facturi="'+str(nrfact)+'" baza="'+str(baza)+'" tva="'+str(tva)+'" nr_facturi_primite="'+str(factprimite)+'" baza_primite="'+str(bazaprimite)+'" tva_primite="'+str(tvaprimite)+'" nr_fact_emise="'+str(nrfactemise)+'" total_baza="'+str(total_baza)+'" total_precedent ="'+str(total_precedent)+'" total_curent ="'+str(total_curent)+'" total_tva="'+str(total_tva)+'" valoare_a="'+str(valoare_a)+'" tva_a="'+str(tva_a)+'" valoare_a1="'+str(valoare_a1)+'" tva_a1="'+str(tva_a1)+'" valoare_b="'+str(valoare_b)+'" tva_b="'+str(tva_b)+'" valoare_b1="'+str(valoare_b1)+'" tva_b1="'+str(tva_b1)+'" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="mfp:anaf:dgti:d300:declaratie:v7 d300.xsd" xmlns="mfp:anaf:dgti:d300:declaratie:v7"></declaratie300>'
		#print(text)
		with open("/home/mirus_app/storage/D300.xml", "w", encoding="utf-8") as f:
			f.write(text)
		# f=open("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage/D300.xml", "w").write(text).encode('utf-8')
		# 	# f=open("C:/Users/Cristian.Iordache/Documents/D300 to XML Final CI/D300 to XML 2/storage/D300.xml", "w").write(text)
	except:
		pass
	try:
		amount=temp['D390 for XML']
		info=temp['Other info']

		an=info.cell(row=2,column=3).value
		luna=info.cell(row=3,column=3).value
		cif=info.cell(row=5,column=3).value
		denu=info.cell(row=4,column=3).value
		telefon=info.cell(row=7,column=3).value
		email=info.cell(row=9,column=3).value
		pren=info.cell(row=54,column=3).value
		nume=info.cell(row=55,column=3).value
		funct=info.cell(row=56,column=3).value
		adresa=info.cell(row=6,column=3).value
		numar=info.cell(row=102,column=3).value
		localitate=info.cell(row=103,column=3).value
		judet=info.cell(row=104,column=3).value

		sector=info.cell(row=105,column=3).value
		if(sector==None):
			sector=""
		bloc=info.cell(row=106,column=3).value
		if(bloc==None):
			bloc=""
		scara=info.cell(row=107,column=3).value
		if(scara==None):
			scara=""
		etaj=info.cell(row=108,column=3).value
		if(etaj==None):
			etaj=""
		apt=info.cell(row=109,column=3).value
		if(apt==None):
			apt=""
		adresa=adresa

		totalplata=amount.cell(row=11,column=3).value
		nrpag=amount.cell(row=12,column=3).value
		nropi=amount.cell(row=13,column=3).value
		bazal=amount.cell(row=14,column=3).value
		bazat=amount.cell(row=15,column=3).value
		bazaa=amount.cell(row=16,column=3).value
		bazaP=amount.cell(row=17,column=3).value
		bazas=amount.cell(row=18,column=3).value
		bazar=amount.cell(row=19,column=3).value
		totalb=amount.cell(row=23,column=3).value

		for row in amount.iter_rows():
				for cell in row:
					if cell.value=="TIP":
						tipc=cell.column
						tipr=cell.row


		for row in amount.iter_rows():
				for cell in row:
					if cell.value=="ŢARA":
						tarc=cell.column
						tarr=cell.row
		for row in amount.iter_rows():
				for cell in row:
					if cell.value=="COD OPERATOR INTRACOMUNITAR":
						coic=cell.column
						coir=cell.row

		for row in amount.iter_rows():
				for cell in row:
					if cell.value=="Denumire":
						denc=cell.column
						denr=cell.row

		for row in amount.iter_rows():
				for cell in row:
					if cell.value=="BAZA IMPOZABILĂ":
						bazc=cell.column
						bazr=cell.row

		luntb=len(amount[tipc])

		try:
			tip=[b.value for b in amount[tipc][tipr:luntb+1]]
		except:
			flash("Please insert the correct header for 'TIP'")
			return render_template("index.html")


		try:
			tara=[b.value for b in amount[tarc][tipr:luntb+1]]
		except:
			flash("Please insert the correct header for 'ŢARA'")
			return render_template("index.html")

		try:
			cod=[b.value for b in amount[coic][tipr:luntb+1]]
		except:
			flash("Please insert the correct header for 'COD OPERATOR INTRACOMUNITAR'")
			return render_template("index.html")

		try:
			den=[b.value for b in amount[denc][tipr:luntb+1]]
		except:
			flash("Please insert the correct header for 'Denumire'")
			return render_template("index.html")

		try:
			baza=[b.value for b in amount[bazc][tipr:luntb+1]]
		except:
			flash("Please insert the correct header for 'BAZA IMPOZABILA'")
			return render_template("index.html")

		txt="390,"+str(luna)+","+str(an)+","+str(cif)+",#"+str(denu)+"#,#"+str(adresa)+"#,#"+str(telefon)+"#,#"+str(email)+"#"+"\n"

		string=""

		text=""
		for i in range(0,len(tip)):
			if(tip[i]!= None):
				text1="#"+str(tip[i])+"#,#"+str(tara[i])+"#,#"+str(cod[i])+"#,#"+str(den[i])+"#,"+str(baza[i])+"\n"
				text=text+text1
				string=string+'<operatie tip="'+str(tip[i])+'" tara="'+str(tara[i])+'" codO="'+str(cod[i])+'" denO="'+str(den[i])+'" baza="'+str(baza[i])+'" />'
		# for i in range(0 ,len(tip))

		texttxt=txt+text

		#print(texttxt)
		textxml='<?xml version="1.0"?> <declaratie390  luna="'+str(luna)+'" an="'+str(an)+'" d_rec="0" nume_declar="'+str(nume)+'" prenume_declar="'+str(pren)+'" functie_declar="'+str(funct)+'" cui="'+str(cif)+'" den="'+str(denu)+'" adresa="'+str(adresa)+'" telefon="'+str(telefon)+'" mail="'+str(email)+'" totalPlata_A="'+str(totalplata)+'" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="mfp:anaf:dgti:d390:declaratie:v3 D390.xsd" xmlns="mfp:anaf:dgti:d390:declaratie:v3">'+'<rezumat nr_pag="'+str(nrpag)+'" nrOPI="'+str(nropi)+'" bazaL="'+str(bazal)+'" bazaT="'+str(bazat)+'" bazaA="'+str(bazaa)+'" bazaP="'+str(bazaP)+'" bazaS="'+str(bazas)+'" bazaR="'+str(bazar)+'" total_baza="'+str(totalb)+'" /> '+string+'</declaratie390>'
		# folderpath="D:/apps/TEST D3APPS/Test 21.03.2022/output"
		folderpath="/home/mirus_app/storage"
		# folderpath="C:/Users/Cristian.Iordache/Documents/D300 to XML Final CI/D300 to XML 2/storage"
		# f=open("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage/D390.txt", "w",encoding="utf-8").write(texttxt)
		f=open("/home/mirus_app/storage/D390.txt", "w",encoding="utf-8").write(texttxt)
	except:
		pass
	sheet1=temp['Other info']
	tipp=str(sheet1.cell(row=58,column=3).value)
	tippok=tipp[:1]
	an=sheet1.cell(row=2,column=3).value
	luna=sheet1.cell(row=3,column=3).value
	sisnormaldetva=sheet1.cell(row=61,column=3).value
	if(sisnormaldetva=="DA"):
		sistem=0
	else:
		sistem=1
	sistvalainc=sheet1.cell(row=62,column=3).value
	operinper=sheet1.cell(row=63,column=3).value
	opcuperafi=sheet1.cell(row=64,column=3).value
	if(opcuperafi=="Da"):	
		persafi=1
	else:
		persafi=0
	coddeinregistrare=sheet1.cell(row=65,column=3).value
	caen=sheet1.cell(row=66,column=3).value
	denumirefirma=sheet1.cell(row=67,column=3).value
	domiciliulfiscalfirma=sheet1.cell(row=6,column=3).value
	telefonfirma=sheet1.cell(row=69,column=3).value
	faxfirma=sheet1.cell(row=70,column=3).value
	emailfirma=sheet1.cell(row=71,column=3).value
	ciffirma=sheet1.cell(row=5,column=3).value
	denumireadmin=sheet1.cell(row=73,column=3).value
	domiciliulfiscaladmin=sheet1.cell(row=74,column=3).value
	telefonadmin=sheet1.cell(row=75,column=3).value
	faxadmin=sheet1.cell(row=76,column=3).value
	emailadmin=sheet1.cell(row=77,column=3).value
	cifadmin=sheet1.cell(row=78,column=3).value
	functiedecl=sheet1.cell(row=56,column=3).value
	try:
		facturi=temp['Sectiunea 2.1&2.2']
		val3=0
	except:
		val3=1

		# for row in facturi.iter_rows():
		# 	for cell in row:
		# 		if cell.value == "Serie":
		# 			row_det = cell.row
		# 			column_serie = cell.column
		# 			lun = len(facturi[cell.column])
		# serief = [b.value for b in facturi[column_serie][row_det:lun]]
	if(val3==0):
		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Inceput Emise":
					row_det = cell.row
					column_inceput = cell.column
					lun = len(facturi[cell.column])
		try:
			inceput = [b.value for b in facturi[column_inceput][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Inceput Alocate' in Facturi sheet")
			return render_template("index.html")

		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Final Emise":
					row_det = cell.row
					column_final = cell.column
					lun = len(facturi[cell.column])
		try:
			final = [b.value for b in facturi[column_final][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Final Alocate' in Facturi sheet")
			return render_template("index.html")
		
		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Tip Emise":
					row_det = cell.row
					column_final = cell.column
					lun = len(facturi[cell.column])
		try:
			tipfacturi = [b.value for b in facturi[column_final][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Final Alocate' in Facturi sheet")
			return render_template("index.html")			


		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Inceput Alocate":
					row_det = cell.row
					column_inceput = cell.column
					lun = len(facturi[cell.column])
		try:
			incepute = [b.value for b in facturi[column_inceput][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Inceput Alocate' in Facturi sheet")
			return render_template("index.html")

		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Final Alocate":
					row_det = cell.row
					column_final = cell.column
					lun = len(facturi[cell.column])
		try:
			finale = [b.value for b in facturi[column_final][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Final Alocate' in Facturi sheet")
			return render_template("index.html")
		
		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Tip Alocate":
					row_det = cell.row
					column_final = cell.column
					lun = len(facturi[cell.column])
		try:
			tipfacturie = [b.value for b in facturi[column_final][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Final Alocate' in Facturi sheet")
			return render_template("index.html")
		tranz=temp['Tranzactii']


		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Cui partener":
					row_det = cell.row
					column_cui = cell.column
					lun = len(tranz[cell.column])
		try:
			cuip= [b.value for b in tranz[column_cui][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Cui partener' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Nume partener":
					row_det = cell.row
					column_nume = cell.column
					lun = len(tranz[cell.column])
		try:
			nume= [b.value for b in tranz[column_nume][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Nume partener' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Tip tranzactie":
					row_det = cell.row
					column_tipt = cell.column
					lun = len(tranz[cell.column])
		try:
			tiptranza= [b.value for b in tranz[column_tipt][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Tip tranzactie' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Tip partener":
					row_det = cell.row
					column_tipt = cell.column
					lun = len(tranz[cell.column])
		try:
			tip_partener= [b.value for b in tranz[column_tipt][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Tip partener' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Baza TVA":
					row_det = cell.row
					column_btv = cell.column
					lun = len(tranz[cell.column])
		try:
			bazatv= [b.value for b in tranz[column_btv][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Baza TVA' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Cota TVA":
					row_det = cell.row
					column_ctva = cell.column
					lun = len(tranz[cell.column]) 
		try:
			cotatva= [b.value for b in tranz[column_ctva][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Cota TVA' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "TVA":
					row_det = cell.row
					column_ctva = cell.column
					lun = len(tranz[cell.column]) 
		try:
			stva= [b.value for b in tranz[column_ctva][row_det:lun]]
		except:
			flash("Please insert the correct header for 'TVA' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Nr Facturi":
					row_det = cell.row
					column_nrf = cell.column
					lun = len(tranz[cell.column])
		try:
			nrfacturi= [b.value for b in tranz[column_nrf][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Nr Facturi' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Neexigibile - nu se vor raporta":
					row_det = cell.row
					column_nrf = cell.column
					lun = len(tranz[cell.column])
		try:
			neex= [b.value for b in tranz[column_nrf][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Neexigibile - nu se vor raporta' in Tranzactii sheet")
			return render_template("index.html")
		
		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Cod si denumire NC produs(TIP V)":
					row_det = cell.row
					column_nrf = cell.column
					lun = len(tranz[cell.column])
		try:
			codv= [b.value for b in tranz[column_nrf][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Cod si denumire NC produs(TIP V)' in Tranzactii sheet")
			return render_template("index.html")

		# #print(asdasdasdadcuip,nume,tiptranza,bazatv,cotatva,nrfacturi)

		text=""

		sumaL5=0
		sumaL9=0
		sumaL19=0
		sumaL20=0
		sumaL24=0
		sumaA5=0
		sumaA9=0
		sumaA19=0
		sumaA20=0
		sumaA24=0

		sumaAI5=0
		sumaAI9=0
		sumaAI19=0
		sumaAI20=0
		sumaAI24=0

		sumaC5=0
		sumaC9=0
		sumaC19=0
		sumaC20=0
		sumaC24=0

		tvaC5=0
		tvaC9=0
		tvaC19=0
		tvaC20=0
		tvaC24=0

		tvaL5=0
		tvaL9=0
		tvaL19=0
		tvaL20=0
		tvaL24=0

		tvaA5=0
		tvaA9=0
		tvaA19=0
		tvaA20=0
		tvaA24=0

		tvaAI5=0
		tvaAI9=0
		tvaAI19=0
		tvaAI20=0
		tvaAI24=0

		nrL5=0
		nrL9=0
		nrL19=0
		nrL19=0
		nrL20=0
		nrL24=0
		nrA5=0
		nrA9=0
		nrA19=0
		nrA20=0
		nrA24=0

		nrAI5=0
		nrAI9=0
		nrAI19=0
		nrAI20=0
		nrAI24=0

		nrC5=0
		nrC9=0
		nrC19=0
		nrC20=0
		nrC24=0
		numar1=0

		nrV0=0
		nrLS0=0
		nrAS0=0
		valV0=0
		valLS0=0
		valAS0=0


		numarcui1=0
		# #print(tip_partener)
		# #print(cotatva)
		count=0
		#print(bazatv)
		tippartener1=[]
		for i in range(0,len(tip_partener)):
			if(str(tip_partener[i])=="1" and tiptranza[i]!="Not applicable for D394" and str(bazatv[i])!="0"):
				tippartener1.append(cuip[i])
			if(str(tip_partener[i])=="1" and neex[i]!="Yes"):

				if(bazatv[i]>0):
					if(str(cotatva[i])=="0"):
						if(str(tiptranza[i])=="V"):
							nrV0=nrV0+1
							valV0=valV0+bazatv[i]
						if(str(tiptranza[i])=="AS"):
							nrLS0=nrLS0+1
							valLS0=valLS0+bazatv[i]						
						if(str(tiptranza[i])=="LS"):
							nrAS0=nrAS0+1
							valAS0=valAS0+bazatv[i]							
					if(str(cotatva[i])=="5"):
						if(str(tiptranza[i])=="L"):

							nrL5=nrL5+int(nrfacturi[i])
							sumaL5=sumaL5+int(bazatv[i])
							tvaL5=tvaL5+int(stva[i])
						if(str(tiptranza[i])=="AI"):
						
							nrAI5=nrAI5+int(nrfacturi[i])
							sumaAI5=sumaAI5+int(bazatv[i])
							tvaAI5=tvaAI5+int(stva[i])
						if(str(tiptranza[i])=="A"):
						
							nrA5=nrA5+int(nrfacturi[i])
							sumaA5=sumaA5+int(bazatv[i])
							tvaA5=tvaA5+int(stva[i])
						if(str(tiptranza[i])=="LS"):
						
							nrLS5=nrLS5+int(nrfacturi[i])
							sumaLS5=sumaLS5+int(bazatv[i])
							tvaLS5=tvaLS5+int(stva[i])
						if(str(tiptranza[i])=="C"):
						
							nrC5=nrC5+int(nrfacturi[i])
							sumaC5=sumaC5+int(bazatv[i])
							tvaC5=tvaC5+int(stva[i])
					if(str(cotatva[i])=="9"):
						if(str(tiptranza[i])=="L"):
						
							nrL9=nrL9+int(nrfacturi[i])
							sumaL9=sumaL9+int(bazatv[i])
							tvaL9=tvaL9+int(stva[i])
						if(str(tiptranza[i])=="LS"):
						
							nrLS9=nrLS9+int(nrfacturi[i])
							sumaLS9=sumaLS9+int(bazatv[i])
							tvaLS9=tvaLS9+int(stva[i])
						if(str(tiptranza[i])=="AI"):
						
							nrAI9=nrAI9+int(nrfacturi[i])
							sumaAI9=sumaAI9+int(bazatv[i])
							tvaAI9=tvaAI9+int(stva[i])
						if(str(tiptranza[i])=="A"):
						
							nrA9=nrA9+int(nrfacturi[i])
							sumaA9=sumaA9+int(bazatv[i])
							tvaA9=tvaA9+int(stva[i])		
						if(str(tiptranza[i])=="C"):
						
							nrC9=nrC9+int(nrfacturi[i])
							sumaC9=sumaC9+int(bazatv[i])
							tvaC9=tvaC9+int(stva[i])
					if(str(cotatva[i])=="19"):
						# #print("YES")
						if(str(tiptranza[i])=="L"):
						
							nrL19=nrL19+int(int(nrfacturi[i]))
							sumaL19=sumaL19+int(bazatv[i])
							tvaL19=tvaL19+int(stva[i])
						if(str(tiptranza[i])=="LS"):
						
							nrLS19=nrLS19+int(nrfacturi[i])
							sumaLS19=sumaLS19+int(bazatv[i])
							tvaLS19=tvaLS19+int(stva[i])

						if(str(tiptranza[i])=="AI"):
						
							nrAI19=nrAI19+int(nrfacturi[i])
							sumaAI19=sumaAI19+int(bazatv[i])
							tvaAI19=tvaAI19+int(stva[i])
							#print(nrAI19)
						if(str(tiptranza[i])=="A"):
						
							#print(tiptranza[i],bazatv[i],stva[i])
							nrA19=nrA19+int(nrfacturi[i])
							sumaA19=sumaA19+int(bazatv[i])
							tvaA19=tvaA19+int(stva[i])
						if(str(tiptranza[i])=="C"):
						
							nrC19=nrC19+int(nrfacturi[i])
							sumaC19=sumaC19+int(bazatv[i])
							tvaC19=tvaC19+int(stva[i])
					if(str(cotatva[i])=="20"):
						if(str(tiptranza[i])=="L"):
						
							nrL20=nrL20+int(nrfacturi[i])
							sumaL20=sumaL20+int(bazatv[i])
							tvaL20=tvaL20+int(stva[i])
						if(str(tiptranza[i])=="LS"):
						
							nrLS20=nrLS20+int(nrfacturi[i])
							sumaLS20=sumaLS20+int(bazatv[i])
							tvaLS20=tvaLS20+int(stva[i])
						if(str(tiptranza[i])=="AI"):
						
							nrAI20=nrAI20+int(nrfacturi[i])
							sumaAI20=sumaAI20+int(bazatv[i])
							tvaAI20=tvaAI20+int(stva[i])
						if(str(tiptranza[i])=="A"):
						
							nrA20=nrA20+int(nrfacturi[i])
							sumaA20=sumaA20+int(bazatv[i])
							tvaA20=tvaA20+int(stva[i])		
						if(str(tiptranza[i])=="C"):
						
							nrC20=nrC20+int(nrfacturi[i])
							sumaC20=sumaC20+int(bazatv[i])
							tvaC20=tvaC20+int(stva[i])
					if(str(cotatva[i])=="24"):
						if(str(tiptranza[i])=="L"):
						
							nrL24=nrL24+int(nrfacturi[i])
							sumaL24=sumaL24+int(bazatv[i])
							tvaL24=tvaL24+int(stva[i])
						if(str(tiptranza[i])=="LS"):
						
							nrLS24=nrLS24+int(nrfacturi[i])
							sumaLS24=sumaLS24+int(bazatv[i])
							tvaLS24=tvaLS24+int(stva[i])		
						if(str(tiptranza[i])=="C"):
						
							nrC24=nrC24+int(nrfacturi[i])
							sumaC24=sumaC24+int(bazatv[i])
							tvaC24=tvaC24+int(stva[i])
						if(str(tiptranza[i])=="AI"):
						
							nrAI24=nrAI24+int(nrfacturi[i])
							sumaAI24=sumaAI24+int(bazatv[i])
							tvaAI24=tvaAI24+int(stva[i])
						if(str(tiptranza[i])=="A"):
						
							nrA24=nrA24+int(nrfacturi[i])
							sumaA24=sumaA24+int(bazatv[i])
							tvaA24=tvaA24+int(stva[i])
		numart=(list(set(tippartener1)))
		print(numart)
		numartotal=len(numart)
		#print(nrAI19)
		numarcui1=numartotal
		print(nrV0)
		if(nrL5>0 or nrA5>0 or nrAI5>0 or nrC5>0):
			text15='''<rezumat1 tip_partener="1" cota="5" facturiL="'''+str(nrL5)+'''" bazaL="'''+str(sumaL5)+'''" tvaL="'''+str(tvaL5)+'''" facturiA="'''+str(nrA5)+'''" bazaA="'''+str(sumaA5)+'''" tvaA="'''+str(tvaA5)+'''" facturiAI="'''+str(nrAI5)+'''" bazaAI="'''+str(sumaAI5)+'''" tvaAI="'''+str(tvaAI5)+'''" facturiC="'''+str(nrC5)+'''" bazaC="'''+str(sumaC5)+'''" tvaC="'''+str(tvaC5)+'''"/>'''
		else:
			text15=""
		if(nrL9>0 or nrA9>0 or nrAI5>0 or nrC9>0):
			text19='''<rezumat1 tip_partener="1" cota="9" facturiL="'''+str(nrL9)+'''" bazaL="'''+str(sumaL9)+'''" tvaL="'''+str(tvaL9)+'''" facturiA="'''+str(nrA9)+'''" bazaA="'''+str(sumaA9)+'''" tvaA="'''+str(tvaA9)+'''" facturiAI="'''+str(nrAI9)+'''" bazaAI="'''+str(sumaAI9)+'''" tvaAI="'''+str(tvaAI9)+'''" facturiC="'''+str(nrC9)+'''" bazaC="'''+str(sumaC9)+'''" tvaC="'''+str(tvaC9)+'''"/>'''
		else:
			text19=""
		if(nrL19>0 or nrA19>0 or nrAI9>0 or nrC19>0):
			text119='''<rezumat1 tip_partener="1" cota="19" facturiL="'''+str(nrL19)+'''" bazaL="'''+str(sumaL19)+'''" tvaL="'''+str(tvaL19)+'''" facturiA="'''+str(nrA19)+'''" bazaA="'''+str(sumaA19)+'''" tvaA="'''+str(tvaA19)+'''" facturiAI="'''+str(nrAI19)+'''" bazaAI="'''+str(sumaAI19)+'''" tvaAI="'''+str(tvaAI19)+'''" facturiC="'''+str(nrC19)+'''" bazaC="'''+str(sumaC19)+'''" tvaC="'''+str(tvaC19)+'''"/>'''
		else:
			text119=""
		if(nrL20>0 or nrA20>0 or nrAI20>0 or nrC20>0):
			text120='''<rezumat1 tip_partener="1" cota="20" facturiL="'''+str(nrL20)+'''" bazaL="'''+str(sumaL20)+'''" tvaL="'''+str(tvaL20)+'''" facturiA="'''+str(nrA20)+'''" bazaA="'''+str(sumaA20)+'''" tvaA="'''+str(tvaA20)+'''" facturiAI="'''+str(nrAI20)+'''" bazaAI="'''+str(sumaAI20)+'''" tvaAI="'''+str(tvaAI20)+'''" facturiC="'''+str(nrC20)+'''" bazaC="'''+str(sumaC20)+'''" tvaC="'''+str(tvaC20)+'''"/>'''
		else:
			text120=""
		if(nrL24>0 or nrA24>0 or nrAI24>0 or nrC24>0):
			text124='''<rezumat1 tip_partener="1" cota="24" facturiL="'''+str(nrL24)+'''" bazaL="'''+str(sumaL24)+'''" tvaL="'''+str(tvaL24)+'''" facturiA="'''+str(nrA24)+'''" bazaA="'''+str(sumaA24)+'''" tvaA="'''+str(tvaA24)+'''" facturiAI="'''+str(nrAI24)+'''" bazaAI="'''+str(sumaAI24)+'''" tvaAI="'''+str(tvaAI24)+'''" facturiC="'''+str(nrC24)+'''" bazaC="'''+str(sumaC24)+'''" tvaC="'''+str(tvaC24)+'''"/>'''
		else:
			text124=""
		text10=''
		if(nrV0>0 or nrAS0>0 or nrLS0>0):
			text10='''<rezumat1 tip_partener="1" cota="0" facturiLS="'''+str(nrLS0)+'''" bazaLS="'''+str(valLS0)+'''" facturiAS="'''+str(nrAS0)+'''" bazaAS="'''+str(valAS0)+'''" facturiV="'''+str(nrV0)+'''" bazaV="'''+str(valV0)+'''">
	<detaliu  bun="22" nrLivV="'''+str(nrV0)+'''" bazaLivV="'''+str(valV0)+'''"/>
	</rezumat1>'''
		sumaLS0=0
		sumaL5=0
		sumaL9=0
		sumaL19=0
		sumaL20=0
		sumaL24=0
		sumaLS5=0
		sumaLS9=0
		sumaLS19=0
		sumaLS20=0
		sumaLS24=0

		tvaL5=0
		tvaL9=0
		tvaL19=0
		tvaL20=0
		tvaL24=0

		tvaLS5=0
		tvaLS9=0
		tvaLS19=0
		tvaLS20=0
		tvaLS24=0

		nrN0=0
		bazaN0=0

		nrLS0=0
		nrL5=0
		nrL9=0
		nrL19=0
		nrL20=0
		nrL24=0
		nrLS5=0
		nrLS9=0
		nrLS19=0
		nrLS20=0
		nrLS24=0

		numarcui2=0
		for i in range(0,len(tip_partener)):
			if(str(tip_partener[i])=="2"):
				if(bazatv[i]>0):
					if(str(cotatva[i])=="5"):
						if(str(tiptranza[i])=="L"):
							nrL5=nrL5+int(nrfacturi[i])
							sumaL5=sumaL5+int(bazatv[i])
							tvaL5=tvaL5+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS5=nrLS5+int(nrfacturi[i])
							sumaLS5=sumaLS5+int(bazatv[i])
							tvaLS5=tvaLS5+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC5=nrC5+int(nrfacturi[i])
							sumaC5=sumaC5+int(bazatv[i])
							tvaC5=tvaC5+int(stva[i])
					if(str(cotatva[i])=="9"):
						if(str(tiptranza[i])=="L"):
							nrL9=nrL9+int(nrfacturi[i])
							sumaL9=sumaL9+int(bazatv[i])
							tvaL9=tvaL9+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS9=nrLS9+int(nrfacturi[i])
							sumaLS9=sumaLS9+int(bazatv[i])
							tvaLS9=tvaLS9+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC9=nrC9+int(nrfacturi[i])
							sumaC9=sumaC9+int(bazatv[i])
							tvaC9=tvaC9+int(stva[i])
					if(str(cotatva[i])=="19"):
						if(str(tiptranza[i])=="L"):
							nrL19=nrL19+int(nrfacturi[i])
							sumaL19=sumaL19+int(bazatv[i])
							tvaL19=tvaL19+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS19=nrLS19+int(nrfacturi[i])
							sumaLS19=sumaLS19+int(bazatv[i])
							tvaLS19=tvaLS19+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC19=nrC19+int(nrfacturi[i])
							sumaC19=sumaC19+int(bazatv[i])
							tvaC19=tvaC19+int(stva[i])
					if(str(cotatva[i])=="20"):
						if(str(tiptranza[i])=="L"):
							nrL20=nrL20+int(nrfacturi[i])
							sumaL20=sumaL20+int(bazatv[i])
							tvaL20=tvaL20+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS20=nrLS20+int(nrfacturi[i])
							sumaLS20=sumaLS20+int(bazatv[i])
							tvaLS20=tvaLS20+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC20=nrC20+int(nrfacturi[i])
							sumaC20=sumaC20+int(bazatv[i])
							tvaC20=tvaC20+int(stva[i])
					if(str(cotatva[i])=="24"):
						if(str(tiptranza[i])=="L"):
							nrL24=nrL24+int(nrfacturi[i])
							sumaL24=sumaL24+int(bazatv[i])
							tvaL24=tvaL24+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS24=nrLS24+int(nrfacturi[i])
							sumaLS24=sumaLS24+int(bazatv[i])
							tvaLS24=tvaLS24+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC24=nrC24+int(nrfacturi[i])
							sumaC24=sumaC24+int(bazatv[i])
							tvaC24=tvaC24+int(stva[i])	
		numarcui2=nrLS24+nrL24+nrLS20+nrL20+nrLS19+nrL19+nrLS9+nrL9+nrLS5+nrL5
		if(nrL5>0):
			text25='''<rezumat1 tip_partener="2" cota="5" facturiL="'''+str(nrL5)+'''" bazaL="'''+str(sumaL5)+'''" tvaL="'''+str(tvaL5)+'''"/>'''
		else:
			text25=""
		if(nrL9>0):
			text29='''<rezumat1 tip_partener="2" cota="9" facturiL="'''+str(nrL9)+'''" bazaL="'''+str(sumaL9)+'''" tvaL="'''+str(tvaL9)+'''"/>'''
		else:
			text29=""
		if(nrL19>0):
			text219='''<rezumat1 tip_partener="2" cota="19" facturiL="'''+str(nrL19)+'''" bazaL="'''+str(sumaL19)+'''" tvaL="'''+str(tvaL19)+'''"/>'''
		else:
			text219=""
		if(nrL20>0):
			text220='''<rezumat1 tip_partener="2" cota="20" facturiL="'''+str(nrL20)+'''" bazaL="'''+str(sumaL20)+'''" tvaL="'''+str(tvaL20)+'''"/>'''
		else:
			text220=""
		if(nrL24>0):
			text224='''<rezumat1 tip_partener="2" cota="24" facturiL="'''+str(nrL24)+'''" bazaL="'''+str(sumaL24)+'''" tvaL="'''+str(tvaL24)+'''"/>'''
		else:
			text224=""
		if(bazaN0>0):
			text20='''<rezumat1 tip_partener="2" cota="0" facturiN="'''+str(nrLS24)+'''" documentN="1" bazaN="'''+str(bazaN0)+'''"/>'''
		else:
			text20=""
		sumaL5=0
		sumaL9=0
		sumaL19=0
		sumaL20=0
		sumaL24=0
		sumaLS5=0
		sumaLS9=0
		sumaLS19=0
		sumaLS20=0
		sumaLS24=0

		sumaC5=0
		sumaC9=0
		sumaC19=0
		sumaC20=0
		sumaC24=0

		tvaC5=0
		tvaC9=0
		tvaC19=0
		tvaC20=0
		tvaC24=0

		tvaL5=0
		tvaL9=0
		tvaL19=0
		tvaL20=0
		tvaL24=0

		tvaLS5=0
		tvaLS9=0
		tvaLS19=0
		tvaLS20=0
		tvaLS24=0


		nrL5=0
		nrL9=0
		nrL19=0
		nrL20=0
		nrL24=0

		nrLS5=0
		nrLS9=0
		nrLS19=0
		nrLS20=0
		nrLS24=0

		nrC5=0
		nrC9=0
		nrC19=0
		nrC20=0
		nrC24=0
		numarcui3=0
		for i in range(0,len(tip_partener)):
			if(str(tip_partener[i])=="3"):
				if(bazatv[i]>0):
					if(str(cotatva[i])=="5"):
						if(str(tiptranza[i])=="L"):
							nrL5=nrL5+int(nrfacturi[i])
							sumaL5=sumaL5+int(bazatv[i])
							tvaL5=tvaL5+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS5=nrLS5+int(nrfacturi[i])
							sumaLS5=sumaLS5+int(bazatv[i])
							tvaLS5=tvaLS5+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC5=nrC5+int(nrfacturi[i])
							sumaC5=sumaC5+int(bazatv[i])
							tvaC5=tvaC5+int(stva[i])
					if(str(cotatva[i])=="9"):
						if(str(tiptranza[i])=="L"):
							nrL9=nrL9+int(nrfacturi[i])
							sumaL9=sumaL9+int(bazatv[i])
							tvaL9=tvaL9+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS9=nrLS9+int(nrfacturi[i])
							sumaLS9=sumaLS9+int(bazatv[i])
							tvaLS9=tvaLS9+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC9=nrC9+int(nrfacturi[i])
							sumaC9=sumaC9+int(bazatv[i])
							tvaC9=tvaC9+int(stva[i])
					if(str(cotatva[i])=="19"):
						if(str(tiptranza[i])=="L"):
							nrL19=nrL19+int(nrfacturi[i])
							sumaL19=sumaL19+int(bazatv[i])
							tvaL19=tvaL19+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS19=nrLS19+int(nrfacturi[i])
							sumaLS19=sumaLS19+int(bazatv[i])
							tvaLS19=tvaLS19+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC19=nrC19+int(nrfacturi[i])
							sumaC19=sumaC19+int(bazatv[i])
							tvaC19=tvaC19+int(stva[i])
					if(str(cotatva[i])=="20"):
						if(str(tiptranza[i])=="L"):
							nrL20=nrL20+int(nrfacturi[i])
							sumaL20=sumaL20+int(bazatv[i])
							tvaL20=tvaL20+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS20=nrLS20+int(nrfacturi[i])
							sumaLS20=sumaLS20+int(bazatv[i])
							tvaLS20=tvaLS20+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC20=nrC20+int(nrfacturi[i])
							sumaC20=sumaC20+int(bazatv[i])
							tvaC20=tvaC20+int(stva[i])
					if(str(cotatva[i])=="24"):
						if(str(tiptranza[i])=="L"):
							nrL24=nrL24+int(nrfacturi[i])
							sumaL24=sumaL24+int(bazatv[i])
							tvaL24=tvaL24+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS24=nrLS24+int(nrfacturi[i])
							sumaLS24=sumaLS24+int(bazatv[i])
							tvaLS24=tvaLS24+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC24=nrC24+int(nrfacturi[i])
							sumaC24=sumaC24+int(bazatv[i])
							tvaC24=tvaC24+int(stva[i])
		numarcui3=nrLS24+nrC24+nrL24+nrC20+nrLS20+nrL20+nrC19+nrLS19+nrL19+nrC19+nrLS9+nrL9+nrC9+nrLS5+nrL5+nrC5
		if(nrL5>0 or nrC5>0):
			text35='''<rezumat1 tip_partener="3" cota="5" facturiL="'''+str(nrL5)+'''" bazaL="'''+str(sumaL5)+'''" tvaL="'''+str(tvaL5)+'''" facturiC="'''+str(nrC5)+'''" bazaC="'''+str(sumaC5)+'''" tvaC="'''+str(tvaC5)+'''"/>'''
		else:
			text35=""
		if(nrL9>0 or nrC9>0):
			text39='''<rezumat1 tip_partener="3" cota="9" facturiL="'''+str(nrL9)+'''" bazaL="'''+str(sumaL9)+'''" tvaL="'''+str(tvaL9)+'''" facturiC="'''+str(nrC9)+'''" bazaC="'''+str(sumaC9)+'''" tvaC="'''+str(tvaC9)+'''"/>'''
		else:
			text39=""
		if(nrL19>0 or nrC19>0):	
			text319='''<rezumat1 tip_partener="3" cota="19" facturiL="'''+str(nrL19)+'''" bazaL="'''+str(sumaL19)+'''" tvaL="'''+str(tvaL19)+'''" facturiC="'''+str(nrC19)+'''" bazaC="'''+str(sumaC19)+'''" tvaC="'''+str(tvaC19)+'''"/>'''
		else:
			text319=""
		if(nrL20>0 or nrC20>0):	
			text320='''<rezumat1 tip_partener="3" cota="20" facturiL="'''+str(nrL20)+'''" bazaL="'''+str(sumaL20)+'''" tvaL="'''+str(tvaL20)+'''" facturiC="'''+str(nrC20)+'''" bazaC="'''+str(sumaC20)+'''" tvaC="'''+str(tvaC20)+'''"/>'''
		else:
			text320=""
		if(nrL24>0  or nrC24>0):	
			text324='''<rezumat1 tip_partener="3" cota="24" facturiL="'''+str(nrL24)+'''" bazaL="'''+str(sumaL24)+'''" tvaL="'''+str(tvaL24)+'''" facturiC="'''+str(nrC24)+'''" bazaC="'''+str(sumaC24)+'''" tvaC="'''+str(tvaC24)+'''"/>'''
		else:
			text324=""
		sumaL5=0
		sumaL9=0
		sumaL19=0
		sumaL20=0
		sumaL24=0
		sumaLS5=0
		sumaLS9=0
		sumaLS19=0
		sumaLS20=0
		sumaLS24=0

		sumaC5=0
		sumaC9=0
		sumaC19=0
		sumaC20=0
		sumaC24=0

		tvaC5=0
		tvaC9=0
		tvaC19=0
		tvaC20=0
		tvaC24=0

		tvaL5=0
		tvaL9=0
		tvaL19=0
		tvaL20=0
		tvaL24=0

		tvaLS5=0
		tvaLS9=0
		tvaLS19=0
		tvaLS20=0
		tvaLS24=0


		nrL5=0
		nrL9=0
		nrL19=0
		nrL20=0
		nrL24=0

		nrLS5=0
		nrLS9=0
		nrLS19=0
		nrLS20=0
		nrLS24=0

		nrC5=0
		nrC9=0
		nrC19=0
		nrC20=0
		nrC24=0
		numarcui4=0
		for i in range(0,len(tip_partener)):
			if(bazatv[i]>0):
				if(str(tip_partener[i])=="4"):
					if(str(cotatva[i])=="5"):
						if(str(tiptranza[i])=="L"):
							nrL5=nrL5+int(nrfacturi[i])
							sumaL5=sumaL5+int(bazatv[i])
							tvaL5=tvaL5+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS5=nrLS5+int(nrfacturi[i])
							sumaLS5=sumaLS5+int(bazatv[i])
							tvaLS5=tvaLS5+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC5=nrC5+int(nrfacturi[i])
							sumaC5=sumaC5+int(bazatv[i])
							tvaC5=tvaC5+int(stva[i])
					if(str(cotatva[i])=="9"):
						if(str(tiptranza[i])=="L"):
							nrL9=nrL9+int(nrfacturi[i])
							sumaL9=sumaL9+int(bazatv[i])
							tvaL9=tvaL9+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS9=nrLS9+int(nrfacturi[i])
							sumaLS9=sumaLS9+int(bazatv[i])
							tvaLS9=tvaLS9+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC9=nrC9+int(nrfacturi[i])
							sumaC9=sumaC9+int(bazatv[i])
							tvaC9=tvaC9+int(stva[i])
					if(str(cotatva[i])=="19"):
						if(str(tiptranza[i])=="L"):
							nrL19=nrL19+int(nrfacturi[i])
							sumaL19=sumaL19+int(bazatv[i])
							tvaL19=tvaL19+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS19=nrLS19+int(nrfacturi[i])
							sumaLS19=sumaLS19+int(bazatv[i])
							tvaLS19=tvaLS19+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC19=nrC19+int(nrfacturi[i])
							sumaC19=sumaC19+int(bazatv[i])
							tvaC19=tvaC19+int(stva[i])
					if(str(cotatva[i])=="20"):
						if(str(tiptranza[i])=="L"):
							nrL20=nrL20+int(nrfacturi[i])
							sumaL20=sumaL20+int(bazatv[i])
							tvaL20=tvaL20+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS20=nrLS20+int(nrfacturi[i])
							sumaLS20=sumaLS20+int(bazatv[i])
							tvaLS20=tvaLS20+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC20=nrC20+int(nrfacturi[i])
							sumaC20=sumaC20+int(bazatv[i])
							tvaC20=tvaC20+int(stva[i])
					if(str(cotatva[i])=="24"):
						if(str(tiptranza[i])=="L"):
							nrL24=nrL24+int(nrfacturi[i])
							sumaL24=sumaL24+int(bazatv[i])
							tvaL24=tvaL24+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS24=nrLS24+int(nrfacturi[i])
							sumaLS24=sumaLS24+int(bazatv[i])
							tvaLS24=tvaLS24+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC24=nrC24+int(nrfacturi[i])
							sumaC24=sumaC24+int(bazatv[i])
							tvaC24=tvaC24+int(stva[i])
		numarcui4=nrLS24+nrC24+nrL24+nrC20+nrLS20+nrL20+nrC19+nrLS19+nrL19+nrC19+nrLS9+nrL9+nrC9+nrLS5+nrL5+nrC5
		if(nrL5>0 or nrC5>0):
			text45='''<rezumat1 tip_partener="4" cota="5" facturiL="'''+str(nrL5)+'''" bazaL="'''+str(sumaL5)+'''" tvaL="'''+str(tvaL5)+'''" facturiC="'''+str(nrC5)+'''" bazaC="'''+str(sumaC5)+'''" tvaC="'''+str(tvaC5)+'''"/>'''
		else:
			text45=""
		if(nrL9>0 or nrC9>0):
			text49='''<rezumat1 tip_partener="4" cota="9" facturiL="'''+str(nrL9)+'''" bazaL="'''+str(sumaL9)+'''" tvaL="'''+str(tvaL9)+'''" facturiC="'''+str(nrC9)+'''" bazaC="'''+str(sumaC9)+'''" tvaC="'''+str(tvaC9)+'''"/>'''
		else:
			text49=""
		if(nrL19>0 or nrC19>0):	
			text419='''<rezumat1 tip_partener="4" cota="19" facturiL="'''+str(nrL19)+'''" bazaL="'''+str(sumaL19)+'''" tvaL="'''+str(tvaL19)+'''" facturiC="'''+str(nrC19)+'''" bazaC="'''+str(sumaC19)+'''" tvaC="'''+str(tvaC19)+'''"/>'''
		else:
			text419=""
		if(nrL20>0 or nrC20>0):	
			text420='''<rezumat1 tip_partener="4" cota="20" facturiL="'''+str(nrL20)+'''" bazaL="'''+str(sumaL20)+'''" tvaL="'''+str(tvaL20)+'''" facturiC="'''+str(nrC20)+'''" bazaC="'''+str(sumaC20)+'''" tvaC="'''+str(tvaC20)+'''"/>'''
		else:
			text420=""
		if(nrL24>0  or nrC24>0):	
			text424='''<rezumat1 tip_partener="4" cota="24" facturiL="'''+str(nrL24)+'''" bazaL="'''+str(sumaL24)+'''" tvaL="'''+str(tvaL24)+'''" facturiC="'''+str(nrC24)+'''" bazaC="'''+str(sumaC24)+'''" tvaC="'''+str(tvaC24)+'''"/>'''
		else:
			text424=""
		sheetG=temp['Sectiunea G. Manual input']
		sheetI=temp['Sectiunea I 1. Manual input']

		L24nr=0
		L20nr=0
		L19nr=0
		L9nr=0
		L5nr=0

		L24b=0
		L20b=0
		L19b=0
		L9b=0
		L5b=0

		L24t=0
		L20t=0
		L19t=0
		L9t=0
		L5t=0

		A24nr=0
		A20nr=0
		A19nr=0
		A9nr=0
		A5nr=0

		A24b=0
		A20b=0
		A19b=0
		A9b=0
		A5b=0

		A24t=0
		A20t=0
		A19t=0
		A9t=0
		A5t=0

		AI24nr=0
		AI20nr=0
		AI19nr=0
		AI9nr=0
		AI5nr=0

		AI24b=0
		AI20b=0
		AI19b=0
		AI9b=0
		AI5b=0

		AI24t=0
		AI20t=0
		AI19t=0
		AI9t=0
		AI5t=0




		for i in range(0,len(cotatva)):
			if(bazatv[i]):
				if(str(cotatva[i])=="24"):
					if(str(tiptranza[i])=="L"):
						L24nr=L24nr+int(nrfacturi[i])
						L24b=L24b+int(bazatv[i])
						L24t=L24t+int(stva[i])
					if(str(tiptranza[i])=="A"):
						A24nr=A24nr+int(nrfacturi[i])
						A24b=A24b+int(bazatv[i])
						A24t=A24t+int(stva[i])
					if(str(tiptranza[i])=="AI"):
						AI24nr=AI24nr+int(nrfacturi[i])
						AI24b=AI24b+int(bazatv[i])
						AI24t=AI24t+int(stva[i])
				if(str(cotatva[i])=="20"):
					if(str(tiptranza[i])=="L"):
						L20nr=L20nr+int(nrfacturi[i])
						L20b=L20b+int(bazatv[i])
						L20t=L20t+int(stva[i])
					if(str(tiptranza[i])=="A"):
						A20nr=A20nr+int(nrfacturi[i])
						A20b=A20b+int(bazatv[i])
						A20t=A20t+int(stva[i])
					if(str(tiptranza[i])=="AI"):
						AI20nr=AI20nr+int(nrfacturi[i])
						AI20b=AI20b+int(bazatv[i])
						AI20t=AI20t+int(stva[i])
				if(str(cotatva[i])=="19"):
					if(str(tiptranza[i])=="L"):
						L19nr=L19nr+int(nrfacturi[i])
						L19b=L19b+int(bazatv[i])
						L19t=L19t+int(stva[i])
					if(str(tiptranza[i])=="A"):
						A19nr=A19nr+int(nrfacturi[i])
						A19b=A19b+int(bazatv[i])
						A19t=A19t+int(stva[i])
					if(str(tiptranza[i])=="AI"):
						AI19nr=AI19nr+int(nrfacturi[i])
						AI19b=AI19b+int(int(bazatv[i]))
						AI19t=AI19t+int(int(stva[i]))
				if(str(cotatva[i])=="9"):
					if(str(tiptranza[i])=="L"):
						L9nr=L9nr+int(nrfacturi[i])
						L9b=L9b+int(int(bazatv[i]))
						L9t=L9t+int(int(stva[i]))
					if(str(tiptranza[i])=="A"):
						A9nr=A9nr+int(nrfacturi[i])
						A9b=A9b+int(int(bazatv[i]))
						A9t=A9t+int(int(stva[i]))
					if(str(tiptranza[i])=="AI"):
						AI9nr=AI9nr+int(nrfacturi[i])
						AI9b=AI9b+int(int(bazatv[i]))
						AI9t=AI9t+int(int(stva[i]))
				if(str(cotatva[i])=="5"):
					if(str(tiptranza[i])=="L"):
						L5nr=L5nr+int(nrfacturi[i])
						L5b=L5b+int(int(bazatv[i]))
						L5t=L5t+int(int(stva[i]))
					if(str(tiptranza[i])=="A"):
						A5nr=A5nr+int(nrfacturi[i])
						A5b=A5b+int(int(bazatv[i]))
						A5t=A5t+int(int(stva[i]))
					if(str(tiptranza[i])=="AI"):
						AI5nr=AI5nr+int(nrfacturi[i])
						AI5b=AI5b+int(int(bazatv[i]))
						AI5t=AI5t+int(int(stva[i]))

		rez224='''<rezumat2 cota="24"  bazaFSLcod="'''+str(sheetI.cell(row=8,column=2).value)+'''" TVAFSLcod="'''+str(sheetI.cell(row=8,column=3).value)+'''" bazaFSL="'''+str(sheetI.cell(row=18,column=2).value)+'''" TVAFSL="'''+str(sheetI.cell(row=18,column=3).value)+'''" bazaFSA="'''+str(sheetI.cell(row=28,column=2).value)+'''" TVAFSA="'''+str(sheetI.cell(row=28,column=3).value)+'''" bazaFSAI="'''+str(sheetI.cell(row=38,column=2).value)+'''" TVAFSAI="'''+str(sheetI.cell(row=38,column=3).value)+'''" bazaBFAI="'''+str(sheetI.cell(row=48,column=2).value)+'''" TVABFAI="'''+str(sheetI.cell(row=48,column=3).value)+'''" nrFacturiL="'''+str(L24nr)+'''" bazaL="'''+str(L24b)+'''" tvaL="'''+str(L24t)+'''" nrFacturiA="'''+str(A24nr)+'''" bazaA="'''+str(A24b)+'''" tvaA="'''+str(A24t)+'''" nrFacturiAI="'''+str(AI24nr)+'''" bazaAI="'''+str(AI24b)+'''" tvaAI="'''+str(AI24t)+'''" bazaL_PF="0" tvaL_PF="0" />'''
		rez220='''<rezumat2 cota="20"  bazaFSLcod="'''+str(sheetI.cell(row=9,column=2).value)+'''" TVAFSLcod="'''+str(sheetI.cell(row=9,column=3).value)+'''" bazaFSL="'''+str(sheetI.cell(row=19,column=2).value)+'''" TVAFSL="'''+str(sheetI.cell(row=19,column=3).value)+'''" bazaFSA="'''+str(sheetI.cell(row=29,column=2).value)+'''" TVAFSA="'''+str(sheetI.cell(row=29,column=3).value)+'''" bazaFSAI="'''+str(sheetI.cell(row=39,column=2).value)+'''" TVAFSAI="'''+str(sheetI.cell(row=39,column=3).value)+'''" bazaBFAI="'''+str(sheetI.cell(row=49,column=2).value)+'''" TVABFAI="'''+str(sheetI.cell(row=49,column=3).value)+'''" nrFacturiL="'''+str(L20nr)+'''" bazaL="'''+str(L20b)+'''" tvaL="'''+str(L20t)+'''" nrFacturiA="'''+str(A20nr)+'''" bazaA="'''+str(A20b)+'''" tvaA="'''+str(A20t)+'''" nrFacturiAI="'''+str(AI20nr)+'''" bazaAI="'''+str(AI20b)+'''" tvaAI="'''+str(AI20t)+'''" baza_incasari_i1="'''+str(sheetG.cell(row=14,column=2).value)+'''" tva_incasari_i1="'''+str(sheetG.cell(row=14,column=3).value)+'''" baza_incasari_i2="'''+str(sheetG.cell(row=24,column=2).value)+'''" tva_incasari_i2="'''+str(sheetG.cell(row=24,column=3).value)+'''" bazaL_PF="0" tvaL_PF="0"/>'''
		rez219='''<rezumat2 cota="19"  bazaFSLcod="'''+str(sheetI.cell(row=10,column=2).value)+'''" TVAFSLcod="'''+str(sheetI.cell(row=10,column=3).value)+'''" bazaFSL="'''+str(sheetI.cell(row=20,column=2).value)+'''" TVAFSL="'''+str(sheetI.cell(row=20,column=3).value)+'''" bazaFSA="'''+str(sheetI.cell(row=30,column=2).value)+'''" TVAFSA="'''+str(sheetI.cell(row=30,column=3).value)+'''" bazaFSAI="'''+str(sheetI.cell(row=40,column=2).value)+'''" TVAFSAI="'''+str(sheetI.cell(row=40,column=3).value)+'''" bazaBFAI="'''+str(sheetI.cell(row=50,column=2).value)+'''" TVABFAI="'''+str(sheetI.cell(row=50,column=3).value)+'''" nrFacturiL="'''+str(L19nr)+'''" bazaL="'''+str(L19b)+'''" tvaL="'''+str(L19t)+'''" nrFacturiA="'''+str(A19nr)+'''" bazaA="'''+str(A19b)+'''" tvaA="'''+str(A19t)+'''" nrFacturiAI="'''+str(AI19nr)+'''" bazaAI="'''+str(AI19b)+'''" tvaAI="'''+str(AI19t)+'''" baza_incasari_i1="'''+str(sheetG.cell(row=15,column=2).value)+'''" tva_incasari_i1="'''+str(sheetG.cell(row=15,column=3).value)+'''" baza_incasari_i2="'''+str(sheetG.cell(row=25,column=2).value)+'''" tva_incasari_i2="'''+str(sheetG.cell(row=25,column=3).value)+'''" bazaL_PF="0" tvaL_PF="0"/>'''
		rez29='''<rezumat2 cota="9"  bazaFSLcod="'''+str(sheetI.cell(row=11,column=2).value)+'''" TVAFSLcod="'''+str(sheetI.cell(row=11,column=3).value)+'''" bazaFSL="'''+str(sheetI.cell(row=21,column=2).value)+'''" TVAFSL="'''+str(sheetI.cell(row=21,column=3).value)+'''" bazaFSA="'''+str(sheetI.cell(row=31,column=2).value)+'''" TVAFSA="'''+str(sheetI.cell(row=31,column=3).value)+'''" bazaFSAI="'''+str(sheetI.cell(row=41,column=2).value)+'''" TVAFSAI="'''+str(sheetI.cell(row=41,column=3).value)+'''" bazaBFAI="'''+str(sheetI.cell(row=51,column=2).value)+'''" TVABFAI="'''+str(sheetI.cell(row=51,column=3).value)+'''" nrFacturiL="'''+str(L9nr)+'''" bazaL="'''+str(L9b)+'''" tvaL="'''+str(L9t)+'''" nrFacturiA="'''+str(A9nr)+'''" bazaA="'''+str(A9b)+'''" tvaA="'''+str(A9t)+'''" nrFacturiAI="'''+str(AI9nr)+'''" bazaAI="'''+str(AI9b)+'''" tvaAI="'''+str(AI9t)+'''" baza_incasari_i1="'''+str(sheetG.cell(row=16,column=2).value)+'''" tva_incasari_i1="'''+str(sheetG.cell(row=16,column=3).value)+'''" baza_incasari_i2="'''+str(sheetG.cell(row=26,column=2).value)+'''" tva_incasari_i2="'''+str(sheetG.cell(row=26,column=3).value)+'''" bazaL_PF="0" tvaL_PF="0"/>'''
		rez25='''<rezumat2 cota="5"  bazaFSLcod="'''+str(sheetI.cell(row=12,column=2).value)+'''" TVAFSLcod="'''+str(sheetI.cell(row=12,column=3).value)+'''" bazaFSL="'''+str(sheetI.cell(row=22,column=2).value)+'''" TVAFSL="'''+str(sheetI.cell(row=22,column=3).value)+'''" bazaFSA="'''+str(sheetI.cell(row=32,column=2).value)+'''" TVAFSA="'''+str(sheetI.cell(row=32,column=3).value)+'''" bazaFSAI="'''+str(sheetI.cell(row=42,column=2).value)+'''" TVAFSAI="'''+str(sheetI.cell(row=42,column=3).value)+'''" bazaBFAI="'''+str(sheetI.cell(row=52,column=2).value)+'''" TVABFAI="'''+str(sheetI.cell(row=52,column=3).value)+'''" nrFacturiL="'''+str(L5nr)+'''" bazaL="'''+str(L5b)+'''" tvaL="'''+str(L5t)+'''" nrFacturiA="'''+str(A5nr)+'''" bazaA="'''+str(A5b)+'''" tvaA="'''+str(A5t)+'''" nrFacturiAI="'''+str(AI5nr)+'''" bazaAI="'''+str(AI5b)+'''" tvaAI="'''+str(AI5t)+'''" baza_incasari_i1="'''+str(sheetG.cell(row=17,column=2).value)+'''" tva_incasari_i1="'''+str(sheetG.cell(row=17,column=3).value)+'''" baza_incasari_i2="'''+str(sheetG.cell(row=27,column=2).value)+'''" tva_incasari_i2="'''+str(sheetG.cell(row=27,column=3).value)+'''" bazaL_PF="0" tvaL_PF="0"/>'''
		totalplata394=numarcui1+numarcui2+numarcui3+numarcui4+L24b+L20b+L19b+L9b+L5b+A5b+A9b+A19b+A20b+A24b+AI5b+AI9b+AI19b+AI20b+AI24b
		#print(totalplata394)
		textinfo='''<?xml version="1.0"?><declaratie394 luna="'''+str(luna)+'''" an="'''+str(an)+'''" tip_D394="'''+str(tippok)+'''" sistemTVA="'''+str(sistem)+'''" op_efectuate="1" prsAfiliat="'''+str(persafi)+'''" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="mfp:anaf:dgti:d394:declaratie:v4 D394.xsd" xmlns="mfp:anaf:dgti:d394:declaratie:v4" cui="'''+str(coddeinregistrare)+'''" den="'''+str(denumirefirma)+'''" adresa="'''+str(domiciliulfiscalfirma)+'''" telefon="'''+str(telefonfirma)+'''" mail="'''+str(emailfirma)+'''" caen="'''+str(caen)+'''" totalPlata_A="'''+str(int(totalplata394))+'''" denR="'''+str(denumireadmin)+'''" functie_reprez="'''+str(functiedecl)+'''" adresaR="'''+str(domiciliulfiscaladmin)+'''" tip_intocmit="0" den_intocmit="Grant Thornton " cif_intocmit="27512924" calitate_intocmit="IMPUTERNICIT" optiune="1" schimb_optiune="1">
		<informatii nrCui1="'''+str(numarcui1)+'''" nrCui2="'''+str(numarcui2)+'''" nrCui3="'''+str(numarcui3)+'''" nrCui4="'''+str(numarcui4)+'''" nr_BF_i1="'''+str(sheetG.cell(row=5,column=2).value)+'''" incasari_i1="'''+str(sheetG.cell(row=6,column=2).value)+'''" incasari_i2="'''+str(sheetG.cell(row=7,column=2).value)+'''" nrFacturi_terti="0" nrFacturi_benef="0" nrFacturi="175" nrFacturiL_PF="0" nrFacturiLS_PF="0" val_LS_PF="0" tvaDedAI24="0" tvaDedAI20="0" tvaDedAI19="0" tvaDedAI9="0" tvaDedAI5="0" incasari_ag="0" costuri_ag="0" marja_ag="0" tva_ag="0" pret_vanzare="0" pret_cumparare="0" marja_antic="0" tva_antic="0" solicit="0"/>'''
		text=text+textinfo+text15+text10+text19+text119+text120+text124+text25+text29+text219+text220+text224+text20+text35+text39+ text319+text320+text324+text45+text49+text419+text420+text424+rez224+rez220+rez219+rez29+rez25+"\n"


		#print(tipfacturi)
		for k in range(0,len(inceput)):
			#print(tipfacturi[k],"aici")
			if(inceput[k]!=None and tipfacturi[k]==1):
				text=text+'''<serieFacturi tip="1" nrI="'''+str(inceput[k])+''' " nrF="'''+str(final[k])+''' "/>'''+"\n"
		for kk in range(0,len(inceput)):
			#print(inceput[kk],"aici emise")
			if(inceput[kk]!=None and tipfacturi[kk]==2):
					text=text+'''<serieFacturi tip="2" nrI="'''+str(inceput[kk])+''' " nrF="'''+str(final[kk])+''' "/>'''+"\n"
		sheet7=temp[' Sectiunea 7 ']
		for j in range(13,18):
			if(sheet7.cell(row=j,column=2).value!=None):
				text=text+'<lista  caen="'+str(sheet7.cell(row=4,column=1).value)+'" cota="'+str(sheet7.cell(row=6,column=1).value)+'" operat="'+str(sheet7.cell(row=4,column=2).value)+'" valoare="'+str(sheet7.cell(row=4,column=3).value)+'" tva="'+str(sheet7.cell(row=j,column=2).value)+'" />'+"\n"

		sheet2=temp['Facturi storno si anulate']
		# de modificat seria sa o bagam in functie de faptul ca e sau nu blnak
		for k in range(4,sheet2.max_row):
			if(sheet2.cell(row=k,column=1).value=="Stornata"):
				text=text+'<facturi  tip_factura="1" nr="'''+str(sheet2.cell(row=k,column=3).value)+'''"/>'''+"\n"
		numep=[]
		for j in range(0,len(nume)):

			numep.append(nume[j].replace("&",""))
			# nume[j].replace("<","")
			# nume[j].replace(">","")
		for i in range(0,len(tiptranza)):
			if(tiptranza[i]!="Not applicable for D394"):
				if(tiptranza[i]=="V"):
					text=text+'<op1 tip="'+str(tiptranza[i])+'" tip_partener="'+str(tip_partener[i])+'" cota="'+str(cotatva[i])+'" cuiP="'+str(cuip[i])+'" denP="'+str(numep[i])+'"  nrFact="'+str(int(nrfacturi[i]))+'" baza="'+str(int(bazatv[i]))+'">'+'<op11  nrFactPR="'+str(int(nrfacturi[i]))+'" codPR="'+str(codv[i])+'" bazaPR="'+str(int(bazatv[i]))+'" /> </op1>'+"\n"
				else:

					if("-" in str(cuip[i]) and "RO" in str(cuip[i])):
						text=text+'<op1  tip="'+str(tiptranza[i])+'" tip_partener="'+str(tip_partener[i])+'" cota="'+str(cotatva[i])+'" denP="'+str(numep[i])+'" taraP="'+str(cuip[i][:2])+'" locP="'+str(cuip[i][3:])+'" judP="'+str(cuip[i][3:])+'" nrFact="'+str(int(nrfacturi[i]))+'" baza="'+str(int(bazatv[i]))+'" tva="'+str(int(stva[i]))+'" />'+"\n"
					else:
						if("-" in str(cuip[i])):
							text=text+'<op1  tip="'+str(tiptranza[i])+'" tip_partener="'+str(tip_partener[i])+'" cota="'+str(cotatva[i])+'" denP="'+str(numep[i])+'" taraP="'+str(cuip[i][:2])+'" locP="'+str(cuip[i][3:])+'" nrFact="'+str(int(nrfacturi[i]))+'" baza="'+str(int(bazatv[i]))+'" tva="'+str(int(stva[i]))+'" />'+"\n"
						else:
							if(bazatv[i]>0):
								text=text+'<op1 tip="'+str(tiptranza[i])+'" tip_partener="'+str(tip_partener[i])+'" cota="'+str(cotatva[i])+'" cuiP="'+str(cuip[i])+'" denP="'+str(numep[i])+'"  nrFact="'+str(int(nrfacturi[i]))+'" baza="'+str(int(bazatv[i]))+'" tva="'+str(int(stva[i]))+'"/>'+"\n"

		text=text+"</declaratie394>"
		# text='<?xml version="1.0"?><declaratie394 luna="'+str(luna)+'" an="'+str(an)+'" tip_D394="'+str(tip)+'" sistemTVA="'+str(sisnormaldetva)+'" op_efectuate="'+str(op_efectuate)+'" prsAfiliat="'+str(prsAfiliat)+'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="mfp:anaf:dgti:d394:declaratie:v3 D394.xsd" xmlns="mfp:anaf:dgti:d394:declaratie:v3" cui="'+str(cui)+'" den="'+str(den)+""

		# f=open("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage/D394.xml", "w",encoding='utf-8').write(text)
		# make_archive("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage","C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/arhiva VAT apps.zip")
		# return send_from_directory("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2","arhiva VAT apps.zip",as_attachment=True) 
		f=open("/home/mirus_app/storage/D394.xml", "w",encoding='utf-8').write(text)
	make_archive("/home/mirus_app/storage","/home/mirus_app/storage/arhiva VAT apps.zip")
	return send_from_directory("/home/mirus_app/storage","arhiva VAT apps.zip",as_attachment=True)



#====================================================================THALES=============================================================================================








	
if __name__ == '__main__':
   app.run()