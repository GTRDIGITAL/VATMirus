from flask import Flask, render_template, request, send_from_directory , url_for
from jinja2 import pass_eval_context
from numpy import append
import shutil
import re
import PyPDF2
from werkzeug.utils import secure_filename
from tkinter import *
from flask import flash
from openpyxl.worksheet.datavalidation import DataValidation 
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
#import pypiwin32
import itertools as it
import datetime
from datetime import datetime
import tkinter as tk
from tkinter.filedialog import askopenfilename
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl import *
import openpyxl
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import  Font
from openpyxl.styles import  Color
from openpyxl.styles import Alignment
import requests
import time

from tkinter import filedialog

from openpyxl.chart import (
    LineChart,
    Series,
    Reference,
)

from openpyxl.cell import Cell
from openpyxl.descriptors import (
    String,
    Sequence,
    Integer,
)
from openpyxl.descriptors.serialisable import Serialisable
import sys, string, os
import os
from flask import Flask, render_template, request, send_from_directory
from datetime import datetime
from werkzeug.utils import secure_filename
import PyPDF2 as pf
import tkinter
# import PIL
from tkinter import *
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import *
from openpyxl.styles import Border, Side
from openpyxl import Workbook
from openpyxl.cell import cell
from openpyxl.descriptors import (
    String,
    Sequence,
    Integer,
)
from openpyxl.chart import LineChart, Reference
import string
# from PIL import ImageTk, Image
from tkinter.filedialog import askopenfilename
from openpyxl.styles import Color, PatternFill, Font, borders
from openpyxl.worksheet.dimensions import ColumnDimension


# from openpyxl.styles import colors
# from PIL import ImageTk, Image
# import pandas as pd
from openpyxl.styles import Alignment, alignment

import os
import xml.etree.ElementTree as ET

import base64
from tkinter import filedialog
import io

import webbrowser

def make_archive(source, destination):
        base = os.path.basename(destination)
        name = base.split('.')[0]
        format = base.split('.')[1]
        archive_from = os.path.dirname(source)
        archive_to = os.path.basename(source.strip(os.sep))
        shutil.make_archive(name, format, archive_from, archive_to)
        shutil.move('%s.%s'%(name,format), destination)

thin = Side(border_style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_left = Border(left=thin, right=None, top=thin, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_right = Border(left=None, right=thin, top=thin, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_centered = Border(left=None, right=None, top=thin, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_upperleft = Border(left=thin, top=thin)

thin = Side(border_style='thin', color='000000')
border_lowerleft = Border(left=thin, right=None, top=None, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_upperright = Border(right=thin, top=thin)

thin = Side(border_style='thin', color='000000')
border_lowerright = Border(right=thin, bottom=thin)

thin = Side(border_style='thin', color='000000')
border_left1 = Border(left=thin)

thin = Side(border_style='thin', color='000000')
border_right1 = Border(right=thin)

thin = Side(border_style='thin', color='000000')
border_top = Border(top=thin)

thin = Side(border_style='thin', color='000000')
border_bottom = Border(bottom=thin)

app=Flask(__name__, template_folder='template')
app.secret_key = "GT ROMANIA Delivery Center"

@app.route('/')
def my_form():
    return render_template('D3APPS dashboard.html')

global LL_g
@app.route('/D3APPS')
def my_form_D300():
	return render_template('D3APPS.html')

@app.route('/D3APPS', methods=['POST', 'GET'])
def D300xml():
	if request.method == 'POST':
		clientname=request.form.get('client')
		D300 = request.files["far"]
		val1 = request.form.get('D300')
		val2 = request.form.get('D390')
		val3 = request.form.get('D394')
		val4 = request.form.get('xyz')
		soldLunaTrecuta = request.form.get('largeAm')
		#print(soldLunaTrecuta)
	if val1=="":
		#print("Da")  # daca e bifat
		val1 = 1
	else:
		#print(val1)            
		val1 = 0
		#print("Nu")

	if val2=="":  # daca e bifat
		val2 = 1
	else:
		#print(val2)            
		val2 = 0

	if val3=="":  # daca e bifat
		val3 = 1
	else:
		#print(val3)            
		val3 = 0
	#print(val4)
	#print(val4,"--------------------------------------------")		
	cap_tabel_color_verde = PatternFill(start_color = '00B050', end_color ='00B050', fill_type = 'solid')
	cap_tabel_color_verde_deschis = PatternFill(start_color = '92D050', end_color ='92D050', fill_type = 'solid')
	cap_tabel_color_black = PatternFill(start_color = '000000', end_color ='000000', fill_type = 'solid')
	cap_tabel = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
	cap_tabelbold = Font(name='Calibri', size=10, bold=True)
	cap_tabeltitlu = Font(name='Calibri', size=15, bold=True,underline='single')
	scrisincredibildemare = Font(name='Calibri', size=30, bold=True)
	cap_tabel_color_GT_movdeschis = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')
	cap_tabel_color_GT_movinchis = PatternFill(start_color='CCC0DA', end_color='CCC0DA', fill_type='solid')

	temp = openpyxl.load_workbook(D300,data_only=True)
	sales=temp['Sales']
	purchases=temp['Purchases']
	info=temp['Other info']
	if(val1==1):
		sheetinutil1=temp.create_sheet('D300--->>>')
		sheetinutil1.sheet_view.showGridLines=False
		sheetinutil1.cell(row=2,column=1).value="Switch to next sheet for D300 Workings draft"
		sheetinutil1.cell(row=2,column=1).font=scrisincredibildemare
		amount=temp.create_sheet('D300 draft figures')
		amount.freeze_panes = 'A8'
		amount.auto_filter.ref = "A7:G71"
		amount.sheet_view.showGridLines = False

		amount.cell(row=6, column=2).value="1"
		amount.cell(row=6, column=3).value="2"
		amount.cell(row=7, column=1).value="Row"
		amount.cell(row=7, column=2).value="Taxable basis"
		amount.cell(row=7, column=3).value="VAT amount"
		amount.cell(row=7, column=4).value="Comments"
		amount.cell(row=7, column=5).value="Journal Source"
		amount.cell(row=7, column=6).value="Flag Suma Control"
		amount.cell(row=7, column=7).value="Suma Control"
		amount.cell(row=8, column=1).value="1"
		amount.cell(row=9, column=1).value="2"
		amount.cell(row=10, column=1).value="3"
		amount.cell(row=11, column=1).value="3.1"
		amount.cell(row=12, column=1).value="4"
		amount.cell(row=13, column=1).value="5"
		amount.cell(row=14, column=1).value="5.1"
		amount.cell(row=15, column=1).value="6"
		amount.cell(row=16, column=1).value="7"
		amount.cell(row=17, column=1).value="7.1"
		amount.cell(row=18, column=1).value="8"
		amount.cell(row=19, column=1).value="9"
		amount.cell(row=20, column=1).value="10"
		amount.cell(row=21, column=1).value="11"
		amount.cell(row=22, column=1).value="12"
		amount.cell(row=23, column=1).value="12.1"
		amount.cell(row=24, column=1).value="12.2"
		amount.cell(row=25, column=1).value="12.3"
		amount.cell(row=26, column=1).value="13"
		amount.cell(row=27, column=1).value="14"
		amount.cell(row=28, column=1).value="15"
		amount.cell(row=29, column=1).value="16"
		amount.cell(row=30, column=1).value="17"
		amount.cell(row=31, column=1).value="18"
		amount.cell(row=32, column=1).value="19"
		amount.cell(row=33, column=1).value="20"
		amount.cell(row=34, column=1).value="20.1"
		amount.cell(row=35, column=1).value="21"
		amount.cell(row=36, column=1).value="22"
		amount.cell(row=37, column=1).value="22.1"
		amount.cell(row=38, column=1).value="23"
		amount.cell(row=39, column=1).value="24"
		amount.cell(row=40, column=1).value="25"
		amount.cell(row=41, column=1).value="26"
		amount.cell(row=42, column=1).value="27"
		amount.cell(row=43, column=1).value="27.1"
		amount.cell(row=44, column=1).value="27.2"
		amount.cell(row=45, column=1).value="27.3"
		amount.cell(row=46, column=1).value="28"
		amount.cell(row=47, column=1).value="29"
		amount.cell(row=48, column=1).value="30"
		amount.cell(row=49, column=1).value="30.1"
		amount.cell(row=50, column=1).value="31"
		amount.cell(row=51, column=1).value="31.1"
		amount.cell(row=52, column=1).value="32"
		amount.cell(row=53, column=1).value="33"
		amount.cell(row=54, column=1).value="34"
		amount.cell(row=55, column=1).value="35"
		amount.cell(row=56, column=1).value="36"
		amount.cell(row=57, column=1).value="37"
		amount.cell(row=58, column=1).value="38"
		amount.cell(row=59, column=1).value="39"
		amount.cell(row=60, column=1).value="40"
		amount.cell(row=61, column=1).value="41"
		amount.cell(row=62, column=1).value="42"
		amount.cell(row=63, column=1).value="43"
		amount.cell(row=64, column=1).value="44"
		amount.cell(row=65, column=1).value="45"
		amount.cell(row=66, column=1).value="46"

		amount.cell(row=68, column=1).value="A"
		amount.cell(row=69, column=1).value="A1"
		amount.cell(row=70, column=1).value="B"
		amount.cell(row=71, column=1).value="B.1"

		for a in range(8, 13):
			amount.cell(row=a, column=2).value='=ROUND(SUMIF(Sales!$7:$7,$A{0}&"."&B$6,Sales!$5:$5),0)'.format(a)
		amount.cell(row=10, column=2).value='=ROUND(SUMIF(Sales!$7:$7,$A10&"."&B$6,Sales!$5:$5)-SUMIF(Sales!$7:$7,$A12&"."&B$6,Sales!$5:$5),0)'.format(1)

		for b in range(13, 19):
			amount.cell(row=b, column=2).value='=ROUND(SUMIF(Purchases!$7:$7,$A{0}&"."&B$6,Purchases!$5:$5),0)'.format(b)
		amount.cell(row=13, column=2).value='=ROUND(SUMIF(Purchases!$7:$7,$A13&"."&B$6,Purchases!$5:$5)-SUMIF(Purchases!$7:$7,$A15&"."&B$6,Purchases!$5:$5),0)'.format(1)
		amount.cell(row=16, column=2).value='=ROUND(SUMIF(Purchases!$7:$7,$A16&"."&B$6,Purchases!$5:$5)-SUMIF(Purchases!$7:$7,$A18&"."&B$6,Purchases!$5:$5),0)'.format(1)
		
		for c in range(19, 22):
			amount.cell(row=c, column=2).value='=ROUND(SUMIF(Sales!$7:$7,$A{0}&"."&B$6,Sales!$5:$5),0)'.format(c)
		
		amount.cell(row=22, column=2).value='=SUM(B23:B25)'

		for d in range(23, 26):
			amount.cell(row=d, column=2).value='=ROUND(SUMIF(Purchases!$7:$7,$A{0}&"."&B$6,Purchases!$5:$5),0)'.format(d)

		for e in range(26, 31):
			amount.cell(row=e, column=2).value='=ROUND(SUMIF(Sales!$7:$7,$A{0}&"."&B$6,Sales!$5:$5),0)'.format(e)

		amount.cell(row=31, column=2).value=0
		amount.cell(row=32, column=2).value='=SUM(B8:B31)-B11-B14-B17-B23-B24-B25'
		amount.cell(row=33, column=2).value='=B13'
		amount.cell(row=34, column=2).value='=B14'
		amount.cell(row=35, column=2).value='=B15'
		amount.cell(row=36, column=2).value='=B16'
		amount.cell(row=37, column=2).value='=B17'
		amount.cell(row=38, column=2).value='=B18'

		for f in range(39, 42):
			amount.cell(row=f, column=2).value='=ROUND(SUMIF(Purchases!$7:$7,$A{0}&"."&B$6,Purchases!$5:$5),0)'.format(f)
		
		amount.cell(row=42, column=2).value='=SUM(B43:B45)'
		amount.cell(row=43, column=2).value='=B23'
		amount.cell(row=44, column=2).value='=B24'
		amount.cell(row=45, column=2).value='=B25'
		amount.cell(row=46, column=2).value=0
		amount.cell(row=47, column=2).value=0
		amount.cell(row=48, column=2).value='=ROUND(SUMIF(Purchases!$7:$7,$A48&"."&B$6,Purchases!$5:$5),0)'
		amount.cell(row=49, column=2).value=0
		amount.cell(row=50, column=2).value='=SUM(B33:B47)-B34-B37-SUM(B43:B45)'
		amount.cell(row=51, column=2).value='=ROUND(SUMIF(Purchases!$7:$7,$A51&"."&B$6,Purchases!$5:$5),0)'
		amount.cell(row=52, column=2).value='=B50-B51'
		amount.cell(row=53, column=2).value=0
		amount.cell(row=54, column=2).value='=ROUND(SUMIF(Purchases!$7:$7,$A54&"."&B$6,Purchases!$5:$5),0)'
		amount.cell(row=55, column=2).value=0
		amount.cell(row=56, column=2).value='=SUM(B52:B55)'
		amount.cell(row=57, column=2).value=0
		amount.cell(row=58, column=2).value=0
		amount.cell(row=59, column=2).value=0
		amount.cell(row=60, column=2).value=0
		amount.cell(row=61, column=2).value='=SUM(B58:B60)'
		amount.cell(row=62, column=2).value=0
		amount.cell(row=63, column=2).value=0
		amount.cell(row=64, column=2).value='=B57+B62+B63'
		amount.cell(row=65, column=2).value='=IF((B61-B664)<0,0,B61-B64)'
		amount.cell(row=66, column=2).value='=IF((B64-B61)<0,0,B64)'

		amount.cell(row=68, column=2).value=0
		amount.cell(row=69, column=2).value=0
		amount.cell(row=70, column=2).value='=ROUND(SUMIF(Purchases!$7:$7,$A70&"."&B$6,Purchases!$5:$5),0)'
		amount.cell(row=71, column=2).value='=B70'
		
		#coloana TVA----------------------------------------------------

		for g in range(8, 13):
			amount.cell(row=g, column=3).value='=ROUND(SUMIF(Sales!$7:$7,$A{0}&"."&C$6,Sales!$5:$5),0)'.format(g)
		
		amount.cell(row=10, column=3).value='=ROUND(SUMIF(Sales!$7:$7,$A10&"."&C$6,Sales!$5:$5)-SUMIF(Sales!$7:$7,$A12&"."&C$6,Sales!$5:$5),0)'

		for h in range(13, 19):
			amount.cell(row=h, column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A{0}&"."&C$6,Purchases!$5:$5),0)'.format(h)
		
		amount.cell(row=13, column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A13&"."&C$6,Purchases!$5:$5)-SUMIF(Purchases!$7:$7,$A15&"."&C$6,Purchases!$5:$5),0)'
		
		# amount.cell(row=16,column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A16&"."&C$6,Purchases!$5:$5)-SUMIF(Purchases!$7:$7,$A18&"."&C$6,Purchases!$5:$5),0)'

		for i in range(19, 23):
			amount.cell(row=i, column=3).value='=ROUND(SUMIF(Sales!$7:$7,$A{0}&"."&C$6,Sales!$5:$5),0)'.format(i)
		
		amount.cell(row=22, column=3).value='=SUM(C23:C25)'

		for j in range(23, 26):
			amount.cell(row=j, column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A{0}&"."&C$6,Purchases!$5:$5),0)'.format(j)
		
		for k in range(26, 31):
			amount.cell(row=k, column=3).value='=ROUND(SUMIF(Sales!$7:$7,$A{0}&"."&C$6,Sales!$5:$5),0)'.format(k)
		
		# amount.cell(row=16, column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A12&"."&C$2,Purchases!$5:$5),0)-ROUND(SUMIF(Purchases!$7:$7,$A14&"."&C$2,Purchases!$5:$5),0)'
		amount.cell(row=16, column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A16&"."&C$6,Purchases!$5:$5),0)-ROUND(SUMIF(Purchases!$7:$7,$A18&"."&C$6,Purchases!$5:$5),0)'
		amount.cell(row=31, column=3).value=0
		amount.cell(row=32, column=3).value='=SUM(C8:C31)-C11-C14-C17-C23-C24-C25'
		amount.cell(row=33, column=3).value='=C13'
		amount.cell(row=34, column=3).value='=C14'
		amount.cell(row=35, column=3).value='=C15'
		amount.cell(row=36, column=3).value='=C16'
		amount.cell(row=37, column=3).value='=C17'
		amount.cell(row=38, column=3).value='=C18'

		for l in range(39, 42):
			amount.cell(row=l, column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A{0}&"."&C$6,Purchases!$5:$5),0)'.format(l)

		amount.cell(row=42, column=3).value='=SUM(C43:C45)'
		amount.cell(row=43, column=3).value='=C23'
		amount.cell(row=44, column=3).value='=C24'
		amount.cell(row=45, column=3).value='=C25'
		amount.cell(row=46, column=3).value=0
		amount.cell(row=47, column=3).value=0
		amount.cell(row=48, column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A48&"."&C$6,Purchases!$5:$5),0)'
		amount.cell(row=49, column=3).value=0
		amount.cell(row=50, column=3).value='=SUM(C33:C47)-C34-C37-SUM(C43:C45)'
		amount.cell(row=51, column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A51&"."&C$6,Purchases!$5:$5),0)'
		amount.cell(row=52, column=3).value='=C50-C51'
		amount.cell(row=53, column=3).value=0
		amount.cell(row=54, column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A54&"."&C$6,Purchases!$5:$5),0)'
		amount.cell(row=55, column=3).value=0
		amount.cell(row=56, column=3).value='=SUM(C52:C55)'
		amount.cell(row=57, column=3).value='=IF((C56-C32)<0,0,C56-C32)'
		amount.cell(row=58, column=3).value='=IF((C32-C56)<0,0,C32-C56)'
		amount.cell(row=59, column=3).value=0
		amount.cell(row=60, column=3).value=0
		amount.cell(row=61, column=3).value='=SUM(C58:C60)'
		
		if soldLunaTrecuta == None or soldLunaTrecuta == "" or soldLunaTrecuta == " ":
			amount.cell(row=62, column=3).value=0
		else:
			amount.cell(row=62, column=3).value=int(soldLunaTrecuta)
		#print(soldLunaTrecuta, "sold luna trecuta")
		amount.cell(row=63, column=3).value=0
		amount.cell(row=64, column=3).value='=C57+C62+C63'
		amount.cell(row=65, column=3).value='=IF((C61-C64)<0,0,C61-C64)'
		amount.cell(row=66, column=3).value='=IF((C64-C61)<0,0,C64-C61)'
		amount.cell(row=68, column=3).value=0
		amount.cell(row=69, column=3).value=0

		amount.cell(row=70, column=3).value='=ROUND(SUMIF(Purchases!$7:$7,$A70&"."&C$6,Purchases!$5:$5),0)'
		amount.cell(row=71, column=3).value='=C70'

		amount.cell(row=73, column=1).value='Informații privind valoarea totală, fără TVA, a operațiunilor prevăzute la art. 2781 alin. (1) lit. b) din Codul fiscal, respectiv a vânzărilor intracomunitare de bunuri la distanță și a prestărilor de servicii de telecomunicaţii, de radiodifuziune şi televiziune, precum și servicii furnizate pe cale electronică, către persoane neimpozabile din alte state membre UE'
		amount.cell(row=73, column=2).value='Total an precedent'
		amount.cell(row=73, column=3).value='An curent (inclusiv perioada de raportare)'

		amount.cell(row=74, column=2).value=0
		amount.cell(row=74, column=3).value=0

		amount.cell(row=22, column=5).value='Total'

		for m in it.chain(range(8, 13), range(19, 22), range(26, 32)):
			amount.cell(row=m, column=5).value='SALES'

		for n in it.chain(range(13, 19), range(23, 26), range(39, 42), range(46, 50), range(53, 56)):
			amount.cell(row=n, column=5).value='Purchases'
		
		for o in range(32, 39):
			amount.cell(row=o, column=5).value='Total'

		for p in range(42, 46):
			amount.cell(row=p, column=5).value='Total'
		
		amount.cell(row=50, column=5).value='Total'
		amount.cell(row=51, column=5).value='Purchases'
		amount.cell(row=52, column=5).value='Total'
		amount.cell(row=66, column=5).value='Purchases'

		for q in range(68, 71):
			amount.cell(row=q, column=5).value='Total'

		for r in range(56, 67):
			amount.cell(row=r, column=5).value='Total'

		amount.cell(row=71, column=5).value='Purchases'

		for s in it.chain(range(8, 13), range(26, 29), range(48, 50)):
			amount.cell(row=s, column=6).value='no VAT'

		for t in it.chain(range(13, 26), range(29, 46)):
			amount.cell(row=t, column=6).value='Add all'

		for u in it.chain(range(46, 48), range(52, 54), range(55, 67)):
			amount.cell(row=u, column=6).value='No basis'
		
		amount.cell(row=50, column=6).value='Add all'
		amount.cell(row=51, column=6).value='Se pune 0'
		amount.cell(row=54, column=6).value='Add all'

		for a1 in range(8, 13):
			amount.cell(row=a1, column=7).value='=B{0}'.format(a1)

		for b1 in range(13, 26):
			amount.cell(row=b1, column=7).value='=SUM(B{0}:C{0})'.format(b1)

		for c1 in range(26, 29):
			amount.cell(row=c1, column=7).value='=B{0}'.format(c1)
		
		for d1 in range(29, 46):
			amount.cell(row=d1, column=7).value='=SUM(B{0}:C{0})'.format(d1)

		amount.cell(row=46, column=7).value='=C46'
		amount.cell(row=47, column=7).value='=C47'
		amount.cell(row=48, column=7).value='=B48'
		amount.cell(row=49, column=7).value='=B49'
		amount.cell(row=50, column=7).value='=SUM(B50:C50)'
		amount.cell(row=51, column=7).value='=C51'
		amount.cell(row=52, column=7).value='=C52'
		amount.cell(row=53, column=7).value='=SUM(B53:C53)'
		amount.cell(row=54, column=7).value='=SUM(B54:C54)'
		amount.cell(row=2,column=1).value="D300 draft figures"
		amount.cell(row=2,column=1).font=cap_tabeltitlu
		amount.row_dimensions[5].hidden = True
		for e1 in range(55, 67):
			amount.cell(row=e1, column=7).value='=C{0}'.format(e1)
		
		for row in amount['B8:B74']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'
		
		for row in amount['C8:C74']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'

		#------foratare D300

		for row in amount['A7:C7']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.alignment=Alignment(horizontal='center',vertical='center')
		for row in amount['D7:G7']:
			for cell in row:
				cell.fill=cap_tabel_color_black
		
		for row in amount['A7:G7']:
			for cell in row:
				cell.font=cap_tabel

		listanoua=['A','B','C','D','E','F','G']

		for column in listanoua:
			for i in listanoua:
				if (column==i):
					amount.column_dimensions[column].width = 17
		
		amount.column_dimensions['D'].hidden = True
		amount.column_dimensions['E'].hidden = True
		amount.column_dimensions['F'].hidden = True
		amount.column_dimensions['G'].hidden = True

		# for row in range(4, 63):
		# 	for cell in row:
		# 		cell.number_format='#,##0_);(#,##0)'

		#-----------------------SALES



		sales.cell(row=11, column=150).value='L'
		sales.cell(row=11, column=151).value='P'
		sales.cell(row=11, column=152).value='T'

		
		listL_index=get_column_letter(150)		
		listP_index=get_column_letter(151)		
		listT_index=get_column_letter(152)

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Codul de inregistrare in scopuri de TVA al clientului Client VAT ID":
					rand_tb = cell.row
					cuiClient = cell.column
					lun = len(sales[cell.column])
		try:
			listaCUISales = [b.value for b in sales[cuiClient][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Codul de inregistrare in scopuri de TVA al clientului Client VAT ID' in Sales sheet")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Denumirea clientului Client name":
					rand_tb = cell.row
					client = cell.column
					lun = len(sales[cell.column])
		try:
			listaClient = [b.value for b in sales[client][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Denumirea clientului Client name' in Sales sheet")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-LIC-Taxable base (RON)- Intra-community supplies of goods":
					rand_tb = cell.row
					bazal = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaL = [b.value for b in sales[bazal][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-LIC-Taxable base (RON)- Intra-community supplies of goods in Sales sheet'")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services":
					rand_tb = cell.row
					bazaP = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaP = [b.value for b in sales[bazaP][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services' in Sales sheet")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-Livrare in cadul unei operatiuni triunghiulara-Taxable basis (RON)-Supplies within a triangular transaction":
					rand_tb = cell.row
					bazaT = cell.column
					lun = len(sales[cell.column])
		try:
			listBazaT = [b.value for b in sales[bazaT][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services' in Sales sheet")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Baza facturi in urma inspectiei fiscale (RON)-Invoice base following tax inspection (RON)":
					rand_tb = cell.row
					bazaFI = cell.column
					lun = len(sales[cell.column])
		try:
			bazaFacturiInspectie = [b.value for b in sales[bazaFI][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza facturi in urma inspectiei fiscale (RON)-Invoice base following tax inspection (RON)' in Sales sheet")
			return render_template("index.html")
		

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "TVA facturi in urma inspectiei fiscale (RON)-VAT invoices following tax inspection (RON)":
					rand_tb = cell.row
					bazaTFI = cell.column
					lun = len(sales[cell.column])
		try:
			tvaFacturiInspectie = [b.value for b in sales[bazaTFI][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA facturi in urma inspectiei fiscale (RON)-VAT invoices following tax inspection (RON)' in Sales sheet")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Baza facturi emise in urma reactivarii codului de TVA (RON)-Base of invoices issued following the reactivation of the VAT code (RON)":
					rand_tb = cell.row
					bazaFR = cell.column
					lun = len(sales[cell.column])
		try:
			bazaFacturiReactivate = [b.value for b in sales[bazaFR][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza facturi emise in urma reactivarii codului de TVA (RON)-Base of invoices issued following the reactivation of the VAT code (RON)' in Sales sheet")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "TVA facturi emise in urma reactivarii codului de TVA (RON)-VAT invoices issued following the reactivation of the VAT code (RON)":
					rand_tb = cell.row
					bazaTFR = cell.column
					lun = len(sales[cell.column])
		try:
			tvaFacturiReactivate = [b.value for b in sales[bazaTFR][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA facturi emise in urma reactivarii codului de TVA (RON)-VAT invoices issued following the reactivation of the VAT code (RON)' in Sales sheet")
			return render_template("index.html")
		
		listBazaL2=[]
		for c in range(0, len(listBazaL)):
			if listBazaL[c] == None:
				listBazaL2.append(0)
			else:
				listBazaL2.append(listBazaL[c])

		listBazaP2=[]
		for c in range(0, len(listBazaP)):
			if listBazaP[c] == None:
				listBazaP2.append(0)
			else:
				listBazaP2.append(listBazaP[c])

		listaBazaT2=[]
		for c in range(0, len(listBazaT)):
			if listBazaT[c] == None:
				listaBazaT2.append(0)
			else:
				listaBazaT2.append(listBazaT[c])
		

		bazaFacturiInspectie1=[]
		
		for i in range(0, len(bazaFacturiInspectie)):
			if bazaFacturiInspectie[i] == None:
				bazaFacturiInspectie1.append(0)
			else:
				bazaFacturiInspectie1.append(bazaFacturiInspectie[i])
		#print(bazaFacturiInspectie1, "BAZA FACTURI INSPECTIE")

		tvaFacturiInspectie1=[]

		for i in range(0, len(tvaFacturiInspectie)):
			if tvaFacturiInspectie[i] == None:
				tvaFacturiInspectie1.append(0)
			else:
				tvaFacturiInspectie1.append(tvaFacturiInspectie[i])

		bazaFacturiReactivate1=[]
		for i in range(0, len(bazaFacturiReactivate)):
			if bazaFacturiReactivate[i] == None:
				bazaFacturiReactivate1.append(0)
			else:
				bazaFacturiReactivate1.append(bazaFacturiReactivate[i])
		
		tvaFacturiReactivate1=[]
		for i in range(0, len(tvaFacturiReactivate)):
			if tvaFacturiReactivate[i] == None:
				tvaFacturiReactivate1.append(0)
			else:
				tvaFacturiReactivate1.append(tvaFacturiReactivate[i])
		# #print(bazaFacturiInspectie1, 'baza fact ins')
		lenbfi=0
		for i in bazaFacturiInspectie1:
			if i != 0:
				sumabfi=sum(bazaFacturiInspectie1)
				lenbfi+=1
		sumaBazaFactIns=s
		lentbfi=0
		for i in tvaFacturiInspectie1:
			if i != 0:
				sumatbfi=sum(tvaFacturiInspectie1)
				lentbfi+=1
		lenbfr=0
		for i in bazaFacturiReactivate1:
			if i != 0:
				sumabfr=sum(bazaFacturiReactivate1)
				lenbfr+=1

		lentbfr=0
		for i in tvaFacturiReactivate1:
			if i != 0:
				sumatbfr=sum(tvaFacturiReactivate1)
				lentbfr+=1

		listL=[]
		for q in listBazaL2:
			if q != 0:
				listL.append('L')
			else:
				listL.append('')
				# listL.remove('')
		# #print(listL, 'bazaL')

		listP=[]
		for q in listBazaP2:
			if q != 0:
				listP.append('P')
			else:
				listP.append('')
				# listP.remove('')
		
		listT=[]
		for q in listaBazaT2:
			if q != 0:
				listT.append('T')
			else:
				listT.append('')
				# listT.remove('')
		
		for a in range(0, len(listL)):
			sales.cell(row=12+a, column=150).value=listL[a]
		
		for a in range(0, len(listP)):
			sales.cell(row=12+a, column=151).value=listP[a]
		
		for a in range(0, len(listT)):
			sales.cell(row=12+a, column=152).value=listT[a]
		
		#-----------------------------PURCHASES
		


		purchases.cell(row=11, column=150).value='A'
		purchases.cell(row=11, column=151).value='S'

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Codul de înregistrare în scopuri de TVA VAT number":
					rand_tb = cell.row
					cuiPurchases = cell.column
					lun = len(purchases[cell.column])
		try:
			listaCUIPurchases = [b.value for b in purchases[cuiPurchases][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Codul de înregistrare în scopuri de TVA VAT number' in Purchases sheet")
			return render_template("index.html")

		# for row in purchases.iter_rows():
		# 	for cell in row:
		# 		if cell.value == "Bază de impozitare (RON)-AIC-Taxable base (RON)":
		# 			rand_tb = cell.row
		# 			bazaA = cell.column
		# 			lun = len(purchases[cell.column])
		# try:
		# 	listaBazaA = [b.value for b in purchases[bazaA][rand_tb:lun]]
		# except:
		# 	flash("Please insert the correct header for 'Bază de impozitare (RON)-AIC-Taxable base (RON)' in Purchases sheet")
		# 	return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Bază de impozitare (RON)-AIC-Taxable base (RON) (perioada curenta)":
					rand_tb = cell.row
					bazaA = cell.column
					lun = len(purchases[cell.column])
		try:
			listaBazaA = [b.value for b in purchases[bazaA][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Bază de impozitare (RON)-AIC-Taxable base (RON) (perioada curenta)' in Purchases sheet")
			return render_template("index.html")

		listaBazaA2=[]
		for c in range(0, len(listaBazaA)):
			if listaBazaA[c] == None:
				listaBazaA2.append(0)
			else:
				listaBazaA2.append(listaBazaA[c])

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Bază de impozitare (RON)-AIS-Taxable base (RON) (perioada curenta)":
					rand_tb = cell.row
					bazaS = cell.column
					lun = len(purchases[cell.column])
		try:
			listaBazaS = [b.value for b in purchases[bazaS][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Bază de impozitare (RON)-AIS-Taxable base (RON)' in Purchases sheet")
			return render_template("index.html")


		listaBazaS2=[]
		for c in range(0, len(listaBazaS)):
			if listaBazaS[c] == None:
				listaBazaS2.append(0)
			else:
				listaBazaS2.append(listaBazaS[c])

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Furnizor Supplier":
					rand_tb = cell.row
					furnizor = cell.column
					lun = len(purchases[cell.column])
		try:
			listaFurnizor = [b.value for b in purchases[furnizor][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Furnizor Supplier' in Purchases sheet")
			return render_template("index.html")

		# #print(listaFurnizor)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Baza facturi in urma inspectiei fiscale (RON)-Invoice base following tax inspection (RON)":
					rand_tb = cell.row
					bazaIF = cell.column
					lun = len(purchases[cell.column])
		try:
			bazaInsFisc = [b.value for b in purchases[bazaIF][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza facturi in urma inspectiei fiscale (RON)-Invoice base following tax inspection (RON)' in Purchases sheet")
			return render_template("index.html")
		# #print(bazaInsFisc, 'bazainsfisc')

		bazaInsFisc1=[]
		for c in range(0, len(bazaInsFisc)):
			if bazaInsFisc[c] == None:
				bazaInsFisc1.append(0)
			else:
				bazaInsFisc1.append(bazaInsFisc[c])
		#print(bazaInsFisc1, 'bazainsfis1')

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TVA facturi in urma inspectiei fiscale (RON)-VAT invoices following tax inspection (RON)":
					rand_tb = cell.row
					bazaTIF = cell.column
					lun = len(purchases[cell.column])
		try:
			tvaInsFisc = [b.value for b in purchases[bazaTIF][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA facturi in urma inspectiei fiscale (RON)-VAT invoices following tax inspection (RON)' in Purchases sheet")
			return render_template("index.html")

		tvaInsFisc1=[]
		for c in range(0, len(tvaInsFisc)):
			if tvaInsFisc[c] == None:
				tvaInsFisc1.append(0)
			else:
				tvaInsFisc1.append(tvaInsFisc[c])

		lenbif=0
		for i in bazaInsFisc1:
			if i != 0:
				sumabif=sum(bazaInsFisc1)
				lenbif+=1

		lentbif=0
		for i in tvaInsFisc1:
			if i != 0:
				sumatbif=sum(tvaInsFisc1)
				lentbif+=1

		listaMare=listaClient + listaFurnizor
		listaF=set(listaMare)
		listaUnicaPartener=list(listaF)

		listA=[]
		for i in listaBazaA2:
			if i != 0:
				listA.append('A')
			else:
				listA.append('')
				# listA.remove('')
		# #print(listA, 'lista A')
		# #print(len(listA), ' lungime lista A')
		
		listS=[]
		for q in listaBazaS2:
			if q != 0:
				listS.append('S')
			else:
				listS.append('')
				# listS.remove('')
		
		for a in range(0, len(listA)):
			purchases.cell(row=12+a, column=150).value=listA[a]
		listA_index=get_column_letter(150)

		for x in range(0, len(listS)):
			purchases.cell(row=12+x, column=151).value=listS[x]
		listS_index=get_column_letter(151)
		# #print(len(listS))
		# #print(listS)
		info=temp['Other info']
	
		listaMapare=["L", "T", "S", "A"]
		Poz="10"
		Fix01="01"
		Fix0000="0000"
		dictMapare={'L':'301', 'T':'302', 'S':'303', 'A':'304'}

		elementMapare=""
		if info.cell(row=53, column=3).value=="L":
			elementMapare="301"
		else:
			if info.cell(row=53, column=3).value == "T":
				elementMapare="302"
			else:
				if info.cell(row=53, column=3).value == "S":
					elementMapare="303"
				else:
					if info.cell(row=53, column=3).value == "A":
						elementMapare="304"
			
		# #print(elementMapare, 'MAPARE')

		LL=""
		AA=""

		if len(str(info.cell(row=3, column=3).value))==2:
			LL=str(info.cell(row=3, column=3).value)
			LL_g=LL
		else:
			LL="0"+str(info.cell(row=3, column=3).value)

		AA=str(info.cell(row=2, column=3).value)[2:]
		strYear=str(info.cell(row=2, column=3).value)[2:]
		intYear=int(strYear)+1
		year=str(intYear)
		
		# #print(year, 'an')

		LLAA=LL+AA
		LL2=""
		AA2=""

		if int(info.cell(row=3, column=3).value)==12:
			LL2="1"
		if len(str(info.cell(row=3, column=3).value))==1:
			LL2="0"+str(int(info.cell(row=3, column=3).value)+1)
		# #print(LL2, 'LL2')
		
		if int(info.cell(row=3, column=3).value)==12:
			AA2=year 
		else:
			AA2=strYear

		ZZLLAA="25"+str(LL2)+str(AA2)
		# #print(ZZLLAA, 'ZZLLAA')

		poz1=""
		LTSA=""
		fix1=""
		LLAA1=""
		ZZLLAA1=""
		fix01=""
		control=""

		poz1 = sum(int(digit) for digit in str(Poz))
		LTSA = sum(int(digit) for digit in str(elementMapare))
		fix1=sum(int(digit) for digit in str(Fix01))
		LLAA1=sum(int(digit) for digit in str(LLAA))
		ZZLLAA1=sum(int(digit) for digit in str(ZZLLAA))
		fix01=sum(int(digit) for digit in str(Fix0000))

		control=poz1+LTSA+fix1+LLAA1+ZZLLAA1+fix01
		#print(control,'control')
		

		nrEvidenta=Poz+elementMapare+Fix01+LLAA+ZZLLAA+Fix0000+str(control)
		#print(nrEvidenta, 'nrEvidenta')
			
		
		info.cell(row=18, column=3).value=nrEvidenta
		info.cell(row=52, column=3).value="=SUM('D300 draft figures'!G8:G66)"
		try:
			info.cell(row=50, column=3).value=lenbfi
			info.cell(row=51, column=3).value=sumabfi
			info.cell(row=52, column=3).value=sumatbfi
		except:
			pass
		try:
			info.cell(row=53, column=3).value=sumabif
			info.cell(row=54, column=3).value=sumatbif
			info.cell(row=55, column=3).value=lentbif
		except:
			pass
		try: 
			# info.cell(row=56, column=3).value=lenbfr
			info.cell(row=57, column=3).value=sumabfr
			info.cell(row=58, column=3).value=sumatbfr
		except:
			pass

	if(val2==1):
		sheetinutil2=temp.create_sheet('D390--->>>')
		sheetinutil2.sheet_view.showGridLines=False
		sheetinutil2.cell(row=2,column=1).value="Switch to next sheet for D390 Workings draft"
		sheetinutil2.cell(row=2,column=1).font=scrisincredibildemare		

		workings=temp.create_sheet('D390 workings')
		workings.cell(row=1,column=1).value="D390 workings"
		workings.cell(row=1,column=1).font=cap_tabelbold
		workings.freeze_panes = 'A4'
		workings.auto_filter.ref = "A3:I10000"
		workings.sheet_view.showGridLines = False
		workings.column_dimensions['I'].hidden = True

		workings.cell(row=3, column=1).value='TIP'
		workings.cell(row=3, column=2).value='ŢARA'
		workings.cell(row=3, column=3).value='COD OPERATOR INTRACOMUNITAR'
		workings.cell(row=3, column=4).value='DENUMIRE'
		workings.cell(row=3, column=5).value='BAZA IMPOZABILĂ'
		workings.cell(row=3, column=6).value='CIF'
		workings.cell(row=3, column=7).value='Country Code'
		workings.cell(row=3, column=8).value='BAZA IMPOZABILĂ'
		workings.cell(row=3, column=9).value='Cheie extragere - filtreaza 1'

		# bazaA_Furnizor=get_column_letter(7)
		# bazaA_index=get_column_letter(bazaA)
		# bazaA_literaA=get_column_letter(bazaA)
		# #print(bazaA_Furnizor, "litera pentru furnizor/ client")
		# #print(bazaA_index, "ASTA E NUMARUL LUI A")
		# #print(bazaA_literaA, "LITERA PENTRU TIP")
		
		a=3
		for x in range(0, len(listA)):
			if str(listA[x])=="A":
				a=a+1
				workings.cell(row=a, column=1).value=listA[x]
				workings.cell(row=a, column=4).value=listaFurnizor[x]
				workings.cell(row=a, column=6).value=listaCUIPurchases[x]
				workings.cell(row=a, column=3).value=listaCUIPurchases[x][2:]
				workings.cell(row=a, column=7).value=listaCUIPurchases[x][0:2]
				# workings.cell(row=a, column=8).value=listaBazaA[x]
				# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BH:BH,Purchases!CK:CK,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
				workings.cell(row=a, column=8).value="=SUMIFS(Purchases!"+str(bazaA)+":"+str(bazaA)+",Purchases!"+str(listA_index)+":"+str(listA_index)+",'D390 workings'!A{0}".format(a)+",Purchases!"+str(cuiPurchases)+":"+str(cuiPurchases)+",'D390 workings'!F{0})".format(a)
				xx="=SUMIFS(Purchases!"+str(bazaA)+":"+str(bazaA)+",Purchases!"+str(listA_index)+":"+str(listA_index)+",'D390 workings'!A{0}".format(a)+",Purchases!"+str(furnizor)+":"+str(furnizor)+",'D390 workings'!F{0})".format(a)
				#print(xx)
				workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
				workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)


		for x in range(0, len(listS)):
			if str(listS[x])=="S":
				a=a+1
				workings.cell(row=a, column=1).value=listS[x]
				workings.cell(row=a, column=4).value=listaFurnizor[x]
				workings.cell(row=a, column=6).value=listaCUIPurchases[x]
				workings.cell(row=a, column=3).value=listaCUIPurchases[x][2:]
				workings.cell(row=a, column=7).value=listaCUIPurchases[x][0:2]
				# workings.cell(row=a, column=8).value=listaBazaS[x]
				# workings.cell(row=a, column=8).value="=SUMIFS(Purchases!BQ:BQ,Purchases!CL:CL,'D390 workings'!A{0},Purchases!F:F,'D390 workings'!F{0})".format(a)
				workings.cell(row=a, column=8).value="=SUMIFS(Purchases!"+str(bazaS)+":"+str(bazaS)+",Purchases!"+str(listS_index)+":"+str(listS_index)+",'D390 workings'!A{0}".format(a)+",Purchases!"+str(cuiPurchases)+":"+str(cuiPurchases)+",'D390 workings'!F{0})".format(a)
				workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
				workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)


		for x in range(0, len(listT)):
			if str(listT[x])=="T":
				a=a+1
				workings.cell(row=a, column=1).value=listT[x]
				workings.cell(row=a, column=4).value=listaClient[x]
				workings.cell(row=a, column=6).value=listaCUISales[x]
				workings.cell(row=a, column=3).value=listaCUISales[x][2:]
				workings.cell(row=a, column=7).value=listaCUISales[x][0:2]
				# workings.cell(row=a, column=8).value=listBazaT[x]
				workings.cell(row=a, column=8).value="=SUMIFS(Sales!"+str(bazaT)+":"+str(bazaT)+",Sales!"+str(listT_index)+":"+str(listT_index)+",'D390 workings'!A{0}".format(a)+",Sales!"+str(cuiClient)+":"+str(cuiClient)+",'D390 workings'!F{0})".format(a)
				workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
				workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)

		for x in range(0, len(listL)):
			if str(listL[x])=="L":
				a=a+1
				workings.cell(row=a, column=1).value=listL[x]
				workings.cell(row=a, column=4).value=listaClient[x]
				workings.cell(row=a, column=6).value=listaCUISales[x]
				workings.cell(row=a, column=3).value=listaCUISales[x][2:]
				workings.cell(row=a, column=7).value=listaCUISales[x][0:2]
				# workings.cell(row=a, column=8).value=listBazaL[x]
				# workings.cell(row=a, column=8).value="=SUMIFS(Sales!AK:AK,Sales!BN:BN,'D390 workings'!A{0},Sales!G:G,'D390 workings'!F{0})".format(a)
				workings.cell(row=a, column=8).value="=SUMIFS(Sales!"+str(bazal)+":"+str(bazal)+",Sales!"+str(listL_index)+":"+str(listL_index)+",'D390 workings'!A{0}".format(a)+",Sales!"+str(cuiClient)+":"+str(cuiClient)+",'D390 workings'!F{0})".format(a)
				workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
				workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)

		for x in range(0, len(listP)):
			if str(listP[x])=="P":
				a=a+1
				workings.cell(row=a, column=1).value=listP[x]
				workings.cell(row=a, column=4).value=listaClient[x]
				workings.cell(row=a, column=6).value=listaCUISales[x]
				workings.cell(row=a, column=3).value=listaCUISales[x][2:]
				workings.cell(row=a, column=7).value=listaCUISales[x][0:2]
				# workings.cell(row=a, column=8).value=listBazaP[x]
				# workings.cell(row=a, column=8).value="=SUMIFS(Sales!AU:AU,Sales!BO:BO,'D390 workings'!A{0},Sales!G:G,'D390 workings'!F{0})".format(a)
				workings.cell(row=a, column=8).value="=SUMIFS(Sales!"+str(bazaP)+":"+str(bazaP)+",Sales!"+(listP_index)+":"+str(listP_index)+",'D390 workings'!A{0}".format(a)+",Sales!"+str(cuiClient)+":"+str(cuiClient)+",'D390 workings'!F{0})".format(a)
				workings.cell(row=a, column=5).value='=ROUND(H{0},0)'.format(a)
				workings.cell(row=a, column=9).value='=IF(F{0}=" "," ",COUNTIFS(F{0}:F10000,F{0},A{0}:A10000,A{0}))'.format(a)
		
		for row in workings.iter_rows():
			for cell in row:
				if cell.value == "TIP":
					rand_tb = cell.row
					tip = cell.column
					lun = len(workings[cell.column])
		try:
			listaTip = [b.value for b in workings[tip][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TIP' in Workings sheet")
			return render_template("index.html")

		for row in workings.iter_rows():
			for cell in row:
				if cell.value == "CIF":
					rand_tb = cell.row
					cod = cell.column
					lun = len(workings[cell.column])
		try:
			codPartener = [b.value for b in workings[cod][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'CIF' in Workings sheet")
			return render_template("index.html")


		for row in workings.iter_rows():
			for cell in row:
				if cell.value == "Country Code":
					rand_tb = cell.row
					country = cell.column
					lun = len(workings[cell.column])
		try:
			countryCode = [b.value for b in workings[country][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Country Code' in Workings sheet")
			return render_template("index.html")

		for row in workings.iter_rows():
			for cell in row:
				if cell.value == "DENUMIRE":
					rand_tb = cell.row
					numep = cell.column
					lun = len(workings[cell.column])
		try:
			partnerName = [b.value for b in workings[numep][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'DENUMIRE' in Workings sheet")
			return render_template("index.html")

		for row in workings.iter_rows():
			for cell in row:
				if cell.value == "BAZA IMPOZABILĂ":
					rand_tb = cell.row
					suma = cell.column
					lun = len(workings[cell.column])
		try:
			sumaTot = [b.value for b in workings[suma][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'BAZA IMPOZABILĂ' in Workings sheet")
			return render_template("index.html")

		for row in workings.iter_rows():
			for cell in row:
				if cell.value == "Cheie extragere - filtreaza 1":
					rand_tb = cell.row
					cheie_sort = cell.column
					lun = len(workings[cell.column])
		try:
			cheie = [b.value for b in workings[cheie_sort][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Cheie extragere - filtreaza 1' in Workings sheet")
			return render_template("index.html")

		for row in workings.iter_rows():
			for cell in row:
				if cell.value == "COD OPERATOR INTRACOMUNITAR":
					rand_tb = cell.row
					coi = cell.column
					lun = len(workings[cell.column])
		try:
			listaCOI = [b.value for b in workings[coi][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'BAZA IMPOZABILĂ' in Workings sheet")
			return render_template("index.html")


		nomenclatorTari={'AT':'Austria', 'BE':'Belgia', 'BG':'Bulgaria','CY':'Cipru','DK':'Danemarca','EE':'Estonia', 'FI':'Finlanda','FR':'Franta', 'DE':'Germania','HR':'Croatia',
						'GR':'Grecia','IE':'Irlanda','IT':'Italia','LV':'Letonia','LT':'Lituania','LU':'Luxemburg','MT':'Malta','XI':'Irlanda de Nord - Regatul Unit','NL':'Olanda',
						'PL':'Polonia','PT':'Portugalia','CZ':'Republica Ceha','RO':'Romania','SK':'Slovacia','SI':'Slovavia','ES':'Spania','SE':'Suedia','HU':'Ungaria'}
		b=3
		for i in countryCode:
			if i in nomenclatorTari:
				b=b+1
				workings.cell(row=b, column=2).value=nomenclatorTari[i]

		for row in workings['A3:I3']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')
		for row in workings['E4:E10000']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'
		
		for row in workings['H4:H10000']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'

		workings.column_dimensions['B'].width=20
		workings.column_dimensions['D'].width=35				

		forxml = temp.create_sheet('D390 for XML')
		forxml.cell(row=1,column=1).value="D390 for XML"
		forxml.cell(row=1,column=1).font=cap_tabelbold
		forxml.freeze_panes = 'A4'
		forxml.auto_filter.ref = "A3:F10000"
		forxml.sheet_view.showGridLines = False

		forxml.cell(row=3, column=1).value="III.B"
		forxml.cell(row=3, column=2).value="TIP"
		forxml.cell(row=3, column=3).value="ŢARA"
		forxml.cell(row=3, column=4).value="COD OPERATOR INTRACOMUNITAR"
		forxml.cell(row=3, column=5).value="Denumire"
		forxml.cell(row=3, column=6).value="BAZA IMPOZABILĂ"

		codeAndType=[]
		codeAndName=[]
		typeAndName=[]
		# typeCodeName=[]

		# for k in range(0,len(listaTip)):
		# 	codeAndType.append(str(listaTip[k])+" "+str(codPartener[k]))
		# 	codeAndName.append(str(listaTip[k])+" "+str(partnerName[k]))

		# #print(codeAndType,'codeandtyp')
		# codeAndTypeUnique=list(set(codeAndType))
		# codeAndNameUnique=list(set(codeAndName))

		for i in range(0, len(listaTip)):
			typeAndName.append(str(listaTip[i])+";;;"+str(partnerName[i])+";;;"+str(countryCode[i])+";;;"+str(listaCOI[i]))
		#print(typeAndName, 'TYPEAndNAME')

		typeAndNameUni=list(set(typeAndName))

		typeAndNameUni=list(set(typeAndName))

		for i in it.chain(range(0, len(typeAndNameUni))):
			x=typeAndNameUni[i].split(";;;")
			forxml.cell(row=4+i, column=2).value=str(x[0])
			forxml.cell(row=4+i, column=3).value=str(x[2])
			forxml.cell(row=4+i, column=4).value=str(x[3])
			forxml.cell(row=4+i, column=5).value=str(x[1])
			forxml.cell(row=4+i, column=6).value="=SUMIFS('D390 workings'!H:H,'D390 workings'!A:A,'D390 for XML'!B{0},'D390 workings'!D:D,'D390 for XML'!E{0},'D390 workings'!I:I,1,'D390 workings'!G:G,'D390 for XML'!C{0})".format(4+i)


		for row in forxml['A3:F3']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		for row in forxml['F4:F10000']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'

		forxml.column_dimensions['D'].width=27
		forxml.column_dimensions['E'].width=35
		forxml.column_dimensions['F'].width=15


#---------------------------NR DE EVIDENTA
	if(val3==1):
		sheetinutil3=temp.create_sheet('D394--->>>')
		sheetinutil3.sheet_view.showGridLines=False
		sheetinutil3.cell(row=2,column=1).value="Switch to next sheet for D394 Workings draft"
		sheetinutil3.cell(row=2,column=1).font=scrisincredibildemare		
		nomenclatorTari={'AT':'Austrie', 'BE':'Belgia', 'BG':'Bulgaria','CY':'Cipru','DK':'Danemarca','EE':'Estonia', 'FI':'Finlanda','FR':'Franta', 'DE':'Germania','HR':'Croatia',
						'GR':'Grecia','IE':'Irlanda','IT':'Italia','LV':'Letonia','LT':'Lituania','LU':'Luxemburg','MT':'Malta','XI':'Irlanda de Nord - Regatul Unit','NL':'Olanda',
						'PL':'Polonia','PT':'Portugalia','CZ':'Republica Ceha','RO':'Romania','SK':'Slovacia','SI':'Slovavia','ES':'Spania','SE':'Suedia','HU':'Ungaria'}


		salesExcel=temp.create_sheet("Mapping tranzactii")
		salesExcel.sheet_view.showGridLines = False
		salesExcel.cell(row=2,column=1).value="Mapping tranzactii"
		salesExcel.cell(row=2,column=1).font=cap_tabeltitlu	
		salesExcel.freeze_panes = 'A10'
		
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value=="Declarat anterior":
					rand_tb = cell.row
					declarateanteriorp = cell.column
					lun = len(purchases[cell.column])
		try:
			listadeclantp = [b.value for b in purchases[declarateanteriorp][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Declarat anterior' in Purchases sheet")
			return render_template("index.html")
		# except:
		# 	listadeclant=[]
		listadeclantp_1=[]
		#print(listadeclantp,"---------")
		for c in range(0, len(listadeclantp)):
			if listadeclantp[c] == None:
				listadeclantp_1.append("No")
			else:
				listadeclantp_1.append(listadeclantp[c])
		#print("-----",listadeclantp_1,"------")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value=="Declarat anterior":
					rand_tb = cell.row
					declarateanterior = cell.column
					lun = len(sales[cell.column])
		try:
			listadeclant = [b.value for b in sales[declarateanterior][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Declarat anterior' in Sales sheet")
			return render_template("index.html")
		# except:
		# 	listadeclant=[]
		listadeclant_1=[]
		#print(len(listadeclant))
		for c in range(0, len(listadeclant)):
			if listadeclant[c] == None:
				listadeclant_1.append("No")
			else:
				listadeclant_1.append(listadeclant[c])

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Denumirea clientului Client name":
					rand_tb = cell.row
					clientCell = cell.column
					lun = len(sales[cell.column])
		try:
			listaClient = [b.value for b in sales[clientCell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Denumirea clientului Client name' in Sales sheet")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Codul de inregistrare in scopuri de TVA al clientului Client VAT ID":
					rand_tb = cell.row
					coloanaClientID = cell.column
					lun = len(sales[cell.column])
		try:
			listaCUISales = [b.value for b in sales[coloanaClientID][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Codul de inregistrare in scopuri de TVA al clientului Client VAT ID' in Sales sheet")
			return render_template("index.html")

		listaCUISales1=[]
		# listadeclant_1=[]
		for val in listaCUISales:
			if val != None:
				# listadeclant_1.append("")
				listaCUISales1.append(val)

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Nr/ document Document no":
					rand_tb = cell.row
					docNumber = cell.column
					lun = len(sales[cell.column])
		try:
			docNoSales = [b.value for b in sales[docNumber][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Nr/ document Document no' in Sales sheet")
			return render_template("index.html")




		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)- Livrari locale-Taxable base (RON)-Local supplies (19%)":
					rand_tb = cell.row
					taxBaseL19 = cell.column
					lun = len(sales[cell.column])
		try:
			taxBaseL19 = [b.value for b in sales[taxBaseL19][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)- Livrari locale-Taxable base (RON)-Local supplies (19%)' in Sales sheet")
			return render_template("index.html")

		taxBaseL19_1=[]
		for c in range(0, len(taxBaseL19)):
			if taxBaseL19[c] == None:
				taxBaseL19_1.append(0)
			else:
				taxBaseL19_1.append(taxBaseL19[c])

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-Livrari locale-VAT (RON)-Local supplies (19%)":
					rand_tb = cell.row
					vatBaseL19 = cell.column
					lun = len(sales[cell.column])
		try:
			vatL19 = [b.value for b in sales[vatBaseL19][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-Livrari locale-VAT (RON)-Local supplies (19%)' in Sales sheet")
			return render_template("index.html")

		vatL19_1=[]
		for c in range(0, len(vatL19)):
			if vatL19[c] == None:
				vatL19_1.append(0)
			else:
				vatL19_1.append(vatL19[c])

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-Livrari locale-Taxable base (RON)-Local supplies (9%)":
					rand_tb = cell.row
					taxBaseL9 = cell.column
					lun = len(sales[cell.column])
		try:
			taxBaseL9 = [b.value for b in sales[taxBaseL9][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-Livrari locale-Taxable base (RON)-Local supplies (9%)' in Sales sheet")
			return render_template("index.html")


		taxBaseL9_1=[]
		for c in range(0, len(taxBaseL9)):
			if taxBaseL9[c] == None:
				taxBaseL9_1.append(0)
			else:
				taxBaseL9_1.append(taxBaseL9[c])


		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-Livrari locale-VAT (RON)-Local supplies (9%)":
					rand_tb = cell.row
					vatBaseL9 = cell.column
					lun = len(sales[cell.column])
		try:
			vatL9 = [b.value for b in sales[vatBaseL9][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-Livrari locale-VAT (RON)-Local supplies (9%)' in Sales sheet")
			return render_template("index.html")

		vatL9_1=[]
		for c in range(0, len(vatL9)):
			if vatL9[c] == None:
				vatL9_1.append(0)
			else:
				vatL9_1.append(vatL9[c])

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (EUR/USD)-Livrari locale-Taxable base (EUR/USD)-Local supplies (5%) ":
					rand_tb = cell.row
					taxBaseL5 = cell.column
					lun = len(sales[cell.column])
		try:
			taxBaseL5 = [b.value for b in sales[taxBaseL5][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (EUR/USD)-Livrari locale-Taxable base (EUR/USD)-Local supplies (5%) ' in Sales sheet")
			return render_template("index.html")


		taxBaseL5_1=[]
		for c in range(0, len(taxBaseL5)):
			if taxBaseL5[c] == None:
				taxBaseL5_1.append(0)
			else:
				taxBaseL5_1.append(taxBaseL5[c])


		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-Livrari locale-VAT (RON)-Local supplies (5%)":
					rand_tb = cell.row
					vatBaseL5 = cell.column
					lun = len(sales[cell.column])
		try:
			vatL5 = [b.value for b in sales[vatBaseL5][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-Livrari locale-VAT (RON)-Local supplies (5%)' in Sales sheet")
			return render_template("index.html")

		vatL5_1=[]
		for c in range(0, len(vatL5)):
			if vatL5[c] == None:
				vatL5_1.append(0)
			else:
				vatL5_1.append(vatL5[c])
		# #print(vatL5_1)

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Bază de impozitare (RON)-Livrari locale taxare inversa-Taxable base (RON)-Local supplies reverse charge":
					rand_tb = cell.row
					taxBV = cell.column
					lun = len(sales[cell.column])
		try:
			taxBaseV = [b.value for b in sales[taxBV][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Bază de impozitare (RON)-Livrari locale taxare inversa-Taxable base (RON)-Local supplies reverse charge' in Sales sheet")
			return render_template("index.html")

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Total document (inclusiv TVA)-Taxable base (RON)":
					rand_tb = cell.row
					totdoc = cell.column
					lun = len(sales[cell.column])
		totdocuments = [b.value for b in sales[totdoc][rand_tb:lun+1]]
		taxBaseV_1=[]
		for c in range(0, len(taxBaseV)):
			if taxBaseV[c] == None:
				taxBaseV_1.append(0)
			else:
				taxBaseV_1.append(taxBaseV[c])
		# #print(taxBaseV_1)

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services":
					rand_tb = cell.row
					taxBi = cell.column
					lun = len(sales[cell.column])
		try:
			taxBaseIntracom = [b.value for b in sales[taxBi][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services' in Sales sheet")
			return render_template("index.html")

		taxBaseIntracom_1=[]
		for c in range(0, len(taxBaseIntracom)):
			if taxBaseIntracom[c] == None:
				taxBaseIntracom_1.append(0)
			else:
				taxBaseIntracom_1.append(taxBaseIntracom[c])

		for row in sales.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services":
					rand_tb = cell.row
					taxBiSc = cell.column
					lun = len(sales[cell.column])
		try:
			taxBaseIntracomScutit = [b.value for b in sales[taxBiSc][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-Prestari servicii UE- Taxable base (RON)-EU services' in Sales sheet")
			return render_template("index.html")

		taxBaseIntracomScutit_1=[]
		for c in range(0, len(taxBaseIntracomScutit)):
			if taxBaseIntracomScutit[c] == None:
				taxBaseIntracomScutit_1.append(0)
			else:
				taxBaseIntracomScutit_1.append(taxBaseIntracomScutit[c])
		sumatotala=[]
		for k in range(0,len(taxBaseL19)):
			sumatotala.append(int(taxBaseL19_1[k])+int(taxBaseL9_1[k])+int(taxBaseL5_1[k])+int(taxBaseIntracom_1[k])+int(taxBaseV_1[k]))
		serieCuiSales=[]
		codTaraCuiSales=[]
		for i in listaCUISales1:
			if(str(i)[:1].isalpha()):
				r = re.compile("([a-zA-Z]+)([0-9]+)")
				m = r.match(str(i))
				try:
					serieCuiSales.append(m.group(2))
					codTaraCuiSales.append(m.group(1))

				except:
					codTaraCuiSales.append(None)
					serieCuiSales.append(str(i))	

			else:
				codTaraCuiSales.append(None)
				serieCuiSales.append(str(i))
			# tara,oras=i.split(',',1)
			# serieCuiSales.append(oras)
			# codTaraCuiSales.append(tara)
		#print(codTaraCuiSales)
		# TIP Furnizor!!!!!
		# #print(len(codTaraCuiSales))
		coteTVAsales=[]
		for i in range(0, len(docNoSales)):
			if (int(vatL19_1[i])>0):
				coteTVAsales.append(19)
			else:
				if (int(vatL19_1[i])<0):
					coteTVAsales.append(19)
				else:
					if (int(vatL9_1[i])> 0):
						coteTVAsales.append(9)
					else:
						if (int(vatL9_1[i]) < 0):
							coteTVAsales.append(9)
						else:
							if (int(vatL5_1[i]) > 0):
								coteTVAsales.append(5)
							else:
								if (int(vatL5_1[i]) < 0):
									coteTVAsales.append(5)
								else:
									coteTVAsales.append(0)		
		codTranzactieSales=[]
		for i in range(0, len(codTaraCuiSales)):
			if str(serieCuiSales[i])[1:2].isalpha():
				codTranzactieSales.append(2)
			else:
				if codTaraCuiSales[i] == "RO" and int(coteTVAsales[i])>0:
					# #print("RO")
					codTranzactieSales.append(1)
				else:	
					if codTaraCuiSales[i] in nomenclatorTari:
						# #print("UE")
						codTranzactieSales.append(3)
					else:
						# #print("nonUE")
						codTranzactieSales.append(4)
		#Cote TVAA




		#TIP TRANZACTIE
		storno=[]
		tipTranzSale = []
		# #print(docNoSales)
		#print(len(docNoSales),len(codTranzactieSales))
		for i in range(0, len(docNoSales)):
			if(listadeclant_1[i]=="Yes"):
				tipTranzSale.append("Declarat anterior")
			else:
				if int(codTranzactieSales[i]) == 1:
					# #print(docNoSales[i]," ",listaCUISales1[i], "", taxBaseL19_1[i], " ", taxBaseL9_1[i], " ", taxBaseL5_1[i])
					if (int(taxBaseL19_1[i])>0 and int(vatL19_1[i])>0) or (int(taxBaseL9_1[i])>0 and int(vatL9_1[i])> 0) or (int(taxBaseL5_1[i])>0 and int(vatL5_1[i])>0):
						tipTranzSale.append('L')
						storno.append("")
					else:
						# None
						if (int(taxBaseL19_1[i])<0 and int(vatL19_1[i])<0) or (int(taxBaseL9_1[i])<0 and int(vatL9_1[i])< 0) or (int(taxBaseL5_1[i])<0 and int(vatL5_1[i])<0):
							tipTranzSale.append('L')
							storno.append("Yes")
						else:
							if int(taxBaseV_1[i]) > 0:
								tipTranzSale.append("V")
								storno.append("")
							else:
								if int(taxBaseV_1[i]) < 0:
									storno.append("Yes")
									tipTranzSale.append('V')
								else:
									tipTranzSale.append(None)
									storno.append("")
				else:
					if int(codTranzactieSales[i]) == 2:
						if (int(taxBaseL19_1[i]) > 0 and int(vatL19_1[i]) > 0) or (int(taxBaseL9_1[i]) > 0 and int(vatL9_1[i]) > 0) or (
								int(taxBaseL5_1[i]) > 0 and int(vatL5_1[i]) > 0):
							tipTranzSale.append('L')
							storno.append("")
						else:
							if (int(taxBaseL19_1[i]) < 0 and int(vatL19_1[i]) < 0) or (
									int(taxBaseL9_1[i]) < 0 and int(vatL9_1[i]) < 0) or (
									int(taxBaseL5_1[i]) < 0 and int(vatL5_1[i]) < 0):
								storno.append("Yes")
								tipTranzSale.append('L')
							else:
								tipTranzSale.append('L')
								storno.append("")
					else:
						if int(codTranzactieSales[i]) == 3:
							if (int(taxBaseIntracom_1[i]) > 0) or (int(taxBaseIntracomScutit_1[i])>0):
								tipTranzSale.append('Not applicable for D394')
								storno.append("")
							else:
								if int(taxBaseIntracom_1[i]) < 0 or int(taxBaseIntracomScutit_1[i])<0 :
									storno.append("Yes")
									tipTranzSale.append('Not applicable for D394')
								else:
									if (int(taxBaseL19_1[i]) > 0 and int(vatL19_1[i]) > 0) or (int(taxBaseL9_1[i]) > 0 and int(vatL9_1[i]) > 0) or (
											int(taxBaseL5_1[i]) > 0 and int(vatL5_1[i]) > 0):
										tipTranzSale.append('L')
										storno.append("")
									else:
										if (int(taxBaseL19_1[i]) < 0 and int(vatL19_1[i]) < 0) or (
												int(taxBaseL9_1[i]) < 0 and int(vatL9_1[i]) < 0) or (
												int(taxBaseL5_1[i]) < 0 and int(vatL5_1[i]) < 0):
											storno.append("Yes")
											tipTranzSale.append('L')
										else:
											tipTranzSale.append('Not applicable for D394')
											storno.append("")

						else:
							if int(codTranzactieSales[i]) == 4:
								if (int(taxBaseL19_1[i]) > 0 and int(vatL19_1[i]) > 0) or (
										int(taxBaseL9_1[i]) > 0 and int(vatL9_1[i]) > 0) or (
										int(taxBaseL5_1[i]) > 0 and int(vatL5_1[i]) > 0):
									tipTranzSale.append('L')
									storno.append("")
								else:
									if (int(taxBaseL19_1[i]) < 0 and int(vatL19_1[i]) < 0) or (
											int(taxBaseL9_1[i]) < 0 and int(vatL9_1[i]) < 0) or (
											int(taxBaseL5_1[i]) < 0 and int(vatL5_1[i]) < 0):
										storno.append("Yes")
										tipTranzSale.append('storno')
									else:
										tipTranzSale.append('L')
										storno.append("")
							else:
								tipTranzSale.append('out of ro')
								storno.append("")



		# #print(docNoSales)
		#Scriere in excel

		salesExcel.cell(row=9, column=1).value = "Cod tara"
		salesExcel.cell(row=9, column=2).value = "Serie cui"
		salesExcel.cell(row=9, column=3).value = "Numar document"
		salesExcel.cell(row=9, column=4).value = "CUI"
		salesExcel.cell(row=9, column=5).value = "Clasa tranzactie"
		salesExcel.cell(row=9, column=6).value = "Tip tranzactie"
		salesExcel.cell(row=9, column=7).value = "Cota TVA"
		salesExcel.cell(row=9, column=8).value = "Total document"
		salesExcel.cell(row=9, column=9).value = "Tip jurnal"
		salesExcel.cell(row=9, column=10).value = "Nume partener"
		salesExcel.cell(row=9, column=11).value = "Check"
		salesExcel.cell(row=9, column=12).value = "Cod si denumire NC produs(TIP V)"

		# dv = DataValidation(
			# type='list', formula1='"Yes,No"', allow_blank=True)


		listahelp=["1002--Secara","1003--Orz","1005--Porumb","1201--Boabe de soia"," 1205--Seminte de rapita sau de rapita salbatica","120600--Seminte de floarea soarelui","121291--Sfecla de zahar","1001-Grau si meslin","1004--Ovaz","10086000--Triticale","22-deseuri feroase si neferoase","23-masa lemnoasa","32-terenuri","33-constructii","34-alte bunuri","35-servicii","24-certificate de emisii de gaze cu efect de sera","25-energie electrica","26-certificate verzi","27-constructii/terenuri","28-aur de investitii","29-telefoane mobile","30-microprocesoare","31-console de jocuri tablete PC si laptopuri"]
		sheethelp=temp.create_sheet("Validation")
		sheethelp.sheet_state = 'hidden'

		# dv = DataValidation(
		# 	type="list", formula1="", allow_blank=True)
		# salesExcel.add_data_validation(dv)

		# dv.add(salesExcel["L2"])


		for i in range(0,len(listahelp)):
			sheethelp.cell(row=i+1,column=1).value=listahelp[i]

		#print(len(tipTranzSale),tipTranzSale)
		for i in range(0, len(codTaraCuiSales)):

			salesExcel.cell(row=10 + i, column=1).value = codTaraCuiSales[i]
			salesExcel.cell(row=10 + i, column=2).value = serieCuiSales[i]
			salesExcel.cell(row=10 + i, column=3).value = docNoSales[i]
			salesExcel.cell(row=10 + i, column=4).value = listaCUISales1[i]
			salesExcel.cell(row=10 + i, column=5).value = codTranzactieSales[i]
			# if(listadeclant_1[i]!=""):
			salesExcel.cell(row=10 + i, column=6).value = tipTranzSale[i]
			# else:
				# salesExcel.cell(row=10 + i, column=6).value = listadeclant_1[i]

			if(tipTranzSale[i]=='V'):
				salesExcel.cell(row=10+i,column=12).value="V"
			else:
				salesExcel.cell(row=10+i,column=12).value="N/A, valid only for V trans."

			salesExcel.cell(row=10 + i, column=8).value = sumatotala[i]
			salesExcel.cell(row=10 + i, column=9).value = "Jurnal vanzari"
			salesExcel.cell(row=10 + i, column=10).value = listaClient[i]

		for i in range(0, len(coteTVAsales)):
			salesExcel.cell(row=10 + i, column=7).value = coteTVAsales[i]
			salesExcel.cell(row=10+i,column=18).value="=B{0}&E{0}&F{0}&G{0}".format(i+10)

		#FORMATARE------------------------------------------------------------------
		red_color = 'ffc7ce'
		green_color='99ff99'
		red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
		green_fill = styles.PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
		row=salesExcel.max_row	
		salesExcel.conditional_formatting.add('K10:K'+str(row-1), formatting.rule.CellIsRule(operator='notEqual', formula=['"OK"'], fill=red_fill))
		for row in salesExcel['A9:L9']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.alignment=Alignment(horizontal='center',vertical='center')				

		for row in salesExcel['A9:L9']:
			for cell in row:
				cell.font = cap_tabel

		# for row in salesExcel['A9:K9']:
		# 	for cell in row:
		# 		cell.border = border_thin

		
		salesExcel.freeze_panes = 'A10'

		salesExcel.column_dimensions['B'].width = 20
		salesExcel.column_dimensions['C'].width = 20
		salesExcel.column_dimensions['D'].width = 20
		salesExcel.column_dimensions['E'].width = 16
		salesExcel.column_dimensions['F'].width = 16
		salesExcel.column_dimensions['F'].width = 20		
		salesExcel.column_dimensions['H'].width = 14
		salesExcel.column_dimensions['J'].width = 35
		salesExcel.column_dimensions['L'].width = 35		
		purchases = temp['Purchases']

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Furnizor Supplier":
					rand_tb = cell.row
					supplierCell = cell.column
					lun = len(purchases[cell.column])
		try:
			supplierName = [b.value for b in purchases[supplierCell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Furnizor Supplier' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Aplica TVA la incasare VAT cash-in system":
					rand_tb = cell.row
					vatCashinSys = cell.column
					lun = len(purchases[cell.column])
		try:
			vatApplies = [b.value for b in purchases[vatCashinSys][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Aplica TVA la incasare VAT cash-in system' in Purchases sheet")
			return render_template("index.html")

		vatApplies_1=[]


		for val in vatApplies:
			vatApplies_1.append(val)
		print(vatApplies_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Codul de înregistrare în scopuri de TVA VAT number":
					rand_tb = cell.row
					suppID = cell.column
					lun = len(purchases[cell.column])
		try:
			suppIDPurch = [b.value for b in purchases[suppID][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Codul de înregistrare în scopuri de TVA VAT number' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Nr. document Document no":
					rand_tb = cell.row
					docNumberPurch = cell.column
					lun = len(purchases[cell.column])
		try:
			docNoPurch = [b.value for b in purchases[docNumberPurch][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Nr. document Document no' in Purchases sheet")
			return render_template("index.html")
		
		docNoPurch1 = []
		for val in docNoPurch:
			if val != None:
				docNoPurch1.append(val)
		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON) -achizitii locale-Taxable base (RON) - local acquisition (19%)":
					rand_tb = cell.row
					taxBaseAch19cell = cell.column
					lun = len(purchases[cell.column])
		try:
			taxBaseAch19 = [b.value for b in purchases[taxBaseAch19cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON) -achizitii locale-Taxable base (RON) - local acquisition (19%)' in Purchases sheet")
			return render_template("index.html")

		taxBaseAch19_1=[]
		for i in range(0, len(taxBaseAch19)):
			if taxBaseAch19[i] == None:
				taxBaseAch19_1.append(0)
			else:
				taxBaseAch19_1.append(taxBaseAch19[i])

		#print(taxBaseAch19,taxBaseAch19_1)
		# #print(taxBaseAch19_1)
		# #print(taxBaseAch19)
		# for item in taxBaseAch19:
		#    #print(type(item))
		# #print(type(taxBaseAch19))

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-achizitii locale-VAT (RON)-local acquisition (19%)":
					rand_tb = cell.row
					vatAch19cell = cell.column
					lun = len(purchases[cell.column])
		try:
			vatAch19 = [b.value for b in purchases[vatAch19cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-achizitii locale-VAT (RON)-local acquisition (19%)' in Purchases sheet")
			return render_template("index.html")

		vatAch19_1=[]
		for i in range(0, len(vatAch19)):
			if vatAch19[i] == None or vatAch19[i]=="":
				vatAch19_1.append(0)
			else:
				vatAch19_1.append(vatAch19[i])
		# #print(vatAch19_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-achizitii locale- Taxable base (RON)-local acquisition (9%)":
					rand_tb = cell.row
					taxBaseAch9cell = cell.column
					lun = len(purchases[cell.column])
		try:
			taxBaseAch9 = [b.value for b in purchases[taxBaseAch9cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-achizitii locale- Taxable base (RON)-local acquisition (9%)' in Purchases sheet")
			return render_template("index.html")

		taxBaseAch9_1=[]
		for i in range(0, len(taxBaseAch9)):
			if taxBaseAch9[i] == None:
				taxBaseAch9_1.append(0)
			else:
				taxBaseAch9_1.append(taxBaseAch9[i])
		# #print(taxBaseAch9_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-achizitii locale-VAT (RON)-local acquisition (9%)":
					rand_tb = cell.row
					vatAch9cell = cell.column
					lun = len(purchases[cell.column])
		try:
			vatAch9 = [b.value for b in purchases[vatAch9cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-achizitii locale-VAT (RON)-local acquisition (9%)' in Purchases sheet")
			return render_template("index.html")

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Total document (inclusiv TVA)-RON":
					rand_tb = cell.row
					totdocp = cell.column
					lun = len(purchases[cell.column])
		try:
			totdocumentp = [b.value for b in purchases[totdocp][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Total document (inclusiv TVA)-RON' in Purchases sheet")
			return render_template("index.html")

		vatAch9_1=[]
		for i in range(0, len(vatAch9)):
			if vatAch9[i] == None:
				vatAch9_1.append(0)
			else:
				vatAch9_1.append(vatAch9[i])
		# #print(vatAch9_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-achizitii locale Taxable base (RON) local acquisition (5%)":
					rand_tb = cell.row
					taxBaseAch5cell = cell.column
					lun = len(purchases[cell.column])
		try:
			taxBaseAch5 = [b.value for b in purchases[taxBaseAch5cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-achizitii locale Taxable base (RON) local acquisition (5%)' in Purchases sheet")
			return render_template("index.html")

		taxBaseAch5_1=[]
		for i in range(0, len(taxBaseAch5)):
			if taxBaseAch5[i] == None:
				taxBaseAch5_1.append(0)
			else:
				taxBaseAch5_1.append(taxBaseAch5[i])
		# #print(taxBaseAch5_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-achizitii locale-VAT (RON)-local acquisition (5%)":
					rand_tb = cell.row
					vatAch5cell = cell.column
					lun = len(purchases[cell.column])
		try:
			vatAch5 = [b.value for b in purchases[vatAch5cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-achizitii locale-VAT (RON)-local acquisition (5%)' in Purchases sheet")
			return render_template("index.html")

		vatAch5_1=[]
		for i in range(0, len(vatAch5)):
			if vatAch5[i] == None:
				vatAch5_1.append(0)
			else:
				vatAch5_1.append(vatAch5[i])


		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Exempt- Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable":
					rand_tb = cell.row
					vatExemptLocAcq = cell.column
					lun = len(purchases[cell.column])
		try:
			vatExempt = [b.value for b in purchases[vatExemptLocAcq][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
			return render_template("index.html")

		vatExempt_1=[]
		for i in range(0, len(vatExempt)):
			if vatExempt[i] == None:
				vatExempt_1.append(0)
			else:
				vatExempt_1.append(vatExempt[i])

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)":
					rand_tb = cell.row
					nonChartb = cell.column
					lun = len(purchases[cell.column])
		try:
			nonCharTaxBase = [b.value for b in purchases[nonChartb][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Achiziţii de bunuri şi servicii scutite de taxă sau neimpozabile / VAT exempt local acquisitions or non-taxable (RON)' in Purchases sheet")
			return render_template("index.html")

		nonCharTaxBase_1=[]
		for i in range(0, len(nonCharTaxBase)):
			if nonCharTaxBase[i] == None:
				nonCharTaxBase_1.append(0)
			else:
				nonCharTaxBase_1.append(nonCharTaxBase[i])
		#print(nonCharTaxBase_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TVA nedeductibila/ Non-deductible VAT (RON)":
					rand_tb = cell.row
					nonChartVATtb = cell.column
					lun = len(purchases[cell.column])
		try:
			nonChartVATBase = [b.value for b in purchases[nonChartVATtb][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA nedeductibila/ Non-deductible VAT (RON)' in Purchases sheet")
			return render_template("index.html")

		nonChartVATBase_1=[]
		for i in range(0, len(nonChartVATBase)):
			if nonChartVATBase[i] == None:
				nonChartVATBase_1.append(0)
			else:
				nonChartVATBase_1.append(nonChartVATBase[i])
		# #print(nonChartVATBase_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge (19%)":
					rand_tb = cell.row
					revTaxBaseAch19cell = cell.column
					lun = len(purchases[cell.column])
		try:
			revTaxBaseAch19 = [b.value for b in purchases[revTaxBaseAch19cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge (19%)' in Purchases sheet")
			return render_template("index.html")

		revTaxBaseAch19_1=[]
		for i in range(0, len(revTaxBaseAch19)):
			if revTaxBaseAch19[i] == None:
				revTaxBaseAch19_1.append(0)
			else:
				revTaxBaseAch19_1.append(revTaxBaseAch19[i])
		# #print(revTaxBaseAch19_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-Achizitii locale taxare inversa-VAT (RON)-Local acquisition reverse charge (19%)":
					rand_tb = cell.row
					revVatAch19cell = cell.column
					lun = len(purchases[cell.column])
		try:
			revVatAch19 = [b.value for b in purchases[revVatAch19cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-Achizitii locale taxare inversa-VAT (RON)-Local acquisition reverse charge (19%)' in Purchases sheet")
			return render_template("index.html")

		revVatAch19_1=[]
		for i in range(0, len(revVatAch19)):
			if revVatAch19[i] == None:
				revVatAch19_1.append(0)
			else:
				revVatAch19_1.append(revVatAch19[i])
		# #print(revVatAch19_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge (9%)":
					rand_tb = cell.row
					revTaxBaseAch9cell = cell.column
					lun = len(purchases[cell.column])
		try:
			revTaxBaseAch9 = [b.value for b in purchases[revTaxBaseAch9cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge (9%)' in Purchases sheet")
			return render_template("index.html")

		revTaxBaseAch9_1=[]
		for i in range(0, len(revTaxBaseAch9)):
			if revTaxBaseAch9[i] == None:
				revTaxBaseAch9_1.append(0)
			else:
				revTaxBaseAch9_1.append(revTaxBaseAch9[i])
		# #print(revTaxBaseAch9_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-Achizitii locale taxare inversa-VAT (RON)-Local acquisition reverse charge (9%)":
					rand_tb = cell.row
					revVatAch9cell = cell.column
					lun = len(purchases[cell.column])
		try:
			revVatAch9 = [b.value for b in purchases[revVatAch9cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-Achizitii locale taxare inversa-VAT (RON)-Local acquisition reverse charge (9%)' in Purchases sheet")
			return render_template("index.html")

		revVatAch9_1=[]
		for i in range(0, len(revVatAch9)):
			if revVatAch9[i] == None:
				revVatAch9_1.append(0)
			else:
				revVatAch9_1.append(revVatAch9[i])
		# #print(revVatAch9_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge(5%)":
					rand_tb = cell.row
					revTaxBaseAch5cell = cell.column
					lun = len(purchases[cell.column])
		try:
			revTaxBaseAch5 = [b.value for b in purchases[revTaxBaseAch5cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Baza de impozitare (RON)-Achizitii locale taxare inversa-Taxable base (RON)-Local acquisition reverse charge(5%)' in Purchases sheet")
			return render_template("index.html")

		revTaxBaseAch5_1=[]
		for i in range(0, len(revTaxBaseAch5)):
			if revTaxBaseAch5[i] == None:
				revTaxBaseAch5_1.append(0)
			else:
				revTaxBaseAch5_1.append(revTaxBaseAch5[i])
		# #print(revTaxBaseAch5_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-Achizitii locale taxare inversa VAT (RON)-Local acquisition reverse charge (5%)":
					rand_tb = cell.row
					revVatAch5cell = cell.column
					lun = len(purchases[cell.column])
		try:
			revVatAch5 = [b.value for b in purchases[revVatAch5cell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-Achizitii locale taxare inversa VAT (RON)-Local acquisition reverse charge (5%)' in Purchases sheet")
			return render_template("index.html")

		revVatAch5_1=[]
		for i in range(0, len(revVatAch5)):
			if revVatAch5[i] == None:
				revVatAch5_1.append(0)
			else:
				revVatAch5_1.append(revVatAch5[i])
		# #print(revVatAch5_1)

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Bază de impozitare (RON)-bunuri nonUE-Taxable base (RON)-nonUE goods":
					rand_tb = cell.row
					nonUEGoodscell = cell.column
					lun = len(purchases[cell.column])
		try:
			taxNonUEgoods = [b.value for b in purchases[nonUEGoodscell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Bază de impozitare (RON)-bunuri nonUE-Taxable base (RON)-nonUE goods' in Purchases sheet")
			return render_template("index.html")

		taxNonUEgoods_1=[]
		for i in range(0, len(taxNonUEgoods)):
			if taxNonUEgoods[i] == None:
				taxNonUEgoods_1.append(0)
			else:
				taxNonUEgoods_1.append(taxNonUEgoods[i])

		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-bunuri nonUE-VAT (RON)-nonUE goods":
					rand_tb = cell.row
					vatNonUEGoodscell = cell.column
					lun = len(purchases[cell.column])
		try:
			vatNonUEGoods = [b.value for b in purchases[vatNonUEGoodscell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-bunuri nonUE-VAT (RON)-nonUE goods' in Purchases sheet")
			return render_template("index.html")

		vatNonUEGoods_1=[]
		for i in range(0, len(vatNonUEGoods)):
			if vatNonUEGoods[i] == None:
				vatNonUEGoods_1.append(0)
			else:
				vatNonUEGoods_1.append(vatNonUEGoods[i])


		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "Bază de impozitare (RON)-servicii nonUE-Taxable base (RON)-nonUE services":
					rand_tb = cell.row
					nonUEServcell = cell.column
					lun = len(purchases[cell.column])
		try:
			taxNonUEservices = [b.value for b in purchases[nonUEServcell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'Bază de impozitare (RON)-servicii nonUE-Taxable base (RON)-nonUE services' in Purchases sheet")
			return render_template("index.html")

		taxNonUEservices_1=[]
		for i in range(0, len(taxNonUEservices)):
			if taxNonUEservices[i] == None:
				taxNonUEservices_1.append(0)
			else:
				taxNonUEservices_1.append(taxNonUEservices[i])


		for row in purchases.iter_rows():
			for cell in row:
				if cell.value == "TVA (RON)-servicii nonUE-VAT (RON)-nonUE services":
					rand_tb = cell.row
					vatNonUEservcell = cell.column
					lun = len(purchases[cell.column])
		try:
			vatNonUEservices = [b.value for b in purchases[vatNonUEservcell][rand_tb:lun]]
		except:
			flash("Please insert the correct header for 'TVA (RON)-servicii nonUE-VAT (RON)-nonUE services' in Purchases sheet")
			return render_template("index.html")

		vatNonUEservices_1=[]
		for i in range(0, len(vatNonUEservices)):
			if vatNonUEservices[i] == None:
				vatNonUEservices_1.append(0)
			else:
				vatNonUEservices_1.append(vatNonUEservices[i])

		serieCuiPurch = []
		codTaraCuiPurch = []
		# #print(suppIDPurch)
		for i in suppIDPurch:
			# #print(i)
			if(str(i)[:1].isalpha()):	
				r = re.compile("([a-zA-Z]+)([0-9]+)")
				m = r.match(str(i))
				try:
					serieCuiPurch.append(m.group(2))
				except:
					serieCuiPurch.append(" ")
				try:
					codTaraCuiPurch.append(m.group(1))
				except:
					codTaraCuiPurch.append(" ")
			else:
				codTaraCuiPurch.append(None)
				serieCuiPurch.append(i)
		# #print(codTaraCuiPurch,serieCuiPurch)
		tipTranzactiePurchases=[]
		#print("Aici vat -------",vatAch19_1,"-----Achizitii")
		#Tip furnizor
		for i in range(0, len(codTaraCuiPurch)):
			if codTaraCuiPurch[i] == "RO":
				# #print("RO")
				tipTranzactiePurchases.append(1)
			else:
				if serieCuiPurch[i] == suppIDPurch[i] and (int(nonCharTaxBase_1[i])>0 or int(nonCharTaxBase_1[i])<0) :
					tipTranzactiePurchases.append(2)
				else:
					if codTaraCuiPurch[i] in nomenclatorTari:
						# #print("UE")
						tipTranzactiePurchases.append(3)
					else:
						# #print("nonUE")
						tipTranzactiePurchases.append(4)
		for j in range(0,len(tipTranzactiePurchases)):
			print(nonCharTaxBase_1[j],tipTranzactiePurchases[j])
		coteTVApurchases=[]
		for i in range(0, len(docNoPurch1)):
			if (int(vatAch19_1[i]) > 0):
				coteTVApurchases.append('19')
			else:
				if (int(vatAch19_1[i])<0):
					coteTVApurchases.append('19')
				else:
					if (int(vatAch9_1[i]) > 0):
						coteTVApurchases.append('9')
					else:
						if (int(vatAch9_1[i]) < 0):
							coteTVApurchases.append('9')
						else:
							if (int(vatAch5_1[i]) > 0):
								coteTVApurchases.append('5')
							else:
								if (int(vatAch5_1[i]) < 0):
									coteTVApurchases.append('5')
								else:
									if (int(revVatAch19_1[i])>0):
										coteTVApurchases.append('19')
									else:
										if (int(revVatAch19_1[i]) < 0):
											coteTVApurchases.append('19')
										else:
											if (int(revVatAch9_1[i]) > 0):
												coteTVApurchases.append('9')
											else:
												if (int(revVatAch9_1[i]) < 0):
													coteTVApurchases.append('9')
												else:
													if (int(revVatAch5_1[i]) > 0):
														coteTVApurchases.append('5')
													else:
														if (int(revVatAch19_1[i]) < 0):
															coteTVApurchases.append('5')
														else:
															coteTVApurchases.append('0')

		#mapare tip tranzactie
		# #print(len(docNoPurch1),len(tipTranzactiePurchases),len(vatApplies))
		# #print(vatApplies)
		tipTranzPurch=[]

		# for i in range(0,len(suppIDPurch)):
			# #print(suppIDPurch[i],tipTranzactiePurchases[i])
		# #print(len(docNoPurch),len(tipTranzactiePurchases))
		#print(len(docNoPurch1),len(listadeclantp_1),"--------------len de lista")
		for i in range(0, len(docNoPurch1)):
			if(listadeclantp_1[i]=="Yes"):
				tipTranzPurch.append("Declarat anterior")
			else:
				# #print(docNoPurch1[i])
				if int(tipTranzactiePurchases[i]) == 1:
					if (taxBaseAch19_1[i] > 0  or taxBaseAch9_1[i] > 0 or taxBaseAch5_1[i] > 0) and vatApplies[i]==None:
						tipTranzPurch.append('A')

						#print(docNoPurch1[i],";;;;;ES 1")
					else:
						if taxBaseAch19_1[i] < 0  or taxBaseAch9_1[i] < 0 or taxBaseAch5_1[i] < 0 :
							tipTranzPurch.append("A")
							#print(docNoPurch1[i],";;;;;es 2")
					if nonCharTaxBase_1[i] > 0 or nonCharTaxBase_1[i]<0 or vatApplies[i] == "YES":
						#print(vatApplies[i])
						tipTranzPurch.append("AI")
						#print(docNoPurch1[i],";;;;;es 3")
					else:
						if (nonCharTaxBase_1[i] < 0 and nonChartVATBase_1[i] < 0):
							tipTranzPurch.append("AI")
							#print(docNoPurch1[i],";;;;;es 4")
						else:
							if (revTaxBaseAch19_1[i] > 0 and revVatAch19_1[i] > 0) or (
									revTaxBaseAch9_1[i] > 0 and revVatAch9_1[i] > 0) or \
									(revTaxBaseAch5_1[i] > 0 and revVatAch5_1[i] > 0):
								tipTranzPurch.append("C")
								#print("Yes 5")
							else:
								if (revTaxBaseAch19_1[i] < 0 and revVatAch19_1[i] < 0) or (revTaxBaseAch9_1[i] < 0 and revVatAch9_1[i] < 0) or \
								(revTaxBaseAch5_1[i] < 0 and revVatAch5_1[i] < 0):
									tipTranzPurch.append("C")
									#print("Yes 6")
								else:
									if (taxNonUEgoods_1[i] > 0 and vatNonUEGoods_1[i] > 0) or (
											taxNonUEservices_1[i] > 0 and vatNonUEservices_1[i] > 0):
										tipTranzPurch.append("C")
										#print("Yes 7")
									else:
										if (taxNonUEgoods_1[i] < 0 and vatNonUEGoods_1[i] < 0) or (
												taxNonUEservices_1[i] < 0 and vatNonUEservices_1[i] < 0):
											tipTranzPurch.append("C")
											#print("Yes 8")
										else:
											if(vatExempt_1[i]>0):
												tipTranzPurch.append("N")
												#print("Yes 8 nou")
				else:
					if int(tipTranzactiePurchases[i]) == 2:
						if int(vatExempt_1[i])>0 or int(nonCharTaxBase_1[i])>0:
							tipTranzPurch.append("N")
							#print(docNoPurch1[i],";;;;es 9")
						else:
							if int(vatExempt_1[i])<0 or int(nonCharTaxBase_1[i])<0:
								tipTranzPurch.append("N")
								#print("Yes 10")
					else:
						if int(tipTranzactiePurchases[i]) == 3:
							if (taxBaseAch19_1[i] > 0 or vatAch19_1[i] > 0) or (taxBaseAch9_1[i] > 0 or vatAch9_1[i] > 0) or (
											taxBaseAch5_1[i] > 0 or vatAch5_1[i] > 0):
										tipTranzPurch.append('A')
										#print(docNoPurch1[i],";;;;;es 12")
							else:
								if (taxBaseAch19_1[i] < 0 or vatAch19_1[i] < 0) or (taxBaseAch9_1[i] < 0 or vatAch9_1[i] < 0) or (
										taxBaseAch5_1[i] < 0 or vatAch5_1[i] < 0):
									tipTranzPurch.append('A')
									#print(docNoPurch1[i],";;;;;es 13")
								else:
									if (taxNonUEgoods_1[i] > 0 and vatNonUEGoods_1[i] > 0) or (
											taxNonUEservices_1[i] > 0 and vatNonUEservices_1[i] > 0):
										tipTranzPurch.append("C")
										#print("Yes 14")
									else:
										if (taxNonUEgoods_1[i] < 0 and vatNonUEGoods_1[i] < 0) or (
												taxNonUEservices_1[i] < 0 and vatNonUEservices_1[i] < 0):
											tipTranzPurch.append("C")
											#print("Yes 15")
										else:
											tipTranzPurch.append("Not applicable for D394")
						else:
								if int(tipTranzactiePurchases[i]) == 4:
									if (taxBaseAch19_1[i] > 0 or vatAch19_1[i] > 0) or (taxBaseAch9_1[i] > 0 or vatAch9_1[i] > 0) or (
											taxBaseAch5_1[i] > 0 or vatAch5_1[i] > 0):
										tipTranzPurch.append('A')
										#print(docNoPurch1[i],";;;;;es 12")
									else:
										if (taxBaseAch19_1[i] < 0 or vatAch19_1[i] < 0) or (taxBaseAch9_1[i] < 0 or vatAch9_1[i] < 0) or (
												taxBaseAch5_1[i] < 0 or vatAch5_1[i] < 0):
											tipTranzPurch.append('A')
											#print(docNoPurch1[i],";;;;;es 13")
										else:
											if (taxNonUEgoods_1[i] > 0 and vatNonUEGoods_1[i] > 0) or (
													taxNonUEservices_1[i] > 0 and vatNonUEservices_1[i] > 0):
												tipTranzPurch.append("C")
												#print("Yes 14")
											else:
												if (taxNonUEgoods_1[i] < 0 and vatNonUEGoods_1[i] < 0) or (
														taxNonUEservices_1[i] < 0 and vatNonUEservices_1[i] < 0):
													tipTranzPurch.append("C")
													#print("Yes 15")
								else:
									tipTranzPurch.append("none")
									#print("Yes 16")
			# #print(docNoPurch1[i],tipTranzPurch[i],docNoPurch[i+1])
		ma=salesExcel.max_row+1
		for i in range(0, len(codTaraCuiPurch)):
			salesExcel.cell(row=ma + i, column=1).value = codTaraCuiPurch[i]

		for i in range(0, len(serieCuiPurch)):
			salesExcel.cell(row=ma + i, column=2).value = serieCuiPurch[i]

		for i in range(0, len(docNoPurch1)):
			salesExcel.cell(row=ma+ i, column=3).value = docNoPurch1[i]

		for i in range(0, len(suppIDPurch)):
			salesExcel.cell(row=ma + i, column=4).value = suppIDPurch[i]

		for i in range(0, len(tipTranzactiePurchases)):
			salesExcel.cell(row=ma+ i, column=5).value = tipTranzactiePurchases[i]

		for i in range(0, len(tipTranzPurch)):
			# if(listadeclantp_1!=""):
			salesExcel.cell(row=ma+ i, column=6).value = tipTranzPurch[i]
			# else:
				# salesExcel.cell(row=ma+ i, column=6).value = "Declarate anterior"
			if(tipTranzPurch[i]=="V"):
				salesExcel.cell(row=ma+i,column=12).value="Add type of tranzactie"
			else:
				salesExcel.cell(row=ma+i,column=12).value="N/A"

		for i in range(0, len(coteTVApurchases)):
			#print(coteTVApurchases[i])
			salesExcel.cell(row=ma+ i, column=7).value = coteTVApurchases[i]
			try:
				salesExcel.cell(row=ma+ i, column=8).value = totdocumentp[i]/(100+int(coteTVApurchases[i]))*100
			except:
				salesExcel.cell(row=ma+ i, column=8).value = totdocumentp[i]
			salesExcel.cell(row=ma+ i, column=9).value = "Jurnal cumparari"
			salesExcel.cell(row=ma+ i, column=10).value = supplierName[i]

		codTaraCUItotal=codTaraCuiPurch+codTaraCuiSales
		for i in range(0, len(codTaraCUItotal)):
			salesExcel.cell(row=10 + i, column=11).value = '=IFERROR(IF(VLOOKUP(B{0}&E{0}&F{0}&G{0},Tranzactii!K:K,1,0)=B{0}&E{0}&F{0}&G{0},"OK","Mapped missing in Tranzactii sheet"),"Mapped missing inTranzactiisheet")'.format(10+i)
		salesExcel.auto_filter.ref = "A9:L9"
		for row in salesExcel['H10:F1000']:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'		
		tranzactii=temp.create_sheet("Tranzactii")
		tranzactii.freeze_panes = 'A6'
		tranzactii.sheet_view.showGridLines = False

												
		setSalesCUI=set(listaCUISales1)
		idSalesUnique=list(setSalesCUI)

		setPurchCUI=set(suppIDPurch)
		idPurchUnique=list(setPurchCUI)

		listaCUIUnique=idSalesUnique+idPurchUnique

		setlistaClient=set(listaClient)
		listaClientUnique=list(setlistaClient)

		setSupplierName=set(supplierName)
		supplierNameUnique=list(setSupplierName)

		# for k in range(0,len(setlistaClient)):
		#    count
		#    for j in range(0,len(listaClient)):


		#print(len(supplierName),len(tipTranzPurch))
		listanouaappendpurch=[]
		# for i in range(0,len(supplierName)):
		print(docNoPurch)
		print(len(serieCuiPurch),len(tipTranzPurch),len(coteTVApurchases),len(tipTranzactiePurchases))
		for k in range(0,len(serieCuiPurch)):
			print(docNoPurch[k],serieCuiPurch[k],tipTranzPurch[k],coteTVApurchases[k],tipTranzactiePurchases[k])
			listanouaappendpurch.append(str(serieCuiPurch[k])+";"+str(tipTranzPurch[k])+";"+str(coteTVApurchases[k])+";"+str(tipTranzactiePurchases[k])+";"+str(listadeclantp_1[k]))

		listanouasetpurch=list(set(listanouaappendpurch))


		listanouaappendsales=[]

		for k in range(0,len(serieCuiSales)):
			listanouaappendsales.append(str(serieCuiSales[k])+";"+str(tipTranzSale[k])+";"+str(coteTVAsales[k])+";"+str(codTranzactieSales[k])+";"+listadeclant_1[k])

		listanouasetsales=list(set(listanouaappendsales))

		countsales=[]
		for p in range(0,len(listanouasetsales)):
			count=0
			for k in range(0,len(listanouaappendsales)):
				if(listanouaappendsales[k]==listanouasetsales[p]):
					count=count+1
			countsales.append(count)
		countpurch=[]
		for p in range(0,len(listanouasetpurch)):
			count=0
			for k in range(0,len(listanouaappendpurch)):
				if(listanouaappendpurch[k]==listanouasetpurch[p]):
					count=count+1
			countpurch.append(count)


		#print(listanouasetsales)

		tranzactii.cell(row=5,column=1).value="Cui partener"
		tranzactii.cell(row=5,column=2).value="Nume partener"
		tranzactii.cell(row=5,column=3).value="Tip partener"
		tranzactii.cell(row=5,column=4).value="Tip tranzactie"
		tranzactii.cell(row=5,column=5).value="Cota TVA"
		tranzactii.cell(row=5,column=6).value="Baza TVA"
		tranzactii.cell(row=5,column=7).value="TVA"
		tranzactii.cell(row=5,column=8).value="Nr Facturi"
		tranzactii.cell(row=5,column=9).value="Neexigibile - nu se vor raporta"
		tranzactii.cell(row=5,column=10).value="Cod si denumire NC produs(TIP V)"
		tranzactii.cell(row=2,column=1).value="Tranzactii"
		tranzactii.cell(row=2,column=1).font=cap_tabeltitlu
		counts=0
		for i in range(0, len(listanouasetsales)):
			x=listanouasetsales[i].split(";")
			#print(x[3],x[4])
			if(int(x[3])<3 or int(x[2])>0):
				counts=counts+1
				y=tranzactii.max_row
				tranzactii.cell(row=y+1,column=1).value=x[0]
				tranzactii.cell(row=y+1,column=2).value="=VLOOKUP(A"+str(y+1)+",'Mapping tranzactii'!B:J,9,0)"
				tranzactii.cell(row=y+1,column=3).value=x[3]
				tranzactii.cell(row=y+1,column=4).value=x[1]
				tranzactii.cell(row=y+1,column=5).value=x[2]
				tranzactii.cell(row=y+1,column=9).value=x[4]
				tranzactii.cell(row=y+1,column=10).value="=xlookup(K"+str(y+1)+",'Mapping tranzactii'!R:R,'Mapping tranzactii'!L:L)"
		countp=0
		for i in range(0, len(listanouasetpurch)):
			x=listanouasetpurch[i].split(";")
			if(int(x[3])<3 or int(x[2])>0):
				countp=countp+1
				y=tranzactii.max_row
				tranzactii.cell(row=y+1,column=1).value=x[0]
				tranzactii.cell(row=y+1,column=2).value="=VLOOKUP(A"+str(y+1)+",'Mapping tranzactii'!B:J,9,0)"   
				try:
					tranzactii.cell(row=y+1,column=3).value=x[3]
				except:
					tranzactii.cell(row=y+1,column=3).value=""
				tranzactii.cell(row=y+1,column=4).value=x[1]
				tranzactii.cell(row=y+1,column=5).value=x[2]
				tranzactii.cell(row=y+1,column=9).value=x[4]
				tranzactii.cell(row=y+1,column=10).value="=xlookup(K"+str(y+1)+",'Mapping tranzactii'!R:R,'Mapping tranzactii'!L:L)"

		countmare=countp+counts
		for i in range(0, countmare):
			tranzactii.cell(row=i+6,column=6).value="=SUMIFS('Mapping tranzactii'!H:H,'Mapping tranzactii'!B:B,A{0},'Mapping tranzactii'!E:E,C{0},'Mapping tranzactii'!F:F,D{0},'Mapping tranzactii'!G:G,E{0})".format(6+i)
			tranzactii.cell(row=i+6,column=7).value="=F{0}/100*E{0}".format(6+i)
			tranzactii.cell(row=i+6,column=8).value="=COUNTIFS('Mapping tranzactii'!B:B,A{0},'Mapping tranzactii'!E:E,C{0},'Mapping tranzactii'!F:F,D{0},'Mapping tranzactii'!G:G,E{0})".format(6+i)
			tranzactii.cell(row=i+6,column=11).value="=A{0}&C{0}&D{0}&E{0}".format(6+i)


		#---------FORMAT-----------------
		for row in tranzactii['A5:I5']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.alignment=Alignment(horizontal='center',vertical='center')				

		for row in tranzactii['A5:I5']:
			for cell in row:
				cell.font = cap_tabel
	
		tranzactii.column_dimensions['K'].hidden = True
		tranzactii.column_dimensions['A'].width = 20
		tranzactii.column_dimensions['I'].width = 27		
		tranzactii.column_dimensions['B'].width = 35
		tranzactii.column_dimensions['C'].width = 13
		tranzactii.column_dimensions['D'].width = 13
		tranzactii.column_dimensions['H'].width = 13
		tranzactii.auto_filter.ref = "A5:H5"
		for row in tranzactii['F6:F'+str(tranzactii.max_row)]:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'
		for row in tranzactii['G6:G'+str(tranzactii.max_row)]:
			for cell in row:
				cell.number_format='#,##0_);(#,##0)'

		saf=temp.create_sheet("Facturi storno si anulate")
		saf.sheet_view.showGridLines = False
		saf.freeze_panes = 'A4'		
		saf.cell(row=1,column=1).value="Facturi storno/anulate"
		saf.cell(row=1,column=1).font=cap_tabelbold
		saf.cell(row=3,column=1).value="Tip"
		saf.cell(row=3,column=2).value="Serie"
		saf.cell(row=3,column=3).value="Numar"
		for row in saf['A3:C3']:
			for cell in row:
				cell.font = cap_tabel
				cell.fill = cap_tabel_color_black
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		ind=saf.max_row
		for k in range(0,len(storno)):
			if(storno[k]=="Yes"):
				saf.cell(row=ind+1,column=1).value="Stornata"
				saf.cell(row=ind+1,column=2).value=""
				saf.cell(row=ind+1,column=3).value=docNoSales[k]
				ind=ind+1
		xx=saf.max_row		
		saf.cell(row=xx+1,column=1).value="Anulata"
		saf.cell(row=xx+1,column=2).value="Please input the cancelled invoice number"	
		bonuri=temp.create_sheet("Bonuri fiscale")
		bonuri.cell(row=4,column=1).value="Luna"
		bonuri.cell(row=4,column=2).value="Nr. bon fiscal"
		bonuri.cell(row=4,column=3).value="Baza 5%"
		bonuri.cell(row=4,column=4).value="TVA 5%"
		bonuri.cell(row=4,column=5).value="Baza 9%"
		bonuri.cell(row=4,column=6).value="TVA 9%"
		bonuri.cell(row=4,column=7).value="Baza 19%"
		bonuri.cell(row=4,column=8).value="TVA 19%"
		bonuri.cell(row=4,column=9).value="Baza 20%"
		bonuri.cell(row=4,column=10).value="TVA 20%"

		for row in bonuri['A4:K4']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.alignment=Alignment(horizontal='center',vertical='center')				

		for row in bonuri['A4:K4']:
			for cell in row:
				cell.font = cap_tabel
		bonuri.sheet_view.showGridLines = False

		facturi=temp.create_sheet("Sectiunea 2.1&2.2")
		facturi.sheet_view.showGridLines = False
		facturi.column_dimensions['A'].width = 12
		facturi.column_dimensions['B'].width = 18
		facturi.column_dimensions['C'].width = 15

		facturi.cell(row=1,column=1).value="Serie Emise"
		facturi.cell(row=1,column=2).value="Inceput Emise"
		facturi.cell(row=1,column=3).value="Final Emise"
		facturi.cell(row=1,column=4).value="Tip Emise"		

		docNoSales2=[]
		seriefacturi=[]
		for i in range(0,len(docNoSales)):
			# docNoSales[i].replaceAll("[^a-zA-Z0-9]", ")
			if(int(codTranzactieSales[i])<3):
				numere=re.sub("[^0-9]", "",str(docNoSales[i]))
				result = ''.join([i for i in str(docNoSales[i]) if not i.isdigit()])
				docNoSales2.append(numere)
				seriefacturi.append(result)
		#print(seriefacturi)
		initial=0
		final=0
		docNoSales2.sort()
		docNo=[]
		for k in range(0,len(docNoSales)):
			docNo.append(str(docNoSales[k]))
		docNo.sort()

		# #print(docNoSales)

		for p in range(0,len(docNoSales)-1):
			#print(docNo[p])
			if(p==0):
				initial=initial+1
				facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
				facturi.cell(row=1+initial,column=2).value=docNo[p]
				if(int(docNo[p])-int(docNo[p+1])< -1):
					final=final+1
					facturi.cell(row=1+final,column=3).value=docNo[p]
			else:
				try:
					if(int(docNo[p])-int(docNo[p-1])==1 and int(docNo[p])-int(docNo[p+1])==-1):
						print("bailando")
					
				except:
					if(int(docNo[p][3:])-int(docNo[p-1][3:])==1 and int(docNo[p][3:])-int(docNo[p+1][3:])==-1):
						print("bailando")
				try:
					if(int(docNo[p])-int(docNo[p-1])>1 and int(docNo[p])-int(docNo[p+1])==-1):
						initial=initial+1
						facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
						facturi.cell(row=1+initial,column=2).value=docNo[p]
				except:
					try:
						if(int(docNo[p][3:])-int(docNo[p-1][3:])>1 and int(docNo[p][3:])-int(docNo[p+1][3:])==-1):
							initial=initial+1
							facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
							facturi.cell(row=1+initial,column=2).value=docNo[p]
					except:
						print(docNo[p])
				try:
					if(int(docNo[p])-int(docNo[p-1])==1 and int(docNo[p])-int(docNo[p+1])<-1):
						final=final+1
						facturi.cell(row=1+final,column=3).value=docNo[p]
				except:
					if(int(docNo[p][3:])-int(docNo[p-1][3:])==1 and int(docNo[p][3:])-int(docNo[p+1][3:])<-1):
						final=final+1
						facturi.cell(row=1+final,column=3).value=docNo[p]
				try:
					if(int(docNo[p])-int(docNo[p-1])>1 and int(docNo[p])-int(docNo[p+1])<-1):
						initial=initial+1
						facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
						facturi.cell(row=1+initial,column=2).value=docNo[p]
						final=final+1
						facturi.cell(row=1+final,column=3).value=docNo[p]
				except:
					if(int(docNo[p][3:])-int(docNo[p-1][3:])>1 and int(docNo[p][3:])-int(docNo[p+1][3:])<-1):
						initial=initial+1
						facturi.cell(row=1+initial,column=1).value=seriefacturi[0]
						facturi.cell(row=1+initial,column=2).value=docNo[p]
						final=final+1
						facturi.cell(row=1+final,column=3).value=docNo[p]

		x=facturi.max_row
		facturi.auto_filter.ref = "A1:C1"
		if(int(docNoSales2[len(docNoSales2)-1])-int(docNoSales2[len(docNoSales2)-2])>1):
			facturi.cell(row=x+1,column=1).value=seriefacturi[0]
			facturi.cell(row=x+1,column=2).value=docNoSales2[len(docNoSales2)-1]
			facturi.cell(row=x+1,column=3).value=docNoSales2[len(docNoSales2)-1]
		else:
			facturi.cell(row=x+1,column=1).value=seriefacturi[0]
			facturi.cell(row=x, column=3).value = docNoSales2[len(docNoSales2) - 1]

		for k in range(2,facturi.max_row):
			facturi.cell(row=k,column=4).value=2
		yy=facturi.max_row+1
		facturi.cell(row=yy,column=1).value="Serie Alocate"
		facturi.cell(row=yy,column=2).value="Inceput Alocate"
		facturi.cell(row=yy,column=3).value="Final Alocate"
		facturi.cell(row=yy,column=4).value="Tip Alocate"		
		for kk in range(1,5):
			facturi.cell(row=yy,column=kk).font=cap_tabel
			facturi.cell(row=yy,column=kk).fill=cap_tabel_color_black	
		for pp in range(2,yy):
			facturi.cell(row=yy+pp-1,column=2).value=facturi.cell(row=pp,column=2).value
			facturi.cell(row=yy+pp-1,column=3).value=facturi.cell(row=pp,column=3).value
			facturi.cell(row=yy+pp-1,column=4).value=1						
		a23=temp.create_sheet("Sectiunea 2.3,2.4")
		dv = DataValidation(
			type='list', formula1='"Yes,No"', allow_blank=True,showDropDown=False)		
		dv.add(a23["A24"])			
		a23.sheet_view.showGridLines = False
		a23.column_dimensions['A'].width=18
		a23.column_dimensions['B'].width=13		
		a23.cell(row=1,column=1).value="Sectiunea 2.3"
		a23.cell(row=3,column=1).value="Denumire beneficiar"
		a23.cell(row=3,column=2).value="CUI beneficiar"
		a23.cell(row=3,column=1).fill=cap_tabel_color_black
		a23.cell(row=3,column=1).font=cap_tabel		
		a23.cell(row=36,column=1).border=border_bottom
		a23.cell(row=36,column=2).border=border_lowerleft
		a23.cell(row=35,column=2).border=border_right
		a23.cell(row=34,column=2).border=border_right
		a23.cell(row=33,column=2).border=border_right
		a23.cell(row=32,column=2).border=border_right						
		a23.cell(row=3,column=2).fill=cap_tabel_color_black
		a23.cell(row=3,column=2).font=cap_tabel


		a23.cell(row=6,column=1).value="Seria"
		a23.cell(row=6,column=2).value="De la"
		a23.cell(row=6,column=3).value="La"



		for row in a23['A6:C6']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		for row in a23['A14:C14']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				

		a23.cell(row=9,column=1).value="Sectiunea 2.4"
		a23.cell(row=9,column=1).font=cap_tabelbold
		a23.cell(row=1,column=1).font=cap_tabelbold		
		a23.cell(row=28,column=1).font=cap_tabelbold		
		a23.cell(row=4,column=1).border=border_lowerright
		a23.cell(row=4,column=2).border=border_lowerright		
		a23.cell(row=12,column=1).border=border_lowerright
		a23.cell(row=12,column=2).border=border_lowerright		
		a23.cell(row=11,column=1).fill=cap_tabel_color_black
		a23.cell(row=11,column=1).font=cap_tabel		
		a23.cell(row=7,column=1).border=border_lowerright
		a23.cell(row=7,column=2).border=border_lowerright
		a23.cell(row=7,column=3).border=border_lowerright
		a23.cell(row=15,column=1).border=border_lowerright
		a23.cell(row=15,column=2).border=border_lowerright
		a23.cell(row=15,column=3).border=border_lowerright				
		a23.cell(row=11,column=2).fill=cap_tabel_color_black
		a23.cell(row=11,column=2).font=cap_tabel
		a23.cell(row=11,column=1).value="Denumire tert"

		a23.cell(row=11,column=2).value="CUI tert"

		a23.cell(row=14,column=1).value="Seria"
		a23.cell(row=14,column=2).value="De la"
		a23.cell(row=14,column=3).value="La"
		for row in a23['A31:B31']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.font=cap_tabel
		try:
			if int(val4)== 1:
				a3=temp.create_sheet("Sectiunea 3")
				a3.cell(row=1,column=1).value="Sectiunea 3"
				a3.cell(row=1,column=1).font=cap_tabeltitlu
				a3.cell(row=3,column=1).value="In cazul in care soldul sumei negative inregistrate in decontul de TVA aferent perioadei de raportare este solicitat la rambursare , se vor selecta datele cu privire la natura operatiunilor din care provine acesta"
				a3.cell(row=7,column=1).value="Achizitii de bunuri si servicii legate direct de bunurile imobile din urmatoarele categorii"
				a3.cell(row=8,column=1).value="a) parcuri eoliene"
				a3.cell(row=9,column=1).value="b) constructii rezidentiale"



				a3.cell(row=10,column=1).value="c) cladiri de birouri"
				a3.cell(row=11,column=1).value="d) constructii industriale"
				a3.cell(row=12,column=1).value="e) altele"
				a3.cell(row=14,column=1).value="Achizitii de bunuri, cu exceptia celor legate direct de bunuri imobile:"
				a3.cell(row=15,column=1).value="a) cu cota de TVA de 24%"
				a3.cell(row=16,column=1).value="b) cu cota standard de TVA de 20%"
				a3.cell(row=17,column=1).value="c) cu cota de TVA de 19%"
				a3.cell(row=18,column=1).value="d) cu cota de TVA de 9%"
				a3.cell(row=19,column=1).value="e) cu cota de TVA de 5%"
				a3.cell(row=21,column=1).value="Achizitii de servicii, cu exceptia celor legate direct de bunurile imobile:"
				a3.cell(row=22,column=1).value="a) cu cota de TVA de 24%"
				a3.cell(row=23,column=1).value="b) cu cota standard de TVA de 20%"
				a3.cell(row=24,column=1).value="c) cu cota de TVA de 19%"
				a3.cell(row=25,column=1).value="d) cu cota de TVA de 9%"
				a3.cell(row=26,column=1).value="e) cu cota de TVA de 5%"
				a3.cell(row=28,column=1).value="Importuri de bunuri"
				a3.cell(row=30,column=1).value="Achizitii imobilizari necorporale"
				a3.cell(row=32,column=1).value="Livrari de bunuri imobile"
				a3.cell(row=34,column=1).value="Livrari de bunuri, cu exceptia bunurilor imobile:"
				a3.cell(row=35,column=1).value="a) cu cota de TVA de 24%"
				a3.cell(row=36,column=1).value="b) cu cota standard de TVA de 20%"
				a3.cell(row=37,column=1).value="c) cu cota de TVA de 19%"
				a3.cell(row=38,column=1).value="d) cu cota de TVA de 9%"
				a3.cell(row=39,column=1).value="e) cu cota de TVA de 5%"
				a3.cell(row=41,column=1).value="Livrari de bunuri scutite de TVA"
				a3.cell(row=43,column=1).value="Livrari de bunuri/prestari de servicii pt care se aplica taxarea inversa"
				a3.cell(row=45,column=1).value="Prestari de servicii:"
				a3.cell(row=46,column=1).value="a) cu cota de TVA de 24%"
				a3.cell(row=47,column=1).value="b) cu cota standard de TVA de 20%"
				a3.cell(row=48,column=1).value="c) cu cota de TVA de 19%"
				a3.cell(row=49,column=1).value="d) cu cota de TVA de 9%"
				a3.cell(row=50,column=1).value="e) cu cota de TVA de 5%"
				a3.cell(row=52,column=1).value="Prestari de servicii scutite de TVA"
				a3.cell(row=54,column=1).value="Livrari intracomunitare de bunuri"
				a3.cell(row=56,column=1).value="Prestari intracomunitare de servicii"
				a3.cell(row=58,column=1).value="Exporturi de bunuri"
				a3.cell(row=60,column=1).value="Livrari imobilizari necorporale"
				a3.cell(row=62,column=1).value="Persoana impozabila nu a efectuat livrari de bunuri/prestari de servicii in perioada de raportare"																					
		except:
			pass


																																																																					
		a5=temp.create_sheet("Sectiunea 5")
		a5.sheet_view.showGridLines = False
		a5.cell(row=2,column=1).value="Sectiunea 5"
		a5.cell(row=2,column=1).font=cap_tabeltitlu
		a5.cell(row=4,column=1).font=cap_tabelbold
		a5.cell(row=6,column=1).font=cap_tabelbold
		a5.column_dimensions['B'].width=13
		a5.merge_cells('A6:J7')
		a5.merge_cells('A18:J19')		
		
		for pop in range(1,11):
			a5.cell(row=5,column=pop).border=border_bottom
			a5.cell(row=7,column=pop).border=border_bottom
			a5.cell(row=17,column=pop).border=border_bottom
			a5.cell(row=19,column=pop).border=border_bottom
		a5.cell(row=6,column=10).border=border_upperright
		a5.cell(row=7,column=10).border=border_lowerright
		a5.cell(row=18,column=10).border=border_upperright
		a5.cell(row=19,column=10).border=border_lowerright									
		a5['A6'].alignment=Alignment(wrap_text=True)
		a5['A18'].alignment=Alignment(wrap_text=True)
		a5['A30'].alignment=Alignment(wrap_text=True)						
		a5.cell(row=4,column=1).value="Sectiune 5.2"
		a5.cell(row=16,column=1).font=cap_tabelbold		
		a5.cell(row=16,column=1).value="Sectiune 5.3"
		a5.cell(row=28,column=1).font=cap_tabelbold		
		a5.cell(row=10,column=1).value="cota 24%"
		a5.cell(row=11,column=1).value="cota 20%"
		a5.cell(row=12,column=1).value="cota 19%"
		a5.cell(row=13,column=1).value="cota 9%"
		a5.cell(row=14,column=1).value="cota 5%"
		a5.cell(row=10,column=2).value=0
		a5.cell(row=11,column=2).value=0
		a5.cell(row=12,column=2).value=0
		a5.cell(row=13,column=2).value=0
		a5.cell(row=14,column=2).value=0

		a5.cell(row=9,column=2).value="Valoare TVA"
		a5.cell(row=9,column=1).value="Cota"		

		a5.cell(row=6,column=1).value="5.2 TVA deductibila aferenta facturilor achitate in perioada de raportare indiferent de data in care acestea au fost primite de la persoane impozabile care aplica sistemul normal de TVA, defalcata pe fiecare cota de TVA"
		a5.cell(row=18,column=1).font=cap_tabelbold		

		a5.cell(row=22,column=1).value="cota 24%"
		a5.cell(row=23,column=1).value="cota 20%"
		a5.cell(row=24,column=1).value="cota 19%"
		a5.cell(row=25,column=1).value="cota 9%"
		a5.cell(row=26,column=1).value="cota 5%"

		a5.cell(row=22,column=2).value=0
		a5.cell(row=23,column=2).value=0
		a5.cell(row=24,column=2).value=0
		a5.cell(row=25,column=2).value=0
		a5.cell(row=26,column=2).value=0

		a5.cell(row=21,column=1).value="Cota"		
		a5.cell(row=21,column=2).value="Valoare TVA"
		for row in a5['A9:B9']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		for row in a5['A21:B21']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
							
		a5.cell(row=18,column=1).value="5.3 TVA deductibila aferenta facturilor achitate in perioada de raportare indiferent de data in care acestea au fost primite de la persoane impozabile care aplica sistemul de TVA la incasare, defalcata pe fiecare cota de TVA"

		for kk in range(22,27):
			a5.cell(row=kk,column=1).border=border_lowerright
			a5.cell(row=kk,column=2).border=border_lowerright						
		for kk in range(10,15):
			a5.cell(row=kk,column=1).border=border_lowerright
			a5.cell(row=kk,column=2).border=border_lowerright		
		a6=temp.create_sheet("Sectiunea 6")
		a6.sheet_view.showGridLines = False
		a6.cell(row=2,column=1).value="Sectiunea 6"
		a6.cell(row=2,column=1).font=cap_tabeltitlu
		a6.cell(row=4,column=1).font=cap_tabelbold
		a6.cell(row=6,column=1).font=cap_tabelbold
		a6.cell(row=15,column=1).font=cap_tabelbold
		a6.cell(row=13,column=1).font=cap_tabelbold
		a6.column_dimensions['A'].width=17			
		a6.column_dimensions['B'].width=28
		a6.column_dimensions['C'].width=17
															
		a6.cell(row=4,column=1).value="Sectiunea 6.1"		
		a6.cell(row=6,column=1).value="6.1 Persoanele impozabile care aplica regimul special pt agentiile de turism, vor completa:"
		a6.merge_cells('A6:F7')
		for l in range(1,7):
			a6.cell(row=5,column=l).border=border_bottom
			a6.cell(row=7,column=l).border=border_bottom
		a6.cell(row=6,column=6).border=border_upperright
		a6.cell(row=7,column=6).border=border_lowerright						

		a6.merge_cells('A15:G16')

		for l in range(1,8):
			a6.cell(row=14,column=l).border=border_bottom
			a6.cell(row=16,column=l).border=border_bottom
		a6.cell(row=15,column=7).border=border_upperright
		a6.cell(row=16,column=7).border=border_lowerright




		a6.cell(row=9,column=1).value="Incasarile agentiei"
		a6.cell(row=9,column=2).value="Costurile agentiei de turism"
		a6.cell(row=9,column=3).value="Marja de profit"
		a6.cell(row=9,column=4).value="TVA"

		a6.cell(row=13,column=1).value="Sectiunea 6.2"

		a6.cell(row=15,column=1).value="6.2 Persoanele impozabile care aplica regimul special pt bunurile second-hand, opere de arta, obiecte de colectie si antichitati , vor completa:"

		for row in a6['A18:D18']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		for row in a6['A9:D9']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')										


		a6.cell(row=18,column=1).value="Pret de vanzare"
		a6.cell(row=18,column=2).value="Pret de cumparare"
		a6.cell(row=18,column=3).value="Marja de profit"	
		a6.cell(row=18,column=4).value="TVA"
		for oo in range(1,5):
			a6.cell(row=10,column=oo).border=border_lowerright
			a6.cell(row=19,column=oo).border=border_lowerright		

		a6['A6'].alignment=Alignment(wrap_text=True)

		a6['A15'].alignment=Alignment(wrap_text=True)
		a7=temp.create_sheet(" Sectiunea 7 ")

		a7.cell(row=2,column=1).value="Sectiunea 7"
		a7.sheet_view.showGridLines = False
		a7.cell(row=2,column=1).font=cap_tabeltitlu
		a7.cell(row=5,column=1).value="7. In situatia in care ati desfasurat, in perioada de raportare, activitati dintre cele inscrise in lista veti selecta activitatea corespunzatoare si veti inscrie valoarea livrarilor/prestarilor, precum si TVA aferenta"
		a7['A5'].alignment=Alignment(wrap_text=True)
		a7.merge_cells('A5:I6')
		a7.cell(row=5,column=1).font=cap_tabelbold
		for ii in range(1,10):
			a7.cell(row=4,column=ii).border=border_bottom
			a7.cell(row=6,column=ii).border=border_bottom
		a7.cell(row=5,column=9).border=border_upperright
		a7.cell(row=6,column=9).border=border_lowerright		
		a7.cell(row=5,column=8).border=border_right
		a7.cell(row=6,column=8).border=border_right

		for jj in range(1,4):
			a7.cell(row=9,column=jj).border=border_bottom
			a7.cell(row=9,column=jj).border=border_right
		
		for pp in range(12,17):
			a7.cell(row=pp,column=2).border=border_lowerright
			a7.cell(row=pp,column=1).border=border_lowerright
		a7.cell(row=8,column=1).value="Activitate"
		a7.cell(row=8,column=2).value="Tip operatiune"
		a7.cell(row=8,column=3).value="Valoarea livrarilor/prestarilor"

		a7.cell(row=11,column=1).value="Cota"
		a7.cell(row=12,column=1).value="cota 24%"
		a7.cell(row=13,column=1).value="cota 20%"
		a7.cell(row=14,column=1).value="cota 19%"
		a7.cell(row=15,column=1).value="cota 9%"
		a7.cell(row=16,column=1).value="cota 5%"
		for row in a7['A8:C8']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.font = cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')										
		for row in a7['A11:B11']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.font = cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		a7.cell(row=11,column=2).value="Valoare TVA"

		for row in facturi['A1:D1']:
			for cell in row:
				cell.fill = cap_tabel_color_black
				cell.alignment=Alignment(horizontal='center',vertical='center')				

		for row in facturi['A1:D1']:
			for cell in row:
				cell.font = cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				

		facturi.column_dimensions['C'].width = 14
		facturi.column_dimensions['D'].width = 14
		a23.column_dimensions['A'].width = 18
		a23.column_dimensions['B'].width = 13
		a23.column_dimensions['C'].width = 10
		a6.column_dimensions['A'].width = 17
		a6.column_dimensions['B'].width = 25
		a6.column_dimensions['C'].width = 15				
		sumaryG=temp.create_sheet("Sectiunea G. Manual input")
		sumaryG.sheet_view.showGridLines = False
		sumaryG.cell(row=2,column=1).value="Sectiunea G"
		sumaryG.cell(row=2,column=1).font=cap_tabeltitlu		
		sumaryG.column_dimensions['A'].width =45
		sumaryG.cell(row=5,column=1).value="Total Nr. Bonuri Fiscale"
		sumaryG.cell(row=5,column=1).font=cap_tabelbold
		sumaryG.cell(row=6,column=1).font=cap_tabelbold
		sumaryG.cell(row=7,column=1).font=cap_tabelbold						
		sumaryG.cell(row=6,column=1).value="Total incasari in perioada de raportare prin intermediul AMEF ( aparate de marcatelectronice fiscale ) inclusiv incasarile prin intermediul bonurilor fiscale care indeplinesc conditiile unei facturi simplificate indiferent daca au/nu au inscris codul de inregistrare in scopuri de TVA al beneficiarului (i1)"
		sumaryG['A6'].alignment=Alignment(wrap_text=True)
		sumaryG['A7'].alignment=Alignment(wrap_text=True)				
		sumaryG.cell(row=7,column=1).value="Total incasari in perioada de raportare efectuate din activitati exceptate de la obligatia utilizarii AMEF***) (i2) conform prevederilor legale in vigoare )"
		sumaryG.merge_cells('A9:H11')
		sumaryG.cell(row=9,column=1).font=cap_tabelbold
		for ii in range(1,9):
			sumaryG.cell(row=8,column=ii).border=border_bottom
			sumaryG.cell(row=11,column=ii).border=border_bottom
		sumaryG.cell(row=9,column=8).border=border_upperright
		sumaryG.cell(row=10,column=8).border=border_right		
		sumaryG.cell(row=11,column=8).border=border_lowerright		
		for ii in range(1,9):	
			sumaryG.cell(row=19,column=ii).border=border_bottom
			sumaryG.cell(row=20,column=ii).border=border_bottom			
		sumaryG.cell(row=5,column=2).border=border_right
		sumaryG.cell(row=6,column=2).border=border_right
		sumaryG.cell(row=7,column=2).border=border_right
		sumaryG.cell(row=5,column=1).border=border_right
		sumaryG.cell(row=6,column=1).border=border_right
		sumaryG.cell(row=7,column=1).border=border_right
		sumaryG.cell(row=4,column=1).border=border_bottom
		sumaryG.cell(row=5,column=1).border=border_bottom
		sumaryG.cell(row=6,column=1).border=border_bottom
		sumaryG.cell(row=7,column=1).border=border_bottom
		sumaryG.cell(row=5,column=2).value=0
		sumaryG.cell(row=6,column=2).value=0
		sumaryG.cell(row=7,column=2).value=0
		sumaryG.cell(row=13,column=1).value="Cota"		
		sumaryG.cell(row=13,column=2).value="Total baza impozabila"
		sumaryG.cell(row=13,column=3).value="TVA"
		sumaryG.cell(row=23,column=1).value="Cota"		
		sumaryG.cell(row=23,column=2).value="Total baza impozabila"
		sumaryG.cell(row=23,column=3).value="TVA"		
		for ii in range(1,4):
			sumaryG.cell(row=17,column=ii).border=border_lowerright
			sumaryG.cell(row=16,column=ii).border=border_lowerright
			sumaryG.cell(row=15,column=ii).border=border_lowerright
			sumaryG.cell(row=14,column=ii).border=border_lowerright
			sumaryG.cell(row=24,column=ii).border=border_lowerright
			sumaryG.cell(row=25,column=ii).border=border_lowerright
			sumaryG.cell(row=26,column=ii).border=border_lowerright
			sumaryG.cell(row=27,column=ii).border=border_lowerright																					
						
		for row in sumaryG['A13:C13']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		for row in sumaryG['A23:C23']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')										
		sumaryG.cell(row=9,column=1).value="Incasari in perioada de raportare prin intermediul AMEF ( aparate de marcat electronice fiscale ) inclusiv incasarile prin intermediul bonurilor fiscale care indeplinesc conditiile unei facturi simplificate indiferent daca au/nu au inscris codul de inregistrare in scopuri de TVA al beneficiarului (i1)"
		sumaryG.cell(row=20,column=1).value="Incasari in perioada de raportare efectuate din activitati exceptate de la obligatia utilizarii AMEF***) (i2) conform prevederilor legale in vigoare )"
		sumaryG['A9'].alignment=Alignment(wrap_text=True)		
		sumaryG.cell(row=14,column=1).value="Cota 20%"
		sumaryG.cell(row=15,column=1).value="Cota 19%"
		sumaryG.cell(row=16,column=1).value="Cota 9%"
		sumaryG.cell(row=17,column=1).value="Cota 5%"

		sumaryG.cell(row=24,column=1).value="Cota 20%"
		sumaryG.cell(row=25,column=1).value="Cota 19%"
		sumaryG.cell(row=26,column=1).value="Cota 9%"
		sumaryG.cell(row=27,column=1).value="Cota 5%"

		sumaryG.cell(row=14,column=2).value=0
		sumaryG.cell(row=15,column=2).value=0
		sumaryG.cell(row=16,column=2).value=0
		sumaryG.cell(row=17,column=2).value=0


		sumaryG.cell(row=14,column=3).value=0
		sumaryG.cell(row=15,column=3).value=0
		sumaryG.cell(row=16,column=3).value=0
		sumaryG.cell(row=17,column=3).value=0


		sumaryG.cell(row=24,column=2).value=0
		sumaryG.cell(row=25,column=2).value=0
		sumaryG.cell(row=26,column=2).value=0
		sumaryG.cell(row=27,column=2).value=0

		sumaryG.cell(row=24,column=3).value=0
		sumaryG.cell(row=25,column=3).value=0
		sumaryG.cell(row=26,column=3).value=0
		sumaryG.cell(row=27,column=3).value=0


		sumaryI=temp.create_sheet("Sectiunea I 1. Manual input")

		for row in sumaryI['A7:C7']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		for row in sumaryI['A17:C17']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		for row in sumaryI['A27:C27']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		for row in sumaryI['A37:C37']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')				
		for row in sumaryI['A47:C47']:
			for cell in row:
				cell.fill=cap_tabel_color_black
				cell.font=cap_tabel
				cell.alignment=Alignment(horizontal='center',vertical='center')																
		sumaryI.cell(row=2,column=1).value="Sectiunea I"
		sumaryI.cell(row=2,column=1).font=cap_tabeltitlu
		sumaryI.sheet_view.showGridLines = False
		sumaryI.cell(row=7,column=1).value="Cota"
		sumaryI.cell(row=7,column=2).value="Baza impozabila"
		sumaryI.cell(row=7,column=3).value="TVA"

		sumaryI.cell(row=7,column=1).value="Cota"
		sumaryI.cell(row=7,column=2).value="Baza impozabila"
		sumaryI.cell(row=7,column=3).value="TVA"
		sumaryI.cell(row=17,column=1).value="Cota"
		sumaryI.cell(row=17,column=2).value="Baza impozabila"
		sumaryI.cell(row=17,column=3).value="TVA"
		sumaryI.cell(row=27,column=1).value="Cota"
		sumaryI.cell(row=27,column=2).value="Baza impozabila"
		sumaryI.cell(row=27,column=3).value="TVA"
		sumaryI.cell(row=37,column=1).value="Cota"
		sumaryI.cell(row=37,column=2).value="Baza impozabila"
		sumaryI.cell(row=37,column=3).value="TVA"								
		sumaryI.cell(row=47,column=1).value="Cota"
		sumaryI.cell(row=47,column=2).value="Baza impozabila"
		sumaryI.cell(row=47,column=3).value="TVA"
		sumaryI.cell(row=4,column=1).value="1.1 Livrari de bunuri/prestari de servicii pentru care s-au emis facturi simplificate care au inscris codul de inregistrare in scopuri de TVA al beneficiarului %"
		sumaryI.cell(row=4,column=1).font=cap_tabelbold
		sumaryI.cell(row=14,column=1).font=cap_tabelbold
		sumaryI.cell(row=24,column=1).font=cap_tabelbold
		sumaryI.cell(row=34,column=1).font=cap_tabelbold
		sumaryI.cell(row=44,column=1).font=cap_tabelbold								
		sumaryI.cell(row=14,column=1).value="1.2 Livrari de bunuri/prestari de servicii pentru care s-au emis facturi simplificate fara a avea inscris codul de inregistrare in scopuri de TVA al beneficiarului %"
		sumaryI['A4'].alignment=Alignment(wrap_text=True)
		sumaryI['A14'].alignment=Alignment(wrap_text=True)
		sumaryI['A24'].alignment=Alignment(wrap_text=True)
		sumaryI['A34'].alignment=Alignment(wrap_text=True)
		sumaryI['A44'].alignment=Alignment(wrap_text=True)								
		sumaryI.cell(row=24,column=1).value="1.3 Achizitii de bunuri si servicii pentru care s-au primit facturi simplificate de la persoane impozabile care aplica sistemul normal de TVA si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

		sumaryI.cell(row=34,column=1).value="1.4 Achizitii de bunuri si servicii pentru care s-au primit facturi simplificate de la persoane impozabile care aplica sistemul de TVA la incasare si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

		sumaryI.cell(row=44,column=1).value="1.5 Achizitii de bunuri si servicii pentru care s-au primit bonuri fiscale care indeplinesc conditiile unei facturi simplificate si care au inscris codul de inregistrare in scopuri de TVA al beneficiarului"

		for ip in range(1,11):
			sumaryI.cell(row=3,column=ip).border=border_bottom
			sumaryI.cell(row=6,column=ip).border=border_top
			sumaryI.cell(row=13,column=ip).border=border_bottom
			sumaryI.cell(row=16,column=ip).border=border_top
			sumaryI.cell(row=23,column=ip).border=border_bottom
			sumaryI.cell(row=26,column=ip).border=border_top
			sumaryI.cell(row=33,column=ip).border=border_bottom
			sumaryI.cell(row=36,column=ip).border=border_top
			sumaryI.cell(row=43,column=ip).border=border_bottom
			sumaryI.cell(row=46,column=ip).border=border_top
		sumaryI.cell(row=4,column=10).border=border_upperright
		sumaryI.cell(row=5,column=10).border=border_lowerright
		sumaryI.cell(row=14,column=10).border=border_upperright
		sumaryI.cell(row=15,column=10).border=border_lowerright
		sumaryI.cell(row=24,column=10).border=border_upperright
		sumaryI.cell(row=25,column=10).border=border_lowerright
		sumaryI.cell(row=34,column=10).border=border_upperright
		sumaryI.cell(row=35,column=10).border=border_lowerright
		sumaryI.cell(row=44,column=10).border=border_upperright
		sumaryI.cell(row=45,column=10).border=border_lowerright						

		sumaryI.cell(row=4,column=10).border=border_upperright
		sumaryI.cell(row=5,column=10).border=border_lowerright
		sumaryI.cell(row=14,column=10).border=border_upperright
		sumaryI.cell(row=15,column=10).border=border_lowerright
		sumaryI.cell(row=24,column=10).border=border_upperright
		sumaryI.cell(row=25,column=10).border=border_lowerright
		sumaryI.cell(row=34,column=10).border=border_upperright
		sumaryI.cell(row=35,column=10).border=border_lowerright
		sumaryI.cell(row=44,column=10).border=border_upperright
		sumaryI.cell(row=45,column=10).border=border_lowerright										
		sumaryI.cell(row=4,column=10).border=border_right
		sumaryI.cell(row=5,column=10).border=border_right
		sumaryI.cell(row=14,column=10).border=border_right
		sumaryI.cell(row=15,column=10).border=border_right
		sumaryI.cell(row=24,column=10).border=border_right
		sumaryI.cell(row=25,column=10).border=border_right
		sumaryI.cell(row=34,column=10).border=border_right
		sumaryI.cell(row=35,column=10).border=border_right
		sumaryI.cell(row=44,column=10).border=border_right
		sumaryI.cell(row=45,column=10).border=border_right																																																	
		for io in range(8,13):
			sumaryI.cell(row=io,column=1).border=border_lowerright
			sumaryI.cell(row=io,column=2).border=border_lowerright
			sumaryI.cell(row=io,column=3).border=border_lowerright			
		for io in range(18,23):
			sumaryI.cell(row=io,column=1).border=border_lowerright
			sumaryI.cell(row=io,column=2).border=border_lowerright
			sumaryI.cell(row=io,column=3).border=border_lowerright			
		for io in range(28,33):
			sumaryI.cell(row=io,column=1).border=border_lowerright
			sumaryI.cell(row=io,column=2).border=border_lowerright
			sumaryI.cell(row=io,column=3).border=border_lowerright			
		for io in range(38,43):
			sumaryI.cell(row=io,column=1).border=border_lowerright
			sumaryI.cell(row=io,column=2).border=border_lowerright
			sumaryI.cell(row=io,column=3).border=border_lowerright			
		for io in range(48,53):
			sumaryI.cell(row=io,column=1).border=border_lowerright
			sumaryI.cell(row=io,column=2).border=border_lowerright
			sumaryI.cell(row=io,column=3).border=border_lowerright																		
		sumaryI.merge_cells('A4:J5')
		sumaryI.merge_cells('A14:J15')
		sumaryI.merge_cells('A24:J25')
		sumaryI.merge_cells('A34:J35')
		sumaryI.merge_cells('A44:J45')
								
		sumaryI.cell(row=8,column=1).value="Cota 24%"
		sumaryI.cell(row=9,column=1).value="Cota 20%"
		sumaryI.cell(row=10,column=1).value="Cota 19%"
		sumaryI.cell(row=11,column=1).value="Cota 9%"
		sumaryI.cell(row=12,column=1).value="Cota 5%"

		sumaryI.cell(row=8,column=2).value=0
		sumaryI.cell(row=9,column=2).value=0
		sumaryI.cell(row=10,column=2).value=0
		sumaryI.cell(row=11,column=2).value=0
		sumaryI.cell(row=12,column=2).value=0

		sumaryI.cell(row=18,column=2).value=0
		sumaryI.cell(row=19,column=2).value=0
		sumaryI.cell(row=20,column=2).value=0
		sumaryI.cell(row=21,column=2).value=0
		sumaryI.cell(row=22,column=2).value=0

		sumaryI.cell(row=28,column=2).value=0
		sumaryI.cell(row=29,column=2).value=0
		sumaryI.cell(row=30,column=2).value=0
		sumaryI.cell(row=31,column=2).value=0
		sumaryI.cell(row=32,column=2).value=0

		sumaryI.cell(row=38,column=2).value=0
		sumaryI.cell(row=39,column=2).value=0
		sumaryI.cell(row=40,column=2).value=0
		sumaryI.cell(row=41,column=2).value=0
		sumaryI.cell(row=42,column=2).value=0

		sumaryI.cell(row=48,column=2).value=0
		sumaryI.cell(row=49,column=2).value=0
		sumaryI.cell(row=50,column=2).value=0
		sumaryI.cell(row=51,column=2).value=0
		sumaryI.cell(row=52,column=2).value=0

		sumaryI.cell(row=8,column=3).value=0
		sumaryI.cell(row=9,column=3).value=0
		sumaryI.cell(row=10,column=3).value=0
		sumaryI.cell(row=11,column=3).value=0
		sumaryI.cell(row=12,column=3).value=0

		sumaryI.cell(row=18,column=3).value=0
		sumaryI.cell(row=19,column=3).value=0
		sumaryI.cell(row=20,column=3).value=0
		sumaryI.cell(row=21,column=3).value=0
		sumaryI.cell(row=22,column=3).value=0

		sumaryI.cell(row=28,column=3).value=0
		sumaryI.cell(row=29,column=3).value=0
		sumaryI.cell(row=30,column=3).value=0
		sumaryI.cell(row=31,column=3).value=0
		sumaryI.cell(row=32,column=3).value=0

		sumaryI.cell(row=38,column=3).value=0
		sumaryI.cell(row=39,column=3).value=0
		sumaryI.cell(row=40,column=3).value=0
		sumaryI.cell(row=41,column=3).value=0
		sumaryI.cell(row=42,column=3).value=0

		sumaryI.cell(row=48,column=3).value=0
		sumaryI.cell(row=49,column=3).value=0
		sumaryI.cell(row=50,column=3).value=0
		sumaryI.cell(row=51,column=3).value=0
		sumaryI.cell(row=52,column=3).value=0		

		sumaryI.cell(row=8,column=1).value="Cota 24%"
		sumaryI.cell(row=9,column=1).value="Cota 20%"
		sumaryI.cell(row=10,column=1).value="Cota 19%"
		sumaryI.cell(row=11,column=1).value="Cota 9%"
		sumaryI.cell(row=12,column=1).value="Cota 5%"

		sumaryI.cell(row=18,column=1).value="Cota 24%"
		sumaryI.cell(row=19,column=1).value="Cota 20%"
		sumaryI.cell(row=20,column=1).value="Cota 19%"
		sumaryI.cell(row=21,column=1).value="Cota 9%"
		sumaryI.cell(row=22,column=1).value="Cota 5%"

		sumaryI.cell(row=28,column=1).value="Cota 24%"
		sumaryI.cell(row=29,column=1).value="Cota 20%"
		sumaryI.cell(row=30,column=1).value="Cota 19%"
		sumaryI.cell(row=31,column=1).value="Cota 9%"
		sumaryI.cell(row=32,column=1).value="Cota 5%"

		sumaryI.cell(row=38,column=1).value="Cota 24%"
		sumaryI.cell(row=39,column=1).value="Cota 20%"
		sumaryI.cell(row=40,column=1).value="Cota 19%"
		sumaryI.cell(row=41,column=1).value="Cota 9%"
		sumaryI.cell(row=42,column=1).value="Cota 5%"

		sumaryI.cell(row=48,column=1).value="Cota 24%"
		sumaryI.cell(row=49,column=1).value="Cota 20%"
		sumaryI.cell(row=50,column=1).value="Cota 19%"
		sumaryI.cell(row=51,column=1).value="Cota 9%"
		sumaryI.cell(row=52,column=1).value="Cota 5%"

		sumaryI.column_dimensions['B'].width = 15	
		sumaryG.column_dimensions['B'].width = 20	
		a7.column_dimensions['B'].width = 12						
		a7.column_dimensions['C'].width = 22
		a6.column_dimensions['B'].width = 22						
		a6.column_dimensions['C'].width = 12


		# for row in sumary['A5:D34']:
		# 	for cell in row:
		# 		cell.border = border_thin

		# for row in sumary['F5:I34']:
		# 	for cell in row:
		# 		cell.border = border_thin

		# for row in sumary['K5:N34']:
		# 	for cell in row:
		# 		cell.border = border_thin

		# for row in sumary['P5:S34']:
		# 	for cell in row:
		# 		cell.border = border_thin

		# for row in sumary['U5:X34']:
		# 	for cell in row:
		# 		cell.border = border_thin

		# listanoua=['A','B','C','D','F','G','H','I','K','L','M','N','P','Q','R','S','U','V','W','W','X']
		# for column in listanoua:
		# 	for i in listanoua:
		# 		if (column==i):
# 			sumary.column_dimensions[column].width = 15

		
		# for i in range(0 ,len(tip)):
	# folderpath="D:/D300 to XML/docs"
	folderpath="C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage spreeadsheet"
	# folderpath="C:/Users/Cristian.Iordache/Documents/D300 to XML Final CI/D300 to XML 2/storage"
	file_pathFS = os.path.join(folderpath, "One VAT app spreadsheets " +str(clientname)+".xlsx")
	temp.save(file_pathFS)
	# return send_from_directory("D:/D300 to XML/docs","One VAT app spreadsheets.xlsx",as_attachment=True)
	return send_from_directory("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage spreeadsheet","One VAT app spreadsheets " +str(clientname)+".xlsx",as_attachment=True)
	return render_template('D3APPS2')
@app.route('/D3APPS2')
def my_form2():
    return render_template('D3APPS second step.html')

@app.route('/D3APPS2', methods=['POST', 'GET'])
def D300():
	if request.method == 'POST':
		D300_2= request.files["d300file2"]
	temp = openpyxl.load_workbook(D300_2,data_only=True)
	try:
		amount=temp['D300 draft figures']
		info=temp['Other info']
		an=info.cell(row=20,column=3).value
		#print(an)
		luna=info.cell(row=21,column=3).value
		# temei=info.cell(row=8,column=2).value
		cif=info.cell(row=27,column=3).value
		den=info.cell(row=26,column=3).value

		# #print(an, 'an')

		# cif=info.cell(row=11,column=2).value
		# adresa="strada: "str(info.cell(row=12,column=2).value)+str(info.cell(row=13,column=2).value)+ ",localitate: "+str(info.cell(row=14,column=2).value)+" "+str(info.cell(row=16,column=2).value)+",judet: "+str(info.cell(row=15,column=2).value)+",cod postal: "+str(info.cell(row=21,column=2).value)
		strada=info.cell(row=28,column=3).value
		numar=info.cell(row=29,column=3).value
		localitate=info.cell(row=30,column=3).value
		judet=info.cell(row=31,column=3).value

		sector=info.cell(row=32,column=3).value
		if(sector==None):
			sector=""
		bloc=info.cell(row=33,column=3).value
		if(bloc==None):
			bloc=""
		scara=info.cell(row=34,column=3).value
		if(scara==None):
			scara=""
		etaj=info.cell(row=35,column=3).value
		if(etaj==None):
			etaj=""
		apt=info.cell(row=36,column=3).value
		if(apt==None):
			apt=""
		strada=strada
		# codpostal=info.cell(row=37,column=3).value
		telefon=info.cell(row=29,column=3).value
		email=info.cell(row=31,column=3).value
		banca=info.cell(row=32,column=3).value
		contban=info.cell(row=33,column=3).value
		Caen=info.cell(row=34,column=3).value
		prorata=info.cell(row=35,column=3).value
		cereale=info.cell(row=36,column=3).value
		telmob=info.cell(row=37,column=3).value
		disp=info.cell(row=38,column=3).value
		nr_evid=int(info.cell(row=18,column=3).value)
		cons=info.cell(row=39,column=3).value
		ramburs=info.cell(row=40,column=3).value
		if(ramburs=="N"):
			ramburs2=0
		else:
			ramburs2=1
		# total_plata=R1_1+R2_1+R3_1+R3_1_1+R4_1+R5_1+R5_1_1+R6_1+R7_1+R7_1_1+R8_1+R9_1+R10_1+R11_1+R12_1+R12_1_1+R12_2_1+R12_3_1+R13_1+R14_1+R15_1+R16_1+R64_1+R65_1+R17_1+R18_1+R18_1_1+R19_1+R20_1+R20_1_1+R21_1+R22_1+R23_1+R24_1+R25_1+R25_1_1+R25_2_1+R25_3_1+R43_1+R44_1+R26_1+R26_1_1+R27_1+R28_1+R29_1+R30_1+R31_1+R32_1+R33_1+R34_1+R35_1+R36_1+R37_1+R38_1+R39_1+R40_1+R41_1+R42_1+R1_2+R2_2+R3_2+R3_1_2+R4_2+R5_2+R5_1_2+R6_2+R7_2+R7_1_2+R8_2+R9_2+R10_2+R11_2+R12_2+R12_1_2+R12_2_2+R12_3_2+R13_2+R14_2+R15_2+R16_2+R64_2+R65_2+R17_2+R18_2+R18_1_2+R19_2+R20_2+R20_1_2+R21_2+R22_2+R23_2+R24_2+R25_2+R25_1_2+R25_2_2+R25_3_2+R43_2+R44_2+R26_2+R26_1_2+R27_2+R28_2+R29_2+R30_2+R31_2+R32_2+R33_2+R34_2+R35_2+R36_2+R37_2+R38_2+R39_2+R40_2+R41_2+R42_2+valoare_a+valoare_a1+valoare_b+valoare_b1+tva_a+tva_a1+tva_b+tva_b1

		ramburs=info.cell(row=40,column=3).value
		nrfact=info.cell(row=41,column=3).value
		baza=info.cell(row=42,column=3).value
		tva=info.cell(row=43,column=3).value
		factprimite=info.cell(row=44,column=3).value
		bazaprimite=info.cell(row=45,column=3).value
		tvaprimite=info.cell(row=46,column=3).value
		valoare_a=amount.cell(row=68,column=2).value
		valoare_a1=amount.cell(row=69,column=2).value
		tva_a=amount.cell(row=68,column=3).value
		tva_a1=amount.cell(row=69,column=3).value
		valoare_b=amount.cell(row=70,column=2).value
		valoare_b1=amount.cell(row=71,column=2).value
		tva_b=amount.cell(row=70,column=3).value
		tva_b1=amount.cell(row=71,column=3).value
		nrfactemise=info.cell(row=47,column=3).value
		total_baza=info.cell(row=48,column=3).value
		total_tva=info.cell(row=49,column=3).value
		total_precedent=info.cell(row=50,column=3).value
		total_curent=info.cell(row=51,column=3).value
		tip=info.cell(row=53,column=3).value
		pren=info.cell(row=54,column=3).value
		nume=info.cell(row=55,column=3).value
		funct=info.cell(row=56,column=3).value
		total_precedent=amount.cell(row=74,column=2).value
		total_curent=amount.cell(row=74,column=3).value
		totalp=info.cell(row=52,column=3).value

		for row in amount.iter_rows():
			for cell in row:
				if cell.value == "Taxable basis":
					rand_tb = cell.row
					suma1 = cell.column
					lun = len(amount[cell.column])
		coloana = [b.value for b in amount[suma1][rand_tb:lun]]


		R1_1=amount.cell(row=8,column=2).value
		R2_1=amount.cell(row=9,column=2).value
		R3_1=amount.cell(row=10,column=2).value
		R3_1_1=amount.cell(row=11,column=2).value
		R4_1=amount.cell(row=12,column=2).value
		R5_1=amount.cell(row=13,column=2).value
		R5_1_1=amount.cell(row=14,column=2).value
		R6_1=amount.cell(row=15,column=2).value
		R7_1=amount.cell(row=16,column=2).value
		R7_1_1=amount.cell(row=17,column=2).value
		R8_1=amount.cell(row=18,column=2).value
		R9_1=amount.cell(row=19,column=2).value
		R10_1=amount.cell(row=20,column=2).value
		R11_1=amount.cell(row=21,column=2).value
		R12_1=amount.cell(row=22,column=2).value
		R12_1_1=amount.cell(row=23,column=2).value
		R12_2_1=amount.cell(row=24,column=2).value
		R12_3_1=amount.cell(row=25,column=2).value
		R13_1=amount.cell(row=26,column=2).value
		R14_1=amount.cell(row=27,column=2).value
		R15_1=amount.cell(row=28,column=2).value
		R16_1=amount.cell(row=29,column=2).value
		R64_1=amount.cell(row=30,column=2).value
		R65_1=amount.cell(row=31,column=2).value
		R17_1=amount.cell(row=32,column=2).value
		R18_1=amount.cell(row=33,column=2).value
		R18_1_1=amount.cell(row=34,column=2).value
		R19_1=amount.cell(row=35,column=2).value
		R20_1=amount.cell(row=36,column=2).value
		R20_1_1=amount.cell(row=37,column=2).value
		R21_1=amount.cell(row=38,column=2).value
		R22_1=amount.cell(row=39,column=2).value
		R23_1=amount.cell(row=40,column=2).value
		R24_1=amount.cell(row=41,column=2).value
		R25_1=amount.cell(row=42,column=2).value
		R25_1_1=amount.cell(row=43,column=2).value
		R25_2_1=amount.cell(row=44,column=2).value
		R25_3_1=amount.cell(row=45,column=2).value
		R43_1=amount.cell(row=46,column=2).value
		R44_1=amount.cell(row=47,column=2).value
		R26_1=amount.cell(row=48,column=2).value
		R26_1_1=amount.cell(row=49,column=2).value
		R27_1=amount.cell(row=50,column=2).value
		R28_1=amount.cell(row=52,column=2).value
		R29_1=amount.cell(row=53,column=2).value
		R30_1=amount.cell(row=54,column=2).value
		R31_1=amount.cell(row=55,column=2).value
		R32_1=amount.cell(row=56,column=2).value
		R33_1=amount.cell(row=57,column=2).value
		R34_1=amount.cell(row=58,column=2).value
		R35_1=amount.cell(row=59,column=2).value
		R36_1=amount.cell(row=60,column=2).value
		R37_1=amount.cell(row=61,column=2).value
		R38_1=amount.cell(row=62,column=2).value
		R39_1=amount.cell(row=63,column=2).value
		R40_1=amount.cell(row=64,column=2).value
		R41_1=amount.cell(row=65,column=2).value
		R42_1=amount.cell(row=66,column=2).value
		R1_2=amount.cell(row=8,column=3).value
		R2_2=amount.cell(row=9,column=3).value
		R3_2=amount.cell(row=10,column=3).value
		R3_1_2=amount.cell(row=11,column=3).value
		R4_2=amount.cell(row=12,column=3).value
		R5_2=amount.cell(row=13,column=3).value
		R5_1_2=amount.cell(row=14,column=3).value
		R6_2=amount.cell(row=15,column=3).value
		R7_2=amount.cell(row=16,column=3).value
		R7_1_2=amount.cell(row=17,column=3).value
		R8_2=amount.cell(row=18,column=3).value
		R9_2=amount.cell(row=19,column=3).value
		R10_2=amount.cell(row=20,column=3).value
		R11_2=amount.cell(row=21,column=3).value
		R12_2=amount.cell(row=22,column=3).value


		R12_1_2=amount.cell(row=23,column=3).value
		R12_2_2=amount.cell(row=24,column=3).value
		R12_3_2=amount.cell(row=25,column=3).value
		R13_2=amount.cell(row=26,column=3).value
		R14_2=amount.cell(row=27,column=3).value
		R15_2=amount.cell(row=28,column=3).value
		R16_2=amount.cell(row=29,column=3).value
		R64_2=amount.cell(row=30,column=3).value
		R65_2=amount.cell(row=31,column=3).value
		R17_2=amount.cell(row=32,column=3).value
		R18_2=amount.cell(row=33,column=3).value
		R18_1_2=amount.cell(row=34,column=3).value
		R19_2=amount.cell(row=35,column=3).value
		R20_2=amount.cell(row=36,column=3).value
		R20_1_2=amount.cell(row=37,column=3).value
		R21_2=amount.cell(row=38,column=3).value
		R22_2=amount.cell(row=39,column=3).value
		R23_2=amount.cell(row=40,column=3).value
		R24_2=amount.cell(row=41,column=3).value
		R25_2=amount.cell(row=42,column=3).value
		R25_1_2=amount.cell(row=43,column=3).value
		R25_2_2=amount.cell(row=44,column=3).value
		R25_3_2=amount.cell(row=45,column=3).value


		R43_2=amount.cell(row=46,column=3).value
		R44_2=amount.cell(row=47,column=3).value
		R26_2=amount.cell(row=48,column=3).value
		R26_1_2=amount.cell(row=49,column=3).value
		R27_2=amount.cell(row=50,column=3).value


		R28_2=amount.cell(row=52,column=3).value
		R29_2=amount.cell(row=53,column=3).value
		R30_2=amount.cell(row=54,column=3).value
		R31_2=amount.cell(row=55,column=3).value
		R32_2=amount.cell(row=56,column=3).value


		R33_2=amount.cell(row=57,column=3).value
		R34_2=amount.cell(row=58,column=3).value
		R35_2=amount.cell(row=59,column=3).value
		R36_2=amount.cell(row=60,column=3).value
		R37_2=amount.cell(row=61,column=3).value
		R38_2=amount.cell(row=62,column=3).value
		R39_2=amount.cell(row=63,column=3).value
		R40_2=amount.cell(row=64,column=3).value
		R41_2=amount.cell(row=65,column=3).value
		R42_2=amount.cell(row=66,column=3).value
		


		# for i in range(0 ,len(tip)):
		# folderpath="D:/apps/TEST D3APPS/Test 21.03.2022/output"
		folderpath="C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage"
		# folderpath="C:/Users/Cristian.Iordache/Documents/D300 to XML Final CI/D300 to XML 2/storage"
		# temp.save(folderpath+".xlsx")
		text='<?xml version="1.0"?> <declaratie300  luna="'+str(luna)+'" an="'+str(an)+'" depusReprezentant="'+str(ramburs2)+'" bifa_interne="0" temei="0" prenume_declar="'+str(pren)+'" nume_declar="'+str(nume)+'" functie_declar="'+str(funct)+'" cui="'+str(cif)+'" den="'+str(den)+'" adresa="'+str(strada)+'" telefon="'+str(telefon)+'" mail="'+str(email)+'" banca="'+str(banca)+'" cont="'+str(contban)+'" caen="'+str(Caen)+'" tip_decont="'+str(tip)+'" pro_rata="'+str(prorata)+'" bifa_cereale="'+str(cereale)+'" bifa_mob="'+str(telmob)+'" bifa_disp="'+str(disp)+'" bifa_cons="'+str(cons)+'" solicit_ramb="'+str(ramburs)+'" nr_evid="'+str(nr_evid)+'" totalPlata_A="'+str(totalp)+'" R1_1="'+str(R1_1)+'" R2_1="'+str(R2_1)+'" R3_1="'+str(R3_1)+'" R3_1_1="'+str(R3_1_1)+'" R4_1="'+str(R4_1)+'" R5_1="'+str(R5_1)+'" R5_2="'+str(R5_2)+'" R5_1_1="'+str(R5_1_1)+'" R5_1_2="'+str(R5_1_2)+'" R6_1="'+str(R6_1)+'" R6_2="'+str(R6_2)+'" R7_1="'+str(R7_1)+'" R7_2="'+str(R7_2)+'" R7_1_1="'+str(R7_1_1)+'" R7_1_2="'+str(R7_1_2)+'" R8_1="'+str(R8_1)+'" R8_2="'+str(R8_2)+'" R9_1="'+str(R9_1)+'" R9_2="'+str(R9_2)+'" R10_1="'+str(R10_1)+'" R10_2="'+str(R10_2)+'" R11_1="'+str(R11_1)+'" R11_2="'+str(R11_2)+'" R12_1="'+str(R12_1)+'" R12_2="'+str(R12_2)+'" R12_1_1="'+str(R12_1_1)+'" R12_1_2="'+str(R12_1_2)+'" R12_2_1="'+str(R12_2_1)+'" R12_2_2="'+str(R12_2_2)+'" R12_3_1="'+str(R12_3_1)+'" R12_3_2="'+str(R12_3_2)+'" R13_1="'+str(R13_1)+'" R14_1="'+str(R14_1)+'" R15_1="'+str(R15_1)+'" R16_1="'+str(R16_1)+'" R16_2="'+str(R16_2)+'" R64_1="'+str(R64_1)+'" R64_2="'+str(R64_2)+'" R65_1="'+str(R65_1)+'" R65_2="'+str(R65_2)+'" R17_1="'+str(R17_1)+'" R17_2="'+str(R17_2)+'" R18_1="'+str(R18_1)+'" R18_2="'+str(R18_2)+'" R18_1_1="'+str(R18_1_1)+'" R18_1_2="'+str(R18_1_2)+'" R19_1="'+str(R19_1)+'" R19_2="'+str(R19_2)+'" R20_1="'+str(R20_1)+'" R20_2="'+str(R20_2)+'" R20_1_1="'+str(R20_1_1)+'" R20_1_2="'+str(R20_1_2)+'" R21_1="'+str(R21_1)+'" R21_2="'+str(R21_2)+'" R22_1="'+str(R22_1)+'" R22_2="'+str(R22_2)+'" R23_1="'+str(R23_1)+'" R23_2="'+str(R23_2)+'" R24_1="'+str(R24_1)+'" R24_2="'+str(R24_2)+'" R25_1="'+str(R25_1)+'" R25_2="'+str(R25_2)+'" R25_1_1="'+str(R25_1_1)+'" R25_1_2="'+str(R25_1_2)+'" R25_2_1="'+str(R25_2_1)+'" R25_2_2="'+str(R25_2_2)+'" R25_3_1="'+str(R25_3_1)+'" R25_3_2="'+str(R25_3_2)+'" R43_2="'+str(R43_2)+'" R44_2="'+str(R44_2)+'" R26_1="'+str(R26_1)+'" R26_1_1="'+str(R26_1_1)+'" R27_1="'+str(R27_1)+'" R27_2="'+str(R27_2)+'" R28_2="'+str(R28_2)+'" R29_2="'+str(R29_2)+'" R30_1="'+str(R30_1)+'" R30_2="'+str(R30_2)+'" R31_2="'+str(R31_2)+'" R32_2="'+str(R32_2)+'" R33_2="'+str(R33_2)+'" R34_2="'+str(R34_2)+'" R35_2="'+str(R35_2)+'" R36_2="'+str(R36_2)+'" R37_2="'+str(R37_2)+'" R38_2="'+str(R38_2)+'" R39_2="'+str(R39_2)+'" R40_2="'+str(R40_2)+'" R41_2="'+str(R41_2)+'" R42_2="'+str(R42_2)+'" nr_facturi="'+str(nrfact)+'" baza="'+str(baza)+'" tva="'+str(tva)+'" nr_facturi_primite="'+str(factprimite)+'" baza_primite="'+str(bazaprimite)+'" tva_primite="'+str(tvaprimite)+'" nr_fact_emise="'+str(nrfactemise)+'" total_baza="'+str(total_baza)+'" total_precedent ="'+str(total_precedent)+'" total_curent ="'+str(total_curent)+'" total_tva="'+str(total_tva)+'" valoare_a="'+str(valoare_a)+'" tva_a="'+str(tva_a)+'" valoare_a1="'+str(valoare_a1)+'" tva_a1="'+str(tva_a1)+'" valoare_b="'+str(valoare_b)+'" tva_b="'+str(tva_b)+'" valoare_b1="'+str(valoare_b1)+'" tva_b1="'+str(tva_b1)+'" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="mfp:anaf:dgti:d300:declaratie:v7 d300.xsd" xmlns="mfp:anaf:dgti:d300:declaratie:v7"></declaratie300>'
		#print(text)
		with open("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage/D300.xml", "w", encoding="utf-8") as f:
			f.write(text)
		# f=open("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage/D300.xml", "w").write(text).encode('utf-8')
		# 	# f=open("C:/Users/Cristian.Iordache/Documents/D300 to XML Final CI/D300 to XML 2/storage/D300.xml", "w").write(text)
	except:
		pass
	try:
		amount=temp['D390 for XML']
		info=temp['Other info']

		an=info.cell(row=2,column=3).value
		luna=info.cell(row=3,column=3).value
		cif=info.cell(row=5,column=3).value
		denu=info.cell(row=4,column=3).value
		telefon=info.cell(row=7,column=3).value
		email=info.cell(row=9,column=3).value
		pren=info.cell(row=54,column=3).value
		nume=info.cell(row=55,column=3).value
		funct=info.cell(row=56,column=3).value
		adresa=info.cell(row=6,column=3).value
		numar=info.cell(row=102,column=3).value
		localitate=info.cell(row=103,column=3).value
		judet=info.cell(row=104,column=3).value

		sector=info.cell(row=105,column=3).value
		if(sector==None):
			sector=""
		bloc=info.cell(row=106,column=3).value
		if(bloc==None):
			bloc=""
		scara=info.cell(row=107,column=3).value
		if(scara==None):
			scara=""
		etaj=info.cell(row=108,column=3).value
		if(etaj==None):
			etaj=""
		apt=info.cell(row=109,column=3).value
		if(apt==None):
			apt=""
		adresa=adresa

		totalplata=amount.cell(row=11,column=3).value
		nrpag=amount.cell(row=12,column=3).value
		nropi=amount.cell(row=13,column=3).value
		bazal=amount.cell(row=14,column=3).value
		bazat=amount.cell(row=15,column=3).value
		bazaa=amount.cell(row=16,column=3).value
		bazaP=amount.cell(row=17,column=3).value
		bazas=amount.cell(row=18,column=3).value
		bazar=amount.cell(row=19,column=3).value
		totalb=amount.cell(row=23,column=3).value

		for row in amount.iter_rows():
				for cell in row:
					if cell.value=="TIP":
						tipc=cell.column
						tipr=cell.row


		for row in amount.iter_rows():
				for cell in row:
					if cell.value=="ŢARA":
						tarc=cell.column
						tarr=cell.row
		for row in amount.iter_rows():
				for cell in row:
					if cell.value=="COD OPERATOR INTRACOMUNITAR":
						coic=cell.column
						coir=cell.row

		for row in amount.iter_rows():
				for cell in row:
					if cell.value=="Denumire":
						denc=cell.column
						denr=cell.row

		for row in amount.iter_rows():
				for cell in row:
					if cell.value=="BAZA IMPOZABILĂ":
						bazc=cell.column
						bazr=cell.row

		luntb=len(amount[tipc])

		try:
			tip=[b.value for b in amount[tipc][tipr:luntb+1]]
		except:
			flash("Please insert the correct header for 'TIP'")
			return render_template("index.html")


		try:
			tara=[b.value for b in amount[tarc][tipr:luntb+1]]
		except:
			flash("Please insert the correct header for 'ŢARA'")
			return render_template("index.html")

		try:
			cod=[b.value for b in amount[coic][tipr:luntb+1]]
		except:
			flash("Please insert the correct header for 'COD OPERATOR INTRACOMUNITAR'")
			return render_template("index.html")

		try:
			den=[b.value for b in amount[denc][tipr:luntb+1]]
		except:
			flash("Please insert the correct header for 'Denumire'")
			return render_template("index.html")

		try:
			baza=[b.value for b in amount[bazc][tipr:luntb+1]]
		except:
			flash("Please insert the correct header for 'BAZA IMPOZABILA'")
			return render_template("index.html")

		txt="390,"+str(luna)+","+str(an)+","+str(cif)+",#"+str(denu)+"#,#"+str(adresa)+"#,#"+str(telefon)+"#,#"+str(email)+"#"+"\n"

		string=""

		text=""
		for i in range(0,len(tip)):
			if(tip[i]!= None):
				text1="#"+str(tip[i])+"#,#"+str(tara[i])+"#,#"+str(cod[i])+"#,#"+str(den[i])+"#,"+str(baza[i])+"\n"
				text=text+text1
				string=string+'<operatie tip="'+str(tip[i])+'" tara="'+str(tara[i])+'" codO="'+str(cod[i])+'" denO="'+str(den[i])+'" baza="'+str(baza[i])+'" />'
		# for i in range(0 ,len(tip))

		texttxt=txt+text

		#print(texttxt)
		textxml='<?xml version="1.0"?> <declaratie390  luna="'+str(luna)+'" an="'+str(an)+'" d_rec="0" nume_declar="'+str(nume)+'" prenume_declar="'+str(pren)+'" functie_declar="'+str(funct)+'" cui="'+str(cif)+'" den="'+str(denu)+'" adresa="'+str(adresa)+'" telefon="'+str(telefon)+'" mail="'+str(email)+'" totalPlata_A="'+str(totalplata)+'" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="mfp:anaf:dgti:d390:declaratie:v3 D390.xsd" xmlns="mfp:anaf:dgti:d390:declaratie:v3">'+'<rezumat nr_pag="'+str(nrpag)+'" nrOPI="'+str(nropi)+'" bazaL="'+str(bazal)+'" bazaT="'+str(bazat)+'" bazaA="'+str(bazaa)+'" bazaP="'+str(bazaP)+'" bazaS="'+str(bazas)+'" bazaR="'+str(bazar)+'" total_baza="'+str(totalb)+'" /> '+string+'</declaratie390>'
		# folderpath="D:/apps/TEST D3APPS/Test 21.03.2022/output"
		folderpath="C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage"
		# folderpath="C:/Users/Cristian.Iordache/Documents/D300 to XML Final CI/D300 to XML 2/storage"
		# f=open("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage/D390.txt", "w",encoding="utf-8").write(texttxt)
		f=open("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage/D390.txt", "w",encoding="utf-8").write(texttxt)
	except:
		pass
	sheet1=temp['Other info']
	tipp=str(sheet1.cell(row=58,column=3).value)
	tippok=tipp[:1]
	an=sheet1.cell(row=2,column=3).value
	luna=sheet1.cell(row=3,column=3).value
	sisnormaldetva=sheet1.cell(row=61,column=3).value
	if(sisnormaldetva=="DA"):
		sistem=0
	else:
		sistem=1
	sistvalainc=sheet1.cell(row=62,column=3).value
	operinper=sheet1.cell(row=63,column=3).value
	opcuperafi=sheet1.cell(row=64,column=3).value
	if(opcuperafi=="Da"):	
		persafi=1
	else:
		persafi=0
	coddeinregistrare=sheet1.cell(row=65,column=3).value
	caen=sheet1.cell(row=66,column=3).value
	denumirefirma=sheet1.cell(row=67,column=3).value
	domiciliulfiscalfirma=sheet1.cell(row=6,column=3).value
	telefonfirma=sheet1.cell(row=69,column=3).value
	faxfirma=sheet1.cell(row=70,column=3).value
	emailfirma=sheet1.cell(row=71,column=3).value
	ciffirma=sheet1.cell(row=5,column=3).value
	denumireadmin=sheet1.cell(row=73,column=3).value
	domiciliulfiscaladmin=sheet1.cell(row=74,column=3).value
	telefonadmin=sheet1.cell(row=75,column=3).value
	faxadmin=sheet1.cell(row=76,column=3).value
	emailadmin=sheet1.cell(row=77,column=3).value
	cifadmin=sheet1.cell(row=78,column=3).value
	functiedecl=sheet1.cell(row=56,column=3).value
	try:
		facturi=temp['Sectiunea 2.1&2.2']
		val3=0
	except:
		val3=1

		# for row in facturi.iter_rows():
		# 	for cell in row:
		# 		if cell.value == "Serie":
		# 			row_det = cell.row
		# 			column_serie = cell.column
		# 			lun = len(facturi[cell.column])
		# serief = [b.value for b in facturi[column_serie][row_det:lun]]
	if(val3==0):
		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Inceput Emise":
					row_det = cell.row
					column_inceput = cell.column
					lun = len(facturi[cell.column])
		try:
			inceput = [b.value for b in facturi[column_inceput][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Inceput Alocate' in Facturi sheet")
			return render_template("index.html")

		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Final Emise":
					row_det = cell.row
					column_final = cell.column
					lun = len(facturi[cell.column])
		try:
			final = [b.value for b in facturi[column_final][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Final Alocate' in Facturi sheet")
			return render_template("index.html")
		
		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Tip Emise":
					row_det = cell.row
					column_final = cell.column
					lun = len(facturi[cell.column])
		try:
			tipfacturi = [b.value for b in facturi[column_final][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Final Alocate' in Facturi sheet")
			return render_template("index.html")			


		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Inceput Alocate":
					row_det = cell.row
					column_inceput = cell.column
					lun = len(facturi[cell.column])
		try:
			incepute = [b.value for b in facturi[column_inceput][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Inceput Alocate' in Facturi sheet")
			return render_template("index.html")

		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Final Alocate":
					row_det = cell.row
					column_final = cell.column
					lun = len(facturi[cell.column])
		try:
			finale = [b.value for b in facturi[column_final][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Final Alocate' in Facturi sheet")
			return render_template("index.html")
		
		for row in facturi.iter_rows():
			for cell in row:
				if cell.value == "Tip Alocate":
					row_det = cell.row
					column_final = cell.column
					lun = len(facturi[cell.column])
		try:
			tipfacturie = [b.value for b in facturi[column_final][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Final Alocate' in Facturi sheet")
			return render_template("index.html")
		tranz=temp['Tranzactii']


		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Cui partener":
					row_det = cell.row
					column_cui = cell.column
					lun = len(tranz[cell.column])
		try:
			cuip= [b.value for b in tranz[column_cui][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Cui partener' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Nume partener":
					row_det = cell.row
					column_nume = cell.column
					lun = len(tranz[cell.column])
		try:
			nume= [b.value for b in tranz[column_nume][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Nume partener' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Tip tranzactie":
					row_det = cell.row
					column_tipt = cell.column
					lun = len(tranz[cell.column])
		try:
			tiptranza= [b.value for b in tranz[column_tipt][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Tip tranzactie' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Tip partener":
					row_det = cell.row
					column_tipt = cell.column
					lun = len(tranz[cell.column])
		try:
			tip_partener= [b.value for b in tranz[column_tipt][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Tip partener' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Baza TVA":
					row_det = cell.row
					column_btv = cell.column
					lun = len(tranz[cell.column])
		try:
			bazatv= [b.value for b in tranz[column_btv][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Baza TVA' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Cota TVA":
					row_det = cell.row
					column_ctva = cell.column
					lun = len(tranz[cell.column]) 
		try:
			cotatva= [b.value for b in tranz[column_ctva][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Cota TVA' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "TVA":
					row_det = cell.row
					column_ctva = cell.column
					lun = len(tranz[cell.column]) 
		try:
			stva= [b.value for b in tranz[column_ctva][row_det:lun]]
		except:
			flash("Please insert the correct header for 'TVA' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Nr Facturi":
					row_det = cell.row
					column_nrf = cell.column
					lun = len(tranz[cell.column])
		try:
			nrfacturi= [b.value for b in tranz[column_nrf][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Nr Facturi' in Tranzactii sheet")
			return render_template("index.html")

		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Neexigibile - nu se vor raporta":
					row_det = cell.row
					column_nrf = cell.column
					lun = len(tranz[cell.column])
		try:
			neex= [b.value for b in tranz[column_nrf][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Neexigibile - nu se vor raporta' in Tranzactii sheet")
			return render_template("index.html")
		
		for row in tranz.iter_rows():
			for cell in row:
				if cell.value == "Cod si denumire NC produs(TIP V)":
					row_det = cell.row
					column_nrf = cell.column
					lun = len(tranz[cell.column])
		try:
			codv= [b.value for b in tranz[column_nrf][row_det:lun]]
		except:
			flash("Please insert the correct header for 'Cod si denumire NC produs(TIP V)' in Tranzactii sheet")
			return render_template("index.html")

		# #print(cuip,nume,tiptranza,bazatv,cotatva,nrfacturi)

		text=""

		sumaL5=0
		sumaL9=0
		sumaL19=0
		sumaL20=0
		sumaL24=0
		sumaA5=0
		sumaA9=0
		sumaA19=0
		sumaA20=0
		sumaA24=0

		sumaAI5=0
		sumaAI9=0
		sumaAI19=0
		sumaAI20=0
		sumaAI24=0

		sumaC5=0
		sumaC9=0
		sumaC19=0
		sumaC20=0
		sumaC24=0

		tvaC5=0
		tvaC9=0
		tvaC19=0
		tvaC20=0
		tvaC24=0

		tvaL5=0
		tvaL9=0
		tvaL19=0
		tvaL20=0
		tvaL24=0

		tvaA5=0
		tvaA9=0
		tvaA19=0
		tvaA20=0
		tvaA24=0

		tvaAI5=0
		tvaAI9=0
		tvaAI19=0
		tvaAI20=0
		tvaAI24=0

		nrL5=0
		nrL9=0
		nrL19=0
		nrL19=0
		nrL20=0
		nrL24=0
		nrA5=0
		nrA9=0
		nrA19=0
		nrA20=0
		nrA24=0

		nrAI5=0
		nrAI9=0
		nrAI19=0
		nrAI20=0
		nrAI24=0

		nrC5=0
		nrC9=0
		nrC19=0
		nrC20=0
		nrC24=0
		numar1=0

		numarcui1=0
		# #print(tip_partener)
		# #print(cotatva)
		count=0
		#print(bazatv)
		for i in range(0,len(tip_partener)):
			if(str(tip_partener[i])=="1" and neex[i]!="Yes"):
				if(bazatv[i]>0):
					if(str(cotatva[i])=="5"):
						if(str(tiptranza[i])=="L"):
							numar1=numar1+1
							nrL5=nrL5+int(nrfacturi[i])
							sumaL5=sumaL5+int(bazatv[i])
							tvaL5=tvaL5+int(stva[i])
						if(str(tiptranza[i])=="AI"):
							numar1=numar1+1						
							nrAI5=nrAI5+int(nrfacturi[i])
							sumaAI5=sumaAI5+int(bazatv[i])
							tvaAI5=tvaAI5+int(stva[i])
						if(str(tiptranza[i])=="A"):
							numar1=numar1+1						
							nrA5=nrA5+int(nrfacturi[i])
							sumaA5=sumaA5+int(bazatv[i])
							tvaA5=tvaA5+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							numar1=numar1+1						
							nrLS5=nrLS5+int(nrfacturi[i])
							sumaLS5=sumaLS5+int(bazatv[i])
							tvaLS5=tvaLS5+int(stva[i])
						if(str(tiptranza[i])=="C"):
							numar1=numar1+1						
							nrC5=nrC5+int(nrfacturi[i])
							sumaC5=sumaC5+int(bazatv[i])
							tvaC5=tvaC5+int(stva[i])
					if(str(cotatva[i])=="9"):
						if(str(tiptranza[i])=="L"):
							numar1=numar1+1						
							nrL9=nrL9+int(nrfacturi[i])
							sumaL9=sumaL9+int(bazatv[i])
							tvaL9=tvaL9+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							numar1=numar1+1						
							nrLS9=nrLS9+int(nrfacturi[i])
							sumaLS9=sumaLS9+int(bazatv[i])
							tvaLS9=tvaLS9+int(stva[i])
						if(str(tiptranza[i])=="AI"):
							numar1=numar1+1						
							nrAI9=nrAI9+int(nrfacturi[i])
							sumaAI9=sumaAI9+int(bazatv[i])
							tvaAI9=tvaAI9+int(stva[i])
						if(str(tiptranza[i])=="A"):
							numar1=numar1+1						
							nrA9=nrA9+int(nrfacturi[i])
							sumaA9=sumaA9+int(bazatv[i])
							tvaA9=tvaA9+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							numar1=numar1+1						
							nrC9=nrC9+int(nrfacturi[i])
							sumaC9=sumaC9+int(bazatv[i])
							tvaC9=tvaC9+int(stva[i])
					if(str(cotatva[i])=="19"):
						# #print("YES")
						if(str(tiptranza[i])=="L"):
							numar1=numar1+1						
							nrL19=nrL19+int(int(nrfacturi[i]))
							sumaL19=sumaL19+int(bazatv[i])
							tvaL19=tvaL19+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							numar1=numar1+1						
							nrLS19=nrLS19+int(nrfacturi[i])
							sumaLS19=sumaLS19+int(bazatv[i])
							tvaLS19=tvaLS19+int(stva[i])

						if(str(tiptranza[i])=="AI"):
							numar1=numar1+1						
							nrAI19=nrAI19+int(nrfacturi[i])
							sumaAI19=sumaAI19+int(bazatv[i])
							tvaAI19=tvaAI19+int(stva[i])
							#print(nrAI19)
						if(str(tiptranza[i])=="A"):
							numar1=numar1+1						
							#print(tiptranza[i],bazatv[i],stva[i])
							nrA19=nrA19+int(nrfacturi[i])
							sumaA19=sumaA19+int(bazatv[i])
							tvaA19=tvaA19+int(stva[i])
						if(str(tiptranza[i])=="C"):
							numar1=numar1+1						
							nrC19=nrC19+int(nrfacturi[i])
							sumaC19=sumaC19+int(bazatv[i])
							tvaC19=tvaC19+int(stva[i])
					if(str(cotatva[i])=="20"):
						if(str(tiptranza[i])=="L"):
							numar1=numar1+1						
							nrL20=nrL20+int(nrfacturi[i])
							sumaL20=sumaL20+int(bazatv[i])
							tvaL20=tvaL20+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							numar1=numar1+1						
							nrLS20=nrLS20+int(nrfacturi[i])
							sumaLS20=sumaLS20+int(bazatv[i])
							tvaLS20=tvaLS20+int(stva[i])
						if(str(tiptranza[i])=="AI"):
							numar1=numar1+1						
							nrAI20=nrAI20+int(nrfacturi[i])
							sumaAI20=sumaAI20+int(bazatv[i])
							tvaAI20=tvaAI20+int(stva[i])
						if(str(tiptranza[i])=="A"):
							numar1=numar1+1						
							nrA20=nrA20+int(nrfacturi[i])
							sumaA20=sumaA20+int(bazatv[i])
							tvaA20=tvaA20+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							numar1=numar1+1						
							nrC20=nrC20+int(nrfacturi[i])
							sumaC20=sumaC20+int(bazatv[i])
							tvaC20=tvaC20+int(stva[i])
					if(str(cotatva[i])=="24"):
						if(str(tiptranza[i])=="L"):
							numar1=numar1+1						
							nrL24=nrL24+int(nrfacturi[i])
							sumaL24=sumaL24+int(bazatv[i])
							tvaL24=tvaL24+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							numar1=numar1+1						
							nrLS24=nrLS24+int(nrfacturi[i])
							sumaLS24=sumaLS24+int(bazatv[i])
							tvaLS24=tvaLS24+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							numar1=numar1+1						
							nrC24=nrC24+int(nrfacturi[i])
							sumaC24=sumaC24+int(bazatv[i])
							tvaC24=tvaC24+int(stva[i])
						if(str(tiptranza[i])=="AI"):
							numar1=numar1+1						
							nrAI24=nrAI24+int(nrfacturi[i])
							sumaAI24=sumaAI24+int(bazatv[i])
							tvaAI24=tvaAI24+int(stva[i])
						if(str(tiptranza[i])=="A"):
							numar1=numar1+1						
							nrA24=nrA24+int(nrfacturi[i])
							sumaA24=sumaA24+int(bazatv[i])
							tvaA24=tvaA24+int(stva[i])
		#print(nrAI19)
		numarcui1=numar1
		if(nrL5>0 or nrA5>0 or nrAI5>0 or nrC5>0):
			text15='''<rezumat1 tip_partener="1" cota="5" facturiL="'''+str(nrL5)+'''" bazaL="'''+str(sumaL5)+'''" tvaL="'''+str(tvaL5)+'''" facturiA="'''+str(nrA5)+'''" bazaA="'''+str(sumaA5)+'''" tvaA="'''+str(tvaA5)+'''" facturiAI="'''+str(nrAI5)+'''" bazaAI="'''+str(sumaAI5)+'''" tvaAI="'''+str(tvaAI5)+'''" facturiC="'''+str(nrC5)+'''" bazaC="'''+str(sumaC5)+'''" tvaC="'''+str(tvaC5)+'''"/>'''
		else:
			text15=""
		if(nrL9>0 or nrA9>0 or nrAI5>0 or nrC9>0):
			text19='''<rezumat1 tip_partener="1" cota="9" facturiL="'''+str(nrL9)+'''" bazaL="'''+str(sumaL9)+'''" tvaL="'''+str(tvaL9)+'''" facturiA="'''+str(nrA9)+'''" bazaA="'''+str(sumaA9)+'''" tvaA="'''+str(tvaA9)+'''" facturiAI="'''+str(nrAI9)+'''" bazaAI="'''+str(sumaAI9)+'''" tvaAI="'''+str(tvaAI9)+'''" facturiC="'''+str(nrC9)+'''" bazaC="'''+str(sumaC9)+'''" tvaC="'''+str(tvaC9)+'''"/>'''
		else:
			text19=""
		if(nrL19>0 or nrA19>0 or nrAI9>0 or nrC19>0):
			text119='''<rezumat1 tip_partener="1" cota="19" facturiL="'''+str(nrL19)+'''" bazaL="'''+str(sumaL19)+'''" tvaL="'''+str(tvaL19)+'''" facturiA="'''+str(nrA19)+'''" bazaA="'''+str(sumaA19)+'''" tvaA="'''+str(tvaA19)+'''" facturiAI="'''+str(nrAI19)+'''" bazaAI="'''+str(sumaAI19)+'''" tvaAI="'''+str(tvaAI19)+'''" facturiC="'''+str(nrC19)+'''" bazaC="'''+str(sumaC19)+'''" tvaC="'''+str(tvaC19)+'''"/>'''
		else:
			text119=""
		if(nrL20>0 or nrA20>0 or nrAI20>0 or nrC20>0):
			text120='''<rezumat1 tip_partener="1" cota="20" facturiL="'''+str(nrL20)+'''" bazaL="'''+str(sumaL20)+'''" tvaL="'''+str(tvaL20)+'''" facturiA="'''+str(nrA20)+'''" bazaA="'''+str(sumaA20)+'''" tvaA="'''+str(tvaA20)+'''" facturiAI="'''+str(nrAI20)+'''" bazaAI="'''+str(sumaAI20)+'''" tvaAI="'''+str(tvaAI20)+'''" facturiC="'''+str(nrC20)+'''" bazaC="'''+str(sumaC20)+'''" tvaC="'''+str(tvaC20)+'''"/>'''
		else:
			text120=""
		if(nrL24>0 or nrA24>0 or nrAI24>0 or nrC24>0):
			text124='''<rezumat1 tip_partener="1" cota="24" facturiL="'''+str(nrL24)+'''" bazaL="'''+str(sumaL24)+'''" tvaL="'''+str(tvaL24)+'''" facturiA="'''+str(nrA24)+'''" bazaA="'''+str(sumaA24)+'''" tvaA="'''+str(tvaA24)+'''" facturiAI="'''+str(nrAI24)+'''" bazaAI="'''+str(sumaAI24)+'''" tvaAI="'''+str(tvaAI24)+'''" facturiC="'''+str(nrC24)+'''" bazaC="'''+str(sumaC24)+'''" tvaC="'''+str(tvaC24)+'''"/>'''
		else:
			text124=""
		sumaLS0=0
		sumaL5=0
		sumaL9=0
		sumaL19=0
		sumaL20=0
		sumaL24=0
		sumaLS5=0
		sumaLS9=0
		sumaLS19=0
		sumaLS20=0
		sumaLS24=0

		tvaL5=0
		tvaL9=0
		tvaL19=0
		tvaL20=0
		tvaL24=0

		tvaLS5=0
		tvaLS9=0
		tvaLS19=0
		tvaLS20=0
		tvaLS24=0

		nrN0=0
		bazaN0=0

		nrLS0=0
		nrL5=0
		nrL9=0
		nrL19=0
		nrL20=0
		nrL24=0
		nrLS5=0
		nrLS9=0
		nrLS19=0
		nrLS20=0
		nrLS24=0

		numarcui2=0
		for i in range(0,len(tip_partener)):
			if(str(tip_partener[i])=="2"):
				if(bazatv[i]>0):
					if(str(cotatva[i])=="5"):
						if(str(tiptranza[i])=="L"):
							nrL5=nrL5+int(nrfacturi[i])
							sumaL5=sumaL5+int(bazatv[i])
							tvaL5=tvaL5+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS5=nrLS5+int(nrfacturi[i])
							sumaLS5=sumaLS5+int(bazatv[i])
							tvaLS5=tvaLS5+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC5=nrC5+int(nrfacturi[i])
							sumaC5=sumaC5+int(bazatv[i])
							tvaC5=tvaC5+int(stva[i])
					if(str(cotatva[i])=="9"):
						if(str(tiptranza[i])=="L"):
							nrL9=nrL9+int(nrfacturi[i])
							sumaL9=sumaL9+int(bazatv[i])
							tvaL9=tvaL9+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS9=nrLS9+int(nrfacturi[i])
							sumaLS9=sumaLS9+int(bazatv[i])
							tvaLS9=tvaLS9+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC9=nrC9+int(nrfacturi[i])
							sumaC9=sumaC9+int(bazatv[i])
							tvaC9=tvaC9+int(stva[i])
					if(str(cotatva[i])=="19"):
						if(str(tiptranza[i])=="L"):
							nrL19=nrL19+int(nrfacturi[i])
							sumaL19=sumaL19+int(bazatv[i])
							tvaL19=tvaL19+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS19=nrLS19+int(nrfacturi[i])
							sumaLS19=sumaLS19+int(bazatv[i])
							tvaLS19=tvaLS19+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC19=nrC19+int(nrfacturi[i])
							sumaC19=sumaC19+int(bazatv[i])
							tvaC19=tvaC19+int(stva[i])
					if(str(cotatva[i])=="20"):
						if(str(tiptranza[i])=="L"):
							nrL20=nrL20+int(nrfacturi[i])
							sumaL20=sumaL20+int(bazatv[i])
							tvaL20=tvaL20+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS20=nrLS20+int(nrfacturi[i])
							sumaLS20=sumaLS20+int(bazatv[i])
							tvaLS20=tvaLS20+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC20=nrC20+int(nrfacturi[i])
							sumaC20=sumaC20+int(bazatv[i])
							tvaC20=tvaC20+int(stva[i])
					if(str(cotatva[i])=="24"):
						if(str(tiptranza[i])=="L"):
							nrL24=nrL24+int(nrfacturi[i])
							sumaL24=sumaL24+int(bazatv[i])
							tvaL24=tvaL24+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS24=nrLS24+int(nrfacturi[i])
							sumaLS24=sumaLS24+int(bazatv[i])
							tvaLS24=tvaLS24+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC24=nrC24+int(nrfacturi[i])
							sumaC24=sumaC24+int(bazatv[i])
							tvaC24=tvaC24+int(stva[i])	
		numarcui2=nrLS24+nrL24+nrLS20+nrL20+nrLS19+nrL19+nrLS9+nrL9+nrLS5+nrL5
		if(nrL5>0):
			text25='''<rezumat1 tip_partener="2" cota="5" facturiL="'''+str(nrL5)+'''" bazaL="'''+str(sumaL5)+'''" tvaL="'''+str(tvaL5)+'''"/>'''
		else:
			text25=""
		if(nrL9>0):
			text29='''<rezumat1 tip_partener="2" cota="9" facturiL="'''+str(nrL9)+'''" bazaL="'''+str(sumaL9)+'''" tvaL="'''+str(tvaL9)+'''"/>'''
		else:
			text29=""
		if(nrL19>0):
			text219='''<rezumat1 tip_partener="2" cota="19" facturiL="'''+str(nrL19)+'''" bazaL="'''+str(sumaL19)+'''" tvaL="'''+str(tvaL19)+'''"/>'''
		else:
			text219=""
		if(nrL20>0):
			text220='''<rezumat1 tip_partener="2" cota="20" facturiL="'''+str(nrL20)+'''" bazaL="'''+str(sumaL20)+'''" tvaL="'''+str(tvaL20)+'''"/>'''
		else:
			text220=""
		if(nrL24>0):
			text224='''<rezumat1 tip_partener="2" cota="24" facturiL="'''+str(nrL24)+'''" bazaL="'''+str(sumaL24)+'''" tvaL="'''+str(tvaL24)+'''"/>'''
		else:
			text224=""
		if(bazaN0>0):
			text20='''<rezumat1 tip_partener="2" cota="0" facturiN="'''+str(nrLS24)+'''" documentN="1" bazaN="'''+str(bazaN0)+'''"/>'''
		else:
			text20=""
		sumaL5=0
		sumaL9=0
		sumaL19=0
		sumaL20=0
		sumaL24=0
		sumaLS5=0
		sumaLS9=0
		sumaLS19=0
		sumaLS20=0
		sumaLS24=0

		sumaC5=0
		sumaC9=0
		sumaC19=0
		sumaC20=0
		sumaC24=0

		tvaC5=0
		tvaC9=0
		tvaC19=0
		tvaC20=0
		tvaC24=0

		tvaL5=0
		tvaL9=0
		tvaL19=0
		tvaL20=0
		tvaL24=0

		tvaLS5=0
		tvaLS9=0
		tvaLS19=0
		tvaLS20=0
		tvaLS24=0


		nrL5=0
		nrL9=0
		nrL19=0
		nrL20=0
		nrL24=0

		nrLS5=0
		nrLS9=0
		nrLS19=0
		nrLS20=0
		nrLS24=0

		nrC5=0
		nrC9=0
		nrC19=0
		nrC20=0
		nrC24=0
		numarcui3=0
		for i in range(0,len(tip_partener)):
			if(str(tip_partener[i])=="3"):
				if(bazatv[i]>0):
					if(str(cotatva[i])=="5"):
						if(str(tiptranza[i])=="L"):
							nrL5=nrL5+int(nrfacturi[i])
							sumaL5=sumaL5+int(bazatv[i])
							tvaL5=tvaL5+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS5=nrLS5+int(nrfacturi[i])
							sumaLS5=sumaLS5+int(bazatv[i])
							tvaLS5=tvaLS5+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC5=nrC5+int(nrfacturi[i])
							sumaC5=sumaC5+int(bazatv[i])
							tvaC5=tvaC5+int(stva[i])
					if(str(cotatva[i])=="9"):
						if(str(tiptranza[i])=="L"):
							nrL9=nrL9+int(nrfacturi[i])
							sumaL9=sumaL9+int(bazatv[i])
							tvaL9=tvaL9+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS9=nrLS9+int(nrfacturi[i])
							sumaLS9=sumaLS9+int(bazatv[i])
							tvaLS9=tvaLS9+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC9=nrC9+int(nrfacturi[i])
							sumaC9=sumaC9+int(bazatv[i])
							tvaC9=tvaC9+int(stva[i])
					if(str(cotatva[i])=="19"):
						if(str(tiptranza[i])=="L"):
							nrL19=nrL19+int(nrfacturi[i])
							sumaL19=sumaL19+int(bazatv[i])
							tvaL19=tvaL19+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS19=nrLS19+int(nrfacturi[i])
							sumaLS19=sumaLS19+int(bazatv[i])
							tvaLS19=tvaLS19+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC19=nrC19+int(nrfacturi[i])
							sumaC19=sumaC19+int(bazatv[i])
							tvaC19=tvaC19+int(stva[i])
					if(str(cotatva[i])=="20"):
						if(str(tiptranza[i])=="L"):
							nrL20=nrL20+int(nrfacturi[i])
							sumaL20=sumaL20+int(bazatv[i])
							tvaL20=tvaL20+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS20=nrLS20+int(nrfacturi[i])
							sumaLS20=sumaLS20+int(bazatv[i])
							tvaLS20=tvaLS20+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC20=nrC20+int(nrfacturi[i])
							sumaC20=sumaC20+int(bazatv[i])
							tvaC20=tvaC20+int(stva[i])
					if(str(cotatva[i])=="24"):
						if(str(tiptranza[i])=="L"):
							nrL24=nrL24+int(nrfacturi[i])
							sumaL24=sumaL24+int(bazatv[i])
							tvaL24=tvaL24+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS24=nrLS24+int(nrfacturi[i])
							sumaLS24=sumaLS24+int(bazatv[i])
							tvaLS24=tvaLS24+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC24=nrC24+int(nrfacturi[i])
							sumaC24=sumaC24+int(bazatv[i])
							tvaC24=tvaC24+int(stva[i])
		numarcui3=nrLS24+nrC24+nrL24+nrC20+nrLS20+nrL20+nrC19+nrLS19+nrL19+nrC19+nrLS9+nrL9+nrC9+nrLS5+nrL5+nrC5
		if(nrL5>0 or nrC5>0):
			text35='''<rezumat1 tip_partener="3" cota="5" facturiL="'''+str(nrL5)+'''" bazaL="'''+str(sumaL5)+'''" tvaL="'''+str(tvaL5)+'''" facturiC="'''+str(nrC5)+'''" bazaC="'''+str(sumaC5)+'''" tvaC="'''+str(tvaC5)+'''"/>'''
		else:
			text35=""
		if(nrL9>0 or nrC9>0):
			text39='''<rezumat1 tip_partener="3" cota="9" facturiL="'''+str(nrL9)+'''" bazaL="'''+str(sumaL9)+'''" tvaL="'''+str(tvaL9)+'''" facturiC="'''+str(nrC9)+'''" bazaC="'''+str(sumaC9)+'''" tvaC="'''+str(tvaC9)+'''"/>'''
		else:
			text39=""
		if(nrL19>0 or nrC19>0):	
			text319='''<rezumat1 tip_partener="3" cota="19" facturiL="'''+str(nrL19)+'''" bazaL="'''+str(sumaL19)+'''" tvaL="'''+str(tvaL19)+'''" facturiC="'''+str(nrC19)+'''" bazaC="'''+str(sumaC19)+'''" tvaC="'''+str(tvaC19)+'''"/>'''
		else:
			text319=""
		if(nrL20>0 or nrC20>0):	
			text320='''<rezumat1 tip_partener="3" cota="20" facturiL="'''+str(nrL20)+'''" bazaL="'''+str(sumaL20)+'''" tvaL="'''+str(tvaL20)+'''" facturiC="'''+str(nrC20)+'''" bazaC="'''+str(sumaC20)+'''" tvaC="'''+str(tvaC20)+'''"/>'''
		else:
			text320=""
		if(nrL24>0  or nrC24>0):	
			text324='''<rezumat1 tip_partener="3" cota="24" facturiL="'''+str(nrL24)+'''" bazaL="'''+str(sumaL24)+'''" tvaL="'''+str(tvaL24)+'''" facturiC="'''+str(nrC24)+'''" bazaC="'''+str(sumaC24)+'''" tvaC="'''+str(tvaC24)+'''"/>'''
		else:
			text324=""
		sumaL5=0
		sumaL9=0
		sumaL19=0
		sumaL20=0
		sumaL24=0
		sumaLS5=0
		sumaLS9=0
		sumaLS19=0
		sumaLS20=0
		sumaLS24=0

		sumaC5=0
		sumaC9=0
		sumaC19=0
		sumaC20=0
		sumaC24=0

		tvaC5=0
		tvaC9=0
		tvaC19=0
		tvaC20=0
		tvaC24=0

		tvaL5=0
		tvaL9=0
		tvaL19=0
		tvaL20=0
		tvaL24=0

		tvaLS5=0
		tvaLS9=0
		tvaLS19=0
		tvaLS20=0
		tvaLS24=0


		nrL5=0
		nrL9=0
		nrL19=0
		nrL20=0
		nrL24=0

		nrLS5=0
		nrLS9=0
		nrLS19=0
		nrLS20=0
		nrLS24=0

		nrC5=0
		nrC9=0
		nrC19=0
		nrC20=0
		nrC24=0
		numarcui4=0
		for i in range(0,len(tip_partener)):
			if(bazatv[i]>0):
				if(str(tip_partener[i])=="4"):
					if(str(cotatva[i])=="5"):
						if(str(tiptranza[i])=="L"):
							nrL5=nrL5+int(nrfacturi[i])
							sumaL5=sumaL5+int(bazatv[i])
							tvaL5=tvaL5+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS5=nrLS5+int(nrfacturi[i])
							sumaLS5=sumaLS5+int(bazatv[i])
							tvaLS5=tvaLS5+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC5=nrC5+int(nrfacturi[i])
							sumaC5=sumaC5+int(bazatv[i])
							tvaC5=tvaC5+int(stva[i])
					if(str(cotatva[i])=="9"):
						if(str(tiptranza[i])=="L"):
							nrL9=nrL9+int(nrfacturi[i])
							sumaL9=sumaL9+int(bazatv[i])
							tvaL9=tvaL9+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS9=nrLS9+int(nrfacturi[i])
							sumaLS9=sumaLS9+int(bazatv[i])
							tvaLS9=tvaLS9+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC9=nrC9+int(nrfacturi[i])
							sumaC9=sumaC9+int(bazatv[i])
							tvaC9=tvaC9+int(stva[i])
					if(str(cotatva[i])=="19"):
						if(str(tiptranza[i])=="L"):
							nrL19=nrL19+int(nrfacturi[i])
							sumaL19=sumaL19+int(bazatv[i])
							tvaL19=tvaL19+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS19=nrLS19+int(nrfacturi[i])
							sumaLS19=sumaLS19+int(bazatv[i])
							tvaLS19=tvaLS19+int(stva[i])
						if(str(tiptranza[i])=="C"):
							nrC19=nrC19+int(nrfacturi[i])
							sumaC19=sumaC19+int(bazatv[i])
							tvaC19=tvaC19+int(stva[i])
					if(str(cotatva[i])=="20"):
						if(str(tiptranza[i])=="L"):
							nrL20=nrL20+int(nrfacturi[i])
							sumaL20=sumaL20+int(bazatv[i])
							tvaL20=tvaL20+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS20=nrLS20+int(nrfacturi[i])
							sumaLS20=sumaLS20+int(bazatv[i])
							tvaLS20=tvaLS20+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC20=nrC20+int(nrfacturi[i])
							sumaC20=sumaC20+int(bazatv[i])
							tvaC20=tvaC20+int(stva[i])
					if(str(cotatva[i])=="24"):
						if(str(tiptranza[i])=="L"):
							nrL24=nrL24+int(nrfacturi[i])
							sumaL24=sumaL24+int(bazatv[i])
							tvaL24=tvaL24+int(stva[i])
						if(str(tiptranza[i])=="LS"):
							nrLS24=nrLS24+int(nrfacturi[i])
							sumaLS24=sumaLS24+int(bazatv[i])
							tvaLS24=tvaLS24+int(stva[i])		
						if(str(tiptranza[i])=="C"):
							nrC24=nrC24+int(nrfacturi[i])
							sumaC24=sumaC24+int(bazatv[i])
							tvaC24=tvaC24+int(stva[i])
		numarcui4=nrLS24+nrC24+nrL24+nrC20+nrLS20+nrL20+nrC19+nrLS19+nrL19+nrC19+nrLS9+nrL9+nrC9+nrLS5+nrL5+nrC5
		if(nrL5>0 or nrC5>0):
			text45='''<rezumat1 tip_partener="4" cota="5" facturiL="'''+str(nrL5)+'''" bazaL="'''+str(sumaL5)+'''" tvaL="'''+str(tvaL5)+'''" facturiC="'''+str(nrC5)+'''" bazaC="'''+str(sumaC5)+'''" tvaC="'''+str(tvaC5)+'''"/>'''
		else:
			text45=""
		if(nrL9>0 or nrC9>0):
			text49='''<rezumat1 tip_partener="4" cota="9" facturiL="'''+str(nrL9)+'''" bazaL="'''+str(sumaL9)+'''" tvaL="'''+str(tvaL9)+'''" facturiC="'''+str(nrC9)+'''" bazaC="'''+str(sumaC9)+'''" tvaC="'''+str(tvaC9)+'''"/>'''
		else:
			text49=""
		if(nrL19>0 or nrC19>0):	
			text419='''<rezumat1 tip_partener="4" cota="19" facturiL="'''+str(nrL19)+'''" bazaL="'''+str(sumaL19)+'''" tvaL="'''+str(tvaL19)+'''" facturiC="'''+str(nrC19)+'''" bazaC="'''+str(sumaC19)+'''" tvaC="'''+str(tvaC19)+'''"/>'''
		else:
			text419=""
		if(nrL20>0 or nrC20>0):	
			text420='''<rezumat1 tip_partener="4" cota="20" facturiL="'''+str(nrL20)+'''" bazaL="'''+str(sumaL20)+'''" tvaL="'''+str(tvaL20)+'''" facturiC="'''+str(nrC20)+'''" bazaC="'''+str(sumaC20)+'''" tvaC="'''+str(tvaC20)+'''"/>'''
		else:
			text420=""
		if(nrL24>0  or nrC24>0):	
			text424='''<rezumat1 tip_partener="4" cota="24" facturiL="'''+str(nrL24)+'''" bazaL="'''+str(sumaL24)+'''" tvaL="'''+str(tvaL24)+'''" facturiC="'''+str(nrC24)+'''" bazaC="'''+str(sumaC24)+'''" tvaC="'''+str(tvaC24)+'''"/>'''
		else:
			text424=""
		sheetG=temp['Sectiunea G. Manual input']
		sheetI=temp['Sectiunea I 1. Manual input']

		L24nr=0
		L20nr=0
		L19nr=0
		L9nr=0
		L5nr=0

		L24b=0
		L20b=0
		L19b=0
		L9b=0
		L5b=0

		L24t=0
		L20t=0
		L19t=0
		L9t=0
		L5t=0

		A24nr=0
		A20nr=0
		A19nr=0
		A9nr=0
		A5nr=0

		A24b=0
		A20b=0
		A19b=0
		A9b=0
		A5b=0

		A24t=0
		A20t=0
		A19t=0
		A9t=0
		A5t=0

		AI24nr=0
		AI20nr=0
		AI19nr=0
		AI9nr=0
		AI5nr=0

		AI24b=0
		AI20b=0
		AI19b=0
		AI9b=0
		AI5b=0

		AI24t=0
		AI20t=0
		AI19t=0
		AI9t=0
		AI5t=0




		for i in range(0,len(cotatva)):
			if(bazatv[i]):
				if(str(cotatva[i])=="24"):
					if(str(tiptranza[i])=="L"):
						L24nr=L24nr+int(nrfacturi[i])
						L24b=L24b+int(bazatv[i])
						L24t=L24t+int(stva[i])
					if(str(tiptranza[i])=="A"):
						A24nr=A24nr+int(nrfacturi[i])
						A24b=A24b+int(bazatv[i])
						A24t=A24t+int(stva[i])
					if(str(tiptranza[i])=="AI"):
						AI24nr=AI24nr+int(nrfacturi[i])
						AI24b=AI24b+int(bazatv[i])
						AI24t=AI24t+int(stva[i])
				if(str(cotatva[i])=="20"):
					if(str(tiptranza[i])=="L"):
						L20nr=L20nr+int(nrfacturi[i])
						L20b=L20b+int(bazatv[i])
						L20t=L20t+int(stva[i])
					if(str(tiptranza[i])=="A"):
						A20nr=A20nr+int(nrfacturi[i])
						A20b=A20b+int(bazatv[i])
						A20t=A20t+int(stva[i])
					if(str(tiptranza[i])=="AI"):
						AI20nr=AI20nr+int(nrfacturi[i])
						AI20b=AI20b+int(bazatv[i])
						AI20t=AI20t+int(stva[i])
				if(str(cotatva[i])=="19"):
					if(str(tiptranza[i])=="L"):
						L19nr=L19nr+int(nrfacturi[i])
						L19b=L19b+int(bazatv[i])
						L19t=L19t+int(stva[i])
					if(str(tiptranza[i])=="A"):
						A19nr=A19nr+int(nrfacturi[i])
						A19b=A19b+int(bazatv[i])
						A19t=A19t+int(stva[i])
					if(str(tiptranza[i])=="AI"):
						AI19nr=AI19nr+int(nrfacturi[i])
						AI19b=AI19b+int(int(bazatv[i]))
						AI19t=AI19t+int(int(stva[i]))
				if(str(cotatva[i])=="9"):
					if(str(tiptranza[i])=="L"):
						L9nr=L9nr+int(nrfacturi[i])
						L9b=L9b+int(int(bazatv[i]))
						L9t=L9t+int(int(stva[i]))
					if(str(tiptranza[i])=="A"):
						A9nr=A9nr+int(nrfacturi[i])
						A9b=A9b+int(int(bazatv[i]))
						A9t=A9t+int(int(stva[i]))
					if(str(tiptranza[i])=="AI"):
						AI9nr=AI9nr+int(nrfacturi[i])
						AI9b=AI9b+int(int(bazatv[i]))
						AI9t=AI9t+int(int(stva[i]))
				if(str(cotatva[i])=="5"):
					if(str(tiptranza[i])=="L"):
						L5nr=L5nr+int(nrfacturi[i])
						L5b=L5b+int(int(bazatv[i]))
						L5t=L5t+int(int(stva[i]))
					if(str(tiptranza[i])=="A"):
						A5nr=A5nr+int(nrfacturi[i])
						A5b=A5b+int(int(bazatv[i]))
						A5t=A5t+int(int(stva[i]))
					if(str(tiptranza[i])=="AI"):
						AI5nr=AI5nr+int(nrfacturi[i])
						AI5b=AI5b+int(int(bazatv[i]))
						AI5t=AI5t+int(int(stva[i]))

		rez224='''<rezumat2 cota="24"  bazaFSLcod="'''+str(sheetI.cell(row=8,column=2).value)+'''" TVAFSLcod="'''+str(sheetI.cell(row=8,column=3).value)+'''" bazaFSL="'''+str(sheetI.cell(row=18,column=2).value)+'''" TVAFSL="'''+str(sheetI.cell(row=18,column=3).value)+'''" bazaFSA="'''+str(sheetI.cell(row=28,column=2).value)+'''" TVAFSA="'''+str(sheetI.cell(row=28,column=3).value)+'''" bazaFSAI="'''+str(sheetI.cell(row=38,column=2).value)+'''" TVAFSAI="'''+str(sheetI.cell(row=38,column=3).value)+'''" bazaBFAI="'''+str(sheetI.cell(row=48,column=2).value)+'''" TVABFAI="'''+str(sheetI.cell(row=48,column=3).value)+'''" nrFacturiL="'''+str(L24nr)+'''" bazaL="'''+str(L24b)+'''" tvaL="'''+str(L24t)+'''" nrFacturiA="'''+str(A24nr)+'''" bazaA="'''+str(A24b)+'''" tvaA="'''+str(A24t)+'''" nrFacturiAI="'''+str(AI24nr)+'''" bazaAI="'''+str(AI24b)+'''" tvaAI="'''+str(AI24t)+'''" bazaL_PF="0" tvaL_PF="0" />'''
		rez220='''<rezumat2 cota="20"  bazaFSLcod="'''+str(sheetI.cell(row=9,column=2).value)+'''" TVAFSLcod="'''+str(sheetI.cell(row=9,column=3).value)+'''" bazaFSL="'''+str(sheetI.cell(row=19,column=2).value)+'''" TVAFSL="'''+str(sheetI.cell(row=19,column=3).value)+'''" bazaFSA="'''+str(sheetI.cell(row=29,column=2).value)+'''" TVAFSA="'''+str(sheetI.cell(row=29,column=3).value)+'''" bazaFSAI="'''+str(sheetI.cell(row=39,column=2).value)+'''" TVAFSAI="'''+str(sheetI.cell(row=39,column=3).value)+'''" bazaBFAI="'''+str(sheetI.cell(row=49,column=2).value)+'''" TVABFAI="'''+str(sheetI.cell(row=49,column=3).value)+'''" nrFacturiL="'''+str(L20nr)+'''" bazaL="'''+str(L20b)+'''" tvaL="'''+str(L20t)+'''" nrFacturiA="'''+str(A20nr)+'''" bazaA="'''+str(A20b)+'''" tvaA="'''+str(A20t)+'''" nrFacturiAI="'''+str(AI20nr)+'''" bazaAI="'''+str(AI20b)+'''" tvaAI="'''+str(AI20t)+'''" baza_incasari_i1="'''+str(sheetG.cell(row=14,column=2).value)+'''" tva_incasari_i1="'''+str(sheetG.cell(row=14,column=3).value)+'''" baza_incasari_i2="'''+str(sheetG.cell(row=24,column=2).value)+'''" tva_incasari_i2="'''+str(sheetG.cell(row=24,column=3).value)+'''" bazaL_PF="0" tvaL_PF="0"/>'''
		rez219='''<rezumat2 cota="19"  bazaFSLcod="'''+str(sheetI.cell(row=10,column=2).value)+'''" TVAFSLcod="'''+str(sheetI.cell(row=10,column=3).value)+'''" bazaFSL="'''+str(sheetI.cell(row=20,column=2).value)+'''" TVAFSL="'''+str(sheetI.cell(row=20,column=3).value)+'''" bazaFSA="'''+str(sheetI.cell(row=30,column=2).value)+'''" TVAFSA="'''+str(sheetI.cell(row=30,column=3).value)+'''" bazaFSAI="'''+str(sheetI.cell(row=40,column=2).value)+'''" TVAFSAI="'''+str(sheetI.cell(row=40,column=3).value)+'''" bazaBFAI="'''+str(sheetI.cell(row=50,column=2).value)+'''" TVABFAI="'''+str(sheetI.cell(row=50,column=3).value)+'''" nrFacturiL="'''+str(L19nr)+'''" bazaL="'''+str(L19b)+'''" tvaL="'''+str(L19t)+'''" nrFacturiA="'''+str(A19nr)+'''" bazaA="'''+str(A19b)+'''" tvaA="'''+str(A19t)+'''" nrFacturiAI="'''+str(AI19nr)+'''" bazaAI="'''+str(AI19b)+'''" tvaAI="'''+str(AI19t)+'''" baza_incasari_i1="'''+str(sheetG.cell(row=15,column=2).value)+'''" tva_incasari_i1="'''+str(sheetG.cell(row=15,column=3).value)+'''" baza_incasari_i2="'''+str(sheetG.cell(row=25,column=2).value)+'''" tva_incasari_i2="'''+str(sheetG.cell(row=25,column=3).value)+'''" bazaL_PF="0" tvaL_PF="0"/>'''
		rez29='''<rezumat2 cota="9"  bazaFSLcod="'''+str(sheetI.cell(row=11,column=2).value)+'''" TVAFSLcod="'''+str(sheetI.cell(row=11,column=3).value)+'''" bazaFSL="'''+str(sheetI.cell(row=21,column=2).value)+'''" TVAFSL="'''+str(sheetI.cell(row=21,column=3).value)+'''" bazaFSA="'''+str(sheetI.cell(row=31,column=2).value)+'''" TVAFSA="'''+str(sheetI.cell(row=31,column=3).value)+'''" bazaFSAI="'''+str(sheetI.cell(row=41,column=2).value)+'''" TVAFSAI="'''+str(sheetI.cell(row=41,column=3).value)+'''" bazaBFAI="'''+str(sheetI.cell(row=51,column=2).value)+'''" TVABFAI="'''+str(sheetI.cell(row=51,column=3).value)+'''" nrFacturiL="'''+str(L9nr)+'''" bazaL="'''+str(L9b)+'''" tvaL="'''+str(L9t)+'''" nrFacturiA="'''+str(A9nr)+'''" bazaA="'''+str(A9b)+'''" tvaA="'''+str(A9t)+'''" nrFacturiAI="'''+str(AI9nr)+'''" bazaAI="'''+str(AI9b)+'''" tvaAI="'''+str(AI9t)+'''" baza_incasari_i1="'''+str(sheetG.cell(row=16,column=2).value)+'''" tva_incasari_i1="'''+str(sheetG.cell(row=16,column=3).value)+'''" baza_incasari_i2="'''+str(sheetG.cell(row=26,column=2).value)+'''" tva_incasari_i2="'''+str(sheetG.cell(row=26,column=3).value)+'''" bazaL_PF="0" tvaL_PF="0"/>'''
		rez25='''<rezumat2 cota="5"  bazaFSLcod="'''+str(sheetI.cell(row=12,column=2).value)+'''" TVAFSLcod="'''+str(sheetI.cell(row=12,column=3).value)+'''" bazaFSL="'''+str(sheetI.cell(row=22,column=2).value)+'''" TVAFSL="'''+str(sheetI.cell(row=22,column=3).value)+'''" bazaFSA="'''+str(sheetI.cell(row=32,column=2).value)+'''" TVAFSA="'''+str(sheetI.cell(row=32,column=3).value)+'''" bazaFSAI="'''+str(sheetI.cell(row=42,column=2).value)+'''" TVAFSAI="'''+str(sheetI.cell(row=42,column=3).value)+'''" bazaBFAI="'''+str(sheetI.cell(row=52,column=2).value)+'''" TVABFAI="'''+str(sheetI.cell(row=52,column=3).value)+'''" nrFacturiL="'''+str(L5nr)+'''" bazaL="'''+str(L5b)+'''" tvaL="'''+str(L5t)+'''" nrFacturiA="'''+str(A5nr)+'''" bazaA="'''+str(A5b)+'''" tvaA="'''+str(A5t)+'''" nrFacturiAI="'''+str(AI5nr)+'''" bazaAI="'''+str(AI5b)+'''" tvaAI="'''+str(AI5t)+'''" baza_incasari_i1="'''+str(sheetG.cell(row=17,column=2).value)+'''" tva_incasari_i1="'''+str(sheetG.cell(row=17,column=3).value)+'''" baza_incasari_i2="'''+str(sheetG.cell(row=27,column=2).value)+'''" tva_incasari_i2="'''+str(sheetG.cell(row=27,column=3).value)+'''" bazaL_PF="0" tvaL_PF="0"/>'''
		totalplata394=numarcui1+numarcui2+numarcui3+numarcui4+L24b+L20b+L19b+L9b+L5b+A5b+A9b+A19b+A20b+A24b+AI5b+AI9b+AI19b+AI20b+AI24b
		#print(totalplata394)
		textinfo='''<?xml version="1.0"?><declaratie394 luna="'''+str(luna)+'''" an="'''+str(an)+'''" tip_D394="'''+str(tippok)+'''" sistemTVA="'''+str(sistem)+'''" op_efectuate="1" prsAfiliat="'''+str(persafi)+'''" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="mfp:anaf:dgti:d394:declaratie:v4 D394.xsd" xmlns="mfp:anaf:dgti:d394:declaratie:v4" cui="'''+str(coddeinregistrare)+'''" den="'''+str(denumirefirma)+'''" adresa="'''+str(domiciliulfiscalfirma)+'''" telefon="'''+str(telefonfirma)+'''" mail="'''+str(emailfirma)+'''" caen="'''+str(caen)+'''" totalPlata_A="'''+str(int(totalplata394))+'''" denR="'''+str(denumireadmin)+'''" functie_reprez="'''+str(functiedecl)+'''" adresaR="'''+str(domiciliulfiscaladmin)+'''" tip_intocmit="0" den_intocmit="Grant Thornton " cif_intocmit="27512924" calitate_intocmit="IMPUTERNICIT" optiune="1" schimb_optiune="1">
		<informatii nrCui1="'''+str(numarcui1)+'''" nrCui2="'''+str(numarcui2)+'''" nrCui3="'''+str(numarcui3)+'''" nrCui4="'''+str(numarcui4)+'''" nr_BF_i1="'''+str(sheetG.cell(row=5,column=2).value)+'''" incasari_i1="'''+str(sheetG.cell(row=6,column=2).value)+'''" incasari_i2="'''+str(sheetG.cell(row=7,column=2).value)+'''" nrFacturi_terti="0" nrFacturi_benef="0" nrFacturi="175" nrFacturiL_PF="0" nrFacturiLS_PF="0" val_LS_PF="0" tvaDedAI24="0" tvaDedAI20="0" tvaDedAI19="0" tvaDedAI9="0" tvaDedAI5="0" incasari_ag="0" costuri_ag="0" marja_ag="0" tva_ag="0" pret_vanzare="0" pret_cumparare="0" marja_antic="0" tva_antic="0" solicit="0"/>'''
		text=text+textinfo+text15+text19+text119+text120+text124+text25+text29+text219+text220+text224+text20+text35+text39+ text319+text320+text324+text45+text49+text419+text420+text424+rez224+rez220+rez219+rez29+rez25+"\n"


		#print(tipfacturi)
		for k in range(0,len(inceput)):
			#print(tipfacturi[k],"aici")
			if(inceput[k]!=None and tipfacturi[k]==1):
				text=text+'''<serieFacturi tip="1" nrI="'''+str(inceput[k])+''' " nrF="'''+str(final[k])+''' "/>'''+"\n"
		for kk in range(0,len(inceput)):
			#print(inceput[kk],"aici emise")
			if(inceput[kk]!=None and tipfacturi[kk]==2):
					text=text+'''<serieFacturi tip="2" nrI="'''+str(inceput[kk])+''' " nrF="'''+str(final[kk])+''' "/>'''+"\n"
		sheet7=temp[' Sectiunea 7 ']
		for j in range(13,18):
			if(sheet7.cell(row=j,column=2).value!=None):
				text=text+'<lista  caen="'+str(sheet7.cell(row=4,column=1).value)+'" cota="'+str(sheet7.cell(row=6,column=1).value)+'" operat="'+str(sheet7.cell(row=4,column=2).value)+'" valoare="'+str(sheet7.cell(row=4,column=3).value)+'" tva="'+str(sheet7.cell(row=j,column=2).value)+'" />'+"\n"

		sheet2=temp['Facturi storno si anulate']
		# de modificat seria sa o bagam in functie de faptul ca e sau nu blnak
		for k in range(4,sheet2.max_row):
			if(sheet2.cell(row=k,column=1).value=="Stornata"):
				text=text+'<facturi  tip_factura="1" nr="'''+str(sheet2.cell(row=k,column=3).value)+'''"/>'''+"\n"

		for i in range(0,len(tiptranza)):
			if("-" in str(cuip[i]) and "RO" in str(cuip[i])):
				text=text+'<op1  tip="'+str(tiptranza[i])+'" tip_partener="'+str(tip_partener[i])+'" cota="'+str(cotatva[i])+'" denP="'+str(nume[i])+'" taraP="'+str(cuip[i][:2])+'" locP="'+str(cuip[i][3:])+'" judP="'+str(cuip[i][3:])+'" nrFact="'+str(int(nrfacturi[i]))+'" baza="'+str(int(bazatv[i]))+'" tva="'+str(int(stva[i]))+'" />'+"\n"
			else:
				if("-" in str(cuip[i])):
					text=text+'<op1  tip="'+str(tiptranza[i])+'" tip_partener="'+str(tip_partener[i])+'" cota="'+str(cotatva[i])+'" denP="'+str(nume[i])+'" taraP="'+str(cuip[i][:2])+'" locP="'+str(cuip[i][3:])+'" nrFact="'+str(int(nrfacturi[i]))+'" baza="'+str(int(bazatv[i]))+'" tva="'+str(int(stva[i]))+'" />'+"\n"
				else:
					if(bazatv[i]>0):
						if(int(cotatva[i])>0 or int(tip_partener[i])<3):
							if(tiptranza=="V"):
								text=text+'<op1 tip="'+str(tiptranza[i])+'" tip_partener="'+str(tip_partener[i])+'" cota="'+str(cotatva[i])+'" cuiP="'+str(cuip[i])+'" denP="'+str(nume[i])+'"  nrFact="'+str(int(nrfacturi[i]))+'" baza="'+str(int(bazatv[i]))+'" tva="'+str(int(stva[i]))+'">'+'<op11  nrFactPR="'+str(int(nrfacturi[i]))+'" codPR="'+str(codv[i])+'" bazaPR="'+str(int(bazatv[i]))+'" /> </op1>'+"\n"
							else:
								text=text+'<op1 tip="'+str(tiptranza[i])+'" tip_partener="'+str(tip_partener[i])+'" cota="'+str(cotatva[i])+'" cuiP="'+str(cuip[i])+'" denP="'+str(nume[i])+'"  nrFact="'+str(int(nrfacturi[i]))+'" baza="'+str(int(bazatv[i]))+'" tva="'+str(int(stva[i]))+'"/>'+"\n"

		text=text+"</declaratie394>"
		# text='<?xml version="1.0"?><declaratie394 luna="'+str(luna)+'" an="'+str(an)+'" tip_D394="'+str(tip)+'" sistemTVA="'+str(sisnormaldetva)+'" op_efectuate="'+str(op_efectuate)+'" prsAfiliat="'+str(prsAfiliat)+'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="mfp:anaf:dgti:d394:declaratie:v3 D394.xsd" xmlns="mfp:anaf:dgti:d394:declaratie:v3" cui="'+str(cui)+'" den="'+str(den)+""

		# f=open("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage/D394.xml", "w",encoding='utf-8').write(text)
		# make_archive("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage","C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/arhiva VAT apps.zip")
		# return send_from_directory("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2","arhiva VAT apps.zip",as_attachment=True)
		f=open("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage/D394.xml", "w",encoding='utf-8').write(text)
	make_archive("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage","C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/arhiva VAT apps.zip")
	return send_from_directory("C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2","arhiva VAT apps.zip",as_attachment=True)
@app.route('/CUI_Extractor')
def CUI():
	return render_template('cuiRetrieval.html')

@app.route('/CUI_Extractor', methods=['POST', 'GET'])
def CUI_process():
	dataCautare = datetime.strptime(
        request.form['srcDate'],
        '%Y-%m-%d')

	dataCautare=dataCautare.date()
	if request.method == 'POST':
		file_Details = request.files["cuiFile"]

		details=openpyxl.load_workbook(file_Details, data_only='True')
		details1 = details.active

		wb = openpyxl.Workbook()
		ws = wb.active
		summary = wb.create_sheet("Summary")

        # #print("introduceri data")
        # dataa=input()
		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Cod de inregistrare fiscala":
					row_tb = cell.row
					column_cui = cell.column
					lun = len(details1[cell.column])
		try:
			cui = [b.value for b in details1[column_cui][row_tb:lun]]
		except:
			flash("Please insert the correct header for 'Cod de inregistrare fiscala'")
			return render_template("index.html")
		ccc=[]
		# try:
		for i in cui:
			# try:
			if "RO" or "RO " in i:
				b=str(i).replace("RO", "").replace(" ","")
				ccc.append(b)
			else:
				ccc.append(i)
			# except:
			# 	pass
			# #print(ccc)

		listaUnicaCui=list(set(ccc))
		#print(listaUnicaCui, "lista unica")
		#print(len(listaUnicaCui))
		# #print(dataCautare)

		x=["CUI", "Data Verificare", "Denumire", "Adresa", "Stare inregistrare", "Scop TVA", "Data Inceput Scop TVA", 
		"Mesaj Scop TVA", "Status TVA Incasare", "Status Inactivi", "Status Split TVA", "Status RO e-Factura"]
		for a in range(0, len(x)):
			summary.cell(row=1, column=1+a).value=x[a]

		for i in range(0, len(listaUnicaCui)):
			response = requests.post('https://webservicesp.anaf.ro/PlatitorTvaRest/api/v6/ws/tva', json=[{'cui': int(listaUnicaCui[i]), 'data': str(dataCautare)}])
			# time.sleep(1)

			data=response.json()
			tip=["cui", "data", "denumire", "adresa", "stare_inregistrare", "scpTVA", "data_inceput_ScpTVA", 
		"mesaj_ScpTVA", "statusTvaIncasare", "statusInactivi", "statusSplitTVA", "statusRO_e_Factura"]

			summary.cell(row=2+i, column=1).value=str(data['found'][0]['cui'])
			summary.cell(row=2+i, column=2).value=str(data['found'][0]['data'])
			summary.cell(row=2+i, column=3).value=str(data['found'][0]['denumire'])
			summary.cell(row=2+i, column=4).value=str(data['found'][0]['adresa'])
			summary.cell(row=2+i, column=5).value=str(data['found'][0]['stare_inregistrare'])
			summary.cell(row=2+i, column=6).value=str(data['found'][0]['scpTVA'])
			summary.cell(row=2+i, column=7).value=str(data['found'][0]['data_inceput_ScpTVA'])
			summary.cell(row=2+i, column=8).value=str(data['found'][0]['mesaj_ScpTVA'])
			summary.cell(row=2+i, column=9).value=str(data['found'][0]['statusTvaIncasare'])
			summary.cell(row=2+i, column=10).value=str(data['found'][0]['statusInactivi'])
			summary.cell(row=2+i, column=11).value=str(data['found'][0]['statusSplitTVA'])
			summary.cell(row=2+i, column=12).value=str(data['found'][0]['statusRO_e_Factura'])


		wr = wb["Sheet"]
		wb.remove(wr)

		folderpath="C:/Users/Bogdan.Constantinesc/Documents/D300 to XML Final CI/D300 to XML 2/storage registre"
		file_pathFS = os.path.join(folderpath, "Informatii CUI"+".xlsx")
		wb.save(file_pathFS)
	return send_from_directory(folderpath, "Informatii CUI"+".xlsx", as_attachment=True)

app.run(debug="True",host="0.0.0.0", port=120)